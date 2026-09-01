# -*- encoding: utf-8 -*-

from io import TextIOWrapper
import calendar
from collections import defaultdict
from decimal import Decimal, InvalidOperation
import io
import json
import os
import re
from django.conf import settings
from django.forms import inlineformset_factory
from django.templatetags.static import static
from django import template
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.shortcuts import render, redirect, get_object_or_404 
from django.template import loader
from django.utils.html import format_html
from django.db.models import Min, Max, Prefetch, Count, Q, Avg, IntegerField, Sum, OuterRef, Subquery
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from apps.home_final.utils import (
    _year_as_int,
    antibiotic_print_order,
    apply_final_breakpoints,
    get_breakpoint_panel_abx_codes,
    get_filtered_antibiotics,
    make_cached_breakpoint_resolver,
    resolve_breakpoint,
    resolve_organism_name,
    sort_abx_codes_by_antibiotic,
)
from .models import *
from apps.home_final.models import *
from apps.wgs_app.models import *
from .forms import *
from apps.wgs_app.forms import *
from apps.home_final.forms import *
from apps.home_final.services import concordance as concordance_service
from django.contrib import messages
# imports for generating pdf
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.templatetags.static import static
from reportlab.lib.units import cm
# for paginator
from django.core.paginator import Paginator
# for dropdown items
from django.contrib import messages
#to auto generate clinic_code, egasp id and clinic
from django.http import JsonResponse, FileResponse
#for importation 
import pandas as pd
from django.utils import timezone
from django.db.models import Q, F, Case, When
from django.utils.timezone import now


def _breakpoint_year_filter(bp_qs, specimen_year):
    target_year = _year_as_int(specimen_year)
    years_by_int = {}
    for year in bp_qs.values_list("Year", flat=True).distinct():
        year_int = _year_as_int(year)
        if year_int is not None:
            years_by_int.setdefault(year_int, []).append(year)

    if not years_by_int:
        return bp_qs.none()

    if target_year is not None:
        matching_years = years_by_int.get(target_year)
        if matching_years:
            return bp_qs.filter(Year__in=matching_years)

        selected_year = min(
            years_by_int,
            key=lambda year: (abs(year - target_year), year > target_year, year),
        )
        return bp_qs.filter(Year__in=years_by_int[selected_year])

    selected_year = max(years_by_int)
    return bp_qs.filter(Year__in=years_by_int[selected_year])


def _raw_specific_panel_codes(specimen_year, org_code):
    org_code = (org_code or "").strip()
    org_values = {org_code}
    organism = (
        Organism_List.objects
        .filter(Q(Whonet_Org_Code__iexact=org_code) | Q(Replaced_by__iexact=org_code))
        .values("Whonet_Org_Code", "Replaced_by")
        .first()
    )

    if organism:
        org_values.add((organism.get("Whonet_Org_Code") or "").strip())
        org_values.add((organism.get("Replaced_by") or "").strip())
    org_values = {value for value in org_values if value}

    org_bp_qs = BreakpointsTable.objects.filter(_org_value_filter(org_values))
    bp_qs = _breakpoint_year_filter(org_bp_qs, specimen_year)
    abx_codes = {
        (code or "").strip().upper()
        for code in bp_qs.values_list("Abx_code", flat=True).distinct()
        if (code or "").strip()
    }
    whonet_codes = {
        (code or "").strip().upper()
        for code in bp_qs.values_list("Whonet_Abx", flat=True).distinct()
        if (code or "").strip()
    }

    return abx_codes, whonet_codes


def _org_value_filter(values):
    query = Q()
    has_values = False
    for value in values:
        value = (value or "").strip()
        if not value:
            continue
        query |= Q(Org__iexact=value)
        has_values = True
    return query if has_values else Q(pk__in=[])


def _raw_related_org_values(field_name, field_value):
    values = {(field_value or "").strip()}
    if not field_value:
        return values

    org_rows = Organism_List.objects.filter(
        **{f"{field_name}__iexact": field_value}
    ).values_list("Whonet_Org_Code", "Replaced_by")

    for whonet_code, replaced_by in org_rows:
        values.add((whonet_code or "").strip())
        values.add((replaced_by or "").strip())

    return {value for value in values if value}


def _raw_year_whonet_codes(specimen_year):
    bp_qs = _breakpoint_year_filter(BreakpointsTable.objects.all(), specimen_year)
    return {
        (code or "").strip().upper()
        for code in bp_qs.values_list("Whonet_Abx", flat=True).distinct()
        if (code or "").strip()
    }


def _code_iexact_filter(field_name, codes):
    query = Q()
    has_codes = False
    for code in codes:
        code = (code or "").strip()
        if not code:
            continue
        query |= Q(**{f"{field_name}__iexact": code})
        has_codes = True
    return query if has_codes else Q(pk__in=[])


def _antibiotic_view_mode(request):
    mode = (request.GET.get("antibiotic_view") or request.POST.get("antibiotic_view") or "all").strip().lower()
    return "panel" if mode == "panel" else "all"


def _raw_antibiotics_for_panel(
    *,
    org_code,
    specimen_year,
    show_site=False,
    retest=False,
    require_org=False,
    existing_whonet_codes=None,
    antibiotic_view="all",
):
    qs = Antibiotic_List.objects.all()
    antibiotic_view = (antibiotic_view or "all").strip().lower()
    existing_whonet_codes = {
        (code or "").strip().upper()
        for code in (existing_whonet_codes or [])
        if (code or "").strip()
    }

    if retest:
        qs = qs.filter(Retest=True)
    elif show_site:
        qs = qs.filter(Show=True)

    org_code = (org_code or "").strip()
    if _is_no_organism(org_code):
        return qs.none()

    if antibiotic_view == "panel":
        if not org_code:
            return qs.none()
        breakpoint_year = _year_as_int(specimen_year) or specimen_year
        panel_codes = {
            (code or "").strip().upper()
            for code in get_breakpoint_panel_abx_codes(breakpoint_year, org_code)
            if (code or "").strip()
        }
        panel_filter = (
            _code_iexact_filter("Whonet_Abx", panel_codes)
            | _code_iexact_filter("Abx_code", panel_codes)
        )
        if existing_whonet_codes:
            panel_filter |= _code_iexact_filter("Whonet_Abx", existing_whonet_codes)
        return qs.filter(panel_filter).distinct().order_by("Antibiotic", "Whonet_Abx")

    return qs.order_by("Antibiotic", "Whonet_Abx")
from django.utils.http import urlencode, url_has_allowed_host_and_scheme
import csv
from django.utils.dateparse import parse_date
from datetime import datetime
from django.db import IntegrityError
from collections import OrderedDict, defaultdict
from django.db import transaction
from django.db.models import Count, Prefetch, Q
from django.template.loader import render_to_string
from django.db.models import Max
from itertools import islice
from django.views.decorators.http import require_GET, require_POST
from django.db.models.functions import ExtractYear, TruncMonth
from apps.home.permissions import (
    ROLE_ADMIN,
    ROLE_CHECKER,
    ROLE_ENCODER,
    ROLE_LAB_ENCODER,
    ROLE_LAB_MANAGER,
    ROLE_VERIFIER,
    can_manage_batch,
    get_user_role,
    get_user_roles,
    role_flags,
    role_required,

)
from apps.authentication.models import UserApproval


def _parse_disk_value(value):
    try:
        disk_value = int(value) if value else None
    except (TypeError, ValueError):
        return None

    if disk_value is not None and disk_value < 6:
        return None
    return disk_value


def _entry_values_match(entry, values):
    for field, value in values.items():
        current = getattr(entry, field)
        if current == value:
            continue
        if current is None and value == "":
            continue
        if current == "" and value is None:
            continue
        return False
    return True


MAIN_ANTIBIOTIC_FIELDS = [
    "ab_Antibiotic", "ab_Abx_code", "ab_Abx",
    "ab_Disk_value", "ab_Disk_RIS", "ab_Disk_enRIS",
    "ab_MIC_value", "ab_MIC_RIS", "ab_MIC_enRIS", "ab_MIC_operand",
    "ab_AlertMIC", "ab_Alert_val",
    "ab_Site_Org", "ab_R_breakpoint", "ab_I_breakpoint",
    "ab_SDD_breakpoint", "ab_S_breakpoint",
]

RETEST_ANTIBIOTIC_FIELDS = [
    "ab_Retest_Antibiotic", "ab_Retest_Abx_code", "ab_Retest_Abx",
    "ab_Retest_DiskValue", "ab_Retest_Disk_RIS", "ab_Retest_Disk_enRIS",
    "ab_Retest_MICValue", "ab_Retest_MIC_RIS", "ab_Retest_MIC_enRIS",
    "ab_Retest_MIC_operand", "ab_Retest_AlertMIC", "ab_Retest_Alert_val",
    "ab_Ret_Org", "ab_Org_Flag", "ab_Abx_Flag", "ab_Abx_Phenotype",
    "ab_Abx_Phenotype_Other", "ab_Ret_R_breakpoint", "ab_Ret_I_breakpoint",
    "ab_Ret_SDD_breakpoint", "ab_Ret_S_breakpoint",
]


def _has_main_antibiotic_data(entry):
    return any([
        entry.ab_Disk_value is not None,
        entry.ab_MIC_value is not None,
        bool((entry.ab_Disk_RIS or "").strip()),
        bool((entry.ab_Disk_enRIS or "").strip()),
        bool((entry.ab_MIC_RIS or "").strip()),
        bool((entry.ab_MIC_enRIS or "").strip()),
        bool((entry.ab_MIC_operand or "").strip()),
        bool(entry.ab_AlertMIC),
    ])


def _has_retest_antibiotic_data(entry):
    return any([
        entry.ab_Retest_DiskValue is not None,
        entry.ab_Retest_MICValue is not None,
        bool((entry.ab_Retest_Disk_RIS or "").strip()),
        bool((entry.ab_Retest_Disk_enRIS or "").strip()),
        bool((entry.ab_Retest_MIC_RIS or "").strip()),
        bool((entry.ab_Retest_MIC_enRIS or "").strip()),
        bool((entry.ab_Retest_MIC_operand or "").strip()),
        bool(entry.ab_Retest_AlertMIC),
    ])


def _clear_main_antibiotic_data(entry):
    for field in MAIN_ANTIBIOTIC_FIELDS:
        field_obj = entry._meta.get_field(field)
        if field_obj.get_internal_type() == "BooleanField":
            setattr(entry, field, False)
        elif field_obj.get_internal_type() in {"IntegerField", "PositiveSmallIntegerField", "DecimalField"}:
            setattr(entry, field, None)
        elif field_obj.null:
            setattr(entry, field, None)
        else:
            setattr(entry, field, "")


def _clear_retest_antibiotic_data(entry):
    for field in RETEST_ANTIBIOTIC_FIELDS:
        field_obj = entry._meta.get_field(field)
        if field_obj.get_internal_type() == "BooleanField":
            setattr(entry, field, False)
        elif field_obj.get_internal_type() in {"IntegerField", "PositiveSmallIntegerField", "DecimalField"}:
            setattr(entry, field, None)
        elif field_obj.null:
            setattr(entry, field, None)
        else:
            setattr(entry, field, "")


def _save_or_delete_antibiotic_entry(entry):
    if _has_main_antibiotic_data(entry) or _has_retest_antibiotic_data(entry):
        entry.save()
    else:
        entry.delete()


NO_ORGANISM_CODES = {"", "n/a", "na", "n.a.", "nv", "none", "null", "nan"}


def _is_no_organism(value):
    return (value or "").strip().lower() in NO_ORGANISM_CODES


FASTIDIOUS_PLUS_LAYOUT_SPECIES_GROUPS = {
    "ABI", "GCT", "AGT", "BD-", "BR-", "CAM", "CAR", "EIK", "FRA",
    "HA-", "HEL", "KIN", "LEG", "MOR", "NE-", "NV", "N/A", "NA",
    "N.A.", "NONE", "NULL", "NAN",
}


def _uses_fastidious_plus_layout(organism):
    if not organism:
        return False

    organism_type = str(organism.get("Organism_Type") or "").strip()
    if organism_type == "+":
        return True

    codes = {
        str(organism.get(field) or "").strip().upper()
        for field in ("Whonet_Org_Code", "Replaced_by", "Species_Group", "Genus_Group", "Genus_Code")
    }
    codes.discard("")
    if codes & FASTIDIOUS_PLUS_LAYOUT_SPECIES_GROUPS:
        return True

    return _organism_name_uses_fastidious_plus_layout(organism.get("Organism")) or bool(
        {"HIN", "NME"} & codes
    )


def _organism_name_uses_fastidious_plus_layout(organism_name):
    name = str(organism_name or "").strip().lower()
    return name.startswith(("haemophilus ", "neisseria "))


def _organism_type_is_plus(org_code, organism_name=None):
    code = str(org_code or "").strip()
    if _is_no_organism(code):
        return True
    if _organism_name_uses_fastidious_plus_layout(organism_name or code):
        return True
    name = str(organism_name or "").strip()
    if not code and not name:
        return False
    lookup = Q()
    if code:
        lookup |= Q(Whonet_Org_Code__iexact=code) | Q(Replaced_by__iexact=code) | Q(Organism__iexact=code)
    if name:
        lookup |= Q(Organism__iexact=name)
    organism = (
        Organism_List.objects
        .filter(lookup)
        .values("Whonet_Org_Code", "Replaced_by", "Organism_Type", "Species_Group", "Genus_Group", "Genus_Code", "Organism")
        .first()
    )
    return _uses_fastidious_plus_layout(organism)


RAW_PHENOTYPE_CLEAR_POST_FIELDS = {
    "phenotype_search_site_pre": ("Site_Pre", "Site_Pre_ed"),
    "phenotype_search_site_post": ("Site_Pos", "Site_Pos_ed"),
    "phenotype_search_ars_pre": ("ars_Pre", "ars_Pre_ed"),
    "phenotype_search_ars_post": ("ars_Post", "ars_Post_ed"),
}


def _is_clear_phenotype_value(value):
    return (value or "").strip().lower() in {"n/a", "na", "n.a.", "none", "null"}


def _apply_clear_phenotype_posts(isolate, post_data):
    for search_name, (source_field, edit_field) in RAW_PHENOTYPE_CLEAR_POST_FIELDS.items():
        if not (
            _is_clear_phenotype_value(post_data.get(search_name))
            or _is_clear_phenotype_value(post_data.get(edit_field))
        ):
            continue
        setattr(isolate, source_field, "")
        setattr(isolate, edit_field, "")


@login_required(login_url="login")
# auto generate clinic_code based on javascript
def get_clinic_code(request):
    site_code = request.GET.get('site_code')
    site_name = SiteData.objects.filter(SiteCode=site_code).values_list('SiteName', flat=True).first()
    return JsonResponse({'site_name': site_name})



@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER)
def settings_page(request):
    active_tab = request.GET.get("tab", "")
    edit_tat_config_id = request.GET.get("edit_tat_config") or request.GET.get("edit")
    edit_concordance_rule_id = request.GET.get("edit_concordance_rule")
    edit_non_working_id = request.GET.get("edit_non_working")
    edit_tat_config = None
    edit_concordance_rule = None
    edit_non_working_day = None
    tat_config_form = TATStepConfigForm()
    tat_overall_setting = TATOverallSetting.get_solo()

    if active_tab == "tat_config" and edit_tat_config_id:
        edit_tat_config = get_object_or_404(TATStepConfig, pk=edit_tat_config_id)
        tat_config_form = TATStepConfigForm(instance=edit_tat_config)

    users = User.objects.all().order_by("last_name", "first_name", "username")
    profile_map = {
        profile.user_id: profile
        for profile in UserProfile.objects.filter(user__in=users)
    }
    approval_map = {
        approval.user_id: approval
        for approval in UserApproval.objects.select_related("user", "reviewed_by")
    }
    staff_map = {
        staff.User_Account_id: staff
        for staff in arsStaff_Details.objects.filter(User_Account__in=users).select_related("User_Account")
    }
    available_staff = (
        arsStaff_Details.objects
        .filter(User_Account__isnull=True)
        .order_by("Staff_Name", "Staff_License")
    )
    account_rows = []
    for account in users:
        profile = profile_map.get(account.id)
        staff = staff_map.get(account.id)
        approval = approval_map.get(account.id)
        needs_staff_link = (
            account.is_active
            and not account.is_superuser
            and not account.is_staff
            and not staff
        )
        account_rows.append({
            "user": account,
            "middle_name": profile.Middle_Name if profile else "",
            "staff": staff,
            "approval": approval,
            "needs_staff_link": needs_staff_link,
            "resolved_role": get_user_role(account),
            "resolved_roles": sorted(get_user_roles(account)),
        })

    def build_wgs_pipeline_cards():
        cards = []
        for pipeline in CustomWGSPipeline.objects.prefetch_related("fields").order_by("sequencing_type", "name"):
            records = pipeline.records.all()
            cards.append({
                "pipeline": pipeline,
                "field_count": pipeline.fields.count(),
                "record_count": records.count(),
                "matched_count": records.filter(match_status="matched").count(),
                "unmatched_count": records.exclude(match_status="matched").count(),
                "last_upload": pipeline.upload_batches.first(),
            })
        return cards

    if active_tab == "concordance_settings" and edit_concordance_rule_id:
        edit_concordance_rule = get_object_or_404(
            ConcordanceOptions,
            pk=edit_concordance_rule_id,
            report__isnull=True,
        )
        concordance_options_form = ConcordanceOptionsForm(instance=edit_concordance_rule)
    else:
        concordance_options_form = ConcordanceOptionsForm()

    non_working_form = NonWorkingDayForm()
    if active_tab == "non_working" and edit_non_working_id:
        edit_non_working_day = get_object_or_404(NonWorkingDay, pk=edit_non_working_id)
        non_working_form = NonWorkingDayForm(instance=edit_non_working_day)

    concordance_org_labels = {
        value: label
        for value, label in concordance_options_form.fields["applied_org"].choices
    }
    concordance_rules = []
    for rule in concordance_service.get_global_concordance_rules():
        rule.group_codes = [
            code.strip()
            for code in (rule.applied_org_grp or "").split(",")
            if code.strip()
        ]
        normalized_applied_org = concordance_service.normalize_applied_org(rule.applied_org)
        if concordance_service.is_all_organisms_option(rule.applied_org):
            rule.applied_org_label = "All organisms"
        elif not normalized_applied_org or rule.applied_org == "-":
            rule.applied_org_label = "n/a"
        else:
            rule.applied_org_label = concordance_org_labels.get(
                normalized_applied_org,
                normalized_applied_org,
            )
        concordance_rules.append(rule)

    non_working_queryset = NonWorkingDay.objects.all()
    non_working_paginator = Paginator(non_working_queryset, 12)
    non_working_page_number = request.GET.get("non_working_page")
    non_working_page_obj = non_working_paginator.get_page(non_working_page_number)

    context = {
        "antibiotic_form": AntibioticsForm(),
        "org_form": OrganismForm(),
        "breakpoint_form": BreakpointsForm(),
        "site_form": SiteCode_Form(),
        "specimen_form": SpecimenTypeForm(),
        "contact_form": ContactForm(can_edit_roles=role_flags(request.user)["can_edit_staff_roles"]),
        "emerging_form":Emerge_Pheno_Form(),
        "abx_upload_form": Antibiotics_uploadForm(),
        'bp_upload_form': Breakpoint_uploadForm(),
        "site_upload_form": SiteCode_uploadForm(),
        "org_upload_form": Organism_uploadForm(),
        "eme_upload_form": Emerging_Crit_upload(),
        "specimen_upload": SpecimenUploadForm(),
        "pheno_pre_form": Phenotype_Pre_Form(),
        "phenotype_pre_upload": Pheno_pre_upForm(),
        "pheno_post_form": Phenotype_Post_Form(),
        "phenotype_post_upload": Pheno_post_upForm(),
        "reco_desc_form": Recco_item_Form(),
        "reco_desc_upload": Reco_item_upForm(),
        "tat_config_form": tat_config_form,
        "tat_overall_form": TATOverallSettingForm(instance=tat_overall_setting),
        "edit_tat_config": edit_tat_config,
        "tat_config_editing": active_tab == "tat_config" and edit_tat_config is not None,
        "tat_upload_form": TATStepConfigUploadForm(),
        "tat_location_form": TATLocationForm(),
        "concordance_options_form": concordance_options_form,
        "edit_concordance_rule": edit_concordance_rule,
        "concordance_rule_editing": edit_concordance_rule is not None,
        "concordance_rules": concordance_rules,
        "concordance_org_choices": concordance_options_form.fields["applied_org"].choices,
        "concordance_group_choices": concordance_options_form.fields["applied_org_grp"].choices,
        "tat_locations": TATLocation.objects.all(),
        "non_working_form": non_working_form,
        "edit_non_working_day": edit_non_working_day,
        "non_working_editing": edit_non_working_day is not None,
        "non_working_days": non_working_page_obj,
        "non_working_page_obj": non_working_page_obj,
        "non_working_total": non_working_paginator.count,
        "account_rows": account_rows,
        "available_staff": available_staff,
        "approval_require_staff_link": getattr(settings, "ACCOUNT_APPROVAL_REQUIRE_STAFF_LINK", False),
        "builtin_wgs_settings": BuiltinWGSPipelineSetting.objects.all(),
        "wgs_pipeline_cards": build_wgs_pipeline_cards(),
        "user_access": role_flags(request.user),



        "editing": False,  # default state
    }

    return render(request, "home/Settings.html", context)


@login_required
@role_required(ROLE_ADMIN)
@require_POST
def approve_account(request, approval_id):
    approval = get_object_or_404(
        UserApproval.objects.select_related("user"),
        pk=approval_id,
        status=UserApproval.STATUS_PENDING,
    )
    staff_id = request.POST.get("staff_id")
    if not staff_id:
        messages.error(request, "Select an unlinked ARSP staff record before approving the account.")
        return redirect("/settings/?tab=accounts_tab")

    try:
        with transaction.atomic():
            staff = arsStaff_Details.objects.select_for_update().get(pk=staff_id)
            if staff.User_Account_id is not None:
                messages.error(request, "That ARSP staff record is already linked to another account.")
                return redirect("/settings/?tab=accounts_tab")

            user = approval.user
            staff.User_Account = user
            staff.save(update_fields=["User_Account"])

            user.is_active = True
            user.save(update_fields=["is_active"])

            approval.status = UserApproval.STATUS_APPROVED
            approval.reviewed_by = request.user
            approval.reviewed_at = timezone.now()
            approval.save(update_fields=["status", "reviewed_by", "reviewed_at"])
    except arsStaff_Details.DoesNotExist:
        messages.error(request, "Selected ARSP staff record was not found.")
        return redirect("/settings/?tab=accounts_tab")

    messages.success(request, f"{approval.user.username} was approved and linked to {staff.display_name}.")
    return redirect("/settings/?tab=accounts_tab")


@login_required
@role_required(ROLE_ADMIN)
@require_POST
def decline_account(request, approval_id):
    approval = get_object_or_404(
        UserApproval.objects.select_related("user"),
        pk=approval_id,
        status=UserApproval.STATUS_PENDING,
    )
    user = approval.user
    user.is_active = False
    user.save(update_fields=["is_active"])

    approval.status = UserApproval.STATUS_DECLINED
    approval.reviewed_by = request.user
    approval.reviewed_at = timezone.now()
    approval.save(update_fields=["status", "reviewed_by", "reviewed_at"])

    messages.success(request, f"{user.username} was declined. No ARSP staff record was changed.")
    return redirect("/settings/?tab=accounts_tab")


@login_required
@role_required(ROLE_ADMIN)
@require_POST
def delete_account_registration(request, approval_id):
    approval = get_object_or_404(
        UserApproval.objects.select_related("user"),
        pk=approval_id,
    )
    if approval.status == UserApproval.STATUS_APPROVED:
        messages.error(request, "Approved accounts cannot be deleted from the registration approval workflow.")
        return redirect("/settings/?tab=accounts_tab")

    username = approval.user.username
    approval.user.delete()
    messages.success(request, f"Registration for {username} was deleted. ARSP staff records were not changed.")
    return redirect("/settings/?tab=accounts_tab")


@login_required
@role_required(ROLE_ADMIN)
@require_POST
def link_account_staff(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    staff_id = request.POST.get("staff_id")
    if not staff_id:
        messages.error(request, "Select an unlinked ARSP staff record before linking the account.")
        return redirect("/settings/?tab=accounts_tab")

    if arsStaff_Details.objects.filter(User_Account=user).exists():
        messages.error(request, "That account is already linked to an ARSP staff record.")
        return redirect("/settings/?tab=accounts_tab")

    try:
        with transaction.atomic():
            staff = arsStaff_Details.objects.select_for_update().get(pk=staff_id)
            if staff.User_Account_id is not None:
                messages.error(request, "That ARSP staff record is already linked to another account.")
                return redirect("/settings/?tab=accounts_tab")
            staff.User_Account = user
            staff.save(update_fields=["User_Account"])
    except arsStaff_Details.DoesNotExist:
        messages.error(request, "Selected ARSP staff record was not found.")
        return redirect("/settings/?tab=accounts_tab")

    messages.success(request, f"{user.username} was linked to {staff.display_name}.")
    return redirect("/settings/?tab=accounts_tab")


@login_required
@role_required(ROLE_ADMIN)
@require_POST
def set_account_active_status(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    action = (request.POST.get("action") or "").strip().lower()

    if user.is_superuser:
        messages.error(request, "Superuser accounts must be managed in Django Admin.")
        return redirect("/settings/?tab=accounts_tab")

    if user == request.user and action == "deactivate":
        messages.error(request, "You cannot deactivate your own account.")
        return redirect("/settings/?tab=accounts_tab")

    if action == "activate":
        user.is_active = True
        user.save(update_fields=["is_active"])
        messages.success(request, f"{user.username} was reactivated.")
    elif action == "deactivate":
        user.is_active = False
        user.save(update_fields=["is_active"])
        messages.success(request, f"{user.username} was deactivated. ARSP staff records were not changed.")
    else:
        messages.error(request, "Unsupported account status action.")

    return redirect("/settings/?tab=accounts_tab")


@login_required
@role_required(ROLE_ADMIN)
@require_POST
def delete_unlinked_account(request, user_id):
    user = get_object_or_404(User, pk=user_id)

    if user.is_superuser:
        messages.error(request, "Superuser accounts must be managed in Django Admin.")
        return redirect("/settings/?tab=accounts_tab")

    if user == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect("/settings/?tab=accounts_tab")

    if arsStaff_Details.objects.filter(User_Account=user).exists():
        messages.error(request, "Linked accounts cannot be deleted here. Deactivate the account instead.")
        return redirect("/settings/?tab=accounts_tab")

    username = user.username
    user.delete()
    messages.success(request, f"Unlinked account {username} was deleted. ARSP staff records were not changed.")
    return redirect("/settings/?tab=accounts_tab")


@login_required(login_url="login")
@require_GET
def check_setting_duplicate(request):
    setting_type = request.GET.get("type", "").strip()
    value = request.GET.get("value", "").strip()
    extra = request.GET.dict()

    if not setting_type or not value:
        return JsonResponse({"exists": False})

    checks = {
        "site_code": lambda: SiteData.objects.filter(SiteCode__iexact=value).exists(),
        "antibiotic": lambda: Antibiotic_List.objects.filter(Whonet_Abx__iexact=value).exists(),
        "organism": lambda: Organism_List.objects.filter(Whonet_Org_Code__iexact=value).exists(),
        "specimen": lambda: SpecimenTypeModel.objects.filter(Specimen_code__iexact=value).exists(),
        "staff_name": lambda: arsStaff_Details.objects.filter(Staff_Name__iexact=value).exists(),
        "staff_email": lambda: arsStaff_Details.objects.filter(Staff_EmailAdd__iexact=value).exists(),
        "staff_user": lambda: value.isdigit() and arsStaff_Details.objects.filter(User_Account_id=int(value)).exists(),
        "emerging_age": lambda: value.isdigit() and Emerging_Filter_Age.objects.filter(Eme_Age=int(value)).exists(),
        "phenotype_pre": lambda: Phenotype_Pre.objects.filter(Pre_Phenotypes__iexact=value).exists(),
        "phenotype_post": lambda: Phenotype_Post.objects.filter(Post_Phenotypes__iexact=value).exists(),
        "recommendation": lambda: Recommendation_items.objects.filter(RecoCode__iexact=value).exists(),
        "non_working_day": lambda: NonWorkingDay.objects.filter(date=value).exists(),
        "tat_location": lambda: TATLocation.objects.filter(name__iexact=value).exists(),
        "tat_config": lambda: TATStepConfig.objects.filter(
            step_type__iexact=value,
            step_owner__iexact=(extra.get("step_owner") or "").strip(),
        ).exists() if (extra.get("step_owner") or "").strip() else False,
    }

    checker = checks.get(setting_type)
    if not checker:
        return JsonResponse({"exists": False})

    return JsonResponse({"exists": bool(checker())})



@login_required(login_url="login")
def index(request):
    current_year = timezone.localdate().year
    isolates = Final_Data.objects.all().order_by('-f_Date_of_Entry')
    tat_records = (
        TATform.objects
        .select_related("tat_Batch_Isolates")
        .prefetch_related("steps__step_config")
        .order_by("-tat_Referral_Date", "tat_Batch_Code")
    )
    tat_configs = TATStepConfig.objects.all().order_by("order")

    tat_q = request.GET.get("tat_q", "").strip()
    tat_status = request.GET.get("tat_status", "").strip()
    tat_year = request.GET.get("tat_year", str(current_year)).strip() or str(current_year)
    records_for_stats = Referred_Data.objects.all()
    if tat_year != "all":
        records_for_stats = records_for_stats.filter(Referral_Date__year=tat_year)

    if tat_q:
        tat_records = tat_records.filter(
            Q(tat_Batch_Code__icontains=tat_q)
            | Q(tat_SiteCode__icontains=tat_q)
            | Q(tat_Batch_Isolates__bat_Batch_Name__icontains=tat_q)
            | Q(tat_Batch_Isolates__bat_RefNo__icontains=tat_q)
        )

    if tat_status:
        tat_records = tat_records.filter(tat_Status_Release=tat_status)

    if tat_year and tat_year != "all":
        tat_records = tat_records.filter(tat_Referral_Date__year=tat_year)

    tat_running_limit = 20
    tat_running_total = tat_records.count()
    tat_preview_records = tat_records[:tat_running_limit]
    tat_running_rows = _build_tat_running_rows(tat_preview_records, tat_configs)
    tat_years = list(
        TATform.objects
        .exclude(tat_Referral_Date__isnull=True)
        .annotate(year=ExtractYear("tat_Referral_Date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )
    if current_year not in tat_years:
        tat_years.insert(0, current_year)

    tat_overdue_count = tat_records.filter(
        tat_Target_Days__gt=0,
    ).filter(
        Q(tat_Status_Release="Released", tat_Final_TAT__gt=F("tat_Target_Days"))
        | (Q(tat_Running_TAT__gt=F("tat_Target_Days")) & ~Q(tat_Status_Release="Released"))
    ).count()
    tat_near_due_count = tat_records.filter(
        tat_Status_Release="Ongoing",
        tat_Target_Days__gt=0,
        tat_Running_TAT__gte=F("tat_Target_Days") - 5,
        tat_Running_TAT__lte=F("tat_Target_Days"),
    ).count()

    tat_status_counts = {
        "total": tat_running_total,
        "ongoing": tat_records.filter(tat_Status_Release="Ongoing").exclude(
            tat_Target_Days__gt=0,
            tat_Running_TAT__gt=F("tat_Target_Days"),
        ).count(),
        "released": tat_records.filter(tat_Status_Release="Released").count(),
        "overdue": tat_overdue_count,
        "near_due": tat_near_due_count,
    }
    tat_yearly_status_rows = []
    if tat_year == "all":
        for row in (
            tat_records
            .exclude(tat_Referral_Date__isnull=True)
            .annotate(year=ExtractYear("tat_Referral_Date"))
            .values("year")
            .annotate(total=Count("id"))
            .order_by("year")
        ):
            year_records = tat_records.filter(tat_Referral_Date__year=row["year"])
            overdue_count = year_records.filter(
                tat_Target_Days__gt=0,
            ).filter(
                Q(tat_Status_Release="Released", tat_Final_TAT__gt=F("tat_Target_Days"))
                | (Q(tat_Running_TAT__gt=F("tat_Target_Days")) & ~Q(tat_Status_Release="Released"))
            ).count()
            near_due_count = year_records.filter(
                tat_Status_Release="Ongoing",
                tat_Target_Days__gt=0,
                tat_Running_TAT__gte=F("tat_Target_Days") - 5,
                tat_Running_TAT__lte=F("tat_Target_Days"),
            ).count()
            tat_yearly_status_rows.append({
                "year": row["year"],
                "total": row["total"] or 0,
                "ongoing": year_records.filter(tat_Status_Release="Ongoing").exclude(
                    tat_Target_Days__gt=0,
                    tat_Running_TAT__gt=F("tat_Target_Days"),
                ).count(),
                "near_due": near_due_count,
                "overdue": overdue_count,
                "released": year_records.filter(tat_Status_Release="Released").count(),
            })
    # Count per clinic
    site_count = records_for_stats.exclude(SiteCode__isnull=True).exclude(SiteCode__exact="").values('SiteCode').distinct().count()

    # Count per city (assuming you have a 'Current_City' field)
    record_count = records_for_stats.values('AccessionNo').distinct().count()

    # Count per sex
    male_count = records_for_stats.filter(Sex='Male').count()
    female_count = records_for_stats.filter(Sex='Female').count()

    # Count per age group
    age_0_18 = records_for_stats.filter(Age__lte=18).count()
    age_19_35 = records_for_stats.filter(Age__range=(19, 35)).count()
    age_36_60 = records_for_stats.filter(Age__range=(36, 60)).count()
    age_60_plus = records_for_stats.filter(Age__gte=61).count()

    concordance_reports = (
        ConcordanceReport.objects
        .filter(final_data__isnull=True, batch__isnull=False)
        .select_related("batch")
    )
    if tat_year and tat_year != "all":
        concordance_reports = concordance_reports.filter(batch__bat_Referral_Date__year=tat_year)

    concordance_summary = concordance_reports.aggregate(
        report_count=Count("id"),
        avg_genus_rate=Avg("genus_rate"),
        avg_species_rate=Avg("species_rate"),
        total_pairs=Sum("total_pairs"),
        concordant_pairs=Sum("concordant_pairs"),
        total_deviation=Sum("total_deviation"),
        critical_deviation=Sum("critical_deviation"),
    )
    concordance_total_pairs = concordance_summary["total_pairs"] or 0
    concordance_concordant_pairs = concordance_summary["concordant_pairs"] or 0
    concordance_total_deviation = concordance_summary["total_deviation"] or 0
    concordance_critical_deviation = concordance_summary["critical_deviation"] or 0
    concordance_ast_rate = (
        round((concordance_concordant_pairs / concordance_total_pairs) * 100, 2)
        if concordance_total_pairs else 0
    )
    concordance_kpis = {
        "year": "All Years" if tat_year == "all" else tat_year,
        "report_count": concordance_summary["report_count"] or 0,
        "genus_rate": round(concordance_summary["avg_genus_rate"] or 0, 2),
        "species_rate": round(concordance_summary["avg_species_rate"] or 0, 2),
        "ast_rate": concordance_ast_rate,
        "total_deviation": concordance_total_deviation,
        "critical_deviation": concordance_critical_deviation,
        "total_pairs": concordance_total_pairs,
    }
    concordance_kpi_year_rows = []
    if tat_year == "all":
        for row in (
            concordance_reports
            .exclude(batch__bat_Referral_Date__isnull=True)
            .annotate(year=ExtractYear("batch__bat_Referral_Date"))
            .values("year")
            .annotate(
                report_count=Count("id"),
                genus_rate=Avg("genus_rate"),
                species_rate=Avg("species_rate"),
                total_pairs=Sum("total_pairs"),
                concordant_pairs=Sum("concordant_pairs"),
                total_deviation=Sum("total_deviation"),
                critical_deviation=Sum("critical_deviation"),
            )
            .order_by("year")
        ):
            total_pairs = row["total_pairs"] or 0
            concordant_pairs = row["concordant_pairs"] or 0
            concordance_kpi_year_rows.append({
                "year": row["year"],
                "report_count": row["report_count"] or 0,
                "genus_rate": round(row["genus_rate"] or 0, 2),
                "species_rate": round(row["species_rate"] or 0, 2),
                "ast_rate": round((concordant_pairs / total_pairs) * 100, 2) if total_pairs else 0,
                "total_deviation": row["total_deviation"] or 0,
                "critical_deviation": row["critical_deviation"] or 0,
                "total_pairs": total_pairs,
            })

    site_concordance_rows = []
    site_report_rows = concordance_reports.values(
        "batch__bat_SiteCode",
        "batch__bat_Site_Name",
    ).annotate(
        report_count=Count("id"),
        genus_rate=Avg("genus_rate"),
        species_rate=Avg("species_rate"),
        total_pairs=Sum("total_pairs"),
        concordant_pairs=Sum("concordant_pairs"),
        total_deviation=Sum("total_deviation"),
    )
    for row in site_report_rows:
        total_pairs = row["total_pairs"] or 0
        concordant_pairs = row["concordant_pairs"] or 0
        site_concordance_rows.append({
            "year": "All Years" if tat_year == "all" else tat_year,
            "site_code": row["batch__bat_SiteCode"] or "N/A",
            "site_name": row["batch__bat_Site_Name"] or "",
            "report_count": row["report_count"] or 0,
            "genus_rate": round(row["genus_rate"] or 0, 2),
            "species_rate": round(row["species_rate"] or 0, 2),
            "ast_rate": round((concordant_pairs / total_pairs) * 100, 2) if total_pairs else 0,
            "total_deviation": row["total_deviation"] or 0,
        })
    site_concordance_rows = sorted(
        site_concordance_rows,
        key=lambda item: item["genus_rate"],
        reverse=True,
    )[:12]
    site_concordance_year_rows = []
    if tat_year == "all":
        site_year_report_rows = (
            concordance_reports
            .exclude(batch__bat_Referral_Date__isnull=True)
            .annotate(year=ExtractYear("batch__bat_Referral_Date"))
            .values("year", "batch__bat_SiteCode", "batch__bat_Site_Name")
            .annotate(
                report_count=Count("id"),
                genus_rate=Avg("genus_rate"),
                species_rate=Avg("species_rate"),
                total_pairs=Sum("total_pairs"),
                concordant_pairs=Sum("concordant_pairs"),
                total_deviation=Sum("total_deviation"),
            )
            .order_by("year", "batch__bat_SiteCode")
        )
        for row in site_year_report_rows:
            total_pairs = row["total_pairs"] or 0
            concordant_pairs = row["concordant_pairs"] or 0
            site_concordance_year_rows.append({
                "year": row["year"],
                "site_code": row["batch__bat_SiteCode"] or "N/A",
                "site_name": row["batch__bat_Site_Name"] or "",
                "report_count": row["report_count"] or 0,
                "genus_rate": round(row["genus_rate"] or 0, 2),
                "species_rate": round(row["species_rate"] or 0, 2),
                "ast_rate": round((concordant_pairs / total_pairs) * 100, 2) if total_pairs else 0,
                "total_deviation": row["total_deviation"] or 0,
            })

    monthly_referral_rows = [
        {
            "year": row["month"].year,
            "month": row["month"].strftime("%b %Y"),
            "referrals": row["referrals"] or 0,
            "sites": row["sites"] or 0,
        }
        for row in records_for_stats.exclude(Referral_Date__isnull=True)
        .annotate(month=TruncMonth("Referral_Date"))
        .values("month")
        .annotate(
            referrals=Count("AccessionNo", distinct=True),
            sites=Count("SiteCode", distinct=True),
        )
        .order_by("month")
    ]

    # WGS_Project is the central connector across WGS uploads/pipelines.
    # A final accession can have more than one connector row, so keep row
    # volume and unique linked accessions as separate dashboard metrics.
    linked_wgs_projects = WGS_Project.objects.filter(Ref_Accession__isnull=False)
    if tat_year and tat_year != "all":
        linked_wgs_projects = linked_wgs_projects.filter(Ref_Accession__f_Referral_Date__year=tat_year)
    wgs_project_count = linked_wgs_projects.count()
    wgs_matched_count = linked_wgs_projects.values("Ref_Accession").distinct().count()
    wgs_unmatched_count = WGS_Project.objects.filter(Ref_Accession__isnull=True).count()
    default_builtin_overview_pipeline_keys = {
        "bactscout",
        "gambit",
        "mlst",
        "checkm2",
        "assembly",
        "amrfinder",
    }
    builtin_active_pipeline_count = BuiltinWGSPipelineSetting.objects.filter(show_in_overview=True).count()
    if not BuiltinWGSPipelineSetting.objects.exists():
        builtin_active_pipeline_count = len(default_builtin_overview_pipeline_keys)
    custom_active_pipeline_count = CustomWGSPipeline.objects.filter(is_active=True).count()
    wgs_active_pipeline_count = builtin_active_pipeline_count + custom_active_pipeline_count
    emerging_records = Emerging_Table.fully_emerging()
    final_emerging_records = Final_Data.objects.filter(
        Q(f_Spec_Emerging=True)
        | Q(f_Emerging_Flag_Age=True)
    )
    if tat_year and tat_year != "all":
        emerging_records = emerging_records.filter(eme_primary_key__f_Referral_Date__year=tat_year)
        final_emerging_records = final_emerging_records.filter(f_Referral_Date__year=tat_year)
    emerging_count = emerging_records.count()
    final_emerging_count = final_emerging_records.count()
    surveillance_snapshot = {
        "year": "All Years" if tat_year == "all" else tat_year,
        "wgs_projects": wgs_project_count,
        "wgs_matched": wgs_matched_count,
        "wgs_unmatched": wgs_unmatched_count,
        "wgs_active_pipelines": wgs_active_pipeline_count,
        "emerging_records": emerging_count,
        "final_emerging_flags": final_emerging_count,
        "concordance_reports": concordance_kpis["report_count"],
        "tat_overdue": tat_status_counts["overdue"],
    }
    surveillance_snapshot_year_rows = []
    if tat_year == "all":
        linked_wgs_years = {
            row["year"]: row
            for row in (
                linked_wgs_projects
                .exclude(Ref_Accession__f_Referral_Date__isnull=True)
                .annotate(year=ExtractYear("Ref_Accession__f_Referral_Date"))
                .values("year")
                .annotate(
                    wgs_projects=Count("id"),
                    wgs_matched=Count("Ref_Accession", distinct=True),
                )
                .order_by("year")
            )
        }
        emerging_years = {
            row["year"]: row["count"] or 0
            for row in (
                Emerging_Table.fully_emerging()
                .exclude(eme_primary_key__f_Referral_Date__isnull=True)
                .annotate(year=ExtractYear("eme_primary_key__f_Referral_Date"))
                .values("year")
                .annotate(count=Count("id"))
                .order_by("year")
            )
        }
        final_emerging_years = {
            row["year"]: row["count"] or 0
            for row in (
                Final_Data.objects
                .filter(Q(f_Spec_Emerging=True) | Q(f_Emerging_Flag_Age=True))
                .exclude(f_Referral_Date__isnull=True)
                .annotate(year=ExtractYear("f_Referral_Date"))
                .values("year")
                .annotate(count=Count("id"))
                .order_by("year")
            )
        }
        concordance_years = {
            row["year"]: row["report_count"] or 0
            for row in concordance_kpi_year_rows
        }
        tat_overdue_years = {
            row["year"]: row["overdue"] or 0
            for row in tat_yearly_status_rows
        }
        referral_years = set(
            Referred_Data.objects
            .exclude(Referral_Date__isnull=True)
            .annotate(year=ExtractYear("Referral_Date"))
            .values_list("year", flat=True)
            .distinct()
        )
        available_dashboard_years = {
            year_value for year_value in tat_years if year_value
        } | referral_years
        for year_value in sorted(
            available_dashboard_years
            | set(linked_wgs_years)
            | set(emerging_years)
            | set(final_emerging_years)
            | set(concordance_years)
            | set(tat_overdue_years)
        ):
            linked_row = linked_wgs_years.get(year_value, {})
            surveillance_snapshot_year_rows.append({
                "year": year_value,
                "wgs_projects": linked_row.get("wgs_projects", 0),
                "wgs_matched": linked_row.get("wgs_matched", 0),
                "wgs_unmatched": "",
                "wgs_active_pipelines": wgs_active_pipeline_count,
                "emerging_records": emerging_years.get(year_value, 0),
                "final_emerging_flags": final_emerging_years.get(year_value, 0),
                "concordance_reports": concordance_years.get(year_value, 0),
                "tat_overdue": tat_overdue_years.get(year_value, 0),
            })
        if wgs_unmatched_count:
            surveillance_snapshot_year_rows.append({
                "year": "No linked accession year",
                "wgs_projects": "",
                "wgs_matched": "",
                "wgs_unmatched": wgs_unmatched_count,
                "wgs_active_pipelines": "",
                "emerging_records": "",
                "final_emerging_flags": "",
                "concordance_reports": "",
                "tat_overdue": "",
            })

    # Include all context variables
    context = {
        'isolates': isolates,
        'site_count': site_count,
        'record_count': record_count,
        'male_count': male_count,
        'female_count': female_count,
        'age_0_18': age_0_18,
        'age_19_35': age_19_35,
        'age_36_60': age_36_60,
        'age_60_plus': age_60_plus,
        'tat_entries': tat_records,
        'tat_configs': tat_configs,
        'tat_running_rows': tat_running_rows,
        'tat_running_total': tat_running_total,
        'tat_running_limit': tat_running_limit,
        'tat_years': tat_years,
        'tat_status_choices': ["Ongoing", "Released", "Overdue"],
        'tat_q': tat_q,
        'tat_status': tat_status,
        'tat_year': tat_year,
        'tat_status_counts': tat_status_counts,
        'dashboard_year_label': "All Years" if tat_year == "all" else tat_year,
        'concordance_kpis': concordance_kpis,
        'concordance_kpis_json': json.dumps([concordance_kpis]),
        'concordance_kpi_year_rows_json': json.dumps(concordance_kpi_year_rows),
        'site_concordance_rows': site_concordance_rows,
        'site_concordance_rows_json': json.dumps(site_concordance_rows),
        'site_concordance_year_rows_json': json.dumps(site_concordance_year_rows),
        'monthly_referral_rows': monthly_referral_rows,
        'monthly_referral_rows_json': json.dumps(monthly_referral_rows),
        'surveillance_snapshot': surveillance_snapshot,
        'surveillance_snapshot_json': json.dumps([surveillance_snapshot]),
        'surveillance_snapshot_year_rows_json': json.dumps(surveillance_snapshot_year_rows),
        'dashboard_all_years': tat_year == "all",
    }

    return render(request, 'home/index.html', context)




@login_required(login_url="login")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:
        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:
        # Redirect to a different view or render a different template
        return redirect('home')  # Redirect to the home view or any other view

    except Exception as e:
        # Log the exception if needed
        print(f"Error: {e}")
        # Redirect to a different view or render a different template
        return redirect('home')  # Redirect to the home view or any other view





#activate bat_seq
@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
def batch_create_view(request):
    """
    Creates a batch:
    - Re-links Referred_Data
    - Assigns clean auto sequence (bat_seq) per batch
    """

    if request.method == "POST":
        form = BatchTable_form(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)

            # extract values
            site_code = (instance.bat_SiteCode or "").strip()
            referral_date = instance.bat_Referral_Date
            ref_no_raw = (instance.bat_RefNo or "").strip()
            batch_no = (instance.bat_BatchNo or "").strip()
            total_batch = (instance.bat_Total_batch or "").strip()
            site_name = (instance.bat_Site_NameGen or "").strip()

            if not (site_code and referral_date and ref_no_raw):
                messages.error(request, "Missing required fields.")
                return redirect("batch_create_view")

            # generate accession numbers
            try:
                year_short = referral_date.strftime("%y")
                year_long = referral_date.strftime("%m%d%Y")

                if "-" in ref_no_raw:
                    start_ref, end_ref = map(int, ref_no_raw.split("-"))
                    if start_ref > end_ref:
                        start_ref, end_ref = end_ref, start_ref
                else:
                    start_ref = end_ref = int(ref_no_raw)

            except ValueError:
                messages.error(request, "Invalid Ref No format.")
                return redirect("batch_create_view")

            accession_numbers = [
                f"{year_short}ARS_{site_code}{str(ref).zfill(4)}"
                for ref in range(start_ref, end_ref + 1)
            ]

           # generate batch code and name
            batch_codegen = f"{site_code}_{year_long}_{batch_no}.{total_batch}_{ref_no_raw}"
            auto_batch_name = batch_codegen

            if not site_name:
                site_obj = SiteData.objects.filter(SiteCode=site_code).first()
                if site_obj:
                    site_name = site_obj.SiteName

            orphan_batch = None

          # Do not overwrite existing batches from the create form.
            existing_batch = Batch_Table.objects.filter(
                bat_Batch_Code=batch_codegen
            ).first()
            if existing_batch:
                raw_count = Referred_Data.objects.filter(
                    Q(Batch_Code=existing_batch.bat_Batch_Code) |
                    Q(Batch_id=existing_batch)
                ).count()
                final_count = Final_Data.objects.filter(
                    Q(f_Batch_Code=existing_batch.bat_Batch_Code) |
                    Q(f_Batch_id=existing_batch)
                ).count()
                if raw_count == 0 and final_count == 0:
                    orphan_batch = existing_batch
                    messages.warning(
                        request,
                        f"Recovered orphan batch shell '{existing_batch.bat_Batch_Name}'. Raw accession record(s) will be recreated."
                    )
                else:
                    query_string = urlencode({"q": existing_batch.bat_Batch_Code})
                    raw_url = f"{reverse('show_data')}?{query_string}"
                    final_url = f"{reverse('show_final_table')}?{query_string}"
                    messages.error(
                        request,
                        format_html(
                            "Batch '<strong>{}</strong>' already exists in Raw Batches "
                            "({} raw record(s), {} final record(s)). Batch creation will not overwrite existing data. "
                            "<a href=\"{}\">View Raw Batch</a> | <a href=\"{}\">Search Final Records</a>. "
                            "Use Copy to Final or Copy Batch to Final to overwrite Final records.",
                            existing_batch.bat_Batch_Name,
                            raw_count,
                            final_count,
                            raw_url,
                            final_url,
                        )
                    )
                    return render(request, "home/Batchname_form.html", {"form": form, "editing": False})

            accession_conflicts = (
                Referred_Data.objects
                .filter(AccessionNo__in=accession_numbers)
                .exclude(Q(Batch_id__isnull=True) & Q(Batch_Code=""))
                .exclude(Q(Batch_id=orphan_batch))
                .values("AccessionNo", "Batch_Code", "Batch_id__bat_Batch_Code")
            )
            if accession_conflicts.exists():
                conflict_list = ", ".join(
                    f"{row['AccessionNo']} ({row['Batch_Code'] or row['Batch_id__bat_Batch_Code'] or 'existing batch'})"
                    for row in accession_conflicts[:5]
                )
                messages.error(
                    request,
                    f"Cannot create batch because accession(s) already belong to another batch: {conflict_list}."
                )
                return render(request, "home/Batchname_form.html", {"form": form, "editing": False})

           # start atomic transaction
            with transaction.atomic():

                batch_defaults = {
                    "created_by": request.user,
                    "bat_Batch_Name": auto_batch_name,
                    "bat_AccessionNo": _batch_accession_summary(accession_numbers),
                    "bat_AccessionNoGen": _batch_accession_summary(accession_numbers),
                    "bat_Batch_Code": batch_codegen,
                    "bat_Site_Name": site_name,
                    "bat_SiteCode": site_code,
                    "bat_Referral_Date": referral_date,
                    "bat_RefNo": ref_no_raw,
                    "bat_BatchNo": batch_no,
                    "bat_Total_batch": total_batch,
                    "bat_Encoder": instance.bat_Encoder or "",
                    "bat_Enc_Lic": instance.bat_Enc_Lic or "",
                    "bat_Checker": instance.bat_Checker or "",
                    "bat_Chec_Lic": instance.bat_Chec_Lic or "",
                    "bat_Verifier": instance.bat_Verifier or "",
                    "bat_Ver_Lic": instance.bat_Ver_Lic or "",
                    "bat_LabManager": instance.bat_LabManager or "",
                    "bat_Lab_Lic": instance.bat_Lab_Lic or "",
                    "bat_Head": instance.bat_Head or "",
                    "bat_Head_Lic": instance.bat_Head_Lic or "",
                }

                if orphan_batch:
                    batch_obj = orphan_batch
                    for field, value in batch_defaults.items():
                        setattr(batch_obj, field, value)
                    batch_obj.save(update_fields=list(batch_defaults.keys()))
                else:
                    batch_obj = Batch_Table.objects.create(**batch_defaults)

                synced_count, _ = _sync_batch_membership(batch_obj, accession_numbers)

                tat_obj, _ = TATform.objects.get_or_create(
                    tat_Batch_Isolates=batch_obj,
                    defaults={
                        "tat_SiteCode": batch_obj.bat_SiteCode,
                        "tat_Batch_Code": batch_obj.bat_Batch_Code,
                        "tat_Referral_Date": batch_obj.bat_Referral_Date,
                        "tat_Num_Isolate": synced_count,
                        "tat_BatchNumber": batch_obj.bat_BatchNo,
                        "tat_Total_Batch": batch_obj.bat_Total_batch,
                    },
                )
                tat_obj.tat_SiteCode = batch_obj.bat_SiteCode
                tat_obj.tat_Batch_Code = batch_obj.bat_Batch_Code
                tat_obj.tat_Referral_Date = batch_obj.bat_Referral_Date
                tat_obj.tat_Num_Isolate = synced_count
                tat_obj.tat_BatchNumber = batch_obj.bat_BatchNo
                tat_obj.tat_Total_Batch = batch_obj.bat_Total_batch
                tat_obj.save(update_fields=[
                    "tat_SiteCode",
                    "tat_Batch_Code",
                    "tat_Referral_Date",
                    "tat_Num_Isolate",
                    "tat_BatchNumber",
                    "tat_Total_Batch",
                ])
                    
            messages.success(
                request,
                f"Batch '{auto_batch_name}' saved with {synced_count} isolates."
            )

            return redirect(
                f"{reverse('show_batches')}?batch_code={batch_obj.bat_Batch_Code}"
            )

        else:
            messages.error(request, "Batch creation failed.")
    else:
        form = BatchTable_form()

    return render(request, "home/Batchname_form.html", {
    "form": form,
    "editing": False,
})




def _parse_batch_ref_range(ref_no_raw):
    ref_no_raw = (ref_no_raw or "").strip()
    if not ref_no_raw:
        raise ValueError("Missing reference number.")

    parts = [part.strip() for part in ref_no_raw.split("-", 1)]
    if len(parts) == 1:
        start_raw = end_raw = parts[0]
    else:
        start_raw, end_raw = parts

    if not (start_raw.isdigit() and end_raw.isdigit()):
        raise ValueError("Invalid Ref No format.")

    start_ref = int(start_raw)
    end_ref = int(end_raw)
    if start_ref > end_ref:
        start_ref, end_ref = end_ref, start_ref

    width = max(len(start_raw), len(end_raw), 4)
    return start_ref, end_ref, width


def _ref_no_from_batch_code(batch_code):
    batch_code = (batch_code or "").strip()
    if "_" not in batch_code:
        return ""
    return batch_code.rsplit("_", 1)[-1].strip()


def _batch_ref_no(batch):
    if not batch:
        return ""
    return (batch.bat_RefNo or "").strip() or _ref_no_from_batch_code(batch.bat_Batch_Code)


def _isolate_ref_no(isolate):
    return (
        (getattr(isolate, "RefNo", None) or "").strip()
        or _batch_ref_no(getattr(isolate, "Batch_id", None))
        or _ref_no_from_batch_code(getattr(isolate, "Batch_Code", ""))
    )


def _build_batch_accessions(site_code, referral_date, ref_no_raw):
    start_ref, end_ref, width = _parse_batch_ref_range(ref_no_raw)
    year_short = referral_date.strftime("%y")
    return [
        f"{year_short}ARS_{site_code}{str(ref).zfill(width)}"
        for ref in range(start_ref, end_ref + 1)
    ]


def _build_batch_code(site_code, referral_date, batch_no, total_batch, ref_no_raw):
    year_long = referral_date.strftime("%m%d%Y")
    return f"{site_code}_{year_long}_{batch_no}.{total_batch}_{ref_no_raw}"


def _batch_accession_summary(accession_numbers):
    if not accession_numbers:
        return ""
    if len(accession_numbers) == 1:
        return accession_numbers[0]
    return f"{accession_numbers[0]} - {accession_numbers[-1]}"


def _batch_personnel_defaults(batch):
    return {
        "arsp_Encoder": batch.bat_Encoder or "",
        "arsp_Enc_Lic": batch.bat_Enc_Lic or "",
        "arsp_Checker": batch.bat_Checker or "",
        "arsp_Chec_Lic": batch.bat_Chec_Lic or "",
        "arsp_Verifier": batch.bat_Verifier or "",
        "arsp_Ver_Lic": batch.bat_Ver_Lic or "",
        "arsp_LabManager": batch.bat_LabManager or "",
        "arsp_Lab_Lic": batch.bat_Lab_Lic or "",
        "arsp_Head": batch.bat_Head or "",
        "arsp_Head_Lic": batch.bat_Head_Lic or "",
        "Date_Accomplished_ARSP": batch.bat_Date_Accomplished,
    }


def _default_signature_staff(default_field):
    return arsStaff_Details.objects.filter(**{default_field: True}).order_by("Staff_Name").first()


def _apply_signature_defaults_to_batch(batch):
    changed_fields = []
    default_lab_manager = _default_signature_staff("Is_Default_Lab_Manager")
    if default_lab_manager and not (batch.bat_LabManager or "").strip():
        batch.bat_LabManager = default_lab_manager.Staff_Name or ""
        batch.bat_Lab_Lic = default_lab_manager.Staff_License or ""
        changed_fields.extend(["bat_LabManager", "bat_Lab_Lic"])
    elif default_lab_manager and not (batch.bat_Lab_Lic or "").strip() and batch.bat_LabManager == default_lab_manager.Staff_Name:
        batch.bat_Lab_Lic = default_lab_manager.Staff_License or ""
        changed_fields.append("bat_Lab_Lic")

    default_head = _default_signature_staff("Is_Default_Head")
    if default_head and not (batch.bat_Head or "").strip():
        batch.bat_Head = default_head.Staff_Name or ""
        batch.bat_Head_Lic = default_head.Staff_License or ""
        changed_fields.extend(["bat_Head", "bat_Head_Lic"])
    elif default_head and not (batch.bat_Head_Lic or "").strip() and batch.bat_Head == default_head.Staff_Name:
        batch.bat_Head_Lic = default_head.Staff_License or ""
        changed_fields.append("bat_Head_Lic")

    return list(dict.fromkeys(changed_fields))


def _sync_batch_membership(batch, accession_numbers):
    new_accessions = list(dict.fromkeys(accession_numbers))
    ref_no = _batch_ref_no(batch)
    modified_at = timezone.now()

    removed_isolates = list(
        Referred_Data.objects
        .filter(Batch_id=batch)
        .exclude(AccessionNo__in=new_accessions)
    )
    removed_accessions = [iso.AccessionNo for iso in removed_isolates]

    if removed_isolates:
        Referred_Data.objects.filter(pk__in=[iso.pk for iso in removed_isolates]).update(
            Batch_id=None,
            bat_seq=None,
            Batch_Code="",
            Batch_Name="",
            RefNo="",
            BatchNo="",
            Total_batch="",
            Date_Modified=modified_at,
        )
        Final_Data.objects.filter(
            f_Batch_id=batch,
            f_AccessionNo__in=removed_accessions,
        ).update(
            f_Batch_id=None,
            f_bat_seq=None,
            f_Batch_Code="",
            f_Batch_Name="",
            f_RefNo="",
            f_BatchNo="",
            f_Total_batch="",
            f_Date_Modified=modified_at,
        )

    personnel_defaults = _batch_personnel_defaults(batch)
    for acc in new_accessions:
        Referred_Data.objects.get_or_create(AccessionNo=acc)
        Referred_Data.objects.filter(AccessionNo=acc).update(
            bat_seq=accession_ref_sequence(acc),
            Batch_id_id=batch.id,
            Batch_Code=batch.bat_Batch_Code,
            Referral_Date=batch.bat_Referral_Date,
            RefNo=ref_no,
            BatchNo=batch.bat_BatchNo,
            Total_batch=batch.bat_Total_batch,
            SiteCode=batch.bat_SiteCode,
            Site_Name=batch.bat_Site_Name,
            Batch_Name=batch.bat_Batch_Name,
            Date_Modified=modified_at,
            **personnel_defaults,
        )

    active_isolates = Referred_Data.objects.filter(
        Batch_id=batch,
        AccessionNo__in=new_accessions,
    )
    for iso in active_isolates:
        Final_Data.objects.filter(f_AccessionNo=iso.AccessionNo).update(
            f_bat_seq=iso.bat_seq,
            f_Batch_id_id=batch.id,
            f_Batch_Code=iso.Batch_Code,
            f_Batch_Name=iso.Batch_Name,
            f_RefNo=_isolate_ref_no(iso),
            f_BatchNo=iso.BatchNo,
            f_Total_batch=iso.Total_batch,
            f_SiteCode=iso.SiteCode,
            f_Site_Name=iso.Site_Name,
            f_Referral_Date=iso.Referral_Date,
            f_arsp_Encoder=iso.arsp_Encoder,
            f_arsp_Enc_Lic=iso.arsp_Enc_Lic,
            f_arsp_Checker=iso.arsp_Checker,
            f_arsp_Chec_Lic=iso.arsp_Chec_Lic,
            f_arsp_Verifier=iso.arsp_Verifier,
            f_arsp_Ver_Lic=iso.arsp_Ver_Lic,
            f_arsp_LabManager=iso.arsp_LabManager,
            f_arsp_Lab_Lic=iso.arsp_Lab_Lic,
            f_arsp_Head=iso.arsp_Head,
            f_arsp_Head_Lic=iso.arsp_Head_Lic,
            f_Date_Accomplished_ARSP=iso.Date_Accomplished_ARSP,
            f_Date_Modified=modified_at,
        )

    return len(new_accessions), len(removed_accessions)


## edit the batch
@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@transaction.atomic
def batch_edit_view(request, pk):

    batch = get_object_or_404(Batch_Table, pk=pk)
    if not can_manage_batch(request.user, batch):
        messages.error(request, "You can only update batches that you created.")
        return redirect("show_data")

    if request.method == "POST":
        prior_ref_no = _batch_ref_no(batch)
        form = BatchEditForm(request.POST, instance=batch)

        if form.is_valid():
            batch = form.save(commit=False)
            _apply_signature_defaults_to_batch(batch)

            site_code = str(batch.bat_SiteCode or "").strip()
            referral_date = batch.bat_Referral_Date
            ref_no_raw = str(batch.bat_RefNo or "").strip() or prior_ref_no
            batch_no = str(batch.bat_BatchNo or "").strip()
            total_batch = str(batch.bat_Total_batch or "").strip()
            site_name = str(batch.bat_Site_NameGen or "").strip()

            if not site_name:
                site_obj = SiteData.objects.filter(SiteCode=site_code).first()
                site_name = site_obj.SiteName if site_obj else batch.bat_Site_Name

            if not (site_code and referral_date and ref_no_raw and batch_no and total_batch):
                messages.error(request, "Site code, referral date, batch no, total batch, and reference no are required.")
                return redirect("batch_edit_view", pk=batch.pk)

            try:
                accession_numbers = _build_batch_accessions(
                    site_code,
                    referral_date,
                    ref_no_raw,
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("batch_edit_view", pk=batch.pk)

            batch_codegen = _build_batch_code(
                site_code,
                referral_date,
                batch_no,
                total_batch,
                ref_no_raw,
            )
            duplicate_batch = (
                Batch_Table.objects
                .filter(bat_Batch_Code=batch_codegen)
                .exclude(pk=batch.pk)
                .first()
            )
            if duplicate_batch:
                messages.error(request, f"Batch '{batch_codegen}' already exists. Please use another reference number or batch number.")
                return redirect("batch_edit_view", pk=batch.pk)

            batch.bat_SiteCode = site_code
            batch.bat_Site_Name = site_name
            batch.bat_Site_NameGen = site_name
            batch.bat_RefNo = ref_no_raw
            batch.bat_BatchNo = batch_no
            batch.bat_Total_batch = total_batch
            batch.bat_AccessionNo = _batch_accession_summary(accession_numbers)
            batch.bat_AccessionNoGen = _batch_accession_summary(accession_numbers)
            batch.bat_Batch_Code = batch_codegen
            batch.bat_Batch_Name = batch_codegen
            batch.save()

            total_isolates, removed_count = _sync_batch_membership(
                batch,
                accession_numbers,
            )

            tat_obj, _ = TATform.objects.get_or_create(
                tat_Batch_Isolates=batch,
                defaults={
                    "tat_SiteCode": batch.bat_SiteCode or "",
                    "tat_Batch_Code": batch.bat_Batch_Code or "",
                    "tat_Referral_Date": batch.bat_Referral_Date,
                    "tat_Num_Isolate": total_isolates,
                    "tat_BatchNumber": batch.bat_BatchNo or "",
                    "tat_Total_Batch": batch.bat_Total_batch or "",
                }
            )
            tat_obj.tat_SiteCode = batch.bat_SiteCode or tat_obj.tat_SiteCode
            tat_obj.tat_Batch_Code = batch.bat_Batch_Code or tat_obj.tat_Batch_Code
            tat_obj.tat_Referral_Date = batch.bat_Referral_Date or tat_obj.tat_Referral_Date
            tat_obj.tat_Num_Isolate = total_isolates
            tat_obj.tat_BatchNumber = batch.bat_BatchNo or tat_obj.tat_BatchNumber
            tat_obj.tat_Total_Batch = batch.bat_Total_batch or tat_obj.tat_Total_Batch
            tat_obj.tat_Date_Released = batch.bat_Date_Accomplished
            tat_obj.save()

            messages.success(
                request,
                f"Batch updated successfully with {total_isolates} accessions. Removed {removed_count} old accessions from this batch."
            )
            return redirect("show_data")

    else:
        form = BatchEditForm(instance=batch)

    return render(
        request,
        "home/Batchname_author.html",
        {
            "form": form,
            "editing": True,
            "batch": batch,
        }
    )



@login_required(login_url="login")
def show_batches(request):
    batch_code = request.GET.get("batch_code")

    if not batch_code:
        last_batch = Batch_Table.objects.order_by("-id").first()
        batch_code = last_batch.bat_Batch_Code if last_batch else None

    isolates = (
        Referred_Data.objects
        .filter(Batch_Code=batch_code)
        .exclude(bat_seq__isnull=True)   
        .order_by("bat_seq")             
        .prefetch_related("antibiotic_entries")
    )

    paginator = Paginator(isolates, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    page_groups = {}
    for isolate in page_obj.object_list:
        batch_code = (isolate.Batch_Code or "Unbatched").strip() or "Unbatched"
        if batch_code not in page_groups:
            page_groups[batch_code] = {
                "code": batch_code,
                "records": [],
            }
        page_groups[batch_code]["records"].append(isolate)

    batch = Batch_Table.objects.filter(bat_Batch_Code=batch_code).first()

    return render(request, "home/Batch_isolates.html", {
        "page_obj": page_obj,
        "batch_code": batch_code,
        "batch": batch,
    })
# @login_required(login_url="login")
# def review_batches(request):

#     isolate_qs = (
#         Referred_Data.objects
#         .only(
#             "id",
#             "AccessionNo",
#             "SiteCode",
#             "Patient_ID",
#             "Site_Org",
#             "Spec_Type",
#             "Batch_id",
#         )
#         .exclude(
#             Q(Site_Org__isnull=False) & ~Q(Site_Org="")
#             |
#             Q(Spec_Type__isnull=False)
#         )
#     )

#     batches = (
#         Batch_Table.objects.all()
#         .order_by("-bat_Referral_Date")
#         .prefetch_related(
#             Prefetch(
#                 "Batch_isolates",
#                 queryset=isolate_qs,
#                 to_attr="prefetched_isolates",
#             )
#         )
#     )

#     total_accessions_all = 0

#     #  PRELOAD Final_Data batch IDs
#     final_batch_ids = set(
#         Final_Data.objects
#         .values_list("f_Batch_id", flat=True)
#         .distinct()
#     )

#     for batch in batches:
#         isolates = getattr(batch, "prefetched_isolates", [])
#         batch.display_isolates = isolates
#         batch.total_isolates = len(isolates)
#         total_accessions_all += len(isolates)

#         #  THIS IS THE KEY FLAG
#         batch.is_copied_to_final = batch.id in final_batch_ids

#     return render(
#         request,
#         "home/review_batches.html",
#         {
#             "batches": batches,
#             "total_batches": batches.count(),
#             "total_accessions_all": total_accessions_all,
#         },
#     )




@login_required(login_url="login")
def review_batches(request):

    q = request.GET.get("q", "").strip()

    isolate_qs = (
        Referred_Data.objects
        .only(
            "id",
            "AccessionNo",
            "Batch_id",
        )
        .exclude(
            Q(Site_Org__isnull=False) & ~Q(Site_Org="")
            |
            Q(Spec_Type__isnull=False)
        )
    )

    # Accession-only filter
    if q:
        isolate_qs = isolate_qs.filter(
            AccessionNo__icontains=q
        )

    batches = Batch_Table.objects.all().order_by("-bat_Referral_Date")

    #  Batch name filter
    if q:
        batches = batches.filter(
            bat_Batch_Code__icontains=q
        )

    batches = batches.prefetch_related(
        Prefetch(
            "Batch_isolates",
            queryset=isolate_qs,
            to_attr="prefetched_isolates",
        )
    )

    total_accessions_all = 0

    final_batch_ids = set(
        Final_Data.objects.values_list("f_Batch_id", flat=True).distinct()
    )

    for batch in batches:
        isolates = getattr(batch, "prefetched_isolates", [])
        batch.display_isolates = isolates
        batch.total_isolates = len(isolates)
        total_accessions_all += len(isolates)
        batch.is_copied_to_final = batch.id in final_batch_ids

    return render(
        request,
        "home/review_batches.html",
        {
            "batches": batches,
            "total_batches": batches.count(),
            "total_accessions_all": total_accessions_all,
            "q": q,
        },
    )






@login_required(login_url="login")
def clean_batch(request, batch_id):
    """
    Deletes a batch and all related Referred_Data records.
    """
    batch = get_object_or_404(Batch_Table, pk=batch_id)
    
    # Delete related isolates manually
    Referred_Data.objects.filter(Batch_Code=batch.bat_Batch_Code).delete()
    
    # Delete the batch itself
    batch.delete()
    
    messages.success(request, f"Batch '{batch.bat_Batch_Name}' and all related isolates have been deleted.")
    
    return redirect('review_batches')




@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@transaction.atomic
def delete_batch(request, batch_id):
    """
    Deletes a batch and all related Referred_Data and Final_Data records.
    """

    batch = get_object_or_404(Batch_Table, pk=batch_id)
    if not can_manage_batch(request.user, batch):
        messages.error(request, "You can only delete batches that you created.")
        return redirect("show_data")

    Referred_Data.objects.filter(
        Batch_Code=batch.bat_Batch_Code
    ).delete()

    Final_Data.objects.filter(
        f_Batch_Code=batch.bat_Batch_Code
    ).delete()

    batch.delete()

    messages.success(
        request,
        f"Batch '{batch.bat_Batch_Name}' and all related records have been deleted."
    )

    return redirect("show_batches")



@login_required(login_url="login")
@transaction.atomic
def delete_all_batches(request):
    """
    Deletes all batches and all related Referred_Data and Final_Data records.
    """

    batches = Batch_Table.objects.all()

    deleted_batches = batches.count()

    if deleted_batches == 0:
        messages.warning(request, "No batches found to delete.")
        return redirect("review_batches")

    # collect batch codes first
    batch_codes = list(
        batches.values_list("bat_Batch_Code", flat=True)
    )

    # delete related data
    Referred_Data.objects.filter(
        Batch_Code__in=batch_codes
    ).delete()

    Final_Data.objects.filter(
        f_Batch_Code__in=batch_codes
    ).delete()

    # delete the batches
    batches.delete()

    messages.success(
        request,
        f"All {deleted_batches} batches and related records have been deleted."
    )

    return redirect("review_batches")




@login_required(login_url="login")
@transaction.atomic
def delete_blank_batches(request):

    blank_isolates = (
        Referred_Data.objects
        .exclude(
            Q(Site_Org__isnull=False) & ~Q(Site_Org="")
            |
            Q(Spec_Type__isnull=False)
        )
    )

    batch_ids = list(
        blank_isolates.values_list("Batch_id", flat=True).distinct()
    )

    batches = Batch_Table.objects.filter(id__in=batch_ids)

    deleted_batches = batches.count()

    if deleted_batches == 0:
        messages.warning(request, "No blank batches found.")
        return redirect("review_batches")

    Referred_Data.objects.filter(Batch_id__in=batch_ids).delete()
    Final_Data.objects.filter(f_Batch_id__in=batch_ids).delete()
    batches.delete()

    messages.success(
        request,
        f"{deleted_batches} blank batch(es) deleted."
    )

    return redirect("review_batches")






@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@transaction.atomic
def delete_record_in_batch(request, id):

    isolate = get_object_or_404(Referred_Data, pk=id)
    if not can_manage_batch(request.user, isolate.Batch_id):
        messages.error(request, "You can only delete records from batches that you created.")
        return redirect("show_data")

    accession_no = isolate.AccessionNo

    Final_Data.objects.filter(
        f_AccessionNo=accession_no
    ).delete()

    isolate.delete()

    messages.success(
        request,
        f"Isolate '{accession_no}' and related final record have been deleted."
    )

    return redirect("show_batches")


############# filtering of antibiotics via ajax 
@login_required(login_url="login")
def get_antibiotic_name(request):
    whonet_code = request.GET.get("whonet")
    try:
        abx = Antibiotic_List.objects.get(Whonet_Abx=whonet_code)
        return JsonResponse({"name": abx.Antibiotic})
    except Antibiotic_List.DoesNotExist:
        return JsonResponse({"name": ""})



# used in breakopints dropdown for organism 
@require_GET
def get_organism_group(request):
    org_code = request.GET.get("org_code")

    if not org_code:
        return JsonResponse({"genus_group": ""})

    try:
        organism = Organism_List.objects.get(
            Whonet_Org_Code=org_code
        )
        return JsonResponse({
            "genus_group": organism.Genus_Group or ""
        })
    except Organism_List.DoesNotExist:
        return JsonResponse({"genus_group": ""})


## aut fill  abx code and tier
@require_GET
def get_antibiotic_details(request):
    whonet_abx = request.GET.get("whonet_abx", "").strip()
    
    # Use filter().first() to avoid DoesNotExist exceptions crashing the AJAX call
    abx = Antibiotic_List.objects.filter(Whonet_Abx=whonet_abx).first()
    
    if abx:
        return JsonResponse({
            "antibiotic": abx.Antibiotic,
            "abx_code": abx.Abx_code,
            "tier": abx.Tier,
        })
    
    return JsonResponse({"error": "Not found"}, status=404)


@login_required
def ajax_filter_antibiotics(request):
    isolate_id = request.GET.get("isolate_id")
    org_code   = request.GET.get("org", "").strip().lower()
    retest     = request.GET.get("retest") == "1"
    antibiotic_view = _antibiotic_view_mode(request)

    isolate = get_object_or_404(Referred_Data, pk=isolate_id)

   # determine breakpoint year
    specimen_year = isolate.Spec_Date.year if isolate.Spec_Date else None

    if specimen_year:
        breakpoint_year = (
            BreakpointsTable.objects
            .filter(Year__lte=specimen_year)
            .order_by("-Year")
            .values_list("Year", flat=True)
            .first()
        )
    else:
        breakpoint_year = (
            BreakpointsTable.objects
            .order_by("-Year")
            .values_list("Year", flat=True)
            .first()
        )


    
    # filter antibiotics
    antibiotics = get_filtered_antibiotics(
        breakpoint_year,
        org_code,     # ✅ ALWAYS CODE (matches raw_data + edit_data)
        retest=retest
    )

    # fetch existing entries
    entries = AntibioticEntry.objects.filter(
        ab_idNum_referred=isolate
    )
    existing_codes = entries.exclude(
        **{f"{'ab_Retest_Abx_code' if retest else 'ab_Abx_code'}__isnull": True}
    ).values_list("ab_Retest_Abx_code" if retest else "ab_Abx_code", flat=True)

    antibiotics = _raw_antibiotics_for_panel(
        org_code=org_code,
        specimen_year=specimen_year,
        show_site=not retest,
        retest=retest,
        existing_whonet_codes=existing_codes,
        antibiotic_view=antibiotic_view,
    )

    # Map entries by Whonet code
    entry_map = {}
    for e in entries:
        code = e.ab_Retest_Abx_code if retest else e.ab_Abx_code
        if code:
            entry_map[code.upper()] = e

    # build payload JSON
    payload = []
    seen_payload_codes = set()

    for abx in antibiotics:
        code = (abx.Whonet_Abx or "").strip().upper()
        method_key = "disk" if abx.Disk_Abx else "mic"
        payload_key = (code, method_key)
        if not code or payload_key in seen_payload_codes:
            continue
        seen_payload_codes.add(payload_key)
        entry = entry_map.get(code)

        if retest:
            payload.append({
                "whonet": code,
                "name": abx.Antibiotic,
                "is_disk": abx.Disk_Abx,

                # RETEST VALUES
                "disk": entry.ab_Retest_DiskValue if entry else "",
                "disk_enris": entry.ab_Retest_Disk_enRIS if entry else "",
                "mic": entry.ab_Retest_MICValue if entry else "",
                "mic_enris": entry.ab_Retest_MIC_enRIS if entry else "",
                "mic_operand": entry.ab_Retest_MIC_operand if entry else "",
                "alert_mic": entry.ab_Retest_AlertMIC if entry else False,
            })
        else:
            payload.append({
                "whonet": code,
                "name": abx.Antibiotic,
                "is_disk": abx.Disk_Abx,

                # MAIN VALUES
                "disk": entry.ab_Disk_value if entry else "",
                "disk_enris": entry.ab_Disk_enRIS if entry else "",
                "mic": entry.ab_MIC_value if entry else "",
                "mic_enris": entry.ab_MIC_enRIS if entry else "",
                "mic_operand": entry.ab_MIC_operand if entry else "",
                "alert_mic": entry.ab_AlertMIC if entry else False,
            })

    return JsonResponse({"antibiotics": payload})





################ Raw data view (final version with dynamic breakpoints)

# new version with auto filter antibiotics + breakpoint assignment
# aligns with new edit data view

@login_required(login_url="login")
@transaction.atomic
def raw_data(request, id):

   # Fetch the isolate
    isolates = get_object_or_404(Referred_Data, pk=id)
    antibiotic_view = _antibiotic_view_mode(request)

   # get the display form
    if request.method == "GET":

        form = Referred_Form(instance=isolates)
        existing_entries = AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates
        )
        specimen_year = isolates.Spec_Date.year if isolates.Spec_Date else None
        antibiotics_main = _raw_antibiotics_for_panel(
            org_code=isolates.Site_Org,
            specimen_year=specimen_year,
            show_site=True,
            existing_whonet_codes=existing_entries.exclude(ab_Abx_code__isnull=True).values_list("ab_Abx_code", flat=True),
            antibiotic_view=antibiotic_view,
        )
        antibiotics_retest = _raw_antibiotics_for_panel(
            org_code=isolates.ars_OrgCode,
            specimen_year=specimen_year,
            retest=True,
            require_org=True,
            existing_whonet_codes=existing_entries.exclude(ab_Retest_Abx_code__isnull=True).values_list("ab_Retest_Abx_code", flat=True),
            antibiotic_view=antibiotic_view,
        )

        return render(request, "home/Referred_form.html", {
            "form": form,
            "isolates": isolates,
            "batch_nav": _raw_batch_navigation(isolates),
            "antibiotics_main": antibiotics_main,
            "antibiotics_retest": antibiotics_retest,
            "existing_entries": existing_entries,
            "retest_entries": existing_entries,
            "edit_mode": True,
            "antibiotic_view": antibiotic_view,
        })

    # save form first before processing antibiotics
    original_phenotypes = {
        "Site_Pre": isolates.Site_Pre,
        "Site_Pos": isolates.Site_Pos,
        "ars_Pre": isolates.ars_Pre,
        "ars_Post": isolates.ars_Post,
    }
    form = Referred_Form(request.POST, instance=isolates)


    if not form.is_valid():
        messages.error(request, "Please check the highlighted fields.")
        existing_entries = AntibioticEntry.objects.filter(ab_idNum_referred=isolates)
        specimen_year = isolates.Spec_Date.year if isolates.Spec_Date else None
        return render(request, "home/Referred_form.html", {
            "form": form,
            "isolates": isolates,
            "batch_nav": _raw_batch_navigation(isolates),
            "antibiotics_main": _raw_antibiotics_for_panel(
                org_code=isolates.Site_Org,
                specimen_year=specimen_year,
                show_site=True,
                existing_whonet_codes=existing_entries.exclude(ab_Abx_code__isnull=True).values_list("ab_Abx_code", flat=True),
                antibiotic_view=antibiotic_view,
            ),
            "antibiotics_retest": _raw_antibiotics_for_panel(
                org_code=isolates.ars_OrgCode,
                specimen_year=specimen_year,
                retest=True,
                require_org=True,
                existing_whonet_codes=existing_entries.exclude(ab_Retest_Abx_code__isnull=True).values_list("ab_Retest_Abx_code", flat=True),
                antibiotic_view=antibiotic_view,
            ),
            "existing_entries": existing_entries,
            "retest_entries": existing_entries,
            "edit_mode": True,
            "antibiotic_view": antibiotic_view,
        })

    isolates = form.save(commit=False)
    for field_name, original_value in original_phenotypes.items():
        setattr(isolates, field_name, original_value)
    _apply_clear_phenotype_posts(isolates, request.POST)
    isolates.save()

    form = Referred_Form(request.POST, instance=isolates)



    specimen_year = isolates.Spec_Date.year if isolates.Spec_Date else None

    # resolve organism codes (do not lowercase)
    resolved_site_org = (isolates.Site_Org or "").strip()
    resolved_ars_org  = (isolates.ars_OrgCode or "").strip()
    site_org_is_na = _is_no_organism(resolved_site_org)
    ars_org_is_na = _is_no_organism(resolved_ars_org)
    resolve_bp = make_cached_breakpoint_resolver()

    existing_main_entries = {
        (entry.ab_Abx_code or "").strip().upper(): entry
        for entry in AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates,
            ab_Abx_code__isnull=False,
        )
    }
    existing_retest_entries = {
        (entry.ab_Retest_Abx_code or "").strip().upper(): entry
        for entry in AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates,
            ab_Retest_Abx_code__isnull=False,
        )
    }
    antibiotics_main = list(_raw_antibiotics_for_panel(
        org_code=resolved_site_org,
        specimen_year=specimen_year,
        show_site=True,
        existing_whonet_codes=existing_main_entries.keys(),
        antibiotic_view=antibiotic_view,
    ))
    antibiotics_retest = list(_raw_antibiotics_for_panel(
        org_code=resolved_ars_org,
        specimen_year=specimen_year,
        retest=True,
        require_org=True,
        existing_whonet_codes=existing_retest_entries.keys(),
        antibiotic_view=antibiotic_view,
    ))
    if site_org_is_na:
        for entry in AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates,
            ab_Abx_code__isnull=False,
        ):
            _clear_main_antibiotic_data(entry)
            _save_or_delete_antibiotic_entry(entry)
        antibiotics_main = []

    if ars_org_is_na:
        for entry in AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates,
            ab_Retest_Abx_code__isnull=False,
        ):
            _clear_retest_antibiotic_data(entry)
            _save_or_delete_antibiotic_entry(entry)
        antibiotics_retest = []

  ## FOR MAIN ANIBIOTICS
    for abx in antibiotics_main:

        abx_code = (abx.Whonet_Abx or "").strip().upper()

        disk_value  = request.POST.get(f"disk_{abx_code}")
        mic_value   = request.POST.get(f"mic_{abx_code}")
        disk_enris  = request.POST.get(f"disk_enris_{abx_code}", "").strip()
        mic_enris   = request.POST.get(f"mic_enris_{abx_code}", "").strip()
        mic_operand = request.POST.get(f"mic_operand_{abx_code}", "").strip()
        alert_mic   = f"alert_mic_{abx_code}" in request.POST

        disk_value = _parse_disk_value(disk_value)

        try:
            mic_value = float(mic_value) if mic_value else None
        except ValueError:
            mic_value = None

        # Save ONLY if something was entered
        if disk_value is None and mic_value is None:
            continue

        entry_defaults = {
            "ab_AccessionNo": isolates.AccessionNo,
            "ab_Antibiotic": abx.Antibiotic,
            "ab_Abx": abx.Abx_code,
            "ab_Disk_value": disk_value,
            "ab_Disk_enRIS": disk_enris,
            "ab_MIC_value": mic_value,
            "ab_MIC_enRIS": mic_enris,
            "ab_MIC_operand": mic_operand,
            "ab_AlertMIC": alert_mic,
        }
        existing_entry = existing_main_entries.get(abx_code)
        if existing_entry and _entry_values_match(existing_entry, entry_defaults):
            continue

        entry, _ = AntibioticEntry.objects.update_or_create(
            ab_idNum_referred=isolates,
            ab_Abx_code=abx_code,
            defaults=entry_defaults,
        )

        # ALWAYS reset breakpoints before re-applying
        entry.ab_breakpoints_id.clear()

        bp_applied = False

        # apply the disk breakpoints if existing
        if disk_value is not None:
            bp_disk = resolve_bp(
                abx_code,
                specimen_year,
                resolved_site_org,
                "DISK",
            )

            if bp_disk:
                entry.ab_breakpoints_id.add(bp_disk)  # for disk
                entry.ab_Site_Org = bp_disk.Org

                entry.ab_R_breakpoint   = bp_disk.R_val
                entry.ab_I_breakpoint   = bp_disk.I_val
                entry.ab_SDD_breakpoint = bp_disk.SDD_val
                entry.ab_S_breakpoint   = bp_disk.S_val
                bp_applied = True

        # apply the mic breakpoints if existing
        if mic_value is not None:
            bp_mic = resolve_bp(
                abx_code,
                specimen_year,
                resolved_site_org,
                "MIC",
            )

            if bp_mic:
                entry.ab_breakpoints_id.add(bp_mic)  # for mic
                entry.ab_Site_Org = bp_mic.Org
                entry.ab_R_breakpoint   = bp_mic.R_val
                entry.ab_I_breakpoint   = bp_mic.I_val
                entry.ab_SDD_breakpoint = bp_mic.SDD_val
                entry.ab_S_breakpoint   = bp_mic.S_val
                entry.ab_Alert_val = bp_mic.Alert_val if alert_mic else ""
                bp_applied = True
            
        if not bp_applied:
            entry.ab_Site_Org = None
            entry.ab_R_breakpoint = None
            entry.ab_I_breakpoint = None
            entry.ab_SDD_breakpoint = None
            entry.ab_S_breakpoint = None
            entry.ab_Alert_val = ""

        entry.save(update_fields=[
            "ab_AccessionNo",
            "ab_Site_Org",
            "ab_Disk_value",
            "ab_Disk_enRIS",
            "ab_MIC_value",
            "ab_MIC_enRIS",
            "ab_MIC_operand",
            "ab_AlertMIC",
            "ab_Alert_val",
            "ab_R_breakpoint",
            "ab_I_breakpoint",
            "ab_SDD_breakpoint",
            "ab_S_breakpoint",
        ])


    ## FOR RETEST ANTIBIOITCS
    for abx in antibiotics_retest:
        abx_code = (abx.Whonet_Abx or "").strip().upper()
        disk_value  = request.POST.get(f"retest_disk_{abx_code}")
        disk_enris  = request.POST.get(f"retest_disk_enris_{abx_code}", "").strip()
        mic_value   = request.POST.get(f"retest_mic_{abx_code}")
        mic_enris   = request.POST.get(f"retest_mic_enris_{abx_code}", "").strip()
        mic_operand = request.POST.get(f"retest_mic_operand_{abx_code}", "").strip()
        alert_mic   = f"retest_alert_mic_{abx_code}" in request.POST

        
        disk_value = _parse_disk_value(disk_value)


        try:
            mic_value = float(mic_value) if mic_value else None
        except ValueError:
            mic_value = None

        if disk_value is None and mic_value is None:
            continue

        entry_defaults = {
            "ab_Retest_Antibiotic": abx.Antibiotic,
            "ab_Retest_Abx": abx.Abx_code,
            "ab_Retest_DiskValue": disk_value,
            "ab_Retest_Disk_enRIS": disk_enris,
            "ab_Retest_MICValue": mic_value,
            "ab_Retest_MIC_enRIS": mic_enris,
            "ab_Retest_MIC_operand": mic_operand,
            "ab_Retest_AlertMIC": alert_mic,
        }
        existing_entry = existing_retest_entries.get(abx_code)
        if existing_entry and _entry_values_match(existing_entry, entry_defaults):
            continue

        entry, _ = AntibioticEntry.objects.update_or_create(
            ab_idNum_referred=isolates,
            ab_Retest_Abx_code=abx_code,
            defaults=entry_defaults,
        )

        # ALWAYS reset breakpoints before re-applying
        entry.ab_breakpoints_id.clear()

        ret_bp_applied = False
        # apply disk breakpoints for retest antibiotics if existing
        if disk_value is not None:
            bp_disk = resolve_bp(
                abx_code,
                specimen_year,
                resolved_ars_org,
                "DISK",
            )

            if bp_disk:
                entry.ab_breakpoints_id.add(bp_disk)  # for disk
                entry.ab_Ret_Org = bp_disk.Org
                entry.ab_Org_Flag = bp_disk.Emerging_Org_Flag
                entry.ab_Abx_Flag = bp_disk.Emerging_Abx_Flag
                entry.ab_Abx_Phenotype = bp_disk.Emerging_Pheno_Flag
                entry.ab_Abx_Phenotype_Other = bp_disk.Emerging_Pheno_Flag_Other
                entry.ab_Ret_R_breakpoint   = bp_disk.R_val
                entry.ab_Ret_I_breakpoint   = bp_disk.I_val
                entry.ab_Ret_SDD_breakpoint = bp_disk.SDD_val
                entry.ab_Ret_S_breakpoint   = bp_disk.S_val
                ret_bp_applied = True


        # apply mic breakpoints for retest antibiotics if existing
        if mic_value is not None:
            bp_mic = resolve_bp(
                abx_code,
                specimen_year,
                resolved_ars_org,
                "MIC",
            )

            if bp_mic:
                entry.ab_breakpoints_id.add(bp_mic)  # for mic
                entry.ab_Ret_Org = bp_mic.Org
                entry.ab_Org_Flag = bp_mic.Emerging_Org_Flag
                entry.ab_Abx_Flag = bp_mic.Emerging_Abx_Flag
                entry.ab_Abx_Phenotype = bp_mic.Emerging_Pheno_Flag
                entry.ab_Abx_Phenotype_Other = bp_mic.Emerging_Pheno_Flag_Other
                entry.ab_Ret_R_breakpoint   = bp_mic.R_val
                entry.ab_Ret_I_breakpoint   = bp_mic.I_val
                entry.ab_Ret_SDD_breakpoint = bp_mic.SDD_val
                entry.ab_Ret_S_breakpoint   = bp_mic.S_val
                entry.ab_Retest_Alert_val = bp_mic.Alert_val if alert_mic else ""
                ret_bp_applied = True
            
        if not ret_bp_applied:
            entry.ab_Ret_Org = None
            entry.ab_Org_Flag = False
            entry.ab_Abx_Flag = False
            entry.ab_Abx_Phenotype = None
            entry.ab_Abx_Phenotype_Other = None
            entry.ab_Ret_R_breakpoint   = None
            entry.ab_Ret_I_breakpoint   = None
            entry.ab_Ret_SDD_breakpoint = None
            entry.ab_Ret_S_breakpoint   = None
            entry.ab_Retest_Alert_val = ""


        entry.save(update_fields=[
            "ab_AccessionNo",
            "ab_Ret_Org",
            "ab_Org_Flag",
            "ab_Abx_Flag",
            "ab_Abx_Phenotype",
            "ab_Abx_Phenotype_Other",
            "ab_Retest_DiskValue",
            "ab_Retest_Disk_enRIS",
            "ab_Retest_MICValue",
            "ab_Retest_MIC_enRIS",
            "ab_Retest_MIC_operand",
            "ab_Retest_AlertMIC",
            "ab_Ret_R_breakpoint",
            "ab_Ret_I_breakpoint",
            "ab_Ret_SDD_breakpoint",
            "ab_Ret_S_breakpoint",
            "ab_Retest_Alert_val",
        ])

    messages.success(request, "Data saved successfully.")
    next_after_save = (request.POST.get("next_after_save") or "").strip()
    if next_after_save and url_has_allowed_host_and_scheme(
        next_after_save,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_after_save)

    batch_nav = _raw_batch_navigation(isolates)
    if batch_nav.get("next_id"):
        return redirect("raw_data", id=batch_nav["next_id"])
    return redirect("show_data")







################ Retrieve all raw data
# @login_required(login_url="login")
# def show_data(request):
    
#     query = request.GET.get("q", "")
#     sort_by = request.GET.get('sort', 'Date_of_Entry')  # Default sort field
#     order = request.GET.get('order', 'desc')  # Default sort order

#     sort_field = f"-{sort_by}" if order == 'desc' else sort_by

#     isolates = Referred_Data.objects.prefetch_related(
#         'antibiotic_entries'
#     ).order_by(sort_field)

#     if query:
#         isolates = isolates.filter(
#             Q(AccessionNo__icontains=query) |
#             Q(First_Name__icontains=query) |
#             Q(Last_Name__icontains=query) |
#             Q(Patient_ID__icontains=query) |
#             Q(Spec_Type__Specimen_code__icontains=query) |  # search in specimen code as well
#             Q(Spec_Type__Specimen_name__icontains=query) |  
#             Q(Site_Org__icontains=query) |
#             Q(Batch_Code__icontains=query) 
#         )

#     copied_ids = Final_Data.objects.values_list("f_AccessionNo", flat=True)

#     paginator = Paginator(isolates, 20)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)

#     context = {
#         'page_obj': page_obj,
#         'current_sort': sort_by,
#         'current_order': order,
#         'copied_ids': copied_ids,
#         'query': query,
#     }

#     return render(request, 'home/tables.html', context)




def _raw_records_base_queryset(request):
    records = Referred_Data.objects.all()
    if get_user_role(request.user) == ROLE_ENCODER:
        records = records.filter(Batch_id__created_by=request.user)
    return records


def _apply_raw_table_filters(records, query="", year=None):
    if query:
        records = records.filter(
            Q(AccessionNo__icontains=query) |
            Q(First_Name__icontains=query) |
            Q(Last_Name__icontains=query) |
            Q(Patient_ID__icontains=query) |
            Q(Spec_Type__Specimen_code__icontains=query) |
            Q(Spec_Type__Specimen_name__icontains=query) |
            Q(Site_Org__icontains=query) |
            Q(Batch_Code__icontains=query) |
            Q(Batch_id__bat_Batch_Code__icontains=query) |
            Q(Batch_id__bat_Batch_Name__icontains=query)
        )

    if year and str(year).isdigit():
        records = records.filter(Referral_Date__year=int(year))

    return records


def _raw_sort_field(sort_by, order):
    allowed_sort_fields = [
        "Referral_Date",
        "Date_of_Entry",
        "Date_Modified",
        "AccessionNo",
        "Last_Name",
        "First_Name",
        "Patient_ID",
        "Batch_Code",
        "SiteCode",
        "Site_Org",
        "Spec_Date",
        "Date_Admis",
        "Date_Birth",
        "Spec_Type__Specimen_code",
        "Spec_Type__Specimen_name",
        "bat_seq",
    ]
    if sort_by not in allowed_sort_fields:
        sort_by = "Date_Modified"
    if order not in ["asc", "desc"]:
        order = "desc"
    return sort_by, order, f"-{sort_by}" if order == "desc" else sort_by


# with year filters
@login_required(login_url="login")
def show_data(request):

    query = request.GET.get("q", "")
    year = request.GET.get("year", None)

    sort_by = request.GET.get("sort", "Date_Modified")
    order = request.GET.get("order", "desc")

    sort_by, order, sort_field = _raw_sort_field(sort_by, order)
    isolates = _apply_raw_table_filters(_raw_records_base_queryset(request), query, year)

    total_records = isolates.count()

    batch_sort_map = {
        "Batch_Code": "batch_code",
        "Referral_Date": "latest_referral_date",
        "Date_of_Entry": "latest_entry_date",
        "Date_Modified": "latest_modified",
        "Spec_Date": "latest_specimen_date",
        "SiteCode": "site_code",
    }
    batch_sort_field = batch_sort_map.get(sort_by, "Batch_Code")
    if order == "desc":
        batch_sort_field = f"-{batch_sort_field}"

    batch_summaries = (
        isolates
        .values("Batch_id")
        .annotate(
            record_count=Count("id"),
            batch_code=Max("Batch_id__bat_Batch_Code"),
            fallback_batch_code=Max("Batch_Code"),
            site_code=Max("SiteCode"),
            latest_referral_date=Max("Referral_Date"),
            latest_entry_date=Max("Date_of_Entry"),
            latest_modified=Max("Date_Modified"),
            latest_specimen_date=Max("Spec_Date"),
        )
        .order_by(batch_sort_field, "-latest_entry_date", "-batch_code")
    )

    paginator = Paginator(batch_summaries, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    page_batch_ids = [
        item.get("Batch_id")
        for item in page_obj.object_list
        if item.get("Batch_id")
    ]
    page_batch_codes = [
        (item.get("batch_code") or item.get("fallback_batch_code") or "").strip()
        for item in page_obj.object_list
        if (item.get("batch_code") or item.get("fallback_batch_code") or "").strip()
    ]
    tat_by_batch_id = {
        tat.tat_Batch_Isolates_id: {
            "days": tat.tat_Running_TAT,
            "pressure": _tat_pressure_for_days(tat.tat_Running_TAT, tat.tat_Target_Days),
        }
        for tat in TATform.objects.filter(tat_Batch_Isolates_id__in=page_batch_ids)
    }
    tat_by_batch_code = {
        tat.tat_Batch_Code.strip(): {
            "days": tat.tat_Running_TAT,
            "pressure": _tat_pressure_for_days(tat.tat_Running_TAT, tat.tat_Target_Days),
        }
        for tat in TATform.objects.filter(tat_Batch_Code__in=page_batch_codes)
        if tat.tat_Batch_Code
    }

    page_groups = []
    for batch_summary in page_obj.object_list:
        raw_batch_code = batch_summary.get("batch_code") or batch_summary.get("fallback_batch_code") or ""
        batch_code = raw_batch_code.strip() or "Unbatched"
        batch_id = batch_summary.get("Batch_id") or ""
        tat_summary = tat_by_batch_id.get(batch_id, tat_by_batch_code.get(batch_code, {}))
        page_groups.append({
            "batch_id": batch_id,
            "code": batch_code,
            "batch_code": raw_batch_code,
            "count": batch_summary["record_count"],
            "site_code": batch_summary.get("site_code") or "",
            "tat_days": tat_summary.get("days"),
            "tat_pressure": tat_summary.get("pressure", "none"),
        })

    # Get available years
    available_years = (
        Referred_Data.objects
        .annotate(year=ExtractYear("Referral_Date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )

    context = {
        "page_obj": page_obj,
        "page_groups": page_groups,
        "total_records": total_records,
        "current_sort": sort_by,
        "current_order": order,
        "query": query,
        "year": year,
        "available_years": available_years,
    }

    return render(request, "home/tables.html", context)


@login_required(login_url="login")
@require_GET
def raw_batch_rows(request):
    batch_id = request.GET.get("batch_id", "")
    batch_dom_id = request.GET.get("target", "")
    query = request.GET.get("q", "")
    year = request.GET.get("year")
    sort_by = request.GET.get("sort", "Date_Modified")
    order = request.GET.get("order", "desc")
    _, _, sort_field = _raw_sort_field(sort_by, order)

    records = _apply_raw_table_filters(_raw_records_base_queryset(request), query, year)
    if batch_id:
        records = records.filter(Batch_id_id=batch_id)
    else:
        records = records.filter(Batch_id__isnull=True)
    records = (
        records
        .select_related("Batch_id", "Spec_Type")
        .prefetch_related("antibiotic_entries")
        .order_by("bat_seq", "AccessionNo", "id")
    )
    copied_ids = set(
        Final_Data.objects
        .filter(f_AccessionNo__in=records.values_list("AccessionNo", flat=True))
        .values_list("f_AccessionNo", flat=True)
    )
    html = render_to_string(
        "home/partials/raw_batch_rows.html",
        {
            "records": records,
            "copied_ids": copied_ids,
            "batch_dom_id": batch_dom_id,
            "request": request,
        },
        request=request,
    )
    return JsonResponse({"html": html})





########### edit data view


### working and orgiginal version, but might run into a problem in the future -filtering of specimen year....

# @login_required(login_url="login")
# @transaction.atomic
# def edit_data(request, id):

#    # fetch the isolate
#     isolates = get_object_or_404(Referred_Data, pk=id)

#     # # determine breakpoint year based on specimen date
#     specimen_year = isolates.Spec_Date.year if isolates.Spec_Date else None

#     if specimen_year:
#         breakpoint_year = (
#             BreakpointsTable.objects
#             .filter(Year__lte=specimen_year)
#             .order_by("-Year")
#             .values_list("Year", flat=True)
#             .first()
#         )
#     else:
#         breakpoint_year = None

#     # helper to get breakpoint year
#     def get_effective_breakpoint_year():
#         if breakpoint_year:
#             return breakpoint_year
#         return (
#             BreakpointsTable.objects
#             .order_by("-Year")
#             .values_list("Year", flat=True)
#             .first()
#         )



#    # get the display form
#     if request.method == "GET":

#         form = Referred_Form(instance=isolates)

#         antibiotics_main = (
#             Antibiotic_List.objects
#             .filter(Show=True)
#             .order_by("Antibiotic")
#         )

#         antibiotics_retest = (
#             Antibiotic_List.objects
#             .filter(Retest=True)
#             .order_by("Antibiotic")
#         )

#         existing_entries = AntibioticEntry.objects.filter(
#             ab_idNum_referred=isolates
#         )

#         retest_entries = existing_entries.exclude(
#             ab_Retest_Abx_code__isnull=True
#         )

#         return render(request, "home/edit.html", {
#             "form": form,
#             "isolates": isolates,
#             "antibiotics_main": antibiotics_main,
#             "antibiotics_retest": antibiotics_retest,
#             "existing_entries": existing_entries,
#             "retest_entries": retest_entries,
#             "edit_mode": True,
#         })

#     # save form first before processing antibiotics
#     old_site_org = (isolates.Site_Org or "").strip()
#     old_ars_org  = (isolates.ars_OrgCode or "").strip()

#     form = Referred_Form(request.POST, instance=isolates)
#     if not form.is_valid():
#         messages.error(request, "Error: Saving unsuccessful")
#         return redirect("edit_data", id=id)

#     isolates = form.save()

#     new_site_org = (isolates.Site_Org or "").strip()
#     new_ars_org  = (isolates.ars_OrgCode or "").strip()

    
#     # resolve organism codes
#     resolved_site_org = (isolates.Site_Org or "").strip()
#     resolved_ars_org  = (isolates.ars_OrgCode or "").strip()




#     # delete existing antibiotic entries if organism code changed
#     if old_site_org != new_site_org:
#         AntibioticEntry.objects.filter(
#             ab_idNum_referred=isolates,
#             ab_Abx_code__isnull=False
#         ).delete()

#     if old_ars_org != new_ars_org:
#         AntibioticEntry.objects.filter(
#             ab_idNum_referred=isolates,
#             ab_Retest_Abx_code__isnull=False
#         ).delete()

#     effective_year = get_effective_breakpoint_year()

#    # save the MAIN ANTIBIOTICS (RAW DATA)
#     antibiotics_main = Antibiotic_List.objects.filter(Show=True)

#     for abx in antibiotics_main:
#         abx_code = (abx.Whonet_Abx or "").strip().upper()

#         disk_value  = request.POST.get(f"disk_{abx_code}")
#         mic_value   = request.POST.get(f"mic_{abx_code}")
#         disk_enris  = request.POST.get(f"disk_enris_{abx_code}", "").strip()
#         mic_enris   = request.POST.get(f"mic_enris_{abx_code}", "").strip()
#         mic_operand = request.POST.get(f"mic_operand_{abx_code}", "").strip()
#         alert_mic   = f"alert_mic_{abx_code}" in request.POST

#         try:
#             disk_value = int(disk_value) if disk_value else None
#         except ValueError:
#             disk_value = None

#         try:
#             mic_value = float(mic_value) if mic_value else None
#         except ValueError:
#             mic_value = None

#         if disk_value is None and mic_value is None:
#             continue

#         entry, _ = AntibioticEntry.objects.update_or_create(
#             ab_idNum_referred=isolates,
#             ab_Abx_code=abx_code,
#             defaults={
#                 "ab_AccessionNo": isolates.AccessionNo,
#                 "ab_Antibiotic": abx.Antibiotic,
#                 "ab_Abx": abx.Abx_code,
#                 "ab_Disk_value": disk_value,
#                 "ab_Disk_enRIS": disk_enris,
#                 "ab_MIC_value": mic_value,
#                 "ab_MIC_enRIS": mic_enris,
#                 "ab_MIC_operand": mic_operand,
#                 "ab_AlertMIC": alert_mic,
#             }
#         )


#         entry.ab_breakpoints_id.clear()  

#         bp_applied = False
      
#       # apply disk breakpoints
#         if disk_value is not None:
#             bp_disk = (
#                 BreakpointsTable.objects
#                 .filter(
#                     Antibiotic_list_id=abx_code,
#                     Year=effective_year,
#                     Test_Method="DISK"
#                 )
#                 .filter(
#                     Q(Org__iexact=resolved_site_org) |
#                     Q(Org='')
#                 )
#                 .order_by(
#                     Case(
#                         When(Org__iexact=resolved_site_org, then=0),
#                         When(Org='', then=1),
#                         default=2,
#                     )
#                 )
#                 .first()
#             )

#             if bp_disk:
#                 entry.ab_breakpoints_id.set([bp_disk])  # for disk
#                 entry.ab_Site_Org = bp_disk.Org
#                 entry.ab_R_breakpoint   = bp_disk.R_val
#                 entry.ab_I_breakpoint   = bp_disk.I_val
#                 entry.ab_SDD_breakpoint = bp_disk.SDD_val
#                 entry.ab_S_breakpoint   = bp_disk.S_val
#                 bp_applied = True
        

#        # apply mic breakpoints
#         if mic_value is not None:
#             bp_mic = (
#                 BreakpointsTable.objects
#                 .filter(
#                     Antibiotic_list_id=abx_code,
#                     Year=effective_year,
#                     Test_Method="MIC"
#                 )
#                 .filter(
#                     Q(Org__iexact=resolved_site_org) |
#                     Q(Org='')
#                 )
#                 .order_by(
#                     Case(
#                         When(Org__iexact=resolved_site_org, then=0),
#                         When(Org='', then=1),
#                         default=2,
#                     )
#                 )
#                 .first()
#             )

#             if bp_mic:
#                 entry.ab_breakpoints_id.set([bp_mic])   # for mic (overrides)
#                 entry.ab_Site_Org = bp_mic.Org
#                 entry.ab_R_breakpoint   = bp_mic.R_val
#                 entry.ab_I_breakpoint   = bp_mic.I_val
#                 entry.ab_SDD_breakpoint = bp_mic.SDD_val
#                 entry.ab_S_breakpoint   = bp_mic.S_val

#                 entry.ab_Alert_val = bp_mic.Alert_val if alert_mic else ""
#                 bp_applied = True
        
#         if not bp_applied:
#             entry.ab_Site_Org = None
#             entry.ab_R_breakpoint = None
#             entry.ab_I_breakpoint = None
#             entry.ab_SDD_breakpoint = None
#             entry.ab_S_breakpoint = None
#             entry.ab_Alert_val = ""


#         entry.save(update_fields=[
#             "ab_AccessionNo",
#             "ab_Site_Org",
#             "ab_Disk_value",
#             "ab_Disk_enRIS",
#             "ab_MIC_value",
#             "ab_MIC_enRIS",
#             "ab_MIC_operand",
#             "ab_AlertMIC",
#             "ab_Alert_val",
#             "ab_R_breakpoint",
#             "ab_I_breakpoint",
#             "ab_SDD_breakpoint",
#             "ab_S_breakpoint",
#         ])


#     # save the RETEST ANTIBIOTICS arsrl data
#     antibiotics_retest = Antibiotic_List.objects.filter(Retest=True)

#     for abx in antibiotics_retest:
#         abx_code = (abx.Whonet_Abx or "").strip().upper()

#         disk_value  = request.POST.get(f"retest_disk_{abx_code}")
#         disk_enris  = request.POST.get(f"retest_disk_enris_{abx_code}", "").strip()
       
#         try:
#             disk_value = int(disk_value) if disk_value else None
#         except ValueError:
#             disk_value = None

        
#         mic_value   = request.POST.get(f"retest_mic_{abx_code}")
#         mic_enris   = request.POST.get(f"retest_mic_enris_{abx_code}", "").strip()
#         mic_operand = request.POST.get(f"retest_mic_operand_{abx_code}", "").strip()
#         alert_mic   = f"retest_alert_mic_{abx_code}" in request.POST


        
#         try:
#             mic_value = float(mic_value) if mic_value else None
#         except ValueError:
#             mic_value = None


#         if disk_value is None and mic_value is None:
#             continue

#         entry, _ = AntibioticEntry.objects.update_or_create(
#             ab_idNum_referred=isolates,
#             ab_Retest_Abx_code=abx_code,
#             defaults={
#                 "ab_AccessionNo": isolates.AccessionNo,
#                 "ab_Retest_Antibiotic": abx.Antibiotic,
#                 "ab_Retest_Abx": abx.Abx_code,
#                 "ab_Retest_DiskValue": disk_value,
#                 "ab_Retest_Disk_enRIS": disk_enris,
#                 "ab_Retest_MICValue": mic_value,
#                 "ab_Retest_MIC_enRIS": mic_enris,
#                 "ab_Retest_MIC_operand": mic_operand,
#                 "ab_Retest_AlertMIC": alert_mic,
#             }
#         )

#         entry.ab_breakpoints_id.clear()  

#         ret_bp_applied = False
#        # apply retest disk breakpoints
#         if disk_value is not None:
#             bp_disk = (
#                 BreakpointsTable.objects
#                 .filter(
#                     Antibiotic_list_id=abx_code,
#                     Year=effective_year,
#                     Test_Method="DISK"
#                 )
#                 .filter(
#                     Q(Org__iexact=resolved_ars_org) |
#                     Q(Org='')
#                 )
#                 .order_by(
#                     Case(
#                         When(Org__iexact=resolved_ars_org, then=0),
#                         When(Org='', then=1),
#                         default=2,
#                     )
#                 )
#                 .first()
#             )

#             if bp_disk:
#                 entry.ab_breakpoints_id.set([bp_disk])  # for disk
#                 entry.ab_Ret_Org = bp_disk.Org
#                 entry.ab_Ret_R_breakpoint   = bp_disk.R_val
#                 entry.ab_Ret_I_breakpoint   = bp_disk.I_val
#                 entry.ab_Ret_SDD_breakpoint = bp_disk.SDD_val
#                 entry.ab_Ret_S_breakpoint   = bp_disk.S_val
#                 ret_bp_applied = True


#         # apply retest mic breakpoints
#         if mic_value is not None:
#             bp_mic = (
#                 BreakpointsTable.objects
#                 .filter(
#                     Antibiotic_list_id=abx_code,
#                     Year=effective_year,
#                     Test_Method="MIC"
#                 )
#                 .filter(
#                     Q(Org__iexact=resolved_ars_org) |
#                     Q(Org='')
#                 )
#                 .order_by(
#                     Case(
#                         When(Org__iexact=resolved_ars_org, then=0),
#                         When(Org='', then=1),
#                         default=2,
#                     )
#                 )
#                 .first()
#             )

#             if bp_mic:
#                 entry.ab_breakpoints_id.set([bp_mic])   # for mic (overrides)
#                 entry.ab_Ret_Org = bp_mic.Org
#                 entry.ab_Ret_R_breakpoint   = bp_mic.R_val
#                 entry.ab_Ret_I_breakpoint   = bp_mic.I_val
#                 entry.ab_Ret_SDD_breakpoint = bp_mic.SDD_val
#                 entry.ab_Ret_S_breakpoint   = bp_mic.S_val
#                 entry.ab_Retest_Alert_val = bp_mic.Alert_val if alert_mic else ""
#                 ret_bp_applied = True
        
#         if not ret_bp_applied:
#             entry.ab_Ret_Org = None
#             entry.ab_Ret_R_breakpoint = None
#             entry.ab_Ret_I_breakpoint = None
#             entry.ab_Ret_SDD_breakpoint = None
#             entry.ab_Ret_S_breakpoint = None
#             entry.ab_Retest_Alert_val = ""


#         entry.save(update_fields=[
#             "ab_AccessionNo",
#             "ab_Ret_Org",
#             "ab_Retest_DiskValue",
#             "ab_Retest_Disk_enRIS",
#             "ab_Retest_MICValue",
#             "ab_Retest_MIC_enRIS",
#             "ab_Retest_MIC_operand",
#             "ab_Retest_AlertMIC",
#             "ab_Ret_R_breakpoint",
#             "ab_Ret_I_breakpoint",
#             "ab_Ret_SDD_breakpoint",
#             "ab_Ret_S_breakpoint",
#             "ab_Retest_Alert_val",
#         ])

#     messages.success(request, "Data saved successfully.")
#     return redirect("show_data")



def _raw_batch_navigation(isolate):
    batch = isolate.Batch_id
    if not batch:
        return {
            "batch_id": None,
            "first_id": None,
            "previous_id": None,
            "next_id": None,
            "last_id": None,
            "position": 1,
            "total": 1,
        }

    isolate_ids = list(
        Referred_Data.objects
        .filter(Batch_id=batch)
        .order_by("bat_seq", "AccessionNo", "id")
        .values_list("id", flat=True)
    )

    try:
        current_index = isolate_ids.index(isolate.id)
    except ValueError:
        current_index = 0

    return {
        "batch_id": batch.id,
        "first_id": isolate_ids[0] if isolate_ids else None,
        "previous_id": isolate_ids[current_index - 1] if current_index > 0 else None,
        "next_id": isolate_ids[current_index + 1] if current_index + 1 < len(isolate_ids) else None,
        "last_id": isolate_ids[-1] if isolate_ids else None,
        "position": current_index + 1,
        "total": len(isolate_ids),
    }


### much safer error-proof version 

@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@transaction.atomic
def edit_data(request, id):

   # fetch the isolate
    isolates = get_object_or_404(Referred_Data, pk=id)
    if not can_manage_batch(request.user, isolates.Batch_id):
        messages.error(request, "You can only update records from batches that you created.")
        return redirect("show_data")
    antibiotic_view = _antibiotic_view_mode(request)


   # get the display form
    if request.method == "GET":

        form = Referred_Form(instance=isolates)
        existing_entries = AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates
        )
        specimen_year = isolates.Spec_Date.year if isolates.Spec_Date else None
        antibiotics_main = _raw_antibiotics_for_panel(
            org_code=isolates.Site_Org,
            specimen_year=specimen_year,
            show_site=True,
            existing_whonet_codes=existing_entries.exclude(ab_Abx_code__isnull=True).values_list("ab_Abx_code", flat=True),
            antibiotic_view=antibiotic_view,
        )
        antibiotics_retest = _raw_antibiotics_for_panel(
            org_code=isolates.ars_OrgCode,
            specimen_year=specimen_year,
            retest=True,
            require_org=True,
            existing_whonet_codes=existing_entries.exclude(ab_Retest_Abx_code__isnull=True).values_list("ab_Retest_Abx_code", flat=True),
            antibiotic_view=antibiotic_view,
        )

        retest_entries = existing_entries.exclude(
            ab_Retest_Abx_code__isnull=True
        )

        return render(request, "home/edit.html", {
            "form": form,
            "isolates": isolates,
            "batch_nav": _raw_batch_navigation(isolates),
            "antibiotics_main": antibiotics_main,
            "antibiotics_retest": antibiotics_retest,
            "existing_entries": existing_entries,
            "retest_entries": retest_entries,
            "edit_mode": True,
            "antibiotic_view": antibiotic_view,
        })

    # save form first before processing antibiotics
    old_site_org = (isolates.Site_Org or "").strip()
    old_ars_org  = (isolates.ars_OrgCode or "").strip()
    old_specimen_year = isolates.Spec_Date.year if isolates.Spec_Date else None
    original_phenotypes = {
        "Site_Pre": isolates.Site_Pre,
        "Site_Pos": isolates.Site_Pos,
        "ars_Pre": isolates.ars_Pre,
        "ars_Post": isolates.ars_Post,
    }

    post_data = request.POST.copy()
    if old_site_org and not (post_data.get("Site_Org") or "").strip():
        old_site_choice = resolve_organism_choice(old_site_org, isolates.Site_OrgName)
        post_data["Site_Org"] = (
            old_site_choice.Whonet_Org_Code if old_site_choice else old_site_org
        )
    if old_ars_org and not (post_data.get("ars_OrgCode") or "").strip():
        old_ars_choice = resolve_organism_choice(old_ars_org, isolates.ars_OrgName)
        post_data["ars_OrgCode"] = (
            old_ars_choice.Whonet_Org_Code if old_ars_choice else old_ars_org
        )
    if isolates.Site_OrgName and not (post_data.get("Site_OrgName") or "").strip():
        post_data["Site_OrgName"] = isolates.Site_OrgName
    if isolates.ars_OrgName and not (post_data.get("ars_OrgName") or "").strip():
        post_data["ars_OrgName"] = isolates.ars_OrgName

    form = Referred_Form(post_data, instance=isolates)
    if not form.is_valid():
        messages.error(request, "Please check the highlighted fields.")
        existing_entries = AntibioticEntry.objects.filter(ab_idNum_referred=isolates)
        specimen_year = isolates.Spec_Date.year if isolates.Spec_Date else None
        return render(request, "home/edit.html", {
            "form": form,
            "isolates": isolates,
            "batch_nav": _raw_batch_navigation(isolates),
            "antibiotics_main": _raw_antibiotics_for_panel(
                org_code=isolates.Site_Org,
                specimen_year=specimen_year,
                show_site=True,
                existing_whonet_codes=existing_entries.exclude(ab_Abx_code__isnull=True).values_list("ab_Abx_code", flat=True),
                antibiotic_view=antibiotic_view,
            ),
            "antibiotics_retest": _raw_antibiotics_for_panel(
                org_code=isolates.ars_OrgCode,
                specimen_year=specimen_year,
                retest=True,
                require_org=True,
                existing_whonet_codes=existing_entries.exclude(ab_Retest_Abx_code__isnull=True).values_list("ab_Retest_Abx_code", flat=True),
                antibiotic_view=antibiotic_view,
            ),
            "existing_entries": existing_entries,
            "retest_entries": existing_entries.exclude(ab_Retest_Abx_code__isnull=True),
            "edit_mode": True,
            "antibiotic_view": antibiotic_view,
        })

    isolates = form.save(commit=False)
    for field_name, original_value in original_phenotypes.items():
        setattr(isolates, field_name, original_value)
    _apply_clear_phenotype_posts(isolates, request.POST)
    isolates.save()


    specimen_year = isolates.Spec_Date.year if isolates.Spec_Date else None



    new_site_org = (isolates.Site_Org or "").strip()
    new_ars_org  = (isolates.ars_OrgCode or "").strip()
    site_org_is_na = _is_no_organism(new_site_org)
    ars_org_is_na = _is_no_organism(new_ars_org)

    
    # resolve organism codes
    resolved_site_org = (isolates.Site_Org or "").strip()
    resolved_ars_org  = (isolates.ars_OrgCode or "").strip()
    resolve_bp = make_cached_breakpoint_resolver()




    # delete existing antibiotic entries if organism code changed
    if old_site_org != new_site_org or site_org_is_na:
        for entry in AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates,
            ab_Abx_code__isnull=False
        ):
            _clear_main_antibiotic_data(entry)
            _save_or_delete_antibiotic_entry(entry)

    if old_ars_org != new_ars_org or ars_org_is_na:
        for entry in AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates,
            ab_Retest_Abx_code__isnull=False
        ):
            _clear_retest_antibiotic_data(entry)
            _save_or_delete_antibiotic_entry(entry)


   # save the MAIN ANTIBIOTICS (RAW DATA)
    existing_main_entries = {
        (entry.ab_Abx_code or "").strip().upper(): entry
        for entry in AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates,
            ab_Abx_code__isnull=False,
        )
    }
    existing_retest_entries_for_main = {
        (entry.ab_Retest_Abx_code or "").strip().upper(): entry
        for entry in AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates,
            ab_Retest_Abx_code__isnull=False,
        )
    }
    can_skip_main_entries = (
        old_site_org == new_site_org
        and old_specimen_year == specimen_year
    )
    antibiotics_main = list(_raw_antibiotics_for_panel(
        org_code=resolved_site_org,
        specimen_year=specimen_year,
        show_site=True,
        existing_whonet_codes=existing_main_entries.keys(),
        antibiotic_view=antibiotic_view,
    ))

    if not site_org_is_na:
      for abx in antibiotics_main:
        abx_code = (abx.Whonet_Abx or "").strip().upper()

        disk_value  = request.POST.get(f"disk_{abx_code}")
        mic_value   = request.POST.get(f"mic_{abx_code}")
        disk_enris  = request.POST.get(f"disk_enris_{abx_code}", "").strip()
        mic_enris   = request.POST.get(f"mic_enris_{abx_code}", "").strip()
        mic_operand = request.POST.get(f"mic_operand_{abx_code}", "").strip()
        alert_mic   = f"alert_mic_{abx_code}" in request.POST

        disk_value = _parse_disk_value(disk_value)

        try:
            mic_value = float(mic_value) if mic_value else None
        except ValueError:
            mic_value = None

        if disk_value is None and mic_value is None:
            for entry in AntibioticEntry.objects.filter(
                ab_idNum_referred=isolates,
                ab_Abx_code=abx_code
            ):
                _clear_main_antibiotic_data(entry)
                _save_or_delete_antibiotic_entry(entry)
            continue

        entry_defaults = {
            "ab_AccessionNo": isolates.AccessionNo,
            "ab_Antibiotic": abx.Antibiotic,
            "ab_Abx": abx.Abx_code,
            "ab_Disk_value": disk_value,
            "ab_Disk_enRIS": disk_enris,
            "ab_MIC_value": mic_value,
            "ab_MIC_enRIS": mic_enris,
            "ab_MIC_operand": mic_operand,
            "ab_AlertMIC": alert_mic,
        }
        existing_entry = existing_main_entries.get(abx_code)
        if (
            can_skip_main_entries
            and existing_entry
            and _entry_values_match(existing_entry, entry_defaults)
        ):
            continue

        entry = existing_entry or existing_retest_entries_for_main.get(abx_code)
        if entry is None:
            entry = AntibioticEntry(
                ab_idNum_referred=isolates,
                ab_Abx_code=abx_code,
            )
        entry.ab_Abx_code = abx_code
        for field, value in entry_defaults.items():
            setattr(entry, field, value)
        entry.save()


        entry.ab_breakpoints_id.clear()  

        bp_applied = False

      # Apply breakpoints
        if disk_value is not None:
            bp_disk = resolve_bp(
                abx_code,
                specimen_year,
                resolved_site_org,
                "DISK",
            )

            if bp_disk:
                entry.ab_breakpoints_id.set([bp_disk])  # for disk
                entry.ab_Site_Org = bp_disk.Org
                entry.ab_R_breakpoint   = bp_disk.R_val
                entry.ab_I_breakpoint   = bp_disk.I_val
                entry.ab_SDD_breakpoint = bp_disk.SDD_val
                entry.ab_S_breakpoint   = bp_disk.S_val
                bp_applied = True
        

        # apply mic breakpoints
        if mic_value is not None:
            bp_mic = resolve_bp(
                abx_code,
                specimen_year,
                resolved_site_org,
                "MIC",
            )

            if bp_mic:
                entry.ab_breakpoints_id.set([bp_mic])   # for mic (overrides)
                entry.ab_Site_Org = bp_mic.Org
                entry.ab_R_breakpoint   = bp_mic.R_val
                entry.ab_I_breakpoint   = bp_mic.I_val
                entry.ab_SDD_breakpoint = bp_mic.SDD_val
                entry.ab_S_breakpoint   = bp_mic.S_val

                entry.ab_Alert_val = bp_mic.Alert_val if alert_mic else ""
                bp_applied = True
        
        if not bp_applied:
            entry.ab_Site_Org = None
            entry.ab_R_breakpoint = None
            entry.ab_I_breakpoint = None
            entry.ab_SDD_breakpoint = None
            entry.ab_S_breakpoint = None
            entry.ab_Alert_val = ""


        entry.save(update_fields=[
            "ab_AccessionNo",
            "ab_Site_Org",
            "ab_Disk_value",
            "ab_Disk_enRIS",
            "ab_MIC_value",
            "ab_MIC_enRIS",
            "ab_MIC_operand",
            "ab_AlertMIC",
            "ab_Alert_val",
            "ab_R_breakpoint",
            "ab_I_breakpoint",
            "ab_SDD_breakpoint",
            "ab_S_breakpoint",
        ])


    # save the RETEST ANTIBIOTICS arsrl data
    existing_retest_entries = {
        (entry.ab_Retest_Abx_code or "").strip().upper(): entry
        for entry in AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates,
            ab_Retest_Abx_code__isnull=False,
        )
    }
    existing_main_entries_for_retest = {
        (entry.ab_Abx_code or "").strip().upper(): entry
        for entry in AntibioticEntry.objects.filter(
            ab_idNum_referred=isolates,
            ab_Abx_code__isnull=False,
        )
    }
    can_skip_retest_entries = (
        old_ars_org == new_ars_org
        and old_specimen_year == specimen_year
    )
    antibiotics_retest = list(_raw_antibiotics_for_panel(
        org_code=resolved_ars_org,
        specimen_year=specimen_year,
        retest=True,
        require_org=True,
        existing_whonet_codes=existing_retest_entries.keys(),
        antibiotic_view=antibiotic_view,
    ))

    if not ars_org_is_na:
      for abx in antibiotics_retest:
        abx_code = (abx.Whonet_Abx or "").strip().upper()

        disk_value  = request.POST.get(f"retest_disk_{abx_code}")
        disk_enris  = request.POST.get(f"retest_disk_enris_{abx_code}", "").strip()
       
        disk_value = _parse_disk_value(disk_value)

        
        mic_value   = request.POST.get(f"retest_mic_{abx_code}")
        mic_enris   = request.POST.get(f"retest_mic_enris_{abx_code}", "").strip()
        mic_operand = request.POST.get(f"retest_mic_operand_{abx_code}", "").strip()
        alert_mic   = f"retest_alert_mic_{abx_code}" in request.POST


        
        try:
            mic_value = float(mic_value) if mic_value else None
        except ValueError:
            mic_value = None


        if disk_value is None and mic_value is None:
            for entry in AntibioticEntry.objects.filter(
                ab_idNum_referred=isolates,
                ab_Retest_Abx_code=abx_code
            ):
                _clear_retest_antibiotic_data(entry)
                _save_or_delete_antibiotic_entry(entry)
            continue

        entry_defaults = {
            "ab_AccessionNo": isolates.AccessionNo,
            "ab_Retest_Antibiotic": abx.Antibiotic,
            "ab_Retest_Abx": abx.Abx_code,
            "ab_Retest_DiskValue": disk_value,
            "ab_Retest_Disk_enRIS": disk_enris,
            "ab_Retest_MICValue": mic_value,
            "ab_Retest_MIC_enRIS": mic_enris,
            "ab_Retest_MIC_operand": mic_operand,
            "ab_Retest_AlertMIC": alert_mic,
        }
        existing_entry = existing_retest_entries.get(abx_code)
        if (
            can_skip_retest_entries
            and existing_entry
            and _entry_values_match(existing_entry, entry_defaults)
        ):
            continue

        entry = existing_entry or existing_main_entries_for_retest.get(abx_code)
        if entry is None:
            entry = AntibioticEntry(
                ab_idNum_referred=isolates,
                ab_Retest_Abx_code=abx_code,
            )
        entry.ab_Retest_Abx_code = abx_code
        for field, value in entry_defaults.items():
            setattr(entry, field, value)
        entry.save()

        entry.ab_breakpoints_id.clear()  

        ret_bp_applied = False
       # apply retest disk value breakpoints
        if disk_value is not None:
            bp_disk = resolve_bp(
                abx_code,
                specimen_year,
                resolved_ars_org,
                "DISK",
            )

            if bp_disk:
                entry.ab_breakpoints_id.set([bp_disk])  # for disk
                entry.ab_Ret_Org = bp_disk.Org
                entry.ab_Org_Flag = bp_disk.Emerging_Org_Flag
                entry.ab_Abx_Flag = bp_disk.Emerging_Abx_Flag
                entry.ab_Abx_Phenotype = bp_disk.Emerging_Pheno_Flag
                entry.ab_Abx_Phenotype_Other = bp_disk.Emerging_Pheno_Flag_Other
                entry.ab_Ret_R_breakpoint   = bp_disk.R_val
                entry.ab_Ret_I_breakpoint   = bp_disk.I_val
                entry.ab_Ret_SDD_breakpoint = bp_disk.SDD_val
                entry.ab_Ret_S_breakpoint   = bp_disk.S_val
                ret_bp_applied = True


        # apply retest mic breakpoints
        if mic_value is not None:
            bp_mic = resolve_bp(
                abx_code,
                specimen_year,
                resolved_ars_org,
                "MIC",
            )

            if bp_mic:
                entry.ab_breakpoints_id.set([bp_mic])   # for mic (overrides)
                entry.ab_Ret_Org = bp_mic.Org
                entry.ab_Org_Flag = bp_mic.Emerging_Org_Flag
                entry.ab_Abx_Flag = bp_mic.Emerging_Abx_Flag
                entry.ab_Abx_Phenotype = bp_mic.Emerging_Pheno_Flag
                entry.ab_Abx_Phenotype_Other = bp_mic.Emerging_Pheno_Flag_Other
                entry.ab_Ret_R_breakpoint   = bp_mic.R_val
                entry.ab_Ret_I_breakpoint   = bp_mic.I_val
                entry.ab_Ret_SDD_breakpoint = bp_mic.SDD_val
                entry.ab_Ret_S_breakpoint   = bp_mic.S_val
                entry.ab_Retest_Alert_val = bp_mic.Alert_val if alert_mic else ""
                ret_bp_applied = True
        
        if not ret_bp_applied:
            entry.ab_Ret_Org = None
            entry.ab_Org_Flag = False
            entry.ab_Abx_Flag = False
            entry.ab_Abx_Phenotype = None
            entry.ab_Abx_Phenotype_Other = None
            entry.ab_Ret_R_breakpoint = None
            entry.ab_Ret_I_breakpoint = None
            entry.ab_Ret_SDD_breakpoint = None
            entry.ab_Ret_S_breakpoint = None
            entry.ab_Retest_Alert_val = ""


        entry.save(update_fields=[
            "ab_AccessionNo",
            "ab_Ret_Org",
            "ab_Org_Flag",
            "ab_Abx_Flag",
            "ab_Abx_Phenotype",
            "ab_Abx_Phenotype_Other",
            "ab_Retest_DiskValue",
            "ab_Retest_Disk_enRIS",
            "ab_Retest_MICValue",
            "ab_Retest_MIC_enRIS",
            "ab_Retest_MIC_operand",
            "ab_Retest_AlertMIC",
            "ab_Ret_R_breakpoint",
            "ab_Ret_I_breakpoint",
            "ab_Ret_SDD_breakpoint",
            "ab_Ret_S_breakpoint",
            "ab_Retest_Alert_val",
        ])

    messages.success(request, "Data saved successfully.")
    next_after_save = (request.POST.get("next_after_save") or "").strip()
    if next_after_save and url_has_allowed_host_and_scheme(
        next_after_save,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_after_save)

    batch_nav = _raw_batch_navigation(isolates)
    if batch_nav.get("next_id"):
        return redirect("edit_data", id=batch_nav["next_id"])
    return redirect("show_data")




# DELETE DATA 
@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@transaction.atomic
def delete_data(request, id):

    isolate = get_object_or_404(Referred_Data, pk=id)
    if not can_manage_batch(request.user, isolate.Batch_id):
        messages.error(request, "You can only delete records from batches that you created.")
        return redirect("show_data")
    accession = isolate.AccessionNo

    if request.method == "POST":

        # 🔹 Delete related Final_Data (if exists)
        Final_Data.objects.filter(
            f_AccessionNo=accession
        ).delete()

        # 🔹 Delete isolate (will cascade to AntibioticEntry)
        isolate.delete()

        messages.success(
            request,
            f"Accession {accession} deleted successfully."
        )

    return redirect("show_data")



########## PDF GENERATION VIEWS

# link callback for xhtml2pdf to handle static and media files DO NOT DELETE!!
def link_callback(uri, rel):
    """
    Convert HTML URIs to absolute system paths so xhtml2pdf can access images and static files.
    """
    sUrl = settings.STATIC_URL      # Typically /static/
    sRoot = settings.STATIC_ROOT    # Path to static folder
    mUrl = settings.MEDIA_URL       # Typically /media/
    mRoot = settings.MEDIA_ROOT     # Path to media folder

    if uri.startswith(mUrl):
        path = os.path.join(mRoot, uri.replace(mUrl, ""))
    elif uri.startswith(sUrl):
        path = os.path.join(sRoot, uri.replace(sUrl, ""))
    else:
        return uri  # Absolute URL (http://...)

    if not os.path.isfile(path):
        raise Exception('File not found: %s' % path)

    return path


def _is_nonviable_result(*values):
    """True when organism/result text says the isolate was non-viable."""
    for value in values:
        normalized = (value or "").strip().lower().replace("-", " ")
        if normalized in {"not viable", "non viable", "nonviable"}:
            return True
    return False


def _tested_pdf_abx_code_groups(entries, abx_map, print_order, site_printable, ars_printable):
    site_codes = set()
    ars_codes = set()

    for entry in entries:
        if entry.ab_Abx_code:
            abx_code = abx_map.get(entry.ab_Abx_code.strip().upper())
            if abx_code and abx_code in site_printable:
                site_codes.add(abx_code)

        if entry.ab_Retest_Abx_code:
            abx_code = abx_map.get(entry.ab_Retest_Abx_code.strip().upper())
            if abx_code and abx_code in ars_printable:
                ars_codes.add(abx_code)

    ordered_site_codes = [code for code in print_order if code in site_codes]
    ordered_site_codes.extend(
        sort_abx_codes_by_antibiotic(
            code for code in site_codes if code not in ordered_site_codes
        )
    )
    ordered_ars_codes = [code for code in print_order if code in ars_codes]
    ordered_ars_codes.extend(
        sort_abx_codes_by_antibiotic(
            code for code in ars_codes if code not in ordered_ars_codes
        )
    )
    return ordered_site_codes, ordered_ars_codes


def _tested_pdf_abx_codes(entries, abx_map, print_order, site_printable, ars_printable):
    site_codes, ars_codes = _tested_pdf_abx_code_groups(
        entries,
        abx_map,
        print_order,
        site_printable,
        ars_printable,
    )
    tested_codes = []
    for code in [*site_codes, *ars_codes]:
        if code not in tested_codes:
            tested_codes.append(code)
    return tested_codes


def _aligned_pdf_abx_codes(site_codes, ars_codes, site_print_order, ars_print_order):
    combined = set(site_codes) | set(ars_codes)
    aligned = []

    for code in [*site_print_order, *ars_print_order]:
        if code in combined and code not in aligned:
            aligned.append(code)

    aligned.extend(
        sort_abx_codes_by_antibiotic(
            code for code in combined if code not in aligned
        )
    )
    return aligned


def _blank_no_organism_report_fields(isolate):
    """Blank only the printed result pane for whichever organism side is n/a."""
    if _is_no_organism(isolate.Site_Org) and not _is_nonviable_result(
        isolate.Site_Pre,
        isolate.Site_OrgName,
        isolate.Site_Pos,
    ):
        isolate.Site_OrgName = ""

    if _is_no_organism(isolate.ars_OrgCode) and not _is_nonviable_result(
        isolate.ars_Pre,
        isolate.ars_OrgName,
        isolate.ars_Post,
    ):
        isolate.ars_OrgName = ""

    if _is_no_organism(isolate.ars_OrgCode):
        isolate.ars_ct_ctl = ""
        isolate.ars_tz_tzl = ""
        isolate.ars_cn_cni = ""
        isolate.ars_ip_ipi = ""


def _chunk_pdf_isolates_for_print(isolates, recommendation_attr, max_per_page=2):
    isolate_list = list(isolates)
    return [
        isolate_list[i:i + max_per_page]
        for i in range(0, len(isolate_list), max_per_page)
    ]


def _pdf_has_long_recommendation(value):
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return False
    numbered_items = len(re.findall(r"(?:^|\s)\d+\.", text))
    return numbered_items >= 4 or len(text) >= 330



# Generate single PDF report
@login_required(login_url="login")
def generate_pdf_panel_old(request, id):

    isolate = get_object_or_404(Referred_Data, pk=id)

    antibiotic_order = list(
        Antibiotic_List.objects
        .exclude(Abx_code__exact="")
        .values("Whonet_Abx", "Abx_code", "Show_Site", "Show_Ars")
        .order_by("id")
    )
    abx_map = {
        (row["Whonet_Abx"] or "").strip().upper(): (row["Abx_code"] or "").strip()
        for row in antibiotic_order
        if (row["Whonet_Abx"] or "").strip() and (row["Abx_code"] or "").strip()
    }
    site_print_order = antibiotic_print_order(show_site=True)
    ars_print_order = antibiotic_print_order(show_ars=True)

    # FETCH ALL ANTIBIOTIC ENTRIES FOR THIS ISOLATE
    entries = list(AntibioticEntry.objects.filter(ab_idNum_referred=isolate))
    site_org = (isolate.Site_Org or "").strip()
    ars_org = (isolate.ars_OrgCode or "").strip() or site_org

    specimen_year = isolate.Spec_Date.year if isolate.Spec_Date else None
    site_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, site_org)
    ars_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, ars_org)

    encoded_site_abx = {
        abx_map.get((entry.ab_Abx_code or "").strip().upper())
        for entry in entries
        if entry.ab_Abx_code
    } - {None, ""}
    encoded_ars_abx = {
        abx_map.get((entry.ab_Retest_Abx_code or "").strip().upper())
        for entry in entries
        if entry.ab_Retest_Abx_code
    } - {None, ""}
    encoded_print_abx = encoded_site_abx | encoded_ars_abx

    printable_site_abx = set(site_print_order)
    printable_ars_abx = set(ars_print_order)
    site_nonviable = _is_nonviable_result(
        isolate.Site_Org,
        isolate.Site_OrgName,
        isolate.Growth,
    )
    ars_nonviable = _is_nonviable_result(
        isolate.ars_OrgCode,
        isolate.ars_OrgName,
    )

    site_candidates = (set(site_panel_abx) | encoded_print_abx) & printable_site_abx
    ars_candidates = (set(ars_panel_abx) | encoded_print_abx) & printable_ars_abx
    site_abx_codes = [
        code for code in site_print_order if code in site_candidates
    ]
    site_abx_codes.extend(
        sort_abx_codes_by_antibiotic(
            code for code in site_candidates if code not in site_abx_codes
        )
    )
    ars_abx_codes = [
        code for code in ars_print_order if code in ars_candidates
    ]
    ars_abx_codes.extend(
        sort_abx_codes_by_antibiotic(
            code for code in ars_candidates if code not in ars_abx_codes
        )
    )

    aligned_abx_codes = _aligned_pdf_abx_codes(
        site_abx_codes,
        ars_abx_codes,
        site_print_order,
        ars_print_order,
    )
    site_abx_codes = aligned_abx_codes
    ars_abx_codes = aligned_abx_codes

    # FETCHED GROUPED ENTRIES
    grouped_entries = {
        code: {"disk": None, "mic": None}
        for code in site_abx_codes
    }

    for e in entries:
        abx_code = abx_map.get((e.ab_Abx_code or "").strip().upper())
        if abx_code not in grouped_entries:
            continue
        if e.ab_Disk_value is not None:
            grouped_entries[abx_code]["disk"] = e
        if e.ab_MIC_value is not None:
            grouped_entries[abx_code]["mic"] = e

    # FETCHED GROUPED RETEST ENTRIES
    grouped_retest = {
        code: {"disk": None, "mic": None}
        for code in ars_abx_codes
    }

    for e in entries:
        abx_code = abx_map.get((e.ab_Retest_Abx_code or "").strip().upper())
        if abx_code not in grouped_retest:
            continue
        if e.ab_Retest_DiskValue is not None:
            grouped_retest[abx_code]["disk"] = e
        if e.ab_Retest_MICValue is not None:
            grouped_retest[abx_code]["mic"] = e

    # CHUNKING FOR TABLE DISPLAY
    MAX_COLS = 29
    site_uses_plus_layout = _organism_type_is_plus(site_org, isolate.Site_OrgName)
    ars_uses_plus_layout = _organism_type_is_plus(ars_org, isolate.ars_OrgName)
    site_max_cols = 32 if site_uses_plus_layout else MAX_COLS
    ars_max_cols = 32 if ars_uses_plus_layout else MAX_COLS
    MAX_ROWS = 2

    def chunked(items, size):
        for i in range(0, len(items), size):
            yield items[i:i + size]

    grouped_rows = list(
        chunked(list(grouped_entries.items()), site_max_cols)
    )[:MAX_ROWS]

    grouped_ars_rows = list(
        chunked(list(grouped_retest.items()), ars_max_cols)
    )[:MAX_ROWS]

    # CONTEXT FOR TEMPLATE
    context = {
        "isolate": isolate,
        "grouped_rows": grouped_rows,
        "grouped_ars_rows": grouped_ars_rows,
        "site_uses_plus_layout": site_uses_plus_layout,
        "ars_uses_plus_layout": ars_uses_plus_layout,
        "now": timezone.now(),
        "logo_path": static("assets/img/brand/arsplogo.jpg"),
    }

    # CREATE PDF RESPONSE
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'filename="Lab_Result_Report.pdf"'

    template = get_template("home/Lab_result.html")
    html = template.render(context)

    pisa.CreatePDF(html, dest=response, link_callback=link_callback)

    return response






@login_required(login_url="login")
def generate_pdf(request, id):

    isolate = get_object_or_404(Referred_Data, pk=id)
    _blank_no_organism_report_fields(isolate)

    antibiotic_order = list(
        Antibiotic_List.objects
        .exclude(Abx_code__exact="")
        .values("Whonet_Abx", "Abx_code")
        .order_by("id")
    )
    abx_map = {
        (row["Whonet_Abx"] or "").strip().upper(): (row["Abx_code"] or "").strip()
        for row in antibiotic_order
        if (row["Whonet_Abx"] or "").strip() and (row["Abx_code"] or "").strip()
    }
    site_print_order = antibiotic_print_order(show_site=True)
    ars_print_order = antibiotic_print_order(show_ars=True)
    site_printable = set(site_print_order)
    ars_printable = set(ars_print_order)

    entries = list(AntibioticEntry.objects.filter(ab_idNum_referred=isolate))

    # Determine the organism codes and specimen year for breakpoint lookup
    site_org = (isolate.Site_Org or "").strip()
    ars_org = (isolate.ars_OrgCode or "").strip() or site_org
    specimen_year = isolate.Spec_Date.year if isolate.Spec_Date else None

    # Get breakpoint panel antibiotic codes for the given organism and specimen year
    site_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, site_org)
    ars_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, ars_org)
    encoded_site_abx = {
        abx_map.get((entry.ab_Abx_code or "").strip().upper())
        for entry in entries
        if entry.ab_Abx_code
    } - {None, ""}
    encoded_ars_abx = {
        abx_map.get((entry.ab_Retest_Abx_code or "").strip().upper())
        for entry in entries
        if entry.ab_Retest_Abx_code
    } - {None, ""}
    encoded_print_abx = encoded_site_abx | encoded_ars_abx

    # Keep panel columns, and also include encoded antibiotics found on either side.
    # Show_Site and Show_Ars still decide whether each side may print the antibiotic.
    site_candidates = (set(site_panel_abx) | encoded_print_abx) & site_printable
    ars_candidates = (set(ars_panel_abx) | encoded_print_abx) & ars_printable

    # Preserve configured print order and append any panel-only codes deterministically.
    site_abx_codes = [abx for abx in site_print_order if abx in site_candidates]
    site_abx_codes.extend(
        sort_abx_codes_by_antibiotic(
            abx for abx in site_candidates if abx not in site_abx_codes
        )
    )
    ars_abx_codes = [abx for abx in ars_print_order if abx in ars_candidates]
    ars_abx_codes.extend(
        sort_abx_codes_by_antibiotic(
            abx for abx in ars_candidates if abx not in ars_abx_codes
        )
    )
    if _is_no_organism(isolate.Site_Org):
        site_abx_codes = []
    if _is_no_organism(isolate.ars_OrgCode):
        ars_abx_codes = []
    aligned_abx_codes = _aligned_pdf_abx_codes(
        site_abx_codes,
        ars_abx_codes,
        site_print_order,
        ars_print_order,
    )
    site_abx_codes = aligned_abx_codes
    ars_abx_codes = aligned_abx_codes

    grouped_entries = {code: {"disk": None, "mic": None} for code in site_abx_codes}
    grouped_retest = {code: {"disk": None, "mic": None} for code in ars_abx_codes}

    for e in entries:
        site_abx = abx_map.get((e.ab_Abx_code or "").strip().upper())
        if site_abx in grouped_entries:
            if e.ab_Disk_value is not None:
                grouped_entries[site_abx]["disk"] = e
            if e.ab_MIC_value is not None:
                grouped_entries[site_abx]["mic"] = e

        ars_abx = abx_map.get((e.ab_Retest_Abx_code or "").strip().upper())
        if ars_abx in grouped_retest:
            if e.ab_Retest_DiskValue is not None:
                grouped_retest[ars_abx]["disk"] = e
            if e.ab_Retest_MICValue is not None:
                grouped_retest[ars_abx]["mic"] = e

    MAX_COLS = 29
    site_uses_plus_layout = _organism_type_is_plus(site_org, isolate.Site_OrgName)
    ars_uses_plus_layout = _organism_type_is_plus(ars_org, isolate.ars_OrgName)
    SITE_MAX_COLS = 32 if site_uses_plus_layout else MAX_COLS
    ARS_MAX_COLS = 32 if ars_uses_plus_layout else MAX_COLS
    MAX_ROWS = 2

    def chunked(items, size):
        for i in range(0, len(items), size):
            yield items[i:i + size]

    blank_site_rows = [
        [("", {"disk": None, "mic": None}) for _ in range(SITE_MAX_COLS)]
        for _ in range(MAX_ROWS)
    ]
    blank_ars_rows = [
        [("", {"disk": None, "mic": None}) for _ in range(ARS_MAX_COLS)]
        for _ in range(MAX_ROWS)
    ]
    grouped_rows = list(chunked(list(grouped_entries.items()), SITE_MAX_COLS))[:MAX_ROWS]
    grouped_ars_rows = list(chunked(list(grouped_retest.items()), ARS_MAX_COLS))[:MAX_ROWS]
    if not grouped_rows:
        grouped_rows = blank_site_rows
    if not grouped_ars_rows:
        grouped_ars_rows = blank_ars_rows

    context = {
        "isolate": isolate,
        "grouped_rows": grouped_rows,
        "grouped_ars_rows": grouped_ars_rows,
        "site_uses_plus_layout": site_uses_plus_layout,
        "ars_uses_plus_layout": ars_uses_plus_layout,
        "now": timezone.now(),
        "logo_path": static("assets/img/brand/arsplogo.jpg"),
    }

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'filename="Lab_Result_Report.pdf"'

    template = get_template("home/Lab_result.html")
    html = template.render(context)

    pisa.CreatePDF(html, dest=response, link_callback=link_callback)

    return response



# Generate batch PDF (2 isolates per page)


# @login_required(login_url="login")
# def generate_batch_pdf(request, id):

#    # Fetch batch
#     batch = get_object_or_404(Batch_Table, pk=id)

#     isolates = (
#         Referred_Data.objects
#         .filter(Batch_id=batch)
#         .order_by("bat_seq")
#     )


#   # chunk isolates into pages (2 per page)
#     def chunked(qs, size):
#         for i in range(0, qs.count(), size):
#             yield qs[i:i + size]

#     isolate_pages = list(chunked(isolates, 2))

#    # antibiotic master list
#     abx_master = (
#         Antibiotic_List.objects
#         .filter(Show_Print=True, Show_Value=True)
#         .order_by("Abx_code")
#         .values("Whonet_Abx", "Abx_code")
#     )

#     abx_map = {
#         row["Whonet_Abx"]: row["Abx_code"]
#         for row in abx_master
#     }

#     retest_master = (
#         Antibiotic_List.objects
#         .filter(Show_Print=True, Show_Value=True, Retest=True)
#         .order_by("Abx_code")
#         .values("Whonet_Abx", "Abx_code")
#     )

#     retest_map = {
#         row["Whonet_Abx"]: row["Abx_code"]
#         for row in retest_master
#     }

#    # chunking constants
#     MAX_COLS = 29
#     MAX_ROWS = 2

#     def chunk_list(items, size):
#         for i in range(0, len(items), size):
#             yield items[i:i + size]

#    # build pages data
#     pages_data = []

#     for page_isolates in isolate_pages:
#         page_entries = []

#         for isolate in page_isolates:

#             entries = AntibioticEntry.objects.filter(
#                 ab_idNum_referred=isolate
#             )

#             # grouped MAIN
#             grouped_entries = {
#                 abx_map[w]: {"disk": None, "mic": None}
#                 for w in abx_map
#             }

#             # grouped RETEST
#             grouped_retest = {
#                 retest_map[w]: {"disk": None, "mic": None}
#                 for w in retest_map
#             }

#             for e in entries:
#                 # MAIN - site
#                 if e.ab_Abx_code in abx_map:
#                     abx_code = abx_map[e.ab_Abx_code]
#                     if e.ab_Disk_Abx:
#                         grouped_entries[abx_code]["disk"] = e
#                     else:
#                         grouped_entries[abx_code]["mic"] = e

#                 # RETEST - arsrl
#                 if e.ab_Retest_Abx_code in retest_map:
#                     abx_code = retest_map[e.ab_Retest_Abx_code]
#                     if e.ab_Disk_Abx:
#                         grouped_retest[abx_code]["disk"] = e
#                     else:
#                         grouped_retest[abx_code]["mic"] = e

#             # show chunked rows
#             grouped_rows = list(
#                 chunk_list(sorted(grouped_entries.items()), MAX_COLS)
#             )[:MAX_ROWS]

#             grouped_ars_rows = list(
#                 chunk_list(sorted(grouped_retest.items()), MAX_COLS)
#             )[:MAX_ROWS]

#             page_entries.append({
#                 "isolate": isolate,
#                 "grouped_rows": grouped_rows,
#                 "grouped_ars_rows": grouped_ars_rows,
#             })

#         pages_data.append(page_entries)


#     context = {
#         "batch": batch,
#         "pages": pages_data,
#         "now": timezone.now(),
#         "logo_path": static("assets/img/brand/arsplogo.jpg"),
#     }

# # pdf response
#     response = HttpResponse(content_type="application/pdf")
#     response["Content-Disposition"] = 'filename="Batch_Report.pdf"'

#     template = get_template("home/Lab_result_batch.html")
#     html = template.render(context)

#     pisa.CreatePDF(html, dest=response, link_callback=link_callback)

#     return response




########### Version 2 - for All antibiotics printing


# for portrait pdf generation, we need to fetch all antibiotics for the batch isolates, then group them by site and arsrl panels, and then chunk them for display. This is more complex but ensures we show all relevant antibiotics in the report.

# @transaction.atomic
# def generate_batch_pdf(request, id):

#     # fetch batch isolates
#     batch = get_object_or_404(Batch_Table, pk=id)

#     isolates = (
#         Referred_Data.objects
#         .filter(Batch_id=batch)
#         .order_by("bat_seq")
#     )

#     # paginate: 2 isolates per page
#     def chunked(qs, size):
#         for i in range(0, qs.count(), size):
#             yield qs[i:i + size]

#     isolate_pages = list(chunked(isolates, 2))

#     # these are the constants
#     MAX_COLS = 29
#     MAX_ROWS = 2

#     def chunk_list(items, size):
#         for i in range(0, len(items), size):
#             yield items[i:i + size]

#     pages_data = []

#     # whonet code map
#     abx_map = dict(
#         Antibiotic_List.objects
#         .values_list("Whonet_Abx", "Abx_code")
#     )

#     # build pdf
#     for page_isolates in isolate_pages:
#         page_entries = []

#         for isolate in page_isolates:

#             site_org = isolate.Site_Org
#             ars_org = isolate.ars_OrgCode

#             # fetch the entries
#             entries = AntibioticEntry.objects.filter(
#                 ab_idNum_referred=isolate
#             )

#             # panels
#             site_panel_abx = set(
#                 BreakpointsTable.objects
#                 .filter(Org=site_org)
#                 .values_list("Abx_code", flat=True)
#                 .distinct()
#             )

#             ars_panel_abx = set(
#                 BreakpointsTable.objects
#                 .filter(Org=ars_org)
#                 .values_list("Abx_code", flat=True)
#                 .distinct()
#             )

#             # extract printable antibiotics
#             printable_abx_site = set(
#                 Antibiotic_List.objects
#                 .filter(Show_Site=True)
#                 .values_list("Abx_code", flat=True)
#             )

#             printable_abx_ars = set(
#                 Antibiotic_List.objects
#                 .filter(Show_Ars=True)
#                 .values_list("Abx_code", flat=True)
#             )

#             # find the encoded antiboitcs
#             encoded_site_abx = set()
#             encoded_ars_abx = set()



#             for e in entries:
#                 if e.ab_Abx_code:
#                     abx = abx_map.get(e.ab_Abx_code.strip().upper())
#                     if abx:
#                         encoded_site_abx.add(abx)

#                 if e.ab_Retest_Abx_code:
#                     abx = abx_map.get(e.ab_Retest_Abx_code.strip().upper())
#                     if abx:
#                         encoded_ars_abx.add(abx)

#             # create the panels 
#             site_abx_codes = sorted(
#                 (site_panel_abx | encoded_site_abx) & printable_abx_site
#             )

#             ars_abx_codes = sorted(
#                 (ars_panel_abx | encoded_ars_abx) & printable_abx_ars
#             )

#             # group the antibiotics based on panels
#             grouped_site = {
#                 abx: {"disk": None, "mic": None}
#                 for abx in site_abx_codes
#             }

#             grouped_ars = {
#                 abx: {"disk": None, "mic": None}
#                 for abx in ars_abx_codes
#             }

#            # assign the values into groups
#             for e in entries:

#                 # -sentinel site-
#                 if e.ab_Abx_code:
#                     abx = abx_map.get(e.ab_Abx_code.strip().upper())
#                     if abx and abx in grouped_site:
#                         # Prefer DISK if a disk value exists
#                         if e.ab_Disk_value is not None:
#                             grouped_site[abx]["disk"] = e

#                         # Otherwise assign MIC only if MIC value exists
#                         elif e.ab_MIC_value is not None:
#                             grouped_site[abx]["mic"] = e


#                 # -arsrl / retest-
#                 if e.ab_Retest_Abx_code:
#                     abx = abx_map.get(e.ab_Retest_Abx_code.strip().upper())
#                     if abx and abx in grouped_ars:
#                         if e.ab_Retest_DiskValue is not None:
#                             grouped_ars[abx]["disk"] = e
#                         elif e.ab_Retest_MICValue is not None:
#                             grouped_ars[abx]["mic"] = e


#             # a chunk for display
#             grouped_rows = list(
#                 chunk_list(list(grouped_site.items()), MAX_COLS)
#             )[:MAX_ROWS]

#             grouped_ars_rows = list(
#                 chunk_list(list(grouped_ars.items()), MAX_COLS)
#             )[:MAX_ROWS]

#             page_entries.append({
#                 "isolate": isolate,
#                 "grouped_rows": grouped_rows,
#                 "grouped_ars_rows": grouped_ars_rows,
#             })

#         pages_data.append(page_entries)

#     # render the pdf
#     context = {
#         "batch": batch,
#         "pages": pages_data,
#         "now": timezone.now(),
#         "logo_path": static("assets/img/brand/arsplogo.jpg"),
#     }

#     response = HttpResponse(content_type="application/pdf")
#     response["Content-Disposition"] = 'filename="Batch_Panel_Report.pdf"'

#     template = get_template("home/Lab_result_panel_portrait.html")
#     html = template.render(context)

#     pisa.CreatePDF(
#         html,
#         dest=response,
#         link_callback=link_callback
#     )

#     return response






# for landscape orientation (all antibiotics)

@transaction.atomic
def generate_batch_pdf_panel_old(request, id):

    # fetch batch isolates
    batch = get_object_or_404(Batch_Table, pk=id)
    changed_fields = _apply_signature_defaults_to_batch(batch)
    if changed_fields:
        batch.save(update_fields=changed_fields)

    isolates = (
        Referred_Data.objects
        .filter(Batch_id=batch)
        .order_by("bat_seq")
    )

# was supposed to sort by accession number, but some accession numbers are not numeric, so we sort by the numeric part of the accession number if possible, otherwise fallback to bat_seq. This ensures a more logical order for printing, but this is resulting to an error.
# @transaction.atomic
# def generate_batch_pdf_panel_old(request, id):

#     batch = get_object_or_404(Batch_Table, pk=id)

#     def accession_sort_key(isolate):
#         accession = (isolate.AccessionNo or "").strip().upper()

#         match = re.search(r"(\d+)$", accession)

#         if match:
#             return int(match.group(1))

#         try:
#             return int(isolate.bat_seq)
#         except (TypeError, ValueError):
#             return 999999

#     isolates = sorted(
#         Referred_Data.objects.filter(Batch_id=batch),
#         key=accession_sort_key
#     )

    isolate_pages = _chunk_pdf_isolates_for_print(isolates, "ars_reco", max_per_page=2)

    # these are the constants
    MAX_COLS = 29
    MAX_ROWS = 2

    def chunk_list(items, size):
        for i in range(0, len(items), size):
            yield items[i:i + size]

    pages_data = []

    # whonet code map
    antibiotic_order = list(
        Antibiotic_List.objects
        .exclude(Abx_code__exact="")
        .values("Whonet_Abx", "Antibiotic", "Show_Site", "Show_Ars", "Abx_code")
        .order_by("id")
    )
    abx_map = {
        (row["Whonet_Abx"] or "").strip().upper(): (row["Abx_code"] or "").strip()
        for row in antibiotic_order
        if (row["Whonet_Abx"] or "").strip() and (row["Abx_code"] or "").strip()
    }
    site_print_order = antibiotic_print_order(show_site=True)
    ars_print_order = antibiotic_print_order(show_ars=True)

    def fixed_panel_rows(grouped_antibiotics, max_cols=MAX_COLS):
        rows = list(chunk_list(list(grouped_antibiotics.items()), max_cols))[:MAX_ROWS]
        while len(rows) < MAX_ROWS:
            rows.append([])
        return rows

    # build pdf
    for page_isolates in isolate_pages:
        page_entries = []
        compact_page = any(
            _pdf_has_long_recommendation(getattr(isolate, "ars_reco", ""))
            for isolate in page_isolates
        )

        for isolate in page_isolates:
            _blank_no_organism_report_fields(isolate)

            site_org = (isolate.Site_Org or "").strip()
            ars_org = (isolate.ars_OrgCode or "").strip() or site_org

            # fetch the entries
            entries = list(AntibioticEntry.objects.filter(
                ab_idNum_referred=isolate
            ))

            specimen_year = isolate.Spec_Date.year if isolate.Spec_Date else None

            # panels
            site_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, site_org)
            ars_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, ars_org)

            encoded_site_abx = {
                abx_map.get((entry.ab_Abx_code or "").strip().upper())
                for entry in entries
                if entry.ab_Abx_code
            }
            encoded_site_abx.discard(None)
            encoded_site_abx.discard("")


            encoded_ars_abx = {
                abx_map.get((entry.ab_Retest_Abx_code or "").strip().upper())
                for entry in entries
                if entry.ab_Retest_Abx_code
            }
            encoded_ars_abx.discard(None)
            encoded_ars_abx.discard("")
            encoded_print_abx = encoded_site_abx | encoded_ars_abx

            printable_site_abx = set(site_print_order)
            printable_ars_abx = set(ars_print_order)
            site_nonviable = _is_nonviable_result(
                isolate.Site_Org,
                isolate.Site_OrgName,
                isolate.Growth,
            )
            ars_nonviable = _is_nonviable_result(
                isolate.ars_OrgCode,
                isolate.ars_OrgName,
            )

            site_candidates = (set(site_panel_abx) | encoded_print_abx) & printable_site_abx
            ars_candidates = (set(ars_panel_abx) | encoded_print_abx) & printable_ars_abx

            site_abx_codes = [
                abx for abx in site_print_order
                if abx in site_candidates
            ]
            site_abx_codes.extend(
                sort_abx_codes_by_antibiotic(
                    abx for abx in site_candidates if abx not in site_abx_codes
                )
            )
            ars_abx_codes = [
                abx for abx in ars_print_order
                if abx in ars_candidates
            ]
            ars_abx_codes.extend(
                sort_abx_codes_by_antibiotic(
                    abx for abx in ars_candidates if abx not in ars_abx_codes
                )
            )
            if _is_no_organism(isolate.Site_Org):
                site_abx_codes = []
            if _is_no_organism(isolate.ars_OrgCode):
                ars_abx_codes = []
            aligned_abx_codes = _aligned_pdf_abx_codes(
                site_abx_codes,
                ars_abx_codes,
                site_print_order,
                ars_print_order,
            )
            site_abx_codes = aligned_abx_codes
            ars_abx_codes = aligned_abx_codes

            # group the antibiotics based on panels
            grouped_site = {
                abx: {"disk": None, "mic": None}
                for abx in site_abx_codes
            }

            grouped_ars = {
                abx: {"disk": None, "mic": None}
                for abx in ars_abx_codes
            }

           # assign the values into groups
            for e in entries:

                # -sentinel site-
                if e.ab_Abx_code:
                    abx = abx_map.get(e.ab_Abx_code.strip().upper())
                    if abx and abx in grouped_site:
                        if e.ab_Disk_value is not None:
                            grouped_site[abx]["disk"] = e

                        if e.ab_MIC_value is not None:
                            grouped_site[abx]["mic"] = e


                # -arsrl / retest-
                if e.ab_Retest_Abx_code:
                    abx = abx_map.get(e.ab_Retest_Abx_code.strip().upper())
                    if abx and abx in grouped_ars:
                        if e.ab_Retest_DiskValue is not None:
                            grouped_ars[abx]["disk"] = e

                        if e.ab_Retest_MICValue is not None:
                            grouped_ars[abx]["mic"] = e


            site_uses_plus_layout = _organism_type_is_plus(site_org, isolate.Site_OrgName)
            ars_uses_plus_layout = _organism_type_is_plus(ars_org, isolate.ars_OrgName)
            site_max_cols = 32 if site_uses_plus_layout else MAX_COLS
            ars_max_cols = 32 if ars_uses_plus_layout else MAX_COLS
            grouped_rows = fixed_panel_rows(grouped_site, site_max_cols)
            grouped_ars_rows = fixed_panel_rows(grouped_ars, ars_max_cols)

            page_entries.append({
                "isolate": isolate,
                "grouped_rows": grouped_rows,
                "grouped_ars_rows": grouped_ars_rows,
                "site_uses_plus_layout": site_uses_plus_layout,
                "ars_uses_plus_layout": ars_uses_plus_layout,
                "compact_page": compact_page,
            })

        pages_data.append(page_entries)

    # render the pdf
    context = {
        "batch": batch,
        "pages": pages_data,
        "now": timezone.now(),
        "logo_path": static("assets/img/brand/arsplogo.jpg"),
    }

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'filename="Batch_Panel_Report.pdf"'

    template = get_template("home/Lab_result_panel.html")
    html = template.render(context)

    pisa.CreatePDF(
        html,
        dest=response,
        link_callback=link_callback
    )

    return response






# to refactor into a batch panel pdf generator for Gram Positive isolates, we can create a new function that filters the isolates based on their Gram stain result. This function will generate a PDF report specifically for Gram Positive isolates in the batch.

@transaction.atomic
def generate_batch_pdf_panel_gram_pos(request, id):

    # fetch batch isolates
    batch = get_object_or_404(Batch_Table, pk=id)
    changed_fields = _apply_signature_defaults_to_batch(batch)
    if changed_fields:
        batch.save(update_fields=changed_fields)

    isolates = (
        Referred_Data.objects
        .filter(Batch_id=batch)
        .order_by("bat_seq")
    )

    isolate_pages = _chunk_pdf_isolates_for_print(isolates, "ars_reco", max_per_page=2)

    # these are the constants
    MAX_COLS = 32
    MAX_ROWS = 2

    def chunk_list(items, size):
        for i in range(0, len(items), size):
            yield items[i:i + size]

    pages_data = []

    # whonet code map
    antibiotic_order = list(
        Antibiotic_List.objects
        .exclude(Abx_code__exact="")
        .values("Whonet_Abx", "Antibiotic", "Show_Site", "Show_Ars", "Abx_code")
        .order_by("id")
    )
    abx_map = {
        (row["Whonet_Abx"] or "").strip().upper(): (row["Abx_code"] or "").strip()
        for row in antibiotic_order
        if (row["Whonet_Abx"] or "").strip() and (row["Abx_code"] or "").strip()
    }
    site_print_order = antibiotic_print_order(show_site=True)
    ars_print_order = antibiotic_print_order(show_ars=True)

    def fixed_panel_rows(grouped_antibiotics):
        rows = list(chunk_list(list(grouped_antibiotics.items()), MAX_COLS))[:MAX_ROWS]
        while len(rows) < MAX_ROWS:
            rows.append([])
        return rows

    # build pdf
    for page_isolates in isolate_pages:
        page_entries = []
        compact_page = any(
            _pdf_has_long_recommendation(getattr(isolate, "ars_reco", ""))
            for isolate in page_isolates
        )

        for isolate in page_isolates:
            _blank_no_organism_report_fields(isolate)

            site_org = (isolate.Site_Org or "").strip()
            ars_org = (isolate.ars_OrgCode or "").strip() or site_org

            # fetch the entries
            entries = list(AntibioticEntry.objects.filter(
                ab_idNum_referred=isolate
            ))

            specimen_year = isolate.Spec_Date.year if isolate.Spec_Date else None

            # panels
            site_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, site_org)
            ars_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, ars_org)

            encoded_site_abx = {
                abx_map.get((entry.ab_Abx_code or "").strip().upper())
                for entry in entries
                if entry.ab_Abx_code
            }
            encoded_site_abx.discard(None)
            encoded_site_abx.discard("")


            encoded_ars_abx = {
                abx_map.get((entry.ab_Retest_Abx_code or "").strip().upper())
                for entry in entries
                if entry.ab_Retest_Abx_code
            }
            encoded_ars_abx.discard(None)
            encoded_ars_abx.discard("")
            encoded_print_abx = encoded_site_abx | encoded_ars_abx

            printable_site_abx = set(site_print_order)
            printable_ars_abx = set(ars_print_order)
            site_nonviable = _is_nonviable_result(
                isolate.Site_Org,
                isolate.Site_OrgName,
                isolate.Growth,
            )
            ars_nonviable = _is_nonviable_result(
                isolate.ars_OrgCode,
                isolate.ars_OrgName,
            )

            site_candidates = (set(site_panel_abx) | encoded_print_abx) & printable_site_abx
            ars_candidates = (set(ars_panel_abx) | encoded_print_abx) & printable_ars_abx

            site_abx_codes = [
                abx for abx in site_print_order
                if abx in site_candidates
            ]
            site_abx_codes.extend(
                sort_abx_codes_by_antibiotic(
                    abx for abx in site_candidates if abx not in site_abx_codes
                )
            )
            ars_abx_codes = [
                abx for abx in ars_print_order
                if abx in ars_candidates
            ]
            ars_abx_codes.extend(
                sort_abx_codes_by_antibiotic(
                    abx for abx in ars_candidates if abx not in ars_abx_codes
                )
            )
            if _is_no_organism(isolate.Site_Org):
                site_abx_codes = []
            if _is_no_organism(isolate.ars_OrgCode):
                ars_abx_codes = []
            aligned_abx_codes = _aligned_pdf_abx_codes(
                site_abx_codes,
                ars_abx_codes,
                site_print_order,
                ars_print_order,
            )
            site_abx_codes = aligned_abx_codes
            ars_abx_codes = aligned_abx_codes

            # group the antibiotics based on panels
            grouped_site = {
                abx: {"disk": None, "mic": None}
                for abx in site_abx_codes
            }

            grouped_ars = {
                abx: {"disk": None, "mic": None}
                for abx in ars_abx_codes
            }

           # assign the values into groups
            for e in entries:

                # -sentinel site-
                if e.ab_Abx_code:
                    abx = abx_map.get(e.ab_Abx_code.strip().upper())
                    if abx and abx in grouped_site:
                        if e.ab_Disk_value is not None:
                            grouped_site[abx]["disk"] = e

                        if e.ab_MIC_value is not None:
                            grouped_site[abx]["mic"] = e


                # -arsrl / retest-
                if e.ab_Retest_Abx_code:
                    abx = abx_map.get(e.ab_Retest_Abx_code.strip().upper())
                    if abx and abx in grouped_ars:
                        if e.ab_Retest_DiskValue is not None:
                            grouped_ars[abx]["disk"] = e

                        if e.ab_Retest_MICValue is not None:
                            grouped_ars[abx]["mic"] = e


            grouped_rows = fixed_panel_rows(grouped_site)
            grouped_ars_rows = fixed_panel_rows(grouped_ars)

            page_entries.append({
                "isolate": isolate,
                "grouped_rows": grouped_rows,
                "grouped_ars_rows": grouped_ars_rows,
                "compact_page": compact_page,
            })

        pages_data.append(page_entries)

    # render the pdf
    context = {
        "batch": batch,
        "pages": pages_data,
        "now": timezone.now(),
        "logo_path": static("assets/img/brand/arsplogo.jpg"),
    }

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'filename="Batch_Panel_Report.pdf"'

    template = get_template("home/Lab_result_panel_gram_pos.html")
    html = template.render(context)

    pisa.CreatePDF(
        html,
        dest=response,
        link_callback=link_callback
    )

    return response






@transaction.atomic
def generate_batch_pdf(request, id):
    return generate_batch_pdf_panel_old(request, id)

    batch = get_object_or_404(Batch_Table, pk=id)
    isolates = (
        Referred_Data.objects
        .filter(Batch_id=batch)
        .order_by("bat_seq")
    )

    def chunked(qs, size):
        for i in range(0, qs.count(), size):
            yield qs[i:i + size]

    def chunk_list(items, size):
        for i in range(0, len(items), size):
            yield items[i:i + size]

    def fixed_rows(grouped_antibiotics, max_cols=MAX_COLS):
        if not grouped_antibiotics:
            return [
                [("", {"disk": None, "mic": None}) for _ in range(max_cols)]
                for _ in range(MAX_ROWS)
            ]
        rows = list(chunk_list(list(grouped_antibiotics.items()), max_cols))[:MAX_ROWS]
        while len(rows) < MAX_ROWS:
            rows.append([])
        return rows

    MAX_COLS = 29
    MAX_ROWS = 2
    isolate_pages = _chunk_pdf_isolates_for_print(list(isolates), "ars_reco", max_per_page=2)
    pages_data = []

    antibiotic_order = list(
        Antibiotic_List.objects
        .exclude(Abx_code__exact="")
        .values("Whonet_Abx", "Abx_code")
        .order_by("id")
    )
    abx_map = {
        (row["Whonet_Abx"] or "").strip().upper(): (row["Abx_code"] or "").strip()
        for row in antibiotic_order
        if (row["Whonet_Abx"] or "").strip() and (row["Abx_code"] or "").strip()
    }
    site_print_order = antibiotic_print_order(show_site=True)
    ars_print_order = antibiotic_print_order(show_ars=True)
    site_printable = set(site_print_order)
    ars_printable = set(ars_print_order)

    for page_isolates in isolate_pages:
        page_entries = []
        compact_page = any(
            _pdf_has_long_recommendation(getattr(isolate, "ars_reco", ""))
            for isolate in page_isolates
        )

        for isolate in page_isolates:
            _blank_no_organism_report_fields(isolate)
            entries = list(AntibioticEntry.objects.filter(
                ab_idNum_referred=isolate
            ))

            site_org = (isolate.Site_Org or "").strip()
            ars_org = (isolate.ars_OrgCode or "").strip() or site_org
            specimen_year = isolate.Spec_Date.year if isolate.Spec_Date else None

            site_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, site_org)
            ars_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, ars_org)
            encoded_site_abx = {
                abx_map.get((entry.ab_Abx_code or "").strip().upper())
                for entry in entries
                if entry.ab_Abx_code
            } - {None, ""}
            encoded_ars_abx = {
                abx_map.get((entry.ab_Retest_Abx_code or "").strip().upper())
                for entry in entries
                if entry.ab_Retest_Abx_code
            } - {None, ""}
            encoded_print_abx = encoded_site_abx | encoded_ars_abx

            site_candidates = (set(site_panel_abx) | encoded_print_abx) & site_printable
            ars_candidates = (set(ars_panel_abx) | encoded_print_abx) & ars_printable

            site_abx_codes = [abx for abx in site_print_order if abx in site_candidates]
            site_abx_codes.extend(
                sort_abx_codes_by_antibiotic(
                    abx for abx in site_candidates if abx not in site_abx_codes
                )
            )
            ars_abx_codes = [abx for abx in ars_print_order if abx in ars_candidates]
            ars_abx_codes.extend(
                sort_abx_codes_by_antibiotic(
                    abx for abx in ars_candidates if abx not in ars_abx_codes
                )
            )
            if _is_no_organism(isolate.Site_Org):
                site_abx_codes = []
            if _is_no_organism(isolate.ars_OrgCode):
                ars_abx_codes = []

            aligned_abx_codes = _aligned_pdf_abx_codes(
                site_abx_codes,
                ars_abx_codes,
                site_print_order,
                ars_print_order,
            )
            site_abx_codes = aligned_abx_codes
            ars_abx_codes = aligned_abx_codes

            grouped_site = {abx: {"disk": None, "mic": None} for abx in site_abx_codes}
            grouped_ars = {abx: {"disk": None, "mic": None} for abx in ars_abx_codes}

            for e in entries:
                site_abx = abx_map.get((e.ab_Abx_code or "").strip().upper())
                if site_abx in grouped_site:
                    if e.ab_Disk_value is not None:
                        grouped_site[site_abx]["disk"] = e
                    if e.ab_MIC_value is not None:
                        grouped_site[site_abx]["mic"] = e

                ars_abx = abx_map.get((e.ab_Retest_Abx_code or "").strip().upper())
                if ars_abx in grouped_ars:
                    if e.ab_Retest_DiskValue is not None:
                        grouped_ars[ars_abx]["disk"] = e
                    if e.ab_Retest_MICValue is not None:
                        grouped_ars[ars_abx]["mic"] = e

            site_uses_plus_layout = _organism_type_is_plus(site_org, isolate.Site_OrgName)
            ars_uses_plus_layout = _organism_type_is_plus(ars_org, isolate.ars_OrgName)
            site_max_cols = 32 if site_uses_plus_layout else MAX_COLS
            ars_max_cols = 32 if ars_uses_plus_layout else MAX_COLS

            page_entries.append({
                "isolate": isolate,
                "grouped_rows": fixed_rows(grouped_site, site_max_cols),
                "grouped_ars_rows": fixed_rows(grouped_ars, ars_max_cols),
                "site_uses_plus_layout": site_uses_plus_layout,
                "ars_uses_plus_layout": ars_uses_plus_layout,
                "compact_page": compact_page,
            })

        pages_data.append(page_entries)

    context = {
        "batch": batch,
        "pages": pages_data,
        "now": timezone.now(),
        "logo_path": static("assets/img/brand/arsplogo.jpg"),
    }

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'filename="Batch_Panel_Report.pdf"'

    template = get_template("home/Lab_result_panel.html")
    html = template.render(context)

    pisa.CreatePDF(
        html,
        dest=response,
        link_callback=link_callback
    )

    return response








# # generate gram stain report  !!! WILL DELETE THIS LATER
# @login_required(login_url="login")

# def generate_gs(request, id):
#     # Get the record from the database using the provided ID
#     try:
#         isolate = Referred_Data.objects.get(pk=id)
#     except Referred_Data.DoesNotExist:
#         return HttpResponse("Error: Data not found.", status=404)
    
#     # Context data to pass to the template
#     context = {
#         'isolate': isolate,
#         'now': timezone.now(),  # Current timestamp
#     }

#     # Create a Django response object with PDF content type
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'inline; filename="Gram_Stain_Report.pdf"'

#     # Load and render the template
#     template_path = 'home/GS_result.html'  # Adjust if needed
#     template = get_template(template_path)
#     html = template.render(context)

#     # Generate PDF using Pisa
#     pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)

#     # Check for errors
#     if pisa_status.err:
#         return HttpResponse(f'Error generating PDF: {html}')

#     return response



######### QUICK SEARCH VIEW

# @login_required(login_url="login")
# # for Quick search
# def search(request):
#    query = request.GET.get('q')
#    items = Referred_Data.objects.filter(AccessionNo__icontains=query)
#    return render (request, 'home/search_results.html',{'items': items, 'query':query})


def _attach_global_search_actions(referred_items, final_items):
    raw_accessions = {
        (item.AccessionNo or "").strip()
        for item in referred_items
        if (item.AccessionNo or "").strip()
    }
    final_accessions = {
        (item.f_AccessionNo or "").strip()
        for item in final_items
        if (item.f_AccessionNo or "").strip()
    }

    matched_finals = {}
    for item in (
        Final_Data.objects
        .filter(f_AccessionNo__in=raw_accessions)
        .select_related("f_Batch_id")
        .order_by("-id")
    ):
        matched_finals.setdefault((item.f_AccessionNo or "").strip(), item)

    matched_raw = {}
    for item in (
        Referred_Data.objects
        .filter(AccessionNo__in=final_accessions)
        .select_related("Batch_id")
        .order_by("-id")
    ):
        matched_raw.setdefault((item.AccessionNo or "").strip(), item)

    all_final_ids = {item.id for item in final_items}
    all_final_ids.update(item.id for item in matched_finals.values())

    raw_batch_ids = {item.Batch_id_id for item in referred_items if item.Batch_id_id}
    final_batch_ids = {item.f_Batch_id_id for item in final_items if item.f_Batch_id_id}
    final_batch_ids.update(
        item.f_Batch_id_id for item in matched_finals.values() if item.f_Batch_id_id
    )
    all_batch_ids = raw_batch_ids | final_batch_ids

    accession_report_ids = {}
    for report in (
        ConcordanceReport.objects
        .filter(final_data_id__in=all_final_ids)
        .exclude(final_data__isnull=True)
        .order_by("final_data_id", "-created_at", "-id")
    ):
        accession_report_ids.setdefault(report.final_data_id, report.id)

    batch_report_ids = {}
    for report in (
        ConcordanceReport.objects
        .filter(batch_id__in=all_batch_ids, final_data__isnull=True)
        .order_by("batch_id", "-created_at", "-id")
    ):
        batch_report_ids.setdefault(report.batch_id, report.id)

    for item in referred_items:
        accession = (item.AccessionNo or "").strip()
        matched_final = matched_finals.get(accession)
        item.search_matched_final = matched_final
        item.search_accession_concordance_id = (
            accession_report_ids.get(matched_final.id) if matched_final else None
        )
        item.search_batch_concordance_id = batch_report_ids.get(item.Batch_id_id)

    for item in final_items:
        accession = (item.f_AccessionNo or "").strip()
        matched_raw_item = matched_raw.get(accession)
        item.search_matched_raw = matched_raw_item
        item.search_accession_concordance_id = accession_report_ids.get(item.id)
        item.search_batch_concordance_id = batch_report_ids.get(item.f_Batch_id_id)


# Quick Search for both
@login_required(login_url="login")
def search(request):
    query = request.GET.get("q", "").strip()

    referred_items = []
    final_items = []

    if query:
        referred_items = Referred_Data.objects.filter(
            AccessionNo__icontains=query
        ).select_related("Batch_id").order_by("-id")

        final_items = Final_Data.objects.filter(
            f_AccessionNo__icontains=query
        ).select_related("f_Batch_id").order_by("-id")

        _attach_global_search_actions(referred_items, final_items)

    return render(
        request,
        "home/search_results.html",
        {
            "query": query,
            "referred_items": referred_items,
            "final_items": final_items,
        }
    )



############## site code dropdown views
def _first_form_error(form, default="Please correct the errors below."):
    form_errors = form.non_field_errors() or [
        error
        for field_errors in form.errors.values()
        for error in field_errors
    ]
    return form_errors[0] if form_errors else default


@login_required(login_url="login")
def add_dropdown(request):
    if request.method != "POST":
        return redirect("/settings/?tab=sitecode")

    site_form = SiteCode_Form(request.POST)

    if site_form.is_valid():
        site_form.save()
        messages.success(request, "Site code added successfully.")
    else:
        messages.warning(request, _first_form_error(site_form, "Failed to add site code. Please check the form."))
        print(site_form.errors)

    return redirect("/settings/?tab=sitecode")


@login_required(login_url="login")
def edit_sitecode(request, pk):
    site = get_object_or_404(SiteData, pk=pk)

    if request.method == "POST":
        site_form = SiteCode_Form(request.POST, instance=site)

        if site_form.is_valid():
            site_form.save()
            messages.success(request, "Site code updated successfully.")
            return redirect("site_view")
        else:
            messages.error(request, "Failed to update site code.")
            print(site_form.errors)

    else:
        site_form = SiteCode_Form(instance=site)

    return render(
        request,
        "home/SiteCodeForm.html",   # ✅ SEPARATE FULL PAGE
        {
            "site_form": site_form,
            "site": site,
            "editing": True,
        },
    )


@login_required(login_url="login")
def delete_dropdown(request, id):
    site_items = get_object_or_404(SiteData, pk=id)
    site_items.delete()
    return redirect('site_view')

def delete_all_dropdown(request):
    SiteData.objects.all().delete()
    messages.success(request, "All site codes were deleted successfully.")
    return redirect('site_view')

@login_required(login_url="login")
def site_view(request):
    q = request.GET.get("q", "").strip()
    site_items = SiteData.objects.all()
    if q:
        site_items = site_items.filter(
            Q(SiteCode__icontains=q)
            | Q(SiteName__icontains=q)
            | Q(Site_Address__icontains=q)
            | Q(Site_Lab_Head__icontains=q)
            | Q(Site_Lab_Head_Credentials__icontains=q)
            | Q(Site_Lab_Head_Email__icontains=q)
            | Q(Site_Med_Ctr_Chief__icontains=q)
            | Q(Site_Med_Ctr_Chief_Credentials__icontains=q)
            | Q(Site_Med_Ctr_Chief_Email__icontains=q)
            | Q(Site_MedTech__icontains=q)
            | Q(Site_MedTech_Credentials__icontains=q)
            | Q(Site_MedTech_Email__icontains=q)
        )
    site_items = site_items.order_by("SiteCode", "SiteName")
    paginator = Paginator(site_items, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    params = request.GET.copy()
    params.pop("page", None)

    return render(
        request,
        'home/SiteCodeView.html',
        {
            'page_obj': page_obj,
            'q': q,
            'preserved_params': params.urlencode(),
        },
    )


def _normalize_site_upload_column(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _site_upload_value(row, *keys):
    for key in keys:
        value = row.get(_normalize_site_upload_column(key), "")
        if pd.notna(value) and str(value).strip():
            return str(value).strip()
    return ""


def _split_site_contact_credentials(name, credentials=""):
    name = str(name or "").strip()
    credentials = str(credentials or "").strip()
    if name and not credentials and "," in name:
        name, credentials = [part.strip() for part in name.split(",", 1)]
    return name, credentials


def _site_upload_defaults(row):
    lab_head = _site_upload_value(row, "site_lab_head", "lab_manager", "lab_head", "laboratory_head")
    lab_head_surname = _site_upload_value(row, "laboratory_head_surname")
    if lab_head and lab_head_surname and lab_head_surname.lower() not in lab_head.lower():
        lab_head = f"{lab_head} {lab_head_surname}".strip()
    lab_head_credentials = _site_upload_value(row, "site_lab_head_credentials", "lab_manager_credentials", "lab_head_credentials", "credentials_1", "credentials.1")
    lab_head, lab_head_credentials = _split_site_contact_credentials(lab_head, lab_head_credentials)

    med_chief, med_chief_credentials = _split_site_contact_credentials(
        _site_upload_value(row, "site_med_ctr_chief", "medical_center_chief"),
        _site_upload_value(row, "site_med_ctr_chief_credentials", "medical_center_chief_credentials", "credentials"),
    )
    medtech, medtech_credentials = _split_site_contact_credentials(
        _site_upload_value(row, "site_medtech", "medtech", "medtech_in_charge"),
        _site_upload_value(row, "site_medtech_credentials", "medtech_credentials", "credentials_2", "credentials.2"),
    )

    return {
        "SiteName": _site_upload_value(row, "sitename", "site_name", "hospital"),
        "Site_Address": _site_upload_value(row, "site_address", "address", "hospital_address"),
        "Site_Lab_Head": lab_head,
        "Site_Lab_Head_Credentials": lab_head_credentials,
        "Site_Lab_Head_Designation": _site_upload_value(row, "site_lab_head_designation", "lab_manager_designation", "lab_head_designation", "position_designation_1", "position_designation.1"),
        "Site_Lab_Head_Email": _site_upload_value(row, "site_lab_head_email", "lab_manager_email", "lab_head_email", "email_address_1", "email_address.1"),
        "Site_Lab_Head_Contact": _site_upload_value(row, "site_lab_head_contact", "lab_manager_contact", "lab_head_contact", "contact_number_1", "contact_number.1"),
        "Site_Med_Ctr_Chief": med_chief,
        "Site_Med_Ctr_Chief_Credentials": med_chief_credentials,
        "Site_Med_Ctr_Chief_Designation": _site_upload_value(row, "site_med_ctr_chief_designation", "medical_center_chief_designation", "position_designation"),
        "Site_Med_Ctr_Chief_Email": _site_upload_value(row, "site_med_ctr_chief_email", "medical_center_chief_email", "email_address"),
        "Site_Med_Ctr_Chief_Contact": _site_upload_value(row, "site_med_ctr_chief_contact", "medical_center_chief_contact", "contact_number"),
        "Site_MedTech": medtech,
        "Site_MedTech_Credentials": medtech_credentials,
        "Site_MedTech_Designation": _site_upload_value(row, "site_medtech_designation", "medtech_designation", "position_designation_2", "position_designation.2"),
        "Site_MedTech_Email": _site_upload_value(row, "site_medtech_email", "medtech_email", "email_address_2", "email_address.2"),
        "Site_MedTech_Contact": _site_upload_value(row, "site_medtech_contact", "medtech_contact", "contact_number_2", "contact_number.2", "laboratory_contact_number"),
    }


@login_required(login_url="login")
def upload_sitecode(request):
    if request.method != "POST":
        return redirect("/settings/?tab=sitecode")

    site_upload_form = SiteCode_uploadForm(request.POST, request.FILES)

    if not site_upload_form.is_valid():
        messages.error(request, "Invalid upload form.")
        return redirect("/settings/?tab=sitecode")

    site_uploaded_file = site_upload_form.save()
    file = site_uploaded_file.File_uploadSite

    try:
        file.open()

        sheets = read_tabular_upload_sheets(file, dtype=str)

        success = 0
        skipped = []
        errors = []

        for sheet_name, df in sheets.items():
            df.columns = [_normalize_site_upload_column(column) for column in df.columns]
            df = df.astype(object).where(pd.notna(df), "")

            for index, row in df.iterrows():
                row_data = row.to_dict()
                site_code = _site_upload_value(row_data, "sitecode", "site_code", "hospital_code").upper()
                defaults = _site_upload_defaults(row_data)
                site_name = defaults["SiteName"]

                if not site_code and not site_name:
                    continue

                if not site_code or not site_name:
                    skipped.append(f"{sheet_name} row {index + 2}")
                    continue

                try:
                    existing = SiteData.objects.filter(SiteCode__iexact=site_code).order_by("id").first()
                    if existing:
                        skipped.append(f"{site_code} already exists")
                        continue
                    else:
                        SiteData.objects.create(SiteCode=site_code, **defaults)
                    success += 1
                except Exception as row_error:
                    errors.append(f"{site_code} ({sheet_name} row {index + 2}): {row_error}")

        messages.success(
            request,
            f"Upload successful. {success} site codes processed."
        )
        if skipped:
            messages.warning(
                request,
                f"Skipped {len(skipped)} row(s) with missing SiteCode/SiteName: {', '.join(skipped[:10])}"
            )
        if errors:
            messages.error(
                request,
                f"{len(errors)} row(s) failed: {'; '.join(errors[:5])}"
            )

    except Exception as e:
        messages.error(request, f"Error processing file: {e}")

    return redirect("/settings/?tab=sitecode")






########## breakpoints views
@login_required(login_url="login")
def add_breakpoints(request):
    if request.method == "POST":
        form = BreakpointsForm(request.POST)

        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Breakpoint added successfully.")
            except IntegrityError:
                messages.error(request, "Required breakpoint fields are missing. Please complete the highlighted fields.")
        else:
            form_errors = form.non_field_errors() or [
                error
                for field_errors in form.errors.values()
                for error in field_errors
            ]
            messages.error(request, form_errors[0] if form_errors else "Please correct the errors below.")

    # ALWAYS redirect back to settings tab
    return redirect("/settings/?tab=breakpoints")


@login_required(login_url="login")
def edit_breakpoints(request, pk):
    breakpoint = get_object_or_404(BreakpointsTable, pk=pk)
    bp_upload_form = Breakpoint_uploadForm()  # keep upload support

    if request.method == "POST":
        form = BreakpointsForm(request.POST, instance=breakpoint)

        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Breakpoint updated successfully.")
                return redirect("/settings/?tab=breakpoints")
            except IntegrityError:
                messages.error(request, "Required breakpoint fields are missing. Please complete the highlighted fields.")

        form_errors = form.non_field_errors() or [
            error
            for field_errors in form.errors.values()
            for error in field_errors
        ]
        messages.error(request, form_errors[0] if form_errors else "Please correct the errors below.")
    else:
        form = BreakpointsForm(instance=breakpoint)

    return render(
        request,
        "home/Breakpoints.html",   # ✅ SEPARATE EDIT PAGE
        {
            "breakpoint_form": form,
            "breakpoint": breakpoint,
            "editing": True,
            "bp_upload_form": bp_upload_form,
        },
    )




@login_required(login_url="login")
def breakpoints_view(request):
    q = request.GET.get('q', '').strip()
    organism = request.GET.get("organism", "").strip()
    test_method = request.GET.get("test_method", "").strip()
    whonet_abx = request.GET.get("whonet_abx", "").strip()
    resistance_phenotype = request.GET.get("resistance_phenotype", "").strip()

    total_breakpoints = BreakpointsTable.objects.count()
    breakpoints = BreakpointsTable.objects.all().order_by('-Date_Modified')

    if q:
        breakpoints = breakpoints.filter(
            Q(Antibiotic__icontains=q) |
            Q(Whonet_Abx__icontains=q) |
            Q(Abx_code__icontains=q) |
            Q(Guidelines__icontains=q) |
            Q(Year__icontains=q) |
            Q(Org__icontains=q) |
            Q(Spec_code__icontains=q) |
            Q(Test_Method__icontains=q) |
            Q(Potency__icontains=q) |
            Q(Tier__icontains=q) |
            Q(R_val__icontains=q) |
            Q(I_val__icontains=q) |
            Q(SDD_val__icontains=q) |
            Q(S_val__icontains=q) |
            Q(Alert_val__icontains=q)
        ).distinct()

    if organism:
        breakpoints = breakpoints.filter(Org__iexact=organism)

    if test_method:
        breakpoints = breakpoints.filter(Test_Method__iexact=test_method)

    if whonet_abx:
        breakpoints = breakpoints.filter(Whonet_Abx__iexact=whonet_abx)

    if resistance_phenotype:
        breakpoints = breakpoints.filter(Emerging_Pheno_Flag__iexact=resistance_phenotype)

    filtered_count = breakpoints.count()

    filter_options = {
        "organisms": (
            BreakpointsTable.objects
            .exclude(Org__isnull=True)
            .exclude(Org__exact="")
            .order_by("Org")
            .values_list("Org", flat=True)
            .distinct()
        ),
        "test_methods": (
            BreakpointsTable.objects
            .exclude(Test_Method__isnull=True)
            .exclude(Test_Method__exact="")
            .order_by("Test_Method")
            .values_list("Test_Method", flat=True)
            .distinct()
        ),
        "whonet_abx": (
            BreakpointsTable.objects
            .exclude(Whonet_Abx__isnull=True)
            .exclude(Whonet_Abx__exact="")
            .order_by("Whonet_Abx")
            .values_list("Whonet_Abx", flat=True)
            .distinct()
        ),
        "resistance_phenotypes": (
            BreakpointsTable.objects
            .exclude(Emerging_Pheno_Flag__isnull=True)
            .exclude(Emerging_Pheno_Flag__exact="")
            .order_by("Emerging_Pheno_Flag")
            .values_list("Emerging_Pheno_Flag", flat=True)
            .distinct()
        ),
    }

    paginator = Paginator(breakpoints, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    params = request.GET.copy()
    params.pop("page", None)
    preserved_params = params.urlencode()

    return render(
        request,
        'home/BreakpointsView.html',
        {
            'breakpoints': breakpoints,
            'page_obj': page_obj,
            'q': q,
            "organism": organism,
            "test_method": test_method,
            "whonet_abx": whonet_abx,
            "resistance_phenotype": resistance_phenotype,
            "total_breakpoints": total_breakpoints,
            "filtered_count": filtered_count,
            "filter_options": filter_options,
            "preserved_params": preserved_params,
        }
    )



@login_required(login_url="login")
#Delete breakpoints
def breakpoints_del(request, id):
    breakpoints = get_object_or_404(BreakpointsTable, pk=id)
    breakpoints.delete()
    return redirect('breakpoints_view')





# works but blindly deletes records with existing whonet_abx code
# @login_required(login_url="login")
# @transaction.atomic
# def upload_breakpoints(request):
#     """
#     Upload and replace BreakpointsTable data from Excel/CSV.
#     Links to existing Antibiotic_List entries via Whonet_Abx.
#     Does NOT create or update Antibiotic_List records.
#     """
#     if request.method == "POST":
#         bp_upload_form = Breakpoint_uploadForm(request.POST, request.FILES)
#         if bp_upload_form.is_valid():
#             uploaded_file = bp_upload_form.save()
#             file = uploaded_file.File_uploadBP
#             print("Uploaded file:", file)

#             try:
#                 # --- Read uploaded file into DataFrame ---
#                 if file.name.endswith(".csv"):
#                     df = pd.read_csv(file)
#                 elif file.name.endswith((".xls", ".xlsx")):
#                     df = pd.read_excel(file)
#                 else:
#                     messages.error(request, "Unsupported file format. Please upload CSV or Excel.")
#                     return redirect("upload_breakpoints")

#                 print("DataFrame contents:\n", df.head())

#                 # Clean data
#                 df.fillna("", inplace=True)
#                 df.columns = df.columns.str.strip()

#                 # Replace existing breakpoints with same WHONET codes
#                 whonet_abx_values = df["Whonet_Abx"].astype(str).str.strip().unique()
#                 BreakpointsTable.objects.filter(Whonet_Abx__in=whonet_abx_values).delete()

#                 # --- Iterate and link ---
#                 skipped = 0
#                 linked = 0
#                 for _, row in df.iterrows():
#                     whonet_code = str(row.get("Whonet_Abx", "")).strip().upper()
#                     if not whonet_code:
#                         continue

#                     # Try to find matching Antibiotic_List record
#                     antibiotic_ref = Antibiotic_List.objects.filter(Whonet_Abx=whonet_code).first()
#                     if not antibiotic_ref:
#                         skipped += 1
#                         print(f"⚠️ Skipped: No Antibiotic_List entry for {whonet_code}")
#                         continue

#                     # Parse date safely
#                     date_modified = pd.to_datetime(row.get("Date_Modified", ""), errors="coerce")
#                     if pd.isna(date_modified):
#                         date_modified = None

#                     # Create BreakpointsTable record linked to Antibiotic_List
#                     BreakpointsTable.objects.create(
#                         # Show=bool(row.get("Show", False)),
#                         # Retest=bool(row.get("Retest", False)),
#                         Disk_Abx=bool(row.get("Disk_Abx", False)),
#                         Emerging_Org_Flag=bool(row.get("Emerging_Org_Flag", False)), 
#                         Emerging_Abx_Flag=bool(row.get("Emerging_Abx_Flag", False)), 
#                         Emerging_Pheno_Flag=row.get("Emerging_Pheno_Flag", ""),
#                         Year=row.get("Year", ""),
#                         Org_Grp=row.get("Org_Grp", ""),
#                         Org=row.get("Org", ""),
#                         Guidelines=row.get("Guidelines", ""),
#                         Tier=row.get("Tier", ""),
#                         Test_Method=row.get("Test_Method", ""),
#                         Potency=row.get("Potency", ""),
#                         Abx_code=row.get("Abx_code", ""),
#                         Antibiotic=row.get("Antibiotic", ""),
#                         Alert_val=row.get("Alert_val", ""),
#                         Whonet_Abx=whonet_code,
#                         R_val=row.get("R_val", ""),
#                         I_val=row.get("I_val", ""),
#                         SDD_val=row.get("SDD_val", ""),
#                         S_val=row.get("S_val", ""),
#                         Date_Modified=date_modified,
#                         Antibiotic_list=antibiotic_ref,  # ✅ Link to existing antibiotic
#                     )
#                     linked += 1

#                 messages.success(
#                     request,
#                     f"✅ Uploaded successfully: {linked} linked, {skipped} skipped (no match in Antibiotic List)."
#                 )
#                 return redirect("breakpoints_view")

#             except Exception as e:
#                 print(" Error during processing:", e)
#                 messages.error(request, f"Error processing file: {e}")
#                 return redirect("add_breakpoints")

#         else:
#             messages.error(request, "Form is not valid.")
#     else:
#         upload_form = Breakpoint_uploadForm()

#     return render(request, "home/Breakpoints.html", {"upload_form": upload_form})




# @login_required(login_url="login")
# @transaction.atomic
# def upload_breakpoints(request):
#     """
#     Upload and replace BreakpointsTable data from Excel/CSV.
#     Links to existing Antibiotic_List entries via Whonet_Abx.
#     Does NOT create or update Antibiotic_List records.
#     """
#     if request.method == "POST":
#         bp_upload_form = Breakpoint_uploadForm(request.POST, request.FILES)
#         if bp_upload_form.is_valid():
#             uploaded_file = bp_upload_form.save()
#             file = uploaded_file.File_uploadBP
#             print("Uploaded file:", file)

#             try:
#                 # --- Read uploaded file into DataFrame ---
#                 if file.name.endswith(".csv"):
#                     df = pd.read_csv(file)
#                 elif file.name.endswith((".xls", ".xlsx")):
#                     df = pd.read_excel(file)
#                 else:
#                     messages.error(request, "Unsupported file format. Please upload CSV or Excel.")
#                     return redirect("upload_breakpoints")

#                 print("DataFrame contents:\n", df.head())

#                 # Clean data
#                 df.fillna("", inplace=True)
#                 df.columns = df.columns.str.strip()

#                 # --------------------------------------------------
#                 # SAFE REPLACEMENT LOGIC (full-coverage check)
#                 # --------------------------------------------------

#                 # Normalize dataframe columns used for comparison
#                 df["Whonet_Abx"] = df["Whonet_Abx"].astype(str).str.strip().str.upper()
#                 df["Year"] = df["Year"].astype(str).str.strip()
#                 df["Test_Method"] = df["Test_Method"].astype(str).str.strip()
#                 df["Org"] = df["Org"].fillna("").astype(str).str.strip()

#                 # Build uploaded breakpoint identity keys
#                 uploaded_keys = set(
#                     zip(
#                         df["Whonet_Abx"],
#                         df["Year"],
#                         df["Test_Method"],
#                         df["Org"],
#                     )
#                 )

#                 # Fetch existing breakpoint identity keys from DB
#                 existing_bps = BreakpointsTable.objects.filter(
#                     Whonet_Abx__in=df["Whonet_Abx"].unique()
#                 ).values_list(
#                     "Whonet_Abx",
#                     "Year",
#                     "Test_Method",
#                     "Org",
#                 )

#                 existing_keys = set(
#                     (w.strip().upper(), y.strip(), t.strip(), (o or "").strip())
#                     for w, y, t, o in existing_bps
#                 )

#                 # Only delete if upload fully replaces existing rows
#                 if existing_keys and not existing_keys.issubset(uploaded_keys):
#                     raise ValueError(
#                         "Upload does not fully replace existing breakpoint definitions. "
#                         "Partial overwrite is not allowed."
#                     )

#                 # Safe to delete
#                 BreakpointsTable.objects.filter(
#                     Whonet_Abx__in=df["Whonet_Abx"].unique()
#                 ).delete()


#                 # --- Iterate and link ---
#                 skipped = 0
#                 linked = 0
#                 for _, row in df.iterrows():
#                     whonet_code = str(row.get("Whonet_Abx", "")).strip().upper()
#                     if not whonet_code:
#                         continue

#                     # Try to find matching Antibiotic_List record
#                     antibiotic_ref = Antibiotic_List.objects.filter(Whonet_Abx=whonet_code).first()
#                     if not antibiotic_ref:
#                         skipped += 1
#                         print(f"⚠️ Skipped: No Antibiotic_List entry for {whonet_code}")
#                         continue

#                     # Parse date safely
#                     date_modified = pd.to_datetime(row.get("Date_Modified", ""), errors="coerce")
#                     if pd.isna(date_modified):
#                         date_modified = None

#                     # Create BreakpointsTable record linked to Antibiotic_List
#                     BreakpointsTable.objects.create(
#                         # Show=bool(row.get("Show", False)),
#                         # Retest=bool(row.get("Retest", False)),
#                         Disk_Abx=bool(row.get("Disk_Abx", False)),
#                         Emerging_Org_Flag=bool(row.get("Emerging_Org_Flag", False)), 
#                         Emerging_Abx_Flag=bool(row.get("Emerging_Abx_Flag", False)), 
#                         Emerging_Pheno_Flag=row.get("Emerging_Pheno_Flag", ""),
#                         Year=row.get("Year", ""),
#                         Org=row.get("Org", ""),
#                         Spec_code=row.get("Spec_code", ""),
#                         Guidelines=row.get("Guidelines", ""),
#                         Tier=row.get("Tier", ""),
#                         Test_Method=row.get("Test_Method", ""),
#                         Potency=row.get("Potency", ""),
#                         Abx_code=row.get("Abx_code", ""),
#                         Antibiotic=row.get("Antibiotic", ""),
#                         Alert_val=row.get("Alert_val", ""),
#                         Whonet_Abx=whonet_code,
#                         R_val=row.get("R_val", ""),
#                         I_val=row.get("I_val", ""),
#                         SDD_val=row.get("SDD_val", ""),
#                         S_val=row.get("S_val", ""),
#                         Date_Modified=date_modified,
#                         Antibiotic_list=antibiotic_ref,  # ✅ Link to existing antibiotic
#                     )
#                     linked += 1

#                 messages.success(
#                     request,
#                     f"✅ Uploaded successfully: {linked} linked, {skipped} skipped (no match in Antibiotic List)."
#                 )
#                 return redirect("breakpoints_view")

#             except Exception as e:
#                 print(" Error during processing:", e)
#                 messages.error(request, f"Error processing file: {e}")
#                 return redirect("add_breakpoints")

#         else:
#             messages.error(request, "Form is not valid.")
#     else:
#         upload_form = Breakpoint_uploadForm()

#     return render(request, "home/Breakpoints.html", {"upload_form": upload_form})


@login_required(login_url="login")
@transaction.atomic
def upload_breakpoints(request):
    """
    Upload and replace BreakpointsTable data from Excel/CSV.
    Links to existing Antibiotic_List entries via Whonet_Abx.
    Does NOT create or update Antibiotic_List records.
    """

    if request.method != "POST":
        return render(
            request,
            "home/Breakpoints.html",
            {"upload_form": Breakpoint_uploadForm()}
        )

    bp_upload_form = Breakpoint_uploadForm(request.POST, request.FILES)
    if not bp_upload_form.is_valid():
        messages.error(request, "Form is not valid.")
        return redirect("add_breakpoints")

    uploaded_file = bp_upload_form.save()
    file = uploaded_file.File_uploadBP
    print("Uploaded file:", file)

    try:
        # --------------------------------------------------
        # READ FILE
        # --------------------------------------------------
        df = read_tabular_upload(file)

        print("DataFrame preview:\n", df.head())

        # --------------------------------------------------
        # CLEAN DATA
        # --------------------------------------------------
        df = df.astype(object).where(pd.notna(df), "")
        df.columns = df.columns.str.strip()

        column_aliases = {
            "whonetabx": "Whonet_Abx",
            "whonet_abx": "Whonet_Abx",
            "abxcode": "Abx_code",
            "abx_code": "Abx_code",
            "antibiotic": "Antibiotic",
            "testmethod": "Test_Method",
            "test_method": "Test_Method",
            "potency": "Potency",
            "year": "Year",
            "org": "Org",
            "speccode": "Spec_code",
            "spec_code": "Spec_code",
            "rval": "R_val",
            "r_val": "R_val",
            "ival": "I_val",
            "i_val": "I_val",
            "sddval": "SDD_val",
            "sdd_val": "SDD_val",
            "sval": "S_val",
            "s_val": "S_val",
        }
        renamed_columns = {}
        for column in df.columns:
            key = re.sub(r"[^a-z0-9]+", "_", str(column).strip().lower()).strip("_")
            canonical = column_aliases.get(key) or column_aliases.get(key.replace("_", ""))
            if canonical and canonical not in df.columns:
                renamed_columns[column] = canonical
        if renamed_columns:
            df.rename(columns=renamed_columns, inplace=True)

        def normalize_cell(value, *, upper=False):
            if value in ("", None):
                return ""
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            value = str(value).strip()
            return value.upper() if upper else value

        def normalize_decimal_text(number_text):
            try:
                number = Decimal(str(number_text).strip())
            except (InvalidOperation, ValueError):
                return str(number_text).strip()

            if number == number.to_integral_value():
                return str(int(number))

            return format(number.normalize(), "f")

        def normalize_breakpoint_value(value):
            value = normalize_cell(value)
            if not value:
                return ""

            match = re.match(r"^(<=|>=|<|>|=)?\s*(-?\d+(?:\.\d+)?)$", value)
            if not match:
                return value

            operator, number_text = match.groups()
            return f"{operator or ''}{normalize_decimal_text(number_text)}"

        df["Whonet_Abx"] = df["Whonet_Abx"].apply(lambda value: normalize_cell(value, upper=True))
        df["Year"] = df["Year"].apply(normalize_cell)
        df["Org"] = df["Org"].apply(normalize_cell)
        df["Test_Method"] = df["Test_Method"].apply(lambda value: normalize_cell(value, upper=True))
        if "Spec_code" not in df.columns:
            df["Spec_code"] = ""
        df["Spec_code"] = df["Spec_code"].apply(normalize_cell)
        for breakpoint_value_col in ["R_val", "I_val", "SDD_val", "S_val"]:
            if breakpoint_value_col not in df.columns:
                df[breakpoint_value_col] = ""
            df[breakpoint_value_col] = df[breakpoint_value_col].apply(
                normalize_breakpoint_value
            )

        def row_value(row, *names, default=""):
            for name in names:
                if name in row.index:
                    return row.get(name, default)
            return default

        def excel_bool(value):
            if isinstance(value, bool):
                return value
            if value in ("", None):
                return False
            return str(value).strip().lower() in {
                "true", "1", "yes", "y", "x"
            }

        def breakpoint_identity(obj):
            return (
                normalize_cell(obj.Whonet_Abx, upper=True),
                normalize_cell(obj.Year),
                normalize_cell(obj.Org),
                normalize_cell(obj.Test_Method, upper=True),
                normalize_cell(obj.Spec_code),
                normalize_breakpoint_value(obj.R_val),
                normalize_breakpoint_value(obj.I_val),
                normalize_breakpoint_value(obj.SDD_val),
                normalize_breakpoint_value(obj.S_val),
            )

        def delete_duplicate_breakpoints(queryset):
            seen = set()
            duplicate_ids = []
            for breakpoint in queryset.order_by("id"):
                key = breakpoint_identity(breakpoint)
                if key in seen:
                    duplicate_ids.append(breakpoint.pk)
                else:
                    seen.add(key)

            if duplicate_ids:
                BreakpointsTable.objects.filter(pk__in=duplicate_ids).delete()

            return len(duplicate_ids)

        # --------------------------------------------------
        # UPSERT BREAKPOINTS
        # --------------------------------------------------
        # Re-uploading should update only the same breakpoint definition.
        # Do not delete every row for the uploaded WHONET antibiotic because
        # other years, organisms, methods, or specimen expressions may be
        # valid separate breakpoint definitions.
        skipped = 0
        duplicate_skipped = 0
        linked = 0
        updated = 0
        deduped = 0
        auto_created_antibiotics = 0

        for _, row in df.iterrows():
            whonet_code = row["Whonet_Abx"]
            if not whonet_code:
                continue

            antibiotic_ref = (
                Antibiotic_List.objects
                .filter(Whonet_Abx=whonet_code)
                .first()
            )

            if not antibiotic_ref:
                antibiotic_ref = Antibiotic_List.objects.create(
                    Show=True,
                    Retest=True,
                    Show_Site=True,
                    Show_Ars=True,
                    Show_Value=True,
                    Disk_Abx=excel_bool(row.get("Disk_Abx", False))
                    or normalize_cell(row.get("Test_Method", ""), upper=True) == "DISK",
                    Test_Method=normalize_cell(row.get("Test_Method", ""), upper=True),
                    Tier=row.get("Tier", ""),
                    Abx_code=normalize_cell(row.get("Abx_code", ""), upper=True),
                    Whonet_Abx=whonet_code,
                    Antibiotic=row.get("Antibiotic", ""),
                    Guidelines=row.get("Guidelines", "CLSI") or "CLSI",
                    Potency=row.get("Potency", ""),
                    Class=row.get("Class", ""),
                    Subclass=row.get("Subclass", ""),
                )
                auto_created_antibiotics += 1
                print(
                    f"Created Antibiotic_List entry for {whonet_code} "
                    f"from breakpoint upload."
                )

            date_modified = pd.to_datetime(
                row.get("Date_Modified", ""),
                errors="coerce"
            )
            if pd.isna(date_modified):
                date_modified = None

            print(
                f"✔ Linking breakpoint → "
                f"{whonet_code} → Antibiotic_List.id={antibiotic_ref.id}"
            )

            year = normalize_cell(row.get("Year", ""))
            org = normalize_cell(row.get("Org", ""))
            test_method = normalize_cell(row.get("Test_Method", ""), upper=True)
            spec_code = normalize_cell(row.get("Spec_code", ""))
            r_val = normalize_breakpoint_value(row.get("R_val", ""))
            i_val = normalize_breakpoint_value(row.get("I_val", ""))
            sdd_val = normalize_breakpoint_value(row.get("SDD_val", ""))
            s_val = normalize_breakpoint_value(row.get("S_val", ""))

            breakpoint_defaults = {
                "Antibiotic_list": antibiotic_ref,
                "Antibiotic": row.get("Antibiotic", ""),
                "Abx_code": row.get("Abx_code", ""),
                "Guidelines": row.get("Guidelines", ""),
                "Tier": row.get("Tier", ""),
                "Potency": row.get("Potency", ""),
                "Disk_Abx": excel_bool(row.get("Disk_Abx", False)),
                "Emerging_specimen": excel_bool(
                    row_value(
                        row,
                        "Emerging_Specimen",
                        "Emerging_specimen",
                    )
                ),
                "Emerging_Org_Flag": excel_bool(row.get("Emerging_Org_Flag", False)),
                "Emerging_Abx_Flag": excel_bool(row.get("Emerging_Abx_Flag", False)),
                "Emerging_Pheno_Flag": row.get("Emerging_Pheno_Flag", ""),
                "Emerging_Pheno_Flag_Other": row.get(
                    "Emerging_Pheno_Flag_Other",
                    "",
                ),
                "R_val": r_val,
                "I_val": i_val,
                "SDD_val": sdd_val,
                "S_val": s_val,
                "Alert_val": row.get("Alert_val", ""),
            }
            if date_modified:
                breakpoint_defaults["Date_Modified"] = date_modified

            candidate_breakpoints = BreakpointsTable.objects.filter(
                Whonet_Abx=whonet_code,
                Year=year,
                Org=org,
                Test_Method=test_method,
                Spec_code=spec_code,
            ).order_by("id")
            matching_breakpoint_ids = [
                breakpoint.pk
                for breakpoint in candidate_breakpoints
                if (
                    normalize_breakpoint_value(breakpoint.R_val) == r_val
                    and normalize_breakpoint_value(breakpoint.I_val) == i_val
                    and normalize_breakpoint_value(breakpoint.SDD_val) == sdd_val
                    and normalize_breakpoint_value(breakpoint.S_val) == s_val
                )
            ]
            matching_breakpoints = BreakpointsTable.objects.filter(
                pk__in=matching_breakpoint_ids
            ).order_by("id")

            breakpoint_obj = matching_breakpoints.first()
            if breakpoint_obj:
                duplicate_ids = list(
                    matching_breakpoints
                    .exclude(pk=breakpoint_obj.pk)
                    .values_list("pk", flat=True)
                )
                if duplicate_ids:
                    BreakpointsTable.objects.filter(pk__in=duplicate_ids).delete()
                    deduped += len(duplicate_ids)
                duplicate_skipped += 1
                continue
            else:
                BreakpointsTable.objects.create(
                    Whonet_Abx=whonet_code,
                    Year=year,
                    Org=org,
                    Test_Method=test_method,
                    Spec_code=spec_code,
                    **breakpoint_defaults,
                )
                linked += 1

        deduped += delete_duplicate_breakpoints(BreakpointsTable.objects.all())

        from apps.wgs_app.views import (
            reapply_final_breakpoints_for_batches,
            reapply_raw_breakpoints_for_batches,
            refresh_emerging_for_batches,
            refresh_tat_for_batches,
            regenerate_batch_concordance,
        )

        batch_ids = list(
            Batch_Table.objects.values_list("id", flat=True)
        )
        refreshed_raw = reapply_raw_breakpoints_for_batches(batch_ids)
        refreshed_final = reapply_final_breakpoints_for_batches(batch_ids)
        refreshed_emerging = refresh_emerging_for_batches(batch_ids)
        regenerated_concordance = regenerate_batch_concordance(
            batch_ids,
            user=request.user,
        )
        refreshed_tat = refresh_tat_for_batches(batch_ids)

        messages.success(
            request,
            f"✅ Uploaded successfully: {linked} created, "
            f"{duplicate_skipped} skipped (already exists), "
            f"{auto_created_antibiotics} antibiotic master record(s) auto-created, "
            f"{skipped} skipped, "
            f"{deduped} duplicate breakpoint record(s) removed."
        )
        return redirect("breakpoints_view")

    except Exception as e:
        print("❌ Error during processing:", e)
        messages.error(request, f"Error processing file: {e}")
        return redirect("add_breakpoints")



@login_required(login_url="login")
#for exporting into excel
def export_breakpoints(request):
    objects = BreakpointsTable.objects.all()
    data = []

    for obj in objects:
        data.append({
            # "Show": obj.Show,
            # "Retest": obj.Retest,
            "Disk_Abx": obj.Disk_Abx,
            "Emerging_Org_Flag": obj.Emerging_Org_Flag,
            "Emerging_Abx_Flag": obj.Emerging_Abx_Flag,
            "Emerging_Pheno_Flag": obj.Emerging_Pheno_Flag,
            "Emerging_Pheno_Flag_Other": obj.Emerging_Pheno_Flag_Other,
            "Emerging_specimen": obj.Emerging_specimen,
            "Spec_code": obj.Spec_code,
            "Year": obj.Year,
            "Org": obj.Org,
            "Org_Code_type": obj.Org_Code_type,
            "Guidelines": obj.Guidelines,
            "Tier": obj.Tier,
            "Test_Method": obj.Test_Method,
            "Potency": obj.Potency,
            "Abx_code": obj.Abx_code,
            "Whonet_Abx": obj.Whonet_Abx,
            "Antibiotic": obj.Antibiotic,
            "R_val": obj.R_val,
            "I_val": obj.I_val,
            "SDD_val": obj.SDD_val,
            "S_val": obj.S_val,
            "Alert_val": obj.Alert_val,
            "Date_Modified": obj.Date_Modified,
        })
    
    # Define file path
    file_path = "Breakpoints_referred.xlsx"

    # Convert data to DataFrame and save as Excel
    df = pd.DataFrame(data)
    df.to_excel(file_path, index=False)

    # Return the file as a response
    return FileResponse(open(file_path, "rb"), as_attachment=True, filename="Breakpoints_referred.xlsx")


def _export_queryset_to_excel(queryset, columns, filename):
    data = []
    for obj in queryset:
        row = {}
        for column, attr in columns:
            value = getattr(obj, attr, "")
            row[column] = value if value is not None else ""
        data.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    pd.DataFrame(data, columns=[column for column, _ in columns]).to_excel(
        response,
        index=False,
    )
    return response


@login_required(login_url="login")
def export_sitecodes(request):
    return _export_queryset_to_excel(
        SiteData.objects.all().order_by("SiteCode"),
        [
            ("SiteCode", "SiteCode"),
            ("SiteName", "SiteName"),
            ("Site_Address", "Site_Address"),
            ("Site_Lab_Head", "Site_Lab_Head"),
            ("Site_Lab_Head_Credentials", "Site_Lab_Head_Credentials"),
            ("Site_Lab_Head_Designation", "Site_Lab_Head_Designation"),
            ("Site_Lab_Head_Email", "Site_Lab_Head_Email"),
            ("Site_Lab_Head_Contact", "Site_Lab_Head_Contact"),
            ("Site_Med_Ctr_Chief", "Site_Med_Ctr_Chief"),
            ("Site_Med_Ctr_Chief_Credentials", "Site_Med_Ctr_Chief_Credentials"),
            ("Site_Med_Ctr_Chief_Designation", "Site_Med_Ctr_Chief_Designation"),
            ("Site_Med_Ctr_Chief_Email", "Site_Med_Ctr_Chief_Email"),
            ("Site_Med_Ctr_Chief_Contact", "Site_Med_Ctr_Chief_Contact"),
            ("Site_MedTech", "Site_MedTech"),
            ("Site_MedTech_Credentials", "Site_MedTech_Credentials"),
            ("Site_MedTech_Designation", "Site_MedTech_Designation"),
            ("Site_MedTech_Email", "Site_MedTech_Email"),
            ("Site_MedTech_Contact", "Site_MedTech_Contact"),
        ],
        "Site_codes.xlsx",
    )


@login_required(login_url="login")
@role_required(ROLE_ADMIN)
def export_contacts(request):
    data = []
    for staff in arsStaff_Details.objects.select_related("User_Account").order_by("Staff_Name"):
        account = staff.User_Account
        try:
            profile = account.profile if account else None
        except UserProfile.DoesNotExist:
            profile = None
        data.append({
            "User_Account": account.username if account else "",
            "First_Name": account.first_name if account else "",
            "Middle_Name": profile.Middle_Name if profile else "",
            "Last_Name": account.last_name if account else "",
            "Email": account.email if account else "",
            "Staff_Name": staff.Staff_Name or "",
            "Staff_Credentials": staff.Staff_Credentials or "",
            "Staff_Designation": staff.Staff_Designation or "",
            "Staff_Telnum": staff.Staff_Telnum or "",
            "Staff_EmailAdd": staff.Staff_EmailAdd or "",
            "Staff_License": staff.Staff_License or "",
            "Staff_Role": staff.display_roles,
        })

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ARSP_staff.xlsx"'
    pd.DataFrame(data).to_excel(response, index=False)
    return response


@login_required(login_url="login")
def delete_all_breakpoints(request):
    if request.method != "POST":
        messages.error(request, "Delete all breakpoints must be confirmed from the page.")
        return redirect('breakpoints_view')

    BreakpointsTable.objects.all().delete()
    messages.success(request, "All records have been deleted successfully.")
    return redirect('breakpoints_view')  # Redirect to the table view


@login_required(login_url="login")
def abxentry_view(request):
    entries = AntibioticEntry.objects.filter(ab_Retest_Abx_code__isnull=True)
    abx_data = {}
    abx_codes = set()

    for entry in entries:
        accession_no = entry.ab_AccessionNo
        abx_code = entry.ab_Abx_code  # Only ordinary antibiotic (excluding retest antibiotics)

        # Get all values and interpretations for ordinary antibiotics
        value = entry.ab_Disk_value or entry.ab_MIC_value
        RIS = entry.ab_Disk_enRIS or entry.ab_MIC_enRIS
        Operand = entry.ab_MIC_operand or None

        if accession_no not in abx_data:
            abx_data[accession_no] = {}

        # Store only **ordinary** antibiotic values
        if abx_code:  
            abx_data[accession_no][abx_code] = {'value': value, 'RIS': RIS, 'Operand': Operand}
            abx_codes.add(abx_code)  # Add only ordinary antibiotics

    context = {
        'abx_data': abx_data,
        'abx_codes': sorted(abx_codes),  # Sorted list of ordinary antibiotics
    }
    
    return render(request, 'home/AntibioticentryView.html', context)


############ Specimen

@login_required(login_url="login")
# View to display all specimen types
def specimen_list(request):
    q = request.GET.get("q", "").strip()
    sort_by = request.GET.get('sort', 'Specimen_code')  # Default sort field
    order = request.GET.get('order', 'desc')  # Default sort order

    sort_field = f"-{sort_by}" if order == 'desc' else sort_by
    specimen_items = SpecimenTypeModel.objects.all().order_by(sort_field)

    if q:
        specimen_items = specimen_items.filter(
            Q(Specimen_code__icontains=q) |
            Q(Specimen_name__icontains=q) 
        )
    paginator = Paginator(specimen_items, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)


    return render(request, 'home/SpecimenView.html', {'specimen_items': specimen_items, 'page_obj': page_obj, 'q': q, 'sort_by': sort_by, 'order': order})




@login_required(login_url="login")
def add_specimen(request):
    
    if request.method != "POST":
        return redirect("/settings/?tab=specimen")

    specimen_form = SpecimenTypeForm(request.POST)

    if specimen_form.is_valid():
        specimen_form.save()
        messages.success(request, "Specimen added successfully.")
    else:
        messages.warning(request, _first_form_error(specimen_form, "Failed to add specimen. Please check the form."))
        print(specimen_form.errors)

    return redirect("/settings/?tab=specimen")




@login_required(login_url="login")
# Edit an existing specimen
@login_required(login_url="login")
def edit_specimen(request, pk):
    specimen = get_object_or_404(SpecimenTypeModel, pk=pk)

    if request.method == "POST":
        form = SpecimenTypeForm(request.POST, instance=specimen)
        if form.is_valid():
            form.save()
            return redirect("specimen_list")
    else:
        form = SpecimenTypeForm(instance=specimen)

    return render(
        request,
        "home/SpecimenEdit.html",
        {
            "form": form,
            "specimen": specimen,
            "editing": True,   
        }
    )



@login_required(login_url="login")
# View to delete a specimen type
def delete_specimen(request, pk):
    specimen = get_object_or_404(SpecimenTypeModel, pk=pk)
    specimen.delete()
    return redirect('specimen_list')



@login_required(login_url="login")
@transaction.atomic
def upload_specimen_code(request):

    if request.method == "POST":
        specimen_upload = SpecimenUploadForm(request.POST, request.FILES)

        if specimen_upload.is_valid():
            uploaded_file = specimen_upload.save()
            file = uploaded_file.File_uploadSpec

            try:
                df = read_tabular_upload(file, dtype=str)

                # Normalize columns
                df.columns = (
                    df.columns
                    .str.strip()
                    .str.lower()
                )

                REQUIRED_COL = "specimen_code"
                if REQUIRED_COL not in df.columns:
                    messages.error(
                        request,
                        "Missing required column: Specimen_code"
                    )
                    return redirect("upload_specimen_code")

                skipped = 0
                duplicate_skipped = 0
                updated = 0
                created = 0

                for idx, row in df.iterrows():
                    specimen_code = str(row.get("specimen_code", "")).strip().lower()

                    if not specimen_code:
                        skipped += 1
                        continue

                    if SpecimenTypeModel.objects.filter(Specimen_code__iexact=specimen_code).exists():
                        duplicate_skipped += 1
                        continue

                    flag = str(row.get("emerging_spec_flag", "")).strip().upper() in [
                        "1", "true", "yes", "y"
                    ]

                    grp_code = str(row.get("specimen_code_grp", "")).strip().lower()
                    grp_obj = None
                    if grp_code:
                        grp_obj, _ = SpecimenTypeModel.objects.get_or_create(
                            Specimen_code=grp_code
                        )

                    SpecimenTypeModel.objects.create(
                        Specimen_code=specimen_code,
                        Emerging_Spec_Flag=flag,
                        Specimen_name=row.get("specimen_name", ""),
                        Specimen_Code_Grp=grp_obj,
                        Specimen_Grp_Name=row.get("specimen_grp_name", ""),
                    )
                    created += 1

                messages.success(
                    request,
                    f"✅ Upload complete: {created} created, {updated} updated, {skipped} skipped."
                )
                if duplicate_skipped:
                    messages.warning(request, f"{duplicate_skipped} specimen code(s) already exist and were skipped.")
                return redirect("specimen_list")

            except Exception as e:
                messages.error(request, f"❌ Error processing file: {e}")
                return redirect("upload_specimen_code")

        messages.error(request, "Form is not valid.")

    return render(request, "Settings.html", {
        "specimen_upload": SpecimenUploadForm()
    })


@login_required(login_url="login")
def delete_all_specimens(request):
    SpecimenTypeModel.objects.all().delete()
    messages.success(request, "All specimen types have been deleted successfully.")
    return redirect('specimen_list')  # Redirect to the specimen list view


@login_required(login_url="login")
def export_specimens(request):
    return _export_queryset_to_excel(
        SpecimenTypeModel.objects.select_related("Specimen_Code_Grp").all().order_by("Specimen_code"),
        [
            ("Emerging_Spec_Flag", "Emerging_Spec_Flag"),
            ("Specimen_code", "Specimen_code"),
            ("Specimen_name", "Specimen_name"),
            ("Specimen_Code_Grp", "Specimen_Code_Grp"),
            ("Specimen_Grp_Name", "Specimen_Grp_Name"),
        ],
        "Specimen_types.xlsx",
    )




####### Download Antibiotic Entries

@login_required(login_url="login")
def export_Antibioticentry(request):
    objects = AntibioticEntry.objects.all()
    data = []

    for obj in objects:
        data.append({
            "ab_idNumber_referred": obj.ab_idNum_referred.AccessionNo if obj.ab_idNum_referred else None,
            "Accession_No": obj.ab_AccessionNo,
            "Site_Org": obj.ab_Site_Org,
            "Antibiotic": obj.ab_Antibiotic,
            "Abx_code": obj.ab_Abx_code,
            "Abx": obj.ab_Abx,
            "Disk_value": obj.ab_Disk_value,
            "Disk_enRIS": obj.ab_Disk_enRIS,
            "MIC_operand": obj.ab_MIC_operand,
            "MIC_value": obj.ab_MIC_value,
            "MIC_enRIS": obj.ab_MIC_enRIS,
            "Ars_Org":obj.ab_Ret_Org,
            "Retest_Antibiotic": obj.ab_Retest_Antibiotic,
            "Retest_Abx_code": obj.ab_Retest_Abx_code,
            "Retest_Abx": obj.ab_Retest_Abx,
            "Retest_DiskValue": obj.ab_Retest_DiskValue,
            "Retest_Disk_enRIS": obj.ab_Retest_Disk_enRIS,
            "Ret_MIC_Operand": obj.ab_Retest_MIC_operand,
            "Retest_MICValue": obj.ab_Retest_MICValue,
            "Retest_MIC_enRIS": obj.ab_Retest_MIC_enRIS,
            "ab_R_breakpoint": obj.ab_R_breakpoint,
            "ab_I_breakpoint": obj.ab_I_breakpoint,
            "ab_SDD_breakpoint": obj.ab_SDD_breakpoint,
            "ab_S_breakpoint":obj.ab_S_breakpoint,
            "ab_Ret_R_breakpoint": obj.ab_Ret_R_breakpoint,
            "ab_Ret_I_breakpoint": obj.ab_Ret_I_breakpoint,
            "ab_Ret_SDD_breakpoint": obj.ab_Ret_SDD_breakpoint,
            "ab_Ret_S_breakpoint": obj.ab_Ret_S_breakpoint
        })
    
    # Define file path
    file_path = "AntibioticEntry_referred.xlsx"

    # Convert data to DataFrame and save as Excel
    df = pd.DataFrame(data)
    df.to_excel(file_path, index=False)

    # Return the file as a response
    return FileResponse(open(file_path, "rb"), as_attachment=True, filename="AntibioticEntry_referred.xlsx")



######## Contact Details views
@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER)
def add_contact(request):
    if request.method != "POST":
        return redirect("/settings/?tab=contact")

    access = role_flags(request.user)
    contact_form = ContactForm(
        request.POST,
        can_edit_roles=access["can_edit_staff_roles"],
    )

    if contact_form.is_valid():
        contact_form.save()
        messages.success(request, "Contact added successfully.")
    else:
        messages.warning(request, _first_form_error(contact_form, "Failed to add contact. Please check the form."))
        print(contact_form.errors)

    return redirect("/settings/?tab=contact")


@login_required(login_url="login")
@role_required(ROLE_ADMIN)
def upload_contacts(request):
    if request.method != "POST":
        return redirect("contact_view")

    uploaded_file = request.FILES.get("contacts_file")
    if not uploaded_file:
        messages.error(request, "Please choose a staff CSV, TSV, or Excel file.")
        return redirect("contact_view")

    try:
        df = read_tabular_upload(uploaded_file)
    except Exception as exc:
        messages.error(request, f"Unable to read staff upload file: {exc}")
        return redirect("contact_view")

    normalized_columns = {
        str(column).strip().lower().replace(" ", "_"): column
        for column in df.columns
    }

    def cell(row, *names):
        for name in names:
            column = normalized_columns.get(name.lower())
            if column is None:
                continue
            value = row.get(column)
            if pd.isna(value):
                return ""
            return str(value).strip()
        return ""

    valid_roles = {"", ROLE_ENCODER, ROLE_LAB_ENCODER, ROLE_VERIFIER, ROLE_CHECKER, ROLE_LAB_MANAGER, ROLE_ADMIN}

    def normalize_uploaded_roles(value):
        role_aliases = {
            "encoder": ROLE_ENCODER,
            "dmu encoder": ROLE_ENCODER,
            "dmu_encoder": ROLE_ENCODER,
            "dmuencoder": ROLE_ENCODER,
            "lab encoder": ROLE_LAB_ENCODER,
            "lab_encoder": ROLE_LAB_ENCODER,
            "labencoder": ROLE_LAB_ENCODER,
            "verifier": ROLE_VERIFIER,
            "checker": ROLE_CHECKER,
            "lab manager": ROLE_LAB_MANAGER,
            "labmanager": ROLE_LAB_MANAGER,
            "lab mgr": ROLE_LAB_MANAGER,
            "laboratory manager": ROLE_LAB_MANAGER,
            "admin": ROLE_ADMIN,
        }
        roles = []
        for part in str(value or "").replace(",", "|").replace(";", "|").split("|"):
            role = role_aliases.get(part.strip().lower(), part.strip())
            if role and role not in valid_roles:
                return None
            if role:
                roles.append(role)
        return "|".join(dict.fromkeys(roles))

    created_count = 0
    updated_count = 0
    skipped_count = 0

    for _, row in df.iterrows():
        staff_name = cell(row, "staff_name", "name", "staff")
        staff_email = cell(row, "staff_emailadd", "staff_email", "email")
        username = cell(row, "user_account", "username", "user")
        role = cell(row, "staff_role", "role")
        role = normalize_uploaded_roles(role)

        if role is None:
            skipped_count += 1
            continue
        if not staff_name and not staff_email and not username:
            skipped_count += 1
            continue

        account = None
        if username:
            account = User.objects.filter(Q(username__iexact=username) | Q(email__iexact=username)).first()

        lookup = {}
        if account:
            lookup["User_Account"] = account
        elif staff_email:
            lookup["Staff_EmailAdd__iexact"] = staff_email
        else:
            lookup["Staff_Name__iexact"] = staff_name

        existing = arsStaff_Details.objects.filter(**lookup).first()
        defaults = {
            "User_Account": account,
            "Staff_Name": staff_name or (account.get_full_name() if account else ""),
            "Staff_Credentials": cell(row, "staff_credentials", "credentials", "suffix"),
            "Staff_Designation": cell(row, "staff_designation", "designation"),
            "Staff_Telnum": cell(row, "staff_telnum", "staff_contact", "contact"),
            "Staff_EmailAdd": staff_email or (account.email if account else ""),
            "Staff_License": cell(row, "staff_license", "license", "license_number"),
            "Staff_Role": role,
        }

        if existing:
            for field, value in defaults.items():
                if field == "User_Account" and value is None:
                    continue
                setattr(existing, field, value)
            existing.save()
            updated_count += 1
        else:
            arsStaff_Details.objects.create(**defaults)
            created_count += 1

    messages.success(
        request,
        f"ARSP staff upload complete. Created: {created_count}, updated: {updated_count}, skipped: {skipped_count}.",
    )
    return redirect("contact_view")



@login_required(login_url="login")
@role_required(ROLE_ADMIN)
def delete_contact(request, id):
    contact_items = get_object_or_404(arsStaff_Details, pk=id)
    contact_items.delete()
    return redirect('contact_view')


@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER)
def edit_contact(request, id):
    contact_item = get_object_or_404(arsStaff_Details, pk=id)

    if request.method == "POST":
        access = role_flags(request.user)
        contact_form = ContactForm(
            request.POST,
            instance=contact_item,
            can_edit_roles=access["can_edit_staff_roles"],
        )
        if contact_form.is_valid():
            contact_form.save()
            messages.success(request, "Staff record updated successfully.")
            return redirect("contact_view")
        messages.error(request, "Failed to update staff record. Please check the form.")
    else:
        access = role_flags(request.user)
        contact_form = ContactForm(
            instance=contact_item,
            can_edit_roles=access["can_edit_staff_roles"],
        )

    return render(
        request,
        "home/Contact_Edit.html",
        {
            "contact_form": contact_form,
            "contact_item": contact_item,
        },
    )


@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER)
def contact_view(request):
    contact_items = arsStaff_Details.objects.select_related("User_Account").all().order_by("Staff_Name")
    return render(request, 'home/Contact_View.html', {'contact_items': contact_items})


@login_required(login_url="login")
def get_ars_staff_details(request):
    ars_staff_name = request.GET.get('ars_staff_id')
    license_field = request.GET.get('license_field')  # NEW: dynamic field key

    ars_staff_details = arsStaff_Details.objects.filter(
        Staff_Name=ars_staff_name
    ).values('Staff_License').first()

    if ars_staff_details:
        return JsonResponse({
            license_field: str(ars_staff_details['Staff_License'])  # dynamic key
        })
    else:
        return JsonResponse({'error': 'Staff not found'}, status=404)






########## Download Combined Table
def is_blank(value):
    return value in [None, '', 0]


def _get_export_date_range(request):
    date_from = parse_date(request.GET.get("date_from", "").strip() or "")
    date_to = parse_date(request.GET.get("date_to", "").strip() or "")
    return date_from, date_to


def _apply_entry_date_range(queryset, entry_field, fallback_field, date_from=None, date_to=None):
    if not date_from and not date_to:
        return queryset

    entry_lookup = f"{entry_field}__date"
    entry_filter = Q()
    fallback_filter = Q(**{f"{entry_field}__isnull": True})

    if date_from:
        entry_filter &= Q(**{f"{entry_lookup}__gte": date_from})
        fallback_filter &= Q(**{f"{fallback_field}__gte": date_from})

    if date_to:
        entry_filter &= Q(**{f"{entry_lookup}__lte": date_to})
        fallback_filter &= Q(**{f"{fallback_field}__lte": date_to})

    return queryset.filter(entry_filter | fallback_filter)


def _apply_referral_date_range(queryset, date_from=None, date_to=None):
    if date_from:
        queryset = queryset.filter(tat_Referral_Date__gte=date_from)
    if date_to:
        queryset = queryset.filter(tat_Referral_Date__lte=date_to)
    return queryset


@login_required(login_url="login")
def download_combined_table(request):
    date_from, date_to = _get_export_date_range(request)
    referred_data_entries = _apply_entry_date_range(
        Referred_Data.objects.all(),
        "Date_of_Entry",
        "Spec_Date",
        date_from,
        date_to,
    )

    # Collect unique antibiotics from both abx and retest
    unique_abx_codes = set()
    for abx_code, rt_code in (
        AntibioticEntry.objects
        .filter(ab_idNum_referred__in=referred_data_entries)
        .values_list('ab_Abx_code', 'ab_Retest_Abx_code')
        .distinct()
    ):
        if abx_code:
            unique_abx_codes.add(abx_code)
        if rt_code:
            unique_abx_codes.add(rt_code)

    sorted_antibiotics = sorted(unique_abx_codes)

    # Pre-check which antibiotics are disk types
    disk_abx_lookup = {
        abx: BreakpointsTable.objects.filter(Whonet_Abx=abx, Disk_Abx=True).exists()
        for abx in sorted_antibiotics
    }

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="combined_raw_data.csv"'
    response.write('\ufeff')  # UTF-8 BOM
    writer = csv.writer(response)


    # Static fields (as you defined)
    static_fields = [
    "bat_seq",
    "Batch_id",
    "Hide",
    "Copy_data",
    "Batch_Name",
    "Batch_Code",
    "Date_of_Entry",
    "Date_Modified",
    "RefNo",
    "BatchNo",
    "Total_batch",
    "AccessionNo",
    "AccessionNoGen",
    "Default_Year",
    "SiteCode",
    "Site_Name",
    "SiteName",
    "Referral_Date",
    "Status",
    "Month_Date",
    "Day_Date",
    "Year_Date",
    "RefDate",
    "Start_AccNo",
    "End_AccNo",
    "No_Isolates",

    "Patient_ID",
    "First_Name",
    "Mid_Name",
    "Last_Name",
    "Date_Birth",
    "Age",
    "Emerging_Flag_Age",
    "Sex",
    "Date_Admis",
    "Nosocomial",
    "Diagnosis",
    "Diagnosis_ICD10",
    "Ward",
    "Ward_Type",
    "Service_Type",

    "Spec_Num",
    "Spec_Date",
    "Spec_Type",
    "Reason",
    "Growth",
    "Urine_ColCt",

    "ampC",
    "ESBL",
    "CARB",
    "MBL",
    "BL",
    "MR",
    "mecA",
    "ICR",
    "OtherResMech",

    "Site_Pre",
    "Site_Org",
    "Site_OrgName",
    "Site_Pos",
    "Comments",

    "ars_ampC",
    "ars_ESBL",
    "ars_CARB",
    "ars_ECIM",
    "ars_MCIM",
    "ars_EC_MCIM",
    "ars_MBL",
    "ars_BL",
    "ars_MR",
    "ars_mecA",
    "ars_ICR",
    "ars_Pre",
    "ars_Post",
    "ars_OrgCode",
    "ars_OrgName",
    "ars_ct_ctl",
    "ars_tz_tzl",
    "ars_cn_cni",
    "ars_ip_ipi",
    "ars_reco_Code",
    "ars_description",
    "ars_reco",
    "Concordance_Check",
    "Concordance_by",
    "Concordance_by_Initials",
    "abx_code",

    "arsp_Encoder",
    "arsp_Enc_Lic",
    "arsp_Checker",
    "arsp_Chec_Lic",
    "arsp_Verifier",
    "arsp_Ver_Lic",
    "arsp_LabManager",
    "arsp_Lab_Lic",
    "arsp_Head",
    "arsp_Head_Lic",
    "Date_Accomplished_ARSP",

    "x_mrse",
    "x_mrsamrse",
    "x_entbac",
    "edta",
]


    header = static_fields[:]
    for abx in sorted_antibiotics:
        header.append(f'{abx}')
        header.append(f'{abx}_RIS')
        header.append(f'{abx}_RT')
        header.append(f'{abx}_RT_RIS')

    writer.writerow(header)

    for referred in referred_data_entries:
        row = [getattr(referred, field, '') for field in static_fields]
        abx_entries = AntibioticEntry.objects.filter(ab_idNum_referred=referred)
        abx_data = {}

        for ab in abx_entries:
            # Initial result
            if ab.ab_Abx_code:
                code = ab.ab_Abx_code
                if code not in abx_data:
                    abx_data[code] = {}
                if not is_blank(ab.ab_MIC_value) or not is_blank(ab.ab_Disk_value):
                    val = ab.ab_Disk_value if not is_blank(ab.ab_Disk_value) else f"{ab.ab_MIC_operand or ''}{ab.ab_MIC_value}"
                    ris = ab.ab_Disk_enRIS or ab.ab_MIC_enRIS
                    abx_data[code].update({
                        '_Val': val,
                        '_RIS': ris,
                    })

            # Retest result
            if ab.ab_Retest_Abx_code:
                code = ab.ab_Retest_Abx_code
                if code not in abx_data:
                    abx_data[code] = {}
                if not is_blank(ab.ab_Retest_MICValue) or not is_blank(ab.ab_Retest_DiskValue):
                    rt_val = ab.ab_Retest_DiskValue if not is_blank(ab.ab_Retest_DiskValue) else f"{ab.ab_Retest_MIC_operand or ''}{ab.ab_Retest_MICValue}"
                    rt_ris = ab.ab_Retest_Disk_enRIS or ab.ab_Retest_MIC_enRIS
                    abx_data[code].update({
                        'RT_Val': rt_val,
                        'RT_RIS': rt_ris,
                    })

        # Populate row with antibiotic data
        for abx in sorted_antibiotics:
            data = abx_data.get(abx, {})
            val = data.get('_Val', '')
            if isinstance(val, (int, float)):
                val = format(val, '.3f')
            rt_val = data.get('RT_Val', '')
            if isinstance(rt_val, (int, float)):
                rt_val = format(rt_val, '.3f')
            row.extend([val, data.get('_RIS', ''), rt_val, data.get('RT_RIS', '')])

        writer.writerow(row)

    return response




def read_uploaded_file(uploaded_file):
    filename = uploaded_file.name.lower()
    if filename.endswith('.csv'):
        return pd.read_csv(uploaded_file)
    elif filename.endswith('.tsv'):
        return pd.read_csv(uploaded_file, sep="\t")
    elif filename.endswith(('.xls', '.xlsx')):
        return pd.read_excel(uploaded_file)
    else:
        raise ValueError("Unsupported file format. Please upload a CSV, TSV, or Excel file.")


def read_tabular_upload(uploaded_file, **kwargs):
    filename = uploaded_file.name.lower()
    if filename.endswith(".csv"):
        return pd.read_csv(uploaded_file, **kwargs)
    if filename.endswith(".tsv"):
        return pd.read_csv(uploaded_file, sep="\t", **kwargs)
    if filename.endswith((".xls", ".xlsx")):
        return pd.read_excel(uploaded_file, **kwargs)
    raise ValueError("Unsupported file format. Please upload CSV, TSV, XLSX, or XLS.")


def read_tabular_upload_sheets(uploaded_file, **kwargs):
    filename = uploaded_file.name.lower()
    if filename.endswith(".csv"):
        return {"CSV": pd.read_csv(uploaded_file, **kwargs)}
    if filename.endswith(".tsv"):
        return {"TSV": pd.read_csv(uploaded_file, sep="\t", **kwargs)}
    if filename.endswith((".xls", ".xlsx")):
        return pd.read_excel(uploaded_file, sheet_name=None, **kwargs)
    raise ValueError("Unsupported file format. Please upload CSV, TSV, XLSX, or XLS.")


### helper for copy data to final
def nz(val):
    """Normalize NULLs for non-nullable CharFields"""
    return val if val not in (None,) else ""





### copy batch




# @login_required(login_url="login")
# @transaction.atomic
# def copy_batch_to_final(request, batch_id):
    
#     def to_bool(val):
#         if isinstance(val, bool):
#             return val
#         if isinstance(val, str):
#             return val.lower() in ("true", "1", "yes", "y")
#         return False

#     batch = get_object_or_404(Batch_Table, pk=batch_id)

#     isolates = Referred_Data.objects.filter(
#         Batch_id=batch
#     )

#     copied = 0

#     for isolate in isolates:

#         raw_entries = AntibioticEntry.objects.filter(
#             ab_idNum_referred=isolate
#         )

#         # final data
#         final_obj, _ = Final_Data.objects.update_or_create(
#             f_AccessionNo=isolate.AccessionNo,
#             defaults={

#                 # ===== BATCH / META =====
#                 "f_bat_seq": isolate.bat_seq,
#                 "f_Batch_id": isolate.Batch_id,
#                 "f_Hide": getattr(isolate, "Hide", False),
#                 "f_Batch_Code": isolate.Batch_Code,
#                 "f_Batch_Name": isolate.Batch_Name,
#                 "f_RefNo": isolate.RefNo,
#                 "f_BatchNo": isolate.BatchNo,
#                 "f_Total_batch": isolate.Total_batch,

#                 "f_AccessionNoGen": getattr(isolate, "AccessionNoGen", ""),

#                 "f_SiteCode": isolate.SiteCode,
#                 "f_Site_Name": isolate.Site_Name,
#                 "f_Referral_Date": isolate.Referral_Date,

#                 # ===== PATIENT =====
#                 "f_Patient_ID": isolate.Patient_ID,
#                 "f_First_Name": isolate.First_Name,
#                 "f_Mid_Name": isolate.Mid_Name,
#                 "f_Last_Name": isolate.Last_Name,
#                 "f_Date_Birth": isolate.Date_Birth,
#                 "f_Age": isolate.Age,
#                 "f_Sex": isolate.Sex,

#                 "f_Date_Admis": isolate.Date_Admis,
#                 "f_Nosocomial": isolate.Nosocomial,
#                 "f_Diagnosis": isolate.Diagnosis,
#                 "f_Diagnosis_ICD10": isolate.Diagnosis_ICD10,
#                 "f_Ward": isolate.Ward,
#                 "f_Ward_Type": isolate.Ward_Type,
#                 "f_Service_Type": isolate.Service_Type,

#                 # ===== SPECIMEN =====
#                 "f_Spec_Num": isolate.Spec_Num,
#                 "f_Spec_Date": isolate.Spec_Date,
#                 "f_Spec_Type": isolate.Spec_Type,
#                 "f_Reason": isolate.Reason,
#                 "f_Growth": isolate.Growth,
#                 "f_Urine_ColCt": isolate.Urine_ColCt,

#                 # ===== PHENOTYPE =====
#                 "f_ampC": isolate.ampC,
#                 "f_ESBL": isolate.ESBL,
#                 "f_CARB": isolate.CARB,
#                 "f_MBL": isolate.MBL,
#                 "f_BL": isolate.BL,
#                 "f_MR": isolate.MR,
#                 "f_mecA": isolate.mecA,
#                 "f_ICR": isolate.ICR,
#                 "f_OtherResMech": isolate.OtherResMech,

#                 # ===== SENTINEL ORGANISM =====
#                 "f_Site_Pre": nz(isolate.Site_Pre),
#                 "f_Site_Org": nz(isolate.Site_Org),
#                 "f_Site_OrgName": nz(isolate.Site_OrgName),
#                 "f_Site_Pos": nz(isolate.Site_Pos),
#                 "f_Comments": nz(isolate.Comments),

#                 # ===== ARSRL =====
#                 "f_ars_ampC": isolate.ars_ampC,
#                 "f_ars_ESBL": isolate.ars_ESBL,
#                 "f_ars_CARB": isolate.ars_CARB,
#                 "f_ars_ECIM": isolate.ars_ECIM,
#                 "f_ars_MCIM": isolate.ars_MCIM,
#                 "f_ars_EC_MCIM": isolate.ars_EC_MCIM,
#                 "f_ars_MBL": isolate.ars_MBL,
#                 "f_ars_BL": isolate.ars_BL,
#                 "f_ars_MR": isolate.ars_MR,
#                 "f_ars_mecA": isolate.ars_mecA,
#                 "f_ars_ICR": isolate.ars_ICR,

#                 "f_ars_Pre": nz(isolate.ars_Pre),
#                 "f_ars_Post": nz(isolate.ars_Post),
#                 "f_ars_OrgCode": nz(isolate.ars_OrgCode),
#                 "f_ars_OrgName": nz(isolate.ars_OrgName),

#                 "f_ars_ct_ctl": isolate.ars_ct_ctl,
#                 "f_ars_tz_tzl": isolate.ars_tz_tzl,
#                 "f_ars_cn_cni": isolate.ars_cn_cni,
#                 "f_ars_ip_ipi": isolate.ars_ip_ipi,

#                 "f_ars_reco_Code": isolate.ars_reco_Code,
#                 "f_ars_description": isolate.ars_description,
#                 "f_ars_reco": isolate.ars_reco,

#                 # ===== SIGNATORIES =====
#                 "f_arsp_Encoder": isolate.arsp_Encoder,
#                 "f_arsp_Enc_Lic": isolate.arsp_Enc_Lic,
#                 "f_arsp_Checker": isolate.arsp_Checker,
#                 "f_arsp_Chec_Lic": isolate.arsp_Chec_Lic,
#                 "f_arsp_Verifier": isolate.arsp_Verifier,
#                 "f_arsp_Ver_Lic": isolate.arsp_Ver_Lic,
#                 "f_arsp_LabManager": isolate.arsp_LabManager,
#                 "f_arsp_Lab_Lic": isolate.arsp_Lab_Lic,
#                 "f_arsp_Head": isolate.arsp_Head,
#                 "f_arsp_Head_Lic": isolate.arsp_Head_Lic,
#                 "f_Date_Accomplished_ARSP": isolate.Date_Accomplished_ARSP,

#                 # ===== EXTRA =====
#                 "f_x_mrse": getattr(isolate, "x_mrse", ""),
#                 "f_x_mrsamrse": getattr(isolate, "x_mrsamrse", ""),
#                 "f_x_entbac": getattr(isolate, "x_entbac", ""),
#                 "f_edta": getattr(isolate, "edta", ""),
#             }
#         )


#         # reset final antibiotics
#         Final_AntibioticEntry.objects.filter(
#             ab_idNum_f_referred=final_obj
#         ).delete()

#         # copy abx and breakopints
#         for e in raw_entries:

#             fe = Final_AntibioticEntry(
#                 ab_idNum_f_referred=final_obj,

#                 ab_AccessionNo=e.ab_AccessionNo,
#                 ab_RefNo=e.ab_RefNo,

#                 ab_Antibiotic=e.ab_Antibiotic,
#                 ab_Abx_code=e.ab_Abx_code,
#                 ab_Abx=e.ab_Abx,

#                 ab_Disk_value=e.ab_Disk_value,
#                 ab_Disk_enRIS=e.ab_Disk_enRIS,

#                 ab_MIC_operand=e.ab_MIC_operand,
#                 ab_MIC_value=e.ab_MIC_value,
#                 ab_MIC_enRIS=e.ab_MIC_enRIS,

#                 ab_AlertMIC=e.ab_AlertMIC,

#                 ab_Retest_Antibiotic=e.ab_Retest_Antibiotic,
#                 ab_Retest_Abx_code=e.ab_Retest_Abx_code,
#                 ab_Retest_Abx=e.ab_Retest_Abx,

#                 ab_Retest_DiskValue=e.ab_Retest_DiskValue,
#                 ab_Retest_Disk_enRIS=e.ab_Retest_Disk_enRIS,

#                 ab_Retest_MIC_operand=e.ab_Retest_MIC_operand,
#                 ab_Retest_MICValue=e.ab_Retest_MICValue,
#                 ab_Retest_MIC_enRIS=e.ab_Retest_MIC_enRIS,

#                 ab_Retest_AlertMIC=e.ab_Retest_AlertMIC,
#             )

#             # REQUIRED FOR M2M
#             fe.save()



#             # COPY BREAKPOINT RELATION
#             raw_bps = e.ab_breakpoints_id.all()
#             if raw_bps.exists():
#                 fe.ab_breakpoints_id.add(*raw_bps)


#             # COPY BREAKPOINT VALUES
#             fe.ab_Site_Org = e.ab_Site_Org
#             fe.ab_R_breakpoint = e.ab_R_breakpoint
#             fe.ab_I_breakpoint = e.ab_I_breakpoint
#             fe.ab_SDD_breakpoint = e.ab_SDD_breakpoint
#             fe.ab_S_breakpoint = e.ab_S_breakpoint

#             fe.ab_Ret_Org = e.ab_Ret_Org
#             fe.ab_Org_Flag = to_bool(e.ab_Org_Flag)
#             fe.ab_Abx_Flag = to_bool(e.ab_Abx_Flag)
#             fe.ab_Abx_Phenotype = (e.ab_Abx_Phenotype or "").strip()
#             fe.ab_Ret_R_breakpoint = e.ab_Ret_R_breakpoint
#             fe.ab_Ret_I_breakpoint = e.ab_Ret_I_breakpoint
#             fe.ab_Ret_SDD_breakpoint = e.ab_Ret_SDD_breakpoint
#             fe.ab_Ret_S_breakpoint = e.ab_Ret_S_breakpoint

#             fe.ab_Alert_val = e.ab_Alert_val
#             fe.ab_Retest_Alert_val = e.ab_Retest_Alert_val
            
            
#             fe.save(update_fields=[
#                 "ab_Site_Org",
#                 "ab_R_breakpoint",
#                 "ab_I_breakpoint",
#                 "ab_SDD_breakpoint",
#                 "ab_S_breakpoint",
#                 "ab_Ret_Org",
#                 "ab_Org_Flag",
#                 "ab_Abx_Flag",
#                 "ab_Abx_Phenotype",
#                 "ab_Ret_R_breakpoint",
#                 "ab_Ret_I_breakpoint",
#                 "ab_Ret_SDD_breakpoint",
#                 "ab_Ret_S_breakpoint",
#                 "ab_Alert_val",
#                 "ab_Retest_Alert_val",
#             ])

#         copied += 1

#     messages.success(
#         request,
#         f"{copied} isolates from batch {batch.bat_Batch_Name} "
#         f"copied to Final successfully. Breakpoints preserved."
#     )

#     return redirect("show_data")



@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@transaction.atomic
def copy_batch_to_final(request, batch_id):

    if request.method != "POST":
        messages.warning(
            request,
            "Use the Copy Batch button so you can choose whether to overwrite existing Final records."
        )
        return redirect("show_data")

    def to_bool(val):
        if isinstance(val, bool):
            return val
        if isinstance(val, str):
            return val.lower() in ("true", "1", "yes", "y")
        return False

    batch = get_object_or_404(Batch_Table, pk=batch_id)
    if not can_manage_batch(request.user, batch):
        messages.error(request, "You can only finalize batches that you created.")
        return redirect("show_data")

    isolates = Referred_Data.objects.filter(
        Batch_id=batch
    )

    existing_final_accessions = set(
        Final_Data.objects.filter(
            f_AccessionNo__in=isolates.values_list("AccessionNo", flat=True)
        ).values_list("f_AccessionNo", flat=True)
    )

    copied = 0
    overwritten = 0
    last_final_obj = None 


    for isolate in isolates:

        spec_obj = None
        ref_no = _isolate_ref_no(isolate)

        raw_entries = AntibioticEntry.objects.filter(
            ab_idNum_referred=isolate
        )

        if isolate.Spec_Type:
            spec_obj = SpecimenTypeModel.objects.filter(
                Specimen_code=isolate.Spec_Type
            ).first()


        # ================= FINAL DATA =================
        final_obj, created = Final_Data.objects.update_or_create(
            f_AccessionNo=isolate.AccessionNo,
            defaults={

                # ===== BATCH / META =====
                "f_bat_seq": isolate.bat_seq,
                "f_Batch_id": isolate.Batch_id,
                "f_Hide": getattr(isolate, "Hide", False),
                "f_Batch_Code": isolate.Batch_Code,
                "f_Batch_Name": isolate.Batch_Name,
                "f_RefNo": ref_no,
                "f_BatchNo": isolate.BatchNo,
                "f_Total_batch": isolate.Total_batch,

                "f_AccessionNoGen": getattr(isolate, "AccessionNoGen", ""),

                "f_SiteCode": isolate.SiteCode,
                "f_Site_Name": isolate.Site_Name,
                "f_Referral_Date": isolate.Referral_Date,

                # ===== PATIENT =====
                "f_Patient_ID": isolate.Patient_ID,
                "f_First_Name": isolate.First_Name,
                "f_Mid_Name": isolate.Mid_Name,
                "f_Last_Name": isolate.Last_Name,
                "f_Date_Birth": isolate.Date_Birth,
                "f_Age": isolate.Age,
                "f_Age_Display": getattr(isolate, "Age_Display", ""),
                "f_Sex": isolate.Sex,

                "f_Date_Admis": isolate.Date_Admis,
                "f_Nosocomial": isolate.Nosocomial,
                "f_Diagnosis": isolate.Diagnosis,
                "f_Diagnosis_ICD10": isolate.Diagnosis_ICD10,
                "f_Ward": isolate.Ward,
                "f_Ward_Type": isolate.Ward_Type,
                "f_Service_Type": isolate.Service_Type,

                # ===== SPECIMEN =====
                "f_Spec_Num": isolate.Spec_Num,
                "f_Spec_Date": isolate.Spec_Date,
                "f_Reason": isolate.Reason,
                "f_Growth": isolate.Growth,
                "f_Urine_ColCt": isolate.Urine_ColCt,

                # ===== PHENOTYPE =====
                "f_ampC": isolate.ampC,
                "f_ESBL": isolate.ESBL,
                "f_CARB": isolate.CARB,
                "f_MBL": isolate.MBL,
                "f_BL": isolate.BL,
                "f_MR": isolate.MR,
                "f_mecA": isolate.mecA,
                "f_ICR": isolate.ICR,
                "f_OtherResMech": isolate.OtherResMech,


                "f_Spec_Type": spec_obj,


                # ===== SENTINEL ORGANISM =====
                "f_Site_Pre": nz(isolate.Site_Pre),
                "f_Site_Pre_ed": nz(isolate.Site_Pre_ed),
                "f_Site_Org": nz(isolate.Site_Org),
                "f_Site_OrgName": nz(isolate.Site_OrgName),
                "f_Site_Pos": nz(isolate.Site_Pos),
                "f_Site_Pos_ed": nz(isolate.Site_Pos_ed),
                "f_Comments": nz(isolate.Comments),

                # ===== ARSRL =====
                "f_ars_ampC": isolate.ars_ampC,
                "f_ars_ESBL": isolate.ars_ESBL,
                "f_ars_CARB": isolate.ars_CARB,
                "f_ars_ECIM": isolate.ars_ECIM,
                "f_ars_MCIM": isolate.ars_MCIM,
                "f_ars_EC_MCIM": isolate.ars_EC_MCIM,
                "f_ars_MBL": isolate.ars_MBL,
                "f_ars_BL": isolate.ars_BL,
                "f_ars_MR": isolate.ars_MR,
                "f_ars_mecA": isolate.ars_mecA,
                "f_ars_ICR": isolate.ars_ICR,

                "f_ars_Pre": nz(isolate.ars_Pre),
                "f_ars_Pre_ed": nz(isolate.ars_Pre_ed),
                "f_ars_Post": nz(isolate.ars_Post),
                "f_ars_Post_ed": nz(isolate.ars_Post_ed),
                "f_ars_OrgCode": nz(isolate.ars_OrgCode),
                "f_ars_OrgName": nz(isolate.ars_OrgName),

                "f_ars_ct_ctl": isolate.ars_ct_ctl,
                "f_ars_tz_tzl": isolate.ars_tz_tzl,
                "f_ars_cn_cni": isolate.ars_cn_cni,
                "f_ars_ip_ipi": isolate.ars_ip_ipi,

                "f_ars_reco_Code": isolate.ars_reco_Code,
                "f_ars_description": isolate.ars_description,
                "f_ars_reco": isolate.ars_reco,

                # ===== SIGNATORIES =====
                "f_arsp_Encoder": isolate.arsp_Encoder,
                "f_arsp_Enc_Lic": isolate.arsp_Enc_Lic,
                "f_arsp_Checker": isolate.arsp_Checker,
                "f_arsp_Chec_Lic": isolate.arsp_Chec_Lic,
                "f_arsp_Verifier": isolate.arsp_Verifier,
                "f_arsp_Ver_Lic": isolate.arsp_Ver_Lic,
                "f_arsp_LabManager": isolate.arsp_LabManager,
                "f_arsp_Lab_Lic": isolate.arsp_Lab_Lic,
                "f_arsp_Head": isolate.arsp_Head,
                "f_arsp_Head_Lic": isolate.arsp_Head_Lic,
                "f_Date_Accomplished_ARSP": isolate.Date_Accomplished_ARSP,

                # ===== EXTRA =====
                "f_x_mrse": getattr(isolate, "x_mrse", ""),
                "f_x_mrsamrse": getattr(isolate, "x_mrsamrse", ""),
                "f_x_entbac": getattr(isolate, "x_entbac", ""),
                "f_edta": getattr(isolate, "edta", ""),
            }
        )


        last_final_obj = final_obj  #  TRACK ACTIVE ISOLATE
        if created:
            copied += 1
        else:
            overwritten += 1

        # ================= RESET FINAL ANTIBIOTICS =================
        Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred=final_obj
        ).delete()


        # =============== COPY ANTIBIOTICS + BREAKPOINTS =================
        for e in raw_entries:

            fe = Final_AntibioticEntry(
                ab_idNum_f_referred=final_obj,

                ab_AccessionNo=e.ab_AccessionNo,
                ab_RefNo=e.ab_RefNo,

                ab_Antibiotic=e.ab_Antibiotic,
                ab_Abx_code=e.ab_Abx_code,
                ab_Abx=e.ab_Abx,

                ab_Disk_value=e.ab_Disk_value,
                ab_Disk_enRIS=e.ab_Disk_enRIS,

                ab_MIC_operand=e.ab_MIC_operand,
                ab_MIC_value=e.ab_MIC_value,
                ab_MIC_enRIS=e.ab_MIC_enRIS,

                ab_AlertMIC=e.ab_AlertMIC,

                ab_Retest_Antibiotic=e.ab_Retest_Antibiotic,
                ab_Retest_Abx_code=e.ab_Retest_Abx_code,
                ab_Retest_Abx=e.ab_Retest_Abx,

                ab_Retest_DiskValue=e.ab_Retest_DiskValue,
                ab_Retest_Disk_enRIS=e.ab_Retest_Disk_enRIS,

                ab_Retest_MIC_operand=e.ab_Retest_MIC_operand,
                ab_Retest_MICValue=e.ab_Retest_MICValue,
                ab_Retest_MIC_enRIS=e.ab_Retest_MIC_enRIS,

                ab_Retest_AlertMIC=e.ab_Retest_AlertMIC,
            )

            fe.save()

            raw_bps = e.ab_breakpoints_id.all()
            if raw_bps.exists():
                fe.ab_breakpoints_id.add(*raw_bps)

            fe.ab_Site_Org = e.ab_Site_Org
            fe.ab_R_breakpoint = e.ab_R_breakpoint
            fe.ab_I_breakpoint = e.ab_I_breakpoint
            fe.ab_SDD_breakpoint = e.ab_SDD_breakpoint
            fe.ab_S_breakpoint = e.ab_S_breakpoint

            fe.ab_Ret_Org = e.ab_Ret_Org
            fe.ab_Org_Flag = to_bool(e.ab_Org_Flag)
            fe.ab_Abx_Flag = to_bool(e.ab_Abx_Flag)
            fe.ab_Abx_Phenotype = (e.ab_Abx_Phenotype or "").strip()
            fe.ab_Abx_Phenotype_Other = (e.ab_Abx_Phenotype_Other or "").strip()

            fe.ab_Ret_R_breakpoint = e.ab_Ret_R_breakpoint
            fe.ab_Ret_I_breakpoint = e.ab_Ret_I_breakpoint
            fe.ab_Ret_SDD_breakpoint = e.ab_Ret_SDD_breakpoint
            fe.ab_Ret_S_breakpoint = e.ab_Ret_S_breakpoint

            fe.ab_Alert_val = e.ab_Alert_val
            fe.ab_Retest_Alert_val = e.ab_Retest_Alert_val

            fe.save(update_fields=[
                "ab_Site_Org",
                "ab_R_breakpoint",
                "ab_I_breakpoint",
                "ab_SDD_breakpoint",
                "ab_S_breakpoint",
                "ab_Ret_Org",
                "ab_Org_Flag",
                "ab_Abx_Flag",
                "ab_Abx_Phenotype",
                "ab_Abx_Phenotype_Other",
                "ab_Ret_R_breakpoint",
                "ab_Ret_I_breakpoint",
                "ab_Ret_SDD_breakpoint",
                "ab_Ret_S_breakpoint",
                "ab_Alert_val",
                "ab_Retest_Alert_val",
            ])

    # SESSION ISOLATE SET ONCE, AFTER COPY
    if last_final_obj:
        request.session["current_final_isolate_id"] = last_final_obj.id

    status_parts = []
    if copied:
        status_parts.append(f"{copied} new")
    if overwritten:
        status_parts.append(f"{overwritten} overwritten")
    status_text = ", ".join(status_parts) or "0 copied"

    messages.success(
        request,
        f"Copy to Final completed for batch {batch.bat_Batch_Name}: "
        f"{status_text}. Breakpoints preserved."
    )

    return redirect("show_data")



####### undo batch copy
@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@transaction.atomic
def undo_copy_batch_to_final(request, batch_id):
    try:
        # Get batch
        batch = get_object_or_404(Batch_Table, pk=batch_id)
        if not can_manage_batch(request.user, batch):
            messages.error(request, "You can only undo batches that you created.")
            return redirect("show_data")

        # Get all referred isolates in this batch
        isolates = Referred_Data.objects.filter(
            Batch_id=batch
        )

        if not isolates.exists():
            messages.warning(
                request,
                "No referred isolates found for this batch."
            )
            return redirect("show_data")

        accession_numbers = isolates.values_list(
            "AccessionNo", flat=True
        )

        # Find corresponding Final_Data records
        final_qs = Final_Data.objects.filter(
            f_AccessionNo__in=accession_numbers
        )

        if not final_qs.exists():
            messages.warning(
                request,
                "No Final Data records found to undo for this batch."
            )
            return redirect("show_data")

        # Delete Final Antibiotic Entries first
        Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred__in=final_qs
        ).delete()

         # 2. CLEAR THE SESSION (The Fix)
        # If the isolate currently being edited is in this batch, remove it from session
        current_session_id = request.session.get("current_final_isolate_id")
        if current_session_id and final_qs.filter(pk=current_session_id).exists():
            request.session.pop("current_final_isolate_id", None)

        # Delete Final Data records
        deleted_count, _ = final_qs.delete()

        messages.success(
            request,
            f"Undo successful: {deleted_count} Final records "
            f"from batch {batch.bat_Batch_Name} were removed."
        )

        return redirect("show_data")

    except Exception as e:
        import traceback
        traceback.print_exc()
        messages.error(
            request,
            f"Error undoing batch copy: {e}"
        )
        return redirect("show_data")





##### copy one isolate only
@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@transaction.atomic
def copy_data_to_final(request, id):
    if request.method != "POST":
        messages.warning(
            request,
            "Use the Copy Accession button so you can choose whether to overwrite an existing Final record."
        )
        return redirect("show_data")

    
    def to_bool(val):
        if isinstance(val, bool):
            return val
        if isinstance(val, str):
            return val.lower() in ("true", "1", "yes", "y")
        return False
        

    isolate = get_object_or_404(Referred_Data, pk=id)
    if not can_manage_batch(request.user, isolate.Batch_id):
        messages.error(request, "You can only finalize records from batches that you created.")
        return redirect("show_data")

    raw_entries = AntibioticEntry.objects.filter(
        ab_idNum_referred=isolate
    )
    ref_no = _isolate_ref_no(isolate)

    final_exists = Final_Data.objects.filter(f_AccessionNo=isolate.AccessionNo).exists()

    # fnial data
    final_obj, created = Final_Data.objects.update_or_create(
        f_AccessionNo=isolate.AccessionNo,
        defaults={

            # batch meta
            "f_bat_seq": isolate.bat_seq,
            "f_Batch_id": isolate.Batch_id,
            "f_Hide": isolate.Hide if hasattr(isolate, "Hide") else False,
            "f_Batch_Code": isolate.Batch_Code,
            "f_Batch_Name": isolate.Batch_Name,
            "f_RefNo": ref_no,
            "f_BatchNo": isolate.BatchNo,
            "f_Total_batch": isolate.Total_batch,

            "f_AccessionNoGen": getattr(isolate, "AccessionNoGen", ""),

            "f_SiteCode": isolate.SiteCode,
            "f_Site_Name": isolate.Site_Name,
            "f_Referral_Date": isolate.Referral_Date,

            # PATIENT 
            "f_Patient_ID": isolate.Patient_ID,
            "f_First_Name": isolate.First_Name,
            "f_Mid_Name": isolate.Mid_Name,
            "f_Last_Name": isolate.Last_Name,
            "f_Date_Birth": isolate.Date_Birth,
            "f_Age": isolate.Age,
            "f_Age_Display": getattr(isolate, "Age_Display", ""),
            "f_Sex": isolate.Sex,

            "f_Date_Admis": isolate.Date_Admis,
            "f_Nosocomial": isolate.Nosocomial,
            "f_Diagnosis": isolate.Diagnosis,
            "f_Diagnosis_ICD10": isolate.Diagnosis_ICD10,
            "f_Ward": isolate.Ward,
            "f_Ward_Type": isolate.Ward_Type,
            "f_Service_Type": isolate.Service_Type,

            # SPECIMEN 
            "f_Spec_Num": isolate.Spec_Num,
            "f_Spec_Date": isolate.Spec_Date,
            "f_Spec_Type": isolate.Spec_Type,
            "f_Reason": isolate.Reason,
            "f_Growth": isolate.Growth,
            "f_Urine_ColCt": isolate.Urine_ColCt,

            # PHENOTYPE 
            "f_ampC": isolate.ampC,
            "f_ESBL": isolate.ESBL,
            "f_CARB": isolate.CARB,
            "f_MBL": isolate.MBL,
            "f_BL": isolate.BL,
            "f_MR": isolate.MR,
            "f_mecA": isolate.mecA,
            "f_ICR": isolate.ICR,
            "f_OtherResMech": isolate.OtherResMech,

            # SENTINEL ORGANISM 
            "f_Site_Pre": nz(isolate.Site_Pre),
            "f_Site_Pre_ed": nz(isolate.Site_Pre_ed),
            "f_Site_Org": nz(isolate.Site_Org),
            "f_Site_OrgName": nz(isolate.Site_OrgName),
            "f_Site_Pos": nz(isolate.Site_Pos),
            "f_Site_Pos_ed": nz(isolate.Site_Pos_ed),
            "f_Comments": nz(isolate.Comments),

            # ARSRL
            "f_ars_ampC": isolate.ars_ampC,
            "f_ars_ESBL": isolate.ars_ESBL,
            "f_ars_CARB": isolate.ars_CARB,
            "f_ars_ECIM": isolate.ars_ECIM,
            "f_ars_MCIM": isolate.ars_MCIM,
            "f_ars_EC_MCIM": isolate.ars_EC_MCIM,
            "f_ars_MBL": isolate.ars_MBL,
            "f_ars_BL": isolate.ars_BL,
            "f_ars_MR": isolate.ars_MR,
            "f_ars_mecA": isolate.ars_mecA,
            "f_ars_ICR": isolate.ars_ICR,

            "f_ars_Pre": nz(isolate.ars_Pre),
            "f_ars_Pre_ed": nz(isolate.ars_Pre_ed),
            "f_ars_Post": nz(isolate.ars_Post),
            "f_ars_Post_ed": nz(isolate.ars_Post_ed),
            "f_ars_OrgCode": nz(isolate.ars_OrgCode),
            "f_ars_OrgName": nz(isolate.ars_OrgName),

            "f_ars_ct_ctl": isolate.ars_ct_ctl,
            "f_ars_tz_tzl": isolate.ars_tz_tzl,
            "f_ars_cn_cni": isolate.ars_cn_cni,
            "f_ars_ip_ipi": isolate.ars_ip_ipi,

            "f_ars_reco_Code": isolate.ars_reco_Code,
            "f_ars_description": isolate.ars_description,
            "f_ars_reco": isolate.ars_reco,

            # SIGNATORIES
            "f_arsp_Encoder": isolate.arsp_Encoder,
            "f_arsp_Enc_Lic": isolate.arsp_Enc_Lic,
            "f_arsp_Checker": isolate.arsp_Checker,
            "f_arsp_Chec_Lic": isolate.arsp_Chec_Lic,
            "f_arsp_Verifier": isolate.arsp_Verifier,
            "f_arsp_Ver_Lic": isolate.arsp_Ver_Lic,
            "f_arsp_LabManager": isolate.arsp_LabManager,
            "f_arsp_Lab_Lic": isolate.arsp_Lab_Lic,
            "f_arsp_Head": isolate.arsp_Head,
            "f_arsp_Head_Lic": isolate.arsp_Head_Lic,
            "f_Date_Accomplished_ARSP": isolate.Date_Accomplished_ARSP,

            # EXTRA
            "f_x_mrse": getattr(isolate, "x_mrse", ""),
            "f_x_mrsamrse": getattr(isolate, "x_mrsamrse", ""),
            "f_x_entbac": getattr(isolate, "x_entbac", ""),
            "f_edta": getattr(isolate, "edta", ""),
        }
    )

    # RESET FINAL ANTIBIOTICS
    Final_AntibioticEntry.objects.filter(
        ab_idNum_f_referred=final_obj
    ).delete()

    # COPY ABX AND BREAKPOINTS
    for e in raw_entries:

        fe = Final_AntibioticEntry(
            ab_idNum_f_referred=final_obj,

            ab_AccessionNo=e.ab_AccessionNo,
            ab_RefNo=e.ab_RefNo,

            ab_Antibiotic=e.ab_Antibiotic,
            ab_Abx_code=e.ab_Abx_code,
            ab_Abx=e.ab_Abx,

            ab_Disk_value=e.ab_Disk_value,
            ab_Disk_enRIS=e.ab_Disk_enRIS,

            ab_MIC_operand=e.ab_MIC_operand,
            ab_MIC_value=e.ab_MIC_value,
            ab_MIC_enRIS=e.ab_MIC_enRIS,

            ab_AlertMIC=e.ab_AlertMIC,

            ab_Retest_Antibiotic=e.ab_Retest_Antibiotic,
            ab_Retest_Abx_code=e.ab_Retest_Abx_code,
            ab_Retest_Abx=e.ab_Retest_Abx,

            ab_Retest_DiskValue=e.ab_Retest_DiskValue,
            ab_Retest_Disk_enRIS=e.ab_Retest_Disk_enRIS,

            ab_Retest_MIC_operand=e.ab_Retest_MIC_operand,
            ab_Retest_MICValue=e.ab_Retest_MICValue,
            ab_Retest_MIC_enRIS=e.ab_Retest_MIC_enRIS,

            ab_Retest_AlertMIC=e.ab_Retest_AlertMIC,
        )

        # SAVE IT FIRST
        fe.save()

        # COPY THE BREAKPOINTS
        raw_bps = e.ab_breakpoints_id.all()
        if raw_bps.exists():
            fe.ab_breakpoints_id.add(*raw_bps)
        

        fe.ab_Site_Org = e.ab_Site_Org
        fe.ab_R_breakpoint = e.ab_R_breakpoint
        fe.ab_I_breakpoint = e.ab_I_breakpoint
        fe.ab_SDD_breakpoint = e.ab_SDD_breakpoint
        fe.ab_S_breakpoint = e.ab_S_breakpoint

        fe.ab_Ret_Org = e.ab_Ret_Org
        fe.ab_Org_Flag = to_bool(e.ab_Org_Flag)
        fe.ab_Abx_Flag = to_bool(e.ab_Abx_Flag)
        fe.ab_Abx_Phenotype = (e.ab_Abx_Phenotype or "").strip()
        fe.ab_Abx_Phenotype_Other = (e.ab_Abx_Phenotype_Other or "").strip()
        fe.ab_Ret_R_breakpoint = e.ab_Ret_R_breakpoint
        fe.ab_Ret_I_breakpoint = e.ab_Ret_I_breakpoint
        fe.ab_Ret_SDD_breakpoint = e.ab_Ret_SDD_breakpoint
        fe.ab_Ret_S_breakpoint = e.ab_Ret_S_breakpoint

        fe.ab_Alert_val = e.ab_Alert_val
        fe.ab_Retest_Alert_val = e.ab_Retest_Alert_val

        fe.save(update_fields=[
            "ab_Site_Org",
            "ab_R_breakpoint",
            "ab_I_breakpoint",
            "ab_SDD_breakpoint",
            "ab_S_breakpoint",
            "ab_Ret_Org",
            "ab_Org_Flag",
            "ab_Abx_Flag",
            "ab_Abx_Phenotype",
            "ab_Abx_Phenotype_Other",
            "ab_Ret_R_breakpoint",
            "ab_Ret_I_breakpoint",
            "ab_Ret_SDD_breakpoint",
            "ab_Ret_S_breakpoint",
            "ab_Alert_val",
            "ab_Retest_Alert_val",
        ])


    messages.success(
        request,
        f"Final data {'copied' if created else 'overwritten'} successfully for Accession {isolate.AccessionNo}. "
        f"Breakpoints copied correctly."
    )

    return redirect("show_data")




#### uploading referred data
@login_required
@transaction.atomic
def upload_combined_table(request):
    """
    Upload referred data (Referred_Data + AntibioticEntry) 
    using user-defined field mappings from FieldMapping.
    """
    form = WGSProjectForm()
    referred_form = ReferredUploadForm()

    if request.method == "POST" and request.FILES.get("ReferredDataFile"):
        try:
            uploaded_file = request.FILES["ReferredDataFile"]
            file_name = uploaded_file.name.lower()

            # --- Load file ---
            if file_name.endswith(".csv"):
                file = TextIOWrapper(uploaded_file.file, encoding="utf-8-sig")
                df = pd.read_csv(file)
            elif file_name.endswith(".tsv"):
                file = TextIOWrapper(uploaded_file.file, encoding="utf-8-sig")
                df = pd.read_csv(file, sep="\t")
            elif file_name.endswith((".xlsx", ".xls")):
                df = pd.read_excel(uploaded_file)
            else:
                messages.error(request, "Unsupported file format. Please upload CSV, TSV, XLSX, or XLS.")
                return render(request, "wgs_app/Add_wgs.html", {
                    "referred_form": referred_form,
                    "form": form,
                })

            # --- Apply user-defined mappings ---
            user_mappings = dict(
                FieldMapping.objects.filter(user=request.user)
                .values_list("raw_field", "mapped_field")
            )
            if user_mappings:
                df.rename(columns=user_mappings, inplace=True)
                print(f"[UPLOAD] Applied {len(user_mappings)} user field mappings.")
            else:
                messages.warning(request, " No saved field mappings found. Using raw headers.")

            # --- Normalize headers (fallback cleanup) ---
            def normalize_header(header):
                key = str(header).strip().lower().replace("_", " ").replace("-", " ")
                return re.sub(r"\s+", " ", key).strip().title()

            df.columns = [normalize_header(c) for c in df.columns]

            # --- Prepare data ---
            rows = df.to_dict("records")
            site_codes = set(SiteData.objects.values_list("SiteCode", flat=True))
            model_fields = [f.name for f in Referred_Data._meta.get_fields()]
            known_abx = set(BreakpointsTable.objects.values_list("Whonet_Abx", flat=True))

            created_ref, updated_ref, created_abx, updated_abx = 0, 0, 0, 0

            # --- Helper Functions ---
            def parse_mic_value(value_str):
                """Extract operator and numeric MIC value (e.g. '<=0.5' → ('<=', 0.5))"""
                if not value_str or pd.isna(value_str):
                    return "", None
                value_str = str(value_str).strip()
                match = re.match(r"^([<>=≤≥]+)?\s*([\d.]+)$", value_str)
                if match:
                    return match.group(1) or "", float(match.group(2))
                try:
                    return "", float(value_str)
                except ValueError:
                    return "", None

            def extract_site_code(accession_no):
                """Extract 3-letter site code from accession number (if exists)."""
                for code in site_codes:
                    if re.search(rf"{code}", str(accession_no), re.IGNORECASE):
                        return code
                return ""

            def parse_batch_info(batch_name):
                """
                Parse batch-related details from a batch name like:
                '1.1 GMH_09122019_1.1_0001-0009'
                """
                if not batch_name or pd.isna(batch_name):
                    return {"BatchNo": "", "TotalBatch": "", "RefNo": ""}
                batch_name = str(batch_name)
                batch_match = re.search(r"(\d+)\.(\d+)", batch_name)
                range_match = re.search(r"_(\d{4}-\d{4})", batch_name)
                return {
                    "BatchNo": batch_match.group(1) if batch_match else "",
                    "TotalBatch": batch_match.group(2) if batch_match else "",
                    "RefNo": range_match.group(1) if range_match else "",
                }

            # --- Process each row ---
            for row in rows:
                cleaned_row = {k: ("" if pd.isna(v) else v) for k, v in row.items()}

                accession = cleaned_row.get("AccessionNo") or cleaned_row.get("ID_Number")
                if not accession:
                    continue

                # Extract site and batch info
                site_code = extract_site_code(accession)
                batch_name = cleaned_row.get("Batch_Name", "")
                batch_info = parse_batch_info(batch_name)

                cleaned_row.update({
                    "Site_Code": site_code,
                    "BatchNo": batch_info["BatchNo"],
                    "Total_Batch": batch_info["TotalBatch"],
                    "RefNo": batch_info["RefNo"],
                })

                # Keep only model fields
                valid_fields = {k: v for k, v in cleaned_row.items() if k in model_fields}

            for field_name, value in valid_fields.items():
                if isinstance(value, str) and len(value) > 255:
                    print(f"[WARNING] {accession} - Field '{field_name}' exceeds 255 chars ({len(value)} chars)")
                # --- Create or update Referred_Data record ---
                ref_obj, ref_created = Referred_Data.objects.update_or_create(
                    AccessionNo=str(accession).strip(),
                    defaults=valid_fields,
                )
                created_ref += int(ref_created)
                updated_ref += int(not ref_created)

                # --- Antibiotic Entries ---
                for abx in known_abx:
                    abx_val = str(cleaned_row.get(abx, "")).strip()
                    abx_ris = str(cleaned_row.get(f"{abx}_RIS", "")).strip()
                    abx_rt_val = str(cleaned_row.get(f"{abx}_RT", "")).strip()
                    abx_rt_ris = str(cleaned_row.get(f"{abx}_RT_RIS", "")).strip()

                    if not any([abx_val, abx_ris, abx_rt_val, abx_rt_ris]):
                        continue

                    mic_op, mic_val = parse_mic_value(abx_val)
                    ret_op, ret_val = parse_mic_value(abx_rt_val)



                    ab_entry, ab_created = AntibioticEntry.objects.update_or_create(
                        ab_idNum_referred=ref_obj,
                        ab_Abx_code=abx,
                        defaults={
                            "ab_MIC_operand": mic_op,
                            "ab_MIC_value": mic_val,
                            "ab_MIC_RIS": abx_ris,
                            "ab_MIC_enRIS": abx_ris,
                            "ab_Retest_MIC_operand": ret_op,
                            "ab_Retest_MICValue": ret_val,
                            "ab_Retest_MIC_RIS": abx_rt_ris,
                            "ab_Retest_MIC_enRIS": abx_rt_ris,
                        },
                    )
                    created_abx += int(ab_created)
                    updated_abx += int(not ab_created)

            # --- Success message ---
            messages.success(
                request,
                f"Upload complete! "
                f"{created_ref} new Referred_Data, {updated_ref} updated; "
                f"{created_abx} new AntibioticEntry, {updated_abx} updated."
            )
            return redirect("show_data")

        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f" Error processing file: {e}")

    # --- Default render (GET request) ---
    return render(request, "wgs_app/Add_wgs.html", {
        "referred_form": referred_form,
        "form": form,
        "bactscout_form": BactScoutUploadForm(),
        "gambit_form": GambitUploadForm(),
        "mlst_form": MlstUploadForm(),
        "checkm2_form": Checkm2UploadForm(),
        "assembly_form": AssemblyUploadForm(),
        "amrfinder_form": AmrUploadForm(),
    })




############ FIELD MAPPER TOOL ############

 # this is the updated field mapper tool with temp file saving and session management
# @login_required
# def field_mapper_tool(request):
#     # upload a raw file and preview headers for mapping.
#     if request.method == "POST" and request.FILES.get("raw_file"):
#         uploaded_file = request.FILES["raw_file"]

#         # --- Read file to extract headers ---
#         try:
#             if uploaded_file.name.endswith(".csv"):
#                 df = pd.read_csv(uploaded_file, nrows=1)
#             else:
#                 df = pd.read_excel(uploaded_file, nrows=1)
#         except Exception as e:
#             messages.error(request, f"Error reading file: {e}")
#             return redirect("field_mapper_tool")

#         raw_headers = df.columns.tolist()

#         # --- Save file temporarily to session ---
#         # Create temp directory if it doesn't exist
#         temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_uploads')
#         os.makedirs(temp_dir, exist_ok=True)
        
#         # Generate unique filename
#         temp_filename = f"{request.user.id}_{uploaded_file.name}"
#         temp_filepath = os.path.join(temp_dir, temp_filename)
        
#         # Save file
#         with open(temp_filepath, 'wb+') as destination:
#             for chunk in uploaded_file.chunks():
#                 destination.write(chunk)
        
#         # Store path in session
#         request.session['temp_file_path'] = temp_filepath
#         request.session['temp_file_name'] = uploaded_file.name

#         # --- Get model field lists ---
#         final_fields = [f.name for f in Final_Data._meta.fields if f.name != "id"]
#         abx_fields = list(
#             Antibiotic_List.objects.filter(Retest=True)
#             .values_list("Whonet_Abx", flat=True)
#             .distinct().order_by("Whonet_Abx")
#         )


#         # --- Load saved mappings ---
#         saved_mappings = FieldMapping.objects.filter(user=request.user)
#         saved_dict = {m.raw_field: m.mapped_field for m in saved_mappings}

#         context = {
#             "raw_headers": raw_headers,
#             "final_fields": final_fields,
#             "abx_fields": abx_fields,
#             "saved_mappings": saved_dict,
#             "file_name": uploaded_file.name,
#         }

#         return render(request, "home/map_fields.html", context)


#     return render(request, "home/upload_raw.html")



@login_required
def field_mapper_tool(request):

    if request.method == "POST" and request.FILES.get("raw_file"):

        uploaded_file = request.FILES["raw_file"]
        target_model = request.POST.get("target_model", "final")
        request.session["target_model"] = target_model

        # -------- READ FILE HEADERS --------
        try:
            df = read_tabular_upload(uploaded_file, nrows=1)

        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect("field_mapper_tool")

        raw_headers = df.columns.tolist()

        # -------- SAVE TEMP FILE --------
        temp_dir = os.path.join(settings.MEDIA_ROOT, "temp_uploads")
        os.makedirs(temp_dir, exist_ok=True)

        temp_filename = f"{request.user.id}_{uploaded_file.name}"
        temp_filepath = os.path.join(temp_dir, temp_filename)

        with open(temp_filepath, "wb+") as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        request.session["temp_file_path"] = temp_filepath
        request.session["temp_file_name"] = uploaded_file.name

        # -------- SELECT MODEL FIELDS --------

        if target_model == "referred":

            model_fields = [
                f.name for f in Referred_Data._meta.fields
                if f.name != "id"
            ]

            antibiotic_fields = list(
                Antibiotic_List.objects
                .values_list("Whonet_Abx", flat=True)
                .distinct()
                .order_by("Whonet_Abx")
            )

        else:

            model_fields = [
                f.name for f in Final_Data._meta.fields
                if f.name != "id"
            ]

            antibiotic_fields = list(
                Antibiotic_List.objects
                .values_list("Whonet_Abx", flat=True)
                .distinct()
                .order_by("Whonet_Abx")
            )

        # -------- ANTIBIOTIC FIELDS --------
        abx_fields = list(
            Antibiotic_List.objects.filter(Retest=True)
            .values_list("Whonet_Abx", flat=True)
            .distinct()
            .order_by("Whonet_Abx")
        )

        # -------- LOAD SAVED MAPPINGS --------
        saved_mappings = FieldMapping.objects.filter(user=request.user)
        saved_dict = {m.raw_field: m.mapped_field for m in saved_mappings}
        saved_retest_headers = {
            m.raw_field
            for m in saved_mappings
            if m.is_retest
        }

        context = {
            "raw_headers": raw_headers,
            "model_fields": model_fields,
            "abx_fields": antibiotic_fields,
            "saved_mappings": saved_dict,
            "saved_retest_headers": saved_retest_headers,
            "file_name": uploaded_file.name,
            "target_model": target_model,
        }

        return render(request, "home/map_fields.html", context)

    return render(request, "home/upload_raw.html")


# AJAX endpoint to save/update a field mapping used in the field mapper tool
# @login_required
# @require_POST
# def update_field_mapping(request):
#     import json

#     data = json.loads(request.body)
#     raw_field = data.get("raw_field")
#     mapped_field = data.get("mapped_field", "").strip()

#     if not raw_field:
#         return JsonResponse({"status": "error", "msg": "Missing raw_field"}, status=400)

#     if mapped_field == "":
#         #  Clear mapping
#         FieldMapping.objects.filter(
#             user=request.user,
#             raw_field=raw_field
#         ).delete()
#     else:
#         FieldMapping.objects.update_or_create(
#             user=request.user,
#             raw_field=raw_field,
#             defaults={
#                 "mapped_field": mapped_field
#             }
#         )

#     return JsonResponse({"status": "ok"})


@login_required
@require_POST
def update_field_mapping(request):
    import json

    data = json.loads(request.body)

    raw_field = data.get("raw_field")
    mapped_field = data.get("mapped_field", "").strip()
    is_retest = data.get("is_retest", False)

    if not raw_field:
        return JsonResponse({"status": "error", "msg": "Missing raw_field"}, status=400)

    if mapped_field == "":
        FieldMapping.objects.filter(
            user=request.user,
            raw_field=raw_field
        ).delete()

    else:
        FieldMapping.objects.update_or_create(
            user=request.user,
            raw_field=raw_field,
            defaults={
                "mapped_field": mapped_field,
                "is_retest": is_retest
            }
        )

    return JsonResponse({"status": "ok"})




# clear all saved mappings for the user
@login_required
def clear_mappings(request):

    if request.method == "POST":
        FieldMapping.objects.filter(user=request.user).delete()
        messages.success(request, "Your saved field mappings were cleared.")
    return redirect("field_mapper_tool")


# helped function to delete temp file after download
@login_required
def cleanup_temp_file(file_path, request):
   
    if file_path and os.path.exists(file_path):
            os.remove(file_path)


@login_required
def download_mapping_summary(request):
    # function to normalize strings for comparison 
    def normalize(val):
        return re.sub(r"[^a-z0-9]", "", str(val).lower())

   # load uploaded file headers
    temp_path = request.session.get("temp_file_path")
    original_name = request.session.get("temp_file_name", "mapping_summary")

    if not temp_path or not os.path.exists(temp_path):
        messages.error(request, "Temporary file not found. Please upload again.")
        return redirect("field_mapper_tool")

    if temp_path.lower().endswith(".csv"):
        df = pd.read_csv(temp_path, nrows=0)
    elif temp_path.lower().endswith(".tsv"):
        df = pd.read_csv(temp_path, sep="\t", nrows=0)
    else:
        df = pd.read_excel(temp_path, nrows=0)

    uploaded_headers = list(df.columns)

  # load saved mappings
    mappings = FieldMapping.objects.filter(user=request.user)

    mapped_norm = {
        normalize(m.raw_field) for m in mappings
    }


    # load database columns
    
    # final_fields = [
    #     f.name for f in Final_Data._meta.fields
    #     if f.name != "id"
    # ]

    target_model = request.session.get("target_model", "final")

    if target_model == "referred":
        model_fields = [
            f.name for f in Referred_Data._meta.fields
            if f.name != "id"
        ]

        abx_fields = list(
            Antibiotic_List.objects
            .values_list("Whonet_Abx", flat=True)
            .distinct()
            .order_by("Whonet_Abx")
        )

    else:
        model_fields = [
            f.name for f in Final_Data._meta.fields
            if f.name != "id"
        ]

        abx_fields = list(
            Antibiotic_List.objects
            .values_list("Whonet_Abx", flat=True)
            .distinct()
            .order_by("Whonet_Abx")
        )

    # build antibiotic column variants
    abx_columns = []

    for abx in abx_fields:
        abx_columns.extend([
            f"{abx}_NM",
            f"{abx}_NM_OP",
            f"{abx}_NM_RIS",
            f"{abx}_ND30",
            f"{abx}_ND30_RIS"
        ])

    database_columns = sorted(set(model_fields + abx_columns))
    # database_columns = sorted(set(final_fields + abx_fields))

    db_norm = {
        normalize(c) for c in database_columns
    }

    # determine unmapped headers
    uploaded_norm_map = {
        normalize(h): h for h in uploaded_headers
    }

    unmapped_headers = [
        original_name
        for norm, original_name in uploaded_norm_map.items()
        if norm not in mapped_norm and norm not in db_norm
    ]

   # build dataframes for each sheet
    df_db = pd.DataFrame({
        "Database_Column_Name": database_columns
    })

    df_uploaded = pd.DataFrame({
        "Uploaded_Column_Name": uploaded_headers
    })

    df_unmapped = pd.DataFrame({
        "Unmapped_Column_Name": unmapped_headers
    })

    # create 3 sheet excel file
    safe_name = os.path.splitext(original_name)[0]
    filename = f"{safe_name}_FIELD_MAPPING_SUMMARY.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    with pd.ExcelWriter(response, engine="openpyxl") as writer:
        df_db.to_excel(writer, index=False, sheet_name="Database_Columns")
        df_uploaded.to_excel(writer, index=False, sheet_name="Uploaded_Columns")
        df_unmapped.to_excel(writer, index=False, sheet_name="Unmapped_Columns")

    return response




def extract_antibiotics(df):
    """
    Antibiotic extraction rules (STRICT):
    - Disk:     <ABX>_ND<potency>
    - Disk RIS: <ABX>_ND<potency>_RIS
    - MIC OP:   <ABX>_NM_OP
    - MIC:      <ABX>_NM
    - MIC RIS:  <ABX>_NM_RIS
    """

    def clean_series(s):
        if isinstance(s, pd.DataFrame):
            s = s.iloc[:, 0]
        if not isinstance(s, pd.Series):
            s = pd.Series(s)
        return (
            s.replace([None, "None", "nan", "NaN", "NAN", "null"], "")
             .fillna("")
             .astype(str)
        )

    def split_operand(val):
        if not val:
            return "", ""
        m = re.match(r"^(<=|>=|<|>|=)?\s*([\d.]+)$", str(val).strip())
        return (m.group(1) or "", m.group(2)) if m else ("", val)

    cols = list(df.columns)
    abx_df = pd.DataFrame(index=df.index)

   # disk VALUES + RIS no operands
    disk_vals = [
        c for c in cols
        if re.fullmatch(r"[A-Za-z]+_ND(?:\d+(?:[._]\d+)?)?", c, re.I)
    ]

    for col in disk_vals:
        abx_df[col.upper()] = clean_series(df[col]).str.upper()

        ris_col = f"{col}_RIS"
        if ris_col in df.columns:
            abx_df[f"{col.upper()}_RIS"] = clean_series(df[ris_col]).str.upper()
        else:
            abx_df[f"{col.upper()}_RIS"] = ""

    # mic VALUES + OPS + RIS
    mic_vals = [c for c in cols if re.fullmatch(r"[A-Za-z]+_NM", c)]

    for col in mic_vals:
        base = col[:-3]  # remove _NM

        ops, vals = [], []
        for v in clean_series(df[col]):
            op, val = split_operand(v)
            ops.append(op)
            vals.append(val.upper())

        abx_df[f"{base.upper()}_NM_OP"] = ops
        abx_df[f"{base.upper()}_NM"] = vals

        ris_col = f"{col}_RIS"
        if ris_col in df.columns:
            abx_df[f"{base.upper()}_NM_RIS"] = clean_series(df[ris_col]).str.upper()
        else:
            abx_df[f"{base.upper()}_NM_RIS"] = ""

    return abx_df



# WORKING VERSION AND CORRECTED ERROR IN DEMOGRAPHICS MAPPING
# @login_required
# def generate_mapped_excel(request):
#     if request.method != "POST":
#         return redirect("field_mapper_tool")

#     try:
#         import os, io, re
#         import pandas as pd
#         from django.http import HttpResponse
#         from django.contrib import messages
#         from apps.home.models import FieldMapping  # this ensures the model is imported

#         # -HELPER FUNCTIONS-

#         # Clean series function to standardize data formatting and missing values
#         def clean_series(s):
#             if isinstance(s, pd.DataFrame):
#                 s = s.iloc[:, 0]
#             if not isinstance(s, pd.Series):
#                 s = pd.Series(s)
#             return s.replace([None, "None", "nan", "NaN", "NAN", "null"], "").fillna("").astype(str)

#         # Split operand function for MIC values
#         def split_operand(val):
#             if not val: return "", ""
#             m = re.match(r"^(<=|>=|<|>|=)?\s*([\d.]+)$", str(val).strip())
#             return (m.group(1) or "", m.group(2)) if m else ("", val)

#         target_model = request.session.get("target_model", "final")

#         # --- Load File ---
#         temp_file_path = request.session.get("temp_file_path")
#         temp_file_name = request.session.get("temp_file_name", "uploaded.xlsx")

        

#         if not temp_file_path or not os.path.exists(temp_file_path):
#             messages.error(request, "File not found.")
#             return redirect("field_mapper_tool")

#         df = pd.read_csv(temp_file_path) if temp_file_name.lower().endswith(".csv") else pd.read_excel(temp_file_path)
#         df.columns = [str(c).strip() for c in df.columns]
        
        
#         # Build original col_map BEFORE any renaming
#         col_map = {c.lower(): c for c in df.columns}

#         # accession and year extraction
#         acc_col = next((c for c in df.columns if "accession" in c.lower()), None)
#         spec_date_col = next((c for c in df.columns if "spec" in c.lower() and "date" in c.lower()), None)
#         year_vals = [""] * len(df)
#         if spec_date_col:
#             year_vals = pd.to_datetime(df[spec_date_col], errors="coerce").dt.year.apply(lambda x: "" if pd.isna(x) else str(int(x))).tolist()

#         # build antibiotic entries
#         abx_df = pd.DataFrame()


#         if target_model == "referred":
#             accession_field = "AccessionNo"
#         else:
#             accession_field = "f_AccessionNo"

#         abx_df[accession_field] = clean_series(df[acc_col]) if acc_col else ""
#         abx_df["Year"] = year_vals

#         # Disk & MIC Loops - Using col_map to find real column names - ensures disk antibiotics are captured correctly

#         # Disk columns
#         disk_cols = [c for c in col_map if re.fullmatch(r"[a-z]+_nd\d+", c)]
#         for lc in disk_cols:
#             real = col_map[lc]
#             ris_lc = f"{lc}_ris"
#             raw_vals = df[real]
#             if isinstance(raw_vals, pd.DataFrame): raw_vals = raw_vals.iloc[:, 0]
#             disk_vals = pd.to_numeric(raw_vals, errors="coerce").astype("Int64").astype(str)
#             disk_vals = ["" if v == "<NA>" else v for v in disk_vals]
#             abx_df[real.upper()] = disk_vals
#             if ris_lc in col_map:
#                 ris_series = df[col_map[ris_lc]]
#                 if isinstance(ris_series, pd.DataFrame): ris_series = ris_series.iloc[:, 0]
#                 abx_df[f"{real.upper()}_RIS"] = clean_series(ris_series).str.upper()
#             else:
#                 abx_df[f"{real.upper()}_RIS"] = ""
#         # MIC columns
#         mic_cols = [c for c in col_map if re.fullmatch(r"[a-z]+_nm", c)]
#         for lc in mic_cols:
#             real = col_map[lc]
#             base = real[:-3]
#             ris_lc = f"{lc}_ris"
#             target_col = df[real]
#             if isinstance(target_col, pd.DataFrame): target_col = target_col.iloc[:, 0]
#             ops, vals = [], []
#             for v in clean_series(target_col):
#                 op, val = split_operand(v)
#                 ops.append(op)
#                 vals.append(val.upper())
#             abx_df[f"{base.upper()}_NM_OP"] = ops
#             abx_df[f"{base.upper()}_NM"] = vals
#             if ris_lc in col_map:
#                 ris_series = df[col_map[ris_lc]]
#                 if isinstance(ris_series, pd.DataFrame): ris_series = ris_series.iloc[:, 0]
#                 abx_df[f"{base.upper()}_NM_RIS"] = clean_series(ris_series).str.upper()
#             else:
#                 abx_df[f"{base.upper()}_NM_RIS"] = ""

#         # Build demogs dataframe
#         user_mappings = FieldMapping.objects.filter(user=request.user, mapped_field__isnull=False)
#         rename_dict = {m.raw_field: m.mapped_field for m in user_mappings}

#         # Filter for strictly non-antibiotic columns
#         demogs_cols = [
#             c for c in df.columns 
#             if not re.search(r"_(nd\d+|nm)(_ris)?$", c, re.I)
#         ]

#         # Create the demogs slice and rename them based on model fields
#         demogs_df = df[demogs_cols].copy()
#         demogs_df.rename(columns=rename_dict, inplace=True)

#         # Clean all data in demogs
#         for c in demogs_df.columns:
#             demogs_df[c] = clean_series(demogs_df[c])

#        # output to excel
#         output = io.BytesIO()

#         with pd.ExcelWriter(output, engine="openpyxl") as writer:

#             if target_model == "referred":
#                 demogs_sheet = "Referred_Demogs"
#                 abx_sheet = "AntibioticEntry"
#             else:
#                 demogs_sheet = "Final_Demogs"
#                 abx_sheet = "Final_AntibioticEntry"

#             demogs_df.to_excel(writer, index=False, sheet_name=demogs_sheet)
#             abx_df.to_excel(writer, index=False, sheet_name=abx_sheet)

#         output.seek(0)
#         response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#         base_name = os.path.splitext(temp_file_name)[0]
#         response["Content-Disposition"] = f'attachment; filename="{base_name}_Mapped.xlsx"'
#         return response

#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#         messages.error(request, f"⚠️ Error: {e}")
#         return redirect("field_mapper_tool")





@login_required
def generate_mapped_excel(request):

    if request.method != "POST":
        return redirect("field_mapper_tool")

    try:

        import os
        import io
        import re
        import json
        import pandas as pd
        from django.http import HttpResponse
        from django.contrib import messages
        from apps.home.models import FieldMapping


        # ---------------- HELPER FUNCTIONS ---------------- #

        def clean_series(s):

            if isinstance(s, pd.DataFrame):
                s = s.iloc[:, 0]

            if not isinstance(s, pd.Series):
                s = pd.Series(s)

            return (
                s.replace(
                    [None, "None", "nan", "NaN", "NAN", "null"],
                    ""
                )
                .fillna("")
                .astype(str)
            )


        def split_operand(val):

            if not val:
                return "", ""

            m = re.match(r"^(<=|>=|<|>|=)?\s*([\d.]+)$", str(val).strip())

            if m:
                return m.group(1) or "", m.group(2)

            return "", val

        def normalize_field_key(value):
            return re.sub(r"[^a-z0-9]", "", str(value).lower())

        def normalize_abx_export_name(value):
            text = str(value or "").strip().upper()
            text = re.sub(r"[\s-]+", "_", text)
            text = re.sub(r"_+", "_", text)
            return text

        def find_column_by_mapped_field(rename_dict, target_fields, fallback_patterns=()):
            normalized_targets = {normalize_field_key(field) for field in target_fields}

            for raw_field, mapped_field in rename_dict.items():
                if normalize_field_key(mapped_field) in normalized_targets:
                    raw_key = normalize_field_key(raw_field)
                    for column in df.columns:
                        if normalize_field_key(column) == raw_key:
                            return column

            for column in df.columns:
                column_key = normalize_field_key(column)
                if any(pattern in column_key for pattern in fallback_patterns):
                    return column

            return None

        def year_series_from_column(column_name):
            if not column_name or column_name not in df.columns:
                return pd.Series([""] * len(df), dtype="object")

            parsed_dates = pd.to_datetime(df[column_name], errors="coerce")
            return parsed_dates.dt.year.apply(
                lambda year: "" if pd.isna(year) else str(int(year))
            )

        def extract_ref_range(value):
            if value is None:
                return ""

            raw_value = str(value).strip()
            if raw_value.lower() in {"", "nan", "nat", "none", "null"}:
                return ""

            match = re.search(r"_(\d{3,6}(?:-\d{3,6})?)\s*$", raw_value)
            return match.group(1) if match else ""

        def extract_accession_ref(value):
            if value is None:
                return ""

            match = re.search(r"(\d+)\s*$", str(value).strip())
            return match.group(1) if match else ""

        def get_demogs_series(frame, column_name):
            if column_name not in frame.columns:
                return None

            values = frame[column_name]
            if isinstance(values, pd.DataFrame):
                values = values.iloc[:, 0]

            return values

        def mapped_abx_target(raw_column, default_base, *, strip_method=False):
            default_base = normalize_abx_export_name(default_base)
            mapped = rename_dict.get(raw_column) or rename_dict.get(raw_column.lower())
            if not mapped:
                return default_base

            mapped = normalize_abx_export_name(mapped)
            if not mapped:
                return default_base

            mapped = re.sub(r"(_RT)?(_VAL|_RIS|_OP)$", "", mapped)
            if strip_method:
                mapped = re.sub(r"_NM(_RT)?$", "", mapped)
                mapped = re.sub(r"_ND(?:\d+(?:[._]\d+)?)?(_RT)?$", "", mapped)
            return mapped or default_base

        def is_retest_column(*raw_columns):
            return any(str(raw).lower() in retest_fields for raw in raw_columns if raw)

        def is_antibiotic_source_column(column_name):
            column_key = str(column_name or "").strip()
            return bool(
                re.fullmatch(
                    r"[A-Za-z]+_ND(?:\d+(?:[._]\d+)?)?(?:_RIS)?",
                    column_key,
                    re.I,
                )
                or re.fullmatch(
                    r"[A-Za-z]+_NM(?:_RIS)?",
                    column_key,
                    re.I,
                )
            )


        # ---------------- TARGET MODEL ---------------- #

        target_model = request.session.get("target_model", "final")


        # ---------------- LOAD FILE ---------------- #

        temp_file_path = request.session.get("temp_file_path")
        temp_file_name = request.session.get("temp_file_name", "uploaded.xlsx")

        if not temp_file_path or not os.path.exists(temp_file_path):
            messages.error(request, "File not found.")
            return redirect("field_mapper_tool")

        if temp_file_name.lower().endswith(".csv"):
            df = pd.read_csv(temp_file_path)
        elif temp_file_name.lower().endswith(".tsv"):
            df = pd.read_csv(temp_file_path, sep="\t")
        else:
            df = pd.read_excel(temp_file_path)

        df.columns = [str(c).strip() for c in df.columns]


        # ---------------- ORIGINAL COLUMN MAP ---------------- #

        col_map = {c.lower(): c for c in df.columns}


        # ---------------- ACCESSION ---------------- #

        acc_col = next(
            (c for c in df.columns if "accession" in c.lower()),
            None
        )


        if target_model == "referred":
            accession_field = "AccessionNo"
        else:
            accession_field = "f_AccessionNo"

        # Collect all antibiotic export columns first, then build the DataFrame
        # once. Repeated DataFrame inserts fragment pandas internals on large
        # antibiotic panels.
        abx_columns = {
            accession_field: (
                clean_series(df[acc_col]) if acc_col else pd.Series([""] * len(df))
            ),
            "Year": pd.Series([""] * len(df), dtype="object"),
        }


        # ---------------- USER FIELD MAPPINGS ---------------- #

        user_mappings = FieldMapping.objects.filter(
            user=request.user,
            mapped_field__isnull=False
        )

        saved_rename_dict = {
            m.raw_field: m.mapped_field
            for m in user_mappings
        }

        try:
            posted_mapping = json.loads(request.POST.get("mapping", "{}") or "{}")
        except (TypeError, ValueError):
            posted_mapping = {}

        posted_mapping = {
            str(raw): str(mapped).strip()
            for raw, mapped in posted_mapping.items()
            if str(mapped).strip()
        }

        try:
            posted_retest_headers = json.loads(
                request.POST.get("retest_headers", "[]") or "[]"
            )
        except (TypeError, ValueError):
            posted_retest_headers = []

        posted_retest_headers = {
            str(header).strip().lower()
            for header in posted_retest_headers
            if str(header).strip()
        }

        if posted_mapping:
            rename_dict = {**saved_rename_dict, **posted_mapping}
            retest_fields = {
                raw.lower(): mapped
                for raw, mapped in posted_mapping.items()
                if raw.lower() in posted_retest_headers
            }
        else:
            rename_dict = saved_rename_dict
            retest_fields = {
                m.raw_field.lower(): m.mapped_field
                for m in user_mappings if m.is_retest
            }

        spec_date_col = find_column_by_mapped_field(
            rename_dict,
            ["Spec_Date", "f_Spec_Date"],
            fallback_patterns=("specdate",),
        )
        year_values = year_series_from_column(spec_date_col)
        abx_columns["Year"] = year_values

        # =====================================================
        # DISK ANTIBIOTICS
        # =====================================================

        disk_cols = [
            c for c in col_map
            if re.fullmatch(r"[a-z]+_nd(?:\d+(?:[._]\d+)?)?", c)
        ]

        for lc in disk_cols:

            real = col_map[lc]
            ris_lc = f"{lc}_ris"

            raw_vals = df[real]

            if isinstance(raw_vals, pd.DataFrame):
                raw_vals = raw_vals.iloc[:, 0]

            disk_vals = (
                pd.to_numeric(raw_vals, errors="coerce")
                .astype("Int64")
                .astype(str)
            )

            disk_vals = ["" if v == "<NA>" else v for v in disk_vals]

            ris_real = col_map.get(ris_lc)
            default_disk_base = (
                mapped_abx_target(ris_real, real.upper())
                if ris_real
                else real.upper()
            )
            base_name = mapped_abx_target(real, default_disk_base)

            # RETEST DETECTION
            if is_retest_column(lc, real):
                base_name = f"{base_name}_RT"

            abx_columns[f"{base_name}_Val"] = pd.Series(disk_vals)

            if ris_lc in col_map:

                ris_series = df[col_map[ris_lc]]

                if isinstance(ris_series, pd.DataFrame):
                    ris_series = ris_series.iloc[:, 0]

                abx_columns[f"{base_name}_RIS"] = clean_series(ris_series).str.upper()

            else:
                abx_columns[f"{base_name}_RIS"] = pd.Series([""] * len(df))


        # =====================================================
        # MIC ANTIBIOTICS
        # =====================================================

        mic_cols = [
            c for c in col_map
            if re.fullmatch(r"[a-z]+_nm", c)
        ]

        for lc in mic_cols:

            real = col_map[lc]
            base = real[:-3]

            ris_lc = f"{lc}_ris"

            target_col = df[real]

            if isinstance(target_col, pd.DataFrame):
                target_col = target_col.iloc[:, 0]

            ops = []
            vals = []

            for v in clean_series(target_col):

                op, val = split_operand(v)

                ops.append(op)
                vals.append(val.upper())

            ris_real = col_map.get(ris_lc)
            default_mic_base = (
                mapped_abx_target(ris_real, base.upper(), strip_method=True)
                if ris_real
                else base.upper()
            )
            base_name = mapped_abx_target(
                real,
                default_mic_base,
                strip_method=True,
            )
            mic_base_name = f"{base_name}_NM"

            if is_retest_column(lc, real):
                mic_base_name = f"{base_name}_NM_RT"

            abx_columns[f"{mic_base_name}_OP"] = pd.Series(ops)
            abx_columns[f"{mic_base_name}_Val"] = pd.Series(vals)

            if ris_lc in col_map:

                ris_series = df[col_map[ris_lc]]

                if isinstance(ris_series, pd.DataFrame):
                    ris_series = ris_series.iloc[:, 0]

                abx_columns[f"{mic_base_name}_RIS"] = clean_series(ris_series).str.upper()

            else:
                abx_columns[f"{mic_base_name}_RIS"] = pd.Series([""] * len(df))

        abx_df = pd.DataFrame(abx_columns).copy()


        # =====================================================
        # DEMOGRAPHICS
        # =====================================================

        demogs_cols = [
            c for c in df.columns
            if not is_antibiotic_source_column(c)
        ]

        demogs_df = df[demogs_cols].copy()

        demogs_df.rename(columns=rename_dict, inplace=True)

        for c in demogs_df.columns:
            demogs_df[c] = clean_series(demogs_df[c])

        default_year_field = "Default_Year" if target_model == "referred" else "f_Default_Year"
        if target_model == "referred" or default_year_field in demogs_df.columns:
            demogs_df[default_year_field] = year_values.astype(str)

        batch_name_field = "Batch_Name" if target_model == "referred" else "f_Batch_Name"
        batch_code_field = "Batch_Code" if target_model == "referred" else "f_Batch_Code"
        ref_no_field = "RefNo" if target_model == "referred" else "f_RefNo"
        accession_no_field = "AccessionNo" if target_model == "referred" else "f_AccessionNo"

        batch_name_series = get_demogs_series(demogs_df, batch_name_field)
        batch_code_series = get_demogs_series(demogs_df, batch_code_field)

        if batch_name_series is not None or batch_code_series is not None:
            if batch_name_series is None:
                batch_name_series = pd.Series(
                    [""] * len(demogs_df),
                    index=demogs_df.index,
                )
            if batch_code_series is None:
                batch_code_series = pd.Series(
                    [""] * len(demogs_df),
                    index=demogs_df.index,
                )

            batch_ref_values = pd.Series(
                [
                    extract_ref_range(batch_name) or extract_ref_range(batch_code)
                    for batch_name, batch_code in zip(
                        batch_name_series,
                        batch_code_series,
                    )
                ],
                index=demogs_df.index,
            )

            accession_no_series = get_demogs_series(demogs_df, accession_no_field)
            if accession_no_series is not None:
                accession_ref_values = accession_no_series.apply(extract_accession_ref)
            else:
                accession_ref_values = pd.Series([""] * len(demogs_df), index=demogs_df.index)

            demogs_df[ref_no_field] = pd.Series([
                batch_ref or accession_ref
                for batch_ref, accession_ref in zip(batch_ref_values, accession_ref_values)
            ], index=demogs_df.index)


        # =====================================================
        # OUTPUT EXCEL
        # =====================================================

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            if target_model == "referred":

                demogs_sheet = "Referred_Demogs"
                abx_sheet = "AntibioticEntry"

            else:

                demogs_sheet = "Final_Demogs"
                abx_sheet = "Final_AntibioticEntry"

            demogs_df.to_excel(
                writer,
                index=False,
                sheet_name=demogs_sheet
            )

            abx_df.to_excel(
                writer,
                index=False,
                sheet_name=abx_sheet
            )

        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        base_name = os.path.splitext(temp_file_name)[0]

        response["Content-Disposition"] = f'attachment; filename="{base_name}_Mapped.xlsx"'

        return response


    except Exception as e:

        import traceback
        traceback.print_exc()

        messages.error(request, f"⚠️ Error: {e}")

        return redirect("field_mapper_tool")





############# Antibiotics Configuration

@login_required(login_url="login")
def add_antibiotics(request):
    if request.method == "POST":
        form = AntibioticsForm(request.POST)

        if form.is_valid():
            antibiotic = form.save()
            messages.success(
                request,
                f"Antibiotic '{antibiotic.Antibiotic}' added successfully."
            )
        else:
            messages.warning(request, _first_form_error(form, "Form validation failed. Please check your inputs."))

    # Always go back to Settings → Antibiotics tab
    return redirect("/settings/?tab=antibiotics")


@login_required(login_url="login")
def edit_antibiotics(request, pk):
    antibiotic = get_object_or_404(Antibiotic_List, pk=pk)
    abx_upload_form = Antibiotics_uploadForm()  
    if request.method == "POST":
        form = AntibioticsForm(request.POST, instance=antibiotic)

        if form.is_valid():
            form.save()
            messages.success(
                request,
                f"Antibiotic '{antibiotic.Antibiotic}' updated successfully."
            )

            return redirect("/settings/?tab=antibiotics")

        messages.error(request, "Form validation failed. Please check your inputs.")
    else:
        form = AntibioticsForm(instance=antibiotic)

    return render(
        request,
        "home/Antibiotic_list.html",   # SEPARATE EDIT PAGE
        {
            "form": form,
            "antibiotic": antibiotic,
            "abx_upload_form": abx_upload_form,
            "editing": True,        # to indicate edit mode # optional but useful
        },
    )


@login_required(login_url="login")
def antibiotics_view(request):
    q = request.GET.get("q", "").strip()

    antibiotics = Antibiotic_List.objects.all().order_by("Whonet_Abx")

    if q:
        antibiotics = antibiotics.filter(
            Q(Antibiotic__icontains=q) |
            Q(Whonet_Abx__icontains=q) |
            Q(Abx_code__icontains=q) |
            Q(Guidelines__icontains=q) |
            Q(Test_Method__icontains=q) |
            Q(Potency__icontains=q) |
            Q(Tier__icontains=q) |
            Q(Class__icontains=q) |
            Q(Subclass__icontains=q)
        )

    paginator = Paginator(antibiotics, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "home/Antibiotic_View.html",
        {
            "antibiotics": antibiotics,
            "page_obj": page_obj,
            "q": q,
        },
    )



@login_required(login_url="login")
#Delete breakpoints
def antibiotics_del(request, id):
    antibiotics = get_object_or_404(Antibiotic_List, pk=id)
    antibiotics.delete()
    return redirect('antibiotics_view')



@login_required(login_url="login")
# for uploading and replacing existing breakpoints data
def upload_antibiotics(request):
    if request.method == "POST":
        abx_upload_form = Antibiotics_uploadForm(request.POST, request.FILES)
        if abx_upload_form.is_valid():
            # Save the uploaded file instance
            uploaded_file = abx_upload_form.save()
            file = uploaded_file.File_uploadAbx  # Get the actual file field
            print("Uploaded file:", file)  # Debugging statement
            try:
                # Load file into a DataFrame using file's temporary path
                df = read_tabular_upload(file)

                # Check the DataFrame for debugging
                print(df)
                
                # Check the DataFrame for debugging
                print("DataFrame contents:\n", df.head())  # Print the first few rows

                # Check column and Replace NaN values with empty strings to avoid validation errors
                df.fillna(value={col: "" for col in df.columns}, inplace=True)


                created_count = 0
                duplicate_count = 0

                for _, row in df.iterrows():
                    whonet_abx = str(row.get('Whonet_Abx', '')).strip().upper()
                    if not whonet_abx:
                        continue
                    if Antibiotic_List.objects.filter(Whonet_Abx__iexact=whonet_abx).exists():
                        duplicate_count += 1
                        continue

                    # Parse Date_Modified if it's present and valid
                    date_modified = None
                    if row.get('Date_Modified'):
                        date_modified = pd.to_datetime(row['Date_Modified'], errors='coerce')
                        if pd.isna(date_modified):
                            date_modified = None

                    Antibiotic_List.objects.create(
                        Show=bool(row.get('Show', False)),
                        Show_Site=bool(row.get('Show_Site', False)),
                        Show_Ars=bool(row.get('Show_Ars', False)),
                        Show_Value=bool(row.get('Show_Value', False)),
                        Retest=bool(row.get('Retest', False)),
                        Disk_Abx=bool(row.get('Disk_Abx', False)),
                        Test_Method=row.get('Test_Method', ''),
                        Tier=row.get('Tier', ''),
                        Abx_code=row.get('Abx_code', ''),
                        Whonet_Abx=whonet_abx,
                        Antibiotic=row.get('Antibiotic', ''),
                        Guidelines=row.get('Guidelines', ''),
                        Potency=row.get('Potency', ''),
                        Class=row.get('Class', ''),
                        Subclass=row.get('Subclass', ''),
                        Date_Modified=date_modified,
                    )
                    created_count += 1

                
                messages.success(request, f"File uploaded. {created_count} created, {duplicate_count} skipped because they already exist.")
                return redirect('antibiotics_view')

            except Exception as e:
                print("Error during processing:", e)  # Debug statement
                messages.error(request, f"Error processing file: {e}")
                return redirect('add_antibiotics')
        else:
            messages.error(request, messages.INFO, "Form is not valid.")

    else:
        abx_upload_form = Antibiotics_uploadForm()

    return render(request, 'settings.html', {'abx_upload_form': abx_upload_form})




@login_required(login_url="login")
#for exporting into excel
def export_antibiotics(request):
    objects = Antibiotic_List.objects.all()
    data = []

    for obj in objects:
        data.append({
            "Show": obj.Show,
            "Retest": obj.Retest,
            "Show_Site": obj.Show_Site,
            "Show_Ars": obj.Show_Ars,
            "Show_Value": obj.Show_Value,
            "Disk_Abx": obj.Disk_Abx,
            "Guidelines": obj.Guidelines,
            "Tier": obj.Tier,
            "Test_Method": obj.Test_Method,
            "Potency": obj.Potency,
            "Abx_code": obj.Abx_code,
            "Whonet_Abx": obj.Whonet_Abx,
            "Antibiotic": obj.Antibiotic,
            "Class": obj.Class,
            "Subclass": obj.Subclass,
            "Date_Modified": obj.Date_Modified,
        })
    
    # Define file path
    file_path = "Antibiotic_list.xlsx"

    # Convert data to DataFrame and save as Excel
    df = pd.DataFrame(data)
    df.to_excel(file_path, index=False)

    # Return the file as a response
    return FileResponse(open(file_path, "rb"), as_attachment=True, filename="Antibiotic_list.xlsx")




@login_required(login_url="login")
def delete_all_antibiotics(request):
    Antibiotic_List.objects.all().delete()
    messages.success(request, "All records have been deleted successfully.")
    return redirect('antibiotics_view')  # Redirect to the table view




######################## Organism 
@login_required(login_url="login")
def add_organism(request):

    if request.method == "POST":
        org_form = OrganismForm(request.POST)

        if org_form.is_valid():
            org_form.save()
            messages.success(request, "Organism added successfully.")
        else:
            messages.warning(request, _first_form_error(org_form, "Form validation failed. Please check your inputs."))

    # Always return to Settings → Organisms tab
    return redirect("/settings/?tab=organisms")



@login_required(login_url="login")
def edit_organism(request, pk):
    organism = get_object_or_404(Organism_List, pk=pk)
    org_upload_form = Organism_uploadForm()  # keep upload support
    old_code = (organism.Whonet_Org_Code or "").strip()
    old_name = organism.Organism or ""

    if request.method == "POST":
        org_form = OrganismForm(request.POST, instance=organism)

        if org_form.is_valid():
            updated_organism = org_form.save()
            new_code = (updated_organism.Whonet_Org_Code or "").strip()
            new_name = updated_organism.Organism or ""

            raw_site_updated = raw_ars_updated = 0
            final_site_updated = final_ars_updated = 0

            if old_code and new_name and (old_code != new_code or old_name != new_name):
                raw_site_updated = Referred_Data.objects.filter(
                    Site_Org__iexact=old_code
                ).update(
                    Site_Org=new_code,
                    Site_OrgName=new_name,
                )
                raw_ars_updated = Referred_Data.objects.filter(
                    ars_OrgCode__iexact=old_code
                ).update(
                    ars_OrgCode=new_code,
                    ars_OrgName=new_name,
                )
                final_site_updated = Final_Data.objects.filter(
                    f_Site_Org__iexact=old_code
                ).update(
                    f_Site_Org=new_code,
                    f_Site_OrgName=new_name,
                )
                final_ars_updated = Final_Data.objects.filter(
                    f_ars_OrgCode__iexact=old_code
                ).update(
                    f_ars_OrgCode=new_code,
                    f_ars_OrgName=new_name,
                )

            affected_records = (
                raw_site_updated
                + raw_ars_updated
                + final_site_updated
                + final_ars_updated
            )

            if affected_records:
                messages.success(
                    request,
                    "Organism updated successfully. "
                    f"{affected_records} associated raw/final organism field(s) were refreshed."
                )
            else:
                messages.success(request, "Organism updated successfully.")
            return redirect("/settings/?tab=organisms")

        messages.error(request, "Form validation failed. Please check your inputs.")
    else:
        org_form = OrganismForm(instance=organism)

    return render(request, "home/Organism.html", {
        "org_form": org_form,
        "organism": organism,
        "org_upload_form": org_upload_form,
    })


@login_required(login_url="login")
def view_organism(request):
    q = request.GET.get("q", "").strip()

    organisms = Organism_List.objects.order_by("Whonet_Org_Code")

    if q:
        organisms = organisms.filter(
            Q(Organism__icontains=q) |
            Q(Whonet_Org_Code__icontains=q) |
            Q(Genus_Group__icontains=q) |
            Q(Genus_Code__icontains=q) |
            Q(Species_Group__icontains=q) |
            Q(Serovar_Group__icontains=q) |
            Q(Organism_Type__icontains=q) |
            Q(Kingdom__icontains=q) |
            Q(Phylum__icontains=q) |
            Q(Class__icontains=q) |
            Q(Order__icontains=q) |
            Q(Family_Code__icontains=q)
        )

    paginator = Paginator(organisms, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "home/Organism_view.html", {
        "page_obj": page_obj,
        "organisms": page_obj.object_list,  # optional cleanup
        "search_query": q,
    })





@login_required(login_url="login")
#Delete Organism
def del_organism (request, id):
    organism = get_object_or_404(Organism_List, pk=id)
    organism.delete()
    return redirect('view_organism')


@login_required(login_url="login")
def del_all_organism(request):
    Organism_List.objects.all().delete()
    messages.success(request, "All records have been deleted successfully.")
    return redirect('view_organism')  # Redirect to the table view




@login_required(login_url="login")
def upload_organisms(request):

    if request.method == "POST":
        org_upload_form = Organism_uploadForm(request.POST, request.FILES)

        if org_upload_form.is_valid():
            uploaded_file = org_upload_form.save()
            file = uploaded_file.File_uploadOrg

            try:
                # Load file depending on extension
                df = read_tabular_upload(file)

                # Fill NaN with empty string
                df = df.fillna("")

                # Required column
                if "Whonet_Org_Code" not in df.columns:
                    messages.error(request, "Missing required column: Whonet_Org_Code")
                    return redirect("upload_organisms")

                created_count = 0
                duplicate_count = 0

                def clean_upload_value(column_name, uppercase=False):
                    value = str(row.get(column_name, "")).strip()
                    return value.upper() if uppercase else value

                for _, row in df.iterrows():
                    org_code = str(row.get("Whonet_Org_Code", "")).strip().lower()
                    if not org_code:
                        continue
                    if Organism_List.objects.filter(Whonet_Org_Code__iexact=org_code).exists():
                        duplicate_count += 1
                        continue

                    Organism_List.objects.create(
                        Whonet_Org_Code=org_code,
                        Replaced_by=clean_upload_value("Replaced_by"),
                        Organism=clean_upload_value("Organism"),
                        Organism_Type=clean_upload_value("Organism_Type"),
                        Family_Code=clean_upload_value("Family_Code", uppercase=True),
                        Genus_Group=clean_upload_value("Genus_Group", uppercase=True),
                        Genus_Code=clean_upload_value("Genus_Code", uppercase=True),
                        Species_Group=clean_upload_value("Species_Group", uppercase=True),
                        Serovar_Group=clean_upload_value("Serovar_Group", uppercase=True),
                        Kingdom=clean_upload_value("Kingdom"),
                        Phylum=clean_upload_value("Phylum"),
                        Class=clean_upload_value("Class"),
                        Order=clean_upload_value("Order"),
                        Family=clean_upload_value("Family"),
                        Genus=clean_upload_value("Genus"),
                    )
                    created_count += 1

                messages.success(request, f"Organism list uploaded. {created_count} created, {duplicate_count} skipped because they already exist.")
                return redirect("view_organism")

            except Exception as e:
                print("Upload error:", e)
                messages.error(request, f"Error processing file: {e}")
                return redirect("add_organism")

        else:
            messages.error(request, "Upload form is not valid.")

    else:
        org_upload_form = Organism_uploadForm()

    return render(request, "Settings.html", {
        "org_upload_form": org_upload_form
    })




@login_required(login_url="login")
#for exporting into excel
def export_organisms(request):
    objects = Organism_List.objects.all()
    data = []

    for obj in objects:
        data.append({
            "Whonet_Org_Code": obj.Whonet_Org_Code or "",
            "Organism": obj.Organism or "",
            "Organism_Type": obj.Organism_Type or "",
            "Replaced_by": obj.Replaced_by or "",

            "Family_Code": obj.Family_Code or "",
            "Genus_Group": obj.Genus_Group or "",
            "Genus_Code": obj.Genus_Code or "",
            "Species_Group": obj.Species_Group or "",
            "Serovar_Group": obj.Serovar_Group or "",

            "Kingdom": obj.Kingdom or "",
            "Phylum": obj.Phylum or "",
            "Class": obj.Class or "",
            "Order": obj.Order or "",
            "Family": obj.Family or "",
            "Genus": obj.Genus or "",
        })

    
    # Define file path
    file_path = "Organism_list.xlsx"

    # Convert data to DataFrame and save as Excel
    df = pd.DataFrame(data)
    df.to_excel(file_path, index=False)

    # Return the file as a response
    return FileResponse(open(file_path, "rb"), as_attachment=True, filename="Organism_list.xlsx")



@login_required(login_url="login")
def get_organism_name(request):
    org_code = request.GET.get("org_code")
    field_key = request.GET.get("field_key")

    if not org_code or not field_key:
        return JsonResponse({"error": "Missing parameters"}, status=400)

    org = Organism_List.objects.filter(
        Whonet_Org_Code=org_code
    ).values().first()

    if not org:
        return JsonResponse({"error": "Organism not found"}, status=404)

    if field_key not in org:
        return JsonResponse({"error": "Invalid field_key"}, status=400)

    return JsonResponse({field_key: org[field_key]})






############ Emerging Resistance
@login_required(login_url="login")
def add_emerging_age(request):

    if request.method == "POST":
        eme_form = Emerge_Pheno_Form(request.POST)

        if eme_form.is_valid():
            eme_form.save()
            messages.success(request, "You have successfully created an Emerging Resistance Criteria")
        else:
            messages.warning(request, _first_form_error(eme_form, "Form validation failed. Please check your inputs."))

    # Always return to Settings → Emerging
    return redirect("/settings/?tab=emerging")


@login_required(login_url="login")
def view_eme_age(request):
    criteria = Emerging_Filter_Age.order_by("Eme_Organism")

    paginator = Paginator(criteria, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "home/Emerge_Phen_View.html", {
        "page_obj": page_obj,
        "criteria": page_obj.object_list,  # optional cleanup
    })


@login_required(login_url="login")
def edit_eme_age(request, pk):
    criteria = get_object_or_404(Emerging_Filter_Age, pk=pk)
    eme_upload_form = Eme_Crit_Upload_Form()

    if request.method == "POST":
        eme_form = Emerge_Pheno_Form(request.POST, instance=criteria)

        if eme_form.is_valid():
            eme_form.save()
            messages.success(request, "Emerging Criteria updated successfully.")
            return redirect("settings_page")  # clean redirect

        messages.error(request, "Form validation failed. Please check your inputs.")
    else:
        eme_form = Emerge_Pheno_Form(instance=criteria)

    return render(request, "settings/settings_page.html", {
        "eme_form": eme_form,
        "criteria": criteria,
        "eme_upload_form": eme_upload_form,
        "active_tab": "emerging",   
    })






# download list of emerging
@login_required(login_url="login")
def download_emerging_csv(request):
    queryset = (
        Referred_Data.objects
        .filter(
            Emerging_Flag_Age=True,
            Specimen_Type__Emerging_Spec_Flag=True,
            antibiotic_entries__ab_breakpoints_id__Emerging_Abx_Flag=True,
        )
        .distinct()
        .prefetch_related("antibiotic_entries")
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="emerging_cases.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Accession No",
        "Age",
        "Specimen",
        "Antibiotics"
    ])

    for case in queryset:
        antibiotics = ", ".join(
            ab.ab_Abx_code for ab in case.antibiotic_entries.all()
        )

        writer.writerow([
            case.AccessionNo,
            case.Age,
            str(case.Specimen_Type),
            antibiotics
        ])

    return response


@login_required(login_url="login")
def export_emerging_age(request):
    return _export_queryset_to_excel(
        Emerging_Filter_Age.objects.all().order_by("Eme_Age"),
        [("Eme_Age", "Eme_Age")],
        "Emerging_age_criteria.xlsx",
    )


########### Phenotypes --- PRE

@login_required(login_url="login")
def add_phenotype_pre(request):

    if request.method != "POST":
        return redirect("/settings/?tab=pheno_pre_tab")

    form = Phenotype_Pre_Form(request.POST)

    
    context = {
        "pheno_pre_form": form,
        "editing": False,
    }


    if form.is_valid():
        form.save()
        messages.success(request, "Phenotype (Pre) added successfully.")
    else:
        messages.warning(request, _first_form_error(form, "Failed to add Phenotype (Pre)."))
        print(form.errors)

    return redirect("/settings/?tab=pheno_pre_tab")




@login_required(login_url="login")
def edit_phenotype_pre(request, pk):

    phenotype = get_object_or_404(Phenotype_Pre, pk=pk)

    form = Phenotype_Pre_Form(instance=phenotype)

    context = {
        "pheno_pre_form": form,
        "editing": True,
        "edit_id": phenotype.id,
    }

    return render(
    request,
    "home/Pheno_Pre.html",
    {
        "form": Phenotype_Pre_Form(instance=phenotype),
        "object": phenotype,
    }
)




@login_required(login_url="login")
def update_phenotype_pre(request, id):

    phenotype = get_object_or_404(Phenotype_Pre, pk=pk)

    if request.method != "POST":
        return redirect("/settings/?tab=pheno_pre_tab")

    form = Phenotype_Pre_Form(request.POST, instance=phenotype)

    if form.is_valid():
        form.save()
        messages.success(request, "Phenotype (Pre) updated successfully.")
    else:
        messages.error(request, "Failed to update Phenotype (Pre).")
        print(form.errors)

    return redirect("/settings/?tab=pheno_pre_tab")



@login_required(login_url="login")
def upload_phenotype_pre(request):

    if request.method != "POST":
        return redirect("/settings/?tab=pheno_pre_tab")

    pheno_pre_upload_form = Pheno_pre_upForm(request.POST, request.FILES)

    if not pheno_pre_upload_form.is_valid():
        messages.error(request, "Invalid upload file.")
        return redirect("/settings/?tab=pheno_pre_tab")

    file = request.FILES["File_Pheno_pre"]


    try:
        df = read_tabular_upload(file)

        created_count = 0
        duplicate_count = 0
        for _, row in df.iterrows():
            phenotype = str(row.get("Pre_Phenotypes") or "").strip()
            if not phenotype:
                continue
            if Phenotype_Pre.objects.filter(Pre_Phenotypes__iexact=phenotype).exists():
                duplicate_count += 1
                continue
            Phenotype_Pre.objects.create(
                Pre_Phenotypes=phenotype
            )
            created_count += 1

        messages.success(request, f"Phenotype (Pre) uploaded. {created_count} created, {duplicate_count} skipped because they already exist.")

    except Exception as e:
        messages.error(request, "Upload failed.")
        print(e)

    return redirect("view_phenotype_pre")


@login_required(login_url="login")
def delete_phenotype_pre(request, pk):

    phenotype = get_object_or_404(Phenotype_Pre, pk=pk)
    phenotype.delete()

    messages.success(request, "Phenotype (Pre) deleted.")
    return redirect("/settings/?tab=pheno_pre_tab")


@login_required(login_url="login")
def delete_all_phenotype_pre(request):

    if request.method == "POST":
        Phenotype_Pre.objects.all().delete()
        messages.success(request, "All Phenotype (Pre) records deleted.")

    return redirect("/settings/?tab=pheno_pre_tab")


@login_required(login_url="login")
def export_phenotype_pre(request):
    return _export_queryset_to_excel(
        Phenotype_Pre.objects.all().order_by("Pre_Phenotypes"),
        [("Pre_Phenotypes", "Pre_Phenotypes")],
        "Phenotype_pre.xlsx",
    )



@login_required(login_url="login")
# View to display all specimen types
def view_phenotype_pre(request):
    q = request.GET.get("q", "").strip()
    sort_by = request.GET.get('sort', 'Pre_Phenotypes')  # Default sort field
    order = request.GET.get('order', 'desc')  # Default sort order

    sort_field = f"-{sort_by}" if order == 'desc' else sort_by
    pheno_pre_items = Phenotype_Pre.objects.all().order_by(sort_field)

    if q:
        pheno_pre_items = pheno_pre_items.filter(
            Q(Pre_Phenotypes__icontains=q)
        )

    paginator = Paginator(pheno_pre_items, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)


    return render(request, 'home/Pheno_Pre_View.html', {'pheno_pre_items': pheno_pre_items, 'page_obj': page_obj, 'q': q, 'sort_by': sort_by, 'order': order})





################ Phenotypes --- POST


@login_required(login_url="login")
def add_phenotype_post(request):

    if request.method != "POST":
        return redirect("/settings/?tab=pheno_post_tab")

    form = Phenotype_Post_Form(request.POST)

    context = {
        "pheno_post_form": form,
        "editing": False,
    }

    if form.is_valid():
        form.save()
        messages.success(request, "Phenotype (Post) added successfully.")
    else:
        messages.warning(request, _first_form_error(form, "Failed to add Phenotype (Post)."))
        print(form.errors)

    return redirect("/settings/?tab=pheno_post_tab")


@login_required(login_url="login")
def view_phenotype_post(request):

    q = request.GET.get("q", "").strip()

    phenotype_posts = Phenotype_Post.objects.all()

    if q:
        phenotype_posts = phenotype_posts.filter(
            Post_Phenotypes__icontains=q
        )

    phenotype_posts = phenotype_posts.order_by("Post_Phenotypes")

    paginator = Paginator(phenotype_posts, 25)  # rows per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "q": q,
    }

    return render(
        request,
        "home/Pheno_Post_View.html",
        context
    )


@login_required(login_url="login")
def upload_phenotype_post(request):

    if request.method != "POST":
        return redirect("/settings/?tab=pheno_post_tab")

    upload_file = request.FILES.get("File_Pheno_post")

    if not upload_file:
        messages.error(request, "No file selected for upload.")
        return redirect("/settings/?tab=pheno_post_tab")

    # Save uploaded file (optional audit trail)
    Pheno_upload_Post.objects.create(File_Pheno_post=upload_file)

    try:
        # Read file
        df = read_tabular_upload(upload_file)

        # Normalize column names
        df.columns = [c.strip() for c in df.columns]

        # Expected column name
        expected_column = "Post_Phenotypes"

        if expected_column not in df.columns:
            messages.error(
                request,
                f"Missing required column: {expected_column}"
            )
            return redirect("/settings/?tab=pheno_post_tab")

        created_count = 0
        duplicate_count = 0
        for value in df[expected_column].dropna():
            phenotype = str(value).strip()
            if not phenotype:
                continue
            if Phenotype_Post.objects.filter(Post_Phenotypes__iexact=phenotype).exists():
                duplicate_count += 1
                continue
            Phenotype_Post.objects.create(Post_Phenotypes=phenotype)
            created_count += 1

        messages.success(
            request,
            f"Phenotype (Post) uploaded. {created_count} created, {duplicate_count} skipped because they already exist."
        )

    except Exception as e:
        messages.error(request, f"Upload failed: {str(e)}")

    return redirect("/settings/?tab=pheno_post_tab")



@login_required(login_url="login")
def edit_phenotype_post(request, pk):

    phenotype_post = get_object_or_404(Phenotype_Post, pk=pk)
    form = Phenotype_Post_Form(instance=phenotype_post)

    context = {
        "form": form,
        "object": phenotype_post,
        "editing": True,
    }

    return render(
        request,
        "home/Pheno_Post.html",
        context
    )



@login_required(login_url="login")
def update_phenotype_post(request, pk):

    phenotype_post = get_object_or_404(Phenotype_Post, pk=pk)

    if request.method != "POST":
        return redirect("view_phenotype_post")

    form = Phenotype_Post_Form(
        request.POST,
        instance=phenotype_post
    )

    if form.is_valid():
        form.save()
        messages.success(
            request,
            "Phenotype (Post) updated successfully."
        )
    else:
        messages.error(
            request,
            "Failed to update Phenotype (Post)."
        )
        print(form.errors)

    return redirect("view_phenotype_post")





@login_required(login_url="login")
def delete_phenotype_post(request, pk):

    phenotype_post = get_object_or_404(Phenotype_Post, pk=pk)

    phenotype_post.delete()

    messages.success(
        request,
        "Phenotype (Post) deleted successfully."
    )

    return redirect("view_phenotype_post")


@login_required(login_url="login")
def export_phenotype_post(request):
    return _export_queryset_to_excel(
        Phenotype_Post.objects.all().order_by("Post_Phenotypes"),
        [("Post_Phenotypes", "Post_Phenotypes")],
        "Phenotype_post.xlsx",
    )

################ Recommendations


@login_required(login_url="login")
def add_recommendation_item(request):

    if request.method != "POST":
        return redirect("/settings/?tab=recommendation_tab")

    form = Recco_item_Form(request.POST)

    if form.is_valid():
        form.save()
        messages.success(request, "Recommendation item added successfully.")
    else:
        messages.warning(request, _first_form_error(form, "Failed to add recommendation item."))
        print(form.errors)

    return redirect("/settings/?tab=recommendation_tab")




@login_required(login_url="login")
def upload_recommendation_items(request):

    if request.method != "POST":
        return redirect("/settings/?tab=recommendation_tab")

    reco_desc_upload = request.FILES.get("File_reco_desc")

    if not reco_desc_upload:
        messages.error(request, "No file selected for upload.")
        return redirect("/settings/?tab=recommendation_tab")

    # optional: keep upload record
    Reco_item_upload.objects.create(File_reco_desc=reco_desc_upload)

    try:
        df = read_tabular_upload(reco_desc_upload)

        df.columns = [c.strip() for c in df.columns]

        required_cols = {"RecoCode", "Description"}
        if not required_cols.issubset(df.columns):
            messages.error(
                request,
                "File must contain columns: RecoCode, Description"
            )
            return redirect("/settings/?tab=recommendation_tab")

        created_count = 0
        duplicate_count = 0
        for _, row in df.iterrows():
            reco_code = str(row["RecoCode"]).strip() if pd.notna(row["RecoCode"]) else ""
            if not reco_code:
                continue
            if Recommendation_items.objects.filter(RecoCode__iexact=reco_code).exists():
                duplicate_count += 1
                continue
            Recommendation_items.objects.create(
                RecoCode=reco_code,
                Description=str(row["Description"]).strip()
            )
            created_count += 1

        messages.success(
            request,
            f"Recommendation items uploaded. {created_count} created, {duplicate_count} skipped because they already exist."
        )

    except Exception as e:
        messages.error(request, f"Upload failed: {e}")

    return redirect("/settings/?tab=recommendation_tab")



@login_required(login_url="login")
def view_recommendation_items(request):

    q = request.GET.get("q", "").strip()

    recommendation_items = Recommendation_items.objects.all()

    if q:
        recommendation_items = recommendation_items.filter(
            RecoCode__icontains=q
        ) | recommendation_items.filter(
            Description__icontains=q
        )

    recommendation_items = recommendation_items.order_by("RecoCode")

    paginator = Paginator(recommendation_items, 25)  # rows per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "q": q,
    }

    return render(
        request,
        "home/Recommendation_View.html",
        context
    )



def edit_recommendation_item(request, pk):

    item = get_object_or_404(Recommendation_items, pk=pk)

    if request.method == "POST":
        form = Recco_item_Form(request.POST, instance=item)

        if form.is_valid():
            form.save()
            messages.success(request, "Recommendation item updated successfully.")
            return redirect("view_recommendation_items")
        else:
            messages.error(request, "Failed to update recommendation item.")
            print(form.errors)
    else:
        form = Recco_item_Form(instance=item)

    context = {
        "form": form,
        "object": item,
        "editing": True,
    }

    return render(
        request,
        "home/Recommendation_Edit.html",
        context
    )


@login_required(login_url="login")
def delete_recommendation_item(request, pk):

    item = get_object_or_404(Recommendation_items, pk=pk)

    if request.method == "POST":
        item.delete()
        messages.success(request, "Recommendation item deleted successfully.")
        return redirect("view_recommendation_items")

    # fallback for GET (confirm via JS already)
    item.delete()
    messages.success(request, "Recommendation item deleted successfully.")
    return redirect("view_recommendation_items")


@login_required(login_url="login")
def export_recommendation_items(request):
    return _export_queryset_to_excel(
        Recommendation_items.objects.all().order_by("RecoCode"),
        [
            ("RecoCode", "RecoCode"),
            ("Description", "Description"),
        ],
        "Recommendation_items.xlsx",
    )




@require_GET
def get_recommendation_description(request):

    reco_code = request.GET.get("reco_code")

    if not reco_code:
        return JsonResponse({"description": ""})

    reco = (
        Recommendation_items.objects
        .filter(RecoCode__iexact=reco_code)
        .order_by("id")
        .first()
    )
    return JsonResponse({"description": reco.Description if reco else ""})



################  PROJECTS

from django.contrib.auth.decorators import login_required
from django.shortcuts import render

# Import your WGS form models — adjust imports to match your actual app paths
from apps.wgs_app.models import *
from apps.wgs_app.forms import (   # adjust to wherever your ModelForms live
    BactScoutUploadForm,
    GambitUploadForm,
    MlstUploadForm,
    Checkm2UploadForm,
    AssemblyUploadForm,
    AmrUploadForm,
)
from apps.home.models import Referred_Data





@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_LAB_ENCODER)
def update_wgs_classification_inline(request, accession_no):

    if request.method != "POST":
        return redirect("/wgs/projects/?tab=wgs_classification")

    from apps.wgs_app.models import SampleInformation, WGS_Project

    isolate = get_object_or_404(
        Final_Data,
        f_AccessionNo=accession_no
    )

    sampleinfo = (
        SampleInformation.objects
        .filter(sample_accession=isolate.f_AccessionNo)
        .order_by("-Date_uploaded_si", "-pk")
        .first()
    )
    if not sampleinfo:
        wgs_project = (
            WGS_Project.objects
            .filter(Q(Ref_Accession=isolate) | Q(WGS_SampleInfo_Acc=isolate.f_AccessionNo))
            .first()
        )
        if not wgs_project:
            wgs_project = WGS_Project.objects.create(
                Ref_Accession=isolate,
                WGS_SampleInfo_Acc=isolate.f_AccessionNo,
                WGS_SampleInfoSummary=True,
            )
        sampleinfo = SampleInformation.objects.create(
            sample_project=wgs_project,
            sample_accession=isolate.f_AccessionNo,
            sample_name=isolate.f_AccessionNo,
        )

    sampleinfo.emerging = "Class_Chk_Emerging" in request.POST
    sampleinfo.structured = "Class_Chk_Structured" in request.POST
    sampleinfo.satscan = "Class_Chk_Satscan" in request.POST
    sampleinfo.serotyping = "Class_Chk_Serotyping" in request.POST
    sampleinfo.ghru_all = "Class_Chk_GHRU_all" in request.POST
    sampleinfo.ghru_neo = "Class_Chk_GHRU_Neo" in request.POST
    sampleinfo.ghru = sampleinfo.ghru_all or sampleinfo.ghru_neo
    sampleinfo.egasp = "Class_Chk_EGASP" in request.POST
    sampleinfo.tricycle = "Class_Chk_Tricycle" in request.POST
    sampleinfo.pulsenet = "Class_Chk_Pulsenet" in request.POST
    sampleinfo.tulip = "Class_Chk_Tulip" in request.POST
    sampleinfo.save(update_fields=[
        "emerging",
        "structured",
        "satscan",
        "serotyping",
        "ghru",
        "ghru_all",
        "ghru_neo",
        "egasp",
        "tricycle",
        "pulsenet",
        "tulip",
    ])

    messages.success(
        request,
        f"WGS classification updated for {accession_no}"
    )

    return redirect("/wgs/projects/?tab=wgs_classification")



@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_LAB_ENCODER)
def wgs_classification_view(request, pk):

    isolate = get_object_or_404(Final_Data, pk=pk)

    # Create or fetch classification row
    classification, created = Classification_Table.objects.get_or_create(
        Class_idNumReferred=isolate,
        defaults={
            "Class_AccessionNo": isolate.f_AccessionNo
        }
    )

    if request.method == "POST":
        form = Classification_Form(
            request.POST,
            instance=classification
        )

        if form.is_valid():
            try:
                with transaction.atomic():
                    obj = form.save(commit=False)
                    obj.Class_AccessionNo = isolate.f_AccessionNo
                    obj.Class_idNumReferred = isolate
                    obj.save()

                messages.success(
                    request,
                    "WGS classification updated successfully."
                )
                return redirect("projects_page")

            except Exception as e:
                messages.error(
                    request,
                    f"Error saving classification: {e}"
                )

        else:
            messages.error(
                request,
                "Failed to update WGS classification."
            )
    else:
        form = Classification_Form(instance=classification)

    return render(
        request,
        "projects/Classification.html",
        {
            "form": form,
            "referred": isolate,
            "editing": True,
        }
    )


############## TAT MONITORING


def _get_tat_effective_days(tat):
    if tat.tat_Status_Release == "Released" and tat.tat_Final_TAT is not None:
        return tat.tat_Final_TAT
    return tat.tat_Running_TAT


def _tat_pressure_for_days(days, target_days=None):
    display_target_days = target_days or 40

    if days is None or not display_target_days:
        return "none"

    remaining_days = display_target_days - days
    tat_ratio = days / display_target_days
    if days > display_target_days:
        return "overdue"
    if 0 <= remaining_days <= 5:
        return "near"
    if tat_ratio >= 0.75:
        return "watch"
    return "safe"


def _build_tat_running_rows(tat_records, configs):
    rows = []
    default_display_target_days = 40

    for tat in tat_records:
        batch = tat.tat_Batch_Isolates
        step_map = {
            step.step_config_id: step
            for step in tat.steps.all()
        }
        effective_days = _get_tat_effective_days(tat)
        target_days = tat.tat_Target_Days or 0
        remaining_days = None
        display_target_days = target_days or default_display_target_days
        display_remaining_days = None
        days_past_target = None

        if effective_days is not None and target_days:
            remaining_days = target_days - effective_days

        if effective_days is not None and display_target_days:
            display_remaining_days = display_target_days - effective_days
            if display_remaining_days < 0:
                days_past_target = abs(display_remaining_days)

        tat_pressure = _tat_pressure_for_days(effective_days, target_days)

        rows.append({
            "tat": tat,
            "batch": batch,
            "batch_name": batch.bat_Batch_Name if batch else tat.tat_Batch_Code,
            "site_code": tat.tat_SiteCode or (batch.bat_SiteCode if batch else ""),
            "referral_date": tat.tat_Referral_Date or (batch.bat_Referral_Date if batch else None),
            "accessions": batch.bat_RefNo if batch else "",
            "num_isolates": tat.tat_Num_Isolate,
            "batch_number": tat.tat_BatchNumber,
            "total_batch": tat.tat_Total_Batch,
            "location": tat.tat_Batch_Location,
            "date_received": min(
                [step.date_received for step in step_map.values() if step.date_received],
                default=None,
            ),
            "action_taken": (
                tat.current_step.step_config.step_type
                if tat.current_step and tat.current_step.step_config
                else ""
            ),
            "assigned_to": (
                tat.current_step.performed_by
                if tat.current_step and tat.current_step.performed_by
                else None
            ),
            "effective_days": effective_days,
            "target_days": target_days,
            "remaining_days": remaining_days,
            "display_target_days": display_target_days,
            "display_remaining_days": display_remaining_days,
            "days_past_target": days_past_target,
            "is_overdue": effective_days is not None and display_target_days and effective_days > display_target_days,
            "near_due": display_remaining_days is not None and 0 <= display_remaining_days <= 5,
            "tat_pressure": tat_pressure,
            "step_cells": [
                {
                    "config": config,
                    "step": step_map.get(config.id),
                }
                for config in configs
            ],
        })

    return rows


TAT_RUNNING_SORT_FIELDS = {
    "batch_code": "tat_Batch_Code",
    "site_code": "tat_SiteCode",
    "referral_date": "tat_Referral_Date",
    "num_isolates": "tat_Num_Isolate",
    "batch_number": "tat_BatchNumber",
    "total_batch": "tat_Total_Batch",
    "running_tat": "tat_Running_TAT",
    "location": "tat_Batch_Location",
    "date_received": "first_date_received",
    "tat_status": "tat_Status_Release",
    "scanning_raw": "tat_Scanning_raw",
    "scanning_ws": "tat_Scanning_ws",
    "scanning_final": "tat_Scanning_final",
    "action_taken": "current_step_type",
    "assigned_to": "current_step_staff",
    "date_released": "tat_Date_Released",
    "final_tat": "tat_Final_TAT",
    "release_status": "tat_Status_Release",
    "remarks": "tat_Remarks",
    "last_update": "tat_Date_Last_Update",
}

TAT_RUNNING_DESC_DEFAULTS = {
    "referral_date",
    "running_tat",
    "date_received",
    "date_released",
    "final_tat",
    "last_update",
}


def _tat_running_sort_context(request, default_sort="referral_date", default_order="desc"):
    current_sort = request.GET.get("sort", default_sort)
    if current_sort not in TAT_RUNNING_SORT_FIELDS:
        current_sort = default_sort

    current_order = request.GET.get("order", default_order).lower()
    if current_order not in {"asc", "desc"}:
        current_order = default_order

    order_field = TAT_RUNNING_SORT_FIELDS[current_sort]
    order_by = [f"-{order_field}" if current_order == "desc" else order_field]
    if order_field != "tat_Batch_Code":
        order_by.append("tat_Batch_Code")
    order_by.append("pk")

    base_params = request.GET.copy()
    base_params.pop("page", None)
    sort_links = {}
    for key in TAT_RUNNING_SORT_FIELDS:
        next_order = "desc" if key in TAT_RUNNING_DESC_DEFAULTS else "asc"
        if key == current_sort:
            next_order = "asc" if current_order == "desc" else "desc"
        params = base_params.copy()
        params["sort"] = key
        params["order"] = next_order
        sort_links[key] = {
            "url": f"?{params.urlencode()}",
            "active": key == current_sort,
            "indicator": "down" if current_order == "desc" else "up",
        }

    page_params = request.GET.copy()
    page_params.pop("page", None)

    return {
        "current_sort": current_sort,
        "current_order": current_order,
        "order_by": order_by,
        "sort_links": sort_links,
        "page_query": page_params.urlencode(),
    }


def _annotate_tat_running_sort_fields(tat_records):
    latest_step = TATStep.objects.filter(tat=OuterRef("pk")).order_by("-id")
    return tat_records.annotate(
        first_date_received=Min("steps__date_received"),
        current_step_type=Subquery(latest_step.values("step_type")[:1]),
        current_step_staff=Subquery(latest_step.values("performed_by__Staff_Name")[:1]),
    )


def _tat_excel_value(row, index, default=None):
    if index >= len(row):
        return default
    value = row.iloc[index]
    if pd.isna(value):
        return default
    return value


def _normalize_tat_column_name(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _tat_column_map(df):
    column_map = {}
    for index, column in enumerate(df.columns):
        key = _normalize_tat_column_name(column)
        if key:
            column_map.setdefault(key, index)
    return column_map


def _tat_value_by_alias(row, column_map, aliases, default_index=None, default=None):
    for alias in aliases:
        index = column_map.get(_normalize_tat_column_name(alias))
        if index is not None:
            value = _tat_excel_value(row, index, default)
            if value not in ("", None):
                return value
    if default_index is None:
        return default
    return _tat_excel_value(row, default_index, default)


def _tat_text(row, index):
    value = _tat_excel_value(row, index, "")
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _tat_int(row, index):
    value = _tat_excel_value(row, index)
    if value in ("", None):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _tat_bool(row, index):
    value = _tat_excel_value(row, index, "")
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(int(value))
    return str(value).strip().lower() in {"1", "true", "yes", "y", "checked", "x"}


def _tat_text_by_alias(row, column_map, aliases, default_index=None):
    value = _tat_value_by_alias(row, column_map, aliases, default_index, "")
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _tat_int_by_alias(row, column_map, aliases, default_index=None):
    value = _tat_value_by_alias(row, column_map, aliases, default_index)
    if value in ("", None):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _tat_bool_by_alias(row, column_map, aliases, default_index=None):
    value = _tat_value_by_alias(row, column_map, aliases, default_index, "")
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(int(value))
    return str(value).strip().lower() in {"1", "true", "yes", "y", "checked", "x"}


def _tat_date(row, index):
    value = _tat_excel_value(row, index)
    if value in ("", None):
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _tat_date_by_alias(row, column_map, aliases, default_index=None):
    value = _tat_value_by_alias(row, column_map, aliases, default_index)
    if value in ("", None):
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _tat_location_value(raw_value, existing_value="n/a"):
    value = str(raw_value or "").strip()
    if not value:
        return existing_value or "n/a"

    configured = {
        item.name.lower(): item.name
        for item in TATLocation.objects.filter(is_active=True)
    }
    configured.setdefault("n/a", "n/a")
    return configured.get(value.lower(), existing_value or "n/a")


def _find_running_tat_sheets(uploaded_file):
    sheets = read_tabular_upload_sheets(uploaded_file)
    valid_sheets = []

    for sheet_name, df in sheets.items():
        normalized = {str(col).strip().lower() for col in df.columns}
        if "batchcode" in normalized and "sitecode" in normalized:
            valid_sheets.append((sheet_name, df))

    if valid_sheets:
        return valid_sheets

    first_name = next(iter(sheets))
    return [(first_name, sheets[first_name])]


@login_required(login_url="login")
@require_POST
def upload_running_tat(request):
    uploaded_file = request.FILES.get("tat_file")
    if not uploaded_file:
        messages.error(request, "Please select a Running TAT CSV, TSV, or Excel file.")
        return redirect("tat_running_list")

    try:
        running_sheets = _find_running_tat_sheets(uploaded_file)
        imported_count = 0
        skipped_count = 0

        with transaction.atomic():
            for sheet_name, df in running_sheets:
                column_map = _tat_column_map(df)
                for _, row in df.iterrows():
                    batch_code = _tat_text_by_alias(row, column_map, ["BatchCode"], 0)
                    if not batch_code or batch_code.lower() in {"nan", "batchcode"}:
                        skipped_count += 1
                        continue

                    site_code = _tat_text_by_alias(row, column_map, ["SiteCode"], 1)
                    referral_date = _tat_date_by_alias(row, column_map, ["RefDate", "Referral Date"], 5)
                    batch_number = _tat_text_by_alias(row, column_map, ["BatchNumber"], 7)
                    total_batch = _tat_text_by_alias(row, column_map, ["TotalBatchNumber", "Total Batch"], 8)
                    num_isolates = _tat_int_by_alias(row, column_map, ["No_Isolates", "No Isolates"], 6)

                    batch = (
                        Batch_Table.objects
                        .filter(Q(bat_Batch_Name=batch_code) | Q(bat_Batch_Code=batch_code))
                        .first()
                    )
                    if not batch:
                        batch = Batch_Table.objects.create(
                            bat_Batch_Name=batch_code,
                            bat_Batch_Code=batch_code,
                            bat_SiteCode=site_code,
                            bat_Referral_Date=referral_date,
                            bat_BatchNo=batch_number,
                            bat_Total_batch=total_batch,
                        )
                    else:
                        batch.bat_SiteCode = site_code or batch.bat_SiteCode
                        batch.bat_Referral_Date = referral_date or batch.bat_Referral_Date
                        batch.bat_BatchNo = batch_number or batch.bat_BatchNo
                        batch.bat_Total_batch = total_batch or batch.bat_Total_batch
                        batch.save(update_fields=[
                            "bat_SiteCode",
                            "bat_Referral_Date",
                            "bat_BatchNo",
                            "bat_Total_batch",
                        ])

                    tat, _ = TATform.objects.get_or_create(
                        tat_Batch_Isolates=batch,
                        defaults={
                            "tat_SiteCode": site_code,
                            "tat_Batch_Code": batch_code,
                            "tat_Referral_Date": referral_date,
                            "tat_BatchNumber": batch_number,
                            "tat_Total_Batch": total_batch,
                            "tat_Num_Isolate": num_isolates,
                        },
                    )

                    batch_location = _tat_location_value(
                        _tat_text_by_alias(
                            row,
                            column_map,
                            [
                                "Batch Location",
                                "LAB/DMU/EXECUTIVE/ MEDICAL SPECIALIST",
                                "LAB/DMU/EXECUTIVE/MEDICALSPECIALIST",
                                "Process of:",
                            ],
                            10,
                        ),
                        tat.tat_Batch_Location,
                    )
                    date_received = _tat_date_by_alias(
                        row,
                        column_map,
                        [
                            "Date Received",
                            "As of",
                            "AS OF = <Actual Date Received by DMU/LAB; start of TAT per draft>",
                        ],
                        11,
                    )
                    date_released = (
                        _tat_date_by_alias(row, column_map, ["Date Released"], 17)
                        or _tat_date_by_alias(row, column_map, ["STATUS/RELEASE"], 13)
                        or date_received
                    )
                    status = (
                        _tat_text_by_alias(row, column_map, ["STATUS/RELEASE"], 19)
                        or _tat_text_by_alias(row, column_map, ["STATUS"], 18)
                    )
                    if status.lower() not in {"released", "ongoing", "overdue"}:
                        status = "Released" if date_released else "Ongoing"
                    else:
                        status = status.title()

                    TATform.objects.filter(pk=tat.pk).update(
                        tat_SiteCode=site_code,
                        tat_Batch_Code=batch_code,
                        tat_Referral_Date=referral_date,
                        tat_BatchNumber=batch_number,
                        tat_Total_Batch=total_batch,
                        tat_Num_Isolate=num_isolates,
                        tat_Running_TAT=_tat_int_by_alias(row, column_map, ["RUNNING TAT"], 9),
                        tat_Batch_Location=batch_location,
                        tat_Date_Released=date_released if status == "Released" else None,
                        tat_Final_TAT=_tat_int_by_alias(row, column_map, ["FINAL TAT"], 18) if status == "Released" else None,
                        tat_Status_Release=status,
                        tat_Remarks=_tat_text_by_alias(row, column_map, ["REMARKS", "STATUS"], 12),
                        tat_Scanning_raw=_tat_bool_by_alias(row, column_map, ["Scanning (Raw)", "Scanning Raw"], 14),
                        tat_Scanning_ws=_tat_bool_by_alias(row, column_map, ["Scanning (Worksheet)", "Scanning Worksheet"], 15),
                        tat_Scanning_final=_tat_bool_by_alias(row, column_map, ["Scanning (FINAL)", "Scanning Final"], 16),
                        tat_Date_Last_Update=timezone.localdate(),
                    )

                    recalculate_tat = False
                    if date_received:
                        first_config = TATStepConfig.objects.order_by("order", "id").first()
                        if first_config:
                            received_step, _ = TATStep.objects.get_or_create(
                                tat_id=tat.pk,
                                step_config=first_config,
                                defaults={
                                    "step_type": first_config.step_type,
                                    "step_owner": first_config.step_owner,
                                },
                            )
                            received_step.date_received = date_received
                            received_step.date_finished = date_released if status == "Released" else None
                            received_step.step_type = first_config.step_type
                            received_step.step_owner = first_config.step_owner
                            received_step.save()
                            recalculate_tat = True

                    action_taken = _tat_text_by_alias(
                        row,
                        column_map,
                        ["ACTION TAKEN", "STATUS/RELEASE"],
                        13,
                    )
                    if action_taken:
                        config, _ = TATStepConfig.objects.get_or_create(
                            step_type=action_taken,
                            step_owner="n/a",
                            defaults={
                                "target_days": 0,
                                "order": (TATStepConfig.objects.aggregate(Max("order")).get("order__max") or 0) + 1,
                            },
                        )
                        step, _ = TATStep.objects.get_or_create(
                            tat_id=tat.pk,
                            step_config=config,
                            defaults={
                                "step_type": config.step_type,
                                "step_owner": config.step_owner,
                            },
                        )
                        step.date_received = date_received
                        step.date_finished = date_released if status == "Released" else None
                        step.step_type = config.step_type
                        step.step_owner = config.step_owner
                        step.save()
                        recalculate_tat = True

                    if recalculate_tat:
                        tat.refresh_from_db()
                        tat.save()

                    imported_count += 1

        messages.success(
            request,
            f"Running TAT upload complete from {len(running_sheets)} sheet(s). {imported_count} imported, {skipped_count} skipped."
        )
    except Exception as exc:
        messages.error(request, f"Running TAT upload failed: {exc}")

    return redirect("tat_running_list")


@login_required(login_url="login")
def tat_monitoring_view(request, batch_id):


    batch = get_object_or_404(Batch_Table, id=batch_id)


    tat_obj, created = TATform.objects.get_or_create(
        tat_Batch_Isolates=batch,
        defaults={
            "tat_SiteCode": batch.bat_SiteCode,
            "tat_Batch_Code": batch.bat_Batch_Name,
            "tat_Referral_Date": batch.bat_Referral_Date,
            "tat_Num_Isolate": (
                batch.referred_data_set.count()
                if hasattr(batch, 'referred_data_set') else None
            ),
            "tat_BatchNumber": batch.bat_BatchNo,
            "tat_Total_Batch": batch.bat_Total_batch,
        }
    )

    batch_sync_fields = []
    batch_number = batch.bat_BatchNo or ""
    total_batch = batch.bat_Total_batch or ""
    if tat_obj.tat_BatchNumber != batch_number:
        tat_obj.tat_BatchNumber = batch_number
        batch_sync_fields.append("tat_BatchNumber")
    if tat_obj.tat_Total_Batch != total_batch:
        tat_obj.tat_Total_Batch = total_batch
        batch_sync_fields.append("tat_Total_Batch")
    if batch_sync_fields:
        tat_obj.save(update_fields=[
            *batch_sync_fields,
            "tat_Running_TAT",
            "tat_Final_TAT",
            "tat_Status_Release",
            "tat_Date_Last_Update",
        ])


    if request.method == "POST":

        form = TATMonitoringForm(
            request.POST,
            instance=tat_obj
        )

        formset = TATStepFormSet(
            request.POST,
            instance=tat_obj,
            prefix='steps'
        )

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                tat_instance = form.save(commit=False)
                formset.instance = tat_instance
                tat_instance.save()
                formset.save()
                tat_instance.save()

                receipt_date = (
                    tat_instance.steps
                    .filter(date_received__isnull=False)
                    .order_by("date_received")
                    .values_list("date_received", flat=True)
                    .first()
                    or tat_instance.tat_Referral_Date
                )
                if (
                    receipt_date
                    and tat_instance.tat_Date_Released
                    and tat_instance.tat_Date_Released < receipt_date
                ):
                    messages.warning(
                        request,
                        "Date Released is earlier than the receipt date. The TAT record was saved, but please verify the dates."
                    )

                ante_dated_steps = []
                for step in tat_instance.steps.all():
                    if step.date_received and step.date_finished and step.date_finished < step.date_received:
                        ante_dated_steps.append(step.step_type or "Unnamed step")

                if ante_dated_steps:
                    messages.warning(
                        request,
                        "Some process steps have a Date Finished earlier than Date Received: "
                        + ", ".join(ante_dated_steps[:5])
                    )

            messages.success(
                request,
                "TAT Monitoring updated successfully."
            )

            return redirect(
                "tat_monitoring_view",
                batch_id=batch.id
            )

        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = TATMonitoringForm(instance=tat_obj)

        formset = TATStepFormSet(
            instance=tat_obj,
            prefix='steps'
        )

    return render(
        request,
        "home/tat_monitoring_form.html",
        {
            "form": form,
            "formset": formset,
            "batch": batch,
            "tat": tat_obj,
            "tat_first_date_received": (
                tat_obj.steps
                .filter(date_received__isnull=False)
                .order_by("date_received")
                .values_list("date_received", flat=True)
                .first()
            ),
        }
    )


@login_required(login_url="login")
def export_tat_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "TAT Records"

    # 🔹 Get all configured process steps (ordered)
    configs = TATStepConfig.objects.all().order_by("order")

    # 🔹 Build dynamic headers
    headers = [
        "Batch Name",
        "Batch Code",
        "Site Code",
        "Referral Date",
        "No. of Isolates",
        "Batch Number",
        "Total Batch",
        "Batch Location",
        "Target Days (Batch)",
        "Running TAT",
        "Final TAT",
        "Date Released",
        "Release Status",
        "Overall Remarks",
        "Last Update",
    ]

    # Add per-process columns dynamically
    for config in configs:
        headers.extend([
            f"{config.step_type} Start",
            f"{config.step_type} End",
            f"{config.step_type} Days",
            f"{config.step_type} Within TAT",
            f"{config.step_type} Performed By",
        ])

    ws.append(headers)

    tat_records = TATform.objects.prefetch_related("steps", "tat_Batch_Isolates").all()

    for tat in tat_records:

        row = [
            tat.tat_Batch_Isolates.bat_Batch_Name if tat.tat_Batch_Isolates else "",
            tat.tat_Batch_Code,
            tat.tat_SiteCode,
            tat.tat_Referral_Date,
            tat.tat_Num_Isolate,
            tat.tat_BatchNumber,
            tat.tat_Total_Batch,
            tat.tat_Batch_Location,
            tat.tat_Target_Days,
            tat.tat_Running_TAT,
            tat.tat_Final_TAT,
            tat.tat_Date_Released,
            tat.tat_Status_Release,
            tat.tat_Remarks,
            tat.tat_Date_Last_Update,
        ]

        # Map steps by config for quick lookup
        step_map = {
            step.step_config_id: step
            for step in tat.steps.all()
        }

        # Fill dynamic step columns
        for config in configs:

            step = step_map.get(config.id)

            if step:
                row.extend([
                    step.date_received,
                    step.date_finished,
                    step.step_days_count,
                    "YES" if step.within_tat else "NO" if step.within_tat is False else "",
                    str(step.performed_by) if step.performed_by else "",
                ])
            else:
                row.extend(["", "", "", "", ""])

        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"TAT_Export_{now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response




@login_required(login_url="login")
def export_tat_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Running TAT"

    configs = TATStepConfig.objects.all().order_by("order")
    date_from, date_to = _get_export_date_range(request)
    q = request.GET.get("q", request.GET.get("tat_q", "")).strip()
    status = request.GET.get("status", request.GET.get("tat_status", "")).strip()
    year = request.GET.get("year", request.GET.get("tat_year", "")).strip()
    site = request.GET.get("site", request.GET.get("tat_site", "")).strip()
    tat_records = (
        TATform.objects
        .select_related("tat_Batch_Isolates")
        .prefetch_related("steps__step_config")
        .order_by("-tat_Referral_Date", "tat_Batch_Code")
    )

    if q:
        tat_records = tat_records.filter(
            Q(tat_Batch_Code__icontains=q)
            | Q(tat_SiteCode__icontains=q)
            | Q(tat_Batch_Isolates__bat_Batch_Name__icontains=q)
            | Q(tat_Batch_Isolates__bat_RefNo__icontains=q)
        )
    if site:
        tat_records = tat_records.filter(tat_SiteCode=site)
    if status:
        tat_records = tat_records.filter(tat_Status_Release=status)
    if year and year != "all":
        tat_records = tat_records.filter(tat_Referral_Date__year=year)

    tat_records = _apply_referral_date_range(tat_records, date_from, date_to)
    running_rows = _build_tat_running_rows(tat_records, configs)

    title_fill = PatternFill("solid", fgColor="0F2744")
    header_fill = PatternFill("solid", fgColor="1E5A96")
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    released_fill = PatternFill("solid", fgColor="D4F4DD")
    ongoing_fill = PatternFill("solid", fgColor="FFF4E6")
    overdue_fill = PatternFill("solid", fgColor="FFD4D4")
    thin_border = Border(
        left=Side(style="thin", color="D9E2EF"),
        right=Side(style="thin", color="D9E2EF"),
        top=Side(style="thin", color="D9E2EF"),
        bottom=Side(style="thin", color="D9E2EF"),
    )

    headers = [
        "BatchCode",
        "SiteCode",
        "RefDate",
        "No_Isolates",
        "BatchNumber",
        "TotalBatchNumber",
        "RUNNING TAT",
        "Batch Location",
        "Date Received",
        "STATUS",
        "ACTION TAKEN",
        "SCANNING RAW",
        "SCANNING WORKSHEET",
        "SCANNING FINAL",
        "ASSIGNED TO",
        "Date Released",
        "FINAL TAT",
        "STATUS/RELEASE",
        "Target Days",
        "Remaining Days",
        "REMARKS",
        "Last update",
    ]

    for config in configs:
        headers.extend([
            f"{config.step_type} Received",
            f"{config.step_type} Finished",
            f"{config.step_type} Days",
            f"{config.step_type} Status",
            f"{config.step_type} Staff",
        ])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1, value="ARSP RUNNING TAT MONITORING")
    title.fill = title_fill
    title.font = Font(color="FFFFFF", bold=True, size=14)
    title.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    range_label = ""
    if date_from or date_to:
        range_label = f" | Referral date: {date_from or 'start'} to {date_to or 'end'}"
    generated = ws.cell(row=2, column=1, value=f"Generated {now().strftime('%Y-%m-%d %H:%M')}{range_label}")
    generated.fill = subheader_fill
    generated.font = Font(color="0F2744", italic=True)
    generated.alignment = Alignment(horizontal="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_data in running_rows:
        tat = row_data["tat"]
        row = [
            row_data["batch_name"],
            row_data["site_code"],
            row_data["referral_date"],
            row_data["num_isolates"],
            row_data["batch_number"],
            row_data["total_batch"],
            tat.tat_Running_TAT,
            row_data["location"],
            row_data["date_received"],
            "Overdue" if row_data["is_overdue"] else tat.tat_Status_Release,
            row_data["action_taken"],
            "YES" if tat.tat_Scanning_raw else "",
            "YES" if tat.tat_Scanning_ws else "",
            "YES" if tat.tat_Scanning_final else "",
            str(row_data["assigned_to"]) if row_data["assigned_to"] else "",
            tat.tat_Date_Released,
            tat.tat_Final_TAT,
            tat.tat_Status_Release,
            row_data["target_days"],
            row_data["remaining_days"],
            tat.tat_Remarks,
            tat.tat_Date_Last_Update,
        ]

        for cell_data in row_data["step_cells"]:
            step = cell_data["step"]
            row.extend([
                step.date_received if step else "",
                step.date_finished if step else "",
                step.step_days_count if step else "",
                "Within" if step and step.within_tat else "Outside" if step and step.within_tat is False else "",
                str(step.performed_by) if step and step.performed_by else "",
            ])

        ws.append(row)
        current_row = ws.max_row
        fill = (
            released_fill
            if tat.tat_Status_Release == "Released"
            else overdue_fill
            if row_data["is_overdue"]
            else ongoing_fill
        )

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_num in {10, 18}:
                cell.fill = fill

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        1: 32, 2: 12, 3: 14, 4: 12, 5: 14, 6: 16, 7: 13, 8: 16,
        9: 14, 10: 14, 11: 24, 12: 14, 13: 20, 14: 16, 15: 18,
        16: 18, 17: 14, 18: 16, 19: 12, 20: 14, 21: 32, 22: 18,
    }
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(idx, 18)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"RUNNING_TAT_{now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def _style_tat_export_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1E5A96")
    thin_border = Border(
        left=Side(style="thin", color="D9E2EF"),
        right=Side(style="thin", color="D9E2EF"),
        top=Side(style="thin", color="D9E2EF"),
        bottom=Side(style="thin", color="D9E2EF"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


@login_required(login_url="login")
def export_tat_analysis_tracking_excel(request):
    date_from, date_to = _get_export_date_range(request)
    year = request.GET.get("tat_year", request.GET.get("year", str(timezone.localdate().year))).strip()
    status = request.GET.get("tat_status", request.GET.get("status", "")).strip()
    q = request.GET.get("tat_q", request.GET.get("q", "")).strip()

    tat_records = (
        TATform.objects
        .select_related("tat_Batch_Isolates")
        .filter(tat_Referral_Date__isnull=False)
        .order_by("tat_Referral_Date", "tat_Batch_Code")
    )
    if year and year != "all":
        tat_records = tat_records.filter(tat_Referral_Date__year=year)
    if status:
        tat_records = tat_records.filter(tat_Status_Release=status)
    if q:
        tat_records = tat_records.filter(
            Q(tat_Batch_Code__icontains=q)
            | Q(tat_SiteCode__icontains=q)
            | Q(tat_Batch_Isolates__bat_Batch_Name__icontains=q)
            | Q(tat_Batch_Isolates__bat_RefNo__icontains=q)
        )
    tat_records = _apply_referral_date_range(tat_records, date_from, date_to)

    monthly_rows = []
    quarter_totals = {
        quarter: {"total": 0, "within": 0}
        for quarter in range(1, 5)
    }

    records_by_month = defaultdict(list)
    for tat in tat_records:
        records_by_month[tat.tat_Referral_Date.month].append(tat)

    for month in range(1, 13):
        month_records = records_by_month.get(month, [])
        batch_count = len(month_records)
        isolate_count = sum(tat.tat_Num_Isolate or 0 for tat in month_records)
        ongoing_count = sum(1 for tat in month_records if tat.tat_Status_Release != "Released" and not tat.tat_Date_Released)
        released_count = sum(1 for tat in month_records if tat.tat_Status_Release == "Released" or tat.tat_Date_Released)
        effective_tats = []
        within_count = 0
        out_count = 0

        for tat in month_records:
            effective_tat = (
                tat.tat_Final_TAT
                if (tat.tat_Status_Release == "Released" or tat.tat_Date_Released) and tat.tat_Final_TAT is not None
                else tat.tat_Running_TAT
            )
            if effective_tat is not None:
                effective_tats.append(effective_tat)
            if effective_tat is not None and tat.tat_Target_Days is not None:
                if effective_tat <= tat.tat_Target_Days:
                    within_count += 1
                else:
                    out_count += 1

        average_tat = round(sum(effective_tats) / len(effective_tats), 2) if effective_tats else ""
        denominator = within_count + out_count
        within_pct = within_count / denominator if denominator else ""
        out_pct = out_count / denominator if denominator else ""
        quarter = ((month - 1) // 3) + 1
        quarter_totals[quarter]["total"] += batch_count
        quarter_totals[quarter]["within"] += within_count

        monthly_rows.append({
            "month": calendar.month_name[month],
            "batch_count": batch_count,
            "isolate_count": isolate_count,
            "average_tat": average_tat,
            "ongoing_count": ongoing_count,
            "released_count": released_count,
            "within_count": within_count,
            "out_count": out_count,
            "within_pct": within_pct,
            "out_pct": out_pct,
            "quarter": quarter,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "TAT Analysis Tracking"
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    headers = [
        "Month",
        "Expected Number of\nBatches",
        "Expected Total\nNumber of Isolates",
        "Average TAT",
        "Ongoing\nBatches",
        "Released\nBatches",
        "Within TAT",
        "Out of TAT",
        "% Within TAT",
        "% Out of TAT",
        "TOTAL Quarterly",
        "Within TAT",
        "Quarterly",
    ]

    ws.merge_cells("A1:M1")
    ws["A1"] = "TAT Analysis Tracking"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].border = Border(left=medium, right=medium, top=medium, bottom=medium)

    report_year = year if year and year != "all" else "All Years"
    ws.merge_cells("D2:M2")
    ws["D2"] = f"TAT Analysis Tracking {report_year}"
    ws["D2"].font = Font(bold=True, size=10)
    ws["D2"].alignment = Alignment(horizontal="center")
    ws["D2"].border = table_border

    if date_from or date_to:
        subtitle = f"Referrals Dated {date_from or 'Start'} to {date_to or 'End'}"
    elif year and year != "all":
        subtitle = f"Referrals Dated January to December {year}"
    else:
        subtitle = "Referrals Dated All Years"
    ws.merge_cells("D3:M3")
    ws["D3"] = subtitle
    ws["D3"].font = Font(bold=True, size=9)
    ws["D3"].alignment = Alignment(horizontal="center")
    ws["D3"].border = table_border

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = table_border

    start_row = 5
    for idx, row_data in enumerate(monthly_rows):
        row = start_row + idx
        values = [
            row_data["month"],
            row_data["batch_count"] or "",
            row_data["isolate_count"] or "",
            row_data["average_tat"],
            row_data["ongoing_count"] or "",
            row_data["released_count"] or "",
            row_data["within_count"] or "",
            row_data["out_count"] or "",
            row_data["within_pct"],
            row_data["out_pct"],
            "",
            "",
            "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = table_border
            cell.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
            if col in {9, 10} and value != "":
                cell.number_format = "0.00%"

    for quarter in range(1, 5):
        first_row = start_row + ((quarter - 1) * 3)
        last_row = first_row + 2
        total = quarter_totals[quarter]["total"]
        within = quarter_totals[quarter]["within"]
        pct = within / total if total else ""
        for col, value in ((11, total or ""), (12, within or ""), (13, pct)):
            ws.merge_cells(start_row=first_row, start_column=col, end_row=last_row, end_column=col)
            cell = ws.cell(row=first_row, column=col, value=value)
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 13 and value != "":
                cell.number_format = "0%"
            for merged_row in range(first_row, last_row + 1):
                ws.cell(row=merged_row, column=col).border = table_border

    totals_row = start_row + 13
    total_batches = sum(row["batch_count"] for row in monthly_rows)
    total_isolates = sum(row["isolate_count"] for row in monthly_rows)
    ws.cell(totals_row, 2, total_batches).font = Font(bold=True)
    ws.cell(totals_row, 3, total_isolates).font = Font(bold=True)
    ws.cell(totals_row + 1, 2, "Total # of Batches").font = Font(italic=True, size=9)
    ws.cell(totals_row + 1, 3, "Total # of Isolates").font = Font(italic=True, size=9)
    ws.cell(totals_row + 2, 2, f"as of {now().strftime('%B %-d, %Y') if os.name != 'nt' else now().strftime('%B %#d, %Y')}").font = Font(italic=True, size=9)

    signature_row = totals_row + 6
    for offset, label in ((0, "Prepared By:"), (4, "Noted By:"), (8, "Approved By:")):
        row = signature_row + offset
        ws.cell(row, 1, label)
        ws.cell(row, 1).font = Font(size=9)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row, 2).border = Border(bottom=thin)

    widths = {
        "A": 13, "B": 20, "C": 20, "D": 14, "E": 12, "F": 12, "G": 12,
        "H": 12, "I": 13, "J": 13, "K": 16, "L": 12, "M": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[4].height = 52
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:M{signature_row + 9}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename_year = year if year and year != "all" else "ALL_YEARS"
    response["Content-Disposition"] = (
        f'attachment; filename="TAT_ANALYSIS_TRACKING_{filename_year}_{now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


def _completed_tat_steps(date_from=None, date_to=None, year=None, include_ongoing=None):
    steps = (
        TATStep.objects
        .select_related("tat", "tat__tat_Batch_Isolates", "step_config", "performed_by")
        .filter(date_finished__isnull=False, step_days_count__isnull=False)
        .order_by("step_owner", "step_config__order", "tat__tat_Batch_Code")
    )
    if include_ongoing == "1":
        steps = steps.filter(
            tat__tat_Running_TAT__isnull=False,
            tat__tat_Date_Released__isnull=True,
        ).exclude(tat__tat_Status_Release="Released")
    else:
        steps = steps.filter(
            Q(tat__tat_Status_Release="Released") | Q(tat__tat_Date_Released__isnull=False),
            tat__tat_Final_TAT__isnull=False,
        )
    if year and year != "all":
        steps = steps.filter(tat__tat_Referral_Date__year=year)
    if date_from:
        steps = steps.filter(tat__tat_Referral_Date__gte=date_from)
    if date_to:
        steps = steps.filter(tat__tat_Referral_Date__lte=date_to)
    return steps


@login_required(login_url="login")
def export_tat_owner_performance_excel(request):
    date_from, date_to = _get_export_date_range(request)
    year = request.GET.get("tat_year", request.GET.get("year", "")).strip()
    include_ongoing = request.GET.get("include_ongoing", "")
    wb = Workbook()
    ws = wb.active
    ws.title = "Owner Summary"
    detail_ws = wb.create_sheet("Owner Details")

    ws.append([
        "Owner",
        "Total Steps",
        "Within Target",
        "Outside Target",
        "Average Days",
        "Compliance %",
    ])

    for owner in ["LAB", "DMU"]:
        steps = _completed_tat_steps(date_from, date_to, year, include_ongoing).filter(step_owner=owner)
        total = steps.count()
        if total == 0:
            continue

        within = steps.filter(within_tat=True).count()
        outside = steps.filter(within_tat=False).count()
        avg_days = steps.aggregate(avg=Avg("step_days_count"))["avg"] or 0
        compliance = round((within / total) * 100, 2) if total else 0

        ws.append([
            owner,
            total,
            within,
            outside,
            round(avg_days, 2),
            compliance,
        ])

    detail_ws.append([
        "Owner",
        "Batch Name",
        "Batch Code",
        "Site",
        "Referral Date",
        "Step",
        "Target Days",
        "Date Received",
        "Date Finished",
        "Days Count",
        "Within TAT",
        "Performed By",
        "Remarks",
    ])

    for step in _completed_tat_steps(date_from, date_to, year, include_ongoing):
        tat = step.tat
        batch = tat.tat_Batch_Isolates
        detail_ws.append([
            step.step_owner,
            batch.bat_Batch_Name if batch else "",
            tat.tat_Batch_Code,
            tat.tat_SiteCode,
            tat.tat_Referral_Date,
            step.step_type,
            step.step_config.target_days if step.step_config else "",
            step.date_received,
            step.date_finished,
            step.step_days_count,
            "Within" if step.within_tat else "Outside" if step.within_tat is False else "",
            str(step.performed_by) if step.performed_by else "",
            step.remarks or "",
        ])

    _style_tat_export_sheet(ws)
    _style_tat_export_sheet(detail_ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"TAT_OWNER_PERFORMANCE_{now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url="login")
def export_tat_ipcr_basis_excel(request):
    date_from, date_to = _get_export_date_range(request)
    year = request.GET.get("tat_year", request.GET.get("year", "")).strip()
    include_ongoing = request.GET.get("include_ongoing", "")
    wb = Workbook()
    ws = wb.active
    ws.title = "IPCR Summary"
    detail_ws = wb.create_sheet("IPCR Details")

    ws.append([
        "Staff",
        "Designation",
        "Total Steps",
        "Average Days",
        "Within Target",
        "Outside Target",
        "Compliance %",
    ])

    staff_summary = (
        _completed_tat_steps(date_from, date_to, year, include_ongoing)
        .filter(performed_by__isnull=False)
        .values(
            "performed_by__id",
            "performed_by__Staff_Name",
            "performed_by__Staff_Designation",
        )
        .annotate(
            total_steps=Count("id"),
            avg_days=Avg("step_days_count"),
            within_count=Count("id", filter=Q(within_tat=True)),
            outside_count=Count("id", filter=Q(within_tat=False)),
        )
        .order_by("performed_by__Staff_Name")
    )

    for staff in staff_summary:
        total = staff["total_steps"] or 0
        within = staff["within_count"] or 0
        compliance = round((within / total) * 100, 2) if total else 0
        ws.append([
            staff["performed_by__Staff_Name"] or "",
            staff["performed_by__Staff_Designation"] or "",
            total,
            round(staff["avg_days"], 2) if staff["avg_days"] else 0,
            within,
            staff["outside_count"] or 0,
            compliance,
        ])

    detail_ws.append([
        "Staff",
        "Designation",
        "Owner",
        "Batch Name",
        "Batch Code",
        "Site",
        "Referral Date",
        "Step",
        "Target Days",
        "Date Received",
        "Date Finished",
        "Days Count",
        "Within TAT",
        "Remarks",
    ])

    for step in _completed_tat_steps(date_from, date_to, year, include_ongoing).filter(performed_by__isnull=False):
        tat = step.tat
        batch = tat.tat_Batch_Isolates
        detail_ws.append([
            step.performed_by.Staff_Name if step.performed_by else "",
            step.performed_by.Staff_Designation if step.performed_by else "",
            step.step_owner,
            batch.bat_Batch_Name if batch else "",
            tat.tat_Batch_Code,
            tat.tat_SiteCode,
            tat.tat_Referral_Date,
            step.step_type,
            step.step_config.target_days if step.step_config else "",
            step.date_received,
            step.date_finished,
            step.step_days_count,
            "Within" if step.within_tat else "Outside" if step.within_tat is False else "",
            step.remarks or "",
        ])

    _style_tat_export_sheet(ws)
    _style_tat_export_sheet(detail_ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"TAT_IPCR_BASIS_{now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url="login")
def download_tat_ipcr_template(request):
    date_from, date_to = _get_export_date_range(request)
    year = request.GET.get("tat_year", request.GET.get("year", str(timezone.localdate().year))).strip()
    include_ongoing = request.GET.get("include_ongoing", "")
    steps = _completed_tat_steps(date_from, date_to, year, include_ongoing)

    def _step_name(step):
        config_name = step.step_config.step_type if step.step_config else ""
        return f"{config_name} {step.step_type or ''}".lower()

    def _matches(step, *needles):
        text = _step_name(step)
        return any(needle.lower() in text for needle in needles)

    step_list = list(steps)

    def _counts(predicate):
        selected = [step for step in step_list if predicate(step)]
        within = sum(1 for step in selected if step.within_tat is True)
        outside = sum(1 for step in selected if step.within_tat is False)
        total = within + outside
        pct = within / total if total else None
        return within, outside, total, pct

    row_specs = [
        (
            "Lab Editing with lab director",
            lambda step: _matches(step, "draft - lab staff", "lab director"),
        ),
        (
            "DMU Editing",
            lambda step: _matches(step, "draft - dmu", "dmu data office", "dmu data officer"),
        ),
        (
            "ID and AST",
            lambda step: _matches(step, "identification and antimicrobial susceptibility", "id and ast"),
        ),
        (
            "Lab staffs editing",
            lambda step: _matches(step, "draft - lab staff"),
        ),
        (
            "Lab Director",
            lambda step: _matches(step, "lab director"),
        ),
        (
            "Concordance",
            lambda step: _matches(step, "concordance"),
        ),
        (
            "Encoding",
            lambda step: _matches(step, "encoding of consolidated data"),
        ),
        (
            "DMU Verification",
            lambda step: _matches(step, "checking of encoded", "printing of results"),
        ),
        (
            "Sorting Labelling batching queuing",
            lambda step: _matches(step, "receiving, checking, sorting", "batching and queuing"),
        ),
        (
            "Executive Secretary",
            lambda step: _matches(step, "secretariat"),
        ),
        (
            "MEDICAL SPECIALIST 4 / OIC",
            lambda step: _matches(step, "medical specialist", "oic"),
        ),
        (
            "Releasing of Result(DMU)",
            lambda step: _matches(step, "releasing of results"),
        ),
        (
            "Scanning",
            lambda step: _matches(step, "scanning"),
        ),
    ]

    if date_from or date_to:
        subtitle = f"{date_from or 'Start'} to {date_to or 'End'}"
    elif year and year != "all":
        subtitle = f"January to December {year}"
    else:
        subtitle = "All Years"

    if include_ongoing == "1":
        subtitle = f"{subtitle} (Ongoing TAT)"

    wb = Workbook()
    ws = wb.active
    ws.title = "IPCR TAT Summary"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="4472C4")
    stripe_fill = PatternFill("solid", fgColor="D9E2F3")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin_blue = Side(style="thin", color="8EAADB")
    medium_blue = Side(style="medium", color="1F4E79")
    black_medium = Side(style="medium", color="000000")
    border = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)

    ws["A1"] = "Confirmatory Test Turn-Around Time Report"
    ws["A2"] = subtitle
    for cell in ("A1", "A2"):
        ws[cell].font = Font(bold=True, size=12)

    headers = ["By Process", "Within", "Out", "Total", "%"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin_blue, right=thin_blue, top=medium_blue, bottom=medium_blue)

    start_row = 5
    for idx, (label, predicate) in enumerate(row_specs):
        row = start_row + idx
        within, outside, total, pct = _counts(predicate)
        values = [label, within, outside, total, pct if pct is not None else "N/A"]
        fill = stripe_fill if idx % 2 == 0 else white_fill
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
            if col == 1:
                cell.font = Font(bold=True)
            elif col == 5 and pct is not None:
                cell.number_format = "0.00%"
                cell.font = Font(bold=True)

    last_data_row = start_row + len(row_specs) - 1
    for col in range(1, 6):
        ws.cell(row=last_data_row, column=col).border = Border(
            left=thin_blue,
            right=thin_blue,
            top=thin_blue,
            bottom=black_medium,
        )

    signature_row = last_data_row + 4
    ws.cell(signature_row, 1, "Prepared By:")
    ws.cell(signature_row + 4, 1, "Noted By:")
    ws.cell(signature_row + 8, 1, "Approved By:")
    for row in (signature_row, signature_row + 4, signature_row + 8):
        ws.cell(row, 1).font = Font(bold=True)

    ws.column_dimensions["A"].width = 45
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 14
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{last_data_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:E{signature_row + 10}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename_year = year if year and year != "all" else "ALL_YEARS"
    mode_label = "ONGOING" if include_ongoing == "1" else "FINAL"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="TAT_IPCR_{mode_label}_{filename_year}_{now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    return response


@login_required(login_url="login")
def tat_review_view(request):

    configs = TATStepConfig.objects.all().order_by("order")

    tat_records = (
        TATform.objects
        .select_related("tat_Batch_Isolates")
        .prefetch_related("steps__step_config")
        .order_by("-tat_Referral_Date", "tat_Batch_Code")
    )

    current_year = timezone.localdate().year
    tat_q = request.GET.get("tat_q", "").strip()
    tat_site = request.GET.get("tat_site", "").strip()
    tat_year = request.GET.get("tat_year", str(current_year)).strip() or str(current_year)
    tat_status = request.GET.get("tat_status", "").strip()
    tat_chart_site = request.GET.get("tat_chart_site", "").strip()
    tat_chart_group = request.GET.get("tat_chart_group", "month").strip() or "month"
    tat_chart_metric = request.GET.get("tat_chart_metric", "avg_tat").strip() or "avg_tat"
    tat_chart_order = request.GET.get("tat_chart_order", "desc").strip() or "desc"
    if tat_chart_order not in {"asc", "desc"}:
        tat_chart_order = "desc"
    if tat_chart_group not in {"month", "assigned", "site", "status"}:
        tat_chart_group = "month"

    if tat_q:
        tat_records = tat_records.filter(
            Q(tat_Batch_Code__icontains=tat_q)
            | Q(tat_Batch_Isolates__bat_Batch_Name__icontains=tat_q)
            | Q(tat_Batch_Isolates__bat_RefNo__icontains=tat_q)
        )

    if tat_site:
        tat_records = tat_records.filter(tat_SiteCode=tat_site)

    if tat_year and tat_year != "all":
        tat_records = tat_records.filter(tat_Referral_Date__year=tat_year)

    if tat_status:
        tat_records = tat_records.filter(tat_Status_Release=tat_status)

    # ================= GLOBAL METRICS =================

    total_batches = tat_records.count()

    total_isolates = tat_records.aggregate(
        total=Sum("tat_Num_Isolate")
    )["total"] or 0

    released_batches = tat_records.filter(tat_Status_Release="Released").count()
    overdue_batches = tat_records.filter(tat_Status_Release="Overdue").count()
    ongoing_batches = tat_records.filter(tat_Status_Release="Ongoing").count()

    avg_running_tat = tat_records.aggregate(
        avg=Avg("tat_Running_TAT")
    )["avg"] or 0

    avg_running_tat = round(avg_running_tat, 1) if avg_running_tat else 0

    # ================= WITHIN TAT % =================

    released_within = tat_records.filter(
        tat_Status_Release="Released",
        tat_Running_TAT__lte=F("tat_Target_Days")
    ).count()

    released_total = released_batches

    within_tat_pct = 0
    if released_total > 0:
        within_tat_pct = round((released_within / released_total) * 100, 1)

    # ================= STEP METRICS =================

    scoped_steps = TATStep.objects.filter(tat__in=tat_records)
    total_steps = scoped_steps.count()
    steps_within = scoped_steps.filter(within_tat=True).count()
    steps_outside = scoped_steps.filter(within_tat=False).count()

    compliance_rate = 0
    if total_steps > 0:
        compliance_rate = round((steps_within / total_steps) * 100, 2)

    # ================= PER PROCESS =================

    process_summary = []

    for config in configs:
        steps = scoped_steps.filter(step_config=config)

        total = steps.count()
        within = steps.filter(within_tat=True).count()
        outside = steps.filter(within_tat=False).count()

        rate = round((within / total) * 100, 2) if total > 0 else 0

        process_summary.append({
            "config": config,
            "total": total,
            "within": within,
            "outside": outside,
            "rate": rate
        })

    tat_paginator = Paginator(tat_records, 50)
    tat_page_obj = tat_paginator.get_page(request.GET.get("page"))
    running_rows = _build_tat_running_rows(tat_page_obj.object_list, configs)
    tat_years = list(
        TATform.objects
        .exclude(tat_Referral_Date__isnull=True)
        .annotate(year=ExtractYear("tat_Referral_Date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )
    if current_year not in tat_years:
        tat_years.insert(0, current_year)

    tat_site_choices = list(
        TATform.objects
        .exclude(tat_SiteCode__isnull=True)
        .exclude(tat_SiteCode__exact="")
        .values_list("tat_SiteCode", flat=True)
        .distinct()
        .order_by("tat_SiteCode")
    )
    chart_metric_options = [
        {"value": "avg_tat", "label": "Average TAT Days"},
        {"value": "within_rate", "label": "Within TAT %"},
        {"value": "released", "label": "Released Batches"},
        {"value": "ongoing", "label": "Ongoing Batches"},
        {"value": "overdue", "label": "Overdue Batches"},
    ]
    chart_metric_labels = {item["value"]: item["label"] for item in chart_metric_options}
    if tat_chart_metric not in chart_metric_labels:
        tat_chart_metric = "avg_tat"
    chart_group_options = [
        {"value": "month", "label": "Month"},
        {"value": "assigned", "label": "Assigned Person"},
        {"value": "site", "label": "Site"},
        {"value": "status", "label": "Status"},
    ]
    chart_group_labels = {item["value"]: item["label"] for item in chart_group_options}

    chart_records = (
        TATform.objects
        .select_related("tat_Batch_Isolates")
        .prefetch_related("steps__performed_by")
        .order_by("tat_SiteCode")
    )
    if tat_year and tat_year != "all":
        chart_records = chart_records.filter(tat_Referral_Date__year=tat_year)
    if tat_status:
        chart_records = chart_records.filter(tat_Status_Release=tat_status)
    if tat_chart_group == "site" and tat_chart_site:
        chart_records = chart_records.filter(tat_SiteCode=tat_chart_site)

    chart_buckets = {}
    for tat in chart_records:
        if tat_chart_group == "month":
            if tat.tat_Referral_Date:
                bucket_label = tat.tat_Referral_Date.strftime("%b %Y")
            else:
                bucket_label = "No referral date"
        elif tat_chart_group == "assigned":
            steps = list(tat.steps.all())
            latest_step = max(steps, key=lambda step: step.id, default=None)
            bucket_label = str(latest_step.performed_by) if latest_step and latest_step.performed_by else "Unassigned"
        elif tat_chart_group == "status":
            bucket_label = tat.tat_Status_Release or "No status"
        else:
            bucket_label = tat.tat_SiteCode or "No site"

        bucket = chart_buckets.setdefault(bucket_label, {
            "label": bucket_label,
            "total": 0,
            "isolates": 0,
            "released": 0,
            "ongoing": 0,
            "overdue": 0,
            "within": 0,
            "tat_total": 0,
            "tat_count": 0,
        })
        effective_days = _get_tat_effective_days(tat)
        bucket["total"] += 1
        bucket["isolates"] += tat.tat_Num_Isolate or 0
        if effective_days is not None:
            bucket["tat_total"] += effective_days
            bucket["tat_count"] += 1
        if tat.tat_Status_Release == "Released":
            bucket["released"] += 1
            if effective_days is not None and tat.tat_Target_Days and effective_days <= tat.tat_Target_Days:
                bucket["within"] += 1
        elif tat.tat_Status_Release == "Overdue":
            bucket["overdue"] += 1
        elif tat.tat_Status_Release == "Ongoing":
            bucket["ongoing"] += 1

    tat_chart_rows = []
    for row in chart_buckets.values():
        released_count = row["released"]
        within_rate = round((row["within"] / released_count) * 100, 1) if released_count else 0
        avg_tat_value = round(row["tat_total"] / row["tat_count"], 1) if row["tat_count"] else 0
        metric_values = {
            "avg_tat": avg_tat_value,
            "within_rate": within_rate,
            "released": row["released"],
            "ongoing": row["ongoing"],
            "overdue": row["overdue"],
        }
        tat_chart_rows.append({
            "label": row["label"],
            "site": row["label"],
            "total": row["total"],
            "isolates": row["isolates"],
            "avg_tat": avg_tat_value,
            "within_rate": within_rate,
            "released": row["released"],
            "ongoing": row["ongoing"],
            "overdue": row["overdue"],
            "metric_value": metric_values[tat_chart_metric],
        })

    tat_chart_rows = sorted(
        tat_chart_rows,
        key=lambda item: (item["metric_value"], item["total"]),
        reverse=tat_chart_order == "desc",
    )[:12]
    chart_max_value = max([row["metric_value"] for row in tat_chart_rows] or [0])
    for row in tat_chart_rows:
        row["bar_pct"] = round((row["metric_value"] / chart_max_value) * 100, 1) if chart_max_value else 0

    chart_totals = chart_records.aggregate(
        total=Count("id"),
        isolates=Sum("tat_Num_Isolate"),
        avg_tat=Avg("tat_Running_TAT"),
        released=Count("id", filter=Q(tat_Status_Release="Released")),
        ongoing=Count("id", filter=Q(tat_Status_Release="Ongoing")),
        overdue=Count("id", filter=Q(tat_Status_Release="Overdue")),
        within=Count(
            "id",
            filter=Q(
                tat_Status_Release="Released",
                tat_Running_TAT__lte=F("tat_Target_Days"),
            ),
        ),
    )
    chart_released_total = chart_totals["released"] or 0
    chart_within_rate = (
        round(((chart_totals["within"] or 0) / chart_released_total) * 100, 1)
        if chart_released_total else 0
    )

    context = {
        "configs": configs,
        "tat_records": tat_records,
        "tat_page_obj": tat_page_obj,
        "tat_q": tat_q,
        "tat_site": tat_site,
        "tat_year": tat_year,
        "tat_status": tat_status,
        "tat_status_choices": ["Ongoing", "Released", "Overdue"],
        "tat_sites": tat_site_choices,
        "tat_chart_site": tat_chart_site,
        "tat_chart_group": tat_chart_group,
        "tat_chart_metric": tat_chart_metric,
        "tat_chart_order": tat_chart_order,
        "tat_chart_metric_label": chart_metric_labels[tat_chart_metric],
        "tat_chart_metric_options": chart_metric_options,
        "tat_chart_group_label": chart_group_labels[tat_chart_group],
        "tat_chart_group_options": chart_group_options,
        "tat_site_chart_rows": tat_chart_rows,
        "tat_chart_totals": {
            "total": chart_totals["total"] or 0,
            "isolates": chart_totals["isolates"] or 0,
            "avg_tat": round(chart_totals["avg_tat"] or 0, 1),
            "released": chart_released_total,
            "ongoing": chart_totals["ongoing"] or 0,
            "overdue": chart_totals["overdue"] or 0,
            "within_rate": chart_within_rate,
        },
        "tat_years": tat_years,
        "current_year": current_year,
        "tat_year_label": "All Years" if tat_year == "all" else tat_year,
        "total_batches": total_batches,
        "total_isolates": total_isolates,
        "released_batches": released_batches,
        "overdue_batches": overdue_batches,
        "ongoing_batches": ongoing_batches,
        "avg_running_tat": avg_running_tat,
        "within_tat_pct": within_tat_pct,
        "total_steps": total_steps,
        "steps_within": steps_within,
        "steps_outside": steps_outside,
        "compliance_rate": compliance_rate,
        "process_summary": process_summary,
        "running_rows": running_rows,
    }

    return render(request, "home/tat_dashboard.html", context)


@login_required(login_url="login")
def tat_running_list(request):
    configs = TATStepConfig.objects.all().order_by("order")
    tat_records = (
        TATform.objects
        .select_related("tat_Batch_Isolates")
        .prefetch_related("steps__step_config")
    )

    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    year = request.GET.get("year", "").strip()

    if q:
        tat_records = tat_records.filter(
            Q(tat_Batch_Code__icontains=q)
            | Q(tat_SiteCode__icontains=q)
            | Q(tat_Batch_Isolates__bat_Batch_Name__icontains=q)
            | Q(tat_Batch_Isolates__bat_RefNo__icontains=q)
        )

    if status:
        tat_records = tat_records.filter(tat_Status_Release=status)

    if year and year != "all":
        tat_records = tat_records.filter(tat_Referral_Date__year=year)

    sort_context = _tat_running_sort_context(request)
    tat_records = _annotate_tat_running_sort_fields(tat_records).order_by(*sort_context["order_by"])

    paginator = Paginator(tat_records, 50)
    page_obj = paginator.get_page(request.GET.get("page"))
    running_rows = _build_tat_running_rows(page_obj.object_list, configs)
    tat_location_choices = list(
        TATLocation.objects
        .filter(is_active=True)
        .order_by("order", "name")
        .values_list("name", flat=True)
    )
    if "n/a" not in tat_location_choices:
        tat_location_choices.append("n/a")
    for tat in page_obj.object_list:
        current_location = (tat.tat_Batch_Location or "").strip()
        if current_location and current_location not in tat_location_choices:
            tat_location_choices.append(current_location)
    years = (
        TATform.objects
        .exclude(tat_Referral_Date__isnull=True)
        .annotate(year=ExtractYear("tat_Referral_Date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )

    return render(
        request,
        "home/tat_running_list.html",
        {
            "configs": configs,
            "running_rows": running_rows,
            "page_obj": page_obj,
            "total_rows": paginator.count,
            "q": q,
            "status": status,
            "year": year,
            "years": years,
            "status_choices": ["Ongoing", "Released", "Overdue"],
            "tat_location_choices": tat_location_choices,
            "sort_links": sort_context["sort_links"],
            "current_sort": sort_context["current_sort"],
            "current_order": sort_context["current_order"],
            "page_query": sort_context["page_query"],
        },
    )


@login_required(login_url="login")
def tat_running_process_list(request):
    configs = TATStepConfig.objects.all().order_by("order")
    tat_records = (
        TATform.objects
        .select_related("tat_Batch_Isolates")
        .prefetch_related("steps__step_config")
    )

    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    year = request.GET.get("year", "").strip()

    if q:
        tat_records = tat_records.filter(
            Q(tat_Batch_Code__icontains=q)
            | Q(tat_SiteCode__icontains=q)
            | Q(tat_Batch_Isolates__bat_Batch_Name__icontains=q)
            | Q(tat_Batch_Isolates__bat_RefNo__icontains=q)
        )

    if status:
        tat_records = tat_records.filter(tat_Status_Release=status)

    if year and year != "all":
        tat_records = tat_records.filter(tat_Referral_Date__year=year)

    sort_context = _tat_running_sort_context(request)
    tat_records = _annotate_tat_running_sort_fields(tat_records).order_by(*sort_context["order_by"])

    paginator = Paginator(tat_records, 50)
    page_obj = paginator.get_page(request.GET.get("page"))
    running_rows = _build_tat_running_rows(page_obj.object_list, configs)
    years = (
        TATform.objects
        .exclude(tat_Referral_Date__isnull=True)
        .annotate(year=ExtractYear("tat_Referral_Date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )

    return render(
        request,
        "home/tat_running_process_list.html",
        {
            "configs": configs,
            "running_rows": running_rows,
            "page_obj": page_obj,
            "total_rows": paginator.count,
            "q": q,
            "status": status,
            "year": year,
            "years": years,
            "status_choices": ["Ongoing", "Released", "Overdue"],
            "sort_links": sort_context["sort_links"],
            "current_sort": sort_context["current_sort"],
            "current_order": sort_context["current_order"],
            "page_query": sort_context["page_query"],
        },
    )


@login_required(login_url="login")
@require_POST
def update_tat_scanning_flags(request, pk):
    tat = get_object_or_404(TATform, pk=pk)
    allowed_fields = {
        "tat_Scanning_raw",
        "tat_Scanning_ws",
        "tat_Scanning_final",
    }
    field = request.POST.get("field", "")
    if field not in allowed_fields:
        return JsonResponse({"ok": False, "error": "Invalid scanning field."}, status=400)

    value = request.POST.get("value", "").strip().lower() in {"1", "true", "yes", "on"}
    setattr(tat, field, value)
    tat.save(update_fields=[
        field,
        "tat_Running_TAT",
        "tat_Final_TAT",
        "tat_Status_Release",
        "tat_Date_Last_Update",
    ])
    return JsonResponse({
        "ok": True,
        "field": field,
        "value": value,
        "last_update": tat.tat_Date_Last_Update.isoformat() if tat.tat_Date_Last_Update else "",
    })


@login_required(login_url="login")
@require_POST
def update_tat_location(request, pk):
    tat = get_object_or_404(TATform, pk=pk)
    location = (request.POST.get("location") or "").strip() or "n/a"
    is_valid_location = (
        location == "n/a"
        or TATLocation.objects.filter(name__iexact=location, is_active=True).exists()
    )

    if not is_valid_location:
        return JsonResponse({"ok": False, "error": "Invalid TAT location."}, status=400)

    configured_location = _tat_location_value(location, tat.tat_Batch_Location)
    tat.tat_Batch_Location = configured_location
    tat.save(update_fields=[
        "tat_Batch_Location",
        "tat_Running_TAT",
        "tat_Final_TAT",
        "tat_Status_Release",
        "tat_Date_Last_Update",
    ])
    return JsonResponse({
        "ok": True,
        "location": tat.tat_Batch_Location,
        "last_update": tat.tat_Date_Last_Update.isoformat() if tat.tat_Date_Last_Update else "",
    })










########## tat configuration 


@login_required(login_url="login")
def add_tat_step_config(request):

    if request.method != "POST":
        return redirect("/settings/?tab=tat_config")

    form = TATStepConfigForm(request.POST)

    if form.is_valid():
        form.save()
        messages.success(request, "TAT Step Configuration added successfully.")
    else:
        messages.warning(request, _first_form_error(form, "Failed to add TAT Step Configuration."))
        print(form.errors)

    return redirect("/settings/?tab=tat_config")



@login_required(login_url="login")
def edit_tat_step_config(request, pk):

    config = get_object_or_404(TATStepConfig, pk=pk)

    if request.method != "POST":
        return redirect(f"/settings/?tab=tat_config&edit_tat_config={config.pk}")

    form = TATStepConfigForm(request.POST, instance=config)

    if form.is_valid():
        form.save()
        messages.success(request, "TAT Step Configuration updated successfully.")
    else:
        messages.error(request, "Failed to update configuration.")
        print(form.errors)

    return redirect("/settings/?tab=tat_config")


@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER)
def update_tat_overall_setting(request):
    if request.method != "POST":
        return redirect("/settings/?tab=tat_config")

    setting = TATOverallSetting.get_solo()
    form = TATOverallSettingForm(request.POST, instance=setting)

    if form.is_valid():
        setting = form.save()
        for tat in TATform.objects.all().iterator():
            tat.tat_Target_Days = setting.target_days
            tat.save()
        messages.success(request, "Overall TAT target days updated successfully.")
    else:
        messages.warning(request, _first_form_error(form, "Failed to update overall TAT target days."))

    return redirect("/settings/?tab=tat_config")


@login_required(login_url="login")
def add_tat_location(request):
    if request.method != "POST":
        return redirect("/settings/?tab=tat_config")

    form = TATLocationForm(request.POST)

    if form.is_valid():
        form.save()
        messages.success(request, "TAT location added successfully.")
    else:
        messages.warning(request, _first_form_error(form, "Failed to add TAT location."))

    return redirect("/settings/?tab=tat_config")


@login_required(login_url="login")
def edit_tat_location(request, pk):
    location = get_object_or_404(TATLocation, pk=pk)

    if request.method != "POST":
        return redirect("/settings/?tab=tat_config")

    form = TATLocationForm(request.POST, instance=location)

    if form.is_valid():
        form.save()
        messages.success(request, "TAT location updated successfully.")
    else:
        messages.warning(request, _first_form_error(form, "Failed to update TAT location."))

    return redirect("/settings/?tab=tat_config")


@login_required(login_url="login")
def delete_tat_location(request, pk):
    location = get_object_or_404(TATLocation, pk=pk)
    name = location.name
    if TATform.objects.filter(tat_Batch_Location=name).exists():
        location.is_active = False
        location.save(update_fields=["is_active"])
        messages.warning(
            request,
            "That location is already used by TAT records, so it was deactivated instead of deleted."
        )
    else:
        location.delete()
        messages.success(request, "TAT location deleted successfully.")

    return redirect("/settings/?tab=tat_config")



def tat_step_config_list(request):
    q = request.GET.get("q", "")

    configs = TATStepConfig.objects.all()

    if q:
        configs = configs.filter(
            Q(step_type__icontains=q) |
            Q(step_owner__icontains=q)
        )

    paginator = Paginator(configs, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "q": q,
    }

    return render(
        request,
        "settings/tat_config_list.html",
        context
    )



@login_required(login_url="login")
def upload_tat_step_config(request):

    if request.method != "POST":
        return redirect("/settings/?tab=tat_config")

    form = TATStepConfigUploadForm(request.POST, request.FILES)

    if form.is_valid():

        upload_instance = form.save()

        try:
            df = read_tabular_upload(upload_instance.tat_file)

            required_columns = ['step_type', 'step_owner', 'target_days', 'order']

            for col in required_columns:
                if col not in df.columns:
                    messages.error(request, f"Missing column: {col}")
                    return redirect("/settings/?tab=tat_config")

            created_count = 0
            duplicate_count = 0
            with transaction.atomic():
                for _, row in df.iterrows():
                    step_type = str(row['step_type']).strip()
                    step_owner = str(row['step_owner']).strip()
                    if TATStepConfig.objects.filter(step_type__iexact=step_type, step_owner__iexact=step_owner).exists():
                        duplicate_count += 1
                        continue
                    TATStepConfig.objects.create(
                        step_type=step_type,
                        step_owner=step_owner,
                        target_days=int(row['target_days']),
                        order=int(row['order']),
                    )
                    created_count += 1

            messages.success(request, f"TAT Step Config uploaded. {created_count} created, {duplicate_count} skipped because they already exist.")

        except Exception as e:
            messages.error(request, f"Upload failed: {str(e)}")

    else:
        messages.error(request, "Invalid upload form.")

    return redirect("/settings/?tab=tat_config")


def delete_all_tat_process(request):
    # This will now work when you click the <a> link
    total = TATStepConfig.objects.count()
    TATStepConfig.objects.all().delete()

    messages.success(
        request, 
        f"Successfully deleted {total} TAT records and all related process steps."
    )
    return redirect('/settings/?tab=tat_config')


@login_required(login_url="login")
def export_tat_step_config(request):
    return _export_queryset_to_excel(
        TATStepConfig.objects.all().order_by("order"),
        [
            ("step_type", "step_type"),
            ("step_owner", "step_owner"),
            ("target_days", "target_days"),
            ("order", "order"),
        ],
        "TAT_step_config.xlsx",
    )


######### TAT NON-WORKING DAYS


@login_required(login_url="login")
def add_non_working_day(request):

    if request.method != "POST":
        return redirect("/settings/?tab=non_working")

    form = NonWorkingDayForm(request.POST)

    if form.is_valid():
        form.save()
        messages.success(request, "Non-working day added successfully.")
    else:
        messages.warning(request, _first_form_error(form, "Failed to add non-working day."))

    return redirect("/settings/?tab=non_working")


@login_required(login_url="login")
def edit_non_working_day(request, pk):
    day = get_object_or_404(NonWorkingDay, pk=pk)

    if request.method != "POST":
        return redirect(f"/settings/?tab=non_working&edit_non_working={pk}")

    form = NonWorkingDayForm(request.POST, instance=day)

    if form.is_valid():
        form.save()
        messages.success(request, "Non-working day updated successfully.")
    else:
        messages.warning(request, _first_form_error(form, "Failed to update non-working day."))
        return redirect(f"/settings/?tab=non_working&edit_non_working={pk}")

    return redirect("/settings/?tab=non_working")


@login_required(login_url="login")
def delete_non_working_day(request, pk):
    day = get_object_or_404(NonWorkingDay, pk=pk)
    day.delete()
    messages.success(request, "Non-working day deleted.")
    return redirect("/settings/?tab=non_working")


@login_required(login_url="login")
@require_POST
def delete_all_non_working_days(request):
    deleted_count = NonWorkingDay.objects.count()
    NonWorkingDay.objects.all().delete()
    messages.success(request, f"Deleted {deleted_count} configured non-working day(s).")
    return redirect("/settings/?tab=non_working")




# @login_required(login_url="login")
# def tat_analysis(request):

#     # Annotate effective TAT (Final if released, Running if ongoing)
#     tat_qs = TATform.objects.annotate(
#         effective_tat=Case(
#             When(tat_Final_TAT__isnull=False, then=F("tat_Final_TAT")),
#             default=F("tat_Running_TAT"),
#             output_field=IntegerField()
#         )
#     )

#     # Only include records that have some TAT value
#     tat_qs = tat_qs.filter(effective_tat__isnull=False)

#     if not tat_qs.exists():
#         return render(request, "home/tat_analysis.html", {
#             "stats": None,
#             "process_analysis": None,
#         })

#     # ===============================
#     # OVERALL BATCH METRICS
#     # ===============================

#     total_batches = tat_qs.count()

#     avg_tat = tat_qs.aggregate(avg=Avg("effective_tat"))["avg"] or 0
#     min_tat = tat_qs.aggregate(min=Min("effective_tat"))["min"] or 0
#     max_tat = tat_qs.aggregate(max=Max("effective_tat"))["max"] or 0

#     within_count = tat_qs.filter(
#         effective_tat__lte=F("tat_Target_Days")
#     ).count()

#     out_count = total_batches - within_count

#     within_pct = round((within_count / total_batches) * 100, 2) if total_batches else 0

#     stats = {
#         "total_batches": total_batches,
#         "average_tat": avg_tat,
#         "min_tat": min_tat,
#         "max_tat": max_tat,
#         "within_tat_count": within_count,
#         "out_of_tat_count": out_count,
#         "within_tat_pct": within_pct,
#         "released_count": tat_qs.filter(tat_Status_Release="Released").count(),
#         "ongoing_count": tat_qs.filter(tat_Status_Release="Ongoing").count(),
#         "overdue_count": tat_qs.filter(tat_Status_Release="Overdue").count(),
#     }

#     # ===============================
#     # PROCESS PERFORMANCE ANALYSIS
#     # ===============================

#     step_qs = (
#         TATStep.objects
#         .filter(step_days_count__isnull=False)
#         .values(
#             "step_config__step_type",
#             "step_config__step_owner",
#             "step_config__target_days",
#         )
#         .annotate(
#             average_tat=Avg("step_days_count"),
#             min_tat=Min("step_days_count"),
#             max_tat=Max("step_days_count"),
#             total_batches=Count("id"),
#             within_target=Count("id", filter=Q(within_tat=True)),
#             out_of_target=Count("id", filter=Q(within_tat=False)),
#         )
#         .order_by("step_config__order")
#     )

#     process_analysis = []

#     for s in step_qs:
#         total = s["total_batches"] or 1
#         pct = round((s["within_target"] / total) * 100, 2)

#         process_analysis.append({
#             "process_name": s["step_config__step_type"],
#             "owner": s["step_config__step_owner"],
#             "target_days": s["step_config__target_days"],
#             "average_tat": s["average_tat"] or 0,
#             "min_tat": s["min_tat"] or 0,
#             "max_tat": s["max_tat"] or 0,
#             "within_target": s["within_target"],
#             "out_of_target": s["out_of_target"],
#             "total_batches": total,
#             "within_target_pct": pct,
#         })


#     # user performance analysis (based on steps performed)

#     user_qs = (
#         TATStep.objects
#         .filter(
#             step_days_count__isnull=False,
#             performed_by__isnull=False
#         )
#         .values(
#             "performed_by__id",
#             "performed_by__Staff_Name",
#             "performed_by__Staff_Designation",
#         )
#         .annotate(
#             total_steps=Count("id"),
#             average_tat=Avg("step_days_count"),
#             min_tat=Min("step_days_count"),
#             max_tat=Max("step_days_count"),
#             within_target=Count("id", filter=Q(within_tat=True)),
#             out_of_target=Count("id", filter=Q(within_tat=False)),
#         )
#         .order_by("-total_steps")
#     )

#     user_analysis = []

#     for u in user_qs:
#         total = u["total_steps"] or 1
#         pct = round((u["within_target"] / total) * 100, 2)

#         user_analysis.append({
#             "user_id": u["performed_by__id"],
#             "staff_name": u["performed_by__Staff_Name"],
#             "designation": u["performed_by__Staff_Designation"],
#             "total_steps": total,
#             "average_tat": u["average_tat"] or 0,
#             "min_tat": u["min_tat"] or 0,
#             "max_tat": u["max_tat"] or 0,
#             "within_target": u["within_target"],
#             "out_of_target": u["out_of_target"],
#             "compliance_pct": pct,
#         })


#     current_batches = (
#         TATform.objects
#         .select_related("tat_Batch_Isolates")
#         .filter(tat_Status_Release__in=["Ongoing", "Overdue"])
#         .order_by("tat_Referral_Date")
#         )

#     return render(request, "home/tat_analysis.html", {
#         "stats": stats,
#         "process_analysis": process_analysis,
#         "user_analysis": user_analysis,
#         "current_batches": current_batches,
#     })




@login_required(login_url="login")
def tat_analysis(request):

    # ==========================================
    # TOGGLE: Released Only vs Include Ongoing
    # ==========================================

    current_year = timezone.localdate().year
    include_ongoing = request.GET.get("include_ongoing")
    tat_year = request.GET.get("tat_year", str(current_year)).strip() or str(current_year)

    if include_ongoing == "1":
        completed = TATform.objects.filter(
            tat_Running_TAT__isnull=False,
            tat_Date_Released__isnull=True,
        ).exclude(tat_Status_Release="Released")
        mode = "ONGOING"
    else:
        completed = TATform.objects.filter(
            Q(tat_Status_Release="Released") | Q(tat_Date_Released__isnull=False),
            tat_Final_TAT__isnull=False
        )
        mode = "RELEASED"

    if tat_year != "all":
        completed = completed.filter(tat_Referral_Date__year=tat_year)

    total_batches = completed.count()

    # ==========================================
    # BATCH-LEVEL STATISTICS
    # ==========================================

    if total_batches > 0:

        effective_values = []
        within_tat_count = 0
        out_of_tat_count = 0

        for tat in completed:
            effective_tat = (
                tat.tat_Final_TAT
                if tat.tat_Status_Release == "Released" and tat.tat_Final_TAT is not None
                else tat.tat_Running_TAT
            )

            if effective_tat is None:
                continue

            effective_values.append(effective_tat)

            if tat.tat_Target_Days is not None:
                if effective_tat <= tat.tat_Target_Days:
                    within_tat_count += 1
                else:
                    out_of_tat_count += 1

        average_tat = sum(effective_values) / len(effective_values) if effective_values else 0
        min_tat = min(effective_values) if effective_values else 0
        max_tat = max(effective_values) if effective_values else 0

        within_pct = round(
            (within_tat_count / total_batches) * 100,
            2
        )

    else:
        average_tat = 0
        min_tat = 0
        max_tat = 0
        within_tat_count = 0
        out_of_tat_count = 0
        within_pct = 0

    stats = {
        "total_batches": total_batches,
        "average_tat": round(average_tat, 2),
        "min_tat": min_tat,
        "max_tat": max_tat,
        "within_tat_count": within_tat_count,
        "out_of_tat_count": out_of_tat_count,
        "within_tat_pct": within_pct,
    }

    # ==========================================
    # PROCESS PERFORMANCE
    # ==========================================

    process_analysis = []

    configs = TATStepConfig.objects.all()

    for config in configs:

        steps = TATStep.objects.filter(
            tat__in=completed,
            step_config=config,
            date_finished__isnull=False,
            step_days_count__isnull=False,
        )

        total = steps.count()

        if total == 0:
            continue

        agg = steps.aggregate(
            avg=Avg("step_days_count"),
            min=Min("step_days_count"),
            max=Max("step_days_count"),
        )

        avg_days = agg["avg"] or 0
        min_days = agg["min"]
        max_days = agg["max"]

        within = steps.filter(within_tat=True).count()
        outside = steps.filter(within_tat=False).count()

        pct = round((within / total) * 100, 1)

        process_analysis.append({
            "process_name": config.step_type,
            "target_days": config.target_days,
            "average_tat": round(avg_days, 2),
            "min_tat": min_days,
            "max_tat": max_days,
            "within_target": within,
            "out_of_target": outside,
            "total_batches": total,
            "within_target_pct": pct,
        })

    # ==========================================
    # OWNER PERFORMANCE (LAB / DMU)
    # ==========================================

    owner_summary = []

    for owner in ["LAB", "DMU"]:

        steps = TATStep.objects.filter(
            tat__in=completed,
            step_owner=owner,
            date_finished__isnull=False,
            step_days_count__isnull=False,
        )

        total = steps.count()

        if total == 0:
            continue

        within = steps.filter(within_tat=True).count()
        outside = steps.filter(within_tat=False).count()

        avg_days = steps.aggregate(
            avg=Avg("step_days_count")
        )["avg"] or 0

        compliance = round((within / total) * 100, 2)

        owner_summary.append({
            "owner": owner,
            "total": total,
            "within": within,
            "outside": outside,
            "average_days": round(avg_days, 2),
            "compliance_pct": compliance,
        })

    # ==========================================
    # STAFF PERFORMANCE (IPCR)
    # ==========================================

    staff_summary = (
        TATStep.objects
        .filter(
            tat__in=completed,
            performed_by__isnull=False,
            date_finished__isnull=False,
            step_days_count__isnull=False,
        )
        .values(
            "performed_by__id",
            "performed_by__Staff_Name",
            "performed_by__Staff_Designation",
        )
        .annotate(
            total_steps=Count("id"),
            avg_days=Avg("step_days_count"),
            within_count=Count("id", filter=Q(within_tat=True)),
            outside_count=Count("id", filter=Q(within_tat=False)),
        )
        .order_by("-total_steps")
    )

    staff_performance = []

    for s in staff_summary:

        total = s["total_steps"]
        within = s["within_count"]

        compliance = round((within / total) * 100, 2) if total > 0 else 0

        staff_performance.append({
            "staff_name": s["performed_by__Staff_Name"],
            "designation": s["performed_by__Staff_Designation"],
            "total_steps": total,
            "average_days": round(s["avg_days"], 2) if s["avg_days"] else 0,
            "within_target": within,
            "outside_target": s["outside_count"],
            "compliance_pct": compliance,
        })

    # ==========================================
    # FINAL CONTEXT
    # ==========================================

    tat_years = list(
        TATform.objects
        .exclude(tat_Referral_Date__isnull=True)
        .annotate(year=ExtractYear("tat_Referral_Date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )
    if current_year not in tat_years:
        tat_years.insert(0, current_year)

    context = {
        "stats": stats,
        "process_analysis": process_analysis,
        "owner_summary": owner_summary,
        "staff_performance": staff_performance,
        "mode": mode,
        "include_ongoing": include_ongoing,
        "tat_year": tat_year,
        "tat_years": tat_years,
        "tat_year_label": "All Years" if tat_year == "all" else tat_year,
    }

    return render(request, "home/tat_analysis.html", context)
