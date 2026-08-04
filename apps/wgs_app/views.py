
from collections import defaultdict
import datetime
from decimal import Decimal, InvalidOperation
from io import TextIOWrapper
import io
import re
from django.db import models, transaction
import csv
from django.db.models import Q, F, Func
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import render, redirect

from apps.home.views import nz
from apps.home.permissions import (
    ROLE_ADMIN,
    ROLE_CHECKER,
    ROLE_ENCODER,
    ROLE_LAB_ENCODER,
    ROLE_LAB_MANAGER,
    ROLE_VERIFIER,
    can_manage_batch,
    role_flags,
    role_required,
)

from .forms import *
from apps.home.forms import *
from apps.home_final.forms import *
import pandas as pd
from apps.home.models import *
from apps.home_final.models import *
from .models import *
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.contrib import messages
import os
from django.core.paginator import Paginator
import re
from .utils import format_accession
from django.contrib.auth.decorators import login_required
from django.conf import settings
from datetime import datetime
from django.utils.dateparse import parse_date
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
import openpyxl
from datetime import date, datetime, timedelta

WGS_WRITE_ALLOWED_ROLES = (ROLE_ADMIN, ROLE_CHECKER, ROLE_LAB_ENCODER)
UPLOAD_CENTER_ALLOWED_ROLES = (ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER, ROLE_LAB_ENCODER)


# helper to read uploaded file (csv or excel)
def read_uploaded_file(uploaded_file, sheet_name=None):
    import pandas as pd

    filename = uploaded_file.name.lower()
    if filename.endswith('.csv'):
        return pd.read_csv(uploaded_file)
    elif filename.endswith(('.xls', '.xlsx')):
        if sheet_name:
            excel_file = pd.ExcelFile(uploaded_file)
            requested_sheets = (
                [sheet_name]
                if isinstance(sheet_name, str)
                else list(sheet_name)
            )
            normalized_sheets = {
                str(name).strip().lower(): name
                for name in excel_file.sheet_names
            }
            for requested_sheet in requested_sheets:
                matched_sheet = normalized_sheets.get(str(requested_sheet).strip().lower())
                if matched_sheet:
                    return pd.read_excel(excel_file, sheet_name=matched_sheet)
            raise ValueError(
                "Required sheet not found. Expected one of "
                f"{requested_sheets}; available sheets: {excel_file.sheet_names}"
            )
        return pd.read_excel(uploaded_file)
    else:
        raise ValueError("Unsupported file format. Please upload a CSV or Excel file.")


def get_or_create_wgs_project_for_upload(
    accession,
    referred_obj,
    accession_field,
    summary_field,
    existing_project=None,
):
    """
    Re-use the WGS connector for a specific WGS sample/run during uploads.

    One Final_Data isolate can have many WGS_Project rows. Therefore this
    helper must not collapse all rows for the same accession into one project.
    Callers should pass existing_project only when they have matched the same
    module sample/run identity (for example BactScout.name/sample_id). If no
    exact sample/run match exists, a new WGS_Project is created.
    """
    accession = (accession or "").strip()
    project = None

    project = existing_project

    if not project:
        project = WGS_Project.objects.create(
            Ref_Accession=referred_obj if referred_obj else None,
            WGS_SampleInfoSummary=False,
            WGS_BactScoutSummary=False,
            WGS_GtdbTkSummary=False,
            WGS_GambitSummary=False,
            WGS_MlstSummary=False,
            WGS_Checkm2Summary=False,
            WGS_AssemblySummary=False,
            WGS_AmrfinderSummary=False,
        )

    if referred_obj and project.Ref_Accession_id != referred_obj.f_AccessionNo:
        project.Ref_Accession = referred_obj

    setattr(project, accession_field, accession)
    setattr(
        project,
        summary_field,
        bool(accession)
        and bool(project.Ref_Accession)
        and accession == getattr(project.Ref_Accession, "f_AccessionNo", None)
    )
    project.save(update_fields=["Ref_Accession", accession_field, summary_field])
    return project


def find_existing_wgs_module_record(model, accession_field, accession, sample_field, sample_value):
    accession = (accession or "").strip()
    sample_value = (sample_value or "").strip()
    if not sample_value:
        return None
    return (
        model.objects
        .filter(**{accession_field: accession, sample_field: sample_value})
        .select_related()
        .order_by("-id")
        .first()
    )


def find_existing_wgs_module_record_by_fields(model, match_fields):
    cleaned_fields = {}
    for field, value in match_fields.items():
        try:
            if pd.isna(value):
                value = ""
        except (TypeError, ValueError):
            pass
        cleaned_fields[field] = value.strip() if isinstance(value, str) else value
    if not any(value not in (None, "") for value in cleaned_fields.values()):
        return None
    return (
        model.objects
        .filter(**cleaned_fields)
        .select_related()
        .order_by("-id")
        .first()
    )


WGS_SAMPLEINFO_LINK_FIELDS = {
    "BactScout": {
        "project_field": "bactscout_project",
        "accession_field": "BactScout_Accession",
        "project_accession_field": "WGS_BactScout_Acc",
        "summary_field": "WGS_BactScoutSummary",
        "sample_fields": ("name",),
    },
    "GtdbTk": {
        "project_field": "gtdbtk_project",
        "accession_field": "GtdbTk_Accession",
        "project_accession_field": "WGS_GtdbTk_Acc",
        "summary_field": "WGS_GtdbTkSummary",
        "sample_fields": ("user_genome",),
    },
    "Gambit": {
        "project_field": "gambit_project",
        "accession_field": "Gambit_Accession",
        "project_accession_field": "WGS_Gambit_Acc",
        "summary_field": "WGS_GambitSummary",
        "sample_fields": ("sample",),
    },
    "Mlst": {
        "project_field": "mlst_project",
        "accession_field": "Mlst_Accession",
        "project_accession_field": "WGS_Mlst_Acc",
        "summary_field": "WGS_MlstSummary",
        "sample_fields": ("name",),
    },
    "Checkm2": {
        "project_field": "checkm2_project",
        "accession_field": "Checkm2_Accession",
        "project_accession_field": "WGS_Checkm2_Acc",
        "summary_field": "WGS_Checkm2Summary",
        "sample_fields": ("Name",),
    },
    "AssemblyScan": {
        "project_field": "assembly_project",
        "accession_field": "Assembly_Accession",
        "project_accession_field": "WGS_Assembly_Acc",
        "summary_field": "WGS_AssemblySummary",
        "sample_fields": ("sample",),
    },
    "Amrfinderplus": {
        "project_field": "amrfinder_project",
        "accession_field": "Amrfinder_Accession",
        "project_accession_field": "WGS_Amrfinder_Acc",
        "summary_field": "WGS_AmrfinderSummary",
        "sample_fields": ("name",),
    },
}


def normalize_wgs_sample_identity(value):
    text = upload_text_or_none(value)
    if not text:
        return ""
    base = os.path.basename(text)
    return os.path.splitext(base)[0].strip()


def build_wgs_sample_identity_filter(link_config, sample_name):
    sample_name = (sample_name or "").strip()
    normalized = normalize_wgs_sample_identity(sample_name)
    if not sample_name:
        return Q(pk__isnull=True)

    identity_filter = Q(**{link_config["accession_field"]: sample_name})
    if normalized and normalized != sample_name:
        identity_filter |= Q(**{link_config["accession_field"]: normalized})

    for field in link_config["sample_fields"]:
        identity_filter |= Q(**{field: sample_name})
        identity_filter |= Q(**{f"{field}__icontains": sample_name})
        if normalized and normalized != sample_name:
            identity_filter |= Q(**{field: normalized})
            identity_filter |= Q(**{f"{field}__icontains": normalized})

    return identity_filter


def sync_wgs_records_from_sampleinfo(sampleinfo):
    sample_name = (sampleinfo.sample_name or "").strip()
    sample_accession = (sampleinfo.sample_accession or "").strip()
    if not sample_name or not sample_accession:
        return 0

    referred_obj = (
        Final_Data.objects.filter(f_AccessionNo=sample_accession).first()
        if sample_accession else None
    )
    project = get_or_create_wgs_project_for_upload(
        sample_accession,
        referred_obj,
        "WGS_SampleInfo_Acc",
        "WGS_SampleInfoSummary",
        existing_project=sampleinfo.sample_project,
    )
    if sampleinfo.sample_project_id != project.pk:
        sampleinfo.sample_project = project
        sampleinfo.save(update_fields=["sample_project"])

    linked_count = 0
    for model_name, link_config in WGS_SAMPLEINFO_LINK_FIELDS.items():
        model = globals().get(model_name)
        if not model:
            continue

        records = list(model.objects.filter(
            build_wgs_sample_identity_filter(link_config, sample_name)
        ))
        for record in records:
            update_fields = []
            if getattr(record, link_config["accession_field"]) != sample_accession:
                setattr(record, link_config["accession_field"], sample_accession)
                update_fields.append(link_config["accession_field"])

            if getattr(record, f"{link_config['project_field']}_id") != project.pk:
                setattr(record, link_config["project_field"], project)
                update_fields.append(link_config["project_field"])

            if update_fields:
                record.save(update_fields=update_fields)
                linked_count += 1

        setattr(project, link_config["project_accession_field"], sample_accession)
        setattr(project, link_config["summary_field"], bool(records))
        project.save(update_fields=[
            link_config["project_accession_field"],
            link_config["summary_field"],
        ])

    return linked_count


def sync_wgs_record_from_sampleinfo(record):
    link_config = WGS_SAMPLEINFO_LINK_FIELDS.get(record.__class__.__name__)
    if not link_config:
        return False

    identities = [
        getattr(record, link_config["accession_field"], None),
        *[getattr(record, field, None) for field in link_config["sample_fields"]],
    ]
    candidates = {
        identity
        for raw_identity in identities
        for identity in (
            (str(raw_identity).strip() if raw_identity is not None else ""),
            normalize_wgs_sample_identity(raw_identity),
        )
        if identity
    }
    if not candidates:
        return False

    sampleinfo = (
        SampleInformation.objects
        .filter(sample_name__in=candidates)
        .exclude(sample_accession__isnull=True)
        .exclude(sample_accession="")
        .select_related("sample_project")
        .order_by("-Date_uploaded_si", "-pk")
        .first()
    )
    if not sampleinfo:
        return False

    sync_wgs_records_from_sampleinfo(sampleinfo)
    return True


def save_or_update_wgs_module_record(existing_record, defaults, overwrite=True):
    model = defaults.pop("_model")
    if existing_record:
        if not overwrite:
            return existing_record, "skipped"
        for field, value in defaults.items():
            setattr(existing_record, field, value)
        existing_record.save(update_fields=list(defaults.keys()))
        sync_wgs_record_from_sampleinfo(existing_record)
        return existing_record, "updated"
    record = model.objects.create(**defaults)
    sync_wgs_record_from_sampleinfo(record)
    return record, "created"


def relink_all_wgs_records_from_sampleinfo():
    linked_count = 0
    for sampleinfo in SampleInformation.objects.exclude(sample_name__isnull=True).exclude(sample_name=""):
        linked_count += sync_wgs_records_from_sampleinfo(sampleinfo)
    return linked_count


def wgs_upload_overwrite_enabled(request):
    return request.POST.get("overwrite", "false").lower() == "true"


def upload_text_or_none(value):
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return text if text and text.lower() != "nan" else None


def upload_number_or_none(value, integer=False):
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, str):
        text = value.strip()
        if not text or text.lower() == "nan" or ";" in text:
            return None
        text = text.replace(",", "")
    else:
        text = value
    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    return int(number) if integer else number


def upload_bool_or_false(value):
    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "checked", "x"}


def has_tested_antibiotic_value(*values):
    for value in values:
        if value is None:
            continue
        try:
            if pd.isna(value):
                continue
        except (TypeError, ValueError):
            pass
        text = str(value).strip()
        if text and text.lower() not in {"nan", "none", "n/a", "na", "-"}:
            return True
    return False
    

# handles the connection of WGS project to referred data
@login_required
@role_required(*UPLOAD_CENTER_ALLOWED_ROLES)
def upload_wgs_view(request):

    if request.method == "POST":
        access = role_flags(request.user)
        form = WGSProjectForm(request.POST)
        sampleinfo_form = SampleInfoUploadForm(request.POST, request.FILES)
        bactscout_form = BactScoutUploadForm(request.POST, request.FILES)
        gtdbtk_form = GtdbTkUploadForm(request.POST, request.FILES)
        gambit_form = GambitUploadForm(request.POST, request.FILES)
        mlst_form = MlstUploadForm(request.POST, request.FILES)
        checkm2_form = Checkm2UploadForm(request.POST, request.FILES)
        assembly_form = AssemblyUploadForm(request.POST, request.FILES)
        amrfinder_form = AmrUploadForm(request.POST, request.FILES)
        demogs_form = DemogsDataUploadForm(request.POST, request.FILES)
        antibiotic_form = FinalAntibioticUploadForm(request.POST, request.FILES)
        raw_antibiotic_form = RawAntibioticUploadForm(request.POST, request.FILES)

        final_data_uploaded = False
        final_antibiotic_uploaded = False
        raw_antibiotic_uploaded = False
        project_saved = False
        sampleinfo_uploaded = False
        bactscout_uploaded = False
        gtdbtk_uploaded = False
        gambit_uploaded = False
        mlst_uploaded = False
        checkm2_uploaded = False
        assembly_uploaded = False
        amrfinder_uploaded = False
        
        
        referred_upload_fields = {"DemogsDataFile", "FinalAntibioticFile", "RawAntibioticFile"}
        wgs_upload_fields = {
            "sampleinfo",
            "bactscoutfile",
            "GtdbTkFile",
            "GambitFile",
            "Mlstfile",
            "Checkm2file",
            "Assemblyfile",
            "Amrfinderfile",
        }

        if not access["can_create"] and referred_upload_fields.intersection(request.FILES):
            messages.error(request, "You do not have permission to upload raw or final referred data.")
            return redirect("upload_wgs_view")

        if not access["can_wgs_create"] and wgs_upload_fields.intersection(request.FILES):
            messages.error(request, "You do not have permission to upload WGS data.")
            return redirect("upload_wgs_view")

        # Final Data upload
        if access["can_create"] and demogs_form.is_valid():
            demogs_form.save()
            final_data_uploaded = True

        # Final antibiotic upload
        if access["can_create"] and antibiotic_form.is_valid():
            antibiotic_form.save()
            final_antibiotic_uploaded = True

        # Raw antibiotic upload
        if access["can_create"] and raw_antibiotic_form.is_valid():
            raw_antibiotic_form.save()
            raw_antibiotic_uploaded = True


        # WGS Project
        if access["can_wgs_create"] and form.is_valid():
            form.save()
            project_saved = True

        # SampleInfo Upload
        if access["can_wgs_create"] and sampleinfo_form.is_valid():
            sampleinfo_form.save()
            sampleinfo_uploaded = True

        # Bacscout Upload
        if access["can_wgs_create"] and bactscout_form.is_valid():
            bactscout_form.save()
            bactscout_uploaded = True


        # Bacscout Upload
        if access["can_wgs_create"] and gtdbtk_form.is_valid():
            gtdbtk_form.save()
            gtdbtk_uploaded = True

        # Gambit Upload
        if access["can_wgs_create"] and gambit_form.is_valid():
            gambit_form.save()
            gambit_uploaded = True
        
        # Mlst Upload
        if access["can_wgs_create"] and mlst_form.is_valid():
            mlst_form.save()
            mlst_uploaded = True
        
        # Checkm2 Upload
        if access["can_wgs_create"] and checkm2_form.is_valid():
            checkm2_form.save()
            checkm2_uploaded = True
        
        # Assembly scan Upload
        if access["can_wgs_create"] and assembly_form.is_valid():
            assembly_form.save()
            assembly_uploaded = True

        
        # Amrfinder Upload
        if access["can_wgs_create"] and amrfinder_form.is_valid():
            amrfinder_form.save()
            amrfinder_uploaded = True

        # If any form worked, refresh
        if project_saved or final_data_uploaded or final_antibiotic_uploaded or raw_antibiotic_uploaded or sampleinfo_uploaded or bactscout_uploaded or gtdbtk_uploaded or gambit_uploaded or mlst_uploaded or checkm2_uploaded or assembly_uploaded or amrfinder_uploaded:
            return redirect("upload_wgs_view")

    else:
        form = WGSProjectForm()
        demogs_form = DemogsDataUploadForm()
        antibiotic_form = FinalAntibioticUploadForm()
        raw_antibiotic_form = RawAntibioticUploadForm()
        sampleinfo_form = SampleInfoUploadForm()
        bactscout_form = BactScoutUploadForm()
        gtdbtk_form = GtdbTkUploadForm()
        gambit_form = GambitUploadForm()
        mlst_form = MlstUploadForm()
        checkm2_form = Checkm2UploadForm()
        assembly_form = AssemblyUploadForm()
        amrfinder_form = AmrUploadForm()

    return render(
        request,
        "wgs_app/Add_wgs.html",
        {
            "form": form,
            "demogs_form": demogs_form,
            "antibiotic_form": antibiotic_form,
            "raw_antibiotic_form": raw_antibiotic_form,
            "sampleinfo_form": sampleinfo_form,
            "bactscout_form": bactscout_form,
            "gtdbtk_form": gtdbtk_form,
            "gambit_form": gambit_form,
            "mlst_form": mlst_form,
            "checkm2_form": checkm2_form,
            "assembly_form": assembly_form,
            "amrfinder_form": amrfinder_form,
            "editing": False,
        },
    )


@login_required
def show_wgs_projects(request):
    # Get all Referred_Data that have associated WGS projects
    referred_with_wgs = Final_Data.objects.filter(
        f_AccessionNo__isnull=False
    ).distinct()
    
    context = {
        'referred_list': referred_with_wgs,
    }
    return render(request, 'wgs_app/view_match.html', context)


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_wgs(request, pk):
    wgs_item = get_object_or_404(WGS_Project, pk=pk)

    if request.method == "POST":
        wgs_item.delete()
        messages.success(request, f"Record {wgs_item.Ref_Accession} deleted successfully!")
        return redirect('show_wgs_projects')  # <-- Correct URL name

    messages.error(request, "Invalid request for deletion.")
    return redirect('show_wgs_projects')  # <-- Correct URL name




#########   Gambit
@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
def upload_gambit(request):
    form = WGSProjectForm()
    gambit_form = GambitUploadForm()
    editing = False  

    if request.method == "POST" and request.FILES.get("GambitFile"):
        overwrite = wgs_upload_overwrite_enabled(request)
        gambit_form = GambitUploadForm(request.POST, request.FILES)
        if gambit_form.is_valid():
            try:
                upload = gambit_form.save()
                df = read_uploaded_file(upload.GambitFile, sheet_name="gambit")
                df.columns = df.columns.str.strip().str.replace(".", "_", regex=False)
            except Exception as e:
                messages.error(request, f"Error processing FASTQ file: {e}")
                return render(request, "wgs_app/Add_wgs.html", {
                    "form": form,
                    "gambit_form": gambit_form,
                    "mlst_form": MlstUploadForm(),
                    "checkm2_form": Checkm2UploadForm(),
                    "amrfinder_form": AmrUploadForm(),
                    "assembly_form": AssemblyUploadForm(),
                    "demogs_form": DemogsDataUploadForm(),
                    "antibiotic_form": FinalAntibioticUploadForm(),
                    "sampleinfo_form" : SampleInfoUploadForm(),
                    "bactscout_form": BactScoutUploadForm(),
                    "gtdbtk_form": GtdbTkUploadForm(),
                    "raw_antibiotic_form": RawAntibioticUploadForm(),
                    "editing": editing,
                })

            site_codes = set(SiteData.objects.values_list("SiteCode", flat=True))
                # helper to build accession
            def format_gambit_accession(raw_name: str, site_codes: set) -> str:
                """
                Returns formatted accession only if BOTH 'ARS' and a valid SiteCode from SiteData exist in the name.
                """
                if not raw_name:
                    return ""

                name = raw_name.strip().upper() # normalize case

                # Reject invalid patterns
                if "UTPR" in name or "UTPN" in name or "BL" in name:
                    return ""
                
                # ✅ Must contain 'ARS' - if not, return empty immediately
                if "ARS" not in name:
                    return ""

                # ✅ Find if any valid SiteCode from DB exists in the sample name
                # Use word boundaries to match complete site codes only
                valid_code = None
                for code in site_codes:
                    code_upper = code.upper()
                    # Look for the site code with word boundaries (hyphens, start/end of string)
                    # Pattern: site code must be followed by a hyphen and digits
                    pattern = rf"[-]?{re.escape(code_upper)}[-]?\d+"
                    if re.search(pattern, name):
                        valid_code = code_upper
                        break

                # No valid site code found → blank
                if not valid_code:
                    return ""

                # ✅ Extract prefix that includes ARS (e.g., "18ARS")
                prefix_match = re.search(r"(\d*ARS)", name)
                prefix = prefix_match.group(1) if prefix_match else "ARS"

                # ✅ Extract numeric digits after the site code (e.g., 0055)
                num_match = re.search(rf"{re.escape(valid_code)}[-]?(\d+)", name)
                digits = num_match.group(1) if num_match else ""

                return f"{prefix}_{valid_code}{digits}" if digits else ""


            for _, row in df.iterrows():
                sample_name = str(row.get("sample", "")).strip()
                gambit_accession = format_gambit_accession(sample_name, site_codes)

                # if invalid accession keep blank
                if not gambit_accession: 
                    gambit_accession = ""

                existing_gambit = find_existing_wgs_module_record(
                    Gambit,
                    "Gambit_Accession",
                    gambit_accession,
                    "sample",
                    sample_name,
                )
                if existing_gambit and not overwrite:
                    continue

                # try to find Referred_Data with this accession
                referred_obj = Final_Data.objects.filter(
                    f_AccessionNo=gambit_accession
                ).first()

                connect_project = get_or_create_wgs_project_for_upload(
                    gambit_accession,
                    referred_obj,
                    "WGS_Gambit_Acc",
                    "WGS_GambitSummary",
                    existing_project=existing_gambit.gambit_project if existing_gambit else None,
                )
                save_or_update_wgs_module_record(
                    existing_gambit,
                    {
                        "_model": Gambit,
                        "Gambit_Accession": gambit_accession,
                        "gambit_project": connect_project,
                        "sample": sample_name,
                        "predicted_name": row.get("predicted_name", ""),
                        "predicted_rank": row.get("predicted_rank", ""),
                        "predicted_ncbi_id": row.get("predicted_ncbi_id", ""),
                        "predicted_threshold": row.get("predicted_threshold", ""),
                        "closest_distance": row.get("closest_distance", ""),
                        "closest_description": row.get("closest_description", ""),
                        "next_name": row.get("next_name", ""),
                        "next_rank": row.get("next_rank", ""),
                        "next_ncbi_id": row.get("next_ncbi_id", ""),
                        "next_threshold": row.get("next_threshold", ""),
                    },
                    overwrite=overwrite,
                )


            messages.success(request, "Gambit records updated successfully.")
            return redirect("show_gambit")

    return render(request, "wgs_app/Add_wgs.html", {
        "form": form,
        "gambit_form": gambit_form,
        "mlst_form": MlstUploadForm(),
        "checkm2_form": Checkm2UploadForm(),
        "assembly_form": AssemblyUploadForm(),
        "amrfinder_form": AmrUploadForm(),
        "demogs_form": DemogsDataUploadForm(),
        "antibiotic_form": FinalAntibioticUploadForm(),
        "sampleinfo_form" : SampleInfoUploadForm(),
        "bactscout_form": BactScoutUploadForm(),
        "gtdbtk_form": GtdbTkUploadForm(),
        "raw_antibiotic_form": RawAntibioticUploadForm(),
        "editing": editing,
    })




@login_required
def show_gambit(request):
    gambit_summaries = Gambit.objects.all().order_by('-Date_uploaded_g')
    upload_dates = (
        Gambit.objects.exclude(Date_uploaded_g__isnull=True)
        .values_list('Date_uploaded_g', flat=True)
        .distinct()
        .order_by('-Date_uploaded_g')
    )

    total_records = Gambit.objects.count()
     # Paginate the queryset to display 20 records per page
    paginator = Paginator(gambit_summaries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "wgs_app/show_gambit.html", {
        "page_obj": page_obj,
        "upload_dates": upload_dates,
        "total_records": total_records,
    })




@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_gambit(request, pk):
    gambit_item = get_object_or_404(Gambit, pk=pk)

    if request.method == "POST":
        # Before deleting, clear related field in WGS_Project
        WGS_Project.objects.filter(WGS_Gambit_Acc=gambit_item.Gambit_Accession).update(
            WGS_Gambit_Acc="",
            WGS_GambitSummary=False
        )

        gambit_item.delete()
        messages.success(request, f"Record {gambit_item.sample} deleted successfully!")
        return redirect('show_gambit')

    messages.error(request, "Invalid request for deletion.")
    return redirect('show_gambit')


# @login_required
# def delete_all_gambit(request):
#     Gambit.objects.all().delete()
#     messages.success(request, "Gambit Records have been deleted successfully.")
#     return redirect('show_gambit')  # Redirect to the table view



@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_all_gambit(request):
    """
    Safely delete all Gambit records but preserve WGS_Project links
    for other WGS data types (BactScout, MLST, CheckM2, Assembly, AMRFinder, etc.).
    """
    # Step 1: Clear only Gambit fields in existing WGS_Project records
    updated_count = WGS_Project.objects.filter(
        WGS_Gambit_Acc__isnull=False
    ).exclude(WGS_Gambit_Acc="").update(
        WGS_Gambit_Acc="",
        WGS_GambitSummary=False
    )

    # Step 2: Delete all Gambit summary data
    Gambit.objects.all().delete()

    # Step 3: Display success message
    messages.success(
        request,
        f"All Gambit records deleted successfully, and {updated_count} WGS Project(s) were unlinked from Gambit data."
    )

    return redirect("show_gambit")




@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_gambit_by_date(request):
    if request.method == "POST":
        upload_date_str = request.POST.get("upload_date")
        print("🕒 Received upload_date_str:", upload_date_str)

        if not upload_date_str:
            messages.error(request, "Please select an upload date to delete.")
            return redirect("show_gambit")

        # Use Django’s date parser
        upload_date = parse_date(upload_date_str)

        if not upload_date:
            messages.error(request, f"Invalid date format: {upload_date_str}")
            return redirect("show_gambit")

        deleted_count, _ = Gambit.objects.filter(Date_uploaded_g=upload_date).delete()
        messages.success(request, f"✅ Deleted {deleted_count} Gambit records uploaded on {upload_date}.")
        return redirect("show_gambit")

    messages.error(request, "Invalid request method.")
    return redirect("show_gambit")


#########   MLST
@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
def upload_mlst(request):
    form = WGSProjectForm()
    mlst_form = MlstUploadForm()
    editing = False  

    if request.method == "POST" and request.FILES.get("Mlstfile"):
        overwrite = wgs_upload_overwrite_enabled(request)
        mlst_form = MlstUploadForm(request.POST, request.FILES)
        try:
            upload = mlst_form.save()
            df = read_uploaded_file(upload.Mlstfile, sheet_name=["mlst", "mlst_new"])
            df.columns = df.columns.str.strip().str.replace(".", "", regex=False)
        except Exception as e:
            messages.error(request, f"Error processing MLST file: {e}")
            return render(request, "wgs_app/Add_wgs.html", {
                "form": form,
                "gambit_form": GambitUploadForm(),
                "mlst_form": mlst_form,
                "checkm2_form": Checkm2UploadForm(),
                "amrfinder_form": AmrUploadForm(),
                "assembly_form": AssemblyUploadForm(),
                "demogs_form": DemogsDataUploadForm(),
                "antibiotic_form": FinalAntibioticUploadForm(),
                "sampleinfo_form" : SampleInfoUploadForm(),
                "bactscout_form": BactScoutUploadForm(),
                "gtdbtk_form": GtdbTkUploadForm(),
                "raw_antibiotic_form": RawAntibioticUploadForm(),
                "editing": editing,
            })

        # ✅ Load all valid site codes from the SiteData table
        site_codes = set(SiteData.objects.values_list("SiteCode", flat=True))

        def to_bool(value):
            if isinstance(value, bool):
                return value
            if value is None or pd.isna(value):
                return False
            return str(value).strip().lower() in {"1", "true", "yes", "y", "checked", "x"}

        # === Helper: build accession from file name ===
        def format_mlst_accession(raw_name: str, site_codes: set) -> str:
            if not raw_name:
                return ""

            base_noext = os.path.splitext(os.path.basename(raw_name))[0].strip()

            # Must contain 'ARS' to be valid
            if "ARS" not in base_noext:
                return ""

            parts = re.split(r"[-_]", base_noext)
            if not parts:
                return ""

            prefix = parts[0]

            # Look for SITE#### pattern where SITE is valid
            for part in parts[1:]:
                match = re.match(r"^([A-Za-z]{2,6})(\d+)$", part)
                if match:
                    letters, digits = match.group(1).upper(), match.group(2)
                    if letters in site_codes:
                        return f"{prefix}_{letters}{digits}"

            # 2Look for a separate valid site code, then grab digits from next part
            for i in range(1, len(parts)):
                part = parts[i]
                if part.upper() in site_codes:
                    letters = part.upper()
                    digits = ""

                    if i + 1 < len(parts):
                        next_part = parts[i + 1]
                        next_match = re.match(r"^([A-Za-z]{2,6})(\d+)$", next_part)
                        if next_match:
                            digits = next_match.group(2)
                        else:
                            digit_match = re.search(r"(\d+)", next_part)
                            if digit_match:
                                digits = digit_match.group(1)

                    # fallback — digits inside current part
                    if not digits:
                        digit_match2 = re.search(r"(\d+)", part)
                        if digit_match2:
                            digits = digit_match2.group(1)

                    return f"{prefix}_{letters}{digits}" if digits else f"{prefix}_{letters}"

            return ""

        print("✅ Total rows in DataFrame:", len(df))

        # === Loop through rows ===
        for _, row in df.iterrows():
            full_path = str(row.get("name", "")).strip()
            mlst_accession = format_mlst_accession(full_path, site_codes)
            scheme = row.get("scheme", "")

            existing_mlst = find_existing_wgs_module_record_by_fields(
                Mlst,
                {
                    "Mlst_Accession": mlst_accession,
                    "name": full_path,
                    "scheme": scheme,
                },
            )
            if existing_mlst and not overwrite:
                continue

            # Find Referred_Data (optional)
            referred_obj = (
                Final_Data.objects.filter(f_AccessionNo=mlst_accession).first()
                if mlst_accession else None
            )

            connect_project = get_or_create_wgs_project_for_upload(
                mlst_accession,
                referred_obj,
                "WGS_Mlst_Acc",
                "WGS_MlstSummary",
                existing_project=existing_mlst.mlst_project if existing_mlst else None,
            )

            save_or_update_wgs_module_record(
                existing_mlst,
                {
                    "_model": Mlst,
                    "Mlst_Accession": mlst_accession,
                    "mlst_project": connect_project,
                    "name": full_path,
                    "scheme": scheme,
                    "mlst": row.get("MLST", ""),
                    "allele1": row.get("allele1", ""),
                    "allele2": row.get("allele2", ""),
                    "allele3": row.get("allele3", ""),
                    "allele4": row.get("allele4", ""),
                    "allele5": row.get("allele5", ""),
                    "allele6": row.get("allele6", ""),
                    "allele7": row.get("allele7", ""),
                },
                overwrite=overwrite,
            )

        messages.success(request, "MLST records updated successfully.")
        return redirect("show_mlst")

    # === GET request fallback ===
    return render(request, "wgs_app/Add_wgs.html", {
        "form": form,
        "gambit_form": GambitUploadForm(),
        "mlst_form": mlst_form,
        "checkm2_form": Checkm2UploadForm(),
        "amrfinder_form": AmrUploadForm(),
        "assembly_form": AssemblyUploadForm(),
        "demogs_form": DemogsDataUploadForm(),
        "antibiotic_form": FinalAntibioticUploadForm(),
        "sampleinfo_form" : SampleInfoUploadForm(),
        "bactscout_form": BactScoutUploadForm(),
        "gtdbtk_form": GtdbTkUploadForm(),
        "raw_antibiotic_form": RawAntibioticUploadForm(),
        "editing": editing
    })



@login_required
def show_mlst(request):
    mlst_summaries = Mlst.objects.all().order_by('-Date_uploaded_m')
    upload_dates = (
        Mlst.objects.exclude(Date_uploaded_m__isnull=True)
        .values_list('Date_uploaded_m', flat=True)
        .distinct()
        .order_by('-Date_uploaded_m')
    )

    total_records = Mlst.objects.count()
     # Paginate the queryset to display 20 records per page
    paginator = Paginator(mlst_summaries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "wgs_app/show_mlst.html", {
        "page_obj": page_obj,
        "upload_dates": upload_dates,
        "total_records": total_records,
    })




# @login_required
# def delete_mlst(request, pk):
#     mlst_item = get_object_or_404(Mlst, pk=pk)

#     if request.method == "POST":
#         mlst_item.delete()
#         messages.success(request, f"Record {mlst_item.sample} deleted successfully!")
#         return redirect('show_mlst')  # <-- Correct URL name

#     messages.error(request, "Invalid request for deletion.")
#     return redirect('show_mlst')  # <-- Correct URL name


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_mlst(request, pk):
    mlst_item = get_object_or_404(Mlst, pk=pk)

    if request.method == "POST":
        # Before deleting, clear related field in WGS_Project
        WGS_Project.objects.filter(WGS_Mlst_Acc=mlst_item.Mlst_Accession).update(
            WGS_Mlst_Acc="",
            WGS_MlstSummary=False
        )

        mlst_item.delete()
        messages.success(request, f"Record {mlst_item.sample} deleted successfully!")
        return redirect('show_mlst')

    messages.error(request, "Invalid request for deletion.")
    return redirect('show_mlst')



# @login_required
# def delete_all_mlst(request):
#     Mlst.objects.all().delete()
#     messages.success(request, "Mlst Records have been deleted successfully.")
#     return redirect('show_mlst')  # Redirect to the table view


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_all_mlst(request):
    """
    Safely delete all MLST records but preserve WGS_Project links
    for other WGS data types (BactScout, CheckM2, Assembly, Gambit, AMRFinder, etc.).
    """
    # Step 1: Clear only MLST fields in existing WGS_Project records
    updated_count = WGS_Project.objects.filter(
        WGS_Mlst_Acc__isnull=False
    ).exclude(WGS_Mlst_Acc="").update(
        WGS_Mlst_Acc="",
        WGS_MlstSummary=False
    )

    # Step 2: Delete all MLST summary data
    Mlst.objects.all().delete()

    # Step 3: Display success message
    messages.success(
        request,
        f"All MLST records deleted successfully, and {updated_count} WGS Project(s) were unlinked from MLST data."
    )

    return redirect("show_mlst")




@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_mlst_by_date(request):
    if request.method == "POST":
        upload_date_str = request.POST.get("upload_date")
        print("🕒 Received upload_date_str:", upload_date_str)

        if not upload_date_str:
            messages.error(request, "Please select an upload date to delete.")
            return redirect("show_mlst")

        # Use Django’s date parser
        upload_date = parse_date(upload_date_str)

        if not upload_date:
            messages.error(request, f"Invalid date format: {upload_date_str}")
            return redirect("show_mlst")

        deleted_count, _ = Mlst.objects.filter(Date_uploaded_m=upload_date).delete()
        messages.success(request, f"✅ Deleted {deleted_count} Mlst records uploaded on {upload_date}.")
        return redirect("show_mlst")

    messages.error(request, "Invalid request method.")
    return redirect("show_mlst")




###################  Checkm2 
@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
def upload_checkm2(request):
    form = WGSProjectForm()
    checkm2_form = Checkm2UploadForm()
    editing = False

    if request.method == "POST" and request.FILES.get("Checkm2file"):
        overwrite = wgs_upload_overwrite_enabled(request)
        checkm2_form = Checkm2UploadForm(request.POST, request.FILES)
        try:
            upload = checkm2_form.save()
            df = read_uploaded_file(upload.Checkm2file, sheet_name="checkm2")
            df.columns = df.columns.str.strip().str.replace(".", "", regex=False)
        except Exception as e:
            messages.error(request, f"Error processing MLST file: {e}")
            return render(request, "wgs_app/Add_wgs.html", {
                "form": form,
                "gambit_form": GambitUploadForm(),
                "mlst_form": MlstUploadForm(),
                "checkm2_form": checkm2_form,
                "amrfinder_form": AmrUploadForm(),
                "assembly_form": AssemblyUploadForm(),
                "demogs_form": DemogsDataUploadForm(),
                "antibiotic_form": FinalAntibioticUploadForm(),
                "sampleinfo_form" : SampleInfoUploadForm(),
                "bactscout_form": BactScoutUploadForm(),
                "gtdbtk_form": GtdbTkUploadForm(),
                "raw_antibiotic_form": RawAntibioticUploadForm(),
                "editing": editing,
            })

        site_codes = set(SiteData.objects.values_list("SiteCode", flat=True))

        # Helper to build accession
        def format_checkm2_accession(raw_name: str) -> str:
            if not raw_name:
                return ""
            # Take basename and remove extension
            base = os.path.basename(raw_name)
            base_noext = os.path.splitext(base)[0].strip()

            if "ARS" not in base_noext:
                return ""

            parts = re.split(r"[-_]", base_noext)
            if not parts:
                return ""

            prefix = parts[0]  # e.g. "18ARS"

            # Look for a part that matches sitecode+digits (e.g. BGH0055, CVM0162)
            for part in parts[1:]:
                m = re.match(r"^([A-Za-z]{2,6})(\d+)", part)
                if m:
                    letters = m.group(1).upper()
                    digits = m.group(2)
                    if letters in site_codes:
                        return f"{prefix}_{letters}{digits}"

            # If sitecode and digits are separated (rare case)
            for i in range(1, len(parts) - 1):
                if parts[i].upper() in site_codes:
                    letters = parts[i].upper()
                    digits_match = re.search(r"(\d+)", parts[i + 1])
                    if digits_match:
                        return f"{prefix}_{letters}{digits_match.group(1)}"
                    return f"{prefix}_{letters}"

            return ""

        print("Total rows in dataframe:", len(df))

        for _, row in df.iterrows():
            sample_name = str(row.get("Name", "")).strip().replace(".fna", "")
            checkm2_accession = format_checkm2_accession(sample_name)

            existing_checkm2 = find_existing_wgs_module_record(
                Checkm2,
                "Checkm2_Accession",
                checkm2_accession,
                "Name",
                sample_name,
            )
            if existing_checkm2 and not overwrite:
                continue

            # Step 1: Try to find Referred_Data with this accession (only if non-blank)
            referred_obj = (
                Final_Data.objects.filter(f_AccessionNo=checkm2_accession).first()
                if checkm2_accession else None
            )

            connect_project = get_or_create_wgs_project_for_upload(
                checkm2_accession,
                referred_obj,
                "WGS_Checkm2_Acc",
                "WGS_Checkm2Summary",
                existing_project=existing_checkm2.checkm2_project if existing_checkm2 else None,
            )

            save_or_update_wgs_module_record(
                existing_checkm2,
                {
                    "_model": Checkm2,
                    "Checkm2_Accession": checkm2_accession,
                    "Name": sample_name,
                    "checkm2_project": connect_project,
                    "Completeness": row.get("Completeness", ""),
                    "Contamination": row.get("Contamination", ""),
                    "Completeness_Model_Used": row.get("Completeness_Model_Used", ""),
                    "Translation_Table_Used": row.get("Translation_Table_Used", ""),
                    "Coding_Density": row.get("Coding_Density", ""),
                    "Contig_N50": row.get("Contig_N50", ""),
                    "Average_Gene_Length": row.get("Average_Gene_Length", ""),
                    "Genome_Size": row.get("Genome_Size", ""),
                    "GC_Content": row.get("GC_Content", ""),
                    "Total_Coding_Sequences": row.get("Total_Coding_Sequences", ""),
                    "Total_Contigs": row.get("Total_Contigs", ""),
                    "Max_Contig_Length": row.get("Max_Contig_Length", ""),
                    "Additional_Notes": row.get("Additional_Notes", ""),
                },
                overwrite=overwrite,
            )

        messages.success(request, "Checkm2 records uploaded successfully.")
        return redirect("show_checkm2")

    return render(request, "wgs_app/Add_wgs.html", {
        "form": form,
        "gambit_form": GambitUploadForm(),
        "mlst_form": MlstUploadForm(),
        "checkm2_form": checkm2_form,
        "assembly_form": AssemblyUploadForm(),
        "amrfinder_form": AmrUploadForm(),
        "demogs_form": DemogsDataUploadForm(),
        "antibiotic_form": FinalAntibioticUploadForm(),
        "sampleinfo_form" : SampleInfoUploadForm(),
        "bactscout_form": BactScoutUploadForm(),
        "gtdbtk_form": GtdbTkUploadForm(),
        "raw_antibiotic_form": RawAntibioticUploadForm(),
        "editing": editing,
    })



@login_required
def show_checkm2(request):
    checkm2_summaries = Checkm2.objects.all().order_by('-Date_uploaded_c')
    upload_dates = (
        Checkm2.objects.exclude(Date_uploaded_c__isnull=True)
        .values_list('Date_uploaded_c', flat=True)
        .distinct()
        .order_by('-Date_uploaded_c')
    )

    total_records = Checkm2.objects.count()
     # Paginate the queryset to display 20 records per page
    paginator = Paginator(checkm2_summaries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "wgs_app/show_checkm2.html", {
        "page_obj": page_obj,
        "upload_dates": upload_dates,
        "total_records": total_records,
    })




# @login_required
# def delete_checkm2(request, pk):
#     checkm2_item = get_object_or_404(Checkm2, pk=pk)

#     if request.method == "POST":
#         checkm2_item.delete()
#         messages.success(request, f"Record {checkm2_item.Name} deleted successfully!")
#         return redirect('show_checkm2')  # <-- Correct URL name

#     messages.error(request, "Invalid request for deletion.")
#     return redirect('show_checkm2')  # <-- Correct URL name



@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_checkm2(request, pk):
    checkm2_item = get_object_or_404(Checkm2, pk=pk)

    if request.method == "POST":
        # Before deleting, clear related field in WGS_Project
        WGS_Project.objects.filter(WGS_Checkm2_Acc=checkm2_item.Checkm2_Accession).update(
            WGS_Checkm2_Acc="",
            WGS_Checkm2Summary=False
        )

        checkm2_item.delete()
        messages.success(request, f"Record {checkm2_item.Name} deleted successfully!")
        return redirect('show_checkm2')

    messages.error(request, "Invalid request for deletion.")
    return redirect('show_checkm2')




# @login_required
# def delete_all_checkm2(request):
#     Checkm2.objects.all().delete()
#     messages.success(request, "Checkm2 Records have been deleted successfully.")
#     return redirect('show_checkm2')  # Redirect to the table view


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_all_checkm2(request):
    """
    Safely delete all CheckM2 records but preserve WGS_Project links
    for other WGS data types (BactScout, MLST, Assembly, Gambit, AMRFinder, etc.).
    """
    # Step 1: Clear only CheckM2 fields in existing WGS_Project records
    updated_count = WGS_Project.objects.filter(
        WGS_Checkm2_Acc__isnull=False
    ).exclude(WGS_Checkm2_Acc="").update(
        WGS_Checkm2_Acc="",
        WGS_Checkm2Summary=False
    )

    # Step 2: Delete all CheckM2 summary data
    Checkm2.objects.all().delete()

    # Step 3: Display success message
    messages.success(
        request,
        f"All CheckM2 records deleted successfully, and {updated_count} WGS Project(s) were unlinked from CheckM2 data."
    )

    return redirect("show_checkm2")




@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_checkm2_by_date(request):
    if request.method == "POST":
        upload_date_str = request.POST.get("upload_date")
        print("🕒 Received upload_date_str:", upload_date_str)

        if not upload_date_str:
            messages.error(request, "Please select an upload date to delete.")
            return redirect("show_checkm2")

        # Use Django’s date parser
        upload_date = parse_date(upload_date_str)

        if not upload_date:
            messages.error(request, f"Invalid date format: {upload_date_str}")
            return redirect("show_checkm2")

        deleted_count, _ = Checkm2.objects.filter(Date_uploaded_c=upload_date).delete()
        messages.success(request, f"✅ Deleted {deleted_count} Checkm2 records uploaded on {upload_date}.")
        return redirect("show_checkm2")

    messages.error(request, "Invalid request method.")
    return redirect("show_checkm2")




###################  Assembly Scan
@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
def upload_assembly(request):
    form = WGSProjectForm()
    assembly_form = AssemblyUploadForm()
    editing = False

    if request.method == "POST" and request.FILES.get("Assemblyfile"):
        overwrite = wgs_upload_overwrite_enabled(request)
        assembly_form = AssemblyUploadForm(request.POST, request.FILES)
        try:
            upload = assembly_form.save()
            df = read_uploaded_file(upload.Assemblyfile, sheet_name="assembly-scan")
            df.columns = df.columns.str.strip().str.replace(".", "", regex=False)
        except Exception as e:
            messages.error(request, f"Error processing Assembly file: {e}")
            return render(request, "wgs_app/Add_wgs.html", {
                "form": form,
                "gambit_form": GambitUploadForm(),
                "mlst_form": MlstUploadForm(),
                "checkm2_form": Checkm2UploadForm(),
                "amrfinder_form": AmrUploadForm(),
                "assembly_form": assembly_form,
                "demogs_form": DemogsDataUploadForm(),
                "antibiotic_form": FinalAntibioticUploadForm(),
                "sampleinfo_form" : SampleInfoUploadForm(),
                "bactscout_form": BactScoutUploadForm(),
                "gtdbtk_form": GtdbTkUploadForm(),
                "raw_antibiotic_form": RawAntibioticUploadForm(),
                "editing": editing,
            })

        site_codes = set(SiteData.objects.values_list("SiteCode", flat=True))

        # Helper to build accession
        def format_assembly_accession(raw_name: str) -> str:
            if not raw_name:
                return ""
            # Take basename and remove extension
            base = os.path.basename(raw_name)
            base_noext = os.path.splitext(base)[0].strip()

            if "ARS" not in base_noext:
                return ""

            parts = re.split(r"[-_]", base_noext)
            if not parts:
                return ""

            prefix = parts[0]  # e.g. "18ARS"

            # Look for a part that matches sitecode+digits (e.g. BGH0055, CVM0162)
            for part in parts[1:]:
                m = re.match(r"^([A-Za-z]{2,6})(\d+)", part)
                if m:
                    letters = m.group(1).upper()
                    digits = m.group(2)
                    if letters in site_codes:
                        return f"{prefix}_{letters}{digits}"

            # If sitecode and digits are separated (rare case)
            for i in range(1, len(parts) - 1):
                if parts[i].upper() in site_codes:
                    letters = parts[i].upper()
                    digits_match = re.search(r"(\d+)", parts[i + 1])
                    if digits_match:
                        return f"{prefix}_{letters}{digits_match.group(1)}"
                    return f"{prefix}_{letters}"

            return ""

        for _, row in df.iterrows():
            sample_name = str(row.get("sample", "")).strip()
            assembly_accession = format_assembly_accession(sample_name)

            existing_assembly = find_existing_wgs_module_record(
                AssemblyScan,
                "Assembly_Accession",
                assembly_accession,
                "sample",
                sample_name,
            )
            if existing_assembly and not overwrite:
                continue

            # Step 1: Try to find Referred_Data with this accession (only if non-blank)
            referred_obj = (
                Final_Data.objects.filter(f_AccessionNo=assembly_accession).first()
                if assembly_accession else None
            )

            connect_project = get_or_create_wgs_project_for_upload(
                assembly_accession,
                referred_obj,
                "WGS_Assembly_Acc",
                "WGS_AssemblySummary",
                existing_project=existing_assembly.assembly_project if existing_assembly else None,
            )

            save_or_update_wgs_module_record(
                existing_assembly,
                {
                    "_model": AssemblyScan,
                    "Assembly_Accession": assembly_accession,
                    "sample": sample_name,
                    "assembly_project": connect_project,
                    "total_contig": row.get("total_contig", ""),
                    "total_contig_length": row.get("total_contig_length", ""),
                    "max_contig_length": row.get("max_contig_length", ""),
                    "mean_contig_length": row.get("mean_contig_length", ""),
                    "median_contig_length": row.get("median_contig_length", ""),
                    "min_contig_length": row.get("min_contig_length", ""),
                    "n50_contig_length": row.get("n50_contig_length", ""),
                    "l50_contig_count": row.get("l50_contig_count", ""),
                    "num_contig_non_acgtn": row.get("num_contig_non_acgtn", ""),
                    "contig_percent_a": row.get("contig_percent_a", ""),
                    "contig_percent_c": row.get("contig_percent_c", ""),
                    "contig_percent_g": row.get("contig_percent_g", ""),
                    "contig_percent_t": row.get("contig_percent_t", ""),
                    "contig_percent_n": row.get("contig_percent_n", ""),
                    "contig_non_acgtn": row.get("contig_non_acgtn", ""),
                    "contigs_greater_1m": row.get("contigs_greater_1m", ""),
                    "contigs_greater_100k": row.get("contigs_greater_100k", ""),
                    "contigs_greater_10k": row.get("contigs_greater_10k", ""),
                    "contigs_greater_1k": row.get("contigs_greater_1k", ""),
                    "percent_contigs_greater_1m": row.get("percent_contigs_greater_1m", ""),
                    "percent_contigs_greater_100k": row.get("percent_contigs_greater_100k", ""),
                    "percent_contigs_greater_10k": row.get("percent_contigs_greater_10k", ""),
                    "percent_contigs_greater_1k": row.get("percent_contigs_greater_1k", ""),
                },
                overwrite=overwrite,
            )

        messages.success(request, "AssemblyScan records uploaded successfully.")
        return redirect("show_assembly")

    return render(request, "wgs_app/Add_wgs.html", {
        "form": form,
        "gambit_form": GambitUploadForm(),
        "mlst_form": MlstUploadForm(),
        "checkm2_form": Checkm2UploadForm(),
        "amrfinder_form": AmrUploadForm(),
        "assembly_form": assembly_form,
        "demogs_form": DemogsDataUploadForm(),
        "antibiotic_form": FinalAntibioticUploadForm(),
        "sampleinfo_form" : SampleInfoUploadForm(),
        "bactscout_form": BactScoutUploadForm(),
        "gtdbtk_form": GtdbTkUploadForm(),
        "raw_antibiotic_form": RawAntibioticUploadForm(),
        "editing": editing,
    })




@login_required
def show_assembly(request):
    assembly_summaries = AssemblyScan.objects.all().order_by('-Date_uploaded_as')
    upload_dates = (
        AssemblyScan.objects.exclude(Date_uploaded_as__isnull=True)
        .values_list('Date_uploaded_as', flat=True)
        .distinct()
        .order_by('-Date_uploaded_as')
    )

    total_records = AssemblyScan.objects.count()
     # Paginate the queryset to display 20 records per page
    paginator = Paginator(assembly_summaries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "wgs_app/show_assembly.html", {
        "page_obj": page_obj,
        "upload_dates": upload_dates,
        "total_records": total_records,
    })




@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_assembly(request, pk):
    assembly_item = get_object_or_404(AssemblyScan, pk=pk)

    if request.method == "POST":
        # Before deleting, clear related field in WGS_Project
        WGS_Project.objects.filter(WGS_Assembly_Acc=assembly_item.Assembly_Accession).update(
            WGS_Assembly_Acc="",
            WGS_AssemblySummary=False
        )

        assembly_item.delete()
        messages.success(request, f"Record {assembly_item.sample} deleted successfully!")
        return redirect('show_assembly')

    messages.error(request, "Invalid request for deletion.")
    return redirect('show_assembly')




@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_all_assembly(request):
    """
    Safely delete all Assembly records but preserve WGS_Project links
    for other WGS data types (BactScout, MLST, CheckM2, Gambit, AMRFinder, etc.).
    """
    # Step 1: Clear only Assembly fields in existing WGS_Project records
    updated_count = WGS_Project.objects.filter(
        WGS_Assembly_Acc__isnull=False
    ).exclude(WGS_Assembly_Acc="").update(
        WGS_Assembly_Acc="",
        WGS_AssemblySummary=False
    )

    # Step 2: Delete all Assembly summary data
    AssemblyScan.objects.all().delete()

    # Step 3: Display success message
    messages.success(
        request,
        f"All Assembly records deleted successfully, and {updated_count} WGS Project(s) were unlinked from Assembly data."
    )

    return redirect("show_assembly")




@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_assembly_by_date(request):
    if request.method == "POST":
        upload_date_str = request.POST.get("upload_date")
        print("🕒 Received upload_date_str:", upload_date_str)

        if not upload_date_str:
            messages.error(request, "Please select an upload date to delete.")
            return redirect("show_assembly")

        # Use Django’s date parser
        upload_date = parse_date(upload_date_str)

        if not upload_date:
            messages.error(request, f"Invalid date format: {upload_date_str}")
            return redirect("show_assembly")

        deleted_count, _ = AssemblyScan.objects.filter(Date_uploaded_as=upload_date).delete()
        messages.success(request, f"✅ Deleted {deleted_count} AssemblyScan records uploaded on {upload_date}.")
        return redirect("show_assembly")

    messages.error(request, "Invalid request method.")
    return redirect("show_assembly")




###################  Amr finder
@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
def upload_amrfinder(request):
    form = WGSProjectForm()
    amrfinder_form = AmrUploadForm()
    editing = False

    if request.method == "POST" and request.FILES.get("Amrfinderfile"):
        overwrite = wgs_upload_overwrite_enabled(request)
        amrfinder_form = AmrUploadForm(request.POST, request.FILES)
        try:
            upload = amrfinder_form.save()
            df = read_uploaded_file(upload.Amrfinderfile, sheet_name="amrfinderplus")
            df.columns = df.columns.str.strip().str.replace(".", "", regex=False)
        except Exception as e:
            messages.error(request, f"Error processing MLST file: {e}")
            return render(request, "wgs_app/Add_wgs.html", {
                "form": form,
                "gambit_form": GambitUploadForm(),
                "mlst_form": MlstUploadForm(),
                "checkm2_form": Checkm2UploadForm(),
                "amrfinder_form": amrfinder_form,
                "assembly_form": AssemblyUploadForm(),
                "demogs_form": DemogsDataUploadForm(),
                "antibiotic_form": FinalAntibioticUploadForm(),
                "sampleinfo_form" : SampleInfoUploadForm(),
                "bactscout_form": BactScoutUploadForm(),
                "gtdbtk_form": GtdbTkUploadForm(),
                "raw_antibiotic_form": RawAntibioticUploadForm(),
                "editing": editing,
            })

        # Clean and standardize column names
        df.columns = (
            df.columns
            .str.strip()
            .str.replace(" ", "_", regex=False)
            .str.replace("%", "pct", regex=False)
            .str.replace(".", "_", regex=False)
            .str.lower()
        )

        # Preload all valid site codes from SiteData (uppercase)
        site_codes = set(SiteData.objects.values_list("SiteCode", flat=True))

        # Helper to build accession
        def format_amrfinder_accession(raw_name: str) -> str:
            if not raw_name:
                return ""
            base_noext = os.path.splitext(os.path.basename(raw_name))[0].strip()

            # Must contain ARS to be eligible
            if "ARS" not in base_noext:
                return ""

            parts = re.split(r"[-_]", base_noext)
            if not parts:
                return ""

            prefix = parts[0]  # e.g., "24ARS"

            # 1) Look for LETTERS + DIGITS where LETTERS is a valid site code
            for part in parts[1:]:
                m = re.match(r"^([A-Za-z]{2,6})(\d+)$", part)
                if m:
                    letters = m.group(1).upper()
                    digits = m.group(2)
                    if letters in site_codes:
                        return f"{prefix}_{letters}{digits}"

            # 2) Check if sitecode is a separate part followed by digits
            for i in range(1, len(parts)):
                part = parts[i]
                if part.upper() in site_codes:
                    letters = part.upper()
                    digits = ""
                    if i + 1 < len(parts):
                        next_part = parts[i + 1]
                        m2 = re.match(r"^([A-Za-z]{2,6})(\d+)$", next_part)
                        if m2:
                            digits = m2.group(2)
                        else:
                            dmatch = re.search(r"(\d+)", next_part)
                            if dmatch:
                                digits = dmatch.group(1)
                    if not digits:
                        dmatch2 = re.search(r"(\d+)", part)
                        if dmatch2:
                            digits = dmatch2.group(1)
                    return f"{prefix}_{letters}{digits}" if digits else f"{prefix}_{letters}"

            # 3) As a fallback, match any valid sitecode prefix
            for part in parts[1:]:
                m = re.match(r"^([A-Za-z]{2,6})(\d+)$", part)
                if m and m.group(1).upper() in site_codes:
                    return f"{prefix}_{m.group(1).upper()}{m.group(2)}"

            return ""

        print("Total rows in dataframe:", len(df))

        for _, row in df.iterrows():
            sample_name = str(row.get("name", "")).strip()
            amrfinder_accession = format_amrfinder_accession(sample_name)
            protein_id = row.get("protein_id", "")
            contig_id = row.get("contig_id", "")
            start = row.get("start", "")
            stop = row.get("stop", "")
            element_symbol = row.get("element_symbol", "")

            existing_amrfinder = find_existing_wgs_module_record_by_fields(
                Amrfinderplus,
                {
                    "Amrfinder_Accession": amrfinder_accession,
                    "name": sample_name,
                    "protein_id": protein_id,
                    "contig_id": contig_id,
                    "start": start,
                    "stop": stop,
                    "element_symbol": element_symbol,
                },
            )
            if existing_amrfinder and not overwrite:
                continue

            # Step 1: Try to find Referred_Data with this accession (only if non-blank)
            referred_obj = (
                Final_Data.objects.filter(f_AccessionNo=amrfinder_accession).first()
                if amrfinder_accession else None
            )

            connect_project = get_or_create_wgs_project_for_upload(
                amrfinder_accession,
                referred_obj,
                "WGS_Amrfinder_Acc",
                "WGS_AmrfinderSummary",
                existing_project=(
                    existing_amrfinder.amrfinder_project
                    if existing_amrfinder else None
                ),
            )

            save_or_update_wgs_module_record(
                existing_amrfinder,
                {
                    "_model": Amrfinderplus,
                    "Amrfinder_Accession": amrfinder_accession,
                    "name": sample_name,
                    "amrfinder_project": connect_project,
                    "protein_id": protein_id,
                    "contig_id": contig_id,
                    "start": start,
                    "stop": stop,
                    "strand": row.get("strand", ""),
                    "element_symbol": element_symbol,
                    "element_name": row.get("element_name", ""),
                    "scope": row.get("scope", ""),
                    "type_field": row.get("type", ""),
                    "subtype": row.get("subtype", ""),
                    "class_field": row.get("class", ""),
                    "subclass": row.get("subclass", ""),
                    "method": row.get("method", ""),
                    "target_length": row.get("target_length", ""),
                    "reference_sequence_length": row.get("reference_sequence_length", ""),
                    "percent_coverage_of_reference": row.get("pct_coverage_of_reference", ""),
                    "percent_identity_to_reference": row.get("pct_identity_to_reference", ""),
                    "alignment_length": row.get("alignment_length", ""),
                    "closest_reference_accession": row.get("closest_reference_accession", ""),
                    "closest_reference_name": row.get("closest_reference_name", ""),
                    "hmm_accession": row.get("hmm_accession", ""),
                    "hmm_description": row.get("hmm_description", ""),
                    "Date_uploaded_am": row.get("date_uploaded_am", ""),
                },
                overwrite=overwrite,
            )

        messages.success(request, "Amrfinder records uploaded successfully.")
        return redirect("show_amrfinder")

    return render(request, "wgs_app/Add_wgs.html", {
        "form": form,
        "gambit_form": GambitUploadForm(),
        "mlst_form": MlstUploadForm(),
        "checkm2_form": Checkm2UploadForm(),
        "assembly_form": AssemblyUploadForm(),
        "amrfinder_form": amrfinder_form,
        "demogs_form": DemogsDataUploadForm(),
        "antibiotic_form": FinalAntibioticUploadForm(),
        "sampleinfo_form" : SampleInfoUploadForm(),
        "bactscout_form": BactScoutUploadForm(),
        "gtdbtk_form": GtdbTkUploadForm(),
        "raw_antibiotic_form": RawAntibioticUploadForm(),
        "editing": editing,
    })




@login_required
def show_amrfinder(request):
    amrfinder_summaries = Amrfinderplus.objects.all().order_by('-Date_uploaded_am')
    upload_dates = (
        Amrfinderplus.objects.exclude(Date_uploaded_am__isnull=True)
        .values_list('Date_uploaded_am', flat=True)
        .distinct()
        .order_by('-Date_uploaded_am')
    )

    total_records = Amrfinderplus.objects.count()
     # Paginate the queryset to display 20 records per page
    paginator = Paginator(amrfinder_summaries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "wgs_app/show_amrfinder.html", {
        "page_obj": page_obj,
        "upload_dates": upload_dates,
        "total_records": total_records,
    })






@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_amrfinder(request, pk):
    amrfinder_item = get_object_or_404(Amrfinderplus, pk=pk)

    if request.method == "POST":
        # Before deleting, clear related field in WGS_Project
        WGS_Project.objects.filter(WGS_Amrfinder_Acc=amrfinder_item.Amrfinder_Accession).update(
            WGS_Amrfinder_Acc="",
            WGS_AmrfinderSummary=False
        )

        amrfinder_item.delete()
        messages.success(request, f"Record {amrfinder_item.name} deleted successfully!")
        return redirect('show_amrfinder')

    messages.error(request, "Invalid request for deletion.")
    return redirect('show_amrfinder')





@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_all_amrfinder(request):
    """
    Safely delete all AMRFinder records but preserve WGS_Project links
    for other WGS data types (BactScout, MLST, CheckM2, Assembly, Gambit, etc.).
    """
    # Step 1: Clear only AMRFinder fields in existing WGS_Project records
    updated_count = WGS_Project.objects.filter(
        WGS_Amrfinder_Acc__isnull=False
    ).exclude(WGS_Amrfinder_Acc="").update(
        WGS_Amrfinder_Acc="",
        WGS_AmrfinderSummary=False
    )

    # Step 2: Delete all AMRFinder summary data
    Amrfinderplus.objects.all().delete()

    # Step 3: Display success message
    messages.success(
        request,
        f"All AMRFinder records deleted successfully, and {updated_count} WGS Project(s) were unlinked from AMRFinder data."
    )

    return redirect("show_amrfinder")




@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_amrfinder_by_date(request):
    if request.method == "POST":
        upload_date_str = request.POST.get("upload_date")
        if not upload_date_str:
            messages.error(request, "Please select an upload date to delete.")
            return redirect("show_amrfinder")

        try:
            target_date = datetime.strptime(upload_date_str, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
            return redirect("show_amrfinder")

        start = datetime.combine(target_date, datetime.min.time())
        end = datetime.combine(target_date + timedelta(days=1), datetime.min.time())

        deleted_count, _ = Amrfinderplus.objects.filter(
            Date_uploaded_am__gte=start,
            Date_uploaded_am__lt=end
        ).delete()

        messages.success(request, f"✅ Deleted {deleted_count} records uploaded on {target_date}.")

    else:
        messages.error(request, "Invalid request method.")

    return redirect("show_amrfinder")




############## Sample Information
@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
def upload_sample_information(request):

    form = WGSProjectForm()
    sampleinfo_form = SampleInfoUploadForm()
    editing = False

    if request.method == "POST" and request.FILES.get("sampleinfo"):
        overwrite = wgs_upload_overwrite_enabled(request)

        sampleinfo_form = SampleInfoUploadForm(request.POST, request.FILES)

        try:
            upload = sampleinfo_form.save()

            df = read_uploaded_file(upload.sampleinfo, sheet_name="batch")

            df.columns = df.columns.str.strip().str.replace(".", "", regex=False)

        except Exception as e:

            messages.error(request, f"Error processing Sample Information file: {e}")

            return render(request, "wgs_app/Add_wgs.html", {
                "form": form,
                "sampleinfo_form": sampleinfo_form,
                "gambit_form": GambitUploadForm(),
                "mlst_form": MlstUploadForm(),
                "checkm2_form": Checkm2UploadForm(),
                "amrfinder_form": AmrUploadForm(),
                "assembly_form": AssemblyUploadForm(),
                "demogs_form": DemogsDataUploadForm(),
                "antibiotic_form": FinalAntibioticUploadForm(),
                "bactscout_form": BactScoutUploadForm(),
                "gtdbtk_form": GtdbTkUploadForm(),
                "raw_antibiotic_form": RawAntibioticUploadForm(),
                "editing": editing,
            })

        site_codes = set(SiteData.objects.values_list("SiteCode", flat=True))

        # Same extraction logic used in Assembly
        def format_sample_accession(raw_name: str) -> str:

            if not raw_name:
                return ""

            base = os.path.basename(raw_name)

            base_noext = os.path.splitext(base)[0].strip()

            if "ARS" not in base_noext:
                return ""

            parts = re.split(r"[-_]", base_noext)

            if not parts:
                return ""

            prefix = parts[0]

            for part in parts[1:]:

                m = re.match(r"^([A-Za-z]{2,6})(\d+)", part)

                if m:

                    letters = m.group(1).upper()

                    digits = m.group(2)

                    if letters in site_codes:
                        return f"{prefix}_{letters}{digits}"

            for i in range(1, len(parts) - 1):

                if parts[i].upper() in site_codes:

                    letters = parts[i].upper()

                    digits_match = re.search(r"(\d+)", parts[i + 1])

                    if digits_match:
                        return f"{prefix}_{letters}{digits_match.group(1)}"

                    return f"{prefix}_{letters}"

            return ""

        # Process rows
        for _, row in df.iterrows():

            sample_name = str(row.get("sample_name", "")).strip()

            sample_accession = format_sample_accession(sample_name)
            batch_code = row.get("batch_code", "")

            existing_sample = find_existing_wgs_module_record_by_fields(
                SampleInformation,
                {
                    "sample_accession": sample_accession,
                    "sample_name": sample_name,
                    "batch_code": batch_code,
                },
            )
            if existing_sample and not overwrite:
                continue

            referred_obj = (
                Final_Data.objects.filter(f_AccessionNo=sample_accession).first()
                if sample_accession else None
            )
            classification = None
            if referred_obj:
                classification = Classification_Table.objects.filter(
                    Class_idNumReferred=referred_obj
                ).first()
            if not classification and sample_accession:
                classification = Classification_Table.objects.filter(
                    Class_AccessionNo=sample_accession
                ).first()

            connect_project = get_or_create_wgs_project_for_upload(
                sample_accession,
                referred_obj,
                "WGS_SampleInfo_Acc",
                "WGS_SampleInfoSummary",
                existing_project=existing_sample.sample_project if existing_sample else None,
            )

            def row_bool(*column_names):
                for column_name in column_names:
                    if column_name in row:
                        return upload_bool_or_false(row.get(column_name, False))
                return False

            def classification_bool(field_name):
                return bool(getattr(classification, field_name, False)) if classification else False

            def row_bool_or_class(classification_field, *column_names):
                for column_name in column_names:
                    if column_name in row:
                        return upload_bool_or_false(row.get(column_name, False))
                return classification_bool(classification_field)

            ghru_all = row_bool_or_class(
                "Class_Chk_GHRU_all",
                "ghru_all",
                "GHRU_all",
                "Class_Chk_GHRU_all",
                "ghru",
            )
            ghru_neo = row_bool_or_class("Class_Chk_GHRU_Neo", "ghru_neo", "GHRU_neo", "Class_Chk_GHRU_Neo")

            sampleinfo_record, _ = save_or_update_wgs_module_record(
                existing_sample,
                {
                    "_model": SampleInformation,
                    "sample_project": connect_project,
                    "sample_accession": sample_accession,
                    "batch_code": batch_code,
                    "sample_name": sample_name,
                    "status": row.get("status", ""),
                    "emerging": row_bool_or_class("Class_Chk_Emerging", "emerging", "Class_Chk_Emerging"),
                    "structured": row_bool_or_class("Class_Chk_Structured", "structured", "Class_Chk_Structured"),
                    "satscan": row_bool_or_class("Class_Chk_Satscan", "satscan", "Class_Chk_Satscan"),
                    "serotyping": row_bool_or_class("Class_Chk_Serotyping", "serotyping", "Class_Chk_Serotyping"),
                    "ghru": row_bool("ghru") or ghru_all,
                    "ghru_all": ghru_all,
                    "ghru_neo": ghru_neo,
                    "egasp": row_bool_or_class("Class_Chk_EGASP", "egasp", "Class_Chk_EGASP"),
                    "tricycle": row_bool_or_class("Class_Chk_Tricycle", "tricycle", "Class_Chk_Tricycle"),
                    "pulsenet": row_bool_or_class("Class_Chk_Pulsenet", "pulsenet", "Class_Chk_Pulsenet"),
                    "tulip": row_bool_or_class("Class_Chk_Tulip", "tulip", "Class_Chk_Tulip"),
                },
                overwrite=overwrite,
            )
            sync_wgs_records_from_sampleinfo(sampleinfo_record)

        messages.success(request, "Sample Information records uploaded successfully.")

        return redirect("show_sample_information")

    return render(request, "wgs_app/Add_wgs.html", {

        "form": form,
        "sampleinfo_form": sampleinfo_form,
        "bactscout_form": BactScoutUploadForm(),
        "gtdbtk_form" : GtdbTkUploadForm(),
        "gambit_form": GambitUploadForm(),
        "mlst_form": MlstUploadForm(),
        "checkm2_form": Checkm2UploadForm(),
        "amrfinder_form": AmrUploadForm(),
        "assembly_form": AssemblyUploadForm(),
        "demogs_form": DemogsDataUploadForm(),
        "antibiotic_form": FinalAntibioticUploadForm(),
        "raw_antibiotic_form": RawAntibioticUploadForm(),
        "editing": editing,
    })


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def relink_sample_information_matches(request):
    linked_count = relink_all_wgs_records_from_sampleinfo()
    messages.success(
        request,
        f"{linked_count} WGS record(s) relinked from Sample Information sample names."
    )
    return redirect("upload_wgs_view")






@login_required
def show_sample_information(request):

    records = SampleInformation.objects.all().order_by('-Date_uploaded_si')

    upload_dates = (
        SampleInformation.objects
        .exclude(Date_uploaded_si__isnull=True)
        .values_list('Date_uploaded_si', flat=True)
        .distinct()
        .order_by('-Date_uploaded_si')
    )

    total_records = SampleInformation.objects.count()

    paginator = Paginator(records, 20)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    return render(request, "wgs_app/show_sampleinfo.html", {

        "page_obj": page_obj,

        "upload_dates": upload_dates,

        "total_records": total_records,

    })



@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_sample_information(request, pk):

    item = get_object_or_404(SampleInformation, pk=pk)

    if request.method == "POST":

        item.delete()

        messages.success(
            request,
            f"Sample {item.sample_name} deleted successfully!"
        )

        return redirect("show_sample_information")

    messages.error(request, "Invalid request.")

    return redirect("show_sample_information")



@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_all_sample_information(request):

    deleted_count, _ = SampleInformation.objects.all().delete()

    messages.success(
        request,
        f"{deleted_count} Sample Information records deleted successfully."
    )

    return redirect("show_sample_information")


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_sample_information_by_date(request):

    if request.method == "POST":

        upload_date_str = request.POST.get("upload_date")

        if not upload_date_str:

            messages.error(request, "Please select a date.")

            return redirect("show_sample_information")

        upload_date = parse_date(upload_date_str)

        if not upload_date:

            messages.error(request, "Invalid date format.")

            return redirect("show_sample_information")

        deleted_count, _ = SampleInformation.objects.filter(
            Date_uploaded_si=upload_date
        ).delete()

        messages.success(
            request,
            f"Deleted {deleted_count} records uploaded on {upload_date}."
        )

        return redirect("show_sample_information")

    messages.error(request, "Invalid request method.")

    return redirect("show_sample_information")


################## Bactscout
@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
def upload_bactscout(request):

    form = WGSProjectForm()
    bactscout_form = BactScoutUploadForm()
    editing = False

    if request.method == "POST" and request.FILES.get("bactscoutfile"):
        overwrite = wgs_upload_overwrite_enabled(request)

        bactscout_form = BactScoutUploadForm(request.POST, request.FILES)

        try:
            upload = bactscout_form.save()
            df = read_uploaded_file(upload.bactscoutfile, sheet_name="bactscout")

            df.columns = df.columns.str.strip().str.replace(".", "", regex=False)
            if "genome_size_expected_status" in df.columns:
                cutoff_index = df.columns.get_loc("genome_size_expected_status")
                df = df.iloc[:, :cutoff_index + 1]

        except Exception as e:

            messages.error(request, f"Error processing BactScout file: {e}")

            return render(request, "wgs_app/Add_wgs.html", {
                "form": form,
                "sampleinfo_form": SampleInfoUploadForm(),
                "bactscout_form": bactscout_form,
                "gtdbtk_form" : GtdbTkUploadForm(),
                "gambit_form": GambitUploadForm(),
                "mlst_form": MlstUploadForm(),
                "checkm2_form": Checkm2UploadForm(),
                "amrfinder_form": AmrUploadForm(),
                "assembly_form": AssemblyUploadForm(),
                "demogs_form": DemogsDataUploadForm(),
                "antibiotic_form": FinalAntibioticUploadForm(),
                "raw_antibiotic_form": RawAntibioticUploadForm(),
                "editing": editing,
            })

        site_codes = set(SiteData.objects.values_list("SiteCode", flat=True))

        # accession extractor
        def format_bactscout_accession(raw_name: str):

            if not raw_name:
                return ""

            base = os.path.basename(raw_name)
            base_noext = os.path.splitext(base)[0].strip()

            if "ARS" not in base_noext:
                return ""

            parts = re.split(r"[-_]", base_noext)

            if not parts:
                return ""

            prefix = parts[0]

            for part in parts[1:]:

                m = re.match(r"^([A-Za-z]{2,6})(\d+)", part)

                if m:

                    letters = m.group(1).upper()
                    digits = m.group(2)

                    if letters in site_codes:
                        return f"{prefix}_{letters}{digits}"

            for i in range(1, len(parts) - 1):

                if parts[i].upper() in site_codes:

                    letters = parts[i].upper()

                    digits_match = re.search(r"(\d+)", parts[i + 1])

                    if digits_match:
                        return f"{prefix}_{letters}{digits_match.group(1)}"

                    return f"{prefix}_{letters}"

            return ""

        for _, row in df.iterrows():

            name = str(row.get("name") or row.get("sample_id") or "").strip()

            bactscout_accession = format_bactscout_accession(name)

            existing_bactscout = (
                BactScout.objects
                .filter(BactScout_Accession=bactscout_accession, name=name)
                .select_related("bactscout_project")
                .order_by("-id")
                .first()
                if name else None
            )
            if existing_bactscout and not overwrite:
                continue

            referred_obj = (
                Final_Data.objects.filter(f_AccessionNo=bactscout_accession).first()
                if bactscout_accession else None
            )

            connect_project = get_or_create_wgs_project_for_upload(
                bactscout_accession,
                referred_obj,
                "WGS_BactScout_Acc",
                "WGS_BactScoutSummary",
                existing_project=(
                    existing_bactscout.bactscout_project
                    if existing_bactscout else None
                ),
            )

            bactscout_defaults = {
                "bactscout_project": connect_project,
                "BactScout_Accession": bactscout_accession,
                "name": name,
                "status": row.get("status") or row.get("a_final_status"),

                "completeness": upload_number_or_none(row.get("completeness")),
                "contamination": upload_number_or_none(row.get("contamination")),
                "completeness_model_used": row.get("completeness_model_used"),
                "translation_table_used": upload_number_or_none(row.get("translation_table_used"), integer=True),
                "coding_density": upload_number_or_none(row.get("coding_density")),
                "contig_n50": upload_number_or_none(row.get("contig_n50"), integer=True),
                "average_gene_length": upload_number_or_none(row.get("average_gene_length")),
                "genome_size": upload_number_or_none(row.get("genome_size"), integer=True),
                "checkm2_gc_content": upload_number_or_none(row.get("checkm2_gc_content")),
                "total_coding_sequences": upload_number_or_none(row.get("total_coding_sequences"), integer=True),
                "total_contigs": upload_number_or_none(row.get("total_contigs"), integer=True),
                "max_contig_length": upload_number_or_none(row.get("max_contig_length"), integer=True),
                "additional_notes": row.get("additional_notes"),

                "a_final_status": row.get("a_final_status"),
                "adapter_detection_status": row.get("adapter_detection_status"),
                "contamination_status": row.get("contamination_status"),
                "species_status": row.get("species_status"),
                "coverage_status": row.get("coverage_status"),
                "coverage_estimate_qualibact_status": row.get("coverage_estimate_qualibact_status"),
                "duplication_status": row.get("duplication_status"),
                "gc_content_status": row.get("gc_content_status"),
                "mlst_status": row.get("mlst_status"),
                "n_content_status": row.get("n_content_status"),
                "read_length_status": row.get("read_length_status"),
                "read_q30_status": row.get("read_q30_status"),

                "species": row.get("species"),
                "species_abundance": upload_text_or_none(row.get("species_abundance")),
                "species_coverage": upload_text_or_none(row.get("species_coverage")),
                "species_message": row.get("species_message"),

                "contamination_message": row.get("contamination_message"),

                "coverage_estimate_sylph": upload_number_or_none(row.get("coverage_estimate_sylph")),
                "coverage_estimate_sylph_message": row.get("coverage_estimate_sylph_message"),
                "coverage_estimate_qualibact": upload_number_or_none(row.get("coverage_estimate_qualibact")),
                "coverage_estimate_qualibact_message": row.get("coverage_estimate_qualibact_message"),

                "duplication_rate": upload_number_or_none(row.get("duplication_rate")),
                "duplication_message": row.get("duplication_message"),

                "gc_content": upload_number_or_none(row.get("gc_content")),
                "gc_content_lower": upload_number_or_none(row.get("gc_content_lower"), integer=True),
                "gc_content_upper": upload_number_or_none(row.get("gc_content_upper"), integer=True),
                "gc_content_message": row.get("gc_content_message"),

                "n_content_rate": upload_number_or_none(row.get("n_content_rate")),
                "n_content_message": row.get("n_content_message"),

                "mlst_st": row.get("mlst_st"),
                "mlst_message": row.get("mlst_message"),

                "read1_mean_length": upload_number_or_none(row.get("read1_mean_length"), integer=True),
                "read2_mean_length": upload_number_or_none(row.get("read2_mean_length"), integer=True),
                "read_length_message": row.get("read_length_message"),

                "read_q20_bases": upload_number_or_none(row.get("read_q20_bases")),
                "read_q20_rate": upload_number_or_none(row.get("read_q20_rate")),
                "read_q30_bases": upload_number_or_none(row.get("read_q30_bases")),
                "read_q30_rate": upload_number_or_none(row.get("read_q30_rate")),
                "read_q30_message": row.get("read_q30_message"),
                "read_total_bases": upload_number_or_none(row.get("read_total_bases")),
                "read_total_reads": upload_number_or_none(row.get("read_total_reads"), integer=True),

                "adapter_detection_message": row.get("adapter_detection_message"),

                "ref_genome": row.get("ref_genome"),
                "genome_size_expected": upload_number_or_none(row.get("genome_size_expected"), integer=True),
                "genome_size_expected_status": row.get("genome_size_expected_status"),
            }

            save_or_update_wgs_module_record(
                existing_bactscout,
                {"_model": BactScout, **bactscout_defaults},
                overwrite=overwrite,
            )

        messages.success(request, "BactScout records uploaded successfully.")

        return redirect("show_bactscout")

    return render(request, "wgs_app/Add_wgs.html", {

        "form": form,
        "sampleinfo_form": SampleInfoUploadForm(),
        "bactscout_form": bactscout_form,
        "gtdbtk_form" : GtdbTkUploadForm(),
        "gambit_form": GambitUploadForm(),
        "mlst_form": MlstUploadForm(),
        "checkm2_form": Checkm2UploadForm(),
        "amrfinder_form": AmrUploadForm(),
        "assembly_form": AssemblyUploadForm(),
        "demogs_form": DemogsDataUploadForm(),
        "antibiotic_form": FinalAntibioticUploadForm(),
        "raw_antibiotic_form": RawAntibioticUploadForm(),
        "editing": editing,

    })


@login_required
def show_bactscout(request):

    records = BactScout.objects.all().order_by('-Date_uploaded_bs')

    upload_dates = (
        BactScout.objects.exclude(Date_uploaded_bs__isnull=True)
        .values_list('Date_uploaded_bs', flat=True)
        .distinct()
        .order_by('-Date_uploaded_bs')
    )

    total_records = BactScout.objects.count()

    paginator = Paginator(records, 20)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    return render(request, "wgs_app/show_bactscout.html", {

        "page_obj": page_obj,
        "upload_dates": upload_dates,
        "total_records": total_records,

    })


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_bactscout(request, pk):

    item = get_object_or_404(BactScout, pk=pk)

    if request.method == "POST":

        item.delete()

        messages.success(request, f"Record {item.name} deleted successfully!")

        return redirect("show_bactscout")

    messages.error(request, "Invalid request.")

    return redirect("show_bactscout")


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_all_bactscout(request):

    deleted_count, _ = BactScout.objects.all().delete()

    messages.success(request, f"{deleted_count} BactScout records deleted.")

    return redirect("show_bactscout")



@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_bactscout_by_date(request):

    if request.method == "POST":

        upload_date_str = request.POST.get("upload_date")

        if not upload_date_str:
            messages.error(request, "Please select an upload date.")
            return redirect("show_bactscout")

        upload_date = parse_date(upload_date_str)

        if not upload_date:
            messages.error(request, "Invalid date format.")
            return redirect("show_bactscout")

        deleted_count, _ = BactScout.objects.filter(
            Date_uploaded_bs=upload_date
        ).delete()

        messages.success(
            request,
            f"Deleted {deleted_count} records uploaded on {upload_date}."
        )

        return redirect("show_bactscout")

    messages.error(request, "Invalid request.")

    return redirect("show_bactscout")

############## GtdbTk 

@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
def upload_gtdbtk(request):

    form = WGSProjectForm()
    gtdbtk_form = GtdbTkUploadForm()
    editing = False

    if request.method == "POST" and request.FILES.get("GtdbTkFile"):
        overwrite = wgs_upload_overwrite_enabled(request)

        gtdbtk_form = GtdbTkUploadForm(request.POST, request.FILES)

        try:
            upload = gtdbtk_form.save()
            df = read_uploaded_file(upload.GtdbTkFile, sheet_name=["gtdbtk", "gtdb-tk", "gtdb_tk"])

            df.columns = df.columns.str.strip().str.replace(".", "", regex=False)

        except Exception as e:

            messages.error(request, f"Error processing GTDB-Tk file: {e}")

            return render(request, "wgs_app/Add_wgs.html", {
            "form": form,
            "sampleinfo_form": SampleInfoUploadForm(),
            "bactscout_form": BactScoutUploadForm(),
            "gtdbtk_form" : gtdbtk_form,
            "gambit_form": GambitUploadForm(),
            "mlst_form": MlstUploadForm(),
            "checkm2_form": Checkm2UploadForm(),
            "amrfinder_form": AmrUploadForm(),
            "assembly_form": AssemblyUploadForm(),
            "demogs_form": DemogsDataUploadForm(),
            "antibiotic_form": FinalAntibioticUploadForm(),
            "raw_antibiotic_form": RawAntibioticUploadForm(),
            "editing": editing,

            })

        site_codes = set(SiteData.objects.values_list("SiteCode", flat=True))

        # accession extractor
        def format_gtdbtk_accession(raw_name: str):

            if not raw_name:
                return ""

            base = os.path.basename(raw_name)
            base_noext = os.path.splitext(base)[0].strip()

            if "ARS" not in base_noext:
                return ""

            parts = re.split(r"[-_]", base_noext)

            if not parts:
                return ""

            prefix = parts[0]

            for part in parts[1:]:

                m = re.match(r"^([A-Za-z]{2,6})(\d+)", part)

                if m:

                    letters = m.group(1).upper()
                    digits = m.group(2)

                    if letters in site_codes:
                        return f"{prefix}_{letters}{digits}"

            for i in range(1, len(parts) - 1):

                if parts[i].upper() in site_codes:

                    letters = parts[i].upper()

                    digits_match = re.search(r"(\d+)", parts[i + 1])

                    if digits_match:
                        return f"{prefix}_{letters}{digits_match.group(1)}"

                    return f"{prefix}_{letters}"

            return ""

        for _, row in df.iterrows():

            user_genome = str(row.get("user_genome", "")).strip()

            gtdbtk_accession = format_gtdbtk_accession(user_genome)

            existing_gtdbtk = find_existing_wgs_module_record(
                GtdbTk,
                "GtdbTk_Accession",
                gtdbtk_accession,
                "user_genome",
                user_genome,
            )
            if existing_gtdbtk and not overwrite:
                continue

            referred_obj = (
                Final_Data.objects.filter(f_AccessionNo=gtdbtk_accession).first()
                if gtdbtk_accession else None
            )

            connect_project = get_or_create_wgs_project_for_upload(
                gtdbtk_accession,
                referred_obj,
                "WGS_GtdbTk_Acc",
                "WGS_GtdbTkSummary",
                existing_project=existing_gtdbtk.gtdbtk_project if existing_gtdbtk else None,
            )

            save_or_update_wgs_module_record(
                existing_gtdbtk,
                {
                    "_model": GtdbTk,
                    "gtdbtk_project": connect_project,
                    "GtdbTk_Accession": gtdbtk_accession,
                    "user_genome": user_genome,
                    "classification": row.get("classification"),
                    "closest_genome_reference": row.get("closest_genome_reference"),
                    "closest_genome_reference_radius": row.get("closest_genome_reference_radius"),
                    "closest_genome_taxonomy": row.get("closest_genome_taxonomy"),
                    "closest_genome_ani": row.get("closest_genome_ani"),
                    "closest_genome_af": row.get("closest_genome_af"),
                    "closest_placement_reference": row.get("closest_placement_reference"),
                    "closest_placement_radius": row.get("closest_placement_radius"),
                    "closest_placement_taxonomy": row.get("closest_placement_taxonomy"),
                    "closest_placement_ani": row.get("closest_placement_ani"),
                    "closest_placement_af": row.get("closest_placement_af"),
                    "pplacer_taxonomy": row.get("pplacer_taxonomy"),
                    "classification_method": row.get("classification_method"),
                    "note": row.get("note"),
                    "other_related_references": row.get("other_related_references"),
                    "msa_percent": row.get("msa_percent"),
                    "translation_table": row.get("translation_table"),
                    "red_value": row.get("red_value"),
                    "warnings": row.get("warnings"),
                },
                overwrite=overwrite,
            )

        messages.success(request, "GTDB-Tk records uploaded successfully.")

        return redirect("show_gtdbtk")

    return render(request, "wgs_app/Add_wgs.html", {

        "form": form,
        "sampleinfo_form": SampleInfoUploadForm(),
        "bactscout_form": BactScoutUploadForm(),
        "gtdbtk_form" : gtdbtk_form,
        "gambit_form": GambitUploadForm(),
        "mlst_form": MlstUploadForm(),
        "checkm2_form": Checkm2UploadForm(),
        "amrfinder_form": AmrUploadForm(),
        "assembly_form": AssemblyUploadForm(),
        "demogs_form": DemogsDataUploadForm(),
        "antibiotic_form": FinalAntibioticUploadForm(),
        "raw_antibiotic_form": RawAntibioticUploadForm(),
        "editing": editing,


    })



@login_required
def show_gtdbtk(request):

    records = GtdbTk.objects.all().order_by('-Date_uploaded_gt')

    upload_dates = (
        GtdbTk.objects.exclude(Date_uploaded_gt__isnull=True)
        .values_list('Date_uploaded_gt', flat=True)
        .distinct()
        .order_by('-Date_uploaded_gt')
    )

    total_records = GtdbTk.objects.count()

    paginator = Paginator(records, 20)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    return render(request, "wgs_app/show_gtdbtk.html", {

        "page_obj": page_obj,
        "upload_dates": upload_dates,
        "total_records": total_records,

    })


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_gtdbtk(request, pk):

    item = get_object_or_404(GtdbTk, pk=pk)

    if request.method == "POST":

        item.delete()

        messages.success(request, f"Record {item.user_genome} deleted successfully!")

        return redirect("show_gtdbtk")

    messages.error(request, "Invalid request.")

    return redirect("show_gtdbtk")


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_all_gtdbtk(request):

    deleted_count, _ = GtdbTk.objects.all().delete()

    messages.success(request, f"{deleted_count} GTDB-Tk records deleted.")

    return redirect("show_gtdbtk")


@login_required
@role_required(*WGS_WRITE_ALLOWED_ROLES)
@require_POST
def delete_gtdbtk_by_date(request):

    if request.method == "POST":

        upload_date_str = request.POST.get("upload_date")

        if not upload_date_str:
            messages.error(request, "Please select an upload date.")
            return redirect("show_gtdbtk")

        upload_date = parse_date(upload_date_str)

        if not upload_date:
            messages.error(request, "Invalid date format.")
            return redirect("show_gtdbtk")

        deleted_count, _ = GtdbTk.objects.filter(
            Date_uploaded_gt=upload_date
        ).delete()

        messages.success(
            request,
            f"Deleted {deleted_count} records uploaded on {upload_date}."
        )

        return redirect("show_gtdbtk")

    messages.error(request, "Invalid request.")

    return redirect("show_gtdbtk")


########### Show all WGS project entries for one Referred_Data AccessionNo,
##########  including BactScout, CheckM2, AMRFinder tables.

# @login_required

@login_required
def view_wgs_overview(request):
    """
    Displays only isolates (Final_Data) that have matched WGS data
    across any WGS table (BactScout, MLST, CheckM2, Assembly, Gambit, AMRFinder).
    Keeps the first page light by paginating before loading per-accession flags.
    """
    q = request.GET.get("q", "").strip()
    selected_year = request.GET.get("year", "").strip()

    wgs_sources = [
        (BactScout, "BactScout_Accession", "Date_uploaded_bs"),
        (Mlst, "Mlst_Accession", "Date_uploaded_m"),
        (Checkm2, "Checkm2_Accession", "Date_uploaded_c"),
        (AssemblyScan, "Assembly_Accession", "Date_uploaded_as"),
        (Gambit, "Gambit_Accession", "Date_uploaded_g"),
        (Amrfinderplus, "Amrfinder_Accession", "Date_uploaded_am"),
    ]

    def accession_year(value):
        text = str(value or "").strip().upper()
        match = re.match(r"^(\d{2})ARS", text)
        if match:
            return str(2000 + int(match.group(1)))
        match = re.match(r"^(\d{4})ARS", text)
        if match:
            return match.group(1)
        return ""

    def source_queryset(model, accession_field):
        return model.objects.exclude(**{f"{accession_field}__isnull": True}).exclude(
            **{accession_field: ""}
        )

    def accession_set(model, field_name):
        return {
            str(acc).strip()
            for acc in source_queryset(model, field_name)
            .values_list(field_name, flat=True)
            .distinct()
            if acc and str(acc).strip()
        }

    all_wgs_accessions = set().union(*[
        accession_set(model, accession_field)
        for model, accession_field, date_field in wgs_sources
    ])
    available_years = sorted(
        {year for year in (accession_year(acc) for acc in all_wgs_accessions) if year},
        reverse=True,
    )

    wgs_sets = {
        "bactscout": accession_set(BactScout, "BactScout_Accession"),
        "mlst": accession_set(Mlst, "Mlst_Accession"),
        "checkm2": accession_set(Checkm2, "Checkm2_Accession"),
        "assembly": accession_set(AssemblyScan, "Assembly_Accession"),
        "gambit": accession_set(Gambit, "Gambit_Accession"),
        "amrfinder": accession_set(Amrfinderplus, "Amrfinder_Accession"),
    }
    wgs_accessions = set().union(*wgs_sets.values())
    if selected_year:
        wgs_accessions = {
            accession
            for accession in wgs_accessions
            if accession_year(accession) == selected_year
        }

    referred_list = Final_Data.objects.only(
        "f_AccessionNo",
        "f_Patient_ID",
        "f_Last_Name",
        "f_First_Name",
        "f_Mid_Name",
        "f_Age",
        "f_Sex",
        "f_Ward",
        "f_Spec_Type",
        "f_ars_OrgCode",
        "f_SiteCode",
        "f_Diagnosis_ICD10",
        "f_Growth",
        "f_Spec_Date",
        "f_Referral_Date",
    ).filter(
        f_AccessionNo__in=wgs_accessions
    )

    if q:
        referred_list = referred_list.filter(
            Q(f_AccessionNo__icontains=q) |
            Q(f_Patient_ID__icontains=q) |
            Q(f_Last_Name__icontains=q) |
            Q(f_First_Name__icontains=q) |
            Q(f_ars_OrgCode__icontains=q) |
            Q(f_SiteCode__icontains=q)
        )

    referred_list = referred_list.order_by("f_AccessionNo")

    paginator = Paginator(referred_list, 50)
    page_obj = paginator.get_page(request.GET.get("page"))
    page_accessions = [
        item.f_AccessionNo.strip()
        for item in page_obj.object_list
        if item.f_AccessionNo and item.f_AccessionNo.strip()
    ]

    antibiotics_qs = Final_AntibioticEntry.objects.filter(
        ab_idNum_f_referred__f_AccessionNo__in=page_accessions
    ).select_related("ab_idNum_f_referred").only(
        "ab_idNum_f_referred__f_AccessionNo",
        "ab_Abx_code",
        "ab_MIC_RIS",
        "ab_MIC_value",
        "ab_Disk_value",
    )
    abx_map = {}
    for ab in antibiotics_qs:
        if not has_tested_antibiotic_value(ab.ab_Disk_value, ab.ab_MIC_value):
            continue
        acc = getattr(ab.ab_idNum_f_referred, "f_AccessionNo", None)
        if acc:
            abx_map.setdefault(acc, []).append({
                "code": ab.ab_Abx_code,
                "ris": ab.ab_MIC_RIS or "",
                "disk": ab.ab_Disk_value or "",
                "mic": ab.ab_MIC_value or "",
            })

    sampleinfo_map = {}
    for sampleinfo in (
        SampleInformation.objects
        .filter(sample_accession__in=page_accessions)
        .order_by("sample_accession", "-Date_uploaded_si", "-pk")
    ):
        accession = (sampleinfo.sample_accession or "").strip()
        if accession:
            sampleinfo_map.setdefault(accession, sampleinfo)

    table_data = []

    for referred in page_obj.object_list:
        acc = referred.f_AccessionNo.strip() if referred.f_AccessionNo else None
        if not acc:
            continue

        summary_flags = {
            key: acc in accessions for key, accessions in wgs_sets.items()
        }

        abx_entries = abx_map.get(acc, [])

        table_data.append({
            "accession": acc,
            "patient_id": referred.f_Patient_ID,
            "patient_name": f"{referred.f_Last_Name}, {referred.f_First_Name} {referred.f_Mid_Name or ''}".strip(),
            "age": referred.f_Age,
            "sex": referred.f_Sex,
            "ward": referred.f_Ward,
            "specimen": referred.f_Spec_Type,
            "organism": referred.f_ars_OrgCode,
            "sitecode": referred.f_SiteCode,
            "diagnosis": referred.f_Diagnosis_ICD10,
            "growth": referred.f_Growth,
            "date_collected": referred.f_Spec_Date,
            "referral_date": referred.f_Referral_Date,
            "summary_flags": summary_flags,
            "antibiotics": abx_entries,
            "sample_information": sampleinfo_map.get(acc),
        })

    filtered_accessions = {
        str(acc).strip()
        for acc in referred_list.values_list("f_AccessionNo", flat=True)
        if acc and str(acc).strip()
    }
    counts = {
        "total": paginator.count,
        "bactscout": len(wgs_sets["bactscout"] & filtered_accessions),
        "mlst": len(wgs_sets["mlst"] & filtered_accessions),
        "checkm2": len(wgs_sets["checkm2"] & filtered_accessions),
        "assembly": len(wgs_sets["assembly"] & filtered_accessions),
        "gambit": len(wgs_sets["gambit"] & filtered_accessions),
        "amrfinder": len(wgs_sets["amrfinder"] & filtered_accessions),
        "with_antibiotics": Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred__f_AccessionNo__in=filtered_accessions
        ).values("ab_idNum_f_referred__f_AccessionNo").distinct().count(),
    }

    params = request.GET.copy()
    params.pop("page", None)
    preserved_params = params.urlencode()

    return render(request, "wgs_app/Wgs_overview.html", {
        "table_data": table_data,
        "page_obj": page_obj,
        "counts": counts,
        "q": q,
        "available_years": available_years,
        "selected_year": selected_year,
        "preserved_params": preserved_params,
    })




# View WGS overview with antibiotic entries but optimized to reduce queries
# Shows isolates that have WGS data in ANY of the WGS tables

@login_required
def get_wgs_details(request, accession):
    print(f"\n=== FETCHING DETAILS FOR ACCESSION: {accession} ===")

    # Fetch the referred isolate
    referred = Final_Data.objects.filter(f_AccessionNo=accession).first()
    if not referred:
        return JsonResponse({"error": "Accession not found."}, status=404)

    # ============================================================
    # ✔ FIX: Convert antibiotic entries to SAFE dictionary format
    # ============================================================
    antibiotics_qs = Final_AntibioticEntry.objects.filter(
        ab_idNum_f_referred__f_AccessionNo=accession
    ).only(
        "ab_Abx_code",
        "ab_Disk_value", "ab_Disk_enRIS",
        "ab_MIC_value", "ab_MIC_enRIS",
        "ab_MIC_operand",
    )

    antibiotics = []
    for ab in antibiotics_qs:
        if not has_tested_antibiotic_value(ab.ab_Disk_value, ab.ab_MIC_value):
            continue
        antibiotics.append({
            "code": ab.ab_Abx_code,

            # Disk
            "disk": ab.ab_Disk_value or "",
            "d_ris": ab.ab_Disk_enRIS or "",

            # MIC
            "mic": ab.ab_MIC_value or "",
            "m_ris": ab.ab_MIC_enRIS or "",
            "m_op": ab.ab_MIC_operand or "",
        })

    print(f" Found {len(antibiotics)} antibiotic entries (SAFE FORMAT)")

    classification = (
        Classification_Table.objects
        .filter(Class_idNumReferred=referred)
        .first()
    )
    classification_items = []
    if classification:
        classification_labels = [
            ("Emerging", classification.Class_Chk_Emerging),
            ("Structured", classification.Class_Chk_Structured),
            ("SatScan", classification.Class_Chk_Satscan),
            ("Serotyping", classification.Class_Chk_Serotyping),
            ("GHRU All", classification.Class_Chk_GHRU_all),
            ("GHRU Neo", classification.Class_Chk_GHRU_Neo),
            ("EGASP", classification.Class_Chk_EGASP),
            ("Tricycle", classification.Class_Chk_Tricycle),
            ("PulseNet", classification.Class_Chk_Pulsenet),
            ("TULIP", classification.Class_Chk_Tulip),
        ]
        classification_items = [
            {"label": label}
            for label, active in classification_labels
            if active
        ]

    sample_information = (
        SampleInformation.objects
        .filter(sample_accession=accession)
        .order_by("-Date_uploaded_si", "-pk")
        .first()
    )

    # ============================================================
    # WGS Related Data
    # ============================================================
    projects = WGS_Project.objects.filter(
        Q(WGS_BactScout_Acc=accession)
        | Q(WGS_Mlst_Acc=accession)
        | Q(WGS_Checkm2_Acc=accession)
        | Q(WGS_Assembly_Acc=accession)
        | Q(WGS_Gambit_Acc=accession)
        | Q(WGS_Amrfinder_Acc=accession)
    ).distinct()

    related_data = {
        "bactscout": list(BactScout.objects.filter(
            Q(bactscout_project__in=projects) | Q(BactScout_Accession=accession)
        )),
        "mlst": list(Mlst.objects.filter(
            Q(mlst_project__in=projects) | Q(mlst_project__WGS_Mlst_Acc=accession)
        )),
        "checkm2": list(Checkm2.objects.filter(
            Q(checkm2_project__in=projects) | Q(checkm2_project__WGS_Checkm2_Acc=accession)
        )),
        "assembly": list(AssemblyScan.objects.filter(
            Q(assembly_project__in=projects) | Q(assembly_project__WGS_Assembly_Acc=accession)
        )),
        "gambit": list(Gambit.objects.filter(
            Q(gambit_project__in=projects) | Q(Gambit_Accession=accession)
        )),
        "amrfinder": list(Amrfinderplus.objects.filter(
            Q(amrfinder_project__in=projects) | Q(amrfinder_project__WGS_Amrfinder_Acc=accession)
        )),
    }

    # ============================================================
    # Build context required by the template
    # ============================================================
    context = {
        "entry": {
            "accession": accession,
            "patient_id": referred.f_Patient_ID,
            "patient_name": f"{referred.f_Last_Name}, {referred.f_First_Name} {referred.f_Mid_Name or ''}".strip(),
            "age": referred.f_Age,
            "sex": referred.f_Sex,
            "ward": referred.f_Ward,
            "specimen": referred.f_Spec_Type,
            "organism": referred.f_ars_OrgCode,
            "sitecode": referred.f_SiteCode,
            "diagnosis": referred.f_Diagnosis_ICD10,
            "growth": referred.f_Growth,
            "date_collected": referred.f_Spec_Date,
            "referral_date": referred.f_Referral_Date,
            "classification": classification,
            "classification_items": classification_items,
            "sample_information": sample_information,
            "antibiotics": antibiotics,  
            "related_data": related_data,
        }
    }

    # Render HTML for AJAX response
    html = render_to_string(
        "wgs_app/Wgs_detail.html",
        context,
        request=request,
    )

    return JsonResponse({"html": html})




@login_required
def download_matched_wgs_data(request):
    """
    Export Final_Data + Antibiotic results (MIC, Disk, RIS)
    along with WGS data (BactScout, MLST, CheckM2, Assembly, AMRFinder, Gambit).

    Mode options:
        ?mode=all  → Complete sets (present in ALL WGS tables)
        ?mode=any  → Partial sets (present in ANY WGS table)
    """
    import io
    import pandas as pd
    from django.http import HttpResponse

    mode = request.GET.get("mode", "any").lower()
    selected_year = request.GET.get("year", "").strip()

    def request_date(value):
        parsed = pd.to_datetime(value or "", errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()

    date_from = request_date(request.GET.get("date_from"))
    date_to = request_date(request.GET.get("date_to"))

    def normalize_accession(value):
        return str(value or "").strip().upper()

    def accession_year(value):
        text = normalize_accession(value)
        match = re.match(r"^(\d{2})ARS", text)
        if match:
            return str(2000 + int(match.group(1)))
        match = re.match(r"^(\d{4})ARS", text)
        if match:
            return match.group(1)
        return ""

    final_acc_map = {
        normalize_accession(acc): acc
        for acc in Final_Data.objects.exclude(f_AccessionNo__isnull=True)
        .exclude(f_AccessionNo="")
        .values_list("f_AccessionNo", flat=True)
        if normalize_accession(acc)
    }

    wgs_sources = [
        (BactScout, "BactScout_Accession", "Date_uploaded_bs"),
        (Mlst, "Mlst_Accession", "Date_uploaded_m"),
        (Checkm2, "Checkm2_Accession", "Date_uploaded_c"),
        (AssemblyScan, "Assembly_Accession", "Date_uploaded_as"),
        (Amrfinderplus, "Amrfinder_Accession", "Date_uploaded_am"),
        (Gambit, "Gambit_Accession", "Date_uploaded_g"),
    ]

    def wgs_accession_set(model, accession_field, date_field, apply_date_filter=True):
        qs = model.objects.exclude(**{f"{accession_field}__isnull": True}).exclude(
            **{accession_field: ""}
        )
        if apply_date_filter and (date_from or date_to):
            if date_from:
                qs = qs.filter(**{f"{date_field}__gte": date_from})
            if date_to:
                qs = qs.filter(**{f"{date_field}__lte": date_to})
        return {
            normalized
            for acc in qs.values_list(accession_field, flat=True).distinct()
            for normalized in [normalize_accession(acc)]
            if normalized in final_acc_map
        }

    def collect_wgs_sets(apply_date_filter=True):
        return [
            wgs_accession_set(model, accession_field, date_field, apply_date_filter)
            for model, accession_field, date_field in wgs_sources
        ]

    # ---- Step 1-2: Collect matching Final_Data/WGS accessions ----
    wgs_sets = collect_wgs_sets(apply_date_filter=True)

    # ---- Step 3: Combine or intersect ----
    if mode == "all":
        matched_accessions = set.intersection(*wgs_sets) if wgs_sets else set()
        filename_suffix = "Complete"
    else:
        matched_accessions = set().union(*wgs_sets) if wgs_sets else set()
        filename_suffix = "Partial"

    date_filter_was_used = bool(date_from or date_to)
    if not matched_accessions and date_filter_was_used:
        wgs_sets = collect_wgs_sets(apply_date_filter=False)
        if mode == "all":
            matched_accessions = set.intersection(*wgs_sets) if wgs_sets else set()
        else:
            matched_accessions = set().union(*wgs_sets) if wgs_sets else set()
        filename_suffix = f"{filename_suffix}_AllDates"

    if not matched_accessions:
        return HttpResponse(
            "No matching WGS accessions found in Final Referred_Data.",
            content_type="text/plain"
        )

    if selected_year:
        matched_accessions = {
            accession
            for accession in matched_accessions
            if accession_year(accession) == selected_year
        }
        if not matched_accessions:
            return HttpResponse(
                f"No matching WGS accessions found for {selected_year}.",
                content_type="text/plain"
            )

    matched_accessions = {
        final_acc_map[normalized]
        for normalized in matched_accessions
        if normalized in final_acc_map
    }

    # ---- Step 4: Query datasets ----
    final_qs = Final_Data.objects.filter(f_AccessionNo__in=matched_accessions)
    abx_qs = Final_AntibioticEntry.objects.filter(
        ab_idNum_f_referred__f_AccessionNo__in=matched_accessions
    )
    classification_qs = Classification_Table.objects.filter(
        Class_idNumReferred__f_AccessionNo__in=matched_accessions
    )
    sampleinfo_qs = SampleInformation.objects.filter(
        sample_accession__in=matched_accessions
    ).order_by("sample_accession", "-Date_uploaded_si", "-pk")
    bactscout_qs = BactScout.objects.filter(BactScout_Accession__in=matched_accessions)
    mlst_qs = Mlst.objects.filter(Mlst_Accession__in=matched_accessions)
    checkm2_qs = Checkm2.objects.filter(Checkm2_Accession__in=matched_accessions)
    assembly_qs = AssemblyScan.objects.filter(Assembly_Accession__in=matched_accessions)
    amrfinder_qs = Amrfinderplus.objects.filter(Amrfinder_Accession__in=matched_accessions)
    gambit_qs = Gambit.objects.filter(Gambit_Accession__in=matched_accessions)

    # ---- Step 5: Convert querysets to DataFrames ----
    final_df = pd.DataFrame.from_records(final_qs.values())
    abx_df = pd.DataFrame.from_records(abx_qs.values())
    classification_fields = [
        "Class_idNumReferred_id",
        "Class_AccessionNo",
        "Class_Chk_Emerging",
        "Class_Chk_Structured",
        "Class_Chk_Satscan",
        "Class_Chk_Serotyping",
        "Class_Chk_GHRU_all",
        "Class_Chk_GHRU_Neo",
        "Class_Chk_EGASP",
        "Class_Chk_Tricycle",
        "Class_Chk_Pulsenet",
        "Class_Chk_Tulip",
    ]
    classification_export_labels = [
        ("Class_Chk_Emerging", "Emerging"),
        ("Class_Chk_Structured", "Structured"),
        ("Class_Chk_Satscan", "SatScan"),
        ("Class_Chk_Serotyping", "Serotyping"),
        ("Class_Chk_GHRU_all", "GHRU All"),
        ("Class_Chk_GHRU_Neo", "GHRU Neo"),
        ("Class_Chk_EGASP", "EGASP"),
        ("Class_Chk_Tricycle", "Tricycle"),
        ("Class_Chk_Pulsenet", "PulseNet"),
        ("Class_Chk_Tulip", "TULIP"),
    ]
    classification_df = pd.DataFrame.from_records(
        classification_qs.values(*classification_fields)
    )
    sampleinfo_fields = [
        "sample_accession",
        "batch_code",
        "sample_name",
        "status",
        "DNA_extraction",
        "dna_performed_by",
        "dna_performed_date",
        "library_preparation",
        "library_performed_by",
        "library_performed_date",
        "sequencing_platform",
        "Date_uploaded_si",
        "id",
    ]
    sampleinfo_df = pd.DataFrame.from_records(sampleinfo_qs.values(*sampleinfo_fields))

    def qs_to_df(qs, model_name, acc_field):
        if not qs.exists():
            return pd.DataFrame()
        df = pd.DataFrame.from_records(qs.values())
        df.insert(0, "Table", model_name)
        df.insert(1, "f_AccessionNo", df[acc_field])
        return df

    bactscout_df = qs_to_df(bactscout_qs, "BactScout", "BactScout_Accession")
    mlst_df = qs_to_df(mlst_qs, "Mlst", "Mlst_Accession")
    checkm2_df = qs_to_df(checkm2_qs, "Checkm2", "Checkm2_Accession")
    assembly_df = qs_to_df(assembly_qs, "AssemblyScan", "Assembly_Accession")
    amrfinder_df = qs_to_df(amrfinder_qs, "Amrfinderplus", "Amrfinder_Accession")
    gambit_df = qs_to_df(gambit_qs, "Gambit", "Gambit_Accession")

    final_df["id"] = final_df["id"].astype(str)
    if not classification_df.empty:
        classification_df["Classification_Summary"] = classification_df.apply(
            lambda row: ", ".join(
                label
                for field, label in classification_export_labels
                if bool(row.get(field))
            ),
            axis=1,
        )
        classification_df["Class_idNumReferred_id"] = classification_df[
            "Class_idNumReferred_id"
        ].astype(str)
        combined_classification_df = classification_df.rename(
            columns={"Class_idNumReferred_id": "id"}
        )
        final_df = final_df.merge(
            combined_classification_df,
            on="id",
            how="left",
        )

    sampleinfo_export_columns = [
        "SampleInfo_batch_code",
        "SampleInfo_sample_name",
        "SampleInfo_status",
        "SampleInfo_DNA_extraction",
        "SampleInfo_dna_performed_by",
        "SampleInfo_dna_performed_date",
        "SampleInfo_library_preparation",
        "SampleInfo_library_performed_by",
        "SampleInfo_library_performed_date",
        "SampleInfo_sequencing_platform",
    ]
    if not sampleinfo_df.empty:
        sampleinfo_df = sampleinfo_df.sort_values(
            ["sample_accession", "Date_uploaded_si", "id"],
            ascending=[True, False, False],
        ).drop_duplicates("sample_accession", keep="first")
        sampleinfo_df = sampleinfo_df.rename(
            columns={
                "sample_accession": "f_AccessionNo",
                "batch_code": "SampleInfo_batch_code",
                "sample_name": "SampleInfo_sample_name",
                "status": "SampleInfo_status",
                "DNA_extraction": "SampleInfo_DNA_extraction",
                "dna_performed_by": "SampleInfo_dna_performed_by",
                "dna_performed_date": "SampleInfo_dna_performed_date",
                "library_preparation": "SampleInfo_library_preparation",
                "library_performed_by": "SampleInfo_library_performed_by",
                "library_performed_date": "SampleInfo_library_performed_date",
                "sequencing_platform": "SampleInfo_sequencing_platform",
            }
        )[["f_AccessionNo", *sampleinfo_export_columns]]
        final_df = final_df.merge(sampleinfo_df, on="f_AccessionNo", how="left")
    else:
        for column in sampleinfo_export_columns:
            final_df[column] = ""

    combined_df = final_df.copy()
    if not abx_df.empty:
        abx_df["ab_idNum_f_referred_id"] = abx_df["ab_idNum_f_referred_id"].astype(str)
        abx_df = abx_df.merge(
            final_df[["id", "f_AccessionNo"]],
            left_on="ab_idNum_f_referred_id",
            right_on="id",
            how="left"
        )

        def pivot_antibiotic(df, value_field, suffix):
            pivot = df.pivot_table(
                index="f_AccessionNo",
                columns="ab_Abx_code",
                values=value_field,
                aggfunc="first"
            )
            pivot.columns = [f"{col}_{suffix}" for col in pivot.columns]
            return pivot

        abx_mic_val = pivot_antibiotic(abx_df, "ab_MIC_value", "MIC")
        abx_mic_ris = pivot_antibiotic(abx_df, "ab_MIC_RIS", "MIC_RIS")
        abx_mic_enris = pivot_antibiotic(abx_df, "ab_MIC_enRIS", "MIC_enRIS")
        abx_disk_val = pivot_antibiotic(abx_df, "ab_Disk_value", "Disk")
        abx_disk_ris = pivot_antibiotic(abx_df, "ab_Disk_RIS", "Disk_RIS")
        abx_disk_enris = pivot_antibiotic(abx_df, "ab_Disk_enRIS", "Disk_enRIS")

        abx_pivot = pd.concat(
            [
                abx_mic_val,
                abx_mic_ris,
                abx_mic_enris,
                abx_disk_val,
                abx_disk_ris,
                abx_disk_enris,
            ],
            axis=1
        )
        abx_pivot.reset_index(inplace=True)
        combined_df = final_df.merge(abx_pivot, on="f_AccessionNo", how="left")

    # ---- Step 7: Make datetimes timezone-naive ----
    def make_tz_naive(df):
        if df.empty:
            return df
        for col in df.select_dtypes(include=["datetimetz", "datetime"]).columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.tz_localize(None)
            df[col] = df[col].dt.strftime("%Y-%m-%d")
        return df

    combined_df = make_tz_naive(combined_df)

    # ---- Step 8: Write all to Excel ----
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        combined_df.to_excel(writer, index=False, sheet_name="Final_Data_With_Antibiotics")
        if not bactscout_df.empty: bactscout_df.to_excel(writer, index=False, sheet_name="BactScout")
        if not mlst_df.empty: mlst_df.to_excel(writer, index=False, sheet_name="MLST")
        if not checkm2_df.empty: checkm2_df.to_excel(writer, index=False, sheet_name="CheckM2")
        if not assembly_df.empty: assembly_df.to_excel(writer, index=False, sheet_name="Assembly")
        if not amrfinder_df.empty: amrfinder_df.to_excel(writer, index=False, sheet_name="AMRFinder")
        if not gambit_df.empty: gambit_df.to_excel(writer, index=False, sheet_name="Gambit")

    output.seek(0)
    filename = f"FinalData_WGS_{filename_suffix}_{pd.Timestamp.now().date()}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def projects_page(request):

    active_tab = request.GET.get("tab", "wgs_classification")

    referred_list = list(
        Final_Data.objects
        .all()
        .order_by("-id")
    )
    referred_accessions = [
        item.f_AccessionNo for item in referred_list if item.f_AccessionNo
    ]
    sampleinfo_by_accession = {}
    for item in (
        SampleInformation.objects
        .filter(sample_accession__in=referred_accessions)
        .order_by("sample_accession", "-Date_uploaded_si", "-pk")
    ):
        sampleinfo_by_accession.setdefault(item.sample_accession, item)
    for referred in referred_list:
        referred.sampleinfo_flags = sampleinfo_by_accession.get(referred.f_AccessionNo)

    context = {
        "active_tab":      active_tab,
        "referred_list":   referred_list,

        # WGS upload forms — passed to the modal forms rendered inside the tab
        "bactscout_form":  BactScoutUploadForm(),
        "gambit_form":     GambitUploadForm(),
        "mlst_form":       MlstUploadForm(),
        "checkm2_form":    Checkm2UploadForm(),
        "assembly_form":   AssemblyUploadForm(),
        "amrfinder_form":  AmrUploadForm(),
    }
    return render(request, "projects/projects.html", context)





##############################  New Upload Final View HELPERS



# ─────────────────────────────────────────────────────────────────────────────
# MAIN UPLOAD VIEW
# ─────────────────────────────────────────────────────────────────────────────
#######  refrain from using this until i polished it
@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@require_POST
def upload_final_data(request):
    """
    POST — receives a multipart Excel file (.xlsx) and saves rows into
    Final_Data and Final_AntibioticEntry (retest antibiotic entries only).

    Returns JSON: { success, created, updated, skipped, errors }
    """
    f = request.FILES.get("final_data_file")
    if not f:
        return JsonResponse({"success": False, "error": "No file uploaded."}, status=400)

    overwrite = request.POST.get("overwrite", "false").lower() == "true"

    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Cannot open file: {e}"}, status=400)

    results = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = rows[0]
        h_lower = [str(c).lower().strip() if c else "" for c in headers]
        col     = {name: idx for idx, name in enumerate(h_lower) if name}

        accession_col = next(
            (
                key for key in (
                    "accession_no",
                    "accession",
                    "f_accessionno",
                    "f_accession_no",
                )
                if key in col
            ),
            None,
        )

        if not accession_col:
            continue

        # Detect retest antibiotic column groups once per sheet
        abx_groups = _parse_abx_columns(headers)

        for row_num, row in enumerate(rows[1:], start=2):

            accession_no = _clean(row[col[accession_col]])
            if not accession_no:
                continue

            # ── Helper: safe column read ──────────────────────────────────────
            def g(key, default=""):
                return _clean(row[col[key]], default) if key in col else default

            def first_present(*keys):
                for key in keys:
                    if key in col:
                        return key
                return None

            def add_text(target, *keys, default=""):
                key = first_present(*keys)
                if key:
                    fd_kwargs[target] = _clean(row[col[key]], default)

            def add_date(target, *keys):
                key = first_present(*keys)
                if key:
                    fd_kwargs[target] = _date(row[col[key]])

            def add_int(target, *keys):
                key = first_present(*keys)
                if key:
                    fd_kwargs[target] = _int(row[col[key]])

            # ── Final_Data fields ─────────────────────────────────────────────
            fd_kwargs = {"f_AccessionNo": accession_no}
            add_text("f_Batch_Code", "batch_code", "f_batch_code")
            add_text("f_Batch_Name", "batch_name", "f_batch_name")
            add_text("f_SiteCode", "site_code", "sitecode", "f_sitecode")
            add_text("f_BatchNo", "batchno", "batch_no", "f_batchno")
            add_text("f_Total_batch", "total_batch", "f_total_batch")
            add_text("f_RefNo", "refno", "ref_no", "f_refno")
            add_date("f_Referral_Date", "referral_date", "f_referral_date")
            add_text("f_Patient_ID", "patient_id", "f_patient_id")
            add_text("f_First_Name", "first_name", "f_first_name")
            add_text("f_Mid_Name", "mid_name", "middle_name", "f_mid_name")
            add_text("f_Last_Name", "last_name", "f_last_name")
            add_date("f_Date_Birth", "date_birth", "birth_date", "f_date_birth")
            add_int("f_Age", "age", "f_age")
            add_text("f_Sex", "sex", "f_sex", default="n/a")
            add_date("f_Date_Admis", "date_admis", "admission_date", "f_date_admis")
            add_text("f_Nosocomial", "nosocomial", "f_nosocomial", default="n/a")
            add_text("f_Diagnosis", "diagnosis", "f_diagnosis")
            add_text("f_Diagnosis_ICD10", "diagnosis_icd10", "f_diagnosis_icd10")
            add_text("f_Ward", "ward", "f_ward")
            add_text("f_Ward_Type", "ward_type", "f_ward_type")
            add_text("f_Service_Type", "service_type", "f_service_type", default="n/a")
            add_text("f_Spec_Num", "spec_num", "f_spec_num")
            add_date("f_Spec_Date", "spec_date", "specimen_date", "year", "f_spec_date")
            add_text("f_Reason", "reason", "f_reason", default="n/a")
            add_text("f_Growth", "growth", "f_growth")
            add_text("f_Urine_ColCt", "urine_colct", "f_urine_colct")
            add_text("f_ars_Pre", "arsrl_pre", "ars_pre", "f_ars_pre")
            add_text("f_ars_OrgName", "arsrl_org", "ars_org", "f_ars_orgname")
            add_text("f_ars_OrgCode", "organismcode", "organism_code", "ars_org_code", "f_ars_orgcode")
            add_text("f_ars_Post", "arsrl_post", "ars_post", "f_ars_post")
            add_text("f_x_mrse", "x_mrse", "f_x_mrse")
            add_text("f_x_mrsamrse", "x_mrsamrse", "f_x_mrsamrse")
            add_text("f_x_entbac", "x_entbac", "f_x_entbac")
            add_text("f_edta", "edta", "f_edta")
            add_text("f_ars_ct_ctl", "ct.ctl", "ct_ctl", "f_ars_ct_ctl")
            add_text("f_ars_tz_tzl", "tz.tzl", "tz_tzl", "f_ars_tz_tzl")
            add_text("f_ars_cn_cni", "cn.cni", "cn_cni", "f_ars_cn_cni")
            add_text("f_ars_ip_ipi", "ip.ipi", "ip_ipi", "f_ars_ip_ipi")
            add_text("f_Comments", "comments", "f_comments")
            add_text("f_ars_reco", "recommendation", "ars_reco", "f_ars_reco")

            # Phenotype fields
            for excel_key, model_field in PHENO_FIELD_MAP.items():
                if excel_key in col:
                    fd_kwargs[model_field] = _pheno(row[col[excel_key]])
            for excel_key, model_field in SITE_PHENO_MAP.items():
                if excel_key in col:
                    fd_kwargs[model_field] = _pheno(row[col[excel_key]])

            # Resolve SpecimenType FK
            spec_type_key = first_present("spec_type", "f_spec_type")
            spec_type_code = _clean(row[col[spec_type_key]]) if spec_type_key else ""
            if spec_type_key and spec_type_code:
                fd_kwargs["f_Spec_Type"] = SpecimenTypeModel.objects.filter(
                    Specimen_code__iexact=spec_type_code
                ).first()

            # Resolve Batch FK
            if fd_kwargs.get("f_Batch_Code"):
                batch_obj = Batch_Table.objects.filter(
                    bat_Batch_Code=fd_kwargs["f_Batch_Code"]
                ).first()
                if batch_obj:
                    fd_kwargs["f_Batch_id"] = batch_obj

            # ── Persist Final_Data ────────────────────────────────────────────
            try:
                with transaction.atomic():
                    existing = Final_Data.objects.filter(
                        f_AccessionNo=accession_no
                    ).first()

                    if existing and not overwrite:
                        results["skipped"] += 1
                        continue

                    if existing:
                        for k, v in fd_kwargs.items():
                            if k != "f_AccessionNo":
                                setattr(existing, k, v)
                        update_fields = [
                            k for k in fd_kwargs
                            if k != "f_AccessionNo"
                        ]
                        if update_fields:
                            existing.save(update_fields=update_fields)
                            results["updated"] += 1
                        else:
                            results["skipped"] += 1
                        isolate = existing
                    else:
                        isolate = Final_Data(**fd_kwargs)
                        isolate.save()
                        results["created"] += 1

                    # ── Effective breakpoint year (same logic as edit view) ────
                    effective_year   = _resolve_effective_year(isolate.f_Spec_Date)
                    # ── ARSRL organism code for breakpoint lookup ─────────────
                    resolved_ars_org = (isolate.f_ars_OrgCode or "").strip()

                    # ── Remove old retest entries on overwrite ────────────────
                    if existing and overwrite and not partial_update:
                        Final_AntibioticEntry.objects.filter(
                            ab_idNum_f_referred=isolate,
                            ab_Retest_Abx_code__isnull=False,
                        ).exclude(ab_Retest_Abx_code="").delete()

                    # ── Save one retest entry per antibiotic with data ─────────
                    for abx_code, grp in abx_groups.items():

                        disk_raw = row[grp['disk_col']]  if grp['disk_col']  is not None else None
                        disk_ris = _clean(row[grp['disk_ris_col']]) if grp['disk_ris_col'] is not None else ""
                        mic_raw  = row[grp['mic_col']]   if grp['mic_col']   is not None else None
                        mic_ris  = _clean(row[grp['mic_ris_col']])  if grp['mic_ris_col']  is not None else ""

                        # Skip entirely empty antibiotic columns
                        if all(
                            v is None or _clean(v) == ""
                            for v in [disk_raw, disk_ris, mic_raw, mic_ris]
                        ):
                            continue

                        disk_int              = _int(disk_raw)
                        mic_value, mic_operand = _decimal(mic_raw)

                        # Nothing numeric to save
                        if disk_int is None and mic_value is None:
                            continue

                        _save_retest_entry(
                            isolate          = isolate,
                            abx_code         = abx_code,
                            disk_int         = disk_int,
                            disk_ris         = disk_ris,
                            mic_value        = mic_value,
                            mic_operand      = mic_operand,
                            mic_ris          = mic_ris,
                            resolved_ars_org = resolved_ars_org,
                            effective_year   = effective_year,
                        )

            except Exception as e:
                results["errors"].append({
                    "row":       row_num,
                    "accession": accession_no,
                    "sheet":     sheet_name,
                    "error":     str(e),
                })

    wb.close()
    results["success"] = True
    return JsonResponse(results)






########### uploading of Final Data - Demographics 


## new version aligns with the new final data model


def parse_int(val):

    if val is None:
        return None

    s = str(val).strip().lower()

    if s in {"", "nan", "none"}:
        return None

    # extract numeric portion (10d → 10)
    match = re.search(r"\d+", s)

    if match:
        return int(match.group())

    return None



# @login_required
# @transaction.atomic
# def upload_final_combined_table(request):



#     if request.method == "POST" and request.FILES.get("DemogsDataFile"):


#         try:

#             uploaded_file = request.FILES["DemogsDataFile"]
#             file_name = uploaded_file.name.lower()

#             # ================= LOAD FILE =================
#             if file_name.endswith(".csv"):
#                 df = pd.read_csv(uploaded_file)

#             elif file_name.endswith((".xlsx", ".xls")):
#                 df = pd.read_excel(uploaded_file)

#             else:
#                 messages.error(request, "Unsupported file format.")
#                 return redirect("upload_final_combined_table")

#             # ================= PREP DATA =================
#             df.columns = [str(c).strip() for c in df.columns]
#             rows = df.to_dict("records")

#             model_fields = {f.name for f in Final_Data._meta.fields}

#             IMMUTABLE_FIELDS = {
#                 "f_Date_of_Entry",
#                 "f_Date_Modified",
#             }

#             INT_FIELDS = {
#                 "f_Age",
#                 "f_bat_seq",
#             }

#             created = 0
#             updated = 0

#             # ================= FK LOOKUP CACHE =================
#             specimen_map = {
#                 str(s.Specimen_code).strip().lower(): s
#                 for s in SpecimenTypeModel.objects.all()
#             }

#             site_map = {
#                 str(s.SiteCode).strip().upper(): s.SiteName
#                 for s in SiteData.objects.all()
#             }

#             # ================= HELPERS =================
#             def parse_date(val):

#                 if not val or str(val).lower() in {"nan", "nat", "none", ""}:
#                     return None

#                 dt = pd.to_datetime(val, errors="coerce")
#                 return None if pd.isna(dt) else dt.date()

#             # ================= PROCESS ROWS =================
#             for row in rows:

#                 accession = str(row.get("f_AccessionNo", "")).strip()
#                 batch_code = str(row.get("f_Batch_Code", "")).strip()

#                 if not accession or not batch_code:
#                     continue

#                 # ---- normalize accession
#                 accession = accession.upper()

#                 # ---- parse date fields
#                 for d in (
#                     "f_Referral_Date",
#                     "f_Spec_Date",
#                     "f_Date_Birth",
#                     "f_Date_Admis",
#                 ):
#                     if d in row:
#                         row[d] = parse_date(row[d])

#                 # ---- keep ONLY Final_Data fields
#                 clean_row = {
#                     k: v
#                     for k, v in row.items()
#                     if k in model_fields and k not in IMMUTABLE_FIELDS
#                 }

#                 # ---- normalize integers
#                 for f in INT_FIELDS:
#                     if f in clean_row:
#                         clean_row[f] = parse_int(clean_row[f])

#                 # ================= FK RESOLUTION =================

#                 # Specimen Type FK
#                 if "f_Spec_Type" in clean_row:

#                     raw_val = clean_row["f_Spec_Type"]

#                     if pd.isna(raw_val) or raw_val is None:
#                         clean_row["f_Spec_Type"] = None

#                     else:
#                         spec_val = str(raw_val).strip().lower()
#                         clean_row["f_Spec_Type"] = specimen_map.get(spec_val)

                
#                 # ================= SITE NAME AUTO ASSIGN =================
#                 site_code = str(clean_row.get("f_SiteCode", "")).strip().upper()

#                 if site_code:
#                     clean_row["f_Site_Name"] = site_map.get(site_code, "")


#                 # NEVER assign batch FK directly
#                 clean_row.pop("f_Batch_id", None)

#                 # ================= UPSERT =================
#                 obj, is_created = Final_Data.objects.update_or_create(
#                     f_AccessionNo=accession,
#                     f_Batch_Code=batch_code,
#                     defaults=clean_row
#                 )

#                 created += int(is_created)
#                 updated += int(not is_created)

#             # ================= SUCCESS =================
#             messages.success(
#                 request,
#                 f"Upload complete! {created} created, {updated} updated."
#             )

#             return redirect("show_final_data")

#         except Exception as e:

#             import traceback
#             traceback.print_exc()

#             messages.error(request, f"Upload failed: {e}")
#             return redirect("upload_final_combined_table")

#     return render(request, "wgs_app/Add_wgs.html")

### HELPER TO RECOGNIZE EITHER DATASET FROM REFERRED_DATA AND FINAL_DATA
def get_value(row, *keys):
    """
    Return first existing value from possible column names
    """
    def normalize_key(value):
        return re.sub(r"[^a-z0-9]", "", str(value).lower())

    for k in keys:
        if k in row and str(row.get(k)).strip() != "":
            return row.get(k)

    normalized_row = {
        normalize_key(k): v
        for k, v in row.items()
    }
    for k in keys:
        normalized = normalize_key(k)
        if normalized in normalized_row and str(normalized_row.get(normalized)).strip() != "":
            return normalized_row.get(normalized)

    return None



# @login_required
# @transaction.atomic
# def upload_final_combined_table(request):

#     if request.method != "POST" or not request.FILES.get("DemogsDataFile"):
#         return render(request, "wgs_app/Add_wgs.html")

#     try:

#         uploaded_file = request.FILES["DemogsDataFile"]
#         file_name = uploaded_file.name.lower()

#         # ================= LOAD FILE =================
#         if file_name.endswith(".csv"):
#             df = pd.read_csv(uploaded_file)

#         elif file_name.endswith((".xlsx", ".xls")):
#             df = pd.read_excel(uploaded_file)

#         else:
#             messages.error(request, "Unsupported file format.")
#             return redirect("upload_final_combined_table")

#         df.columns = [str(c).strip() for c in df.columns]
#         rows = df.to_dict("records")

#         # ================= FIELD DEFINITIONS =================
#         final_fields = {f.name for f in Final_Data._meta.fields}

#         IMMUTABLE_FIELDS = {
#             "f_Date_of_Entry",
#             "f_Date_Modified",
#         }

#         INT_FIELDS = {
#             "f_Age",
#             "f_bat_seq",
#         }

#         # ================= LOOKUP CACHE =================
#         specimen_map = {
#             str(s.Specimen_code).strip().lower(): s
#             for s in SpecimenTypeModel.objects.all()
#         }

#         site_map = {
#             str(s.SiteCode).strip().upper(): s.SiteName
#             for s in SiteData.objects.all()
#         }

#         created = 0
#         updated = 0

#         # ================= PROCESS ROWS =================

#         for row in rows:

#             accession = str(get_value(row, "f_AccessionNo", "AccessionNo") or "").strip().upper()
#             batch_code = str(get_value(row, "f_Batch_Code", "Batch_Code") or "").strip()

#             if not accession:
#                 continue

#             referral_date = parse_date(get_value(row, "f_Referral_Date", "Referral_Date"))
#             spec_date = parse_date(get_value(row, "f_Spec_Date", "Spec_Date"))
#             birth_date = parse_date(get_value(row, "f_Date_Birth", "Date_Birth"))
#             admis_date = parse_date(get_value(row, "f_Date_Admis", "Date_Admis"))

#             site_code = str(get_value(row, "f_SiteCode", "SiteCode") or "").strip().upper()
#             site_name = site_map.get(site_code, "")

#             # ================= REFERRED DATA =================

#             Referred_Data.objects.update_or_create(
#                 AccessionNo=accession,
#                 defaults={
#                     "SiteCode": site_code,
#                     "Site_Name": site_name,
#                     "Referral_Date": referral_date,
#                 }
#             )

#             # ================= PREP FINAL DATA =================

#             clean_row = {}

#             for k, v in row.items():

#                 if k not in final_fields or k in IMMUTABLE_FIELDS:
#                     continue

#                 clean_row[k] = v

#             clean_row["f_AccessionNo"] = accession
#             clean_row["f_Batch_Code"] = batch_code
#             clean_row["f_Referral_Date"] = referral_date
#             clean_row["f_Spec_Date"] = spec_date
#             clean_row["f_Date_Birth"] = birth_date
#             clean_row["f_Date_Admis"] = admis_date
#             clean_row["f_SiteCode"] = site_code
#             clean_row["f_Site_Name"] = site_name

#             # normalize integers
#             for field in INT_FIELDS:
#                 if field in clean_row:
#                     clean_row[field] = parse_int(clean_row[field])

#             # specimen FK
#             if "f_Spec_Type" in clean_row:

#                 raw_val = clean_row["f_Spec_Type"]

#                 if pd.isna(raw_val) or raw_val is None:
#                     clean_row["f_Spec_Type"] = None
#                 else:
#                     spec_val = str(raw_val).strip().lower()
#                     clean_row["f_Spec_Type"] = specimen_map.get(spec_val)

#             # ================= FINAL DATA UPSERT =================

#             obj, is_created = Final_Data.objects.update_or_create(
#                 f_AccessionNo=accession,
#                 defaults=clean_row
#             )

#             if is_created:
#                 created += 1
#             else:
#                 updated += 1

#         # ================= SUCCESS =================
#         created_batches = generate_batches_from_referred()

#         messages.success(
#             request,
#             f"Upload complete! {created} created, {updated} updated. "
#             f"{created_batches} batch(es) generated automatically."
#         )


#         return redirect("show_final_data")

#     except Exception as e:

#         import traceback
#         traceback.print_exc()

#         messages.error(request, f"Upload failed: {e}")
#         return redirect("upload_wgs_view")


def parse_date(val):

    if val is None:
        return None

    if pd.isna(val):
        return None

    if isinstance(val, pd.Timestamp):
        return val.date()

    if isinstance(val, datetime):
        return val.date()

    if isinstance(val, date):
        return val

    try:
        parsed = pd.to_datetime(val, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()
    except Exception:
        return None
    


########### uploding of demogs view

# @login_required
# @transaction.atomic
# def upload_final_combined_table(request):

#     if request.method != "POST" or not request.FILES.get("DemogsDataFile"):
#         return render(request, "wgs_app/Add_wgs.html")

#     try:

#         uploaded_file = request.FILES["DemogsDataFile"]
#         file_name = uploaded_file.name.lower()

#         # ================= READ FILE =================
#         if file_name.endswith(".csv"):
#             df = pd.read_csv(uploaded_file)

#         elif file_name.endswith((".xlsx", ".xls")):
#             df = pd.read_excel(uploaded_file)

#         else:
#             messages.error(request, "Unsupported file format.")
#             return redirect("upload_final_combined_table")

#         df.columns = [str(c).strip() for c in df.columns]
#         rows = df.to_dict("records")

#         # ================= FIELD LISTS =================
#         referred_fields = {
#             f.name for f in Referred_Data._meta.fields
#             if not f.auto_created
#         }

#         final_fields = {
#             f.name for f in Final_Data._meta.fields
#             if not f.auto_created
#         }

#         IMMUTABLE_FIELDS = {
#             "f_Date_of_Entry",
#             "f_Date_Modified",
#         }

#         INT_FIELDS = {
#             "Age",
#             "f_Age",
#             "bat_seq",
#             "f_bat_seq"
#         }

#         # ================= LOOKUP CACHE =================
#         specimen_map = {
#             str(s.Specimen_code).strip().lower(): s
#             for s in SpecimenTypeModel.objects.all()
#         }

#         site_map = {
#             str(s.SiteCode).strip().upper(): s.SiteName
#             for s in SiteData.objects.all()
#         }

#         # ================= EXISTING CACHE =================
#         existing_referred = {
#             obj.AccessionNo: obj
#             for obj in Referred_Data.objects.all()
#         }

#         existing_final = {
#             obj.f_AccessionNo: obj
#             for obj in Final_Data.objects.all()
#         }

#         referred_create = []
#         referred_update = []

#         final_create = []
#         final_update = []

#         # ================= PROCESS ROWS =================
#         for row in rows:

#             row = {str(k).strip(): v for k, v in row.items()}

#             accession = str(get_value(row, "AccessionNo", "accession_no")).strip().upper()

#             if not accession:
#                 continue

#             site_code = str(get_value(row, "SiteCode", "site_code") or "").strip().upper()
#             site_name = site_map.get(site_code, "")

#             referral_date = parse_date(get_value(row, "Referral_Date", "referral_date"))
#             birth_date = parse_date(get_value(row, "Date_Birth", "date_birth"))
#             admis_date = parse_date(get_value(row, "Date_Admis", "date_admis"))
#             spec_date = parse_date(get_value(row, "Spec_Date", "spec_date"))

#             batch_code = get_value(row, "Batch_Code", "batch_code")

#             # ================= REFERRED DATA =================
#             clean_referred = {
#                 k: v
#                 for k, v in row.items()
#                 if k in referred_fields
#             }

#             clean_referred["AccessionNo"] = accession
#             clean_referred["Referral_Date"] = referral_date
#             clean_referred["Date_Birth"] = birth_date
#             clean_referred["Date_Admis"] = admis_date
#             clean_referred["Spec_Date"] = spec_date
#             clean_referred["SiteCode"] = site_code
#             clean_referred["Site_Name"] = site_name
#             clean_referred["Batch_Code"] = batch_code

#             # integer cleanup
#             for field in INT_FIELDS:
#                 if field in clean_referred:
#                     clean_referred[field] = parse_int(clean_referred[field])

#             # specimen FK
#             if "Spec_Type" in clean_referred:

#                 raw_val = clean_referred["Spec_Type"]

#                 if raw_val:
#                     clean_referred["Spec_Type"] = specimen_map.get(
#                         str(raw_val).strip().lower()
#                     )

#             if accession in existing_referred:

#                 obj = existing_referred[accession]

#                 for k, v in clean_referred.items():
#                     setattr(obj, k, v)

#                 referred_update.append(obj)

#             else:

#                 referred_create.append(
#                     Referred_Data(**clean_referred)
#                 )

#             # ================= FINAL DATA =================
#             clean_final = {}

#             for k, v in row.items():

#                 fk = f"f_{k}"

#                 if fk in final_fields and fk not in IMMUTABLE_FIELDS:
#                     clean_final[fk] = v

#             clean_final["f_AccessionNo"] = accession
#             clean_final["f_Referral_Date"] = referral_date
#             clean_final["f_Date_Birth"] = birth_date
#             clean_final["f_Date_Admis"] = admis_date
#             clean_final["f_Spec_Date"] = spec_date
#             clean_final["f_SiteCode"] = site_code
#             clean_final["f_Site_Name"] = site_name
#             clean_final["f_Batch_Code"] = batch_code

#             for field in INT_FIELDS:
#                 if field in clean_final:
#                     clean_final[field] = parse_int(clean_final[field])

#             if "f_Spec_Type" in clean_final:

#                 raw_val = clean_final["f_Spec_Type"]

#                 if raw_val:
#                     clean_final["f_Spec_Type"] = specimen_map.get(
#                         str(raw_val).strip().lower()
#                     )

#             if accession in existing_final:

#                 obj = existing_final[accession]

#                 for k, v in clean_final.items():
#                     setattr(obj, k, v)

#                 final_update.append(obj)

#             else:

#                 final_create.append(
#                     Final_Data(**clean_final)
#                 )

#         # ================= REMOVE DUPLICATES =================
#         referred_update = list({o.AccessionNo: o for o in referred_update}.values())
#         final_update = list({o.f_AccessionNo: o for o in final_update}.values())

#         # ================= BULK SAVE =================
#         if referred_create:
#             Referred_Data.objects.bulk_create(referred_create)

#         if referred_update:
#             Referred_Data.objects.bulk_update(
#                 referred_update,
#                 [f for f in referred_fields if f != "id"]
#             )

#         if final_create:
#             Final_Data.objects.bulk_create(final_create)

#         if final_update:
#             Final_Data.objects.bulk_update(
#                 final_update,
#                 [f for f in final_fields if f not in IMMUTABLE_FIELDS and f != "id"]
#             )

#         created = len(final_create)
#         updated = len(final_update)

#         # ================= BATCH GENERATION =================
#         created_batches = generate_batches_from_referred()

#         messages.success(
#             request,
#             f"Upload complete! {created} created, {updated} updated. "
#             f"{created_batches} batch(es) generated."
#         )

#         return redirect("show_final_data")

#     except Exception as e:

#         import traceback
#         traceback.print_exc()

#         messages.error(request, f"Upload failed: {e}")
#         return redirect("upload_wgs_view")

def copy_batches_to_final(batch_ids):
    from apps.home_final.signals import suspend_import_signals

    copied = 0
    last_final_obj = None
    final_text_fields = {
        field.name
        for field in Final_Data._meta.fields
        if isinstance(field, (models.CharField, models.TextField))
        and not field.null
    }

    def clean_final_text(value):
        if value is None:
            return ""
        try:
            if pd.isna(value):
                return ""
        except TypeError:
            pass
        if str(value).strip().lower() in {"nan", "nat", "none", "null"}:
            return ""
        return str(value).strip()

    isolates = list(
        Referred_Data.objects
        .filter(Batch_id_id__in=batch_ids)
        .select_related("Batch_id", "Spec_Type")
        .order_by("Batch_id_id", "bat_seq")
    )
    accessions = [
        clean_final_text(isolate.AccessionNo)
        for isolate in isolates
        if clean_final_text(isolate.AccessionNo)
    ]
    existing_by_accession = {
        obj.f_AccessionNo: obj
        for obj in Final_Data.objects.filter(f_AccessionNo__in=accessions)
    }
    create_objects = []
    update_objects = []

    for isolate in isolates:

        spec_obj = isolate.Spec_Type

        final_defaults = {

                      # ===== BATCH / META =====
                "f_bat_seq": isolate.bat_seq,
                "f_Batch_id": isolate.Batch_id,
                "f_Hide": getattr(isolate, "Hide", False),
                "f_Batch_Code": isolate.Batch_Code,
                "f_Batch_Name": isolate.Batch_Name,
                "f_RefNo": isolate.RefNo,
                "f_BatchNo": isolate.BatchNo,
                "f_Total_batch": isolate.Total_batch,

                "f_AccessionNoGen": getattr(isolate, "AccessionNoGen", ""),

                "f_SiteCode": isolate.SiteCode,
                "f_Site_Name": isolate.Site_Name,
                "f_Referral_Date": isolate.Referral_Date,

                # ===== PATIENT =====
                "f_Patient_ID": isolate.Patient_ID,
                "f_First_Name": isolate.First_Name,
                "f_Mid_Name": isolate.Mid_Name,
                "f_Last_Name": isolate.Last_Name,
                "f_Date_Birth": isolate.Date_Birth,
                "f_Age": isolate.Age,
                "f_Age_Display": getattr(isolate, "Age_Display", ""),
                "f_Sex": isolate.Sex,

                "f_Date_Admis": isolate.Date_Admis,
                "f_Nosocomial": isolate.Nosocomial,
                "f_Diagnosis": isolate.Diagnosis,
                "f_Diagnosis_ICD10": isolate.Diagnosis_ICD10,
                "f_Ward": isolate.Ward,
                "f_Ward_Type": isolate.Ward_Type,
                "f_Service_Type": isolate.Service_Type,

                # ===== SPECIMEN =====
                "f_Spec_Num": isolate.Spec_Num,
                "f_Spec_Date": isolate.Spec_Date,
                "f_Reason": isolate.Reason,
                "f_Growth": isolate.Growth,
                "f_Urine_ColCt": isolate.Urine_ColCt,

                # ===== PHENOTYPE =====
                "f_ampC": isolate.ampC,
                "f_ESBL": isolate.ESBL,
                "f_CARB": isolate.CARB,
                "f_MBL": isolate.MBL,
                "f_BL": isolate.BL,
                "f_MR": isolate.MR,
                "f_mecA": isolate.mecA,
                "f_ICR": isolate.ICR,
                "f_OtherResMech": isolate.OtherResMech,


                "f_Spec_Type": spec_obj,


                # ===== SENTINEL ORGANISM =====
                "f_Site_Pre_ed": nz(isolate.Site_Pre_ed),
                "f_Site_Org": nz(isolate.Site_Org),
                "f_Site_OrgName": nz(isolate.Site_OrgName),
                "f_Site_Pos": nz(isolate.Site_Pos),
                "f_Comments": nz(isolate.Comments),

                # ===== ARSRL =====
                "f_ars_ampC": isolate.ars_ampC,
                "f_ars_ESBL": isolate.ars_ESBL,
                "f_ars_CARB": isolate.ars_CARB,
                "f_ars_ECIM": isolate.ars_ECIM,
                "f_ars_MCIM": isolate.ars_MCIM,
                "f_ars_EC_MCIM": isolate.ars_EC_MCIM,
                "f_ars_MBL": isolate.ars_MBL,
                "f_ars_BL": isolate.ars_BL,
                "f_ars_MR": isolate.ars_MR,
                "f_ars_mecA": isolate.ars_mecA,
                "f_ars_ICR": isolate.ars_ICR,

                "f_ars_Pre_ed": nz(isolate.ars_Pre_ed),
                "f_ars_Post": nz(isolate.ars_Post),
                "f_ars_OrgCode": nz(isolate.ars_OrgCode),
                "f_ars_OrgName": nz(isolate.ars_OrgName),

                "f_ars_ct_ctl": isolate.ars_ct_ctl,
                "f_ars_tz_tzl": isolate.ars_tz_tzl,
                "f_ars_cn_cni": isolate.ars_cn_cni,
                "f_ars_ip_ipi": isolate.ars_ip_ipi,

                "f_ars_reco_Code": isolate.ars_reco_Code,
                "f_ars_description": isolate.ars_description,
                "f_ars_reco": isolate.ars_reco,

                # ===== SIGNATORIES =====
                "f_arsp_Encoder": isolate.arsp_Encoder,
                "f_arsp_Enc_Lic": isolate.arsp_Enc_Lic,
                "f_arsp_Checker": isolate.arsp_Checker,
                "f_arsp_Chec_Lic": isolate.arsp_Chec_Lic,
                "f_arsp_Verifier": isolate.arsp_Verifier,
                "f_arsp_Ver_Lic": isolate.arsp_Ver_Lic,
                "f_arsp_LabManager": isolate.arsp_LabManager,
                "f_arsp_Lab_Lic": isolate.arsp_Lab_Lic,
                "f_arsp_Head": isolate.arsp_Head,
                "f_arsp_Head_Lic": isolate.arsp_Head_Lic,
                "f_Date_Accomplished_ARSP": isolate.Date_Accomplished_ARSP,

                # ===== EXTRA =====
                "f_x_mrse": getattr(isolate, "x_mrse", ""),
                "f_x_mrsamrse": getattr(isolate, "x_mrsamrse", ""),
                "f_x_entbac": getattr(isolate, "x_entbac", ""),
                "f_edta": getattr(isolate, "edta", ""),
        }

        for field_name in final_text_fields:
            if field_name in final_defaults:
                final_defaults[field_name] = clean_final_text(
                    final_defaults[field_name]
                )

        accession = clean_final_text(isolate.AccessionNo)
        final_obj = existing_by_accession.get(accession)

        if final_obj:
            for field_name, value in final_defaults.items():
                setattr(final_obj, field_name, value)
            update_objects.append(final_obj)
        else:
            final_obj = Final_Data(
                f_AccessionNo=accession,
                **final_defaults,
            )
            create_objects.append(final_obj)
            existing_by_accession[accession] = final_obj

        last_final_obj = final_obj
        copied += 1

    update_fields = sorted({
        field_name
        for obj in update_objects
        for field_name in final_defaults
    })

    with suspend_import_signals():
        if create_objects:
            Final_Data.objects.bulk_create(
                create_objects,
                batch_size=1000,
            )
        if update_objects:
            Final_Data.objects.bulk_update(
                update_objects,
                update_fields,
                batch_size=1000,
            )

        final_by_accession = {
            accession: final_obj
            for accession, final_obj in existing_by_accession.items()
            if final_obj.pk
        }
        final_ids = [obj.pk for obj in final_by_accession.values()]
        if final_ids:
            Final_AntibioticEntry.objects.filter(
                ab_idNum_f_referred_id__in=final_ids
            ).delete()

        raw_entries = list(
            AntibioticEntry.objects.filter(
                ab_idNum_referred_id__in=[
                    isolate.AccessionNo for isolate in isolates
                ]
            ).select_related("ab_idNum_referred")
        )
        raw_field_names = {
            field.name
            for field in AntibioticEntry._meta.concrete_fields
        }
        copy_fields = [
            field
            for field in Final_AntibioticEntry._meta.concrete_fields
            if field.name in raw_field_names
            and field.name not in {
                "id",
                "ab_idNum_f_referred",
                "ab_Date_uploaded_fd",
                "ab_Date_uploaded_rd",
            }
        ]
        final_antibiotics = []
        for raw_entry in raw_entries:
            final_obj = final_by_accession.get(
                raw_entry.ab_idNum_referred.AccessionNo
            )
            if not final_obj:
                continue
            copied_values = {}
            for field in copy_fields:
                value = getattr(raw_entry, field.name)
                if value is None and not field.null:
                    if field.has_default():
                        value = field.get_default()
                    elif isinstance(field, (models.CharField, models.TextField)):
                        value = ""
                    elif isinstance(field, models.BooleanField):
                        value = False
                copied_values[field.name] = value
            final_antibiotics.append(
                Final_AntibioticEntry(
                    ab_idNum_f_referred_id=final_obj.pk,
                    **copied_values,
                )
            )
        if final_antibiotics:
            Final_AntibioticEntry.objects.bulk_create(
                final_antibiotics,
                batch_size=1000,
            )

    return copied, last_final_obj


def regenerate_batch_concordance(batch_ids, user=None):
    from apps.home_final.services import concordance as concordance_service

    regenerated = 0

    for batch_id in batch_ids:
        concordance_service.generate_concordance_for_batch(batch_id, user=user)
        regenerated += 1

    return regenerated


def refresh_tat_for_batches(batch_ids):
    refreshed = 0

    tat_entries = TATform.objects.filter(tat_Batch_Isolates_id__in=batch_ids)

    for tat in tat_entries:
        total_isolates = Referred_Data.objects.filter(
            Batch_id_id=tat.tat_Batch_Isolates_id
        ).count()
        tat.tat_Num_Isolate = total_isolates
        tat.save()
        refreshed += 1

    return refreshed


def refresh_emerging_for_batches(batch_ids):
    from apps.home_final.signals import recalculate_emerging_for_final

    refreshed = 0

    isolates = Final_Data.objects.filter(f_Batch_id_id__in=batch_ids)

    for isolate in isolates:
        recalculate_emerging_for_final(isolate)
        refreshed += 1

    return refreshed


def refresh_emerging_for_final_ids(final_ids):
    from apps.home_final.signals import recalculate_emerging_for_final

    refreshed = 0

    isolates = Final_Data.objects.filter(pk__in=final_ids)

    for isolate in isolates:
        recalculate_emerging_for_final(isolate)
        refreshed += 1

    return refreshed


def regenerate_concordance_for_final_ids(final_ids, user=None):
    from apps.home_final.services import concordance as concordance_service

    regenerated = 0

    isolates = Final_Data.objects.filter(pk__in=final_ids)

    for isolate in isolates:
        concordance_service.generate_concordance_for_isolate(isolate, user=user)
        regenerated += 1

    return regenerated


def reapply_raw_breakpoints_for_batches(batch_ids):
    from apps.home.signals import determine_ris

    updated = 0
    changed_entries = []
    entry_breakpoints = {}
    effective_year_for, resolve_breakpoint = _build_breakpoint_resolver()

    entries = (
        AntibioticEntry.objects
        .filter(ab_idNum_referred__Batch_id_id__in=batch_ids)
        .select_related("ab_idNum_referred")
        .prefetch_related("ab_breakpoints_id")
    )

    for entry in entries:
        isolate = entry.ab_idNum_referred
        effective_year = effective_year_for(isolate.Spec_Date)
        site_org = (isolate.Site_Org or "").strip()
        ars_org = (
            (isolate.ars_OrgCode or "").strip()
            or site_org
        )
        breakpoint_ids = []

        if entry.ab_Abx_code:
            _clear_bp_fields(entry)

            disk_bp = None
            mic_bp = None

            if entry.ab_Disk_value is not None:
                disk_bp = resolve_breakpoint(
                    entry.ab_Abx_code,
                    isolate.Spec_Date,
                    "DISK",
                    site_org,
                )

            if entry.ab_MIC_value is not None:
                mic_bp = resolve_breakpoint(
                    entry.ab_Abx_code,
                    isolate.Spec_Date,
                    "MIC",
                    site_org,
                )

            bp = mic_bp or disk_bp
            if bp:
                if mic_bp:
                    _apply_bp_to_entry(
                        entry, bp, False, alert_mic=True, set_relation=False
                    )
                else:
                    _apply_bp_to_entry(entry, bp, True, set_relation=False)
                breakpoint_ids.append(bp.id)

        if entry.ab_Retest_Abx_code:
            _clear_retest_bp_fields(entry)

            retest_disk_bp = None
            retest_mic_bp = None

            if entry.ab_Retest_DiskValue is not None:
                retest_disk_bp = resolve_breakpoint(
                    entry.ab_Retest_Abx_code,
                    isolate.Spec_Date,
                    "DISK",
                    ars_org,
                )

            if entry.ab_Retest_MICValue is not None:
                retest_mic_bp = resolve_breakpoint(
                    entry.ab_Retest_Abx_code,
                    isolate.Spec_Date,
                    "MIC",
                    ars_org,
                )

            bp = retest_mic_bp or retest_disk_bp
            if bp:
                if retest_mic_bp:
                    _apply_bp_to_retest_entry(
                        entry, bp, False, alert_mic=True, set_relation=False
                    )
                else:
                    _apply_bp_to_retest_entry(
                        entry, bp, True, set_relation=False
                    )
                breakpoint_ids.append(bp.id)

        _recalculate_antibiotic_ris(entry, determine_ris)
        changed_entries.append(entry)
        entry_breakpoints[entry.pk] = breakpoint_ids
        updated += 1

    if changed_entries:
        AntibioticEntry.objects.bulk_update(
            changed_entries,
            [
                "ab_Site_Org", "ab_Ret_Org",
                "ab_Org_Flag", "ab_Abx_Flag", "ab_Abx_Phenotype",
                "ab_R_breakpoint", "ab_I_breakpoint",
                "ab_SDD_breakpoint", "ab_S_breakpoint", "ab_Alert_val",
                "ab_Ret_R_breakpoint", "ab_Ret_I_breakpoint",
                "ab_Ret_SDD_breakpoint", "ab_Ret_S_breakpoint",
                "ab_Retest_Alert_val",
                "ab_Disk_RIS", "ab_MIC_RIS",
                "ab_Retest_Disk_RIS", "ab_Retest_MIC_RIS",
            ],
            batch_size=1000,
        )
        _bulk_replace_breakpoint_links(
            AntibioticEntry,
            entry_breakpoints,
        )

    return updated


def reapply_final_breakpoints_for_batches(batch_ids, debug=False):
    from apps.home_final.signals import determine_ris

    updated = 0
    debug_rows = []
    changed_entries = []
    entry_breakpoints = {}
    effective_year_for, resolve_breakpoint = _build_breakpoint_resolver()

    entries = (
        Final_AntibioticEntry.objects
        .filter(ab_idNum_f_referred__f_Batch_id_id__in=batch_ids)
        .select_related("ab_idNum_f_referred")
        .prefetch_related("ab_breakpoints_id")
    )

    for entry in entries:
        isolate = entry.ab_idNum_f_referred
        effective_year = effective_year_for(isolate.f_Spec_Date)
        site_org = (isolate.f_Site_Org or "").strip()
        ars_org = (
            (isolate.f_ars_OrgCode or "").strip()
            or site_org
        )
        breakpoint_ids = []
        applied_notes = []

        if entry.ab_Abx_code:
            _clear_bp_fields(entry)

            disk_bp = None
            mic_bp = None

            if entry.ab_Disk_value is not None:
                disk_bp = resolve_breakpoint(
                    entry.ab_Abx_code,
                    isolate.f_Spec_Date,
                    "DISK",
                    site_org,
                )

            if entry.ab_MIC_value is not None:
                mic_bp = resolve_breakpoint(
                    entry.ab_Abx_code,
                    isolate.f_Spec_Date,
                    "MIC",
                    site_org,
                )

            bp = mic_bp or disk_bp
            if bp:
                if mic_bp:
                    _apply_bp_to_entry(
                        entry, bp, False, alert_mic=True, set_relation=False
                    )
                    applied_notes.append(
                        f"main MIC {entry.ab_Abx_code} -> BP#{bp.id} {bp.Org or 'generic'}"
                    )
                else:
                    _apply_bp_to_entry(entry, bp, True, set_relation=False)
                    applied_notes.append(
                        f"main DISK {entry.ab_Abx_code} -> BP#{bp.id} {bp.Org or 'generic'}"
                    )
                breakpoint_ids.append(bp.id)
            else:
                applied_notes.append(f"main {entry.ab_Abx_code} -> no breakpoint")

        if entry.ab_Retest_Abx_code:
            _clear_retest_bp_fields(entry)

            retest_disk_bp = None
            retest_mic_bp = None

            if entry.ab_Retest_DiskValue is not None:
                retest_disk_bp = resolve_breakpoint(
                    entry.ab_Retest_Abx_code,
                    isolate.f_Spec_Date,
                    "DISK",
                    ars_org,
                )

            if entry.ab_Retest_MICValue is not None:
                retest_mic_bp = resolve_breakpoint(
                    entry.ab_Retest_Abx_code,
                    isolate.f_Spec_Date,
                    "MIC",
                    ars_org,
                )

            bp = retest_mic_bp or retest_disk_bp
            if bp:
                if retest_mic_bp:
                    _apply_bp_to_retest_entry(
                        entry, bp, False, alert_mic=True, set_relation=False
                    )
                    applied_notes.append(
                        f"retest MIC {entry.ab_Retest_Abx_code} -> BP#{bp.id} {bp.Org or 'generic'}"
                    )
                else:
                    _apply_bp_to_retest_entry(
                        entry, bp, True, set_relation=False
                    )
                    applied_notes.append(
                        f"retest DISK {entry.ab_Retest_Abx_code} -> BP#{bp.id} {bp.Org or 'generic'}"
                    )
                breakpoint_ids.append(bp.id)
            else:
                applied_notes.append(f"retest {entry.ab_Retest_Abx_code} -> no breakpoint")

        _recalculate_antibiotic_ris(entry, determine_ris)
        changed_entries.append(entry)
        entry_breakpoints[entry.pk] = breakpoint_ids
        updated += 1

        if debug:
            debug_rows.append(
                f"{isolate.f_AccessionNo}: year={effective_year or 'none'}, "
                f"site_org={site_org or '-'}, ars_org={ars_org or '-'}, "
                f"{'; '.join(applied_notes) or 'no antibiotic code'}"
            )

    if changed_entries:
        Final_AntibioticEntry.objects.bulk_update(
            changed_entries,
            [
                "ab_Site_Org", "ab_Ret_Org",
                "ab_Org_Flag", "ab_Abx_Flag", "ab_Abx_Phenotype",
                "ab_R_breakpoint", "ab_I_breakpoint",
                "ab_SDD_breakpoint", "ab_S_breakpoint", "ab_Alert_val",
                "ab_Ret_R_breakpoint", "ab_Ret_I_breakpoint",
                "ab_Ret_SDD_breakpoint", "ab_Ret_S_breakpoint",
                "ab_Retest_Alert_val",
                "ab_Disk_RIS", "ab_MIC_RIS",
                "ab_Retest_Disk_RIS", "ab_Retest_MIC_RIS",
            ],
            batch_size=1000,
        )
        _bulk_replace_breakpoint_links(
            Final_AntibioticEntry,
            entry_breakpoints,
        )

    if debug:
        print("[FINAL ABX DEBUG] Breakpoint refresh summary:")
        if debug_rows:
            for row in debug_rows[:100]:
                print(f"[FINAL ABX DEBUG]   {row}")
            if len(debug_rows) > 100:
                print(f"[FINAL ABX DEBUG]   ... {len(debug_rows) - 100} more entries")
        else:
            print("[FINAL ABX DEBUG]   No final antibiotic entries found for affected batches.")

    return updated


########## UPLOADING DEMOGS
@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@transaction.atomic
def upload_referred_table(request):

    if request.method != "POST" or not request.FILES.get("DemogsDataFile"):
        return render(request, "wgs_app/Add_wgs.html")

    try:

        uploaded_file = request.FILES["DemogsDataFile"]
        upload_mode = request.POST.get("demographic_upload_mode", "raw").strip().lower()
        overwrite = request.POST.get("overwrite", "false").lower() == "true"
        file_name = uploaded_file.name.lower()

        # ================= READ FILE =================
        if file_name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)

        elif file_name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(uploaded_file)

        else:
            messages.error(request, "Unsupported file format.")
            return redirect("upload_wgs_view")

        df.columns = [str(c).strip() for c in df.columns]
        rows = df.to_dict("records")

        def is_blank_value(value):
            if value is None:
                return True
            try:
                if pd.isna(value):
                    return True
            except TypeError:
                pass
            return str(value).strip().lower() in {"", "nan", "nat", "none", "null"}

        def clean_text_value(value):
            return "" if is_blank_value(value) else str(value).strip()

        def parse_bool_value(value):
            if is_blank_value(value):
                return False
            if isinstance(value, bool):
                return value
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return bool(value)

            normalized = str(value).strip().lower()
            if normalized in {"true", "t", "yes", "y", "1", "checked"}:
                return True
            if normalized in {"false", "f", "no", "n", "0", "unchecked"}:
                return False

            return False

        def parse_date_value(value):
            if is_blank_value(value):
                return None
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            return parse_date(clean_text_value(value))

        if upload_mode == "final":
            final_fields = {
                f.name for f in Final_Data._meta.fields
                if not f.auto_created
            }
            final_field_map = {
                f.name: f for f in Final_Data._meta.fields
                if not f.auto_created
            }
            final_text_fields = {
                name for name, field in final_field_map.items()
                if isinstance(field, (models.CharField, models.TextField))
            }
            final_boolean_fields = {
                name for name, field in final_field_map.items()
                if isinstance(field, models.BooleanField)
            }
            final_date_fields = {
                name for name, field in final_field_map.items()
                if isinstance(field, models.DateField)
            }
            final_int_fields = {"f_Age", "f_bat_seq"}

            specimen_map = {
                str(s.Specimen_code).strip().lower(): s
                for s in SpecimenTypeModel.objects.all()
            }
            site_map = {
                str(s.SiteCode).strip().upper(): s.SiteName
                for s in SiteData.objects.all()
            }

            def ref_no_from_batch_code(batch_code):
                batch_code = clean_text_value(batch_code)
                if "_" not in batch_code:
                    return ""
                return batch_code.rsplit("_", 1)[-1].strip()

            normalized_rows = []
            uploaded_accessions = set()

            for row in rows:
                row = {str(k).strip(): v for k, v in row.items()}

                accession = clean_text_value(
                    get_value(row, "f_AccessionNo", "AccessionNo", "accession_no", "ID_Number")
                ).upper()
                if not accession:
                    continue

                uploaded_accessions.add(accession)
                normalized_rows.append({"row": row, "accession": accession})

            existing = {
                obj.f_AccessionNo: obj
                for obj in Final_Data.objects.filter(
                    f_AccessionNo__in=uploaded_accessions
                )
            }
            to_create = []
            to_update = []
            skipped = 0
            touched_batch_ids = set()

            for normalized in normalized_rows:
                row = normalized["row"]
                accession = normalized["accession"]
                clean = {}

                for field_name in final_fields:
                    if field_name in {"id", "f_AccessionNo", "f_Batch_id", "f_Spec_Type"}:
                        continue

                    raw_key = field_name[2:] if field_name.startswith("f_") else field_name
                    value = get_value(row, field_name, raw_key, raw_key.lower())
                    if value is None:
                        continue

                    clean[field_name] = value

                for field in final_text_fields:
                    if field in clean:
                        clean[field] = clean_text_value(clean[field])

                for field in final_boolean_fields:
                    if field in clean:
                        clean[field] = parse_bool_value(clean[field])

                for field in final_date_fields:
                    if field in clean:
                        clean[field] = parse_date_value(clean[field])

                for field in final_int_fields:
                    if field in clean:
                        clean[field] = parse_int(clean[field])

                site_code = clean_text_value(
                    get_value(row, "f_SiteCode", "SiteCode", "site_code")
                ).upper()
                if site_code:
                    clean["f_SiteCode"] = site_code
                    clean["f_Site_Name"] = clean.get("f_Site_Name") or site_map.get(site_code, "")

                spec_type = get_value(row, "f_Spec_Type", "Spec_Type", "spec_type")
                if not is_blank_value(spec_type):
                    clean["f_Spec_Type"] = specimen_map.get(
                        clean_text_value(spec_type).lower()
                    )

                batch_code = clean_text_value(
                    get_value(row, "f_Batch_Code", "Batch_Code", "batch_code", "f_Batch_Name", "Batch_Name")
                )
                if batch_code:
                    batch_obj = Batch_Table.objects.filter(
                        bat_Batch_Code=batch_code
                    ).first()
                    if batch_obj:
                        clean["f_Batch_id"] = batch_obj
                        touched_batch_ids.add(batch_obj.id)
                        clean["f_Batch_Code"] = batch_obj.bat_Batch_Code
                        clean["f_Batch_Name"] = batch_obj.bat_Batch_Name
                        clean["f_RefNo"] = (
                            clean.get("f_RefNo")
                            or batch_obj.bat_RefNo
                            or ref_no_from_batch_code(batch_obj.bat_Batch_Code)
                            or ""
                        )
                        clean["f_BatchNo"] = clean.get("f_BatchNo") or batch_obj.bat_BatchNo
                        clean["f_Total_batch"] = clean.get("f_Total_batch") or batch_obj.bat_Total_batch
                        clean["f_Referral_Date"] = clean.get("f_Referral_Date") or batch_obj.bat_Referral_Date

                for field in final_text_fields:
                    if field in clean:
                        clean[field] = clean_text_value(clean[field])

                final_obj = existing.get(accession)
                if final_obj:
                    if not overwrite:
                        skipped += 1
                        continue
                    for field_name, value in clean.items():
                        setattr(final_obj, field_name, value)
                    to_update.append(final_obj)
                else:
                    to_create.append(
                        Final_Data(f_AccessionNo=accession, **clean)
                    )

            update_fields = sorted({
                field_name
                for field_name in final_fields
                if field_name not in {"id", "f_AccessionNo"}
            })

            from apps.home_final.signals import suspend_import_signals

            with suspend_import_signals():
                if to_create:
                    Final_Data.objects.bulk_create(to_create, batch_size=1000)
                if to_update:
                    Final_Data.objects.bulk_update(
                        to_update,
                        update_fields,
                        batch_size=1000,
                    )

            touched_final_ids = list(
                Final_Data.objects
                .filter(f_AccessionNo__in=[
                    obj.f_AccessionNo for obj in to_create + to_update
                ])
                .values_list("id", flat=True)
            )
            refreshed_final_breakpoints = reapply_final_breakpoints_for_batches(
                list(touched_batch_ids)
            )
            refreshed_emerging = refresh_emerging_for_final_ids(touched_final_ids)
            regenerated_concordance = regenerate_concordance_for_final_ids(
                touched_final_ids,
                user=request.user,
            )
            refreshed_tat = refresh_tat_for_batches(list(touched_batch_ids))

            messages.success(
                request,
                f"Final demographic upload complete! "
                f"{len(to_create)} created, {len(to_update)} updated, {skipped} skipped. "
                f"{refreshed_final_breakpoints} final antibiotic breakpoint record(s) refreshed. "
                f"{refreshed_emerging} emerging record(s) refreshed. "
                f"{regenerated_concordance} concordance accession report(s) refreshed. "
                f"{refreshed_tat} TAT record(s) refreshed."
            )

            return redirect("show_final_data")

        # ================= FIELD LIST =================
        referred_fields = {
            f.name for f in Referred_Data._meta.fields
            if not f.auto_created
        }
        referred_field_map = {
            f.name: f for f in Referred_Data._meta.fields
            if not f.auto_created
        }
        text_fields = {
            name for name, field in referred_field_map.items()
            if isinstance(field, (models.CharField, models.TextField))
        }
        boolean_fields = {
            name for name, field in referred_field_map.items()
            if isinstance(field, models.BooleanField)
        }

        INT_FIELDS = {
            "Age",
            "bat_seq"
        }

        # ================= LOOKUPS =================
        specimen_map = {
            str(s.Specimen_code).strip().lower(): s
            for s in SpecimenTypeModel.objects.all()
        }

        site_map = {
            str(s.SiteCode).strip().upper(): s.SiteName
            for s in SiteData.objects.all()
        }

        # ================= EXISTING CACHE =================
        existing = {}

        to_create = []
        to_update = []
        uploaded_batch_codes = set()
        skipped = 0

        def normalize_ref_no(value, accession):
            inferred = infer_ref_from_accession(accession)

            if is_blank_value(value):
                return inferred

            raw_ref = clean_text_value(value)

            try:
                numeric_ref = float(raw_ref)
                if numeric_ref.is_integer():
                    raw_ref = str(int(numeric_ref))
            except ValueError:
                pass

            if not raw_ref.isdigit():
                return inferred

            if inferred and inferred.isdigit():
                return raw_ref.zfill(len(inferred))

            return raw_ref

        def infer_site_code_from_accession(accession):
            match = re.search(r"ARS_([A-Z]+)\d+$", accession or "", re.I)
            return match.group(1).upper() if match else ""

        def infer_ref_from_accession(accession):
            match = re.search(r"(\d+)$", accession or "")
            return match.group(1) if match else ""

        def extract_ref_range_from_batch_name(batch_name):
            if is_blank_value(batch_name):
                return ""

            match = re.search(
                r"_(\d{3,6}(?:-\d{3,6})?)\s*$",
                clean_text_value(batch_name)
            )
            return match.group(1) if match else ""

        def extract_batch_sequence(batch_code):
            if is_blank_value(batch_code):
                return "", ""

            match = re.search(
                r"_(\d+)\.(\d+)_\d{3,6}(?:-\d{3,6})?\s*$",
                clean_text_value(batch_code)
            )
            if not match:
                return "", ""

            return match.group(1), match.group(2)

        def parse_year_as_date(value):
            if is_blank_value(value):
                return None

            if isinstance(value, datetime):
                return date(value.year, 1, 1)

            if isinstance(value, date):
                return date(value.year, 1, 1)

            raw_value = clean_text_value(value)

            try:
                year = int(float(raw_value))
                if 1900 <= year <= 2100:
                    return date(year, 1, 1)
            except ValueError:
                pass

            parsed = parse_date(raw_value)
            return date(parsed.year, 1, 1) if parsed else None

        normalized_rows = []
        batch_groups = {}

        for row in rows:

            row = {str(k).strip(): v for k, v in row.items()}

            # Allow a final-data mapped workbook to be uploaded here without
            # failing on missing raw field names. Most final fields are the raw
            # field name with an "f_" prefix.
            for key, value in list(row.items()):
                if not key.startswith("f_"):
                    continue
                raw_key = key[2:]
                if raw_key in referred_fields and is_blank_value(row.get(raw_key)):
                    row[raw_key] = value

            accession = clean_text_value(
                get_value(row, "AccessionNo", "accession_no", "f_AccessionNo")
            ).upper()

            if not accession:
                continue

            site_code = clean_text_value(
                get_value(row, "SiteCode", "site_code", "f_SiteCode")
            ).upper()
            if not site_code:
                site_code = infer_site_code_from_accession(accession)

            referral_date = parse_date(
                get_value(row, "Referral_Date", "referral_date", "f_Referral_Date")
            )

            explicit_ref_no = clean_text_value(
                get_value(row, "RefNo", "ref_no", "f_RefNo")
            )
            inferred_ref_no = normalize_ref_no(explicit_ref_no, accession)
            explicit_batch_code = clean_text_value(
                get_value(
                    row,
                    "Batch_Code",
                    "batch_code",
                    "f_Batch_Code",
                    "Batch_Name",
                    "batch_name",
                    "f_Batch_Name",
                )
            )
            code_batch_no, code_total_batch = extract_batch_sequence(
                explicit_batch_code
            )
            batch_no = clean_text_value(
                get_value(row, "BatchNo", "batch_no", "f_BatchNo")
            ) or code_batch_no or "1"
            total_batch = clean_text_value(
                get_value(row, "Total_batch", "total_batch", "f_Total_batch")
            ) or code_total_batch or "1"

            group_key = (
                explicit_batch_code,
                site_code,
                referral_date,
                batch_no,
                total_batch,
            )

            normalized = {
                "row": row,
                "accession": accession,
                "site_code": site_code,
                "referral_date": referral_date,
                "batch_no": batch_no,
                "total_batch": total_batch,
                "ref_no": inferred_ref_no,
                "explicit_batch_code": explicit_batch_code,
                "group_key": group_key,
            }
            normalized_rows.append(normalized)
            batch_groups.setdefault(group_key, []).append(normalized)

        batch_codes_by_group = {}

        for group_key, group_rows in batch_groups.items():
            explicit_batch_code, site_code, referral_date, batch_no, total_batch = group_key

            if explicit_batch_code:
                ref_no = (
                    extract_ref_range_from_batch_name(explicit_batch_code)
                    or group_rows[0]["ref_no"]
                )
                batch_code = explicit_batch_code

            else:
                refs = [
                    row["ref_no"]
                    for row in group_rows
                    if row["ref_no"] and row["ref_no"].isdigit()
                ]

                if refs:
                    width = max(len(ref) for ref in refs)
                    ref_numbers = sorted(int(ref) for ref in refs)
                    start_ref = str(ref_numbers[0]).zfill(width)
                    end_ref = str(ref_numbers[-1]).zfill(width)
                    ref_no = (
                        start_ref
                        if start_ref == end_ref
                        else f"{start_ref}-{end_ref}"
                    )
                else:
                    ref_no = ""

                if not (site_code and referral_date and ref_no):
                    missing = []
                    if not site_code:
                        missing.append("SiteCode")
                    if not referral_date:
                        missing.append("Referral_Date")
                    if not ref_no:
                        missing.append("AccessionNo or RefNo")
                    raise ValueError(
                        "Cannot generate batch code. Missing: "
                        f"{', '.join(missing)}. Please map SiteCode, "
                        "Referral_Date, and AccessionNo or RefNo."
                    )

                batch_code = (
                    f"{site_code}_{referral_date.strftime('%m%d%Y')}_"
                    f"{batch_no}.{total_batch}_{ref_no}"
                )

            for row in group_rows:
                row["batch_code"] = batch_code
                row["batch_ref_no"] = ref_no
                row["batch_name"] = batch_code

        uploaded_accessions = {
            row["accession"]
            for row in normalized_rows
        }
        existing = {
            obj.AccessionNo: obj
            for obj in Referred_Data.objects.filter(
                AccessionNo__in=uploaded_accessions
            )
        }

        # ================= PROCESS ROWS =================
        for normalized in normalized_rows:

            row = normalized["row"]
            accession = normalized["accession"]
            site_code = normalized["site_code"]

            site_name = site_map.get(site_code, "")

            referral_date = normalized["referral_date"]

            birth_date = parse_date(
                get_value(row, "Date_Birth", "date_birth")
            )

            admis_date = parse_date(
                get_value(row, "Date_Admis", "date_admis")
            )

            spec_date = parse_date(
                get_value(row, "Spec_Date", "spec_date")
            )

            default_year = parse_year_as_date(
                get_value(row, "Default_Year", "default_year", "f_Default_Year")
            )

            batch_code = normalized["batch_code"]

            clean = {
                k: v for k, v in row.items()
                if k in referred_fields
            }

            for field in text_fields:
                if field in clean:
                    clean[field] = clean_text_value(clean[field])

            for field in boolean_fields:
                if field in clean:
                    clean[field] = parse_bool_value(clean[field])

            clean["AccessionNo"] = accession
            clean["Referral_Date"] = referral_date
            clean["Date_Birth"] = birth_date
            clean["Date_Admis"] = admis_date
            clean["Spec_Date"] = spec_date
            clean["Default_Year"] = default_year or (
                date(spec_date.year, 1, 1) if spec_date else None
            )
            clean["SiteCode"] = site_code
            clean["Site_Name"] = site_name
            clean["Batch_Code"] = batch_code
            clean["Batch_Name"] = normalized["batch_name"]
            clean["RefNo"] = normalized["batch_ref_no"]
            clean["BatchNo"] = normalized["batch_no"]
            clean["Total_batch"] = normalized["total_batch"]

            # integer cleanup
            for field in INT_FIELDS:
                if field in clean:
                    clean[field] = parse_int(clean[field])

            # specimen FK
            if "Spec_Type" in clean:

                raw_val = clean["Spec_Type"]

                if raw_val:
                    clean["Spec_Type"] = specimen_map.get(
                        str(raw_val).strip().lower()
                    )

            # ================= CREATE / UPDATE =================
            if accession in existing:
                if not overwrite:
                    skipped += 1
                    continue

                obj = existing[accession]

                for k, v in clean.items():
                    setattr(obj, k, v)

                to_update.append(obj)
                uploaded_batch_codes.add(batch_code)

            else:

                to_create.append(
                    Referred_Data(**clean)
                )
                uploaded_batch_codes.add(batch_code)

        # ================= REMOVE DUPLICATES =================
        to_update = list({o.AccessionNo: o for o in to_update}.values())

        # ================= SAVE =================
        if to_create:
            Referred_Data.objects.bulk_create(to_create, batch_size=1000)

        if to_update:
            Referred_Data.objects.bulk_update(
                to_update,
                [f for f in referred_fields if f != "id"],
                batch_size=1000,
            )

        created = len(to_create)
        updated = len(to_update)

        # ================= GENERATE BATCH =================
        created_batches = generate_batches_from_referred()
        batch_ids = list(
            Batch_Table.objects
            .filter(bat_Batch_Code__in=uploaded_batch_codes)
            .values_list("id", flat=True)
        )
        refreshed_breakpoints = reapply_raw_breakpoints_for_batches(batch_ids)
        messages.success(
            request,
            f"Upload complete! {created} created, {updated} updated, {skipped} skipped. "
            f"{created_batches} batches generated. "
            f"{refreshed_breakpoints} antibiotic breakpoint record(s) refreshed. "
            "Final data, emerging, concordance, and TAT were not refreshed."
        )

        return redirect("show_batches")

    except Exception as e:

        import traceback
        traceback.print_exc()

        messages.error(request, f"Upload failed: {e}")
        return redirect("upload_wgs_view")



########## end of uploading view

@login_required
def show_final_data(request):
    finaldata_summaries = Final_Data.objects.all().order_by("f_Referral_Date")  # optional ordering

    total_records = Final_Data.objects.count()
     # Paginate the queryset to display 20 records per page
    paginator = Paginator(finaldata_summaries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Render the template with paginated data
    return render(
        request,
        "wgs_app/show_final_data.html",
        {"page_obj": page_obj,
         "total_records": total_records,
         },  # only send page_obj
    )



@login_required
def show_referred_data(request):
    rawdata_summaries = Referred_Data.objects.all().order_by("Referral_Date")  # optional ordering

    total_records = Referred_Data.objects.count()
     # Paginate the queryset to display 20 records per page
    paginator = Paginator(rawdata_summaries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Render the template with paginated data
    return render(
        request,
        "wgs_app/show_referred_data.html",
        {"page_obj": page_obj,
         "total_records": total_records,
         },  # only send page_obj
    )





@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@require_POST
@transaction.atomic
def delete_final_data(request, pk):

    final_item = get_object_or_404(Final_Data, pk=pk)
    if not can_manage_batch(request.user, final_item.f_Batch_id):
        messages.error(request, "You can only delete final records from batches that you created.")
        return redirect('show_final_data')

    if request.method == "POST":
        accession = final_item.f_AccessionNo
        final_item.delete()   #  This triggers the signal

        messages.success(
            request,
            f"Record {accession} deleted. Concordance report invalidated."
        )

        return redirect('show_final_data')

    messages.error(request, "Invalid request for deleting Final data.")
    return redirect('show_final_data')



@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@require_POST
@transaction.atomic
def delete_referred_data(request, pk):

    referred_item = get_object_or_404(Referred_Data, pk=pk)
    if not can_manage_batch(request.user, referred_item.Batch_id):
        messages.error(request, "You can only delete referred records from batches that you created.")
        return redirect('show_referred_data')

    if request.method == "POST":
        accession = referred_item.AccessionNo
        referred_item.delete()   #  This triggers the signal

        messages.success(
            request,
            f"Record {accession} deleted. Concordance report invalidated."
        )

        return redirect('show_referred_data')

    messages.error(request, "Invalid request for deleting Final data.")
    return redirect('show_referred_data')




@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER)
@require_POST
@transaction.atomic
def delete_all_final_data(request):
    Final_Data.objects.all().delete()
    messages.success(request, "Final Referred Isolates have been deleted successfully.")
    return redirect('show_final_data')  # Redirect to the table view



@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER)
@require_POST
def delete_all_referred_data(request):
    Referred_Data.objects.all().delete()
    messages.success(request, "Referred Isolates have been deleted successfully.")
    return redirect('show_referred_data')  # Redirect to the table view



@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER)
@require_POST
@transaction.atomic
def delete_finaldata_by_date(request):
    if request.method == "POST":
        upload_date_str = request.POST.get("upload_date")
        print(" Received upload_date_str:", upload_date_str)

        if not upload_date_str:
            messages.error(request, "Please select an upload date to delete.")
            return redirect("show_final_data")

        # Use Django’s date parser
        upload_date = parse_date(upload_date_str)

        if not upload_date:
            messages.error(request, f"Invalid date format: {upload_date_str}")
            return redirect("show_final_data")

        deleted_count, _ = Final_Data.objects.filter(f_Date_Modified__date=upload_date).delete()
        messages.success(request, f" Deleted {deleted_count} Final Isolates records uploaded on {upload_date}.")
        return redirect("show_final_data")

    messages.error(request, "Invalid request method.")
    return redirect("show_final_data")


@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER)
@require_POST
def delete_referreddata_by_date(request):
    if request.method == "POST":
        upload_date_str = request.POST.get("upload_date")
        print(" Received upload_date_str:", upload_date_str)

        if not upload_date_str:
            messages.error(request, "Please select an upload date to delete.")
            return redirect("show_referred_data")

        # Use Django’s date parser
        upload_date = parse_date(upload_date_str)

        if not upload_date:
            messages.error(request, f"Invalid date format: {upload_date_str}")
            return redirect("show_referred_data")

        deleted_count, _ = Referred_Data.objects.filter(Date_Modified__date=upload_date).delete()
        messages.success(request, f" Deleted {deleted_count} Final Isolates records uploaded on {upload_date}.")
        return redirect("show_referred_data")

    messages.error(request, "Invalid request method.")
    return redirect("show_referred_data")



############# Uploading of Final Antibiotic Entries


########### HELPERS ################
# ─────────────────────────────────────────────────────────────────────────────
# COLUMN PATTERNS  (all retest / ARSRL results)
#   <abx>_nd<n>        → retest disk value
#   <abx>_nd<n>_ris    → retest disk RIS
#   <abx>_nm           → retest MIC value
#   <abx>_nm_ris       → retest MIC RIS
# ─────────────────────────────────────────────────────────────────────────────

DISK_PAT     = re.compile(r'^([a-z0-9]{2,6})_(nd\d+(?:_\d+)?)$')
DISK_RIS_PAT = re.compile(r'^([a-z0-9]{2,6})_(nd\d+(?:_\d+)?)_ris$')
MIC_PAT      = re.compile(r'^([a-z0-9]{2,6})_(nm)$')
MIC_RIS_PAT  = re.compile(r'^([a-z0-9]{2,6})_(nm)_ris$')

# Excel col → f_ars_* (ARSRL phenotype)
PHENO_FIELD_MAP = {
    'ampc':      'f_ars_ampC',
    'esbl':      'f_ars_ESBL',
    'carb':      'f_ars_CARB',
    'mbl':       'f_ars_MBL',
    'bl':        'f_ars_BL',
    'mr':        'f_ars_MR',
    'meca':      'f_ars_mecA',
    'icr':       'f_ars_ICR',
    'ecim':      'f_ars_ECIM',
    'mcim':      'f_ars_MCIM',
    'ecim.mcim': 'f_ars_EC_MCIM',
}

# Excel col → f_* (site phenotype)
SITE_PHENO_MAP = {
    'ampc': 'f_ampC',
    'esbl': 'f_ESBL',
    'carb': 'f_CARB',
    'mbl':  'f_MBL',
    'bl':   'f_BL',
    'mr':   'f_MR',
    'meca': 'f_mecA',
    'icr':  'f_ICR',
}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _clean(val, default=""):
    if val is None:
        return default
    try:
        if pd.isna(val):
            return default
    except TypeError:
        pass
    s = str(val).strip()
    return s if s.lower() not in ("none", "nan", "nat", "null", "") else default


def _date(val):
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
    except Exception:
        return None


def _int(val):
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except TypeError:
        pass
    try:
        return int(float(str(val).strip()))
    except Exception:
        return None


def _decimal(val):

    if val is None:
        return None, ""
    try:
        if pd.isna(val):
            return None, ""
    except TypeError:
        pass

    s = str(val).strip()
    operand = ""

    if s.lower() in {"", "nan", "nat", "none", "null"}:
        return None, ""

    for op in ("<=", ">=", "<", ">"):
        if s.startswith(op):
            operand = op
            s = s[len(op):]
            break

    try:
        value = Decimal(s)
        if not value.is_finite():
            return None, operand
        return value, operand
    except (InvalidOperation, ValueError):
        return None, ""
    

def _pheno(val):
    s = _clean(val)
    return s if s in ("(+)", "(-)", "NT") else "n/a"


# ─────────────────────────────────────────────────────────────────────────────
# ANTIBIOTIC COLUMN DETECTION
# ─────────────────────────────────────────────────────────────────────────────

def _parse_abx_columns(headers):

    groups = {}
    markers = {"rt", "val", "value", "ris", "op", "operand", "enris"}

    def normalize_col(col):
        text = str(col or "").strip().lower()
        text = re.sub(r"[\s\-.]+", "_", text)
        text = re.sub(r"_+", "_", text)
        return text.strip("_")

    def column_info(col):
        parts = normalize_col(col).split("_")
        if len(parts) < 2:
            return None

        method = parts[1]
        if not (method.startswith("nd") or method == "nm"):
            return None

        is_retest = "rt" in parts[2:]
        field = ""
        if "ris" in parts[2:]:
            field = "ris"
        elif "op" in parts[2:] or "operand" in parts[2:]:
            field = "op"
        elif "val" in parts[2:] or "value" in parts[2:]:
            field = "val"
        else:
            field = "val"

        method_tokens = [method]
        for token in parts[2:]:
            if token in markers:
                break
            method_tokens.append(token)

        code = f"{parts[0].upper()}_{'_'.join(method_tokens).upper()}"
        group_key = f"{code}_RT" if is_retest else code

        return {
            "code": code,
            "group_key": group_key,
            "field": field,
            "is_disk": method.startswith("nd"),
            "is_mic": method == "nm",
            "is_retest": is_retest,
        }

    for idx, col in enumerate(headers):
        info = column_info(col)
        if not info:
            continue

        g = groups.setdefault(info["group_key"], {
            "disk_col": None,
            "disk_ris_col": None,
            "mic_col": None,
            "mic_ris_col": None,
            "mic_op_col": None,
            "disk_code": None,
            "mic_code": None,
        })

        code = f"{info['code']}_RT" if info["is_retest"] else info["code"]

        if info["is_disk"] and info["field"] == "ris":
            g["disk_ris_col"] = idx
            g["disk_code"] = code
        elif info["is_disk"]:
            g["disk_col"] = idx
            g["disk_code"] = code
        elif info["is_mic"] and info["field"] == "op":
            g["mic_op_col"] = idx
            g["mic_code"] = code
        elif info["is_mic"] and info["field"] == "ris":
            g["mic_ris_col"] = idx
            g["mic_code"] = code
        elif info["is_mic"]:
            g["mic_col"] = idx
            g["mic_code"] = code

    return groups


def _antibiotic_lookup_key(code):
    text = str(code or "").strip().upper()
    text = re.sub(r"[^A-Z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def _add_antibiotic_aliases(antibiotics, antibiotic):
    for code in (antibiotic.Abx_code, antibiotic.Whonet_Abx):
        raw_code = str(code or "").strip().upper()
        if not raw_code:
            continue
        antibiotics[raw_code] = antibiotic
        antibiotics[_antibiotic_lookup_key(raw_code)] = antibiotic


def _resolve_antibiotic(antibiotics, *codes):
    for code in codes:
        raw_code = str(code or "").strip().upper()
        if not raw_code:
            continue
        antibiotic = antibiotics.get(raw_code) or antibiotics.get(_antibiotic_lookup_key(raw_code))
        if antibiotic:
            return antibiotic
    return None


def _canonical_whonet_code(antibiotic, fallback):
    return (antibiotic.Whonet_Abx or fallback or "").strip().upper()



def _empty_group(abx):
    return {
        'abx_code':     abx,
        'disk_label':   None,
        'disk_col':     None,
        'disk_ris_col': None,
        'mic_col':      None,
        'mic_ris_col':  None,
    }




def _resolve_effective_year(spec_date):
    """
    Returns the nearest breakpoint year relative to the specimen year.

    Rules:
    1. Use closest breakpoint year <= specimen year
    2. If specimen year is earlier than all breakpoints → use earliest
    3. If specimen year is later than all breakpoints → use latest
    """

    years = list(
        BreakpointsTable.objects
        .values_list("Year", flat=True)
        .distinct()
    )

    if not years:
        return None

    years = sorted(int(y) for y in years)

    if spec_date:
        specimen_year = spec_date.year

        # find closest <= specimen year
        eligible = [y for y in years if y <= specimen_year]

        if eligible:
            return str(max(eligible))

        # specimen earlier than all breakpoints
        return str(years[0])

    # if specimen date missing → use newest
    return str(years[-1])


def _year_as_int(value):
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _find_best_breakpoint(abx_code, year, method, org_code):
    if not abx_code or not method:
        return None

    abx_code = str(abx_code).strip().upper()
    org_code = (org_code or "").strip()
    target_year = _year_as_int(year)

    candidates = list(
        BreakpointsTable.objects
        .filter(
            Q(Antibiotic_list_id=abx_code) | Q(Whonet_Abx__iexact=abx_code),
            Test_Method__iexact=method,
        )
        .filter(
            Q(Org__iexact=org_code) |
            Q(Org__isnull=True) |
            Q(Org="")
        )
    )

    if not candidates:
        return None

    by_year = {}
    for bp in candidates:
        bp_year = _year_as_int(bp.Year)
        if bp_year is not None:
            by_year.setdefault(bp_year, []).append(bp)

    if not by_year:
        return None

    if target_year is None:
        selected_year = max(by_year)
    elif target_year in by_year:
        selected_year = target_year
    else:
        previous_years = [candidate_year for candidate_year in by_year if candidate_year <= target_year]
        selected_year = max(previous_years) if previous_years else min(by_year)

    return sorted(
        by_year[selected_year],
        key=lambda bp: (
            0 if (bp.Org or "").strip().lower() == org_code.lower() else 1,
            0 if not (bp.Spec_code or "").strip() else 1,
        ),
    )[0]


def _build_breakpoint_resolver():
    breakpoints = list(BreakpointsTable.objects.all())
    years = sorted({
        year
        for year in (_year_as_int(bp.Year) for bp in breakpoints)
        if year is not None
    })
    by_code_method = {}

    for bp in breakpoints:
        method = (bp.Test_Method or "").strip().upper()
        codes = {
            (bp.Whonet_Abx or "").strip().upper(),
            (bp.Antibiotic_list_id or "").strip().upper(),
        }
        for code in codes - {""}:
            by_code_method.setdefault((code, method), []).append(bp)

    def effective_year(spec_date):
        if not years:
            return None
        if not spec_date:
            return years[-1]
        eligible = [year for year in years if year <= spec_date.year]
        return max(eligible) if eligible else years[0]

    def resolve(abx_code, spec_date, method, org_code):
        code = (abx_code or "").strip().upper()
        method = (method or "").strip().upper()
        org_code = (org_code or "").strip()
        target_year = effective_year(spec_date)
        candidates = [
            bp
            for bp in by_code_method.get((code, method), [])
            if not (bp.Org or "").strip()
            or (bp.Org or "").strip().lower() == org_code.lower()
        ]
        by_year = {}
        for bp in candidates:
            bp_year = _year_as_int(bp.Year)
            if bp_year is not None:
                by_year.setdefault(bp_year, []).append(bp)
        if not by_year:
            return None
        if target_year in by_year:
            selected_year = target_year
        else:
            previous = [year for year in by_year if year <= target_year]
            selected_year = max(previous) if previous else min(by_year)
        return sorted(
            by_year[selected_year],
            key=lambda bp: (
                0
                if (bp.Org or "").strip().lower() == org_code.lower()
                else 1,
                0 if not (bp.Spec_code or "").strip() else 1,
            ),
        )[0]

    return effective_year, resolve




############################
# BREAKPOINT HELPERS
############################

def _apply_bp_to_entry(entry, bp, is_disk, alert_mic=False, set_relation=True):

    if set_relation:
        entry.ab_breakpoints_id.set([bp])

    entry.ab_Site_Org = bp.Org
    entry.ab_Org_Flag = bool(bp.Emerging_Org_Flag)
    entry.ab_Abx_Flag = bool(bp.Emerging_Abx_Flag)
    entry.ab_Abx_Phenotype = bp.Emerging_Pheno_Flag or ""

    entry.ab_R_breakpoint   = bp.R_val
    entry.ab_I_breakpoint   = bp.I_val
    entry.ab_SDD_breakpoint = bp.SDD_val
    entry.ab_S_breakpoint   = bp.S_val

    if not is_disk:
        entry.ab_Alert_val = bp.Alert_val if alert_mic else ""


def _apply_bp_to_retest_entry(entry, bp, is_disk, alert_mic=False, set_relation=True):

    if set_relation:
        entry.ab_breakpoints_id.set([bp])

    entry.ab_Ret_Org = bp.Org
    entry.ab_Org_Flag = bool(bp.Emerging_Org_Flag)
    entry.ab_Abx_Flag = bool(bp.Emerging_Abx_Flag)
    entry.ab_Abx_Phenotype = bp.Emerging_Pheno_Flag or ""

    entry.ab_Ret_R_breakpoint   = bp.R_val
    entry.ab_Ret_I_breakpoint   = bp.I_val
    entry.ab_Ret_SDD_breakpoint = bp.SDD_val
    entry.ab_Ret_S_breakpoint   = bp.S_val

    if not is_disk:
        entry.ab_Retest_Alert_val = bp.Alert_val if alert_mic else ""


############################
# CLEAR BREAKPOINT FIELDS
############################

def _clear_bp_fields(entry):

    entry.ab_Site_Org = None
    entry.ab_Org_Flag = False
    entry.ab_Abx_Flag = False
    entry.ab_Abx_Phenotype = ""

    entry.ab_R_breakpoint   = None
    entry.ab_I_breakpoint   = None
    entry.ab_SDD_breakpoint = None
    entry.ab_S_breakpoint   = None

    entry.ab_Alert_val = ""
    entry.ab_Disk_RIS = ""
    entry.ab_MIC_RIS = ""


def _clear_retest_bp_fields(entry):

    entry.ab_Ret_Org = None
    entry.ab_Org_Flag = False
    entry.ab_Abx_Flag = False
    entry.ab_Abx_Phenotype = ""

    entry.ab_Ret_R_breakpoint   = None
    entry.ab_Ret_I_breakpoint   = None
    entry.ab_Ret_SDD_breakpoint = None
    entry.ab_Ret_S_breakpoint   = None

    entry.ab_Retest_Alert_val = ""
    entry.ab_Retest_Disk_RIS = ""
    entry.ab_Retest_MIC_RIS = ""


def _bulk_upsert_antibiotic_entries(entry_model, pending, overwrite=True):
    """
    Persist the final state of parsed antibiotic cells in batches.

    pending maps (mode, isolate_fk_value, code) to either:
      - None, meaning delete the matching entry
      - {"isolate": obj, "defaults": {...}}, meaning create/update
    """
    if not pending:
        return 0, 0, 0

    is_final = entry_model is Final_AntibioticEntry
    fk_name = "ab_idNum_f_referred" if is_final else "ab_idNum_referred"
    fk_id_name = f"{fk_name}_id"
    isolate_fk_values = {key[1] for key in pending}

    existing_entries = list(
        entry_model.objects.filter(**{f"{fk_id_name}__in": isolate_fk_values})
    )
    existing_main = {}
    existing_retest = {}
    for entry in existing_entries:
        main_code = (entry.ab_Abx_code or "").strip().upper()
        retest_code = (entry.ab_Retest_Abx_code or "").strip().upper()
        if main_code:
            existing_main.setdefault((getattr(entry, fk_id_name), main_code), entry)
        if retest_code:
            existing_retest.setdefault((getattr(entry, fk_id_name), retest_code), entry)

    to_create = []
    to_update = {}
    delete_ids = set()
    update_fields = set()

    for (mode, isolate_fk_value, code), operation in pending.items():
        lookup = existing_retest if mode == "retest" else existing_main
        entry = lookup.get((isolate_fk_value, code))

        if operation is None:
            if overwrite and entry and entry.pk:
                delete_ids.add(entry.pk)
            continue

        defaults = operation["defaults"]
        for ris_field in (
            "ab_Disk_enRIS",
            "ab_MIC_enRIS",
            "ab_Retest_Disk_enRIS",
            "ab_Retest_MIC_enRIS",
        ):
            if ris_field in defaults and defaults[ris_field]:
                defaults[ris_field] = str(defaults[ris_field]).upper()

        if entry is None:
            entry = entry_model(**{fk_id_name: isolate_fk_value}, **defaults)
            to_create.append(entry)
            lookup[(isolate_fk_value, code)] = entry
        elif not overwrite:
            continue
        else:
            for field_name, value in defaults.items():
                setattr(entry, field_name, value)
            to_update[entry.pk] = entry
            update_fields.update(defaults)

    if delete_ids:
        entry_model.objects.filter(pk__in=delete_ids).delete()

    if to_create:
        entry_model.objects.bulk_create(to_create, batch_size=1000)

    if to_update and update_fields:
        entry_model.objects.bulk_update(
            list(to_update.values()),
            sorted(update_fields),
            batch_size=1000,
        )

    return len(to_create), len(to_update), len(delete_ids)


def _bulk_replace_breakpoint_links(entry_model, entry_breakpoints):
    if not entry_breakpoints:
        return

    m2m_field = entry_model._meta.get_field("ab_breakpoints_id")
    through = m2m_field.remote_field.through
    source_field = m2m_field.m2m_field_name()
    target_field = m2m_field.m2m_reverse_field_name()
    entry_ids = list(entry_breakpoints)

    through.objects.filter(
        **{f"{source_field}_id__in": entry_ids}
    ).delete()

    links = [
        through(
            **{
                f"{source_field}_id": entry_id,
                f"{target_field}_id": breakpoint_id,
            }
        )
        for entry_id, breakpoint_ids in entry_breakpoints.items()
        for breakpoint_id in dict.fromkeys(breakpoint_ids)
    ]
    if links:
        through.objects.bulk_create(links, batch_size=2000)


def _recalculate_antibiotic_ris(entry, determine_ris):
    entry.ab_Disk_RIS = determine_ris(
        entry.ab_Disk_value,
        entry.ab_R_breakpoint,
        entry.ab_I_breakpoint,
        entry.ab_S_breakpoint,
        entry.ab_SDD_breakpoint,
        is_disk=True,
    )
    entry.ab_MIC_RIS = determine_ris(
        entry.ab_MIC_value,
        entry.ab_R_breakpoint,
        entry.ab_I_breakpoint,
        entry.ab_S_breakpoint,
        entry.ab_SDD_breakpoint,
    )
    entry.ab_Retest_Disk_RIS = determine_ris(
        entry.ab_Retest_DiskValue,
        entry.ab_Ret_R_breakpoint,
        entry.ab_Ret_I_breakpoint,
        entry.ab_Ret_S_breakpoint,
        entry.ab_Ret_SDD_breakpoint,
        is_disk=True,
    )
    entry.ab_Retest_MIC_RIS = determine_ris(
        entry.ab_Retest_MICValue,
        entry.ab_Ret_R_breakpoint,
        entry.ab_Ret_I_breakpoint,
        entry.ab_Ret_S_breakpoint,
        entry.ab_Ret_SDD_breakpoint,
    )


############################
# SAVE SITE ANTIBIOTIC
############################

def _save_site_entry(
    isolate,
    abx_code,
    disk_int,
    disk_ris,
    mic_value,
    mic_operand,
    mic_ris,
    resolved_org,
    effective_year,
    abx_obj=None
):

    entry, _ = AntibioticEntry.objects.update_or_create(

        ab_idNum_referred=isolate,
        ab_Abx_code=abx_code.upper(),

        defaults={

            "ab_AccessionNo": isolate.AccessionNo,
            "ab_RefNo": isolate.RefNo,

            "ab_Antibiotic": abx_obj.Antibiotic if abx_obj else abx_code,
            "ab_Abx": abx_obj.Abx_code if abx_obj else abx_code,

            "ab_Disk_value": disk_int,
            "ab_Disk_enRIS": disk_ris,

            "ab_MIC_value": mic_value,
            "ab_MIC_operand": mic_operand,
            "ab_MIC_enRIS": mic_ris,
        }
    )

    entry.ab_breakpoints_id.clear()

    bp_applied = False

    if disk_int is not None:

        bp = _find_best_breakpoint(abx_code, effective_year, "DISK", resolved_org)

        if bp:
            _apply_bp_to_entry(entry, bp, True)
            bp_applied = True

    if mic_value is not None:

        bp = _find_best_breakpoint(abx_code, effective_year, "MIC", resolved_org)

        if bp:
            _apply_bp_to_entry(entry, bp, False)
            bp_applied = True

    if not bp_applied:
        _clear_bp_fields(entry)

    entry.save()

    return entry


def _save_final_site_entry(
    isolate,
    abx_code,
    disk_int,
    disk_ris,
    mic_value,
    mic_operand,
    mic_ris,
    resolved_org,
    effective_year,
    abx_obj=None
):

    entry, created = Final_AntibioticEntry.objects.update_or_create(

        ab_idNum_f_referred=isolate,
        ab_Abx_code=abx_code.upper(),

        defaults={

            "ab_AccessionNo": isolate.f_AccessionNo,
            "ab_RefNo": isolate.f_RefNo,

            "ab_Antibiotic": abx_obj.Antibiotic if abx_obj else abx_code,
            "ab_Abx": abx_obj.Abx_code if abx_obj else abx_code,

            "ab_Disk_value": disk_int,
            "ab_Disk_enRIS": disk_ris,

            "ab_MIC_value": mic_value,
            "ab_MIC_operand": mic_operand,
            "ab_MIC_enRIS": mic_ris,
        }
    )

    entry.ab_breakpoints_id.clear()

    bp_applied = False

    if disk_int is not None:

        bp = _find_best_breakpoint(abx_code, effective_year, "DISK", resolved_org)

        if bp:
            _apply_bp_to_entry(entry, bp, True)
            bp_applied = True

    if mic_value is not None:

        bp = _find_best_breakpoint(abx_code, effective_year, "MIC", resolved_org)

        if bp:
            _apply_bp_to_entry(entry, bp, False, alert_mic=True)
            bp_applied = True

    if not bp_applied:
        _clear_bp_fields(entry)

    entry.save()

    return entry, created


############################
# SAVE RETEST ANTIBIOTIC
############################

def _save_retest_entry(
    entry_model,
    isolate,
    abx_code,
    disk_int,
    disk_ris,
    mic_value,
    mic_operand,
    mic_ris,
    resolved_ars_org,
    effective_year,
    abx_obj=None
):

    # --------------------------------
    # Resolve accession + refno
    # --------------------------------
    accession = getattr(isolate, "AccessionNo", None) or getattr(isolate, "f_AccessionNo", "")
    refno = getattr(isolate, "RefNo", None) or getattr(isolate, "f_RefNo", "")

    # --------------------------------
    # Determine correct FK field
    # --------------------------------
    fk_field = (
        "ab_idNum_f_referred"
        if entry_model.__name__ == "Final_AntibioticEntry"
        else "ab_idNum_referred"
    )

    filter_kwargs = {
        fk_field: isolate,
        "ab_Retest_Abx_code": abx_code.upper()
    }

    # --------------------------------
    # Create / Update entry
    # --------------------------------
    entry, created = entry_model.objects.update_or_create(

        **filter_kwargs,

        defaults={

            "ab_AccessionNo": accession,
            "ab_RefNo": refno,

            "ab_Retest_Antibiotic": abx_obj.Antibiotic if abx_obj else abx_code,
            "ab_Retest_Abx": abx_obj.Abx_code if abx_obj else abx_code,

            "ab_Retest_DiskValue": disk_int,
            "ab_Retest_Disk_enRIS": disk_ris,

            "ab_Retest_MICValue": mic_value,
            "ab_Retest_MIC_operand": mic_operand,
            "ab_Retest_MIC_enRIS": mic_ris,
        }
    )

    # --------------------------------
    # Clear breakpoints
    # --------------------------------
    entry.ab_breakpoints_id.clear()

    bp_applied = False

    # --------------------------------
    # DISK BREAKPOINT
    # --------------------------------
    if disk_int is not None:

        bp = _find_best_breakpoint(abx_code, effective_year, "DISK", resolved_ars_org)

        if bp:
            _apply_bp_to_retest_entry(entry, bp, True)
            bp_applied = True

    # --------------------------------
    # MIC BREAKPOINT
    # --------------------------------
    if mic_value is not None:

        bp = _find_best_breakpoint(abx_code, effective_year, "MIC", resolved_ars_org)

        if bp:
            _apply_bp_to_retest_entry(entry, bp, False, alert_mic=True)
            bp_applied = True

    # --------------------------------
    # No breakpoint found
    # --------------------------------
    if not bp_applied:
        _clear_retest_bp_fields(entry)

    entry.save()

    return entry, created

########## END OF HELPERS ################
# @login_required
# @require_POST
# @transaction.atomic
# def upload_final_antibiotics(request):

#     file = request.FILES.get("FinalAntibioticFile")

#     if not file:
#         messages.error(request, "No file uploaded")
#         return redirect("show_final_antibiotic")

#     try:
#         wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
#     except Exception as e:
#         messages.error(request, str(e))
#         return redirect("show_final_antibiotic")

#     created = 0
#     updated = 0
#     skipped = 0

#     # -----------------------------
#     # LOAD REFERENCE DATA
#     # -----------------------------

#     breakpoint_keys = set(
#         (x or "").strip().upper()
#         for x in BreakpointsTable.objects.values_list("Whonet_Abx", flat=True)
#     )

#     antibiotics = {
#         (a.Abx_code or "").strip().upper(): a
#         for a in Antibiotic_List.objects.filter(Retest=True)
#     }

#     def normalize_accession(val):
#         return str(val).strip().replace(" ", "").upper()

#     isolates = {
#         normalize_accession(i.f_AccessionNo): i
#         for i in Final_Data.objects.all()
#     }
#     # -----------------------------
#     # PROCESS WORKBOOK
#     # -----------------------------

#     for sheet in wb.sheetnames:

#         ws = wb[sheet]
#         rows = ws.iter_rows(values_only=True)

#         headers = next(rows)
#         headers_lower = [str(c).lower().strip() if c else "" for c in headers]

#         if "f_accessionno" not in headers_lower:
#             continue

#         accession_idx = headers_lower.index("f_accessionno")

#         abx_groups = {
#             k: v for k, v in _parse_abx_columns(headers).items()
#             if v["disk_col"] is not None or v["mic_col"] is not None
#         }

#         for row in rows:

#             accession = _clean(row[accession_idx]).upper()

#             if not accession:
#                 continue

#             isolate = isolates.get(accession)

#             if not isolate:
#                 skipped += 1
#                 continue

#             effective_year = _resolve_effective_year(isolate.f_Spec_Date)
#             resolved_ars_org = (isolate.f_ars_OrgCode or "").strip()

#             for abx_code, grp in abx_groups.items():

#                 abx_code = abx_code.upper()

#                 # resolve antibiotic object
#                 abx_obj = antibiotics.get(abx_code)

#                 if not abx_obj:
#                     continue

#                 disk_key = headers[grp["disk_col"]].upper() if grp["disk_col"] is not None else None
#                 mic_key = headers[grp["mic_col"]].upper() if grp["mic_col"] is not None else None

#                 # remove accidental _OP
#                 if mic_key and mic_key.endswith("_OP"):
#                     mic_key = mic_key.replace("_OP", "")

#                 # ----------------------------
#                 # EXCEL VALUES
#                 # ----------------------------

#                 disk_raw = row[grp["disk_col"]] if grp["disk_col"] is not None else None
#                 disk_ris = _clean(row[grp["disk_ris_col"]]) if grp["disk_ris_col"] is not None else ""

#                 mic_raw = row[grp["mic_col"]] if grp["mic_col"] is not None else None
#                 mic_ris = _clean(row[grp["mic_ris_col"]]) if grp["mic_ris_col"] is not None else ""

#                 mic_operand = ""
#                 if grp["mic_op_col"] is not None:
#                     mic_operand = _clean(row[grp["mic_op_col"]])

#                 disk_val = _int(disk_raw)
#                 mic_val, parsed_operand = _decimal(mic_raw)

#                 if not mic_operand:
#                     mic_operand = parsed_operand

#                 # ----------------------------
#                 # DELETE IF EMPTY
#                 # --------------------  --------

#                 if disk_val is None and mic_val is None:

#                     Final_AntibioticEntry.objects.filter(
#                         ab_idNum_f_referred=isolate,
#                         ab_Retest_Abx_code=abx_code
#                     ).delete()

#                     continue

#                 # ----------------------------
#                 # CREATE / UPDATE ENTRY
#                 # ----------------------------
#                 whonet_code = disk_key if disk_val is not None else mic_key

#                 entry, created_flag = Final_AntibioticEntry.objects.update_or_create(

#                     ab_idNum_f_referred=isolate,
#                     ab_Retest_Abx_code=whonet_code,

#                     defaults={

#                         "ab_AccessionNo": isolate.f_AccessionNo,
#                         "ab_RefNo": isolate.f_RefNo,

#                         "ab_Retest_Antibiotic": abx_obj.Antibiotic,
#                         "ab_Retest_Abx": abx_obj.Abx_code,
                        
#                         "ab_Retest_DiskValue": disk_val,
#                         "ab_Retest_MICValue": mic_val,

#                         "ab_Retest_Disk_enRIS": disk_ris,
#                         "ab_Retest_MIC_enRIS": mic_ris,
#                         "ab_Retest_MIC_operand": mic_operand,
#                     }
#                 )

#                 entry.ab_breakpoints_id.clear()

#                 ret_bp_applied = False

#                 # ----------------------------
#                 # DISK BREAKPOINT
#                 # ----------------------------

#                 if disk_val is not None and disk_key and disk_key in breakpoint_keys:

#                     bp_disk = BreakpointsTable.objects.filter(
#                         Whonet_Abx__iexact=disk_key,
#                         Year=effective_year,
#                         Test_Method="DISK",
#                         Org__in=[resolved_ars_org, ""]
#                     ).order_by("-Org").first()

#                     if bp_disk:

#                         entry.ab_breakpoints_id.set([bp_disk])

#                         entry.ab_Ret_Org = bp_disk.Org
#                         entry.ab_Org_Flag = bool(bp_disk.Emerging_Org_Flag)
#                         entry.ab_Abx_Flag = bool(bp_disk.Emerging_Abx_Flag)
#                         entry.ab_Abx_Phenotype = bp_disk.Emerging_Pheno_Flag or ""

#                         entry.ab_Ret_R_breakpoint = bp_disk.R_val
#                         entry.ab_Ret_I_breakpoint = bp_disk.I_val
#                         entry.ab_Ret_SDD_breakpoint = bp_disk.SDD_val
#                         entry.ab_Ret_S_breakpoint = bp_disk.S_val

#                         ret_bp_applied = True

#                 # ----------------------------
#                 # MIC BREAKPOINT
#                 # ----------------------------

#                 if mic_val is not None and mic_key and mic_key in breakpoint_keys:

#                     bp_mic = BreakpointsTable.objects.filter(
#                         Whonet_Abx__iexact=mic_key,
#                         Year=effective_year,
#                         Test_Method="MIC",
#                         Org__in=[resolved_ars_org, ""]
#                     ).order_by("-Org").first()

#                     if bp_mic:

#                         entry.ab_breakpoints_id.set([bp_mic])

#                         entry.ab_Ret_Org = bp_mic.Org
#                         entry.ab_Org_Flag = bool(bp_mic.Emerging_Org_Flag)
#                         entry.ab_Abx_Flag = bool(bp_mic.Emerging_Abx_Flag)
#                         entry.ab_Abx_Phenotype = bp_mic.Emerging_Pheno_Flag or ""

#                         entry.ab_Ret_R_breakpoint = bp_mic.R_val
#                         entry.ab_Ret_I_breakpoint = bp_mic.I_val
#                         entry.ab_Ret_SDD_breakpoint = bp_mic.SDD_val
#                         entry.ab_Ret_S_breakpoint = bp_mic.S_val

#                         entry.ab_Retest_Alert_val = bp_mic.Alert_val

#                         ret_bp_applied = True

#                 # ----------------------------
#                 # NO BREAKPOINT
#                 # ----------------------------

#                 if not ret_bp_applied:

#                     entry.ab_Ret_Org = None
#                     entry.ab_Org_Flag = False
#                     entry.ab_Abx_Flag = False
#                     entry.ab_Abx_Phenotype = ""

#                     entry.ab_Ret_R_breakpoint = None
#                     entry.ab_Ret_I_breakpoint = None
#                     entry.ab_Ret_SDD_breakpoint = None
#                     entry.ab_Ret_S_breakpoint = None

#                     entry.ab_Retest_Alert_val = ""

#                 entry.save()

#                 if created_flag:
#                     created += 1
#                 else:
#                     updated += 1

#     wb.close()

#     messages.success(
#         request,
#         f"{created} antibiotics uploaded successfully. "
#         f"{updated} updated, {skipped} skipped."
#     )

#     return redirect("show_final_antibiotic")


########### implements sentinel site & ars antitbioc fields upload
@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@require_POST
@transaction.atomic
def upload_final_antibiotics(request):
    from apps.home_final.signals import suspend_import_signals

    file = request.FILES.get("FinalAntibioticFile")
    overwrite = request.POST.get("overwrite", "false").lower() == "true"
    partial_update = request.POST.get("partial_update", "false").lower() == "true"
    if partial_update:
        overwrite = True

    if not file:
        messages.error(request, "No file uploaded")
        return redirect("show_final_antibiotic")

    print(f"[FINAL ABX DEBUG] Upload started: {file.name}")

    try:
        file_name = file.name.lower()
        if file_name.endswith(".csv"):
            sheets = {"Final_AntibioticEntry": pd.read_csv(file)}
        elif file_name.endswith((".xlsx", ".xls")):
            sheets = pd.read_excel(file, sheet_name=None)
        else:
            messages.error(request, "Unsupported file format. Please upload CSV, XLSX, or XLS.")
            return redirect("show_final_antibiotic")
    except Exception as e:
        print(f"[FINAL ABX DEBUG] Failed to read uploaded file: {e}")
        messages.error(request, str(e))
        return redirect("show_final_antibiotic")

    created = 0
    updated = 0
    skipped = 0
    affected_batch_ids = set()
    pending_entries = {}

    # -----------------------------
    # LOAD REFERENCE DATA
    # -----------------------------

    antibiotics = {}
    for antibiotic in Antibiotic_List.objects.all():
        _add_antibiotic_aliases(antibiotics, antibiotic)


    def normalize_accession(val):
        return str(val).strip().replace(" ", "").upper()

    def normalize_header(val):
        return re.sub(r"[^a-z0-9]", "", str(val).lower())

    isolates = {
        normalize_accession(i.f_AccessionNo): i
        for i in Final_Data.objects.all()
    }
    print(f"[FINAL ABX DEBUG] Loaded {len(isolates)} Final_Data accession(s).")

    # -----------------------------
    # PROCESS WORKBOOK
    # -----------------------------

    processed_sheets = 0
    processed_abx_sheets = 0
    matched_accessions = 0

    for sheet_name, df in sheets.items():

        df.columns = [str(c).strip() for c in df.columns]
        headers = list(df.columns)
        print(
            f"[FINAL ABX DEBUG] Sheet '{sheet_name}': "
            f"{len(df)} row(s), {len(headers)} column(s)."
        )
        header_keys = [normalize_header(c) for c in headers]
        accession_candidates = {"faccessionno", "accessionno", "faccessionnumber"}
        accession_idx = next(
            (
                idx for idx, key in enumerate(header_keys)
                if key in accession_candidates
            ),
            None
        )

        if accession_idx is None:
            print(
                f"[FINAL ABX DEBUG] Sheet '{sheet_name}' skipped: "
                "no f_AccessionNo/AccessionNo column."
            )
            continue

        processed_sheets += 1
        print(
            f"[FINAL ABX DEBUG] Sheet '{sheet_name}' accession column: "
            f"{headers[accession_idx]}"
        )

        abx_groups = {
            k: v for k, v in _parse_abx_columns(headers).items()
            if v["disk_col"] is not None or v["mic_col"] is not None
        }

        if not abx_groups:
            print(
                f"[FINAL ABX DEBUG] Sheet '{sheet_name}' skipped: "
                "no antibiotic value columns detected."
            )
            continue

        processed_abx_sheets += 1
        print(
            f"[FINAL ABX DEBUG] Sheet '{sheet_name}' detected "
            f"{len(abx_groups)} antibiotic group(s): "
            f"{', '.join(sorted(abx_groups.keys())[:20])}"
            f"{'...' if len(abx_groups) > 20 else ''}"
        )

        for row in df.itertuples(index=False, name=None):

            accession = normalize_accession(row[accession_idx])

            if not accession:
                continue

            isolate = isolates.get(accession)

            if not isolate:
                skipped += 1
                print(f"[FINAL ABX DEBUG] Accession skipped: {accession} not found in Final_Data.")
                continue

            matched_accessions += 1

            if isolate.f_Batch_id_id:
                affected_batch_ids.add(isolate.f_Batch_id_id)

            for abx_code, grp in abx_groups.items():

                # ----------------------------
                # EXCEL VALUES
                # ----------------------------

                disk_raw = row[grp["disk_col"]] if grp["disk_col"] is not None else None
                disk_ris = _clean(row[grp["disk_ris_col"]]) if grp["disk_ris_col"] is not None else ""

                mic_raw = row[grp["mic_col"]] if grp["mic_col"] is not None else None
                mic_ris = _clean(row[grp["mic_ris_col"]]) if grp["mic_ris_col"] is not None else ""

                mic_operand = ""
                if grp["mic_op_col"] is not None:
                    mic_operand = _clean(row[grp["mic_op_col"]])

                disk_val = _int(disk_raw)
                mic_val, parsed_operand = _decimal(mic_raw)

                if not mic_operand:
                    mic_operand = parsed_operand

                disk_key = grp.get("disk_code")
                mic_key = grp.get("mic_code")

                if mic_key and mic_key.endswith("_OP"):
                    mic_key = mic_key.replace("_OP", "")

                is_retest = False

                if disk_key and "_RT" in disk_key:
                    is_retest = True

                if mic_key and "_RT" in mic_key:
                    is_retest = True

                if disk_key:
                    disk_key = disk_key.replace("_RT", "")

                if mic_key:
                    mic_key = mic_key.replace("_RT", "")

                whonet_code = disk_key if disk_key else mic_key

                if not whonet_code:
                    continue

                abx_obj = _resolve_antibiotic(antibiotics, whonet_code, abx_code)

                if not abx_obj:
                    skipped += 1
                    continue

                resolved_whonet_code = _canonical_whonet_code(abx_obj, whonet_code)

                # ----------------------------
                # DELETE IF EMPTY
                # ----------------------------

                if disk_val is None and mic_val is None:
                    pending_entries[
                        (
                            "retest" if is_retest else "main",
                            isolate.pk,
                            resolved_whonet_code,
                        )
                    ] = None
                    continue

                # ----------------------------
                # SAVE USING HELPER
                # ----------------------------
                whonet_code = disk_key if disk_val is not None else mic_key
                resolved_whonet_code = _canonical_whonet_code(abx_obj, whonet_code)

                if is_retest:
                    defaults = {
                        "ab_AccessionNo": isolate.f_AccessionNo,
                        "ab_RefNo": isolate.f_RefNo,
                        "ab_Retest_Antibiotic": abx_obj.Antibiotic,
                        "ab_Retest_Abx_code": resolved_whonet_code,
                        "ab_Retest_Abx": abx_obj.Abx_code,
                        "ab_Retest_DiskValue": disk_val,
                        "ab_Retest_Disk_enRIS": disk_ris,
                        "ab_Retest_MICValue": mic_val,
                        "ab_Retest_MIC_operand": mic_operand,
                        "ab_Retest_MIC_enRIS": mic_ris,
                    }
                else:
                    defaults = {
                        "ab_AccessionNo": isolate.f_AccessionNo,
                        "ab_RefNo": isolate.f_RefNo,
                        "ab_Antibiotic": abx_obj.Antibiotic,
                        "ab_Abx_code": resolved_whonet_code,
                        "ab_Abx": abx_obj.Abx_code,
                        "ab_Disk_value": disk_val,
                        "ab_Disk_enRIS": disk_ris,
                        "ab_MIC_value": mic_val,
                        "ab_MIC_operand": mic_operand,
                        "ab_MIC_enRIS": mic_ris,
                    }

                pending_entries[
                        (
                            "retest" if is_retest else "main",
                            isolate.pk,
                            resolved_whonet_code,
                        )
                ] = {
                    "isolate": isolate,
                    "defaults": defaults,
                }

    with suspend_import_signals():
        created, updated, _ = _bulk_upsert_antibiotic_entries(
            Final_AntibioticEntry,
            pending_entries,
            overwrite=overwrite,
        )

    refreshed_breakpoints = 0
    refreshed_emerging = 0
    regenerated_concordance = 0
    refreshed_tat = 0

    if affected_batch_ids:
        with suspend_import_signals():
            refreshed_breakpoints = reapply_final_breakpoints_for_batches(
                affected_batch_ids,
                debug=True
            )
        refreshed_emerging = refresh_emerging_for_batches(affected_batch_ids)
        regenerated_concordance = regenerate_batch_concordance(
            affected_batch_ids,
            user=request.user
        )
        refreshed_tat = refresh_tat_for_batches(affected_batch_ids)

    print(
        "[FINAL ABX DEBUG] Upload complete: "
        f"created={created}, updated={updated}, skipped={skipped}, "
        f"matched_accessions={matched_accessions}, "
        f"affected_batches={len(affected_batch_ids)}, "
        f"breakpoints={refreshed_breakpoints}, emerging={refreshed_emerging}, "
        f"concordance={regenerated_concordance}, tat={refreshed_tat}."
    )

    if processed_sheets == 0:
        messages.error(
            request,
            "No final antibiotic sheet was processed. Include an accession column "
            "named f_AccessionNo or AccessionNo."
        )
        return redirect("show_final_antibiotic")

    if processed_abx_sheets == 0:
        messages.error(
            request,
            "No final antibiotic columns were found. Use mapped columns like "
            "AMP_ND10_Val / AMP_ND10_RIS or retest columns like "
            "AMP_ND10_RT_Val / AMP_ND10_RT_RIS."
        )
        return redirect("show_final_antibiotic")

    if matched_accessions == 0:
        messages.error(
            request,
            "No uploaded accessions matched Final Data. Upload/copy demographics "
            "to Final first, or check the accession column values."
        )
        return redirect("show_final_antibiotic")

    if created == 0 and updated == 0:
        messages.warning(
            request,
            "Final antibiotic upload found matching accessions, but no antibiotic "
            "values were saved. Check that the mapped antibiotic value columns are not blank."
        )
        return redirect("show_final_antibiotic")

    messages.success(
        request,
        f"{created} antibiotics uploaded successfully. "
        f"{updated} updated, {skipped} skipped. "
        f"{refreshed_breakpoints} final breakpoint record(s) refreshed. "
        f"{refreshed_emerging} emerging record(s) refreshed. "
        f"{regenerated_concordance} concordance batch report(s) refreshed. "
        f"{refreshed_tat} TAT record(s) refreshed."
    )

    return redirect("show_final_antibiotic")



# @login_required
# @require_POST
# @transaction.atomic
# def upload_raw_antibiotics(request):

#     file = request.FILES.get("RawAntibioticFile")

#     if not file:
#         messages.error(request, "No file uploaded")
#         return redirect("show_raw_antibiotic")

#     try:
#         wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
#     except Exception as e:
#         messages.error(request, str(e))
#         return redirect("show_raw_antibiotic")

#     created = 0
#     updated = 0
#     skipped = 0

#     # -----------------------------
#     # LOAD REFERENCE DATA
#     # -----------------------------

#     breakpoint_keys = set(
#         (x or "").strip().upper()
#         for x in BreakpointsTable.objects.values_list("Whonet_Abx", flat=True)
#     )

#     antibiotics = {
#         (a.Abx_code or "").strip().upper(): a
#         for a in Antibiotic_List.objects.filter(Show=True)
#     }

#     def normalize_accession(val):
#         return str(val).strip().replace(" ", "").upper()

#     isolates = {
#         normalize_accession(i.AccessionNo): i
#         for i in Referred_Data.objects.all()
#     }
#     # -----------------------------
#     # PROCESS WORKBOOK
#     # -----------------------------

#     for sheet in wb.sheetnames:

#         ws = wb[sheet]
#         rows = ws.iter_rows(values_only=True)

#         headers = next(rows)
#         headers_lower = [str(c).lower().strip() if c else "" for c in headers]

#         if "accessionno" not in headers_lower:
#             continue

#         accession_idx = headers_lower.index("accessionno")

#         abx_groups = {
#             k: v for k, v in _parse_abx_columns(headers).items()
#             if v["disk_col"] is not None or v["mic_col"] is not None
#         }

#         for row in rows:

#             accession = _clean(row[accession_idx]).upper()

#             if not accession:
#                 continue

#             isolate = isolates.get(accession)

#             if not isolate:
#                 skipped += 1
#                 continue

#             effective_year = _resolve_effective_year(isolate.Spec_Date)
#             resolved_ars_org = (isolate.ars_OrgCode or "").strip()

#             for abx_code, grp in abx_groups.items():

#                 abx_code = abx_code.upper()

#                 # resolve antibiotic object
#                 abx_obj = antibiotics.get(abx_code)

#                 if not abx_obj:
#                     continue

#                 disk_key = headers[grp["disk_col"]].upper() if grp["disk_col"] is not None else None
#                 mic_key = headers[grp["mic_col"]].upper() if grp["mic_col"] is not None else None

#                 # remove accidental _OP
#                 if mic_key and mic_key.endswith("_OP"):
#                     mic_key = mic_key.replace("_OP", "")

#                 # ----------------------------
#                 # EXCEL VALUES
#                 # ----------------------------

#                 disk_raw = row[grp["disk_col"]] if grp["disk_col"] is not None else None
#                 disk_ris = _clean(row[grp["disk_ris_col"]]) if grp["disk_ris_col"] is not None else ""

#                 mic_raw = row[grp["mic_col"]] if grp["mic_col"] is not None else None
#                 mic_ris = _clean(row[grp["mic_ris_col"]]) if grp["mic_ris_col"] is not None else ""

#                 mic_operand = ""
#                 if grp["mic_op_col"] is not None:
#                     mic_operand = _clean(row[grp["mic_op_col"]])

#                 disk_val = _int(disk_raw)
#                 mic_val, parsed_operand = _decimal(mic_raw)

#                 if not mic_operand:
#                     mic_operand = parsed_operand

#                 # ----------------------------
#                 # DELETE IF EMPTY
#                 # --------------------  --------

#                 if disk_val is None and mic_val is None:

#                     AntibioticEntry.objects.filter(
#                         ab_idNum_referred=isolate,
#                         ab_Retest_Abx_code=abx_code
#                     ).delete()

#                     continue

#                 # ----------------------------
#                 # CREATE / UPDATE ENTRY
#                 # ----------------------------
#                 whonet_code = disk_key if disk_val is not None else mic_key

#                 entry, created_flag = AntibioticEntry.objects.update_or_create(

#                     ab_idNum_referred=isolate,
#                     ab_Abx_code=whonet_code,

#                     defaults={

#                         "ab_AccessionNo": isolate.AccessionNo,
#                         "ab_RefNo": isolate.RefNo,

#                         "ab_Antibiotic": abx_obj.Antibiotic,
#                         "ab_Abx_code": abx_obj.Abx_code,
                        
#                         "ab_Disk_value": disk_val,
#                         "ab_MIC_value": mic_val,

#                         "ab_Disk_enRIS": disk_ris,
#                         "ab_MIC_enRIS": mic_ris,
#                         "ab_MIC_operand": mic_operand,
#                     }
#                 )

#                 entry.ab_breakpoints_id.clear()

#                 bp_applied = False

#                 # ----------------------------
#                 # DISK BREAKPOINT
#                 # ----------------------------

#                 if disk_val is not None and disk_key and disk_key in breakpoint_keys:

#                     bp_disk = BreakpointsTable.objects.filter(
#                         Whonet_Abx__iexact=disk_key,
#                         Year=effective_year,
#                         Test_Method="DISK",
#                         Org__in=[resolved_ars_org, ""]
#                     ).order_by("-Org").first()

#                     if bp_disk:

#                         entry.ab_breakpoints_id.set([bp_disk])

#                         entry.ab_Site_Org = bp_disk.Org
#                         entry.ab_Org_Flag = bool(bp_disk.Emerging_Org_Flag)
#                         entry.ab_Abx_Flag = bool(bp_disk.Emerging_Abx_Flag)
#                         entry.ab_Abx_Phenotype = bp_disk.Emerging_Pheno_Flag or ""

#                         entry.ab_R_breakpoint = bp_disk.R_val
#                         entry.ab_I_breakpoint = bp_disk.I_val
#                         entry.ab_SDD_breakpoint = bp_disk.SDD_val
#                         entry.ab_S_breakpoint = bp_disk.S_val

#                         bp_applied = True

#                 # ----------------------------
#                 # MIC BREAKPOINT
#                 # ----------------------------

#                 if mic_val is not None and mic_key and mic_key in breakpoint_keys:

#                     bp_mic = BreakpointsTable.objects.filter(
#                         Whonet_Abx__iexact=mic_key,
#                         Year=effective_year,
#                         Test_Method="MIC",
#                         Org__in=[resolved_ars_org, ""]
#                     ).order_by("-Org").first()

#                     if bp_mic:

#                         entry.ab_breakpoints_id.set([bp_mic])

#                         entry.ab_Site_Org = bp_mic.Org
#                         entry.ab_Org_Flag = bool(bp_mic.Emerging_Org_Flag)
#                         entry.ab_Abx_Flag = bool(bp_mic.Emerging_Abx_Flag)
#                         entry.ab_Abx_Phenotype = bp_mic.Emerging_Pheno_Flag or ""

#                         entry.ab_R_breakpoint = bp_mic.R_val
#                         entry.ab_I_breakpoint = bp_mic.I_val
#                         entry.ab_SDD_breakpoint = bp_mic.SDD_val
#                         entry.ab_S_breakpoint = bp_mic.S_val

#                         entry.ab_Alert_val = bp_mic.Alert_val

#                         bp_applied = True

#                 # ----------------------------
#                 # NO BREAKPOINT
#                 # ----------------------------

#                 if not bp_applied:

#                     entry.ab_Site_Org = None
#                     entry.ab_Org_Flag = False
#                     entry.ab_Abx_Flag = False
#                     entry.ab_Abx_Phenotype = ""

#                     entry.ab_R_breakpoint = None
#                     entry.ab_I_breakpoint = None
#                     entry.ab_SDD_breakpoint = None
#                     entry.ab_S_breakpoint = None

#                     entry.ab_Alert_val = ""

#                 entry.save()

#                 if created_flag:
#                     created += 1
#                 else:
#                     updated += 1

#     wb.close()

#     messages.success(
#         request,
#         f"{created} antibiotics uploaded successfully. "
#         f"{updated} updated, {skipped} skipped."
#     )

#     return redirect("show_raw_antibiotic")


############## implements uploading on both sentinel and ars antibiotic fields

@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@require_POST
@transaction.atomic
def upload_raw_antibiotics(request):

    file = request.FILES.get("RawAntibioticFile")
    overwrite = request.POST.get("overwrite", "false").lower() == "true"

    if not file:
        messages.error(request, "No file uploaded")
        return redirect("show_raw_antibiotic")

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        messages.error(request, str(e))
        return redirect("show_raw_antibiotic")

    created = 0
    updated = 0
    skipped = 0
    affected_batch_ids = set()
    pending_entries = {}

    # -----------------------------
    # LOAD REFERENCE DATA
    # -----------------------------

    antibiotics = {}
    for antibiotic in Antibiotic_List.objects.all():
        _add_antibiotic_aliases(antibiotics, antibiotic)

    def normalize_accession(val):
        return str(val).strip().replace(" ", "").upper()

    isolates = {
        normalize_accession(i.AccessionNo): i
        for i in Referred_Data.objects.all()
    }

    # -----------------------------
    # PROCESS WORKBOOK
    # -----------------------------

    for sheet in wb.sheetnames:

        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)

        headers = next(rows)
        headers_lower = [str(c).lower().strip() if c else "" for c in headers]

        if "accessionno" not in headers_lower:
            continue

        accession_idx = headers_lower.index("accessionno")

        abx_groups = {
            k: v for k, v in _parse_abx_columns(headers).items()
            if v["disk_col"] is not None or v["mic_col"] is not None
        }

        for row in rows:

            accession = normalize_accession(row[accession_idx])

            if not accession:
                continue

            isolate = isolates.get(accession)

            if not isolate:
                skipped += 1
                continue

            if isolate.Batch_id_id:
                affected_batch_ids.add(isolate.Batch_id_id)

            for abx_code, grp in abx_groups.items():

                abx_code = abx_code.upper()

                disk_key = grp.get("disk_code")
                mic_key = grp.get("mic_code")

                if mic_key and mic_key.endswith("_OP"):
                    mic_key = mic_key.replace("_OP", "")

                # ----------------------------
                # RT DETECTION
                # ----------------------------

                is_retest = False

                if disk_key and "_RT" in disk_key:
                    is_retest = True

                if mic_key and "_RT" in mic_key:
                    is_retest = True

                if disk_key:
                    disk_key = disk_key.replace("_RT", "")

                if mic_key:
                    mic_key = mic_key.replace("_RT", "")

                # ----------------------------
                # EXCEL VALUES
                # ----------------------------

                disk_raw = row[grp["disk_col"]] if grp["disk_col"] is not None else None
                disk_ris = _clean(row[grp["disk_ris_col"]]) if grp["disk_ris_col"] is not None else ""

                mic_raw = row[grp["mic_col"]] if grp["mic_col"] is not None else None
                mic_ris = _clean(row[grp["mic_ris_col"]]) if grp["mic_ris_col"] is not None else ""

                mic_operand = ""
                if grp["mic_op_col"] is not None:
                    mic_operand = _clean(row[grp["mic_op_col"]])

                disk_val = _int(disk_raw)
                mic_val, parsed_operand = _decimal(mic_raw)

                if not mic_operand:
                    mic_operand = parsed_operand

                whonet_code = disk_key if disk_val is not None else mic_key

                if not whonet_code:
                    continue

                abx_obj = _resolve_antibiotic(antibiotics, whonet_code, abx_code)

                if not abx_obj:
                    skipped += 1
                    continue

                resolved_whonet_code = _canonical_whonet_code(abx_obj, whonet_code)

                # ----------------------------
                # DELETE IF EMPTY
                # ----------------------------

                if disk_val is None and mic_val is None:
                    pending_entries[
                        (
                            "retest" if is_retest else "main",
                            isolate.AccessionNo,
                            resolved_whonet_code,
                        )
                    ] = None
                    continue

                # ----------------------------
                # CREATE / UPDATE ENTRY
                # ----------------------------

                if is_retest:
                    defaults = {
                        "ab_AccessionNo": isolate.AccessionNo,
                        "ab_RefNo": isolate.RefNo,
                        "ab_Retest_Antibiotic": abx_obj.Antibiotic,
                        "ab_Retest_Abx_code": resolved_whonet_code,
                        "ab_Retest_Abx": abx_obj.Abx_code,
                        "ab_Retest_DiskValue": disk_val,
                        "ab_Retest_MICValue": mic_val,
                        "ab_Retest_Disk_enRIS": disk_ris,
                        "ab_Retest_MIC_enRIS": mic_ris,
                        "ab_Retest_MIC_operand": mic_operand,
                    }
                else:
                    defaults = {
                        "ab_AccessionNo": isolate.AccessionNo,
                        "ab_RefNo": isolate.RefNo,
                        "ab_Antibiotic": abx_obj.Antibiotic,
                        "ab_Abx_code": resolved_whonet_code,
                        "ab_Abx": abx_obj.Abx_code,
                        "ab_Disk_value": disk_val,
                        "ab_MIC_value": mic_val,
                        "ab_Disk_enRIS": disk_ris,
                        "ab_MIC_enRIS": mic_ris,
                        "ab_MIC_operand": mic_operand,
                    }

                pending_entries[
                        (
                            "retest" if is_retest else "main",
                            isolate.AccessionNo,
                            resolved_whonet_code,
                        )
                ] = {
                    "isolate": isolate,
                    "defaults": defaults,
                }

    wb.close()

    created, updated, _ = _bulk_upsert_antibiotic_entries(
        AntibioticEntry,
        pending_entries,
        overwrite=overwrite,
    )

    refreshed_breakpoints = 0

    if affected_batch_ids:
        refreshed_breakpoints = reapply_raw_breakpoints_for_batches(
            affected_batch_ids
        )

    messages.success(
        request,
        f"{created} antibiotics uploaded successfully. "
        f"{updated} updated, {skipped} skipped. "
        f"{refreshed_breakpoints} raw breakpoint record(s) refreshed. "
        "Final antibiotic data was not changed."
    )

    return redirect("show_raw_antibiotic")

















# updated version
@login_required
def show_final_antibiotic(request):

    entries = (
        Final_AntibioticEntry.objects
        .select_related("ab_idNum_f_referred")
        .order_by("ab_idNum_f_referred__f_AccessionNo")
    )

    abx_data = {}
    abx_columns = set()

    for entry in entries:

        # Skip invalid entries
        if not entry.ab_Abx_code or not entry.ab_idNum_f_referred:
            continue

        acc = entry.ab_idNum_f_referred.f_AccessionNo
        abx_code = entry.ab_Abx_code.strip().upper()

        # Column naming (value / operand / RIS)
        col_value = abx_code
        col_op    = f"{abx_code}_OP"
        col_ris   = f"{abx_code}_RIS"

        abx_columns.update([col_value, col_op, col_ris])

        if acc not in abx_data:
            abx_data[acc] = {}

        # Determine MIC vs DISK based on actual data
        if entry.ab_MIC_value is not None:
            abx_data[acc][col_value] = entry.ab_MIC_value or ""
            abx_data[acc][col_op]    = entry.ab_MIC_operand or ""
            abx_data[acc][col_ris]   = entry.ab_MIC_enRIS or ""
        else:
            abx_data[acc][col_value] = entry.ab_Disk_value or ""
            abx_data[acc][col_op]    = ""
            abx_data[acc][col_ris]   = entry.ab_Disk_enRIS or ""

    # Sort antibiotic columns consistently
    abx_columns = sorted(abx_columns)

    # Pagination (per accession)
    paginated_list = list(abx_data.items())
    paginator = Paginator(paginated_list, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "wgs_app/show_final_antibiotic.html",
        {
            "page_obj": page_obj,
            "abx_data": dict(page_obj.object_list),
            "abx_codes": abx_columns,
            "total_records": len(abx_data),  # number of isolates
        }
    )



# updated version
@login_required
def show_raw_antibiotic(request):

    entries = (
        AntibioticEntry.objects
        .select_related("ab_idNum_referred")
        .order_by("ab_idNum_referred__AccessionNo")
    )

    abx_data = {}
    abx_columns = set()

    for entry in entries:

        # Skip invalid entries
        if not entry.ab_Abx_code or not entry.ab_idNum_referred:
            continue

        acc = entry.ab_idNum_referred.AccessionNo
        abx_code = entry.ab_Abx_code.strip().upper()

        # Column naming (value / operand / RIS)
        col_value = abx_code
        col_op    = f"{abx_code}_OP"
        col_ris   = f"{abx_code}_RIS"

        abx_columns.update([col_value, col_op, col_ris])

        if acc not in abx_data:
            abx_data[acc] = {}

        # Determine MIC vs DISK based on actual data
        if entry.ab_MIC_value is not None:
            abx_data[acc][col_value] = entry.ab_MIC_value or ""
            abx_data[acc][col_op]    = entry.ab_MIC_operand or ""
            abx_data[acc][col_ris]   = entry.ab_MIC_enRIS or ""
        else:
            abx_data[acc][col_value] = entry.ab_Disk_value or ""
            abx_data[acc][col_op]    = ""
            abx_data[acc][col_ris]   = entry.ab_Disk_enRIS or ""

    # Sort antibiotic columns consistently
    abx_columns = sorted(abx_columns)

    # Pagination (per accession)
    paginated_list = list(abx_data.items())
    paginator = Paginator(paginated_list, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "wgs_app/show_raw_antibiotic.html",
        {
            "page_obj": page_obj,
            "abx_data": dict(page_obj.object_list),
            "abx_codes": abx_columns,
            "total_records": len(abx_data),  # number of isolates
        }
    )






### updated version
@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@require_POST
def delete_final_antibiotic(request, pk):

    target = get_object_or_404(Final_AntibioticEntry, pk=pk)
    acc = target.ab_idNum_f_referred.f_AccessionNo

    if request.method == "POST":
        Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred__f_AccessionNo=acc
        ).delete()

        messages.success(
            request,
            f"All final antibiotic records for accession {acc} deleted successfully!"
        )
        return redirect("show_final_antibiotic")

    messages.error(request, "Invalid request method.")
    return redirect("show_final_antibiotic")



@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@require_POST
def delete_raw_antibiotic(request, pk):

    target = get_object_or_404(AntibioticEntry, pk=pk)
    acc = target.ab_idNum_referred.AccessionNo

    if request.method == "POST":
        AntibioticEntry.objects.filter(
            ab_idNum_referred__AccessionNo=acc
        ).delete()

        messages.success(
            request,
            f"All antibiotic records for accession {acc} deleted successfully!"
        )
        return redirect("show_raw_antibiotic")

    messages.error(request, "Invalid request method.")
    return redirect("show_raw_antibiotic")





@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER)
@require_POST
def delete_all_final_antibiotic(request):
    Final_AntibioticEntry.objects.all().delete()
    messages.success(request, "Final antibiotic entries have been deleted successfully.")
    return redirect('show_final_antibiotic')  # Redirect to the table view




@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER)
@require_POST
def delete_all_raw_antibiotic(request):
    AntibioticEntry.objects.all().delete()
    messages.success(request, "Raw antibiotic entries have been deleted successfully.")
    return redirect('show_raw_antibiotic')  # Redirect to the table view





#### updated version
@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER)
@require_POST
def delete_finalantibiotic_by_date(request):

    if request.method == "POST":
        upload_date_str = request.POST.get("upload_date")

        if not upload_date_str:
            messages.error(request, "Please select an upload date to delete.")
            return redirect("show_final_antibiotic")

        upload_date = parse_date(upload_date_str)

        if not upload_date:
            messages.error(request, f"Invalid date format: {upload_date_str}")
            return redirect("show_final_antibiotic")

        deleted_count, _ = Final_AntibioticEntry.objects.filter(
            ab_Date_uploaded_fd=upload_date
        ).delete()

        messages.success(
            request,
            f"Deleted {deleted_count} Final Antibiotic entries uploaded on {upload_date}."
        )
        return redirect("show_final_antibiotic")

    messages.error(request, "Invalid request method.")
    return redirect("show_final_antibiotic")



@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER)
@require_POST
def delete_rawantibiotic_by_date(request):

    if request.method == "POST":
        upload_date_str = request.POST.get("upload_date")

        if not upload_date_str:
            messages.error(request, "Please select an upload date to delete.")
            return redirect("show_raw_antibiotic")

        upload_date = parse_date(upload_date_str)

        if not upload_date:
            messages.error(request, f"Invalid date format: {upload_date_str}")
            return redirect("show_raw_antibiotic")

        deleted_count, _ = AntibioticEntry.objects.filter(
            ab_Date_uploaded_rd=upload_date
        ).delete()

        messages.success(
            request,
            f"Deleted {deleted_count} Antibiotic entries uploaded on {upload_date}."
        )
        return redirect("show_raw_antibiotic")

    messages.error(request, "Invalid request method.")
    return redirect("show_raw_antibiotic")




@login_required(login_url="login")
def export_Final_Antibioticentry(request):
    objects = Final_AntibioticEntry.objects.all()
    data = []

    for obj in objects:
        data.append({
            "ab_idNum_f_referred": obj.ab_idNum_f_referred.f_AccessionNo if obj.ab_idNum_f_referred else None,
            "f_Accession_No": obj.ab_AccessionNo,
            "f_Site_Org": obj.ab_Site_Org,
            "Antibiotic": obj.ab_Antibiotic,
            "Abx_code": obj.ab_Abx_code,
            "Abx": obj.ab_Abx,
            "Disk_value": obj.ab_Disk_value,
            "Disk_enRIS": obj.ab_Disk_enRIS,
            "MIC_operand": obj.ab_MIC_operand,
            "MIC_value": obj.ab_MIC_value,
            "MIC_enRIS": obj.ab_MIC_enRIS,
            "f_Ars_Org":obj.ab_Ret_Org,
            "Retest_Antibiotic": obj.ab_Retest_Antibiotic,
            "Retest_Abx_code": obj.ab_Retest_Abx_code,
            "Retest_Abx": obj.ab_Retest_Abx,
            "Retest_DiskValue": obj.ab_Retest_DiskValue,
            "Retest_Disk_enRIS": obj.ab_Retest_Disk_enRIS,
            "Ret_MIC_Operand": obj.ab_Retest_MIC_operand,
            "Retest_MICValue": obj.ab_Retest_MICValue,
            "Retest_MIC_enRIS": obj.ab_Retest_MIC_enRIS,
            "ab_R_breakpoint": obj.ab_R_breakpoint,
            "ab_I_breakpoint": obj.ab_I_breakpoint,
            "ab_SDD_breakpoint": obj.ab_SDD_breakpoint,
            "ab_S_breakpoint":obj.ab_S_breakpoint,
            "ab_Ret_R_breakpoint": obj.ab_Ret_R_breakpoint,
            "ab_Ret_I_breakpoint": obj.ab_Ret_I_breakpoint,
            "ab_Ret_SDD_breakpoint": obj.ab_Ret_SDD_breakpoint,
            "ab_Ret_S_breakpoint": obj.ab_Ret_S_breakpoint
        })
    
    # Define file path
    file_path = "Final_AntibioticEntry.xlsx"

    # Convert data to DataFrame and save as Excel
    df = pd.DataFrame(data)
    df.to_excel(file_path, index=False)

    # Return the file as a response
    return FileResponse(open(file_path, "rb"), as_attachment=True, filename="Final_AntibioticEntry.xlsx")





@login_required(login_url="login")
def export_raw_Antibioticentry(request):
    objects = AntibioticEntry.objects.all()
    data = []

    for obj in objects:
        data.append({
            "ab_idNum_f_referred": obj.ab_idNum_referred.AccessionNo if obj.ab_idNum_referred else None,
            "f_Accession_No": obj.ab_AccessionNo,
            "f_Site_Org": obj.ab_Site_Org,
            "Antibiotic": obj.ab_Antibiotic,
            "Abx_code": obj.ab_Abx_code,
            "Abx": obj.ab_Abx,
            "Disk_value": obj.ab_Disk_value,
            "Disk_enRIS": obj.ab_Disk_enRIS,
            "MIC_operand": obj.ab_MIC_operand,
            "MIC_value": obj.ab_MIC_value,
            "MIC_enRIS": obj.ab_MIC_enRIS,
            "f_Ars_Org":obj.ab_Ret_Org,
            "Retest_Antibiotic": obj.ab_Retest_Antibiotic,
            "Retest_Abx_code": obj.ab_Retest_Abx_code,
            "Retest_Abx": obj.ab_Retest_Abx,
            "Retest_DiskValue": obj.ab_Retest_DiskValue,
            "Retest_Disk_enRIS": obj.ab_Retest_Disk_enRIS,
            "Ret_MIC_Operand": obj.ab_Retest_MIC_operand,
            "Retest_MICValue": obj.ab_Retest_MICValue,
            "Retest_MIC_enRIS": obj.ab_Retest_MIC_enRIS,
            "ab_R_breakpoint": obj.ab_R_breakpoint,
            "ab_I_breakpoint": obj.ab_I_breakpoint,
            "ab_SDD_breakpoint": obj.ab_SDD_breakpoint,
            "ab_S_breakpoint":obj.ab_S_breakpoint,
            "ab_Ret_R_breakpoint": obj.ab_Ret_R_breakpoint,
            "ab_Ret_I_breakpoint": obj.ab_Ret_I_breakpoint,
            "ab_Ret_SDD_breakpoint": obj.ab_Ret_SDD_breakpoint,
            "ab_Ret_S_breakpoint": obj.ab_Ret_S_breakpoint
        })
    
    # Define file path
    file_path = "AntibioticEntry.xlsx"

    # Convert data to DataFrame and save as Excel
    df = pd.DataFrame(data)
    df.to_excel(file_path, index=False)

    # Return the file as a response
    return FileResponse(open(file_path, "rb"), as_attachment=True, filename="AntibioticEntry.xlsx")



################ BATCH GENERATOR
############ Auto-Generate Batch
def generate_batches_from_referred():

    from itertools import groupby

    isolates = (
        Referred_Data.objects
        .filter(Batch_id__isnull=True)
        .exclude(Batch_Code__exact="")
        .order_by("Batch_Code", "AccessionNo")
    )

    created_batches = 0

    for batch_code, group in groupby(isolates, key=lambda x: x.Batch_Code):

        group = list(group)

        if not group:
            continue

        first = group[0]

        site_code = first.SiteCode
        referral_date = first.Referral_Date
        batch_name = first.Batch_Name or batch_code
        batch_no = first.BatchNo or "1"
        total_batch = first.Total_batch or "1"
        ref_no = first.RefNo or ""

        site_name = ""
        site_obj = SiteData.objects.filter(SiteCode=site_code).first()
        if site_obj:
            site_name = site_obj.SiteName

        batch_obj, created = Batch_Table.objects.get_or_create(
            bat_Batch_Code=batch_code,
            defaults={
                "bat_Batch_Name": batch_name,
                "bat_Site_Name": site_name,
                "bat_SiteCode": site_code,
                "bat_Referral_Date": referral_date,
                "bat_RefNo": ref_no,
                "bat_BatchNo": batch_no,
                "bat_Total_batch": total_batch
            }
        )

        # -------- Assign bat_seq --------
        seq = 1
        updated_isolates = []

        for iso in group:

            iso.Batch_id = batch_obj
            iso.Batch_Code = batch_code
            iso.Batch_Name = batch_name
            iso.bat_seq = seq

            updated_isolates.append(iso)

            seq += 1

        Referred_Data.objects.bulk_update(
            updated_isolates,
            ["Batch_id", "Batch_Code", "Batch_Name", "bat_seq"]
        )

        total_isolates = Referred_Data.objects.filter(Batch_id=batch_obj).count()
        tat_obj, _ = TATform.objects.get_or_create(
            tat_Batch_Isolates=batch_obj,
            defaults={
                "tat_SiteCode": batch_obj.bat_SiteCode,
                "tat_Batch_Code": batch_obj.bat_Batch_Code,
                "tat_Referral_Date": batch_obj.bat_Referral_Date,
                "tat_Num_Isolate": str(total_isolates),
                "tat_BatchNumber": batch_obj.bat_BatchNo,
                "tat_Total_Batch": batch_obj.bat_Total_batch,
            }
        )

        tat_obj.tat_SiteCode = batch_obj.bat_SiteCode
        tat_obj.tat_Batch_Code = batch_obj.bat_Batch_Code
        tat_obj.tat_Referral_Date = batch_obj.bat_Referral_Date
        tat_obj.tat_Num_Isolate = str(total_isolates)
        tat_obj.tat_BatchNumber = batch_obj.bat_BatchNo
        tat_obj.tat_Total_Batch = batch_obj.bat_Total_batch
        tat_obj.save(update_fields=[
            "tat_SiteCode",
            "tat_Batch_Code",
            "tat_Referral_Date",
            "tat_Num_Isolate",
            "tat_BatchNumber",
            "tat_Total_Batch",
        ])

        created_batches += 1

    return created_batches





# @login_required
# @transaction.atomic
# def create_batch_from_referred(request):

#     isolates = (
#         Referred_Data.objects
#         .filter(Batch_id__isnull=True)
#         .order_by("SiteCode", "Referral_Date", "AccessionNo")
#     )

#     from itertools import groupby

#     created_batches = 0

#     for (site, ref_date), group in groupby(
#         isolates,
#         key=lambda x: (x.SiteCode, x.Referral_Date)
#     ):

#         group = list(group)

#         if not group:
#             continue

#         year_short = ref_date.strftime("%y")
#         year_long = ref_date.strftime("%m%d%Y")

#         ref_numbers = [
#             int(i.AccessionNo[-4:])
#             for i in group
#             if i.AccessionNo[-4:].isdigit()
#         ]

#         start_ref = min(ref_numbers)
#         end_ref = max(ref_numbers)

#         ref_range = f"{start_ref}-{end_ref}"

#         batch_code = f"{site}_{year_long}_1.1_{ref_range}"

#         site_name = ""
#         site_obj = SiteData.objects.filter(SiteCode=site).first()
#         if site_obj:
#             site_name = site_obj.SiteName

#         batch_obj = Batch_Table.objects.create(
#             bat_Batch_Name=batch_code,
#             bat_AccessionNo=", ".join(i.AccessionNo for i in group),
#             bat_Batch_Code=batch_code,
#             bat_Site_Name=site_name,
#             bat_SiteCode=site,
#             bat_Referral_Date=ref_date,
#             bat_RefNo=ref_range,
#             bat_BatchNo="1",
#             bat_Total_batch="1"
#         )

#         seq = 1

#         for iso in group:

#             iso.Batch_id = batch_obj
#             iso.Batch_Code = batch_code
#             iso.Batch_Name = batch_code
#             iso.bat_seq = seq

#             iso.save(update_fields=[
#                 "Batch_id",
#                 "Batch_Code",
#                 "Batch_Name",
#                 "bat_seq"
#             ])

#             seq += 1

#         created_batches += 1

#     messages.success(
#         request,
#         f"{created_batches} batch(es) created successfully."
#     )

#     return redirect("show_batches")


@login_required
@transaction.atomic
def create_batch_from_referred(request):

    created_batches = generate_batches_from_referred()

    messages.success(
        request,
        f"{created_batches} batch(es) created successfully."
    )

    return redirect("show_batches")
