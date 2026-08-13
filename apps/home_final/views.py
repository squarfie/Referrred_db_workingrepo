# ================================
# Standard Library
# ================================
import os
import re
import csv
from decimal import Decimal, InvalidOperation
from datetime import date, datetime
from collections import OrderedDict, defaultdict
from urllib.parse import urlparse
from django.templatetags.static import static
# ================================
# Django Core
# ================================
from django.shortcuts import render, redirect, get_object_or_404
from django.http import (
    HttpResponse,
    HttpResponseRedirect,
    JsonResponse,
    FileResponse,
)
from django.urls import reverse
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.template.loader import get_template, render_to_string
from django.utils import timezone
from django.utils.timezone import now
from django.utils.dateparse import parse_date
from django.utils.http import url_has_allowed_host_and_scheme
from django.db import IntegrityError, transaction
from django.db.models import (
    Q,
    Count,
    Prefetch,
    Case,
    When,
    F,
    Max,
    OuterRef,
    Subquery,
)
from django.core.paginator import Paginator
from django.views.decorators.http import require_GET, require_POST
from django.apps import apps as django_apps

# ================================
# Third-Party Libraries
# ================================

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font as XLFont  # ✅ Safe alias (NO tkinter)
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter

# PDF (ReportLab)
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm

# Optional PDF (HTML to PDF)
from xhtml2pdf import pisa
from openpyxl.styles import Font

# ================================
# Project Imports
# ================================
from apps.home.views import _apply_signature_defaults_to_batch, link_callback
from apps.home.models import *
from apps.home.forms import *
from apps.home.permissions import (
    ROLE_ADMIN,
    ROLE_CHECKER,
    ROLE_ENCODER,
    ROLE_LAB_ENCODER,
    ROLE_LAB_MANAGER,
    ROLE_VERIFIER,
    can_manage_batch,
    get_user_role,
    role_required,
)
from apps.wgs_app.models import *
from apps.wgs_app.forms import *

from .models import *
from .forms import *
from .models import ConcordanceReport
from .utils import (
    _year_as_int,
    antibiotic_print_order,
    get_filtered_antibiotics,
    apply_final_breakpoints,
    make_cached_breakpoint_resolver,
    resolve_breakpoint,
    get_breakpoint_panel_abx_codes,
    sort_abx_codes_by_antibiotic,
)
from .services import concordance as concordance_service
from django.db.models.functions import ExtractYear
from django.contrib.staticfiles import finders


def HomeAntibioticList():
    return django_apps.get_model("home", "Antibiotic_List")


def HomeOrganismList():
    return django_apps.get_model("home", "Organism_List")


def HomeBreakpointsTable():
    return django_apps.get_model("home", "BreakpointsTable")


def _breakpoint_year_filter(bp_qs, specimen_year):
    target_year = _year_as_int(specimen_year)
    years_by_int = {}
    for year in bp_qs.values_list("Year", flat=True).distinct():
        year_int = _year_as_int(year)
        if year_int is not None:
            years_by_int.setdefault(year_int, []).append(year)

    if not years_by_int:
        return bp_qs.none()

    if target_year is not None:
        matching_years = years_by_int.get(target_year)
        if matching_years:
            return bp_qs.filter(Year__in=matching_years)

        selected_year = min(
            years_by_int,
            key=lambda year: (abs(year - target_year), year > target_year, year),
        )
        return bp_qs.filter(Year__in=years_by_int[selected_year])

    selected_year = max(years_by_int)
    return bp_qs.filter(Year__in=years_by_int[selected_year])


def _final_specific_panel_codes(specimen_year, org_code):
    org_code = (org_code or "").strip()
    OrganismList = HomeOrganismList()
    org_values = {org_code}
    organism = (
        OrganismList.objects
        .filter(Q(Whonet_Org_Code__iexact=org_code) | Q(Replaced_by__iexact=org_code))
        .values("Whonet_Org_Code", "Replaced_by")
        .first()
    )

    if organism:
        org_values.add((organism.get("Whonet_Org_Code") or "").strip())
        org_values.add((organism.get("Replaced_by") or "").strip())
    org_values = {value for value in org_values if value}

    Breakpoints = HomeBreakpointsTable()
    org_bp_qs = Breakpoints.objects.filter(_org_value_filter(org_values))
    bp_qs = _breakpoint_year_filter(org_bp_qs, specimen_year)
    abx_codes = {
        (code or "").strip().upper()
        for code in bp_qs.values_list("Abx_code", flat=True).distinct()
        if (code or "").strip()
    }
    whonet_codes = {
        (code or "").strip().upper()
        for code in bp_qs.values_list("Whonet_Abx", flat=True).distinct()
        if (code or "").strip()
    }

    return abx_codes, whonet_codes


def _org_value_filter(values):
    query = Q()
    has_values = False
    for value in values:
        value = (value or "").strip()
        if not value:
            continue
        query |= Q(Org__iexact=value)
        has_values = True
    return query if has_values else Q(pk__in=[])


def _final_related_org_values(field_name, field_value):
    values = {(field_value or "").strip()}
    if not field_value:
        return values

    OrganismList = HomeOrganismList()
    org_rows = OrganismList.objects.filter(
        **{f"{field_name}__iexact": field_value}
    ).values_list("Whonet_Org_Code", "Replaced_by")

    for whonet_code, replaced_by in org_rows:
        values.add((whonet_code or "").strip())
        values.add((replaced_by or "").strip())

    return {value for value in values if value}


def _final_year_whonet_codes(specimen_year):
    Breakpoints = HomeBreakpointsTable()
    bp_qs = _breakpoint_year_filter(Breakpoints.objects.all(), specimen_year)
    return {
        (code or "").strip().upper()
        for code in bp_qs.values_list("Whonet_Abx", flat=True).distinct()
        if (code or "").strip()
    }


def _code_iexact_filter(field_name, codes):
    query = Q()
    has_codes = False
    for code in codes:
        code = (code or "").strip()
        if not code:
            continue
        query |= Q(**{f"{field_name}__iexact": code})
        has_codes = True
    return query if has_codes else Q(pk__in=[])


def _antibiotic_view_mode(request):
    mode = (request.GET.get("antibiotic_view") or request.POST.get("antibiotic_view") or "all").strip().lower()
    return "panel" if mode == "panel" else "all"


def _final_antibiotics_for_panel(
    *,
    org_code,
    specimen_year,
    show_site=False,
    retest=False,
    require_org=False,
    existing_whonet_codes=None,
    antibiotic_view="all",
):
    AntibioticList = HomeAntibioticList()
    qs = AntibioticList.objects.all()
    antibiotic_view = (antibiotic_view or "all").strip().lower()
    existing_whonet_codes = {
        (code or "").strip().upper()
        for code in (existing_whonet_codes or [])
        if (code or "").strip()
    }

    if retest:
        qs = qs.filter(Retest=True)
    elif show_site:
        qs = qs.filter(Show=True)

    org_code = (org_code or "").strip()
    if _is_no_organism(org_code):
        return qs.none()

    if antibiotic_view == "panel":
        if not org_code:
            return qs.none()
        breakpoint_year = _year_as_int(specimen_year) or specimen_year
        panel_codes = {
            (code or "").strip().upper()
            for code in get_breakpoint_panel_abx_codes(breakpoint_year, org_code)
            if (code or "").strip()
        }
        panel_filter = (
            _code_iexact_filter("Whonet_Abx", panel_codes)
            | _code_iexact_filter("Abx_code", panel_codes)
        )
        if existing_whonet_codes:
            panel_filter |= _code_iexact_filter("Whonet_Abx", existing_whonet_codes)
        return qs.filter(panel_filter).distinct().order_by("Antibiotic", "Whonet_Abx")

    qs = qs.filter(Show_All=True)
    return qs.order_by("Antibiotic", "Whonet_Abx")


def _parse_disk_value(value):
    try:
        disk_value = int(value) if value else None
    except (TypeError, ValueError):
        return None

    if disk_value is not None and disk_value < 6:
        return None
    return disk_value


def _entry_values_match(entry, values):
    for field, value in values.items():
        current = getattr(entry, field)
        if current == value:
            continue
        if current is None and value == "":
            continue
        if current == "" and value is None:
            continue
        return False
    return True


MAIN_ANTIBIOTIC_FIELDS = [
    "ab_Antibiotic", "ab_Abx_code", "ab_Abx",
    "ab_Disk_value", "ab_Disk_RIS", "ab_Disk_enRIS",
    "ab_MIC_value", "ab_MIC_RIS", "ab_MIC_enRIS", "ab_MIC_operand",
    "ab_AlertMIC", "ab_Alert_val",
    "ab_Site_Org", "ab_R_breakpoint", "ab_I_breakpoint",
    "ab_SDD_breakpoint", "ab_S_breakpoint",
]

RETEST_ANTIBIOTIC_FIELDS = [
    "ab_Retest_Antibiotic", "ab_Retest_Abx_code", "ab_Retest_Abx",
    "ab_Retest_DiskValue", "ab_Retest_Disk_RIS", "ab_Retest_Disk_enRIS",
    "ab_Retest_MICValue", "ab_Retest_MIC_RIS", "ab_Retest_MIC_enRIS",
    "ab_Retest_MIC_operand", "ab_Retest_AlertMIC", "ab_Retest_Alert_val",
    "ab_Ret_Org", "ab_Org_Flag", "ab_Abx_Flag", "ab_Abx_Phenotype",
    "ab_Abx_Phenotype_Other", "ab_Ret_R_breakpoint", "ab_Ret_I_breakpoint",
    "ab_Ret_SDD_breakpoint", "ab_Ret_S_breakpoint",
]


def _has_main_antibiotic_data(entry):
    return any([
        entry.ab_Disk_value is not None,
        entry.ab_MIC_value is not None,
        bool((entry.ab_Disk_RIS or "").strip()),
        bool((entry.ab_Disk_enRIS or "").strip()),
        bool((entry.ab_MIC_RIS or "").strip()),
        bool((entry.ab_MIC_enRIS or "").strip()),
        bool((entry.ab_MIC_operand or "").strip()),
        bool(entry.ab_AlertMIC),
    ])


def _has_retest_antibiotic_data(entry):
    return any([
        entry.ab_Retest_DiskValue is not None,
        entry.ab_Retest_MICValue is not None,
        bool((entry.ab_Retest_Disk_RIS or "").strip()),
        bool((entry.ab_Retest_Disk_enRIS or "").strip()),
        bool((entry.ab_Retest_MIC_RIS or "").strip()),
        bool((entry.ab_Retest_MIC_enRIS or "").strip()),
        bool((entry.ab_Retest_MIC_operand or "").strip()),
        bool(entry.ab_Retest_AlertMIC),
    ])


def _is_blank_print_value(value):
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"", "n/a", "na", "none", "null", "-", "--"}
    return False


def _has_printable_main_antibiotic_result(entry):
    return any(
        not _is_blank_print_value(getattr(entry, field, None))
        for field in (
            "ab_Disk_value",
            "ab_Disk_RIS",
            "ab_Disk_enRIS",
            "ab_MIC_value",
            "ab_MIC_RIS",
            "ab_MIC_enRIS",
        )
    )


def _has_printable_retest_antibiotic_result(entry):
    return any(
        not _is_blank_print_value(getattr(entry, field, None))
        for field in (
            "ab_Retest_DiskValue",
            "ab_Retest_Disk_RIS",
            "ab_Retest_Disk_enRIS",
            "ab_Retest_MICValue",
            "ab_Retest_MIC_RIS",
            "ab_Retest_MIC_enRIS",
        )
    )


def _clear_main_antibiotic_data(entry):
    for field in MAIN_ANTIBIOTIC_FIELDS:
        field_obj = entry._meta.get_field(field)
        if field_obj.get_internal_type() == "BooleanField":
            setattr(entry, field, False)
        elif field_obj.get_internal_type() in {"IntegerField", "PositiveSmallIntegerField", "DecimalField"}:
            setattr(entry, field, None)
        elif field_obj.null:
            setattr(entry, field, None)
        else:
            setattr(entry, field, "")


def _clear_retest_antibiotic_data(entry):
    for field in RETEST_ANTIBIOTIC_FIELDS:
        field_obj = entry._meta.get_field(field)
        if field_obj.get_internal_type() == "BooleanField":
            setattr(entry, field, False)
        elif field_obj.get_internal_type() in {"IntegerField", "PositiveSmallIntegerField", "DecimalField"}:
            setattr(entry, field, None)
        elif field_obj.null:
            setattr(entry, field, None)
        else:
            setattr(entry, field, "")


def _save_or_delete_antibiotic_entry(entry):
    if _has_main_antibiotic_data(entry) or _has_retest_antibiotic_data(entry):
        entry.save()
    else:
        entry.delete()


NO_ORGANISM_CODES = {"", "n/a", "na", "n.a.", "nv", "none", "null", "nan"}


def _is_no_organism(value):
    return (value or "").strip().lower() in NO_ORGANISM_CODES


FASTIDIOUS_PLUS_LAYOUT_SPECIES_GROUPS = {
    "ABI", "GCT", "AGT", "BD-", "BR-", "CAM", "CAR", "EIK", "FRA",
    "HA-", "HEL", "KIN", "LEG", "MOR", "NE-", "NV", "N/A", "NA",
    "N.A.", "NONE", "NULL", "NAN",
}


def _uses_fastidious_plus_layout(organism):
    if not organism:
        return False

    organism_type = str(organism.get("Organism_Type") or "").strip()
    if organism_type == "+":
        return True

    codes = {
        str(organism.get(field) or "").strip().upper()
        for field in ("Whonet_Org_Code", "Replaced_by", "Species_Group", "Genus_Group", "Genus_Code")
    }
    codes.discard("")
    if codes & FASTIDIOUS_PLUS_LAYOUT_SPECIES_GROUPS:
        return True

    return _organism_name_uses_fastidious_plus_layout(organism.get("Organism")) or bool(
        {"HIN", "NME"} & codes
    )


def _organism_name_uses_fastidious_plus_layout(organism_name):
    name = str(organism_name or "").strip().lower()
    return name.startswith(("haemophilus ", "neisseria "))


def _organism_type_is_plus(org_code, organism_name=None):
    code = str(org_code or "").strip()
    if _is_no_organism(code):
        return True
    if _organism_name_uses_fastidious_plus_layout(organism_name or code):
        return True
    name = str(organism_name or "").strip()
    if not code and not name:
        return False
    lookup = Q()
    if code:
        lookup |= Q(Whonet_Org_Code__iexact=code) | Q(Replaced_by__iexact=code) | Q(Organism__iexact=code)
    if name:
        lookup |= Q(Organism__iexact=name)
    OrganismList = HomeOrganismList()
    organism = (
        OrganismList.objects
        .filter(lookup)
        .values("Whonet_Org_Code", "Replaced_by", "Organism_Type", "Species_Group", "Genus_Group", "Genus_Code", "Organism")
        .first()
    )
    return _uses_fastidious_plus_layout(organism)


PHENOTYPE_CLEAR_POST_FIELDS = {
    "phenotype_search_f_site_pre": ("f_Site_Pre", "f_Site_Pre_ed"),
    "phenotype_search_f_site_post": ("f_Site_Pos", "f_Site_Pos_ed"),
    "phenotype_search_f_ars_pre": ("f_ars_Pre", "f_ars_Pre_ed"),
    "phenotype_search_f_ars_post": ("f_ars_Post", "f_ars_Post_ed"),
}


def _is_clear_phenotype_value(value):
    return (value or "").strip().lower() in {"n/a", "na", "n.a.", "none", "null"}


def _apply_clear_phenotype_posts(isolate, post_data):
    for search_name, (source_field, edit_field) in PHENOTYPE_CLEAR_POST_FIELDS.items():
        if not (
            _is_clear_phenotype_value(post_data.get(search_name))
            or _is_clear_phenotype_value(post_data.get(edit_field))
        ):
            continue
        setattr(isolate, source_field, "")
        setattr(isolate, edit_field, "")


# SHOW FINAL DATA TABLE
# @login_required(login_url="login")
# def show_final_table(request):
#     query = request.GET.get("q", "").strip()

#     # DEFAULT SORT FIELD MUST EXIST
#     sort_by = request.GET.get("sort", "f_Date_Modified")
#     order = request.GET.get("order", "desc")

#     # SAFETY: allow only valid sortable fields
#     allowed_sort_fields = {
#         "f_AccessionNo",
#         "f_First_Name",
#         "f_Last_Name",
#         "f_Batch_Code",
#         "f_SiteCode",
#         "f_Date_Modified",
#         "f_Spec_Date",
#     }

#     if sort_by not in allowed_sort_fields:
#         sort_by = "f_Date_Modified"

#     sort_field = f"-{sort_by}" if order == "desc" else sort_by

#     # BASE QUERYSET
#     records = (
#         Final_Data.objects
#         .select_related("f_Spec_Type", "f_Batch_id")  # FK optimization
#         .prefetch_related("final_entries")           # Antibiotics
#         .order_by(sort_field)
#     )

#     # SEARCH LOGIC (ALIGNED WITH Final_Data)
#     if query:
#         records = records.filter(
#             Q(f_AccessionNo__icontains=query) |
#             Q(f_First_Name__icontains=query) |
#             Q(f_Last_Name__icontains=query) |
#             Q(f_Patient_ID__icontains=query) |
#             Q(f_Batch_Code__icontains=query) |
#             Q(f_SiteCode__icontains=query) |
#             Q(f_Site_Name__icontains=query) |
#             Q(f_Site_Org__icontains=query) |
#             Q(f_Site_OrgName__icontains=query) |
#             Q(f_ars_OrgCode__icontains=query) |
#             Q(f_ars_OrgName__icontains=query) |
#             Q(f_Spec_Num__icontains=query) |
#             Q(f_Spec_Type__Specimen_code__icontains=query) |   # FK SAFE
#             Q(f_Spec_Type__Specimen_name__icontains=query)
#         ).distinct()

#     # PAGINATION
#     paginator = Paginator(records, 20)
#     page_number = request.GET.get("page")
#     page_obj = paginator.get_page(page_number)

#     return render(
#         request,
#         "home_final/tables_final.html",
#         {
#             "page_obj": page_obj,
#             "current_sort": sort_by,
#             "current_order": order,
#             "query": query,
#         }
#     )




def _final_records_base_queryset(request):
    records = Final_Data.objects.all()
    if get_user_role(request.user) == ROLE_ENCODER:
        records = records.filter(f_Batch_id__created_by=request.user)
    return records


def _apply_final_table_filters(records, query="", year=None):
    if query:
        records = records.filter(
            Q(f_AccessionNo__icontains=query) |
            Q(f_First_Name__icontains=query) |
            Q(f_Last_Name__icontains=query) |
            Q(f_Patient_ID__icontains=query) |
            Q(f_Batch_Code__icontains=query) |
            Q(f_SiteCode__icontains=query) |
            Q(f_Site_Name__icontains=query) |
            Q(f_Site_Org__icontains=query) |
            Q(f_Site_OrgName__icontains=query) |
            Q(f_ars_OrgCode__icontains=query) |
            Q(f_ars_OrgName__icontains=query) |
            Q(f_Spec_Num__icontains=query) |
            Q(f_Spec_Type__Specimen_code__icontains=query) |
            Q(f_Spec_Type__Specimen_name__icontains=query)
        ).distinct()

    if year and str(year).isdigit():
        records = records.filter(f_Referral_Date__year=int(year))

    return records


def _final_sort_field(sort_by, order):
    allowed_sort_fields = {
        "f_AccessionNo",
        "f_First_Name",
        "f_Last_Name",
        "f_Batch_Code",
        "f_SiteCode",
        "f_Date_Modified",
        "f_Spec_Date",
        "f_bat_seq",
    }
    if sort_by not in allowed_sort_fields:
        sort_by = "f_Date_Modified"
    if order not in ["asc", "desc"]:
        order = "desc"
    return sort_by, order, f"-{sort_by}" if order == "desc" else sort_by


@login_required(login_url="login")
def show_final_table(request):

    query = request.GET.get("q", "").strip()
    year = request.GET.get("year")

    sort_by = request.GET.get("sort", "f_Date_Modified")
    order = request.GET.get("order", "desc")

    sort_by, order, sort_field = _final_sort_field(sort_by, order)
    records = _apply_final_table_filters(_final_records_base_queryset(request), query, year)

    total_records = records.count()

    batch_sort_map = {
        "f_Batch_Code": "batch_code",
        "f_SiteCode": "site_code",
        "f_Date_Modified": "latest_modified",
        "f_Spec_Date": "latest_specimen_date",
    }
    batch_sort_field = batch_sort_map.get(sort_by, "f_Batch_Code")
    if order == "desc":
        batch_sort_field = f"-{batch_sort_field}"

    batch_summaries = (
        records
        .values("f_Batch_id")
        .annotate(
            record_count=Count("id"),
            batch_code=Max("f_Batch_id__bat_Batch_Code"),
            fallback_batch_code=Max("f_Batch_Code"),
            site_code=Max("f_SiteCode"),
            latest_modified=Max("f_Date_Modified"),
            latest_specimen_date=Max("f_Spec_Date"),
        )
        .order_by(batch_sort_field, "-batch_code")
    )

    paginator = Paginator(batch_summaries, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    page_groups = []
    for batch_summary in page_obj.object_list:
        raw_batch_code = batch_summary.get("batch_code") or batch_summary.get("fallback_batch_code") or ""
        batch_code = raw_batch_code.strip() or "Unbatched"
        page_groups.append({
            "batch_id": batch_summary.get("f_Batch_id") or "",
            "code": batch_code,
            "batch_code": raw_batch_code,
            "count": batch_summary["record_count"],
            "site_code": batch_summary.get("site_code") or "",
        })

    # Available years
    available_years = (
        Final_Data.objects
        .annotate(year=ExtractYear("f_Referral_Date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )

    # Preserve GET parameters
    params = request.GET.copy()
    params.pop("sort", None)
    params.pop("order", None)
    preserved_params = params.urlencode()

    return render(
        request,
        "home_final/tables_final.html",
        {
            "page_obj": page_obj,
            "page_groups": page_groups,
            "total_records": total_records,
            "current_sort": sort_by,
            "current_order": order,
            "query": query,
            "year": year,
            "available_years": available_years,
            "preserved_params": preserved_params,
        }
    )


@login_required(login_url="login")
@require_GET
def final_batch_rows(request):
    batch_id = request.GET.get("batch_id", "")
    batch_dom_id = request.GET.get("target", "")
    query = request.GET.get("q", "").strip()
    year = request.GET.get("year")
    sort_by = request.GET.get("sort", "f_Date_Modified")
    order = request.GET.get("order", "desc")
    _, _, sort_field = _final_sort_field(sort_by, order)

    records = _apply_final_table_filters(_final_records_base_queryset(request), query, year)
    if batch_id:
        records = records.filter(f_Batch_id_id=batch_id)
    else:
        records = records.filter(f_Batch_id__isnull=True)
    records = (
        records
        .select_related("f_Spec_Type", "f_Batch_id")
        .prefetch_related("final_entries")
        .order_by("f_bat_seq", "f_AccessionNo", "id")
    )
    html = render_to_string(
        "home_final/partials/final_batch_rows.html",
        {
            "records": records,
            "batch_dom_id": batch_dom_id,
            "request": request,
        },
        request=request,
    )
    return JsonResponse({"html": html})




# # Create your views here.

# EDIT DATA - NEW VERSION
# @login_required(login_url="login")
# @transaction.atomic
# def edit_final_data(request, id):

#     isolate = get_object_or_404(Final_Data, pk=id)

#     request.session["current_final_isolate_id"] = isolate.id


#     classification, _ = Classification_Table.objects.get_or_create(
#         Class_AccessionNo=isolate.f_AccessionNo,
#         defaults={"Class_idNumReferred": isolate}
#     )


#     # =========================
#     # GET
#     # =========================
#     if request.method == "GET":

#         form = FinalReferred_Form(instance=isolate)

#         antibiotics_main = Antibiotic_List.objects.filter(Show=True).order_by("Antibiotic")
#         antibiotics_retest = Antibiotic_List.objects.filter(Retest=True).order_by("Antibiotic")

#         existing_entries = Final_AntibioticEntry.objects.filter(
#             ab_idNum_f_referred=isolate
#         )

#         retest_entries = existing_entries.exclude(
#             ab_Retest_Abx_code__isnull=True
#         )

#         return render(request, "home_final/edit_final.html", {
#             "form": form,
#             "isolates": isolate,
#             "antibiotics_main": antibiotics_main,
#             "antibiotics_retest": antibiotics_retest,
#             "existing_entries": existing_entries,
#             "retest_entries": retest_entries,
#             "classification": classification,
#             "edit_mode": True,
#         })

#     # =========================
#     # POST
#     # =========================


#     old_site_org = (isolate.f_Site_Org or "").strip()
#     old_ars_org  = (isolate.f_ars_OrgCode or "").strip()

#     form = FinalReferred_Form(request.POST, instance=isolate)
#     if not form.is_valid():
#         messages.error(request, "Error saving Final data.")
#         return redirect("edit_final_data", id=id)

#     # use instance, NOT form.save()
#     isolate = form.instance

#     # --- FORCE SAVE SITE ORG ---
#     site_org_obj = form.cleaned_data.get("f_Site_Org")
#     if site_org_obj:
#         isolate.f_Site_Org = site_org_obj.Whonet_Org_Code
#     else:
#         isolate.f_Site_Org = ""

#     # --- FORCE SAVE ARS ORG ---
#     ars_org_obj = form.cleaned_data.get("f_ars_OrgCode")
#     if ars_org_obj:
#         isolate.f_ars_OrgCode = ars_org_obj.Whonet_Org_Code
#     else:
#         isolate.f_ars_OrgCode = ""

#     # save ONCE
#     isolate.save()

#     classification.Class_Chk_Emerging    = "Class_Chk_Emerging" in request.POST
#     classification.Class_Chk_Satscan     = "Class_Chk_Satscan" in request.POST
#     classification.Class_Chk_Serotyping  = "Class_Chk_Serotyping" in request.POST
#     classification.Class_Chk_GHRU_all    = "Class_Chk_GHRU_all" in request.POST
#     classification.Class_Chk_GHRU_Neo    = "Class_Chk_GHRU_Neo" in request.POST
#     classification.Class_Chk_Tricycle    = "Class_Chk_Tricycle" in request.POST

#     classification.Class_AccessionNo = isolate.f_AccessionNo
#     classification.save()




#     # =========================
#     # Breakpoint year
#     # =========================
#     specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None

#     if specimen_year:
#         effective_year = (
#             BreakpointsTable.objects
#             .filter(Year__lte=str(specimen_year))
#             .order_by("-Year")
#             .values_list("Year", flat=True)
#             .first()
#         )
#     else:
#         effective_year = (
#             BreakpointsTable.objects
#             .order_by("-Year")
#             .values_list("Year", flat=True)
#             .first()
#         )

#     resolved_site_org = (isolate.f_Site_Org or "").strip()
#     resolved_ars_org  = (isolate.f_ars_OrgCode or "").strip()

#     # =========================
#     # ORG CHANGE CLEANUP
#     # =========================
#     if old_site_org != resolved_site_org:
#         Final_AntibioticEntry.objects.filter(
#             ab_idNum_f_referred=isolate,
#             ab_Abx_code__isnull=False
#         ).delete()

#     if old_ars_org != resolved_ars_org:
#         Final_AntibioticEntry.objects.filter(
#             ab_idNum_f_referred=isolate,
#             ab_Retest_Abx_code__isnull=False
#         ).delete()

#     # =========================
#     # MAIN ANTIBIOTICS
#     # =========================
#     for abx in Antibiotic_List.objects.filter(Show=True):

#         code = (abx.Whonet_Abx or "").strip().upper()

#         disk = request.POST.get(f"disk_{code}")
#         mic  = request.POST.get(f"mic_{code}")

#         disk = int(disk) if disk and disk.isdigit() else None
#         mic  = float(mic) if mic else None

#         if disk is None and mic is None:
#             continue

#         entry, _ = Final_AntibioticEntry.objects.update_or_create(
#             ab_idNum_f_referred=isolate,
#             ab_Abx_code=code,
#             defaults={
#                 "ab_AccessionNo": isolate.f_AccessionNo,
#                 "ab_Antibiotic": abx.Antibiotic,
#                 "ab_Abx": abx.Abx_code,
#                 "ab_Disk_value": disk,
#                 "ab_MIC_value": mic,
#                 "ab_Disk_enRIS": request.POST.get(f"disk_enris_{code}", ""),
#                 "ab_MIC_enRIS": request.POST.get(f"mic_enris_{code}", ""),
#                 "ab_MIC_operand": request.POST.get(f"mic_operand_{code}", ""),
#                 "ab_AlertMIC": f"alert_mic_{code}" in request.POST,
#             }
#         )

#         apply_final_breakpoints(
#             entry,
#             org_code=resolved_site_org,
#             year=effective_year,
#             is_retest=False
#         )

#     # =========================
#     # RETEST ANTIBIOTICS
#     # =========================
#     for abx in Antibiotic_List.objects.filter(Retest=True):

#         code = (abx.Whonet_Abx or "").strip().upper()

#         disk = request.POST.get(f"retest_disk_{code}")
#         mic  = request.POST.get(f"retest_mic_{code}")

#         disk = int(disk) if disk and disk.isdigit() else None
#         mic  = float(mic) if mic else None

#         if disk is None and mic is None:
#             continue

#         entry, _ = Final_AntibioticEntry.objects.update_or_create(
#             ab_idNum_f_referred=isolate,
#             ab_Retest_Abx_code=code,
#             defaults={
#                 "ab_AccessionNo": isolate.f_AccessionNo,
#                 "ab_Retest_Antibiotic": abx.Antibiotic,
#                 "ab_Retest_Abx": abx.Abx_code,
#                 "ab_Retest_DiskValue": disk,
#                 "ab_Retest_MICValue": mic,
#                 "ab_Retest_Disk_enRIS": request.POST.get(f"retest_disk_enris_{code}", ""),
#                 "ab_Retest_MIC_enRIS": request.POST.get(f"retest_mic_enris_{code}", ""),
#                 "ab_Retest_MIC_operand": request.POST.get(f"retest_mic_operand_{code}", ""),
#                 "ab_Retest_AlertMIC": f"retest_alert_mic_{code}" in request.POST,
#             }
#         )

#         apply_final_breakpoints(
#             entry,
#             org_code=resolved_ars_org,
#             year=effective_year,
#             is_retest=True
#         )

#     messages.success(request, "Final data saved successfully.")
#     return redirect("show_final_table")


def _final_batch_navigation(isolate):
    batch = isolate.f_Batch_id
    if not batch:
        return {
            "batch_id": None,
            "first_id": None,
            "previous_id": None,
            "next_id": None,
            "last_id": None,
            "position": 1,
            "total": 1,
        }

    isolate_ids = list(
        Final_Data.objects
        .filter(f_Batch_id=batch)
        .order_by("f_bat_seq", "f_AccessionNo", "id")
        .values_list("id", flat=True)
    )

    try:
        current_index = isolate_ids.index(isolate.id)
    except ValueError:
        current_index = 0

    return {
        "batch_id": batch.id,
        "first_id": isolate_ids[0] if isolate_ids else None,
        "previous_id": isolate_ids[current_index - 1] if current_index > 0 else None,
        "next_id": isolate_ids[current_index + 1] if current_index + 1 < len(isolate_ids) else None,
        "last_id": isolate_ids[-1] if isolate_ids else None,
        "position": current_index + 1,
        "total": len(isolate_ids),
    }


@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_ENCODER)
@transaction.atomic
def edit_final_data(request, id):

    isolate = get_object_or_404(Final_Data, pk=id)
    if not can_manage_batch(request.user, isolate.f_Batch_id):
        messages.error(request, "You can only update final records from batches that you created.")
        return redirect("show_final_table")

    request.session["current_final_isolate_id"] = isolate.id
    antibiotic_view = _antibiotic_view_mode(request)

    classification, _ = Classification_Table.objects.get_or_create(
        Class_idNumReferred=isolate,
        defaults={"Class_AccessionNo": isolate.f_AccessionNo}
    )

    # =========================
    # GET
    # =========================
    if request.method == "GET":

        form = FinalReferred_Form(instance=isolate)
        existing_entries = Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred=isolate
        )
        existing_main_codes = existing_entries.exclude(
            ab_Abx_code__isnull=True
        ).values_list("ab_Abx_code", flat=True)
        existing_retest_codes = existing_entries.exclude(
            ab_Retest_Abx_code__isnull=True
        ).values_list("ab_Retest_Abx_code", flat=True)
        specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None

        antibiotics_main = _final_antibiotics_for_panel(
            org_code=isolate.f_Site_Org,
            specimen_year=specimen_year,
            show_site=True,
            existing_whonet_codes=existing_main_codes,
            antibiotic_view=antibiotic_view,
        )
        antibiotics_retest = _final_antibiotics_for_panel(
            org_code=isolate.f_ars_OrgCode,
            specimen_year=specimen_year,
            retest=True,
            require_org=True,
            existing_whonet_codes=existing_retest_codes,
            antibiotic_view=antibiotic_view,
        )

        retest_entries = existing_entries.exclude(
            ab_Retest_Abx_code__isnull=True
        )

        return render(request, "home_final/edit_final.html", {
            "form": form,
            "isolates": isolate,
            "batch_nav": _final_batch_navigation(isolate),
            "antibiotics_main": antibiotics_main,
            "antibiotics_retest": antibiotics_retest,
            "existing_entries": existing_entries,
            "retest_entries": retest_entries,
            "classification": classification,
            "edit_mode": True,
            "antibiotic_view": antibiotic_view,
        })

    # =========================
    # POST
    # =========================

    old_site_org = (isolate.f_Site_Org or "").strip()
    old_ars_org  = (isolate.f_ars_OrgCode or "").strip()
    old_specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None
    original_phenotypes = {
        "f_Site_Pre": isolate.f_Site_Pre,
        "f_Site_Pos": isolate.f_Site_Pos,
        "f_ars_Pre": isolate.f_ars_Pre,
        "f_ars_Post": isolate.f_ars_Post,
    }

    post_data = request.POST.copy()
    if old_site_org and not (post_data.get("f_Site_Org") or "").strip():
        old_site_choice = resolve_organism_choice(old_site_org, isolate.f_Site_OrgName)
        post_data["f_Site_Org"] = (
            old_site_choice.Whonet_Org_Code if old_site_choice else old_site_org
        )
    if old_ars_org and not (post_data.get("f_ars_OrgCode") or "").strip():
        old_ars_choice = resolve_organism_choice(old_ars_org, isolate.f_ars_OrgName)
        post_data["f_ars_OrgCode"] = (
            old_ars_choice.Whonet_Org_Code if old_ars_choice else old_ars_org
        )
    if isolate.f_Site_OrgName and not (post_data.get("f_Site_OrgName") or "").strip():
        post_data["f_Site_OrgName"] = isolate.f_Site_OrgName
    if isolate.f_ars_OrgName and not (post_data.get("f_ars_OrgName") or "").strip():
        post_data["f_ars_OrgName"] = isolate.f_ars_OrgName

    form = FinalReferred_Form(post_data, instance=isolate)

    if not form.is_valid():
        messages.error(request, "Please check the highlighted fields.")
        existing_entries = Final_AntibioticEntry.objects.filter(ab_idNum_f_referred=isolate)
        specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None
        return render(request, "home_final/edit_final.html", {
            "form": form,
            "isolates": isolate,
            "batch_nav": _final_batch_navigation(isolate),
            "antibiotics_main": _final_antibiotics_for_panel(
                org_code=isolate.f_Site_Org,
                specimen_year=specimen_year,
                show_site=True,
                existing_whonet_codes=existing_entries.exclude(ab_Abx_code__isnull=True).values_list("ab_Abx_code", flat=True),
                antibiotic_view=antibiotic_view,
            ),
            "antibiotics_retest": _final_antibiotics_for_panel(
                org_code=isolate.f_ars_OrgCode,
                specimen_year=specimen_year,
                retest=True,
                require_org=True,
                existing_whonet_codes=existing_entries.exclude(ab_Retest_Abx_code__isnull=True).values_list("ab_Retest_Abx_code", flat=True),
                antibiotic_view=antibiotic_view,
            ),
            "existing_entries": existing_entries,
            "retest_entries": existing_entries.exclude(ab_Retest_Abx_code__isnull=True),
            "classification": classification,
            "edit_mode": True,
            "antibiotic_view": antibiotic_view,
        })

    isolate = form.save(commit=False)
    for field_name, original_value in original_phenotypes.items():
        setattr(isolate, field_name, original_value)
    _apply_clear_phenotype_posts(isolate, request.POST)
    isolate.save()

    # =========================
    # SAVE CLASSIFICATION
    # =========================
    classification.Class_Chk_Structured = "Class_Chk_Structured" in request.POST
    classification.Class_Chk_Satscan    = "Class_Chk_Satscan" in request.POST
    classification.Class_Chk_Serotyping = "Class_Chk_Serotyping" in request.POST
    classification.Class_Chk_GHRU_all   = "Class_Chk_GHRU_all" in request.POST
    classification.Class_Chk_GHRU_Neo   = "Class_Chk_GHRU_Neo" in request.POST
    classification.Class_Chk_EGASP      = "Class_Chk_EGASP" in request.POST
    classification.Class_Chk_Tricycle   = "Class_Chk_Tricycle" in request.POST
    classification.Class_Chk_Pulsenet   = "Class_Chk_Pulsenet" in request.POST
    classification.Class_Chk_Tulip      = "Class_Chk_Tulip" in request.POST
    classification.Class_AccessionNo    = isolate.f_AccessionNo
    classification.save()

    # =========================
    # BREAKPOINT YEAR
    # =========================
    specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None

    new_site_org = (isolate.f_Site_Org or "").strip()
    new_ars_org  = (isolate.f_ars_OrgCode or "").strip()
    site_org_is_na = _is_no_organism(new_site_org)
    ars_org_is_na = _is_no_organism(new_ars_org)

    # =========================
    # DELETE IF ORGANISM CHANGED
    # =========================
    if old_site_org != new_site_org or site_org_is_na:
        for entry in Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred=isolate,
            ab_Abx_code__isnull=False
        ):
            _clear_main_antibiotic_data(entry)
            _save_or_delete_antibiotic_entry(entry)

    if old_ars_org != new_ars_org or ars_org_is_na:
        for entry in Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred=isolate,
            ab_Retest_Abx_code__isnull=False
        ):
            _clear_retest_antibiotic_data(entry)
            _save_or_delete_antibiotic_entry(entry)

    resolved_site_org = new_site_org
    resolved_ars_org  = new_ars_org

    # =========================
    # MAIN ANTIBIOTICS
    # =========================
    resolve_bp = make_cached_breakpoint_resolver()
    existing_main_entries = {
        (entry.ab_Abx_code or "").strip().upper(): entry
        for entry in Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred=isolate,
            ab_Abx_code__isnull=False,
        )
    }
    existing_retest_entries_for_main = {
        (entry.ab_Retest_Abx_code or "").strip().upper(): entry
        for entry in Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred=isolate,
            ab_Retest_Abx_code__isnull=False,
        )
    }
    can_skip_main_entries = (
        old_site_org == new_site_org
        and old_specimen_year == specimen_year
    )
    antibiotics_main = list(_final_antibiotics_for_panel(
        org_code=resolved_site_org,
        specimen_year=specimen_year,
        show_site=True,
        existing_whonet_codes=existing_main_entries.keys(),
        antibiotic_view=antibiotic_view,
    ))

    if not site_org_is_na:
      for abx in antibiotics_main:

        abx_code = (abx.Whonet_Abx or "").strip().upper()

        disk_value  = request.POST.get(f"disk_{abx_code}")
        mic_value   = request.POST.get(f"mic_{abx_code}")
        disk_enris  = request.POST.get(f"disk_enris_{abx_code}", "").strip()
        mic_enris   = request.POST.get(f"mic_enris_{abx_code}", "").strip()
        mic_operand = request.POST.get(f"mic_operand_{abx_code}", "").strip()
        alert_mic   = f"alert_mic_{abx_code}" in request.POST

        disk_value = _parse_disk_value(disk_value)

        try:
            mic_value = float(mic_value) if mic_value else None
        except ValueError:
            mic_value = None

        if disk_value is None and mic_value is None:
            for entry in Final_AntibioticEntry.objects.filter(
                ab_idNum_f_referred=isolate,
                ab_Abx_code=abx_code
            ):
                _clear_main_antibiotic_data(entry)
                _save_or_delete_antibiotic_entry(entry)
            continue

        entry_defaults = {
            "ab_AccessionNo": isolate.f_AccessionNo,
            "ab_Antibiotic": abx.Antibiotic,
            "ab_Abx": abx.Abx_code,
            "ab_Disk_value": disk_value,
            "ab_Disk_enRIS": disk_enris,
            "ab_MIC_value": mic_value,
            "ab_MIC_enRIS": mic_enris,
            "ab_MIC_operand": mic_operand,
            "ab_AlertMIC": alert_mic,
        }
        existing_entry = existing_main_entries.get(abx_code)
        if (
            can_skip_main_entries
            and existing_entry
            and _entry_values_match(existing_entry, entry_defaults)
        ):
            continue

        entry = existing_entry or existing_retest_entries_for_main.get(abx_code)
        if entry is None:
            entry = Final_AntibioticEntry(
                ab_idNum_f_referred=isolate,
                ab_Abx_code=abx_code,
            )
        entry.ab_Abx_code = abx_code
        for field, value in entry_defaults.items():
            setattr(entry, field, value)
        entry.save()

        entry.ab_breakpoints_id.clear()
        bp_applied = False

        # DISK
        if disk_value is not None:
            bp_disk = resolve_bp(
                abx_code,
                specimen_year,
                resolved_site_org,
                "DISK",
            )

            if bp_disk:
                entry.ab_breakpoints_id.set([bp_disk])
                entry.ab_Site_Org = bp_disk.Org 
                entry.ab_R_breakpoint = bp_disk.R_val
                entry.ab_I_breakpoint = bp_disk.I_val
                entry.ab_SDD_breakpoint = bp_disk.SDD_val
                entry.ab_S_breakpoint = bp_disk.S_val
                bp_applied = True

        # MIC
        if mic_value is not None:
            bp_mic = resolve_bp(
                abx_code,
                specimen_year,
                resolved_site_org,
                "MIC",
            )

            if bp_mic:
                entry.ab_breakpoints_id.set([bp_mic])
                entry.ab_Site_Org = bp_mic.Org
                entry.ab_R_breakpoint = bp_mic.R_val
                entry.ab_I_breakpoint = bp_mic.I_val
                entry.ab_SDD_breakpoint = bp_mic.SDD_val
                entry.ab_S_breakpoint = bp_mic.S_val
                entry.ab_Alert_val = bp_mic.Alert_val if alert_mic else ""
                bp_applied = True

        if not bp_applied:
            entry.ab_Site_Org = None
            entry.ab_R_breakpoint = None
            entry.ab_I_breakpoint = None
            entry.ab_SDD_breakpoint = None
            entry.ab_S_breakpoint = None
            entry.ab_Alert_val = ""

        entry.save()

    # =========================
    # RETEST ANTIBIOTICS
    # =========================
    existing_retest_entries = {
        (entry.ab_Retest_Abx_code or "").strip().upper(): entry
        for entry in Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred=isolate,
            ab_Retest_Abx_code__isnull=False,
        )
    }
    existing_main_entries_for_retest = {
        (entry.ab_Abx_code or "").strip().upper(): entry
        for entry in Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred=isolate,
            ab_Abx_code__isnull=False,
        )
    }
    can_skip_retest_entries = (
        old_ars_org == new_ars_org
        and old_specimen_year == specimen_year
    )
    antibiotics_retest = list(_final_antibiotics_for_panel(
        org_code=resolved_ars_org,
        specimen_year=specimen_year,
        retest=True,
        require_org=True,
        existing_whonet_codes=existing_retest_entries.keys(),
        antibiotic_view=antibiotic_view,
    ))

    if not ars_org_is_na:
      for abx in antibiotics_retest:

        abx_code = (abx.Whonet_Abx or "").strip().upper()

        disk_value  = request.POST.get(f"retest_disk_{abx_code}")
        mic_value   = request.POST.get(f"retest_mic_{abx_code}")
        disk_enris  = request.POST.get(f"retest_disk_enris_{abx_code}", "").strip()
        mic_enris   = request.POST.get(f"retest_mic_enris_{abx_code}", "").strip()
        mic_operand = request.POST.get(f"retest_mic_operand_{abx_code}", "").strip()
        alert_mic   = f"retest_alert_mic_{abx_code}" in request.POST

        disk_value = _parse_disk_value(disk_value)

        try:
            mic_value = float(mic_value) if mic_value else None
        except ValueError:
            mic_value = None

        if disk_value is None and mic_value is None:
            for entry in Final_AntibioticEntry.objects.filter(
                ab_idNum_f_referred=isolate,
                ab_Retest_Abx_code=abx_code
            ):
                _clear_retest_antibiotic_data(entry)
                _save_or_delete_antibiotic_entry(entry)
            continue

        entry_defaults = {
            "ab_AccessionNo": isolate.f_AccessionNo,
            "ab_Retest_Antibiotic": abx.Antibiotic,
            "ab_Retest_Abx": abx.Abx_code,
            "ab_Retest_DiskValue": disk_value,
            "ab_Retest_MICValue": mic_value,
            "ab_Retest_Disk_enRIS": disk_enris,
            "ab_Retest_MIC_enRIS": mic_enris,
            "ab_Retest_MIC_operand": mic_operand,
            "ab_Retest_AlertMIC": alert_mic,
        }
        existing_entry = existing_retest_entries.get(abx_code)
        if (
            can_skip_retest_entries
            and existing_entry
            and _entry_values_match(existing_entry, entry_defaults)
        ):
            continue

        entry = existing_entry or existing_main_entries_for_retest.get(abx_code)
        if entry is None:
            entry = Final_AntibioticEntry(
                ab_idNum_f_referred=isolate,
                ab_Retest_Abx_code=abx_code,
            )
        entry.ab_Retest_Abx_code = abx_code
        for field, value in entry_defaults.items():
            setattr(entry, field, value)
        entry.save()

        entry.ab_breakpoints_id.clear()
        ret_bp_applied = False

        if disk_value is not None:
            bp_disk = resolve_bp(
                abx_code,
                specimen_year,
                resolved_ars_org,
                "DISK",
            )

            if bp_disk:
                entry.ab_breakpoints_id.set([bp_disk])
                entry.ab_Ret_Org = bp_disk.Org
                entry.ab_Org_Flag = bool(bp_disk.Emerging_Org_Flag)
                entry.ab_Abx_Flag = bool(bp_disk.Emerging_Abx_Flag)
                entry.ab_Abx_Phenotype = bp_disk.Emerging_Pheno_Flag or ""
                entry.ab_Ret_R_breakpoint = bp_disk.R_val
                entry.ab_Ret_I_breakpoint = bp_disk.I_val
                entry.ab_Ret_SDD_breakpoint = bp_disk.SDD_val
                entry.ab_Ret_S_breakpoint = bp_disk.S_val
                ret_bp_applied = True

        if mic_value is not None:
            bp_mic = resolve_bp(
                abx_code,
                specimen_year,
                resolved_ars_org,
                "MIC",
            )

            if bp_mic:
                entry.ab_breakpoints_id.set([bp_mic])
                entry.ab_Ret_Org = bp_mic.Org
                entry.ab_Org_Flag = bool(bp_mic.Emerging_Org_Flag)
                entry.ab_Abx_Flag = bool(bp_mic.Emerging_Abx_Flag)
                entry.ab_Abx_Phenotype = bp_mic.Emerging_Pheno_Flag or ""
                entry.ab_Ret_R_breakpoint = bp_mic.R_val
                entry.ab_Ret_I_breakpoint = bp_mic.I_val
                entry.ab_Ret_SDD_breakpoint = bp_mic.SDD_val
                entry.ab_Ret_S_breakpoint = bp_mic.S_val
                entry.ab_Retest_Alert_val = bp_mic.Alert_val if alert_mic else ""
                ret_bp_applied = True

        if not ret_bp_applied:
            entry.ab_Ret_Org = None
            entry.ab_Org_Flag = False
            entry.ab_Abx_Flag = False
            entry.ab_Abx_Phenotype = ""
            entry.ab_Ret_R_breakpoint = None
            entry.ab_Ret_I_breakpoint = None
            entry.ab_Ret_SDD_breakpoint = None
            entry.ab_Ret_S_breakpoint = None
            entry.ab_Retest_Alert_val = ""

        entry.save()

    messages.success(request, "Final data saved successfully.")
    next_after_save = (request.POST.get("next_after_save") or "").strip()
    if next_after_save and url_has_allowed_host_and_scheme(
        next_after_save,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_after_save)

    batch_nav = _final_batch_navigation(isolate)
    if batch_nav.get("next_id"):
        return redirect("edit_final_data", id=batch_nav["next_id"])
    return redirect("show_final_table")









############## Lab Result


def _is_nonviable_result(*values):
    """True when organism/result text says the isolate was non-viable."""
    for value in values:
        normalized = (value or "").strip().lower().replace("-", " ")
        if normalized in {"not viable", "non viable", "nonviable"}:
            return True
    return False


def _tested_pdf_abx_code_groups(entries, abx_map, print_order, site_printable, ars_printable):
    tested_codes = set()

    for entry in entries:
        if entry.ab_Abx_code:
            abx_code = abx_map.get(entry.ab_Abx_code.strip().upper())
            if abx_code and abx_code in site_printable:
                tested_codes.add(abx_code)

        if entry.ab_Retest_Abx_code:
            abx_code = abx_map.get(entry.ab_Retest_Abx_code.strip().upper())
            if abx_code and abx_code in ars_printable:
                tested_codes.add(abx_code)

    ordered_codes = [code for code in print_order if code in tested_codes]
    ordered_codes.extend(
        sort_abx_codes_by_antibiotic(
            code for code in tested_codes if code not in ordered_codes
        )
    )
    return ordered_codes, ordered_codes


def _tested_pdf_abx_codes(entries, abx_map, print_order, site_printable, ars_printable):
    site_codes, ars_codes = _tested_pdf_abx_code_groups(
        entries,
        abx_map,
        print_order,
        site_printable,
        ars_printable,
    )
    tested_codes = []
    for code in [*site_codes, *ars_codes]:
        if code not in tested_codes:
            tested_codes.append(code)
    return tested_codes


def _aligned_pdf_abx_codes(site_codes, ars_codes, site_print_order, ars_print_order):
    combined = set(site_codes) | set(ars_codes)
    aligned = []

    for code in [*site_print_order, *ars_print_order]:
        if code in combined and code not in aligned:
            aligned.append(code)

    aligned.extend(
        sort_abx_codes_by_antibiotic(
            code for code in combined if code not in aligned
        )
    )
    return aligned


def _blank_no_organism_report_fields(isolate):
    """Blank only the printed result pane for whichever organism side is n/a."""
    if _is_no_organism(isolate.f_Site_Org) and not _is_nonviable_result(
        isolate.f_Site_Pre,
        isolate.f_Site_OrgName,
        isolate.f_Site_Pos,
    ):
        isolate.f_Site_OrgName = ""

    if _is_no_organism(isolate.f_ars_OrgCode) and not _is_nonviable_result(
        isolate.f_ars_Pre,
        isolate.f_ars_OrgName,
        isolate.f_ars_Post,
    ):
        isolate.f_ars_OrgName = ""

    if _is_no_organism(isolate.f_ars_OrgCode):
        isolate.f_ars_ct_ctl = ""
        isolate.f_ars_tz_tzl = ""
        isolate.f_ars_cn_cni = ""
        isolate.f_ars_ip_ipi = ""


def _chunk_pdf_isolates_for_print(isolates, recommendation_attr, max_per_page=2):
    isolate_list = list(isolates)
    return [
        isolate_list[i:i + max_per_page]
        for i in range(0, len(isolate_list), max_per_page)
    ]


def _pdf_has_long_recommendation(value):
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return False
    numbered_items = len(re.findall(r"(?:^|\s)\d+\.", text))
    return numbered_items >= 4 or len(text) >= 330



@transaction.atomic
def generate_final_batch_pdf_panel_old(request, id):

    # fetch batch isolates
    batch = get_object_or_404(Batch_Table, pk=id)
    changed_fields = _apply_signature_defaults_to_batch(batch)
    if changed_fields:
        batch.save(update_fields=changed_fields)

    # isolates = (
    #     Final_Data.objects
    #     .filter(f_Batch_id=batch)
    #     .order_by("f_bat_seq")
    # )

    # # paginate: 2 isolates per page
    # def chunked(qs, size):
    #     for i in range(0, qs.count(), size):
    #         yield qs[i:i + size]

    # isolate_pages = list(chunked(isolates, 2))

    # sort isolates by accession number, then by f_bat_seq
    def final_ars_sort_key(isolate):
        accession = (isolate.f_AccessionNo or "").strip().upper()

        # Example: 26ARS_DLS0017 -> prefix DLS, number 17
        match = re.search(r"([A-Z]+)(\d+)$", accession)

        if match:
            prefix = match.group(1)
            number = int(match.group(2))
            return (prefix, number)

        # fallback to f_bat_seq if accession format is unusual
        try:
            return ("", int(isolate.f_bat_seq))
        except (TypeError, ValueError):
            return ("", 999999)


    isolates = sorted(
        Final_Data.objects.filter(f_Batch_id=batch),
        key=final_ars_sort_key
    )

    isolate_pages = _chunk_pdf_isolates_for_print(isolates, "f_ars_reco", max_per_page=2)


    # these are the constants
    MAX_COLS = 29
    MAX_ROWS = 2

    def chunk_list(items, size):
        for i in range(0, len(items), size):
            yield items[i:i + size]

    pages_data = []

    # whonet code map
    AntibioticList = HomeAntibioticList()
    abx_map = dict(
        AntibioticList.objects
        .values_list("Whonet_Abx", "Abx_code")
    )
    site_print_order = antibiotic_print_order(show_site=True)
    ars_print_order = antibiotic_print_order(show_ars=True)
    printable_abx_site = set(site_print_order)
    printable_abx_ars = set(ars_print_order)

    # build pdf
    for page_isolates in isolate_pages:
        page_entries = []
        compact_page = any(
            _pdf_has_long_recommendation(getattr(isolate, "f_ars_reco", ""))
            for isolate in page_isolates
        )

        for isolate in page_isolates:
            _blank_no_organism_report_fields(isolate)

            site_org = isolate.f_Site_Org
            ars_org = isolate.f_ars_OrgCode or site_org
            specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None

            # fetch the entries
            entries = Final_AntibioticEntry.objects.filter(
                ab_idNum_f_referred=isolate
            )

            site_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, site_org)
            ars_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, ars_org)

            encoded_site_abx = set()
            encoded_ars_abx = set()

            for e in entries:
                if e.ab_Abx_code:
                    abx = abx_map.get(e.ab_Abx_code.strip().upper())
                    if abx:
                        encoded_site_abx.add(abx)

                if e.ab_Retest_Abx_code:
                    abx = abx_map.get(e.ab_Retest_Abx_code.strip().upper())
                    if abx:
                        encoded_ars_abx.add(abx)
            encoded_print_abx = encoded_site_abx | encoded_ars_abx

            site_candidates = (set(site_panel_abx) | encoded_print_abx) & printable_abx_site
            ars_candidates = (set(ars_panel_abx) | encoded_print_abx) & printable_abx_ars

            site_abx_codes = [
                abx for abx in site_print_order
                if abx in site_candidates
            ]
            site_abx_codes.extend(
                sort_abx_codes_by_antibiotic(
                    abx for abx in site_candidates if abx not in site_abx_codes
                )
            )
            ars_abx_codes = [
                abx for abx in ars_print_order
                if abx in ars_candidates
            ]
            ars_abx_codes.extend(
                sort_abx_codes_by_antibiotic(
                    abx for abx in ars_candidates if abx not in ars_abx_codes
                )
            )
            if _is_no_organism(isolate.f_Site_Org):
                site_abx_codes = []
            if _is_no_organism(isolate.f_ars_OrgCode):
                ars_abx_codes = []
            aligned_abx_codes = _aligned_pdf_abx_codes(
                site_abx_codes,
                ars_abx_codes,
                site_print_order,
                ars_print_order,
            )
            site_abx_codes = aligned_abx_codes
            ars_abx_codes = aligned_abx_codes

            # group the antibiotics based on panels
            grouped_site = {
                abx: {"disk": None, "mic": None}
                for abx in site_abx_codes
            }

            grouped_ars = {
                abx: {"disk": None, "mic": None}
                for abx in ars_abx_codes
            }

           # assign the values into groups
            for e in entries:

                # -sentinel site-
                if e.ab_Abx_code:
                    abx = abx_map.get(e.ab_Abx_code.strip().upper())
                    if abx and abx in grouped_site:
                        # Prefer DISK if a disk value exists
                        if e.ab_Disk_value is not None:
                            grouped_site[abx]["disk"] = e

                        # Otherwise assign MIC only if MIC value exists
                        elif e.ab_MIC_value is not None:
                            grouped_site[abx]["mic"] = e


                # -arsrl / retest-
                if e.ab_Retest_Abx_code:
                    abx = abx_map.get(e.ab_Retest_Abx_code.strip().upper())
                    if abx and abx in grouped_ars:
                        if e.ab_Retest_DiskValue is not None:
                            grouped_ars[abx]["disk"] = e
                        elif e.ab_Retest_MICValue is not None:
                            grouped_ars[abx]["mic"] = e


            site_uses_plus_layout = _organism_type_is_plus(site_org, isolate.f_Site_OrgName)
            site_max_cols = 32 if site_uses_plus_layout else MAX_COLS

            # a chunk for display
            grouped_rows = list(
                chunk_list(list(grouped_site.items()), site_max_cols)
            )[:MAX_ROWS]
            while len(grouped_rows) < MAX_ROWS:
                grouped_rows.append([])

            ars_uses_plus_layout = _organism_type_is_plus(ars_org, isolate.f_ars_OrgName)
            ars_max_cols = 32 if ars_uses_plus_layout else MAX_COLS
            grouped_ars_rows = list(
                chunk_list(list(grouped_ars.items()), ars_max_cols)
            )[:MAX_ROWS]
            while len(grouped_ars_rows) < MAX_ROWS:
                grouped_ars_rows.append([])

            site_group_count = len(grouped_rows)
            ars_group_count = len(grouped_ars_rows)

            page_entries.append({
                "isolate": isolate,
                "grouped_rows": grouped_rows,
                "grouped_ars_rows": grouped_ars_rows,
                "site_uses_plus_layout": site_uses_plus_layout,
                "ars_uses_plus_layout": ars_uses_plus_layout,
                "patient_rowspan": (site_group_count + ars_group_count) * 5,
                "site_detail_rowspan": (site_group_count * 5) - 1,
                "ars_detail_rowspan": (ars_group_count * 5) - 1,
                "compact_page": compact_page,
            })

        pages_data.append(page_entries)

    # render the pdf
    context = {
        "batch": batch,
        "pages": pages_data,
        "now": timezone.now(),
        "logo_path": static("assets/img/brand/arsplogo.jpg"),
    }

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'filename="Batch_Panel_Report.pdf"'

    template = get_template("home_final/Lab_result_final.html")

    html = template.render(context)

    pisa.CreatePDF(
        html,
        dest=response,
        link_callback=link_callback
    )

    return response


@transaction.atomic
def generate_final_batch_pdf(request, id):
    return generate_final_batch_pdf_panel_old(request, id)

    batch = get_object_or_404(Batch_Table, pk=id)
    isolates = (
        Final_Data.objects
        .filter(f_Batch_id=batch)
        .order_by("f_bat_seq")
    )

    def chunked(qs, size):
        for i in range(0, qs.count(), size):
            yield qs[i:i + size]

    def chunk_list(items, size):
        for i in range(0, len(items), size):
            yield items[i:i + size]

    def fixed_rows(grouped_antibiotics, max_cols=MAX_COLS):
        if not grouped_antibiotics:
            return [
                [("", {"disk": None, "mic": None}) for _ in range(max_cols)]
                for _ in range(MAX_ROWS)
            ]
        rows = list(chunk_list(list(grouped_antibiotics.items()), max_cols))[:MAX_ROWS]
        while len(rows) < MAX_ROWS:
            rows.append([])
        return rows

    MAX_COLS = 29
    MAX_ROWS = 2
    isolate_pages = _chunk_pdf_isolates_for_print(list(isolates), "f_ars_reco", max_per_page=2)
    pages_data = []

    AntibioticList = HomeAntibioticList()
    antibiotic_order = list(
        AntibioticList.objects
        .exclude(Abx_code__exact="")
        .values("Whonet_Abx", "Abx_code")
        .order_by("id")
    )
    abx_map = {
        (row["Whonet_Abx"] or "").strip().upper(): (row["Abx_code"] or "").strip()
        for row in antibiotic_order
        if (row["Whonet_Abx"] or "").strip() and (row["Abx_code"] or "").strip()
    }
    site_print_order = antibiotic_print_order(show_site=True)
    ars_print_order = antibiotic_print_order(show_ars=True)
    site_printable = set(site_print_order)
    ars_printable = set(ars_print_order)

    for page_isolates in isolate_pages:
        page_entries = []
        compact_page = any(
            _pdf_has_long_recommendation(getattr(isolate, "f_ars_reco", ""))
            for isolate in page_isolates
        )

        for isolate in page_isolates:
            _blank_no_organism_report_fields(isolate)
            entries = list(Final_AntibioticEntry.objects.filter(
                ab_idNum_f_referred=isolate
            ))

            site_org = (isolate.f_Site_Org or "").strip()
            ars_org = (isolate.f_ars_OrgCode or "").strip() or site_org
            specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None

            site_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, site_org)
            ars_panel_abx = get_breakpoint_panel_abx_codes(specimen_year, ars_org)
            encoded_site_abx = {
                abx_map.get((entry.ab_Abx_code or "").strip().upper())
                for entry in entries
                if entry.ab_Abx_code
            } - {None, ""}
            encoded_ars_abx = {
                abx_map.get((entry.ab_Retest_Abx_code or "").strip().upper())
                for entry in entries
                if entry.ab_Retest_Abx_code
            } - {None, ""}
            encoded_print_abx = encoded_site_abx | encoded_ars_abx

            site_candidates = (set(site_panel_abx) | encoded_print_abx) & site_printable
            ars_candidates = (set(ars_panel_abx) | encoded_print_abx) & ars_printable

            site_abx_codes = [abx for abx in site_print_order if abx in site_candidates]
            site_abx_codes.extend(
                sort_abx_codes_by_antibiotic(
                    abx for abx in site_candidates if abx not in site_abx_codes
                )
            )
            ars_abx_codes = [abx for abx in ars_print_order if abx in ars_candidates]
            ars_abx_codes.extend(
                sort_abx_codes_by_antibiotic(
                    abx for abx in ars_candidates if abx not in ars_abx_codes
                )
            )
            if _is_no_organism(isolate.f_Site_Org):
                site_abx_codes = []
            if _is_no_organism(isolate.f_ars_OrgCode):
                ars_abx_codes = []

            aligned_abx_codes = _aligned_pdf_abx_codes(
                site_abx_codes,
                ars_abx_codes,
                site_print_order,
                ars_print_order,
            )
            site_abx_codes = aligned_abx_codes
            ars_abx_codes = aligned_abx_codes

            grouped_site = {abx: {"disk": None, "mic": None} for abx in site_abx_codes}
            grouped_ars = {abx: {"disk": None, "mic": None} for abx in ars_abx_codes}

            for e in entries:
                site_abx = abx_map.get((e.ab_Abx_code or "").strip().upper())
                if site_abx in grouped_site:
                    if e.ab_Disk_value is not None:
                        grouped_site[site_abx]["disk"] = e
                    if e.ab_MIC_value is not None:
                        grouped_site[site_abx]["mic"] = e

                ars_abx = abx_map.get((e.ab_Retest_Abx_code or "").strip().upper())
                if ars_abx in grouped_ars:
                    if e.ab_Retest_DiskValue is not None:
                        grouped_ars[ars_abx]["disk"] = e
                    if e.ab_Retest_MICValue is not None:
                        grouped_ars[ars_abx]["mic"] = e

            site_uses_plus_layout = _organism_type_is_plus(site_org, isolate.f_Site_OrgName)
            ars_uses_plus_layout = _organism_type_is_plus(ars_org, isolate.f_ars_OrgName)
            site_max_cols = 32 if site_uses_plus_layout else MAX_COLS
            ars_max_cols = 32 if ars_uses_plus_layout else MAX_COLS
            grouped_rows = fixed_rows(grouped_site, site_max_cols)
            grouped_ars_rows = fixed_rows(grouped_ars, ars_max_cols)
            site_group_count = len(grouped_rows)
            ars_group_count = len(grouped_ars_rows)

            page_entries.append({
                "isolate": isolate,
                "grouped_rows": grouped_rows,
                "grouped_ars_rows": grouped_ars_rows,
                "site_uses_plus_layout": site_uses_plus_layout,
                "ars_uses_plus_layout": ars_uses_plus_layout,
                "patient_rowspan": (site_group_count + ars_group_count) * 5,
                "site_detail_rowspan": (site_group_count * 5) - 1,
                "ars_detail_rowspan": (ars_group_count * 5) - 1,
                "compact_page": compact_page,
            })

        pages_data.append(page_entries)

    context = {
        "batch": batch,
        "pages": pages_data,
        "now": timezone.now(),
        "logo_path": static("assets/img/brand/arsplogo.jpg"),
    }

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'filename="Batch_Panel_Report.pdf"'

    template = get_template("home_final/Lab_result_final.html")
    html = template.render(context)

    pisa.CreatePDF(
        html,
        dest=response,
        link_callback=link_callback
    )

    return response


# @login_required(login_url="login")
# def export_concordance_report_pdf_v2(request, report_id):
#     report = get_object_or_404(
#         ConcordanceReport.objects.select_related("batch"),
#         id=report_id,
#         final_data__isnull=True
#     )

#     context = _build_batch_concordance_context(report)
#     context["now"] = datetime.now().strftime("%d %B %Y").upper()

#     template = get_template("home_final/concordance_report_pdf.html")
#     html = template.render(context)

#     response = HttpResponse(content_type="application/pdf")
#     response["Content-Disposition"] = (
#         f'inline; filename=Batch_{report.batch.bat_Batch_Name}_Concordance.pdf'
#     )

#     pisa_status = pisa.CreatePDF(html, dest=response)
#     if pisa_status.err:
#         return HttpResponse("PDF generation error", status=500)

#     return response


@login_required(login_url="login")
def export_concordance_report_pdf_v2(request, report_id):
    report = get_object_or_404(
        ConcordanceReport.objects.select_related("batch"),
        id=report_id,
        final_data__isnull=True
    )

    context = _build_batch_concordance_context(report)
    context["now"] = datetime.now().strftime("%d %B %Y").upper()
    context["header_path"] = static("assets/img/brand/header_ed.png")
    context["footer_path"] = static("assets/img/brand/footer.png")

    template = get_template("home_final/concordance_report_pdf.html")
    html = template.render(context)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename=Batch_{report.batch.bat_Batch_Name}_Concordance.pdf'
    )

    pisa_status = pisa.CreatePDF(
        html,
        dest=response,
        link_callback=link_callback,
    )
    if pisa_status.err:
        return HttpResponse("PDF generation error", status=500)

    return response




############ filter antibiotics for panels


# used in breakopints dropdown for organism 
@require_GET
def get_organism_group(request):
    org_code = request.GET.get("org_code")

    if not org_code:
        return JsonResponse({"genus_group": ""})

    try:
        OrganismList = HomeOrganismList()
        organism = OrganismList.objects.get(
            Whonet_Org_Code=org_code
        )
        return JsonResponse({
            "genus_group": organism.Genus_Group or ""
        })
    except HomeOrganismList().DoesNotExist:
        return JsonResponse({"genus_group": ""})


## aut fill  abx code and tier

@login_required(login_url="login")
def get_antibiotic_name(request):
    whonet_code = request.GET.get("whonet")
    AntibioticList = HomeAntibioticList()
    try:
        abx = AntibioticList.objects.get(Whonet_Abx=whonet_code)
        return JsonResponse({"name": abx.Antibiotic})
    except AntibioticList.DoesNotExist:
        return JsonResponse({"name": ""})

@require_GET
def get_antibiotic_details(request):
    whonet_abx = request.GET.get("whonet_abx", "").strip()
    
    # Use filter().first() to avoid DoesNotExist exceptions crashing the AJAX call
    abx = HomeAntibioticList().objects.filter(Whonet_Abx=whonet_abx).first()
    
    if abx:
        return JsonResponse({
            "antibiotic": abx.Antibiotic,
            "abx_code": abx.Abx_code,
            "tier": abx.Tier,
        })
    
    return JsonResponse({"error": "Not found"}, status=404)


# @login_required(login_url="login")
# def ajax_filter_antibiotics(request):
#     if not request.user.is_authenticated:
#         return JsonResponse({"error": "Unauthorized"}, status=401)


#     isolate_id = request.GET.get("isolate_id")
#     org_code   = request.GET.get("org", "").strip().lower()
#     retest     = request.GET.get("retest") == "1"

#     isolate = get_object_or_404(Final_Data, pk=isolate_id)

#     # ================= DETERMINE BREAKPOINT YEAR =================
#     specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None

#     if specimen_year:
#         breakpoint_year = (
#             BreakpointsTable.objects
#             .filter(Year__lte=str(specimen_year))
#             .order_by("-Year")
#             .values_list("Year", flat=True)
#             .first()
#         )
#     else:
#         breakpoint_year = (
#             BreakpointsTable.objects
#             .order_by("-Year")
#             .values_list("Year", flat=True)
#             .first()
#         )

#     # If no breakpoints exist at all, return empty safely
#     if not breakpoint_year:
#         return JsonResponse({"antibiotics": []})

#     # ================= FILTER ANTIBIOTICS =================
#     antibiotics = get_filtered_antibiotics(
#         breakpoint_year,
#         org_code,     # ALWAYS organism code
#         retest=retest
#     )

#     # ================= FETCH EXISTING FINAL ENTRIES =================
#     entries = Final_AntibioticEntry.objects.filter(
#         ab_idNum_f_referred=isolate
#     )

#     # Map entries by WHONET code
#     entry_map = {}
#     for e in entries:
#         code = e.ab_Retest_Abx_code if retest else e.ab_Abx_code
#         if code:
#             entry_map[code.strip().upper()] = e

#     # ================= BUILD RESPONSE =================
#     payload = []

#     for abx in antibiotics:
#         code = (abx.Whonet_Abx or "").strip().upper()
#         entry = entry_map.get(code)

#         if retest:
#             payload.append({
#                 "whonet": code,
#                 "name": abx.Antibiotic,
#                 "is_disk": abx.Disk_Abx,

#                 # RETEST VALUES (FINAL)
#                 "disk": entry.ab_Retest_DiskValue if entry else "",
#                 "disk_enris": entry.ab_Retest_Disk_enRIS if entry else "",
#                 "mic": entry.ab_Retest_MICValue if entry else "",
#                 "mic_enris": entry.ab_Retest_MIC_enRIS if entry else "",
#                 "mic_operand": entry.ab_Retest_MIC_operand if entry else "",
#                 "alert_mic": entry.ab_Retest_AlertMIC if entry else False,
#             })
#         else:
#             payload.append({
#                 "whonet": code,
#                 "name": abx.Antibiotic,
#                 "is_disk": abx.Disk_Abx,

#                 # MAIN VALUES (FINAL)
#                 "disk": entry.ab_Disk_value if entry else "",
#                 "disk_enris": entry.ab_Disk_enRIS if entry else "",
#                 "mic": entry.ab_MIC_value if entry else "",
#                 "mic_enris": entry.ab_MIC_enRIS if entry else "",
#                 "mic_operand": entry.ab_MIC_operand if entry else "",
#                 "alert_mic": entry.ab_AlertMIC if entry else False,
#             })

#     return JsonResponse({"antibiotics": payload})





# @login_required(login_url="login")
# def ajax_filter_antibiotics(request):
#     if not request.user.is_authenticated:
#         return JsonResponse({"error": "Unauthorized"}, status=401)


#     isolate_id = request.GET.get("isolate_id")
#     org_code   = request.GET.get("org", "").strip().lower()
#     retest     = request.GET.get("retest") == "1"

#     isolate = get_object_or_404(Final_Data, pk=isolate_id)

#     # ================= DETERMINE BREAKPOINT YEAR =================
#     specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None

#     if specimen_year:
#         breakpoint_year = (
#             BreakpointsTable.objects
#             .filter(Year__lte=str(specimen_year))
#             .order_by("-Year")
#             .values_list("Year", flat=True)
#             .first()
#         )
#     else:
#         breakpoint_year = (
#             BreakpointsTable.objects
#             .order_by("-Year")
#             .values_list("Year", flat=True)
#             .first()
#         )

#     # If no breakpoints exist at all, return empty safely
#     if not breakpoint_year:
#         return JsonResponse({"antibiotics": []})

#     # ================= FILTER ANTIBIOTICS =================
#     antibiotics = get_filtered_antibiotics(
#         breakpoint_year,
#         org_code,     # ALWAYS organism code
#         retest=retest
#     )

#     # ================= FETCH EXISTING FINAL ENTRIES =================
#     entries = Final_AntibioticEntry.objects.filter(
#         ab_idNum_f_referred=isolate
#     )

#     # Map entries by WHONET code
#     entry_map = {}
#     for e in entries:
#         code = e.ab_Retest_Abx_code if retest else e.ab_Abx_code
#         if code:
#             entry_map[code.strip().upper()] = e

#     # ================= BUILD RESPONSE =================
#     payload = []

#     for abx in antibiotics:
#         code = (abx.Whonet_Abx or "").strip().upper()
#         entry = entry_map.get(code)

#         if retest:
#             payload.append({
#                 "whonet": code,
#                 "name": abx.Antibiotic,
#                 "is_disk": abx.Disk_Abx,

#                 # RETEST VALUES (FINAL)
#                 "disk": entry.ab_Retest_DiskValue if entry else "",
#                 "disk_enris": entry.ab_Retest_Disk_enRIS if entry else "",
#                 "mic": entry.ab_Retest_MICValue if entry else "",
#                 "mic_enris": entry.ab_Retest_MIC_enRIS if entry else "",
#                 "mic_operand": entry.ab_Retest_MIC_operand if entry else "",
#                 "alert_mic": entry.ab_Retest_AlertMIC if entry else False,
#             })
#         else:
#             payload.append({
#                 "whonet": code,
#                 "name": abx.Antibiotic,
#                 "is_disk": abx.Disk_Abx,

#                 # MAIN VALUES (FINAL)
#                 "disk": entry.ab_Disk_value if entry else "",
#                 "disk_enris": entry.ab_Disk_enRIS if entry else "",
#                 "mic": entry.ab_MIC_value if entry else "",
#                 "mic_enris": entry.ab_MIC_enRIS if entry else "",
#                 "mic_operand": entry.ab_MIC_operand if entry else "",
#                 "alert_mic": entry.ab_AlertMIC if entry else False,
#             })

#     return JsonResponse({"antibiotics": payload})




@login_required(login_url="login")
def ajax_filter_antibiotics(request):
    # 1. Get ID from GET parameters (passed by JS)
    isolate_id = request.GET.get("isolate_id")
    
    # 2. Fallback to session ONLY if GET is missing
    if not isolate_id:
        isolate_id = request.session.get("current_final_isolate_id")

    if not isolate_id:
        return JsonResponse({"error": "No Isolate ID provided"}, status=400)



    org_code   = request.GET.get("org", "").strip().lower()
    retest     = request.GET.get("retest") == "1"
    antibiotic_view = _antibiotic_view_mode(request)

    isolate = get_object_or_404(Final_Data, pk=isolate_id)

    # ================= DETERMINE BREAKPOINT YEAR =================
    specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None

    if specimen_year:
        breakpoint_year = (
            BreakpointsTable.objects
            .filter(Year__lte=str(specimen_year))
            .order_by("-Year")
            .values_list("Year", flat=True)
            .first()
        )
    else:
        breakpoint_year = (
            BreakpointsTable.objects
            .order_by("-Year")
            .values_list("Year", flat=True)
            .first()
        )

    # ================= FETCH EXISTING FINAL ENTRIES =================
    entries = Final_AntibioticEntry.objects.filter(
        ab_idNum_f_referred=isolate
    )
    existing_codes = entries.exclude(
        **{f"{'ab_Retest_Abx_code' if retest else 'ab_Abx_code'}__isnull": True}
    ).values_list("ab_Retest_Abx_code" if retest else "ab_Abx_code", flat=True)

    # ================= FILTER ANTIBIOTICS =================
    antibiotics = _final_antibiotics_for_panel(
        org_code=org_code,
        specimen_year=specimen_year,
        show_site=not retest,
        retest=retest,
        existing_whonet_codes=existing_codes,
        antibiotic_view=antibiotic_view,
    )

    # Map entries by WHONET code
    entry_map = {}
    for e in entries:
        code = e.ab_Retest_Abx_code if retest else e.ab_Abx_code
        if code:
            entry_map[code.strip().upper()] = e

    # ================= BUILD RESPONSE =================
    payload = []

    for abx in antibiotics:
        code = (abx.Whonet_Abx or "").strip().upper()
        entry = entry_map.get(code)

        if retest:
            payload.append({
                "whonet": code,
                "name": abx.Antibiotic,
                "is_disk": abx.Disk_Abx,

                # RETEST VALUES (FINAL)
                "disk": entry.ab_Retest_DiskValue if entry else "",
                "disk_enris": entry.ab_Retest_Disk_enRIS if entry else "",
                "mic": entry.ab_Retest_MICValue if entry else "",
                "mic_enris": entry.ab_Retest_MIC_enRIS if entry else "",
                "mic_operand": entry.ab_Retest_MIC_operand if entry else "",
                "alert_mic": entry.ab_Retest_AlertMIC if entry else False,
            })
        else:
            payload.append({
                "whonet": code,
                "name": abx.Antibiotic,
                "is_disk": abx.Disk_Abx,

                # MAIN VALUES (FINAL)
                "disk": entry.ab_Disk_value if entry else "",
                "disk_enris": entry.ab_Disk_enRIS if entry else "",
                "mic": entry.ab_MIC_value if entry else "",
                "mic_enris": entry.ab_MIC_enRIS if entry else "",
                "mic_operand": entry.ab_MIC_operand if entry else "",
                "alert_mic": entry.ab_AlertMIC if entry else False,
            })

    return JsonResponse({"antibiotics": payload})


# @login_required(login_url="login")
# def ajax_filter_antibiotics(request):
#     # 1. Get ID from GET parameters (passed by JS)
#     isolate_id = request.GET.get("isolate_id")
    
#     # 2. Fallback to session ONLY if GET is missing
#     if not isolate_id:
#         isolate_id = request.session.get("current_final_isolate_id")

#     if not isolate_id:
#         return JsonResponse({"error": "No Isolate ID provided"}, status=400)

#     # 3. Fetch the isolate safely
#     isolate = get_object_or_404(Final_Data, pk=isolate_id)

#     org_code = request.GET.get("org", "").strip().upper()
#     retest   = request.GET.get("retest") == "1"

#     # ================= DETERMINE BREAKPOINT YEAR =================
#     specimen_year = isolate.f_Spec_Date.year if isolate.f_Spec_Date else None

#     if specimen_year:
#         breakpoint_year = (
#             BreakpointsTable.objects
#             .filter(Year__lte=str(specimen_year))
#             .order_by("-Year")
#             .values_list("Year", flat=True)
#             .first()
#         )
#     else:
#         breakpoint_year = (
#             BreakpointsTable.objects
#             .order_by("-Year")
#             .values_list("Year", flat=True)
#             .first()
#         )

#     if not breakpoint_year:
#         return JsonResponse({"antibiotics": []})

#     # ================= FILTER ANTIBIOTICS =================
#     antibiotics = get_filtered_antibiotics(
#         breakpoint_year,
#         org_code,
#         retest=retest
#     )

#     # ================= FETCH EXISTING FINAL ENTRIES =================
#     entries = Final_AntibioticEntry.objects.filter(
#         ab_idNum_f_referred=isolate
#     )

#     entry_map = {}
#     for e in entries:
#         code = e.ab_Retest_Abx_code if retest else e.ab_Abx_code
#         if code:
#             entry_map[code.strip().upper()] = e

#     # ================= BUILD RESPONSE =================
#     payload = []

#     for abx in antibiotics:
#         code = (abx.Whonet_Abx or "").strip().upper()
#         entry = entry_map.get(code)

#         if retest:
#             payload.append({
#                 "whonet": code,
#                 "name": abx.Antibiotic,
#                 "is_disk": abx.Disk_Abx,

#                 "disk": entry.ab_Retest_DiskValue if entry else "",
#                 "disk_enris": entry.ab_Retest_Disk_enRIS if entry else "",
#                 "mic": entry.ab_Retest_MICValue if entry else "",
#                 "mic_enris": entry.ab_Retest_MIC_enRIS if entry else "",
#                 "mic_operand": entry.ab_Retest_MIC_operand if entry else "",
#                 "alert_mic": entry.ab_Retest_AlertMIC if entry else False,
#             })
#         else:
#             payload.append({
#                 "whonet": code,
#                 "name": abx.Antibiotic,
#                 "is_disk": abx.Disk_Abx,

#                 "disk": entry.ab_Disk_value if entry else "",
#                 "disk_enris": entry.ab_Disk_enRIS if entry else "",
#                 "mic": entry.ab_MIC_value if entry else "",
#                 "mic_enris": entry.ab_MIC_enRIS if entry else "",
#                 "mic_operand": entry.ab_MIC_operand if entry else "",
#                 "alert_mic": entry.ab_AlertMIC if entry else False,
#             })

#     return JsonResponse({"antibiotics": payload})






########## organism filtering
@login_required(login_url="login")
def get_organism_name(request):
    org_code = request.GET.get("org_code")
    field_key = request.GET.get("field_key")

    if not org_code or not field_key:
        return JsonResponse({"error": "Missing parameters"}, status=400)

    org = HomeOrganismList().objects.filter(
        Whonet_Org_Code=org_code
    ).values().first()

    if not org:
        return JsonResponse({"error": "Organism not found"}, status=404)

    if field_key not in org:
        return JsonResponse({"error": "Invalid field_key"}, status=400)

    return JsonResponse({field_key: org[field_key]})




@login_required(login_url="login")
def download_combined_final_table(request):
    """
    Export FINAL DATA + FINAL ANTIBIOTIC ENTRIES into one wide CSV
    """
    date_from = parse_date(request.GET.get("date_from", "").strip() or "")
    date_to = parse_date(request.GET.get("date_to", "").strip() or "")

    final_data_entries = (
        Final_Data.objects
        .prefetch_related("final_entries")
        .all()
    )
    if date_from or date_to:
        entry_filter = Q()
        fallback_filter = Q(f_Date_of_Entry__isnull=True)

        if date_from:
            entry_filter &= Q(f_Date_of_Entry__date__gte=date_from)
            fallback_filter &= Q(f_Spec_Date__gte=date_from)

        if date_to:
            entry_filter &= Q(f_Date_of_Entry__date__lte=date_to)
            fallback_filter &= Q(f_Spec_Date__lte=date_to)

        final_data_entries = final_data_entries.filter(entry_filter | fallback_filter)

    classification_fields = [
        "Class_AccessionNo",
        "Class_Chk_Emerging",
        "Class_Chk_Structured",
        "Class_Chk_Satscan",
        "Class_Chk_Serotyping",
        "Class_Chk_GHRU_all",
        "Class_Chk_GHRU_Neo",
        "Class_Chk_EGASP",
        "Class_Chk_Tricycle",
        "Class_Chk_Pulsenet",
        "Class_Chk_Tulip",
    ]
    classification_by_final_id = {
        item.Class_idNumReferred_id: item
        for item in Classification_Table.objects.filter(
            Class_idNumReferred__in=final_data_entries
        ).only("Class_idNumReferred_id", *classification_fields)
    }

    # --------------------------------------------------
    # Collect UNIQUE antibiotics (main + retest)
    # --------------------------------------------------
    unique_abx_codes = set()

    for abx, ret in (
        Final_AntibioticEntry.objects
        .filter(ab_idNum_f_referred__in=final_data_entries)
        .values_list("ab_Abx_code", "ab_Retest_Abx_code")
        .distinct()
    ):
        if abx:
            unique_abx_codes.add(abx.upper())
        if ret:
            unique_abx_codes.add(ret.upper())

    sorted_antibiotics = sorted(unique_abx_codes)

    # --------------------------------------------------
    # HTTP CSV RESPONSE
    # --------------------------------------------------
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="final_combined_data.csv"'
    response.write("\ufeff")  # UTF-8 BOM

    writer = csv.writer(response)

    # --------------------------------------------------
    # STATIC FIELDS (Final_Data)
    # --------------------------------------------------
    static_fields = [
        "f_bat_seq",
        "f_Batch_id",
        "f_Hide",
        "f_Batch_Name",
        "f_Batch_Code",
        "f_Date_of_Entry",
        "f_Date_Modified",
        "f_RefNo",
        "f_BatchNo",
        "f_Total_batch",
        "f_AccessionNo",
        "f_AccessionNoGen",
        "f_SiteCode",
        "f_Site_Name",
        "f_Referral_Date",

        "f_Patient_ID",
        "f_First_Name",
        "f_Mid_Name",
        "f_Last_Name",
        "f_Date_Birth",
        "f_Age",
        "f_Emerging_Flag_Age",
        "f_Sex",
        "f_Date_Admis",
        "f_Nosocomial",
        "f_Diagnosis",
        "f_Diagnosis_ICD10",
        "f_Ward",
        "f_Ward_Type",
        "f_Service_Type",

        "f_Spec_Num",
        "f_Spec_Date",
        "f_Spec_Type",
        "f_Spec_Emerging",
        "f_Reason",
        "f_Growth",
        "f_Urine_ColCt",

        "f_ampC",
        "f_ESBL",
        "f_CARB",
        "f_MBL",
        "f_BL",
        "f_MR",
        "f_mecA",
        "f_ICR",
        "f_OtherResMech",

        "f_Site_Pre",
        "f_Site_Pre_ed",
        "f_Site_Org",
        "f_Site_OrgName",
        "f_Site_Pos",
        "f_Site_Pos_ed",
        "f_Comments",

        "f_ars_ampC",
        "f_ars_ESBL",
        "f_ars_CARB",
        "f_ars_ECIM",
        "f_ars_MCIM",
        "f_ars_EC_MCIM",
        "f_ars_MBL",
        "f_ars_BL",
        "f_ars_MR",
        "f_ars_mecA",
        "f_ars_ICR",
        "f_ars_Pre",
        "f_ars_Pre_ed",
        "f_ars_Post",
        "f_ars_Post_ed",
        "f_ars_OrgCode",
        "f_ars_OrgName",
        "f_ars_ct_ctl",
        "f_ars_tz_tzl",
        "f_ars_cn_cni",
        "f_ars_ip_ipi",
        "f_ars_reco_Code",
        "f_ars_description",
        "f_ars_reco",

        "f_arsp_Encoder",
        "f_arsp_Enc_Lic",
        "f_arsp_Checker",
        "f_arsp_Chec_Lic",
        "f_arsp_Verifier",
        "f_arsp_Ver_Lic",
        "f_arsp_LabManager",
        "f_arsp_Lab_Lic",
        "f_arsp_Head",
        "f_arsp_Head_Lic",
        "f_Date_Accomplished_ARSP",

        "f_x_mrse",
        "f_x_mrsamrse",
        "f_x_entbac",
        "f_edta",
    ]

    # --------------------------------------------------
    # HEADER
    # --------------------------------------------------
    header = static_fields[:] + classification_fields

    for abx in sorted_antibiotics:
        header.extend([
            abx,
            f"{abx}_RIS",
            f"{abx}_RT",
            f"{abx}_RT_RIS",
        ])

    writer.writerow(header)

    # --------------------------------------------------
    # ROW DATA
    # --------------------------------------------------
    for final_obj in final_data_entries:

        row = [
            getattr(final_obj, field, "")
            for field in static_fields
        ]
        classification = classification_by_final_id.get(final_obj.id)
        row.extend([
            getattr(classification, field, "")
            if classification else ""
            for field in classification_fields
        ])

        abx_data = {}

        for entry in final_obj.final_entries.all():

            # MAIN RESULT
            if entry.ab_Abx_code:
                code = entry.ab_Abx_code.upper()
                abx_data.setdefault(code, {})

                if entry.ab_Disk_value is not None or entry.ab_MIC_value is not None:
                    val = (
                        entry.ab_Disk_value
                        if entry.ab_Disk_value is not None
                        else f"{entry.ab_MIC_operand or ''}{entry.ab_MIC_value}"
                    )
                    ris = entry.ab_Disk_enRIS or entry.ab_MIC_enRIS

                    abx_data[code].update({
                        "VAL": val,
                        "RIS": ris,
                    })

            # RETEST RESULT
            if entry.ab_Retest_Abx_code:
                code = entry.ab_Retest_Abx_code.upper()
                abx_data.setdefault(code, {})

                if entry.ab_Retest_DiskValue is not None or entry.ab_Retest_MICValue is not None:
                    rt_val = (
                        entry.ab_Retest_DiskValue
                        if entry.ab_Retest_DiskValue is not None
                        else f"{entry.ab_Retest_MIC_operand or ''}{entry.ab_Retest_MICValue}"
                    )
                    rt_ris = entry.ab_Retest_Disk_enRIS or entry.ab_Retest_MIC_enRIS

                    abx_data[code].update({
                        "RT_VAL": rt_val,
                        "RT_RIS": rt_ris,
                    })

        # Expand antibiotic columns
        for abx in sorted_antibiotics:
            data = abx_data.get(abx, {})

            row.extend([
                data.get("VAL", ""),
                data.get("RIS", ""),
                data.get("RT_VAL", ""),
                data.get("RT_RIS", ""),
            ])

        writer.writerow(row)

    return response



@login_required(login_url="login")
def final_abxentry_view(request):
    """
    Displays ORDINARY (non-retest) antibiotic results
    from Final_AntibioticEntry grouped by accession number.
    """

    # Only FINAL antibiotic entries (exclude retest antibiotics)
    entries = (
        Final_AntibioticEntry.objects
        .filter(ab_Retest_Abx_code__isnull=True)
        .select_related("ab_idNum_f_referred")
    )

    abx_data = {}
    abx_codes = set()

    for entry in entries:
        final_obj = entry.ab_idNum_f_referred
        accession_no = final_obj.f_AccessionNo

        abx_code = entry.ab_Abx_code  # ordinary antibiotic only

        # Determine value (Disk preferred, else MIC)
        value = (
            entry.ab_Disk_value
            if entry.ab_Disk_value is not None
            else entry.ab_MIC_value
        )

        # RIS (prefer disk, else MIC)
        ris = (
            entry.ab_Disk_enRIS
            if entry.ab_Disk_enRIS
            else entry.ab_MIC_enRIS
        )

        operand = entry.ab_MIC_operand or None

        if accession_no not in abx_data:
            abx_data[accession_no] = {}

        if abx_code:
            abx_data[accession_no][abx_code] = {
                "value": value,
                "RIS": ris,
                "Operand": operand,
            }
            abx_codes.add(abx_code)

    context = {
        "abx_data": abx_data,
        "abx_codes": sorted(abx_codes),
    }

    return render(
        request,
        "home_final/AntibioticentryFinal.html",
        context
    )


@login_required(login_url="login")
def export_final_antibiotic_entries(request):
    """
    Export ALL Final_AntibioticEntry records to Excel
    aligned with Final_Data + Final_AntibioticEntry models.
    """

    objects = (
        Final_AntibioticEntry.objects
        .select_related("ab_idNum_f_referred")
        .all()
    )

    data = []

    for obj in objects:
        final = obj.ab_idNum_f_referred

        data.append({
            # ================= LINKED FINAL DATA =================
            "Final_AccessionNo": final.f_AccessionNo if final else None,
            "Final_Batch_Code": final.f_Batch_Code if final else None,
            "Final_Site_Name": final.f_Site_Name if final else None,
            "Final_Site_Org": final.f_Site_Org if final else None,
            "Final_ARS_Org": final.f_ars_OrgCode if final else None,

            ######### flag
            "Org_Flag": obj.ab_Org_Flag,
            "Abx_Flag": obj.ab_Abx_Flag,
            "Phenotype": obj.ab_Abx_Phenotype,


            # ================= ANTIBIOTIC META =================
            "Antibiotic": obj.ab_Antibiotic,
            "Abx_code": obj.ab_Abx_code,
            "Abx": obj.ab_Abx,
            "Disk_Abx": obj.ab_Disk_Abx,

            # ================= SENTINEL SITE RESULTS =================
            "Disk_value": obj.ab_Disk_value,
            "Disk_RIS": obj.ab_Disk_RIS,
            "Disk_enRIS": obj.ab_Disk_enRIS,

            "MIC_operand": obj.ab_MIC_operand,
            "MIC_value": obj.ab_MIC_value,
            "MIC_RIS": obj.ab_MIC_RIS,
            "MIC_enRIS": obj.ab_MIC_enRIS,

            "Alert_MIC": obj.ab_AlertMIC,
            "Alert_MIC_Value": obj.ab_Alert_val,

            # ================= BREAKPOINTS =================
            "R_breakpoint": obj.ab_R_breakpoint,
            "I_breakpoint": obj.ab_I_breakpoint,
            "SDD_breakpoint": obj.ab_SDD_breakpoint,
            "S_breakpoint": obj.ab_S_breakpoint,

            # ================= ARSRL / RETEST =================
            "Retest_Antibiotic": obj.ab_Retest_Antibiotic,
            "Retest_Abx_code": obj.ab_Retest_Abx_code,
            "Retest_Abx": obj.ab_Retest_Abx,

            "Retest_DiskValue": obj.ab_Retest_DiskValue,
            "Retest_Disk_RIS": obj.ab_Retest_Disk_RIS,
            "Retest_Disk_enRIS": obj.ab_Retest_Disk_enRIS,

            "Retest_MIC_operand": obj.ab_Retest_MIC_operand,
            "Retest_MICValue": obj.ab_Retest_MICValue,
            "Retest_MIC_RIS": obj.ab_Retest_MIC_RIS,
            "Retest_MIC_enRIS": obj.ab_Retest_MIC_enRIS,

            "Retest_Alert_MIC": obj.ab_Retest_AlertMIC,
            "Retest_Alert_Value": obj.ab_Retest_Alert_val,

            "Ret_R_breakpoint": obj.ab_Ret_R_breakpoint,
            "Ret_I_breakpoint": obj.ab_Ret_I_breakpoint,
            "Ret_SDD_breakpoint": obj.ab_Ret_SDD_breakpoint,
            "Ret_S_breakpoint": obj.ab_Ret_S_breakpoint,

            # ================= FLAGS =================
            "Org_Flag": obj.ab_Org_Flag,
            "Abx_Flag": obj.ab_Abx_Flag,
            "Phenotype_Flag": obj.ab_Abx_Phenotype,

            # ================= SYSTEM =================
            "Date_Uploaded": obj.ab_Date_uploaded_fd,
        })

    # ================= EXPORT =================
    df = pd.DataFrame(data)

    file_path = "Final_Antibiotic_Entries.xlsx"
    df.to_excel(file_path, index=False)

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename="Final_Antibiotic_Entries.xlsx"
    )



######################### Autofill recommendatiaon



@require_GET
def get_recommendation_f_description(request):

    reco_code = request.GET.get("reco_code")

    if not reco_code:
        return JsonResponse({"description": ""})

    reco = (
        Recommendation_items.objects
        .filter(RecoCode__iexact=reco_code)
        .order_by("id")
        .first()
    )
    return JsonResponse({"description": reco.Description if reco else ""})



############### Emerging List


def filter_emerging_by_request(queryset, request):
    q = request.GET.get("q", "").strip()
    date_from_raw = request.GET.get("date_from", "").strip()
    date_to_raw = request.GET.get("date_to", "").strip()
    date_from = parse_date(date_from_raw) if date_from_raw else None
    date_to = parse_date(date_to_raw) if date_to_raw else None

    if q:
        queryset = queryset.filter(eme_Accession__icontains=q)
    if date_from:
        queryset = queryset.filter(
            eme_primary_key__f_Referral_Date__gte=date_from
        )
    if date_to:
        queryset = queryset.filter(
            eme_primary_key__f_Referral_Date__lte=date_to
        )

    return queryset, q, date_from_raw, date_to_raw


def get_emerging_export_queryset(request):
    queryset, q, date_from, date_to = filter_emerging_by_request(
        Emerging_Table.fully_emerging().select_related("eme_primary_key"),
        request,
    )
    return (
        queryset.order_by(
            "-eme_primary_key__f_Referral_Date",
            "eme_Accession",
        ),
        q,
        date_from,
        date_to,
    )


@login_required
def emerging_list_view(request):

    qs, q, date_from, date_to = get_emerging_export_queryset(request)
    download_params = request.GET.copy()
    download_params.pop("page", None)

    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "projects/Emerging_List.html",
        {
            "page_obj": page_obj,
            "q": q,
            "date_from": date_from,
            "date_to": date_to,
            "download_querystring": download_params.urlencode(),
        }
    )

########## for downloading


def is_blank(val):
    return val is None or val == ""


def has_final_retest_result(entry):
    return any(value not in ("", None) for value in (
        entry.ab_Retest_DiskValue,
        entry.ab_Retest_MICValue,
        entry.ab_Retest_Disk_RIS,
        entry.ab_Retest_MIC_RIS,
        entry.ab_Retest_Disk_enRIS,
        entry.ab_Retest_MIC_enRIS,
    ))


@login_required(login_url="login")
def download_emerging_list(request):


    emerging_qs, _, _, _ = get_emerging_export_queryset(request)

   #unique antibiotcs
    unique_abx_codes = set()

    abx_qs = (
        Final_AntibioticEntry.objects
        .filter(
            ab_idNum_f_referred__in=emerging_qs.values_list(
                "eme_primary_key_id", flat=True
            ),
            ab_Retest_Abx_code__isnull=False,
        )
        .exclude(ab_Retest_Abx_code="")
    )

    for entry in abx_qs:
        if has_final_retest_result(entry):
            unique_abx_codes.add(entry.ab_Retest_Abx_code)

    sorted_antibiotics = sorted(unique_abx_codes)


    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = (
        'attachment; filename="emerging_cases_wide.csv"'
    )
    response.write("\ufeff")  # UTF-8 BOM (Excel safe)

    writer = csv.writer(response)


    EXPORT_FIELDS = OrderedDict([
        ("eme_Site_Code",    "Site_Code"),
        ("eme_Accession",    "Accession_No"),
        ("eme_ReferralDate", "Referral_Date"),
        ("eme_DateAdmis",    "Date_Admitted"),
        ("eme_Diagnosis",    "Diagnosis"),
        ("eme_Diag_ICD",     "Diagnosis_ICD"),
        ("eme_ars_Org",      "Organism"),
        ("eme_abx_code_pheno", "Emerging_Antibiotics"),
        ("eme_abx_Phenotype",  "Resistance Phenotype"),

        ("eme_ars_Pre",      "ARS_Pre"),
        ("eme_ars_Post",     "ARS_Post"),

        ("eme_org_Grp",      "Organism_Group"),
        ("eme_org_Genus",    "Organism_Genus"),

        ("eme_spec_Num",     "Specimen_No"),
        ("eme_spec_Date",    "Specimen_Date"),
        ("eme_spec_Type",    "Specimen_Type"),
    ])


    header = list(EXPORT_FIELDS.values())

    #write headers
    for abx in sorted_antibiotics:
        header.extend([
            f"{abx}_RT",
            f"{abx}_RT_RIS",
        ])

    writer.writerow(header)

    #write rows
    for eme in emerging_qs:

        # ---- Static columns ----
        row = [
            getattr(eme, field, "")
            for field in EXPORT_FIELDS.keys()
        ]

        # ---- Antibiotic data ----
        abx_entries = Final_AntibioticEntry.objects.filter(
            ab_idNum_f_referred=eme.eme_primary_key
        )

        abx_data = {}

        for ab in abx_entries:

            if ab.ab_Retest_Abx_code and has_final_retest_result(ab):
                code = ab.ab_Retest_Abx_code
                abx_data.setdefault(code, {})

                if not is_blank(ab.ab_Retest_DiskValue) or not is_blank(ab.ab_Retest_MICValue):
                    rt_val = (
                        ab.ab_Retest_DiskValue
                        if not is_blank(ab.ab_Retest_DiskValue)
                        else f"{ab.ab_Retest_MIC_operand or ''}{ab.ab_Retest_MICValue}"
                    )
                    rt_ris = ab.ab_Retest_Disk_enRIS or ab.ab_Retest_MIC_enRIS

                    abx_data[code].update({
                        "RT_Val": rt_val,
                        "RT_RIS": rt_ris,
                    })

        #format into wide
        for abx in sorted_antibiotics:
            data = abx_data.get(abx, {})

            rt_val = data.get("RT_Val", "")
            if isinstance(rt_val, (int, float)):
                rt_val = format(rt_val, ".3f")

            row.extend([
                rt_val,
                data.get("RT_RIS", ""),
            ])

        writer.writerow(row)

    return response


############# wgs overview


@login_required
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_LAB_ENCODER)
def update_wgs_classification_inline(request, accession_no):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "POST required"}, status=405)

    isolate = get_object_or_404(Final_Data, f_AccessionNo=accession_no)
    sampleinfo = (
        SampleInformation.objects
        .filter(sample_accession=isolate.f_AccessionNo)
        .order_by("-Date_uploaded_si", "-pk")
        .first()
    )
    if not sampleinfo:
        wgs_project = (
            WGS_Project.objects
            .filter(Q(Ref_Accession=isolate) | Q(WGS_SampleInfo_Acc=isolate.f_AccessionNo))
            .first()
        )
        if not wgs_project:
            wgs_project = WGS_Project.objects.create(
                Ref_Accession=isolate,
                WGS_SampleInfo_Acc=isolate.f_AccessionNo,
                WGS_SampleInfoSummary=True,
            )
        sampleinfo = SampleInformation.objects.create(
            sample_project=wgs_project,
            sample_accession=isolate.f_AccessionNo,
            sample_name=isolate.f_AccessionNo,
        )

    sampleinfo.emerging = "Class_Chk_Emerging" in request.POST
    sampleinfo.structured = "Class_Chk_Structured" in request.POST
    sampleinfo.satscan = "Class_Chk_Satscan" in request.POST
    sampleinfo.serotyping = "Class_Chk_Serotyping" in request.POST
    sampleinfo.ghru_all = "Class_Chk_GHRU_all" in request.POST
    sampleinfo.ghru_neo = "Class_Chk_GHRU_Neo" in request.POST
    sampleinfo.ghru = sampleinfo.ghru_all or sampleinfo.ghru_neo
    sampleinfo.egasp = "Class_Chk_EGASP" in request.POST
    sampleinfo.tricycle = "Class_Chk_Tricycle" in request.POST
    sampleinfo.pulsenet = "Class_Chk_Pulsenet" in request.POST
    sampleinfo.tulip = "Class_Chk_Tulip" in request.POST
    sampleinfo.save(update_fields=[
        "emerging",
        "structured",
        "satscan",
        "serotyping",
        "ghru",
        "ghru_all",
        "ghru_neo",
        "egasp",
        "tricycle",
        "pulsenet",
        "tulip",
    ])

    messages.success(
        request,
        f"WGS sample information flags updated for {accession_no}"
    )
    return redirect("/wgs/projects/?tab=wgs_classification")



# @login_required(login_url="login")
# def wgs_classification_view(request, pk):

#     isolate = get_object_or_404(Final_Data, pk=pk)

#     classification, created = Classification_Table.objects.get_or_create(
#         Class_idNumReferred=isolate,
#         defaults={
#             "Class_AccessionNo": isolate.f_AccessionNo,
#         }
#     )

#     if request.method == "POST":
#         form = FinalReferred_Form(request.POST, instance=isolate)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "WGS classification updated successfully.")
#             return redirect("wgs_classification_view", pk=pk)
#     else:
#         form = FinalReferred_Form(instance=isolate)

#     return render(
#         request,
#         "projects/wgs_review.html",
#         {
#             "form": form,
#             "isolates": isolate,
#         }
#     )


@login_required(login_url="login")
@role_required(ROLE_ADMIN, ROLE_CHECKER, ROLE_LAB_ENCODER)
def wgs_classification_view(request, pk):
    isolate = get_object_or_404(Final_Data, pk=pk)

    wgs_project = (
        WGS_Project.objects
        .filter(Ref_Accession=isolate.f_AccessionNo)
        .first()
    )

    bactscout = gambit = mlst = checkm2 = assembly = amrfinder = None

    if wgs_project:
        bactscout = BactScout.objects.filter(
            BactScout_Accession=wgs_project.WGS_BactScout_Acc
        )

        gambit = Gambit.objects.filter(
            Gambit_Accession=wgs_project.WGS_Gambit_Acc
        )

        mlst = Mlst.objects.filter(
            Mlst_Accession=wgs_project.WGS_Mlst_Acc
        )

        checkm2 = Checkm2.objects.filter(
            Checkm2_Accession=wgs_project.WGS_Checkm2_Acc
        )

        assembly = AssemblyScan.objects.filter(
            Assembly_Accession=wgs_project.WGS_Assembly_Acc
        )

        amrfinder = Amrfinderplus.objects.filter(
            Amrfinder_Accession=wgs_project.WGS_Amrfinder_Acc
        )

    context = {
        "isolates": isolate,
        "wgs_project": wgs_project,

        # tool-specific datasets
        "bactscout": bactscout,
        "gambit": gambit,
        "mlst": mlst,
        "checkm2": checkm2,
        "assembly": assembly,
        "amrfinder": amrfinder,
        "bactscout_fields": BactScout.FIELD_LABELS,
        "gambit_fields": Gambit.FIELD_LABELS,
        "mlst_fields": Mlst.FIELD_LABELS,
        "checkm2_fields": Checkm2.FIELD_LABELS,
        "amrfinder_fields": Amrfinderplus.FIELD_LABELS,
    }

    return render(
        request,
        "projects/wgs_review.html",
        context
    )





########## CONCORDANCE ANALYSIS
# ================================
# RIS Extractors (PUT ABOVE VIEW)
# ================================

# =====================================================
# RIS HELPERS
# =====================================================

def get_site_ris(entry):
    if entry.ab_Disk_enRIS:
        return entry.ab_Disk_enRIS.strip().upper()
    if entry.ab_MIC_enRIS:
        return entry.ab_MIC_enRIS.strip().upper()
    return None


def get_ars_ris(entry):
    if entry.ab_Retest_Disk_enRIS:
        return entry.ab_Retest_Disk_enRIS.strip().upper()
    if entry.ab_Retest_MIC_enRIS:
        return entry.ab_Retest_MIC_enRIS.strip().upper()
    return None


########  CLASSIFY ID CONCORDANCE HELPER



def clean_str(val):
    """
    Safely convert any value to a lowercase string.
    Handles None, NaN, numbers, and Django model objects.
    """
    if val is None:
        return ""

    if pd.isna(val):
        return ""

    if hasattr(val, "Pre_Phenotypes"):
        return str(val.Pre_Phenotypes).strip().lower()

    return str(val).strip().lower()


def classify_id_concordance(site_org, ars_pre, ars_org):

    site = clean_str(site_org)
    ars = clean_str(ars_org)
    pre = clean_str(ars_pre)

    # Mixed Culture detection
    if "mixed culture of" in pre:
        return "M", "M"

    if ars == "not viable":
        return "N", "N"

    if site and ars and site == ars:
        return "G", "S"

    if site and ars and site != ars:
        return "X", "X"

    return None, None

# ==============================
# AST CONCORDANCE
# ==============================
def classify_ast_deviation(site_ris, ars_ris):

    if not site_ris or not ars_ris:
        return None, False

    s = site_ris.strip().upper()
    a = ars_ris.strip().upper()

    # Normalize special values
    if s == "SDD":
        s = "S"
    if a == "SDD":
        a = "S"

    if s == "NS":
        s = "R"
    if a == "NS":
        a = "R"

    if s in ["ND", "NA"] or a in ["ND", "NA"]:
        return None, False

    if s == a:
        return "S", False

    if s == "S" and a == "R":
        return "A", True

    if s == "R" and a == "S":
        return "B", True

    if "I" in {s, a}:
        return "C", True

    return None, False



@login_required(login_url="login")
def concordance_analysis_view(request):
    current_year = timezone.localdate().year
    selected_year = request.GET.get("year", str(current_year)).strip() or str(current_year)
    dashboard_year = None if selected_year == "all" else selected_year
    context = concordance_service.collect_concordance_dashboard(dashboard_year)
    year_options = list(
        Final_Data.objects
        .exclude(f_Referral_Date__isnull=True)
        .annotate(year=ExtractYear("f_Referral_Date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )
    if current_year not in year_options:
        year_options.insert(0, current_year)
    context.update({
        "selected_year": selected_year,
        "year_options": year_options,
        "current_year": current_year,
        "year_scope_label": "All Years" if selected_year == "all" else selected_year,
    })

    return render(
        request,
        "home_final/concordance_dashboard.html",
        context
    )

    isolates = Final_Data.objects.prefetch_related("final_entries")

    total_isolates = isolates.count()

    # =============================
    # ORGANISM COUNTERS
    # =============================
    concordant_species = 0
    concordant_genus = 0
    different_org = 0
    mixed_count = 0
    not_viable_count = 0

    # =============================
    # AST COUNTERS
    # =============================
    total_pairs = 0
    concordant_pairs = 0
    vmd = 0
    md = 0
    minor = 0

    print("\n========== CONCORDANCE DEBUG START ==========\n")

    for isolate in isolates:

        print(f"\n--- Isolate {isolate.id} ({isolate.f_AccessionNo}) ---")

        # =============================
        # ORGANISM CONCORDANCE
        # =============================
        genus_con, species_con = classify_id_concordance(
            isolate.f_Site_OrgName,
            isolate.f_ars_Pre,
            isolate.f_ars_OrgName
        )

        if genus_con == "G":
            concordant_genus += 1

        if species_con == "S":
            concordant_species += 1

        if genus_con == "X":
            different_org += 1

        if genus_con == "M":
            mixed_count += 1

        if genus_con == "N":
            not_viable_count += 1

        print(
            f"Genus_Con={genus_con}, Species_Con={species_con}"
        )

        # =============================
        # AST CONCORDANCE
        # =============================
        entries = isolate.final_entries.all()

        site_results = {}
        ars_results = {}

        for entry in entries:

            antibiotic = entry.ab_Abx_code or entry.ab_Retest_Abx_code

            site_ris = get_site_ris(entry)
            ars_ris = get_ars_ris(entry)

            print(
                f"Entry {entry.id} | Abx={antibiotic} | "
                f"Site={site_ris} | ARS={ars_ris}"
            )

            if antibiotic and site_ris:
                site_results[antibiotic] = site_ris

            if antibiotic and ars_ris:
                ars_results[antibiotic] = ars_ris

        print("Site Results:", site_results)
        print("ARS Results:", ars_results)

        for antibiotic, site_ris in site_results.items():

            ars_ris = ars_results.get(antibiotic)

            if not ars_ris:
                print(f"Skipping {antibiotic} (no ARS match)")
                continue

            print(
                f"Comparing {antibiotic} → Site={site_ris} vs ARS={ars_ris}"
            )

            code, is_disc = classify_ast_deviation(site_ris, ars_ris)

            if not code:
                print("→ Skipped (invalid/ND)")
                continue

            total_pairs += 1

            if code == "S":
                concordant_pairs += 1
                print("→ Concordant")

            elif code == "A":
                vmd += 1
                print("→ Very Major")

            elif code == "B":
                md += 1
                print("→ Major")

            elif code == "C":
                minor += 1
                print("→ Minor")

    print("\n========== CONCORDANCE DEBUG END ==========\n")

    # =============================
    # RATE CALCULATIONS
    # =============================
    species_rate = round(
        (concordant_species / total_isolates) * 100, 2
    ) if total_isolates else 0

    genus_rate = round(
        (concordant_genus / total_isolates) * 100, 2
    ) if total_isolates else 0

    ast_concordance_rate = round(
        (concordant_pairs / total_pairs) * 100, 2
    ) if total_pairs else 0

    vmd_rate = round(
        (vmd / total_pairs) * 100, 2
    ) if total_pairs else 0

    md_rate = round(
        (md / total_pairs) * 100, 2
    ) if total_pairs else 0

    minor_rate = round(
        (minor / total_pairs) * 100, 2
    ) if total_pairs else 0

    context = {
        "total_isolates": total_isolates,

        # Organism
        "concordant_species": concordant_species,
        "species_rate": species_rate,

        "concordant_genus": concordant_genus,
        "genus_rate": genus_rate,

        "different_org": different_org,
        "mixed_count": mixed_count,
        "not_viable_count": not_viable_count,

        # AST
        "total_pairs": total_pairs,
        "concordant_pairs": concordant_pairs,
        "ast_concordance_rate": ast_concordance_rate,

        "vmd": vmd,
        "vmd_rate": vmd_rate,

        "md": md,
        "md_rate": md_rate,

        "minor": minor,
        "minor_rate": minor_rate,
        "isolates": isolates,

    }

    return render(
        request,
        "home_final/concordance_dashboard.html",
        context
    )




# this is per batch
@login_required(login_url="login")
@require_POST
@transaction.atomic
def concordance_generate_batch(request):

    batch_id = request.POST.get("batch_id")
    report = concordance_service.generate_concordance_for_batch(
        batch_id,
        request.user,
    )
    return redirect("concordance_batch_detail", report_id=report.id)

    batch = get_object_or_404(Batch_Table, id=batch_id)

    isolates = (
        Final_Data.objects
        .filter(f_Batch_id=batch)
        .prefetch_related("final_entries")
    )

    total_isolates = isolates.count()

    total_pairs = 0
    concordant_pairs = 0
    vmd = 0
    md = 0
    minor = 0

    # =========================
    # GENUS / SPECIES COUNTERS
    # =========================
    genus_match = 0
    species_match = 0
    genus_discordant = 0
    mixed_count = 0
    not_viable_count = 0

    detail_objects = []

    for isolate in isolates:

        # ======================
        # ID CONCORDANCE
        # ======================
        genus_con, species_con = classify_id_concordance(
            isolate.f_Site_OrgName,
            isolate.f_ars_Pre,
            isolate.f_ars_OrgName
        )

        if genus_con == "G":
            genus_match += 1
        elif genus_con == "X":
            genus_discordant += 1
        elif genus_con == "M":
            mixed_count += 1
        elif genus_con == "N":
            not_viable_count += 1

        if species_con == "S":
            species_match += 1

        # ======================
        # AST COMPARISON
        # ======================
        site_results = {}
        ars_results = {}

        for entry in isolate.final_entries.all():

            site_ris = get_site_ris(entry)
            ars_ris = get_ars_ris(entry)

            if entry.ab_Abx_code and site_ris:
                site_results[entry.ab_Abx_code] = site_ris

            if entry.ab_Retest_Abx_code and ars_ris:
                ars_results[entry.ab_Retest_Abx_code] = ars_ris

        for antibiotic, site_ris in site_results.items():

            ars_ris = ars_results.get(antibiotic)
            if not ars_ris:
                continue

            code, is_disc = classify_ast_deviation(site_ris, ars_ris)
            if not code:
                continue

            total_pairs += 1

            if code == "S":
                concordant_pairs += 1
            elif code == "A":
                vmd += 1
            elif code == "B":
                md += 1
            elif code == "C":
                minor += 1

            detail_objects.append(
                ConcordanceDetail(
                    accession_no=isolate.f_AccessionNo,
                    isolate_id=isolate.id,
                    organism=isolate.f_Site_OrgName,
                    antibiotic=antibiotic,
                    site_ris=site_ris,
                    ars_ris=ars_ris,
                    deviation_code=code,
                    is_discordant=is_disc,
                    genus_con=genus_con,
                    species_con=species_con,
                )
            )

    viable = total_isolates - mixed_count - not_viable_count

    genus_rate = round((genus_match / viable) * 100, 2) if viable else 0
    species_rate = round((species_match / viable) * 100, 2) if viable else 0

    total_deviation = vmd + md + minor
    critical_deviation = vmd + md

    ast_concordance_rate = round(
        (concordant_pairs / total_pairs) * 100, 2
    ) if total_pairs else 0

    report, created = ConcordanceReport.objects.update_or_create(
        batch=batch,
        final_data=None,
        defaults={
            "created_by": request.user,
            "total_isolates": total_isolates,
            "total_pairs": total_pairs,
            "concordant_pairs": concordant_pairs,
            "vmd": vmd,
            "md": md,
            "minor": minor,
            "total_deviation": total_deviation,
            "critical_deviation": critical_deviation,
            "ast_concordance_rate": ast_concordance_rate,
            "genus_match": genus_match,
            "species_match": species_match,
            "genus_rate": genus_rate,
            "species_rate": species_rate,
        }
    )

    report.details.all().delete()

    for obj in detail_objects:
        obj.report = report

    ConcordanceDetail.objects.bulk_create(detail_objects)

    return redirect("concordance_batch_detail", report_id=report.id)





@login_required(login_url="login")
def concordance_batch_detail(request, report_id):

    report = get_object_or_404(
        ConcordanceReport.objects.select_related("batch", "created_by"),
        id=report_id,
        final_data__isnull=True
    )

    context = _build_batch_concordance_context(report)
    context["is_batch_report"] = True

    return render(
        request,
        "home_final/concordance_batch_detail.html",
        context
    )



# this is per accession
@login_required(login_url="login")
@require_POST
@transaction.atomic
def concordance_generate_accession(request):

    isolate_id = request.POST.get("isolate_id")
    isolate = get_object_or_404(Final_Data, id=isolate_id)
    report = concordance_service.generate_concordance_for_isolate(
        isolate,
        request.user,
    )
    return redirect("concordance_accession_detail", report_id=report.id)

    isolate = get_object_or_404(
        Final_Data.objects.prefetch_related("final_entries"),
        id=isolate_id
    )

    # ==========================
    # ORGANISM CONCORDANCE
    # ==========================
    genus_con, species_con = classify_id_concordance(
        isolate.f_Site_OrgName,
        isolate.f_ars_Pre,
        isolate.f_ars_OrgName
    )

    genus_match = 1 if genus_con == "G" else 0
    species_match = 1 if species_con == "S" else 0

    genus_rate = 100 if genus_con == "G" else 0
    species_rate = 100 if species_con == "S" else 0

    # ==========================
    # AST COMPARISON
    # ==========================
    total_pairs = 0
    concordant_pairs = 0
    vmd = 0
    md = 0
    minor = 0

    site_results = {}
    ars_results = {}
    detail_objects = []

    for entry in isolate.final_entries.all():

        site_ris = get_site_ris(entry)
        ars_ris = get_ars_ris(entry)

        if entry.ab_Abx_code and site_ris:
            site_results[entry.ab_Abx_code] = site_ris

        if entry.ab_Retest_Abx_code and ars_ris:
            ars_results[entry.ab_Retest_Abx_code] = ars_ris

    for antibiotic, site_ris in site_results.items():

        ars_ris = ars_results.get(antibiotic)
        if not ars_ris:
            continue

        code, is_disc = classify_ast_deviation(site_ris, ars_ris)
        if not code:
            continue

        total_pairs += 1

        if code == "S":
            concordant_pairs += 1
        elif code == "A":
            vmd += 1
        elif code == "B":
            md += 1
        elif code == "C":
            minor += 1

        detail_objects.append(
            ConcordanceDetail(
                accession_no=isolate.f_AccessionNo,
                isolate_id=isolate.id,
                organism=isolate.f_Site_OrgName,
                antibiotic=antibiotic,
                site_ris=site_ris,
                ars_ris=ars_ris,
                deviation_code=code,
                is_discordant=is_disc,
                genus_con=genus_con,
                species_con=species_con,
            )
        )

    total_deviation = vmd + md + minor
    critical_deviation = vmd + md

    ast_concordance_rate = round(
        (concordant_pairs / total_pairs) * 100, 2
    ) if total_pairs else 0

    report, created = ConcordanceReport.objects.update_or_create(
        final_data=isolate,
        defaults={
            "batch": isolate.f_Batch_id,
            "created_by": request.user,
            "total_isolates": 1,
            "total_pairs": total_pairs,
            "concordant_pairs": concordant_pairs,
            "vmd": vmd,
            "md": md,
            "minor": minor,
            "total_deviation": total_deviation,
            "critical_deviation": critical_deviation,
            "ast_concordance_rate": ast_concordance_rate,
            "genus_match": genus_match,
            "species_match": species_match,
            "genus_rate": genus_rate,
            "species_rate": species_rate,
        }
    )

    report.details.all().delete()

    for obj in detail_objects:
        obj.report = report

    ConcordanceDetail.objects.bulk_create(detail_objects)

    return redirect("concordance_accession_detail", report_id=report.id)



@login_required(login_url="login")
def concordance_accession_detail(request, report_id):

    report = get_object_or_404(
        ConcordanceReport.objects.select_related(
            "final_data", "batch", "created_by"
        ),
        id=report_id,
        final_data__isnull=False
    )

    details = report.details.all().order_by("antibiotic")

    context = {
        "report": report,
        "details": details,
        "isolate": report.final_data,
    }

    return render(
        request,
        "home_final/concordance_accession_detail.html",
        context
    )


### used when final data entry is updated to auto-regenerate snapshot for that accession
def regenerate_concordance_snapshot(isolate, user=None):
    """
    Auto-regenerates concordance snapshot using the unified engine.
    """
    return concordance_service.generate_concordance_for_isolate(isolate, user)

    from django.db import transaction

    genus_con, species_con = classify_id_concordance(
        isolate.f_Site_OrgName,
        isolate.f_ars_Pre,
        isolate.f_ars_OrgName
    )

    site_results = {}
    ars_results = {}

    for entry in isolate.final_entries.all():

        site_ris = get_site_ris(entry)
        ars_ris = get_ars_ris(entry)

        site_abx = entry.ab_Abx_code
        ars_abx = entry.ab_Retest_Abx_code

        if site_abx and site_ris:
            site_results[site_abx] = site_ris

        if ars_abx and ars_ris:
            ars_results[ars_abx] = ars_ris

    total_pairs = 0
    concordant_pairs = 0
    vmd = 0
    md = 0
    minor = 0

    detail_objects = []

    for antibiotic, site_ris in site_results.items():

        ars_ris = ars_results.get(antibiotic)
        if not ars_ris:
            continue

        code, is_disc = classify_ast_deviation(site_ris, ars_ris)
        if not code:
            continue

        total_pairs += 1

        if code == "S":
            concordant_pairs += 1
        elif code == "A":
            vmd += 1
        elif code == "B":
            md += 1
        elif code == "C":
            minor += 1

        detail_objects.append(
            ConcordanceDetail(
                accession_no=isolate.f_AccessionNo,
                isolate_id=isolate.id,
                organism=isolate.f_Site_OrgName,
                antibiotic=antibiotic,
                site_ris=site_ris,
                ars_ris=ars_ris,
                deviation_code=code,
                is_discordant=is_disc,
                genus_con=genus_con,
                species_con=species_con,
            )
        )

    total_deviation = vmd + md + minor
    critical_deviation = vmd + md

    ast_concordance_rate = round(
        (concordant_pairs / total_pairs) * 100, 2
    ) if total_pairs else 0

    critical_deviation_rate = round(
        (critical_deviation / total_pairs) * 100, 2
    ) if total_pairs else 0

    total_deviation_rate = round(
        (total_deviation / total_pairs) * 100, 2
    ) if total_pairs else 0

    genus_match = 1 if genus_con == "G" else 0
    species_match = 1 if species_con == "S" else 0
    genus_rate = 100 if genus_con == "G" else 0
    species_rate = 100 if species_con == "S" else 0

    with transaction.atomic():

        report, _ = ConcordanceReport.objects.update_or_create(
            final_data=isolate,
            defaults={
                "batch": isolate.f_Batch_id,
                "created_by": user,
                "total_isolates": 1,
                "total_pairs": total_pairs,
                "concordant_pairs": concordant_pairs,
                "vmd": vmd,
                "md": md,
                "minor": minor,
                "total_deviation": total_deviation,
                "critical_deviation": critical_deviation,
                "ast_concordance_rate": ast_concordance_rate,
                "critical_deviation_rate": critical_deviation_rate,
                "total_deviation_rate": total_deviation_rate,
                "genus_match": genus_match,
                "species_match": species_match,
                "genus_rate": genus_rate,
                "species_rate": species_rate,
            }
        )

        report.details.all().delete()

        for obj in detail_objects:
            obj.report = report

        ConcordanceDetail.objects.bulk_create(detail_objects)


# =====================================================
# HISTORY VIEW
# =====================================================

@login_required(login_url="login")
def concordance_history_view(request):
    q = request.GET.get("q", "").strip()
    report_type = request.GET.get("report_type", "").strip()
    generated_by = request.GET.get("generated_by", "").strip()
    quality = request.GET.get("quality", "").strip()
    year = request.GET.get("year", "").strip()
    show_all = request.GET.get("show_all") == "1"

    reports = (
        ConcordanceReport.objects
        .select_related("batch", "final_data", "created_by")
        .annotate(total_dev_calc=F("vmd") + F("md") + F("minor"))
    )

    if not show_all:
        latest_batch_report = (
            ConcordanceReport.objects
            .filter(final_data__isnull=True, batch_id=OuterRef("batch_id"))
            .order_by("-created_at", "-id")
            .values("id")[:1]
        )
        latest_accession_report = (
            ConcordanceReport.objects
            .filter(final_data__isnull=False, final_data_id=OuterRef("final_data_id"))
            .order_by("-created_at", "-id")
            .values("id")[:1]
        )
        reports = reports.filter(
            Q(final_data__isnull=True, id=Subquery(latest_batch_report))
            | Q(final_data__isnull=False, id=Subquery(latest_accession_report))
        )

    if q:
        reports = reports.filter(
            Q(batch__bat_Batch_Name__icontains=q)
            | Q(batch__bat_SiteCode__icontains=q)
            | Q(batch__bat_RefNo__icontains=q)
            | Q(final_data__f_AccessionNo__icontains=q)
            | Q(final_data__f_SiteCode__icontains=q)
        )

    if report_type == "batch":
        reports = reports.filter(final_data__isnull=True)
    elif report_type == "accession":
        reports = reports.filter(final_data__isnull=False)

    if generated_by:
        reports = reports.filter(created_by_id=generated_by)

    if year:
        reports = reports.filter(
            Q(batch__bat_Referral_Date__year=year)
            | Q(final_data__f_Referral_Date__year=year)
        )

    if quality == "perfect":
        reports = reports.filter(total_dev_calc=0)
    elif quality == "acceptable":
        reports = reports.filter(vmd=0, total_dev_calc__gt=0, total_dev_calc__lte=3)
    elif quality == "review":
        reports = reports.exclude(Q(total_dev_calc=0) | Q(vmd=0, total_dev_calc__gt=0, total_dev_calc__lte=3))

    reports = reports.order_by("-created_at", "-id")
    paginator = Paginator(reports, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    batch_years = (
        Batch_Table.objects
        .exclude(bat_Referral_Date__isnull=True)
        .annotate(year=ExtractYear("bat_Referral_Date"))
        .values_list("year", flat=True)
    )
    accession_years = (
        Final_Data.objects
        .exclude(f_Referral_Date__isnull=True)
        .annotate(year=ExtractYear("f_Referral_Date"))
        .values_list("year", flat=True)
    )
    years = sorted(
        {item for item in list(batch_years) + list(accession_years) if item},
        reverse=True
    )
    users = User.objects.filter(concordancereport__isnull=False).distinct().order_by("username")

    params = request.GET.copy()
    params.pop("page", None)
    preserved_params = params.urlencode()

    return render(
        request,
        "home_final/concordance_history.html",
        {
            "page_obj": page_obj,
            "reports": page_obj.object_list,
            "report_count": paginator.count,
            "q": q,
            "report_type": report_type,
            "generated_by": generated_by,
            "quality": quality,
            "year": year,
            "show_all": show_all,
            "years": years,
            "users": users,
            "preserved_params": preserved_params,
            "quality_choices": [
                ("perfect", "Perfect"),
                ("acceptable", "Acceptable"),
                ("review", "Needs Review"),
            ],
        }
    )



def _format_refno_range(isolates, batch=None):
    if batch and getattr(batch, "bat_RefNo", None):
        return batch.bat_RefNo

    refnos = [
        str(value).strip()
        for value in isolates.exclude(f_RefNo__isnull=True)
        .exclude(f_RefNo__exact="")
        .values_list("f_RefNo", flat=True)
        if str(value).strip()
    ]
    numeric_refs = []
    width = 0

    for refno in refnos:
        match = re.search(r"(\d+)$", refno)
        if match:
            numeric_refs.append(int(match.group(1)))
            width = max(width, len(match.group(1)))

    if numeric_refs:
        start = str(min(numeric_refs)).zfill(width or 4)
        end = str(max(numeric_refs)).zfill(width or 4)
        return f"{start}-{end}"

    return ", ".join(refnos)


def _resolve_site_contact(batch, first_isolate=None):
    site_code = (
        getattr(batch, "bat_SiteCode", "")
        or getattr(first_isolate, "f_SiteCode", "")
        or ""
    )
    site_name = (
        getattr(first_isolate, "f_Site_Name", "")
        or getattr(batch, "bat_Site_Name", "")
        or ""
    )

    site = None
    if str(site_code).strip():
        site = SiteData.objects.filter(SiteCode__iexact=str(site_code).strip()).first()
    if site is None and str(site_name).strip():
        site = SiteData.objects.filter(SiteName__iexact=str(site_name).strip()).first()

    recipient_name = (getattr(site, "Site_Lab_Head", "") or "").strip()
    recipient_credentials = (getattr(site, "Site_Lab_Head_Credentials", "") or "").strip()
    recipient_designation = (getattr(site, "Site_Lab_Head_Designation", "") or "").strip()
    recipient_address = (getattr(site, "Site_Address", "") or "").strip()
    if recipient_name and not recipient_credentials and "," in recipient_name:
        recipient_name, recipient_credentials = [
            part.strip()
            for part in recipient_name.split(",", 1)
        ]

    salutation_name = "Lab Manager"
    if recipient_name:
        name_without_credentials = recipient_name.split(",")[0].strip()
        name_parts = [part for part in re.split(r"\s+", name_without_credentials) if part]
        last_name = name_parts[-1].title() if name_parts else name_without_credentials
        credential_text = f"{recipient_name} {recipient_credentials}"
        has_doctor_credential = bool(re.search(r"\b(MD|DR\.?|DO)\b", credential_text, re.IGNORECASE))
        salutation_name = f"Dr. {last_name}" if has_doctor_credential else name_without_credentials

    return {
        "site_record": site,
        "recipient_name": recipient_name or "Lab Manager",
        "credentials": recipient_credentials,
        "recipient_designation": recipient_designation or "Lab Manager",
        "recipient_address": recipient_address,
        "recipient_salutation": salutation_name,
        "site_address": recipient_address,
    }


def _build_batch_concordance_context(report):
    batch = report.batch
    isolates = Final_Data.objects.filter(f_Batch_id=batch).order_by("f_bat_seq", "f_AccessionNo")
    first_isolate = isolates.first()
    id_stats = concordance_service.build_id_stats(isolates)
    details = report.details.all().order_by("isolate_id", "antibiotic")
    total_pairs = report.total_pairs or 0
    total_deviation = (report.vmd or 0) + (report.md or 0) + (report.minor or 0)
    critical_deviation = (report.vmd or 0) + (report.md or 0)
    different_org_rate = round((id_stats["different_org"] / id_stats["viable_pure"]) * 100, 2) if id_stats["viable_pure"] else 0
    total_pairs_rate = 100 if total_pairs else 0
    isolate_ref_map = {
        isolate.id: isolate.f_RefNo
        for isolate in isolates
    }
    isolate_bat_seq_map = {
        isolate.id: isolate.f_bat_seq if isolate.f_bat_seq is not None else ""
        for isolate in isolates
    }
    discordant_details = [
        {
            "refno": isolate_ref_map.get(detail.isolate_id, ""),
            "bat_seq": isolate_bat_seq_map.get(detail.isolate_id, ""),
            "accession_no": detail.accession_no,
            "organism": detail.organism,
            "antibiotic": detail.antibiotic,
            "site_ris": detail.site_ris,
            "ars_ris": detail.ars_ris,
            "deviation_code": detail.deviation_code,
        }
        for detail in details
        if detail.deviation_code != "S"
    ]

    context = {
        "report": report,
        "batch": batch,
        "isolates": isolates,
        "details": details,
        "site_name": first_isolate.f_Site_Name if first_isolate else getattr(batch, "bat_Site_Name", ""),
        "referral_date": first_isolate.f_Referral_Date if first_isolate else getattr(batch, "bat_Referral_Date", ""),
        "accession_numbers": _format_refno_range(isolates, batch),
        "total_isolates": isolates.count(),
        "mixed_count": id_stats["mixed_count"],
        "nonviable_count": id_stats["nonviable_count"],
        "viable_pure": id_stats["viable_pure"],
        "genus_match": id_stats["genus_match"],
        "species_match": id_stats["species_match"],
        "different_org": id_stats["different_org"],
        "different_org_rate": different_org_rate,
        "genus_rate": id_stats["genus_rate"],
        "species_rate": id_stats["species_rate"],
        "discordant_rows": id_stats["discordant_rows"],
        "total_pairs": total_pairs,
        "total_pairs_rate": total_pairs_rate,
        "concordant": report.concordant_pairs or 0,
        "vmd": report.vmd or 0,
        "md": report.md or 0,
        "minor": report.minor or 0,
        "critical_deviation": critical_deviation,
        "total_deviation": total_deviation,
        "concordance_rate": report.ast_concordance_rate or 0,
        "critical_deviation_rate": round((critical_deviation / total_pairs) * 100, 2) if total_pairs else 0,
        "total_deviation_rate": round((total_deviation / total_pairs) * 100, 2) if total_pairs else 0,
        "abx_summary": concordance_service.calculate_antibiotic_summary(details),
        "isolate_ref_map": isolate_ref_map,
        "isolate_bat_seq_map": isolate_bat_seq_map,
        "discordant_details": discordant_details,
    }
    context.update(_resolve_site_contact(batch, first_isolate))
    return context


@login_required(login_url="login")
def concordance_report_detail_view(request, report_id):

    report = get_object_or_404(
        ConcordanceReport.objects.select_related("batch", "created_by"),
        id=report_id
    )

    # Snapshot detail rows (AST comparison rows)
    details = report.details.all().order_by("isolate_id", "antibiotic")

    context = {
        "report": report,
        "details": details,
    }

    return render(
        request,
        "home_final/concordance_batch_detail.html",
        context
    )



@login_required(login_url="login")
def export_concordance_batch_excel(request, report_id):

    # 🔒 Ensure this is a BATCH report only
    report = get_object_or_404(
        ConcordanceReport.objects.select_related("batch"),
        id=report_id,
        final_data__isnull=True  # <--- IMPORTANT
    )

    batch = report.batch

    if not batch:
        return HttpResponse("Invalid batch report.", status=400)

    isolates = Final_Data.objects.filter(f_Batch_id=batch)

    total_isolates = isolates.count()

    site_name = isolates.first().f_Site_Name if isolates.exists() else ""
    referral_date = isolates.first().f_Referral_Date if isolates.exists() else ""

    ref_no = (
        isolates
        .exclude(f_RefNo__isnull=True)
        .exclude(f_RefNo__exact="")
        .values_list("f_RefNo", flat=True)
        .first()
    )

    accession_numbers = ref_no or ""

    # =====================================================
    # ORGANISM CONCORDANCE CALCULATION (LIVE)
    # =====================================================

    genus_match = 0
    species_match = 0
    different_org = 0
    mixed_count = 0
    nonviable_count = 0

    discordant_rows = []

    for isolate in isolates:

        site_org = (isolate.f_Site_OrgName or "").strip().lower()
        ars_org = (isolate.f_ars_OrgName or "").strip().lower()
        ars_pre = (isolate.f_ars_Pre or "").strip().lower()

        # Mixed Culture
        if ars_pre == "mixed culture":
            mixed_count += 1
            continue

        # Not Viable
        if ars_org == "not viable":
            nonviable_count += 1
            continue

        if site_org and ars_org:

            if site_org == ars_org:
                species_match += 1
                genus_match += 1

            elif site_org.split(" ")[0] == ars_org.split(" ")[0]:
                genus_match += 1

            else:
                different_org += 1
                discordant_rows.append([
                    isolate.f_RefNo,
                    isolate.f_Site_OrgName,
                    isolate.f_ars_OrgName
                ])

    viable_pure = total_isolates - mixed_count - nonviable_count

    genus_rate = round((genus_match / viable_pure) * 100, 2) if viable_pure else 0
    species_rate = round((species_match / viable_pure) * 100, 2) if viable_pure else 0

    id_stats = concordance_service.build_id_stats(isolates)
    genus_match = id_stats["genus_match"]
    species_match = id_stats["species_match"]
    different_org = id_stats["different_org"]
    mixed_count = id_stats["mixed_count"]
    nonviable_count = id_stats["nonviable_count"]
    viable_pure = id_stats["viable_pure"]
    genus_rate = id_stats["genus_rate"]
    species_rate = id_stats["species_rate"]
    discordant_rows = [
        [row["refno"], row["site_org"], row["ars_org"]]
        for row in id_stats["discordant_rows"]
    ]

    # =====================================================
    # AST CALCULATIONS (FROM SNAPSHOT)
    # =====================================================

    total_pairs = report.total_pairs or 0
    concordant = report.concordant_pairs or 0
    vmd = report.vmd or 0
    md = report.md or 0
    minor = report.minor or 0

    total_deviation = vmd + md + minor

    concordance_rate = round((concordant / total_pairs) * 100, 2) if total_pairs else 0
    vmd_rate = round((vmd / total_pairs) * 100, 2) if total_pairs else 0
    total_deviation_rate = round((total_deviation / total_pairs) * 100, 2) if total_pairs else 0

    # =====================================================
    # WORKBOOK
    # =====================================================

    wb = Workbook()
    ws = wb.active
    ws.title = "Concordance Quality Report"

    # HEADER
    ws["A1"] = "Sentinel Site"
    ws["B1"] = site_name

    ws["A2"] = "Date of Referral"
    ws["B2"] = str(referral_date)

    ws["A3"] = "Accession Numbers"
    ws["B3"] = accession_numbers

    for row in range(1, 4):
        ws[f"A{row}"].font = Font(bold=True)

    # =====================================================
    # 1. QUALITY OF REFERRAL
    # =====================================================

    ws["A5"] = "1. QUALITY OF REFERRAL"
    ws["A5"].font = Font(bold=True)

    ws["A6"] = "1.1 Number of Isolates"
    ws["B6"] = total_isolates

    ws["A7"] = "1.2 Number of Nonviable Isolates"
    ws["B7"] = nonviable_count

    ws["A8"] = "1.3 Number of Mixed Cultures"
    ws["B8"] = mixed_count

    ws["A9"] = "1.4 Number of Viable and Pure Isolates"
    ws["B9"] = viable_pure

    # =====================================================
    # 2. ORGANISM IDENTIFICATION
    # =====================================================

    ws["A11"] = "2. ORGANISM IDENTIFICATION"
    ws["A11"].font = Font(bold=True)

    ws["A12"] = "2.1 Concordant ID Genus Level"
    ws["B12"] = ">= 96%"
    ws["C12"] = genus_match
    ws["D12"] = f"{genus_rate}%"

    ws["A13"] = "2.2 Concordant ID Species Level"
    ws["B13"] = ">= 90%"
    ws["C13"] = species_match
    ws["D13"] = f"{species_rate}%"

    ws["A14"] = "2.3 Different Identification"
    ws["C14"] = different_org

    # =====================================================
    # 3. AST RESULTS
    # =====================================================

    ws["A16"] = "3. AST RESULTS"
    ws["A16"].font = Font(bold=True)

    ws["A17"] = "3.1 Total Number of AST pairs analyzed"
    ws["C17"] = total_pairs

    ws["A18"] = "3.2 Concordant AST Results"
    ws["C18"] = concordant
    ws["D18"] = f"{concordance_rate}%"

    ws["A19"] = "A – Very Major Deviations"
    ws["C19"] = vmd
    ws["D19"] = f"{vmd_rate}%"

    ws["A20"] = "B – Major Deviations"
    ws["C20"] = md

    ws["A21"] = "C – Minor Deviations"
    ws["C21"] = minor

    ws["A22"] = "3.4 Critical Deviations (<=5%)"
    ws["C22"] = vmd + md

    ws["A23"] = "3.5 Total Deviations (<=8%)"
    ws["C23"] = total_deviation
    ws["D23"] = f"{total_deviation_rate}%"

    # Auto width
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    filename = f"{batch.bat_Batch_Name}_Concordance.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response






@login_required(login_url="login")
def export_concordance_batch_excel(request, report_id):
    report = get_object_or_404(
        ConcordanceReport.objects.select_related("batch"),
        id=report_id,
        final_data__isnull=True
    )
    if not report.batch:
        return HttpResponse("Invalid batch report.", status=400)

    context = _build_batch_concordance_context(report)

    wb = Workbook()
    ws = wb.active
    ws.title = "Concordance Summary"

    blue = PatternFill("solid", fgColor="1E5A96")
    light_blue = PatternFill("solid", fgColor="D9EAF7")
    gray = PatternFill("solid", fgColor="E6E6E6")
    thin = Border(
        left=Side(style="thin", color="9E9E9E"),
        right=Side(style="thin", color="9E9E9E"),
        top=Side(style="thin", color="9E9E9E"),
        bottom=Side(style="thin", color="9E9E9E"),
    )

    ws.merge_cells("A1:D1")
    ws["A1"] = "Table 1: Summary Concordance Report"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = blue
    ws["A1"].alignment = Alignment(horizontal="center")

    info_rows = [
        ("Sentinel Site", context["site_name"]),
        ("Date of Referral", context["referral_date"]),
        ("Accession Numbers", context["accession_numbers"]),
    ]
    for idx, (label, value) in enumerate(info_rows, 3):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)
        ws.cell(row=idx, column=1).fill = light_blue

    rows = [
        ("1. QUALITY OF REFERRAL", "TARGET", "Number", "%"),
        ("1.1 Number of Isolates", "", context["total_isolates"], ""),
        ("1.2 Number of Nonviable Isolates", "", context["nonviable_count"], ""),
        ("1.3 Number of Mixed Cultures", "", context["mixed_count"], ""),
        ("1.4 Number of Viable and Pure Isolates", "", context["viable_pure"], ""),
        ("2. ORGANISM IDENTIFICATION", "TARGET", "Number", "%"),
        ("2.1 Concordant ID Genus Level", ">= 96%", context["genus_match"], f"{context['genus_rate']}%"),
        ("2.2 Concordant ID Species Level", ">= 90%", context["species_match"], f"{context['species_rate']}%"),
        ("2.3 Different Identification", "", context["different_org"], f"{context['different_org_rate']}%"),
        ("3. AST RESULTS", "TARGET", "Number", "%"),
        ("3.1 Total Number of AST pairs analyzed", "", context["total_pairs"], f"{context['total_pairs_rate']}%"),
        ("3.2 Concordant AST Results", "", context["concordant"], f"{context['concordance_rate']}%"),
        ("3.3 Discordant AST Results", "", "", ""),
        ("A - Very Major Deviations", "", context["vmd"], ""),
        ("B - Major Deviations", "", context["md"], ""),
        ("C - Minor Deviations", "", context["minor"], ""),
        ("3.4 Critical Deviations", "<= 5%", context["critical_deviation"], f"{context['critical_deviation_rate']}%"),
        ("3.5 Total Deviations", "<= 8%", context["total_deviation"], f"{context['total_deviation_rate']}%"),
    ]

    start_row = 7
    for offset, row_values in enumerate(rows):
        row_num = start_row + offset
        for col_num, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if row_values[0].startswith(("1.", "2.", "3.")) and row_values[0].count(".") == 1:
            for col_num in range(1, 5):
                ws.cell(row=row_num, column=col_num).fill = gray
                ws.cell(row=row_num, column=col_num).font = Font(bold=True)

    row_num = start_row + len(rows) + 2
    ws.cell(row=row_num, column=1, value="Table 2. Discordant Identification").font = Font(bold=True)
    row_num += 1
    for col_num, header in enumerate(["Bat Seq", "SITE IDENTIFICATION", "ARSRL IDENTIFICATION"], 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.fill = gray
        cell.font = Font(bold=True)
        cell.border = thin
    for item in context["discordant_rows"]:
        row_num += 1
        for col_num, value in enumerate([item["bat_seq"], item["site_org"], item["ars_org"]], 1):
            ws.cell(row=row_num, column=col_num, value=value).border = thin

    row_num += 2
    ws.cell(row=row_num, column=1, value="Table 3. AST Results with Deviations").font = Font(bold=True)
    row_num += 1
    ast_headers = ["Bat Seq", "Accession No", "ARSRL_FINAL", "Antibiotic", "Site RIS", "ARSRL RIS", "Deviation"]
    for col_num, header in enumerate(ast_headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.fill = gray
        cell.font = Font(bold=True)
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if context["discordant_details"]:
        for item in context["discordant_details"]:
            row_num += 1
            row_values = [
                item["bat_seq"],
                item["accession_no"],
                item["organism"],
                item["antibiotic"],
                item["site_ris"],
                item["ars_ris"],
                item["deviation_code"],
            ]
            for col_num, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    else:
        row_num += 1
        ws.cell(row=row_num, column=1, value="No AST deviation records.")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(ast_headers))
        for col_num in range(1, len(ast_headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")

    detail_ws = wb.create_sheet("AST Snapshot")
    detail_headers = ["Bat Seq", "Accession No", "Organism", "Antibiotic", "Site RIS", "ARSRL RIS", "Deviation"]
    for col_num, header in enumerate(detail_headers, 1):
        cell = detail_ws.cell(row=1, column=col_num, value=header)
        cell.fill = blue
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for detail in context["details"]:
        detail_ws.append([
            context["isolate_bat_seq_map"].get(detail.isolate_id, ""),
            detail.accession_no,
            detail.organism,
            detail.antibiotic,
            detail.site_ris,
            detail.ars_ris,
            detail.deviation_code,
        ])

    for sheet in [ws, detail_ws]:
        sheet.freeze_panes = "A2" if sheet == detail_ws else "A7"
        for column_index in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(column_index)
            max_length = 0
            for row_index in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                if isinstance(cell, MergedCell) or cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 42)

    filename = f"{report.batch.bat_Batch_Name}_Concordance.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url="login")
def export_concordance_accession_excel(request, report_id):

    # 🔒 Ensure this is an ACCESSION report only
    report = get_object_or_404(
        ConcordanceReport.objects.select_related("final_data"),
        id=report_id,
        final_data__isnull=False
    )

    isolate = report.final_data

    # =====================================================
    # ORGANISM CONCORDANCE (USING YOUR EXACT RULES)
    # =====================================================

    site_org = (isolate.f_Site_OrgName or "").strip().lower()
    ars_org = (isolate.f_ars_OrgName or "").strip().lower()
    ars_pre = (isolate.f_ars_Pre or "").strip().lower()

    if ars_pre == "mixed culture":
        genus_con = "M"
        species_con = "M"

    elif ars_org == "not viable":
        genus_con = "N"
        species_con = "N"

    elif site_org and ars_org and site_org == ars_org:
        genus_con = "G"
        species_con = "S"

    else:
        genus_con = "X"
        species_con = "X"

    genus_con, species_con = concordance_service.classify_id_concordance(
        isolate.f_Site_OrgName,
        isolate.f_ars_Pre,
        isolate.f_ars_OrgName,
        isolate.f_ars_Post,
    )

    # =====================================================
    # AST SNAPSHOT VALUES (FROM REPORT)
    # =====================================================

    total_pairs = report.total_pairs or 0
    concordant = report.concordant_pairs or 0
    vmd = report.vmd or 0
    md = report.md or 0
    minor = report.minor or 0

    total_deviation = vmd + md + minor

    concordance_rate = round((concordant / total_pairs) * 100, 2) if total_pairs else 0
    vmd_rate = round((vmd / total_pairs) * 100, 2) if total_pairs else 0
    total_deviation_rate = round((total_deviation / total_pairs) * 100, 2) if total_pairs else 0

    # =====================================================
    # WORKBOOK
    # =====================================================

    wb = Workbook()
    ws = wb.active
    ws.title = "Accession Concordance Report"

    # HEADER SECTION
    ws["A1"] = "Accession Number"
    ws["B1"] = isolate.f_AccessionNo

    ws["A2"] = "Sentinel Site"
    ws["B2"] = isolate.f_Site_Name

    ws["A3"] = "Date of Referral"
    ws["B3"] = str(isolate.f_Referral_Date)

    ws["A4"] = "Organism (Site)"
    ws["B4"] = isolate.f_Site_OrgName

    ws["A5"] = "Organism (ARSRL)"
    ws["B5"] = isolate.f_ars_OrgName

    for row in range(1, 6):
        ws[f"A{row}"].font = Font(bold=True)

    # =====================================================
    # 1. ORGANISM IDENTIFICATION
    # =====================================================

    ws["A7"] = "1. ORGANISM IDENTIFICATION"
    ws["A7"].font = Font(bold=True)

    ws["A8"] = "Genus Concordance"
    ws["B8"] = genus_con

    ws["A9"] = "Species Concordance"
    ws["B9"] = species_con

    # =====================================================
    # 2. AST RESULTS
    # =====================================================

    ws["A11"] = "2. AST RESULTS"
    ws["A11"].font = Font(bold=True)

    ws["A12"] = "Total AST pairs analyzed"
    ws["C12"] = total_pairs

    ws["A13"] = "Concordant Results"
    ws["C13"] = concordant
    ws["D13"] = f"{concordance_rate}%"

    ws["A14"] = "Very Major Deviations (A)"
    ws["C14"] = vmd
    ws["D14"] = f"{vmd_rate}%"

    ws["A15"] = "Major Deviations (B)"
    ws["C15"] = md

    ws["A16"] = "Minor Deviations (C)"
    ws["C16"] = minor

    ws["A17"] = "Total Deviations"
    ws["C17"] = total_deviation
    ws["D17"] = f"{total_deviation_rate}%"

    # =====================================================
    # DETAIL TABLE (FROM SNAPSHOT)
    # =====================================================

    ws["A19"] = "3. AST Detail Comparison"
    ws["A19"].font = Font(bold=True)

    headers = ["Antibiotic", "Site RIS", "ARS RIS", "Deviation"]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=20, column=col_num, value=header).font = Font(bold=True)

    row_index = 21

    for detail in report.details.all().order_by("antibiotic"):
        ws.cell(row=row_index, column=1, value=detail.antibiotic)
        ws.cell(row=row_index, column=2, value=detail.site_ris)
        ws.cell(row=row_index, column=3, value=detail.ars_ris)
        ws.cell(row=row_index, column=4, value=detail.deviation_code)
        row_index += 1

    # AUTO WIDTH
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    filename = f"{isolate.f_AccessionNo}_Concordance.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response



def link_callback(uri, rel):
    """
    Converts static/media URLs to absolute file paths for xhtml2pdf.
    Needed for images like header_ed.jpg and footer.jpg.
    """

    path = urlparse(uri).path

    if path.startswith(settings.MEDIA_URL):
        file_path = os.path.join(
            settings.MEDIA_ROOT,
            path.replace(settings.MEDIA_URL, "")
        )

    elif path.startswith(settings.STATIC_URL):
        relative_path = path.replace(settings.STATIC_URL, "")
        file_path = finders.find(relative_path)

        if not file_path:
            file_path = os.path.join(settings.STATIC_ROOT, relative_path)

    else:
        return uri

    if not os.path.isfile(file_path):
        raise Exception(f"File not found: {file_path}")

    return file_path


# @login_required(login_url="login")
# def export_concordance_report_pdf(request, report_id):

#     report = get_object_or_404(
#         ConcordanceReport.objects.select_related("batch"),
#         id=report_id
#     )

#     batch = report.batch
#     isolates = Final_Data.objects.filter(f_Batch_id=batch)

#     total_isolates = isolates.count()

#     site_name = isolates.first().f_Site_Name if isolates.exists() else ""
#     referral_date = isolates.first().f_Referral_Date if isolates.exists() else ""
#     refno = isolates.first().f_RefNo if isolates.exists() else ""

#     # =====================================================
#     # ORGANISM CONCORDANCE (LIVE CALCULATION)
#     # =====================================================

#     genus_match = 0
#     species_match = 0
#     different_org = 0
#     mixed_count = 0
#     nonviable_count = 0
#     discordant_rows = []

#     for isolate in isolates:

#         site_org = (isolate.f_Site_OrgName or "").strip().lower()
#         ars_org = (isolate.f_ars_OrgName or "").strip().lower()
#         ars_pre = (isolate.f_ars_Pre or "").strip().lower()

#         if "mixed culture" in ars_pre:
#             mixed_count += 1
#             continue

#         if ars_org == "not viable":
#             nonviable_count += 1
#             continue

#         if site_org and ars_org:

#             if site_org == ars_org:
#                 species_match += 1
#                 genus_match += 1

#             elif site_org.split(" ")[0] == ars_org.split(" ")[0]:
#                 genus_match += 1

#             else:
#                 different_org += 1
#                 discordant_rows.append({
#                     "refno": isolate.f_RefNo,
#                     "site_org": isolate.f_Site_OrgName,
#                     "ars_org": isolate.f_ars_OrgName
#                 })

#     viable_pure = total_isolates - mixed_count - nonviable_count

#     genus_rate = round((genus_match / viable_pure) * 100, 2) if viable_pure else 0
#     species_rate = round((species_match / viable_pure) * 100, 2) if viable_pure else 0

#     id_stats = concordance_service.build_id_stats(isolates)
#     genus_match = id_stats["genus_match"]
#     species_match = id_stats["species_match"]
#     different_org = id_stats["different_org"]
#     mixed_count = id_stats["mixed_count"]
#     nonviable_count = id_stats["nonviable_count"]
#     viable_pure = id_stats["viable_pure"]
#     genus_rate = id_stats["genus_rate"]
#     species_rate = id_stats["species_rate"]
#     discordant_rows = id_stats["discordant_rows"]

#     # =====================================================
#     # AST CALCULATIONS (FROM SNAPSHOT)
#     # =====================================================

#     total_pairs = report.total_pairs or 0
#     concordant = report.concordant_pairs or 0
#     vmd = report.vmd or 0
#     md = report.md or 0
#     minor = report.minor or 0

#     total_deviation = vmd + md + minor

#     concordance_rate = round((concordant / total_pairs) * 100, 2) if total_pairs else 0
#     vmd_rate = round((vmd / total_pairs) * 100, 2) if total_pairs else 0
#     total_deviation_rate = round((total_deviation / total_pairs) * 100, 2) if total_pairs else 0

#     # =====================================================
#     # DISCORDANT ANTIBIOTIC SUMMARY
#     # =====================================================

#     details = report.details.all()
#     abx_summary = concordance_service.calculate_antibiotic_summary(details)

#     context = {
#         "report": report,
#         "site_name": site_name,
#         "referral_date": referral_date,
#         "refno": refno,
#         "total_isolates": total_isolates,
#         "mixed_count": mixed_count,
#         "nonviable_count": nonviable_count,
#         "viable_pure": viable_pure,
#         "genus_match": genus_match,
#         "species_match": species_match,
#         "different_org": different_org,
#         "genus_rate": genus_rate,
#         "species_rate": species_rate,
#         "discordant_rows": discordant_rows,
#         "total_pairs": total_pairs,
#         "concordant": concordant,
#         "vmd": vmd,
#         "md": md,
#         "minor": minor,
#         "concordance_rate": concordance_rate,
#         "vmd_rate": vmd_rate,
#         "total_deviation": total_deviation,
#         "total_deviation_rate": total_deviation_rate,
#         "abx_summary": abx_summary,
#         "now": datetime.now().strftime("%d %B %Y"),
#         "header_path": static("assets/img/brand/header_ed.png"),
#         "footer_path": static("assets/img/brand/footer.png"),
        
#     }

#     template = get_template("home_final/concordance_report_pdf.html")
#     html = template.render(context)

#     response = HttpResponse(content_type="application/pdf")
#     response["Content-Disposition"] = (
#         f'inline; filename=Batch_{report.batch.bat_Batch_Name}_Concordance.pdf'
#     )

#     pisa_status = pisa.CreatePDF(
#     html,
#     dest=response,
#     link_callback=link_callback
# )

#     if pisa_status.err:
#         return HttpResponse("PDF generation error", status=500)

#     return response
